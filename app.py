"""
Control Pedidos Princess Canarias — Flask + PostgreSQL (Supabase)
Despliegue: Render.com  |  BD: Supabase  |  Email: EmailJS (frontend)
"""

import os, json, logging, secrets, atexit, hashlib, re, threading, base64, hmac, gzip
from html import unescape as _html_unescape
from datetime import datetime, timedelta, timezone, date as _date
from functools import wraps

from apscheduler.schedulers.background import BackgroundScheduler

import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2.pool import ThreadedConnectionPool
import requests

from flask import Flask, request, jsonify, send_from_directory, session, g, Response
from werkzeug.security import generate_password_hash, check_password_hash
from models import ESTADOS_VALIDOS, ESTADOS_EMAIL_PROVEEDOR, ESTADOS_EMAIL_INTERNO

# ── Configuración ──────────────────────────────────────────────────────────────

DATABASE_URL = os.environ.get("DATABASE_URL", "")          # Supabase → Settings → Database → URI
SECRET_KEY = os.environ["SECRET_KEY"]

# Tamaño del pool de conexiones (v12.7.0). Antes cada request abría y
# cerraba su propia conexión con psycopg2.connect() — con el tráfico actual
# (10 hoteles) ese handshake TCP/TLS repetido en cada petición era el
# principal cuello de botella. Ahora se reutilizan conexiones ya abiertas de
# un pool. MAXCONN limita cuántas conexiones físicas simultáneas se mantienen
# contra Supabase. Ajustable por variable de entorno sin tocar código.
DB_POOL_MAXCONN = int(os.environ.get("DB_POOL_MAXCONN", "15"))

# ── Storage para adjuntos de pedidos cerrados (v12.8.0) ──────────────────────
# pedido_adjuntos.datos (bytea) es, con diferencia, la mayor consumidora del
# tamaño de la base de datos (ver Admin → Integridad → Tamaño de BD). Los
# adjuntos de pedidos ya cerrados (ENTREGADO/CANCELADO) no vuelven a
# escribirse nunca, así que se migran a Supabase Storage — el dato sigue
# siendo consultable exactamente igual desde /api/adjuntos/<id>, solo cambia
# dónde vive el byte. Importante: esto reduce TAMAÑO DE BD, no egress — cada
# descarga desde Storage sigue contando como egress igual que antes.
# SUPABASE_URL: la URL del proyecto (https://xxxx.supabase.co), NO la de la
# base de datos. SUPABASE_SERVICE_ROLE_KEY: Supabase → Settings → API →
# service_role (NUNCA la anon/public key — esta debe quedarse solo en el
# servidor, nunca llegar al navegador).
SUPABASE_URL = os.environ.get("SUPABASE_URL", "").rstrip("/")
SUPABASE_SERVICE_ROLE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_STORAGE_BUCKET = os.environ.get("SUPABASE_STORAGE_BUCKET", "adjuntos-cerrados")
STORAGE_CONFIGURADO = bool(SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY)
ESTADOS_CERRADOS = ("ENTREGADO", "CANCELADO")

# ── SSO hacia el catálogo DALI (v12.30.02) ──────────────────────────────────
# Acceso de un clic desde el menú lateral "Catálogo DALI": este backend
# firma un token de un solo uso (ver _generar_token_sso_dali) con el rol ya
# mapeado (compras -> admin, hotel -> hotel), y el backend de DALI
# (backend/src/controllers/authController.js, DALI_SSO_SECRET debe ser
# IDÉNTICA en los dos servicios de Render) lo verifica, aprovisiona/actualiza
# el usuario en su propia tabla `usuarios` y abre sesión sin pedir
# contraseña. DALI_FRONTEND_URL es la URL pública del frontend de DALI (el
# Worker de Cloudflare que hace de proxy, no la de onrender.com directa).
DALI_SSO_SECRET = os.environ.get("DALI_SSO_SECRET", "")
DALI_FRONTEND_URL = os.environ.get(
    "DALI_FRONTEND_URL", "https://dali-proxy.centralcompras1-canarias.workers.dev"
).rstrip("/")
# compras (comprador) -> admin: gestiona el catálogo igual que en Compras.
# hotel -> hotel: mismo acceso de solo consulta que ya tiene en DALI.
# admin -> admin. Cualquier otro rol (p.ej. "user", legado) no tiene mapeo:
# sin acceso a DALI.
DALI_ROL_MAP = {"admin": "admin", "compras": "admin", "hotel": "hotel"}

# ── Hotel de pruebas (v12.29.41 / v12.29.42) ────────────────────────────────
# El hotel "PR" (⚠️ HOTEL PRUEBAS) existe solo para pruebas internas. Solo
# puede verlo/usarlo: (a) el rol admin, y (b) el usuario dedicado a estas
# pruebas (username 'usuario prueba'), sea cual sea su rol real — el resto de
# usuarios no deben verlo ni interactuar con sus pedidos en ningún listado,
# dropdown, dashboard o alerta automática. Ver _puede_ver_hotel_pruebas(),
# _es_hotel_pruebas_id() y los filtros aplicados en /api/maestros,
# /api/pedidos, /api/stats, /api/techo/resumen y los jobs de alertas.
HOTEL_CODIGO_PRUEBAS = "PR"
USERNAME_HOTEL_PRUEBAS = "usuario prueba"

# Email — gestionado enteramente por EmailJS en el frontend
# EMAILS_INTERNOS eliminado: los destinatarios internos se leen siempre de la BD (rol admin/compras)

app = Flask(__name__)
app.secret_key = SECRET_KEY
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

# ── Compresión gzip de respuestas de texto (2026-08-31, auditoría de
# rendimiento, Etapa 3/3) ────────────────────────────────────────────────
# Ninguna respuesta salía comprimida — ni el JSON de la API (el listado de
# Pedidos o de Proveedores puede pesar cientos de KB) ni, sobre todo,
# templates/index.html (628 KB, servido con Cache-Control: no-cache, así
# que se descarga entero en cada carga/recarga de página).
#
# Se probó primero la librería Flask-Compress, pero se descartó por dos
# motivos concretos, comprobados aquí antes de escribir nada de esto:
# 1) TODAS sus versiones (hasta la 1.10.0, la más antigua publicada)
#    dependen obligatoriamente del paquete `brotli` (una extensión en C) —
#    no hay forma de usarla en modo "solo gzip, sin dependencias nuevas".
# 2) Con esa librería instalada tal cual, index.html —el fichero que más
#    pesa, el que más interesaba comprimir— NO llegaba a comprimirse con
#    gzip: send_from_directory() devuelve una respuesta "en streaming", y
#    Flask-Compress excluye gzip a propósito de su lista de algoritmos
#    para streaming (usa brotli/zstd/deflate ahí, nunca gzip) — así que
#    hacía falta que brotli funcionara de verdad en el build de Render
#    para que el propio objetivo de esta etapa se cumpliera.
#
# Con un `after_request` propio, de una docena de líneas, se evita la
# dependencia nueva y se comprime igual el fichero que más importa. Solo
# actúa sobre contenido de texto (HTML/CSS/JS/JSON — nunca los PDF, Excel
# o imágenes que ya sirve la app, que van comprimidos de otra forma o no
# se benefician de gzip), respeta si el navegador anuncia soporte de gzip
# (cabecera Accept-Encoding) y no toca respuestas ya comprimidas ni por
# debajo de un tamaño mínimo (gzip añade su propia cabecera, no compensa
# para respuestas muy pequeñas).
#
# No se toca el ETag de index() (más abajo) a propósito: esa vista ya
# implementa su propio atajo 304 comparando If-None-Match contra el hash
# tal cual, sin sufijo — si aquí se le añadiera algo como ":gzip" (como
# hacen algunas librerías, para no confundir la versión comprimida con la
# normal bajo el mismo validador), ese atajo dejaría de encajar y CADA
# carga volvería a mandar el HTML entero, exactamente el problema de
# egress que el ETag se creó para evitar (ver comentario de index()).
# Sin caché intermedia compartida delante de esta app (Render sirve
# directo), no hay ningún cliente que pueda recibir la copia comprimida
# de otro por error — así que no hace falta ese sufijo aquí.
_GZIP_MIMETYPES = {
    "text/html", "text/css", "text/javascript", "application/javascript",
    "application/json", "text/plain", "text/xml", "application/xml",
}
_GZIP_MIN_BYTES = 500

@app.after_request
def _comprimir_respuesta_gzip(response):
    vary = response.headers.get("Vary")
    if not vary:
        response.headers["Vary"] = "Accept-Encoding"
    elif "accept-encoding" not in vary.lower():
        response.headers["Vary"] = f"{vary}, Accept-Encoding"

    if (
        "gzip" not in (request.headers.get("Accept-Encoding") or "")
        or response.mimetype not in _GZIP_MIMETYPES
        or "Content-Encoding" in response.headers
        or response.status_code < 200 or response.status_code >= 300
    ):
        return response

    # send_from_directory() (usada para los ficheros de static/, y como
    # respaldo en index() si falla la lectura normal de index.html) marca
    # la respuesta como "direct_passthrough" para poder enviar el fichero
    # sin cargarlo entero en memoria de golpe — hay que desactivarlo antes
    # de poder leer/reemplazar el cuerpo con get_data()/set_data(), si no
    # Werkzeug lanza un error. (La vía normal de index() ya construye la
    # respuesta con Response(content, ...) directamente, sin passthrough,
    # pero desactivarlo aquí no le afecta — sigue funcionando igual.)
    response.direct_passthrough = False
    data = response.get_data()
    if len(data) < _GZIP_MIN_BYTES:
        return response

    response.set_data(gzip.compress(data, compresslevel=6))
    response.headers["Content-Encoding"] = "gzip"
    response.headers["Content-Length"] = str(response.content_length)
    return response

def _auto_migrate():
    """Añade columnas/tablas nuevas de forma idempotente."""
    try:
        db = psycopg2.connect(
            DATABASE_URL, cursor_factory=RealDictCursor,
            connect_timeout=20,
            application_name="control_pedidos_web_migracion",
        )
        db.autocommit = True
        with db.cursor() as cur:
            # ══════════════════════════════════════════════════════════════
            # (2026-08-10) Bloque movido AQUÍ, al principio del todo — antes
            # vivía casi al final de esta función (justo antes de db.close()),
            # y _auto_migrate() tiene 111 sentencias en total, la inmensa
            # mayoría sin try/except propio. Si CUALQUIERA de esas otras 100+
            # sentencias fallaba por el motivo que fuera —sin relación con
            # estos cambios—, el except genérico de toda la función paraba la
            # ejecución ahí mismo y este bloque, al estar casi al final, nunca
            # llegaba a ejecutarse. Bug real confirmado en producción: RLS
            # sin activar en Supabase y /api/proveedores con 500 por
            # "column sujeto_seguimiento does not exist", pese a llevar
            # semanas desplegado. Poniéndolo el primero de todos, se garantiza
            # que se aplique siempre, pase lo que pase más abajo en el resto
            # de la función esa misma ejecución — cada sentencia sigue con su
            # propio try/except además, por si falla alguna de estas 3 en
            # concreto.
            # ── RLS en tablas propias de la app sin política pública ────────
            # Supabase expone TODAS las tablas del esquema public vía su API
            # REST automática (PostgREST) salvo que tengan RLS activado — el
            # Security Advisor lo marca como error ("RLS Disabled in Public").
            # Esta app nunca usa esa API (ni el backend ni el frontend; todo
            # habla por conexión directa a Postgres con DATABASE_URL, nunca
            # con la anon key), así que activar RLS sin ninguna política es
            # 100% seguro para el funcionamiento — simplemente cierra el
            # acceso público accidental por ese otro camino.
            for _tabla_rls in ("egress_tracking", "db_size_tracking", "db_vacuum_log", "agente_heartbeat",
                                "expediente_exceso", "proveedor_contacto_hoteles", "bridge_popup_visto"):
                try:
                    cur.execute(f"ALTER TABLE IF EXISTS {_tabla_rls} ENABLE ROW LEVEL SECURITY")
                except Exception as e:
                    log.warning(f"No se pudo activar RLS en {_tabla_rls}: {e}")
            # ── Índices de búsqueda de proveedores (2026-08-31, auditoría de
            # rendimiento — Víctor: "la ficha proveedores se atasca un poco")
            # ─────────────────────────────────────────────────────────────
            # get_proveedores() busca con ILIKE '%texto%' (comodín al
            # PRINCIPIO) en nombre/codigo/codigo_dali a la vez — ese patrón
            # no puede usar un índice normal (B-tree), así que cada letra
            # tecleada en el buscador de Proveedores obligaba a Postgres a
            # recorrer la tabla entera de proveedores. La extensión pg_trgm
            # (trigramas) sí permite indexar ILIKE con comodín al principio;
            # Supabase la trae disponible de fábrica, sin coste ni permisos
            # especiales. Bloque puesto aquí arriba, cada sentencia con su
            # propio try/except, por el mismo motivo ya documentado para
            # sujeto_seguimiento/codigo_dali más abajo: _auto_migrate() tiene
            # 111+ sentencias, la mayoría sin try/except propio, y un fallo
            # anterior cualquiera abortaría la función antes de llegar aquí
            # si este bloque viviera al final.
            try:
                cur.execute("CREATE EXTENSION IF NOT EXISTS pg_trgm")
            except Exception as e:
                log.warning(f"No se pudo crear la extensión pg_trgm: {e}")
            for _idx_nombre, _columna in (
                ("idx_prov_nombre_trgm",      "nombre"),
                ("idx_prov_codigo_trgm",      "codigo"),
                ("idx_prov_codigo_dali_trgm", "codigo_dali"),
            ):
                try:
                    cur.execute(
                        f"CREATE INDEX IF NOT EXISTS {_idx_nombre} ON proveedores USING gin ({_columna} gin_trgm_ops)"
                    )
                except Exception as e:
                    log.warning(f"No se pudo crear el índice {_idx_nombre}: {e}")
            # ── Índices de búsqueda de pedidos (2026-08-31, auditoría de
            # rendimiento, Etapa 2/3) ───────────────────────────────────────
            # get_pedidos() busca con ILIKE '%texto%' (comodín al principio)
            # en pedido_num/observaciones a la vez (pr.nombre ya quedó
            # cubierto por idx_prov_nombre_trgm, arriba; h.codigo no —
            # hoteles tiene ~10 filas, un índice ahí no aporta nada). Mismo
            # problema que en Proveedores: sin este índice, cada búsqueda
            # recorre la tabla `pedidos` entera — y esa tabla crece muy
            # deprisa (Víctor: +306,7% de pedidos en el último mes visto en
            # el propio dashboard), así que sin esto la búsqueda se pone más
            # lenta cada mes que pasa, aunque no se toque nada más.
            try:
                cur.execute("CREATE EXTENSION IF NOT EXISTS pg_trgm")  # no-op si Proveedores ya la creó
            except Exception as e:
                log.warning(f"No se pudo crear la extensión pg_trgm: {e}")
            for _idx_nombre, _columna in (
                ("idx_pedidos_pedido_num_trgm",   "pedido_num"),
                ("idx_pedidos_observaciones_trgm", "observaciones"),
            ):
                try:
                    cur.execute(
                        f"CREATE INDEX IF NOT EXISTS {_idx_nombre} ON pedidos USING gin ({_columna} gin_trgm_ops)"
                    )
                except Exception as e:
                    log.warning(f"No se pudo crear el índice {_idx_nombre}: {e}")
            # ── Índices B-tree de filtros/orden de pedidos (2026-09-01,
            # revisión "agilizar y limpiar" tras el cierre de la auditoría de
            # rendimiento) ───────────────────────────────────────────────────
            # get_pedidos() (GET /api/pedidos, la pantalla principal) filtra
            # por hotel_id/estado/departamento_id/fecha_solicitud y ordena
            # por creado_en o norden según el caso — ninguna de esas columnas
            # tenía índice propio (los trgm de arriba solo cubren el texto
            # libre de pedido_num/observaciones). Sin esto, tanto el listado
            # como su COUNT(*) de paginación recorren la tabla entera en cada
            # página — coste que solo crece con el tiempo, igual que con los
            # índices trgm. No reduce el tamaño de lo que se devuelve (eso ya
            # lo limita la paginación), así que esto es una mejora de
            # velocidad/cómputo en Supabase, no de egress — el ahorro de
            # egress real está en las Etapas siguientes (expedientes,
            # eliminados).
            for _idx_nombre, _columna in (
                ("idx_pedidos_hotel_id",        "hotel_id"),
                ("idx_pedidos_estado",          "estado"),
                ("idx_pedidos_departamento_id", "departamento_id"),
                ("idx_pedidos_fecha_solicitud", "fecha_solicitud"),
                ("idx_pedidos_creado_en",       "creado_en"),
                ("idx_pedidos_norden",          "norden"),
            ):
                try:
                    cur.execute(
                        f"CREATE INDEX IF NOT EXISTS {_idx_nombre} ON pedidos ({_columna})"
                    )
                except Exception as e:
                    log.warning(f"No se pudo crear el índice {_idx_nombre}: {e}")
            # fecha_tramitacion solo se filtra con "IS NOT NULL" (filtro
            # rápido alerta=1) — índice parcial, más pequeño y más útil que
            # uno completo ya que la mayoría de pedidos no están en trámite.
            try:
                cur.execute(
                    "CREATE INDEX IF NOT EXISTS idx_pedidos_fecha_tramitacion "
                    "ON pedidos (fecha_tramitacion) WHERE fecha_tramitacion IS NOT NULL"
                )
            except Exception as e:
                log.warning(f"No se pudo crear el índice idx_pedidos_fecha_tramitacion: {e}")
            # historial_estados.pedido_id: sin índice, se recorría entera en
            # cada apertura del detalle de un pedido (GET /api/pedidos/<id>,
            # WHERE pedido_id=%s ORDER BY creado_en DESC) — se incluye
            # creado_en en el mismo índice porque es exactamente el orden que
            # pide esa consulta.
            try:
                cur.execute(
                    "CREATE INDEX IF NOT EXISTS idx_historial_estados_pedido_id "
                    "ON historial_estados (pedido_id, creado_en DESC)"
                )
            except Exception as e:
                log.warning(f"No se pudo crear el índice idx_historial_estados_pedido_id: {e}")
            # ── Verificación de listados PDF de SAP — filtro de proveedores ──
            # (2026-08-10) A petición del usuario: el criterio se invierte a
            # "opt-in" — DEFAULT FALSE, ningún proveedor está sujeto a
            # seguimiento hasta que un admin lo marque explícitamente en su
            # ficha. Antes era "opt-out" (DEFAULT TRUE, había que desmarcar
            # uno a uno alimentación/bebida) — con tantísimos proveedores de
            # compra diaria frente a los pocos que sí interesa seguir, es más
            # seguro empezar todos apagados y que el admin encienda solo los
            # que quiere vigilar, que al revés.
            # FIX (mismo día): el SQL de emergencia que se dio para
            # desbloquear /api/proveedores mientras esta migración no se
            # aplicaba (v12.29.66) usaba DEFAULT TRUE — versión anterior a
            # este cambio de criterio. Si se llegó a ejecutar a mano, la
            # columna ya existía con DEFAULT TRUE y TODOS los proveedores en
            # TRUE, y un simple "ADD COLUMN IF NOT EXISTS ... DEFAULT FALSE"
            # no toca nada si la columna ya existe — por eso seguían todos
            # marcados. Se comprueba el DEFAULT real de la columna en
            # information_schema y, si no es FALSE (columna inexistente o
            # con el DEFAULT antiguo), se crea/corrige el DEFAULT y se
            # resetean a FALSE los proveedores que estuvieran en TRUE — algo
            # seguro de hacer aquí porque, al ser una funcionalidad
            # recién nacida, nadie ha podido marcar todavía ninguno a
            # propósito (la pantalla llevaba rota desde que se introdujo).
            # Es correctivo y de una sola vez: en cuanto el DEFAULT quede en
            # FALSE, esta condición deja de cumplirse y no se vuelve a tocar
            # nada en arranques futuros, así que si un admin marca luego
            # proveedores concretos, esos cambios quedan a salvo para
            # siempre.
            try:
                cur.execute("""
                    SELECT column_default FROM information_schema.columns
                    WHERE table_name='proveedores' AND column_name='sujeto_seguimiento'
                """)
                _fila_col = cur.fetchone()
                _default_actual = (_fila_col or {}).get("column_default") or ""
                if "false" not in _default_actual.lower():
                    cur.execute(
                        "ALTER TABLE proveedores ADD COLUMN IF NOT EXISTS sujeto_seguimiento BOOLEAN NOT NULL DEFAULT FALSE"
                    )
                    cur.execute("ALTER TABLE proveedores ALTER COLUMN sujeto_seguimiento SET DEFAULT FALSE")
                    cur.execute("UPDATE proveedores SET sujeto_seguimiento = FALSE WHERE sujeto_seguimiento = TRUE")
                    log.info("[MIGRACION] sujeto_seguimiento corregido a opt-in (DEFAULT FALSE) y proveedores reseteados")
            except Exception as e:
                log.warning(f"No se pudo corregir sujeto_seguimiento a opt-in: {e}")
            # ── Hotel de pruebas "PR" ─────────────────────────────────────
            try:
                cur.execute("""
                    INSERT INTO hoteles (codigo, nombre) VALUES
                        ('PR', '⚠️ HOTEL PRUEBAS — no usar en operativa real')
                    ON CONFLICT DO NOTHING
                """)
                if cur.rowcount:
                    log.info("[MIGRACION] Hotel de pruebas 'PR' insertado")
            except Exception as e:
                log.warning(f"No se pudo insertar el hotel de pruebas 'PR': {e}")
            # ── Total Pedido real (v12.30.30) ────────────────────────────────
            # (2026-08-28) Movido aquí, al bloque protegido del principio —
            # vivía casi al final de la función (justo antes de db.close()),
            # y una sentencia anterior cualquiera de las 100+ sin try/except
            # propio bastaba para abortar la función entera antes de llegar
            # aquí, dejando la columna sin crear en Supabase sin ningún aviso
            # visible salvo un log.warning genérico — bug real confirmado en
            # producción (Víctor: "Comparar listado PDF (SAP)" fallaba con
            # 'column "total_pedido" does not exist' pese a llevar desplegado
            # desde v12.30.30). Mismo motivo exacto que ya obligó a mover
            # aquí el bloque de sujeto_seguimiento (ver comentario de arriba
            # del todo de esta función).
            #
            # Nuevo campo "TOTAL PEDIDO" en la ficha del pedido — el comprador
            # puede rellenarlo a mano si quiere, pero la idea (petición de
            # Víctor) es que no haga falta: al comparar el "Listado de Pedidos"
            # PDF de SAP, el importe base (6ª columna del PDF, ya extraído como
            # `importe_base` en _comparar_listado_pdf_logica) se guarda aquí
            # automáticamente para cada pedido localizado, como el valor real
            # del pedido — sin ningún paso manual de confirmación, a
            # diferencia del resto de "Comparar listado PDF" (que solo
            # propone, nunca escribe solo): este campo es puramente
            # informativo, no dispara notificaciones ni cambia estado, así
            # que no hay nada que confirmar.
            try:
                cur.execute(
                    "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS total_pedido NUMERIC(10,2)"
                )
            except Exception as e:
                log.warning(f"No se pudo añadir la columna pedidos.total_pedido: {e}")
            # ── Código DALI del proveedor (2026-08-31) ────────────────────────
            # Víctor: "en la ficha de proveedores, necesito junto a la casilla
            # CODGIGO SAP, OTRA PARA CODIGO DALI ; Actualmente estamos
            # trabajando con los dos sistemas y vamos asociando tanto
            # artículos como proveedores." — columna nueva, editable a mano
            # (sin integración automática con la app DALI todavía, es solo
            # referencia cruzada mientras dura la migración a los dos
            # sistemas en paralelo). Puesta aquí, en el bloque protegido del
            # principio con su propio try/except, y NO al final de la función
            # (como se hizo primero, error corregido en la misma entrega) —
            # exactamente el mismo motivo, ya documentado arriba, por el que
            # hubo que mover aquí sujeto_seguimiento y total_pedido: son 111
            # sentencias en total, la mayoría sin try/except propio, y
            # cualquier fallo anterior en la función aborta todo lo que venga
            # después. Confirmado en producción: /api/proveedores daba 500
            # con 'column "codigo_dali" does not exist' nada más desplegar,
            # porque la migración vivía al final y nunca llegaba a ejecutarse.
            try:
                cur.execute(
                    "ALTER TABLE proveedores ADD COLUMN IF NOT EXISTS codigo_dali TEXT"
                )
            except Exception as e:
                log.warning(f"No se pudo añadir la columna proveedores.codigo_dali: {e}")
            # ── Marcado automático de "Comunicado A&B" / "Comunicado Jefe
            # Dep." (2026-08-31) ──────────────────────────────────────────
            # Víctor: "cuando el correo interno de 'PEDIDO ENVIADO AL
            # PROVEEDOR' va con copia al departamento A&B se marque
            # automáticamente la casilla y en todos los casos que se ponga
            # en copia al responsable del departamento también se marque la
            # correspondiente (...) no podrán ser modificadas por el
            # usuario, solo con el envío del correo." Estas dos columnas
            # nuevas van en emails_sistema_pendientes (no en pedidos, que ya
            # tiene comunicado_ab/comunicado_jefe_dep desde antes): guardan,
            # en el momento en que se ENCOLA el correo "ENVIADO AL
            # PROVEEDOR", si ese envío concreto va a incluir a A&B y/o al
            # correo del departamento — para poder aplicarlo a pedidos.* SOLO
            # cuando se confirme que el correo se ha enviado de verdad (ver
            # api_marcar_email_sistema_enviado), nunca antes. Puesta aquí, en
            # el bloque protegido, por el mismo motivo que codigo_dali/
            # total_pedido/sujeto_seguimiento arriba.
            try:
                cur.execute(
                    "ALTER TABLE emails_sistema_pendientes "
                    "ADD COLUMN IF NOT EXISTS marca_comunicado_ab BOOLEAN NOT NULL DEFAULT FALSE"
                )
            except Exception as e:
                log.warning(f"No se pudo añadir la columna emails_sistema_pendientes.marca_comunicado_ab: {e}")
            try:
                cur.execute(
                    "ALTER TABLE emails_sistema_pendientes "
                    "ADD COLUMN IF NOT EXISTS marca_comunicado_jefe_dep BOOLEAN NOT NULL DEFAULT FALSE"
                )
            except Exception as e:
                log.warning(f"No se pudo añadir la columna emails_sistema_pendientes.marca_comunicado_jefe_dep: {e}")
            # (2026-09-01) Fix duplicados reales de correo interno de cambio de
            # estado (ENVIADO AL PROVEEDOR / ENTREGA PARCIAL / ENTREGADO):
            # cuando emailjs.send() ya ha entregado el correo DE VERDAD pero la
            # confirmación posterior (marcar-enviado) falla repetidamente, la
            # fila se quedaba enviado=FALSE y la reserva de 2 min caducaba —
            # el siguiente sondeo (5 min después) la reclamaba y volvía a
            # llamar a emailjs.send() de verdad: un segundo correo real, no un
            # simple reintento. Esta columna marca esas filas como "se entregó
            # pero no se pudo confirmar en BD" para (a) que el poller deje de
            # reclamarlas para siempre (ver api_marcar_email_sistema_enviado_
            # no_confirmado, que sube intentos a MAX_INTENTOS_EMAIL_SISTEMA) y
            # (b) que el panel de admin las distinga de un fallo real de envío
            # — con "Marcar como enviado" en vez de "Reactivar", para no
            # arriesgarse a un reenvío real por error.
            try:
                cur.execute(
                    "ALTER TABLE emails_sistema_pendientes "
                    "ADD COLUMN IF NOT EXISTS enviado_no_confirmado BOOLEAN NOT NULL DEFAULT FALSE"
                )
            except Exception as e:
                log.warning(f"No se pudo añadir la columna emails_sistema_pendientes.enviado_no_confirmado: {e}")
            # ── Correo de departamento por hotel (2026-08-28) ────────────────
            # A petición de Víctor: cada hotel tiene un correo distinto para
            # el mismo departamento (RESTAURANTE de JN != RESTAURANTE de GY),
            # así que hace falta esta tabla nueva de relación — ver también
            # models.py (SQL_STATEMENTS, para instalaciones nuevas desde
            # cero) y la nota de arriba sobre por qué toda sentencia de
            # esquema nueva se pone aquí, en el bloque protegido, con su
            # propio try/except.
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS departamento_hotel_email (
                        id              SERIAL PRIMARY KEY,
                        hotel_id        INTEGER NOT NULL REFERENCES hoteles(id) ON DELETE CASCADE,
                        departamento_id INTEGER NOT NULL REFERENCES departamentos(id) ON DELETE CASCADE,
                        email           TEXT,
                        email2          TEXT,
                        UNIQUE (hotel_id, departamento_id)
                    )
                """)
            except Exception as e:
                log.warning(f"No se pudo crear la tabla departamento_hotel_email: {e}")
            # ── Tokens de descarga pública de adjuntos (2026-08-28) ──────────
            # A petición de Víctor: enlace de descarga del PDF del pedido en
            # el correo al proveedor, en vez de adjuntarlo directamente
            # (EmailJS en el plan Free no admite adjuntos) — ver también
            # models.py, _obtener_o_crear_token_adjunto() y
            # /descargas/adjunto/<token> más abajo en este mismo archivo.
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS adjunto_descarga_tokens (
                        id         SERIAL PRIMARY KEY,
                        adjunto_id INTEGER NOT NULL REFERENCES pedido_adjuntos(id) ON DELETE CASCADE,
                        token      TEXT NOT NULL UNIQUE,
                        expira_en  TIMESTAMPTZ NOT NULL,
                        creado_en  TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_adjunto_token ON adjunto_descarga_tokens(token)")
            except Exception as e:
                log.warning(f"No se pudo crear la tabla adjunto_descarga_tokens: {e}")
            # ── Contactos adicionales de notificación (2026-08-28) ───────────
            # A petición de Víctor: además del comprador, rol hotel y correo
            # de departamento (departamento_hotel_email, arriba), quiere
            # poder registrar contactos sueltos que no son usuarios de la
            # app (p. ej. "Chef Ejecutivo", "Director de Compras",
            # "Administrativo A&B") y decidir, por CADA contacto, en qué
            # departamentos y para qué cambios de estado concretos se les
            # pone en copia en el correo interno de cambio de estado —
            # global para toda la cadena (decisión confirmada con Víctor:
            # no varía por hotel, a diferencia de departamento_hotel_email).
            # Ver también models.py (SQL_STATEMENTS, instalaciones nuevas) y
            # el uso real en enviar_emails_estado() y
            # /api/admin/notificaciones-contactos más abajo en este archivo.
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS notificacion_contactos (
                        id        SERIAL PRIMARY KEY,
                        nombre    TEXT NOT NULL,
                        email     TEXT,
                        email2    TEXT,
                        activo    INTEGER NOT NULL DEFAULT 1,
                        creado_en TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
            except Exception as e:
                log.warning(f"No se pudo crear la tabla notificacion_contactos: {e}")
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS notificacion_contacto_reglas (
                        id              SERIAL PRIMARY KEY,
                        contacto_id     INTEGER NOT NULL REFERENCES notificacion_contactos(id) ON DELETE CASCADE,
                        departamento_id INTEGER NOT NULL REFERENCES departamentos(id) ON DELETE CASCADE,
                        estado          TEXT NOT NULL,
                        UNIQUE (contacto_id, departamento_id, estado)
                    )
                """)
                cur.execute(
                    "CREATE INDEX IF NOT EXISTS idx_notif_reglas_depto_estado "
                    "ON notificacion_contacto_reglas(departamento_id, estado)"
                )
            except Exception as e:
                log.warning(f"No se pudo crear la tabla notificacion_contacto_reglas: {e}")
            # ══════════════════════════════════════════════════════════════
            # Columnas legacy de proveedores (para DBs antiguas)
            for col_name, col_type in [("codigo","TEXT"),("movil","TEXT"),("observaciones","TEXT"),
                                        ("contacto","TEXT"),("email","TEXT"),("telefono","TEXT")]:
                cur.execute(f"ALTER TABLE proveedores ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
            # Tabla de contactos múltiples (v9.2)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS proveedor_contactos (
                    id           SERIAL PRIMARY KEY,
                    proveedor_id INTEGER NOT NULL REFERENCES proveedores(id) ON DELETE CASCADE,
                    nombre       TEXT,
                    telefono     TEXT,
                    movil        TEXT,
                    email        TEXT,
                    es_principal INTEGER NOT NULL DEFAULT 0,
                    orden        INTEGER NOT NULL DEFAULT 0
                )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_prov_contactos ON proveedor_contactos(proveedor_id)")
            # Nuevas columnas v9.4 (para DBs existentes sin ellas)
            cur.execute("ALTER TABLE proveedor_contactos ADD COLUMN IF NOT EXISTS movil TEXT")
            # es_principal: añadir sin NOT NULL primero (seguro para tablas con filas existentes)
            cur.execute("ALTER TABLE proveedor_contactos ADD COLUMN IF NOT EXISTS es_principal INTEGER DEFAULT 0")
            cur.execute("UPDATE proveedor_contactos SET es_principal=0 WHERE es_principal IS NULL")
            # Marcar como principal el contacto de orden=0 si ninguno tiene es_principal=1
            cur.execute("""
                UPDATE proveedor_contactos SET es_principal=1
                WHERE id IN (
                    SELECT DISTINCT ON (proveedor_id) id FROM proveedor_contactos
                    WHERE proveedor_id NOT IN (SELECT proveedor_id FROM proveedor_contactos WHERE es_principal=1)
                    ORDER BY proveedor_id, orden, id
                )
            """)
            # Migrar datos legacy: si hay contacto/email/telefono/movil y no hay contactos aún
            cur.execute("""
                INSERT INTO proveedor_contactos (proveedor_id, nombre, telefono, movil, email, es_principal, orden)
                SELECT id,
                       NULLIF(TRIM(COALESCE(contacto,'')), ''),
                       NULLIF(TRIM(COALESCE(telefono,'')), ''),
                       NULLIF(TRIM(COALESCE(movil,'')), ''),
                       NULLIF(TRIM(COALESCE(email,'')), ''),
                       1,
                       0
                FROM proveedores
                WHERE NOT EXISTS (SELECT 1 FROM proveedor_contactos pc WHERE pc.proveedor_id = proveedores.id)
                  AND (TRIM(COALESCE(contacto,''))!='' OR TRIM(COALESCE(email,''))!=''
                       OR TRIM(COALESCE(telefono,''))!='' OR TRIM(COALESCE(movil,''))!='')
            """)
            # ── v12.27.4 — Correos específicos por hotel en contactos de proveedor.
            # Un contacto puede quedar "general" (sin fila aquí → se usa para
            # todos los hoteles del proveedor, comportamiento de siempre) o
            # "restringido" a uno o varios hoteles concretos — ver
            # _get_proveedor_emails_principales(proveedor_id, hotel_id).
            cur.execute("""
                CREATE TABLE IF NOT EXISTS proveedor_contacto_hoteles (
                    contacto_id INTEGER NOT NULL REFERENCES proveedor_contactos(id) ON DELETE CASCADE,
                    hotel_id    INTEGER NOT NULL REFERENCES hoteles(id) ON DELETE CASCADE,
                    PRIMARY KEY (contacto_id, hotel_id)
                )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_prov_contacto_hoteles_hotel ON proveedor_contacto_hoteles(hotel_id)")
            # ── v12.27.6 — Por defecto TODOS los hoteles marcados por contacto ──
            # A petición del usuario: en vez de "vacío = general (todos los
            # hoteles)" como estado invisible, cada contacto debe nacer con
            # TODOS los hoteles marcados explícitamente, y el admin desmarca
            # los que no le correspondan (operación inversa a como se planteó
            # en v12.27.4). Esta migración marca de una vez todos los hoteles
            # a cada contacto que a día de hoy no tenga ninguno asignado —
            # es decir, TODOS los contactos existentes, porque la función es
            # nueva y nadie ha marcado nada todavía. Es idempotente: solo
            # toca contactos sin ninguna fila en proveedor_contacto_hoteles,
            # así que no se repite en contactos ya restringidos a mano más
            # adelante.
            cur.execute("""
                INSERT INTO proveedor_contacto_hoteles (contacto_id, hotel_id)
                SELECT pc.id, h.id
                FROM proveedor_contactos pc
                CROSS JOIN hoteles h
                WHERE NOT EXISTS (
                    SELECT 1 FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id
                )
                ON CONFLICT DO NOTHING
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS pedido_adjuntos (
                    id            SERIAL PRIMARY KEY,
                    pedido_id     INTEGER NOT NULL REFERENCES pedidos(id) ON DELETE CASCADE,
                    tipo          TEXT NOT NULL,
                    nombre        TEXT NOT NULL,
                    mime_type     TEXT NOT NULL,
                    datos         BYTEA NOT NULL,
                    subido_por_id INTEGER REFERENCES usuarios(id),
                    creado_en     TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            cur.execute("CREATE INDEX IF NOT EXISTS idx_adjuntos_pedido ON pedido_adjuntos(pedido_id)")
            # ── Tokens de restablecimiento de contraseña ──────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS password_reset_tokens (
                    id         SERIAL PRIMARY KEY,
                    usuario_id INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
                    token      TEXT NOT NULL UNIQUE,
                    expira_en  TIMESTAMPTZ NOT NULL,
                    usado      INTEGER NOT NULL DEFAULT 0
                )
            """)
            # ── Columna móvil en usuarios (v9.5) ─────────────────────────────
            cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS movil TEXT")
            # ── Columna email2 en usuarios (v12.25.8) — segundo email opcional.
            # El primero (email) sigue siendo obligatorio y es el único que se
            # usa en la firma; email2, si existe, se añade como destinatario
            # adicional (BCC/CC según el correo) en todos los envíos a ese
            # usuario, pero nunca aparece en la firma.
            cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS email2 TEXT")
            # ── Columnas nombre cache en pedidos e historial (v9.9.7) ─────────
            cur.execute("ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS creado_por_nombre TEXT")
            cur.execute("ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS modificado_por_nombre TEXT")
            cur.execute("ALTER TABLE historial_estados ADD COLUMN IF NOT EXISTS usuario_nombre TEXT")
            # Rellenar cache para registros existentes (ejecución única, segura)
            cur.execute("""
                UPDATE pedidos p SET creado_por_nombre = u.nombre
                FROM usuarios u WHERE u.id = p.creado_por_id AND p.creado_por_nombre IS NULL
            """)
            cur.execute("""
                UPDATE pedidos p SET modificado_por_nombre = u.nombre
                FROM usuarios u WHERE u.id = p.modificado_por_id AND p.modificado_por_nombre IS NULL
            """)
            cur.execute("""
                UPDATE historial_estados h SET usuario_nombre = u.nombre
                FROM usuarios u WHERE u.id = h.usuario_id AND h.usuario_nombre IS NULL
            """)
            # ── Tabla asignación hoteles a usuario hotel (v9.9.5) ─────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS usuario_hoteles (
                    usuario_id INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
                    hotel_id   INTEGER NOT NULL REFERENCES hoteles(id)  ON DELETE CASCADE,
                    PRIMARY KEY (usuario_id, hotel_id)
                )
            """)
            # Permite gestionar desde admin qué hoteles atiende cada comprador,
            # reemplazando el diccionario HOTEL_COMPRADOR hardcodeado.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS usuario_comprador_hoteles (
                    usuario_id INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
                    hotel_id   INTEGER NOT NULL REFERENCES hoteles(id)  ON DELETE CASCADE,
                    PRIMARY KEY (usuario_id, hotel_id)
                )
            """)
            cur.execute("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS telegram_chat_id TEXT")
            # Añade índice único sobre hotel_id para garantizar a nivel de BD
            # que ningún hotel pueda tener más de un comprador asignado.
            # Se usa CREATE UNIQUE INDEX IF NOT EXISTS para ser idempotente.
            cur.execute("""
                CREATE UNIQUE INDEX IF NOT EXISTS uq_comprador_hotel
                ON usuario_comprador_hoteles (hotel_id)
            """)
            # ── Log de WhatsApp (v9.5) ───────────────────────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS whatsapp_log (
                    id           SERIAL PRIMARY KEY,
                    pedido_id    INTEGER REFERENCES pedidos(id),
                    tipo         TEXT NOT NULL,
                    destinatario TEXT NOT NULL,
                    mensaje      TEXT,
                    enviado      INTEGER NOT NULL DEFAULT 0,
                    error        TEXT,
                    creado_en    TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            # ── Techo de gastos (v9.0) ───────────────────────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS familias (
                    id     SERIAL PRIMARY KEY,
                    nombre TEXT NOT NULL UNIQUE,
                    activo INTEGER NOT NULL DEFAULT 1
                )
            """)
            cur.execute("ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS familia_id INTEGER REFERENCES familias(id)")
            cur.execute("ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS importe NUMERIC(10,2)")
            cur.execute("""
                DO $$ BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_name='pedidos' AND column_name='sujeto_techo'
                    ) THEN
                        ALTER TABLE pedidos ADD COLUMN sujeto_techo INTEGER NOT NULL DEFAULT 0;
                    END IF;
                END $$;
            """)
            # ── Tabla config_alertas (v10.5) ─────────────────────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS config_alertas (
                    clave  TEXT PRIMARY KEY,
                    valor  TEXT NOT NULL,
                    tipo   TEXT NOT NULL DEFAULT 'numero',
                    label  TEXT NOT NULL,
                    grupo  TEXT NOT NULL DEFAULT 'general',
                    orden  INTEGER NOT NULL DEFAULT 0
                )
            """)
            cur.execute("SELECT COUNT(*) as n FROM config_alertas")
            row = cur.fetchone()
            n = row[0] if isinstance(row, tuple) else row['n']
            if n == 0:
                defaults = [
                    ("enviado_primera",        "15", "numero", "Enviado al proveedor — 1ª alerta (días)",           "estado_enviado",    1),
                    ("enviado_urgente",         "25", "numero", "Enviado al proveedor — Urgente (días)",             "estado_enviado",    2),
                    ("enviado_ciclo",           "10", "numero", "Enviado al proveedor — Ciclo repetición (días)",    "estado_enviado",    3),
                    ("firma_compras_primera",    "8", "numero", "Firma Dir. Compras — 1ª alerta (días)",             "estado_firma",      1),
                    ("firma_compras_urgente",    "0", "numero", "Firma Dir. Compras — Urgente (días, 0=nunca)",      "estado_firma",      2),
                    ("firma_compras_ciclo",      "8", "numero", "Firma Dir. Compras — Ciclo repetición (días)",      "estado_firma",      3),
                    ("firma_hotel_primera",      "5", "numero", "Firma Dir. Hotel — 1ª alerta (días)",               "estado_firma",      4),
                    ("firma_hotel_urgente",      "0", "numero", "Firma Dir. Hotel — Urgente (días, 0=nunca)",        "estado_firma",      5),
                    ("firma_hotel_ciclo",        "5", "numero", "Firma Dir. Hotel — Ciclo repetición (días)",        "estado_firma",      6),
                    ("entrega_parcial_primera", "10", "numero", "Entrega Parcial — 1ª alerta (días)",                "estado_entrega",    1),
                    ("entrega_parcial_urgente",  "0", "numero", "Entrega Parcial — Urgente (días, 0=nunca)",         "estado_entrega",    2),
                    ("entrega_parcial_ciclo",   "10", "numero", "Entrega Parcial — Ciclo repetición (días)",         "estado_entrega",    3),
                    ("cotizacion_primera",       "2", "numero", "Pendiente Cotización — 1ª alerta (días)",           "estado_cotizacion", 1),
                    ("cotizacion_urgente",       "3", "numero", "Pendiente Cotización — Urgente (días)",             "estado_cotizacion", 2),
                    ("cotizacion_ciclo",         "3", "numero", "Pendiente Cotización — Ciclo repetición (días)",    "estado_cotizacion", 3),
                    ("dias_critico",            "60", "numero", "Días crítico global (fuerza reenvío urgente)",      "global",            1),
                    ("activar_uso_plazo_entrega","1",  "bool",   "Activar alertas basadas en plazo de entrega del proveedor", "global", 2),
                    ("plazo_aviso_dias_antes",   "5",  "numero", "Plazo entrega — Aviso previo (días antes de la entrega)",   "plazo_entrega", 1),
                    ("plazo_urgente_ciclo",       "2",  "numero", "Plazo entrega — Ciclo urgente tras vencer (cada N días)",   "plazo_entrega", 2),
                    ("plazo_parcial_aviso_dias_antes", "3", "numero", "Entrega Parcial c/plazo — Aviso previo (días antes)",   "plazo_entrega", 3),
                    ("plazo_parcial_urgente_ciclo",    "2", "numero", "Entrega Parcial c/plazo — Ciclo urgente (cada N días)", "plazo_entrega", 4),
                    ("techo_max_pedido",      "3000", "numero", "Techo — Importe máximo por pedido (€)",             "techo",             1),
                    ("techo_max_mes",         "6000", "numero", "Techo — Importe máximo mensual por hotel (€)",      "techo",             2),
                    ("techo_max_pedidos",        "2", "numero", "Techo — Nº máximo de pedidos por hotel/mes",        "techo",             3),
                    ("techo_max_pedidos_familia", "1", "numero", "Techo — Nº máximo de pedidos por hotel/mes y familia", "techo",           4),
                    ("techo_max_mes_familia",      "0", "numero", "Techo — Importe máximo mensual por hotel y familia (€) (0 = sin límite)", "techo", 5),
                    ("techo_pct_amarillo",      "60", "numero", "Techo — % consumido para alerta 🟡 amarilla (defecto 60%)",  "techo",             6),
                ]
                cur.executemany(
                    "INSERT INTO config_alertas (clave,valor,tipo,label,grupo,orden) VALUES (%s,%s,%s,%s,%s,%s)",
                    defaults
                )
            # ── Solicitudes de acceso en 2 fases (v10.5) ────────────────────
            cur.execute("""
                CREATE TABLE IF NOT EXISTS solicitudes_acceso (
                    id              SERIAL PRIMARY KEY,
                    nombre          TEXT NOT NULL,
                    apellidos       TEXT NOT NULL,
                    email           TEXT NOT NULL,
                    hoteles         TEXT NOT NULL,
                    usuario_windows TEXT,
                    token           TEXT UNIQUE,
                    estado          TEXT NOT NULL DEFAULT 'fase1_pendiente',
                    ip_solicitante  TEXT,
                    creado_en       TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    token_expira    TIMESTAMPTZ,
                    completado_en   TIMESTAMPTZ
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_solicitudes_token ON solicitudes_acceso(token)"
            )
            # Migración v11.6.7: columna movil en solicitudes_acceso
            cur.execute(
                "ALTER TABLE solicitudes_acceso ADD COLUMN IF NOT EXISTS movil TEXT"
            )
            # Migración v12.11.0: registro de envío del email de Fase 2 (para
            # poder ver en el panel admin si el EmailJS del navegador tuvo
            # éxito o no, en vez de asumirlo a ciegas).
            cur.execute(
                "ALTER TABLE solicitudes_acceso ADD COLUMN IF NOT EXISTS fase2_email_estado TEXT"
            )
            cur.execute(
                "ALTER TABLE solicitudes_acceso ADD COLUMN IF NOT EXISTS fase2_email_detalle TEXT"
            )
            cur.execute(
                "ALTER TABLE solicitudes_acceso ADD COLUMN IF NOT EXISTS fase2_email_en TIMESTAMPTZ"
            )
            # ── Tabla cola de notificaciones para el bridge agenda (v10.7.7) ──────
            # Cada fila es un aviso pendiente de entregar a un usuario concreto.
            # El bridge lo consume con GET /api/bridge/notificaciones y marca leído.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS bridge_notificaciones (
                    id           SERIAL PRIMARY KEY,
                    usuario      TEXT NOT NULL,         -- username del destinatario
                    tipo         TEXT NOT NULL,         -- 'cambio_estado' | 'alerta_auto' | 'techo' | 'familia_repetida'
                    pedido_id    INTEGER,               -- puede ser NULL (p.ej. alertas de techo)
                    titulo       TEXT NOT NULL,
                    mensaje      TEXT NOT NULL,
                    nivel        TEXT NOT NULL DEFAULT 'aviso',  -- 'aviso' | 'urgente'
                    leido        BOOLEAN NOT NULL DEFAULT FALSE,
                    creado_en    TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    visible_en   TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_bridge_notif_usuario_leido "
                "ON bridge_notificaciones(usuario, leido)"
            )
            # ── (2026-08-14) Antirrepetición de popups por cambio de estado ────
            # visible_en: hasta qué momento se retiene el aviso antes de que el
            # bridge (main_agenda) pueda recogerlo. Por defecto NOW() (visible
            # de inmediato, comportamiento de siempre). _encolar_bridge_notifi
            # cacion() la adelanta 5 minutos solo para tipo='cambio_estado' —
            # ver esa función para el porqué.
            cur.execute(
                "ALTER TABLE bridge_notificaciones ADD COLUMN IF NOT EXISTS visible_en TIMESTAMPTZ NOT NULL DEFAULT NOW()"
            )
            # ── v12.29.47 (PRUEBA) — Popup de main_agenda: entrega única persistida ──
            # Hasta ahora /api/bridge/alertas devolvía SIEMPRE los pedidos en
            # alerta activa, y era pedidos_agenda_bridge.py (Organizador
            # Princess) quien decidía si tocaba (re)mostrar el popup con un
            # intervalo en horas guardado EN MEMORIA (_estado_popups). Al
            # reiniciarse la app ese historial se perdía y el popup podía
            # reaparecer de golpe -- causa probable de los avisos repetidos
            # reportados por el comprador de INSIRE (2026-08-04).
            # Esta tabla mueve el dedup al servidor: cada fila (usuario,
            # pedido_id, nivel) significa "este popup ya se entregó a este
            # usuario" -- de forma permanente, no reseteable por el cliente.
            # Ver _filtrar_popups_no_vistos().
            cur.execute("""
                CREATE TABLE IF NOT EXISTS bridge_popup_visto (
                    id         SERIAL PRIMARY KEY,
                    usuario    TEXT NOT NULL,
                    pedido_id  INTEGER NOT NULL,
                    nivel      TEXT NOT NULL,
                    visto_en   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    UNIQUE (usuario, pedido_id, nivel)
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_bridge_popup_visto_usuario "
                "ON bridge_popup_visto(usuario)"
            )
            # ── v11.4.0 — Plazo de entrega por pedido ─────────────────────────
            cur.execute(
                "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS plazo_entrega_dias INTEGER"
            )
            # ── v12.29.4 — Fecha de entrega específica (alternativa al plazo
            # en días) — si el proveedor da un día de entrega concreto en vez
            # de "X días", se guarda directamente aquí y las alertas/
            # reclamaciones se calculan a partir de ella en vez de sumar
            # plazo_entrega_dias a fecha_tramitacion. Ver
            # _resolver_fecha_entrega_prevista().
            cur.execute(
                "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS fecha_entrega_especifica DATE"
            )
            # ── v12.29.8 — Rediseño Techo de Gastos, Fase 1 ────────────────────
            # mes_consumo_techo: se rellena SOLO cuando el pedido pasa a
            # ENVIADO AL PROVEEDOR (momento real de consumo del techo, ya no
            # al crear/editar) y se vacía si se cancela después — evita
            # recalcular sobre historial_estados en cada consulta. Formato
            # 'YYYY-MM'.
            # no_autorizado_previo: flag de integridad — TRUE si el pedido
            # llegó a ENVIADO AL PROVEEDOR sin haber pasado por una
            # autorización de Dirección General cuando debería.
            cur.execute(
                "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS mes_consumo_techo TEXT"
            )
            cur.execute(
                "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS no_autorizado_previo BOOLEAN NOT NULL DEFAULT FALSE"
            )
            # ── v12.29.24 — Rediseño Techo de Gastos, Fase 7 (backfill) ────────
            # Los pedidos que YA estaban en ENVIADO AL PROVEEDOR/ENTREGA
            # PARCIAL/ENTREGADO antes de que existiera mes_consumo_techo se
            # quedarían con esa columna vacía para siempre si no se rellena
            # una vez — y entonces desaparecerían del cálculo del techo del
            # mes en que de verdad se enviaron (tanto en el resumen del mes
            # en curso, si coincidiera con el mes actual, como sobre todo en
            # el histórico por meses). Se rellena UNA sola vez con el mismo
            # criterio de fallback que usaba el endpoint histórico antes de
            # simplificarse en la Fase 4 (COALESCE historial_estados →
            # fecha_tramitacion → creado_en), tomando el ÚLTIMO registro de
            # "pasó a ENVIADO AL PROVEEDOR" en historial_estados si existe
            # más de uno. Idempotente por el propio WHERE
            # (mes_consumo_techo IS NULL): en despliegues ya migrados esta
            # sentencia no actualiza ninguna fila y no hace nada.
            cur.execute("""
                UPDATE pedidos p
                SET mes_consumo_techo = to_char(
                    COALESCE(
                        (SELECT hs.creado_en FROM historial_estados hs
                         WHERE hs.pedido_id = p.id AND hs.estado_nuevo = 'ENVIADO AL PROVEEDOR'
                         ORDER BY hs.creado_en DESC LIMIT 1),
                        NULLIF(p.fecha_tramitacion, '')::timestamptz,
                        p.creado_en
                    ), 'YYYY-MM'
                )
                WHERE p.sujeto_techo = 1
                  AND p.mes_consumo_techo IS NULL
                  AND p.estado IN ('ENVIADO AL PROVEEDOR', 'ENTREGA PARCIAL', 'ENTREGADO')
            """)
            _backfill_techo_n = cur.rowcount
            if _backfill_techo_n:
                log.info("[MIGRACION] Backfill mes_consumo_techo (rediseño Techo Fase 7): %s pedido(s) actualizado(s)",
                         _backfill_techo_n)
            # ── v12.29.33 — FIX: tabla expediente_exceso nunca se creó en
            # producción ────────────────────────────────────────────────────
            # Mismo bug que el hotel "PR" de arriba: expediente_exceso (y sus
            # índices) solo estaban en SQL_STATEMENTS (models.py), que únicamente
            # ejecuta init_db.py a mano en el primer despliegue — nadie lo
            # vuelve a correr sobre una base de datos ya existente. Resultado:
            # /api/techo/resumen hacía SELECT sobre una tabla inexistente,
            # el backend devolvía 500, _fetchTecho() lo capturaba y devolvía
            # null, y loadTecho() petaba en `d.mes` dejando la vista
            # "Techo de gastos" colgada en "Cargando…" para siempre.
            # Se repite aquí, en _auto_migrate() (la función que sí corre en
            # cada arranque), con CREATE TABLE/INDEX IF NOT EXISTS — no
            # duplica nada si ya se llegó a ejecutar init_db.py.
            # v12.29.35 — DEBUG temporal: try/except propio para esta migración
            # concreta, con logging detallado (tipo de excepción + repr, no
            # solo str(e)) — así, si vuelve a fallar, sabremos exactamente cuál
            # de las 4 sentencias es y por qué, sin depender de adivinar por
            # la posición en el código. Quitar (o simplificar) una vez
            # confirmado que la tabla se crea correctamente.
            try:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS expediente_exceso (
                        id                              SERIAL PRIMARY KEY,
                        pedido_id                       INTEGER NOT NULL REFERENCES pedidos(id) ON DELETE CASCADE,
                        hotel_id                        INTEGER NOT NULL REFERENCES hoteles(id),
                        familia_id                      INTEGER REFERENCES familias(id),
                        mes                             TEXT NOT NULL,
                        importe_pedido                  NUMERIC(10,2),
                        consumo_previo                  NUMERIC(10,2),
                        exceso                          NUMERIC(10,2),
                        motivo_solicitud                TEXT,
                        usuario_solicitante_id          INTEGER REFERENCES usuarios(id),
                        resultado                       TEXT NOT NULL DEFAULT 'pendiente',
                        usuario_resuelve_id             INTEGER REFERENCES usuarios(id),
                        fecha_resolucion                TIMESTAMPTZ,
                        observaciones_direccion_general TEXT,
                        consumido_en_solicitud          NUMERIC(10,2),
                        disponible_en_solicitud         NUMERIC(10,2),
                        creado_en                       TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                log.info("[MIGRACION] Tabla expediente_exceso — CREATE TABLE ejecutado")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_expediente_pedido ON expediente_exceso(pedido_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_expediente_hotel_familia_mes ON expediente_exceso(hotel_id, familia_id, mes)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_expediente_resultado ON expediente_exceso(resultado)")
                log.info("[MIGRACION] Tabla expediente_exceso — índices OK")
            except Exception as _e_exp:
                log.error(f"[MIGRACION] FALLO creando expediente_exceso — tipo={type(_e_exp).__name__} repr={_e_exp!r}")
                raise
            for _clave, _valor, _tipo, _label, _grupo, _orden in [
                ('activar_uso_plazo_entrega',      '1', 'bool',   'Activar alertas basadas en plazo de entrega del proveedor', 'global',        2),
                ('plazo_aviso_dias_antes',          '5', 'numero', 'Plazo entrega — Aviso previo (días antes de la entrega)',   'plazo_entrega', 1),
                ('plazo_urgente_ciclo',             '2', 'numero', 'Plazo entrega — Ciclo urgente tras vencer (cada N días)',   'plazo_entrega', 2),
                ('plazo_parcial_aviso_dias_antes',  '3', 'numero', 'Entrega Parcial c/plazo — Aviso previo (días antes)',       'plazo_entrega', 3),
                ('plazo_parcial_urgente_ciclo',     '2', 'numero', 'Entrega Parcial c/plazo — Ciclo urgente (cada N días)',     'plazo_entrega', 4),
            ]:
                cur.execute("""
                    INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (clave) DO NOTHING
                """, (_clave, _valor, _tipo, _label, _grupo, _orden))
            # ── v12.19.0 — Reclamación automática por email al proveedor ──────
            # Cuando el plazo de entrega informado por el proveedor vence
            # (nivel_alerta == 'urgente' en _alertas_plazo_entrega) y el
            # pedido sigue en ENVIADO AL PROVEEDOR / ENTREGA PARCIAL, el job
            # diario encola automáticamente el email de reclamación al
            # proveedor (reutilizando _build_alerta_email) en vez de esperar
            # a que un usuario lo envíe manualmente desde la ficha del pedido.
            cur.execute("""
                INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (clave) DO NOTHING
            """, ('activar_reclamacion_proveedor_auto', '0', 'bool',
                  'Enviar reclamación automática por email al proveedor cuando vence el plazo',
                  'plazo_entrega', 5))
            # Columnas para que la cola de emails de sistema pueda llevar CC
            # (compradores del hotel) y quedar vinculada a un pedido concreto
            # — necesario para las reclamaciones automáticas a proveedor.
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes ADD COLUMN IF NOT EXISTS cc_emails TEXT"
            )
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes ADD COLUMN IF NOT EXISTS pedido_id INTEGER"
            )
            # ── v12.23.8 — La etiqueta se queda corta: ese mismo interruptor
            # ahora también gobierna el aviso automático al comprador en
            # Pendiente Cotización sin proveedor y en Pendiente Firma
            # Compras/Hotel — se actualiza el texto para reflejarlo (UPDATE,
            # no INSERT ON CONFLICT, porque la fila ya existe en producción).
            cur.execute("""
                UPDATE config_alertas SET label = %s WHERE clave = %s
            """, ('Enviar avisos automáticos por email (reclamación a proveedor y avisos internos) cuando corresponda',
                  'activar_reclamacion_proveedor_auto'))
            # ── v12.5.0 — Repetición de popups en Agenda por tipo de alerta ────
            # Controla, para cada estado de pedido, si el popup en Organizador
            # Princess se repite mientras el pedido siga en alerta y cada
            # cuántas horas (por separado para nivel 🔴 crítico/urgente y
            # 🟡 normal/aviso). Antes esto era fijo en código (bridge: 1h
            # urgente / 24h aviso) e igual para todos los tipos. Consumido por
            # _clasificar_alertas() → expuesto en /api/bridge/alertas → leído
            # por pedidos_agenda_bridge.py (Organizador Princess).
            for _clave, _valor, _tipo, _label, _grupo, _orden in [
                ('enviado_popup_repetir',           '1',  'bool',   'Enviado al proveedor — Repetir popup en Agenda',        'popup_repeticion', 1),
                ('enviado_popup_horas_critico',     '1',  'numero', 'Enviado al proveedor — Repetir cada (horas) si 🔴 URGENTE', 'popup_repeticion', 2),
                ('enviado_popup_horas_normal',       '24', 'numero', 'Enviado al proveedor — Repetir cada (horas) si 🟡 AVISO',   'popup_repeticion', 3),
                ('firma_compras_popup_repetir',      '1',  'bool',   'Firma Dir. Compras — Repetir popup en Agenda',           'popup_repeticion', 4),
                ('firma_compras_popup_horas_critico', '1', 'numero', 'Firma Dir. Compras — Repetir cada (horas) si 🔴 URGENTE', 'popup_repeticion', 5),
                ('firma_compras_popup_horas_normal',  '24','numero', 'Firma Dir. Compras — Repetir cada (horas) si 🟡 AVISO',   'popup_repeticion', 6),
                ('firma_hotel_popup_repetir',        '1',  'bool',   'Firma Dir. Hotel — Repetir popup en Agenda',             'popup_repeticion', 7),
                ('firma_hotel_popup_horas_critico',  '1',  'numero', 'Firma Dir. Hotel — Repetir cada (horas) si 🔴 URGENTE',   'popup_repeticion', 8),
                ('firma_hotel_popup_horas_normal',   '24', 'numero', 'Firma Dir. Hotel — Repetir cada (horas) si 🟡 AVISO',     'popup_repeticion', 9),
                ('entrega_parcial_popup_repetir',    '1',  'bool',   'Entrega Parcial — Repetir popup en Agenda',              'popup_repeticion', 10),
                ('entrega_parcial_popup_horas_critico','1','numero', 'Entrega Parcial — Repetir cada (horas) si 🔴 URGENTE',    'popup_repeticion', 11),
                ('entrega_parcial_popup_horas_normal', '24','numero','Entrega Parcial — Repetir cada (horas) si 🟡 AVISO',      'popup_repeticion', 12),
                ('cotizacion_popup_repetir',         '1',  'bool',   'Pendiente Cotización — Repetir popup en Agenda',         'popup_repeticion', 13),
                ('cotizacion_popup_horas_critico',   '1',  'numero', 'Pendiente Cotización — Repetir cada (horas) si 🔴 URGENTE','popup_repeticion', 14),
                ('cotizacion_popup_horas_normal',    '24', 'numero', 'Pendiente Cotización — Repetir cada (horas) si 🟡 AVISO', 'popup_repeticion', 15),
            ]:
                cur.execute("""
                    INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (clave) DO NOTHING
                """, (_clave, _valor, _tipo, _label, _grupo, _orden))
            # ── v12.6.0 — Reenvío a Admins (antes hardcodeado a "cada 2 días") ─
            # Ambos jobs (techo urgente y familia/partida repetida) esperaban un
            # número fijo de días naturales entre avisos repetidos al mismo
            # hotel para el rol admin. Se convierte en config editable.
            for _clave, _valor, _tipo, _label, _grupo, _orden in [
                ('techo_urgente_admin_reenvio_dias',    '2', 'numero', 'Techo urgente — Reenvío a Admins (cada N días)',            'reenvio_admin', 1),
                ('familia_repetida_admin_reenvio_dias', '2', 'numero', 'Familia/partida repetida — Reenvío a Admins (cada N días)', 'reenvio_admin', 2),
            ]:
                cur.execute("""
                    INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (clave) DO NOTHING
                """, (_clave, _valor, _tipo, _label, _grupo, _orden))
            # ── v12.28.0 — Techo por familia configurable (antes fijo a 1) ─────
            # Hasta ahora una familia de artículos solo podía usarse UNA vez al
            # mes por hotel (regla fija en código). A petición del usuario se
            # convierte en un límite editable: Nº máximo de pedidos por
            # hotel/mes Y familia (no solo el total de pedidos por hotel/mes,
            # que ya existía en 'techo_max_pedidos'). Se inicializa a 1 para
            # no cambiar el comportamiento actual en producción hasta que un
            # admin lo modifique desde Config alertas → Techo de gastos.
            cur.execute("""
                INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (clave) DO NOTHING
            """, ('techo_max_pedidos_familia', '1', 'numero',
                  'Techo — Nº máximo de pedidos por hotel/mes y familia', 'techo', 4))
            # ── v12.29.0 — Techo de importe (€) también por hotel/mes y familia ──
            # Complementa a techo_max_pedidos_familia (que limita el Nº de
            # pedidos): ahora también se puede limitar el IMPORTE acumulado
            # de una familia concreta en el mes, igual que techo_max_mes ya
            # limitaba el acumulado del hotel entero. Por defecto 0 = sin
            # límite (no cambia nada en producción hasta que un admin ponga
            # un valor > 0 desde Config alertas → Techo de gastos).
            cur.execute("""
                INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (clave) DO NOTHING
            """, ('techo_max_mes_familia', '0', 'numero',
                  'Techo — Importe máximo mensual por hotel y familia (€) (0 = sin límite)', 'techo', 5))
            # ── v12.23.6 — Reclamación automática también para Pendiente Cotización ──
            # A petición del usuario, se extiende la reclamación automática al
            # proveedor (activar_reclamacion_proveedor_auto) al estado
            # PENDIENTE COTIZACIÓN, igual que ya funcionaba para ENVIADO AL
            # PROVEEDOR y ENTREGA PARCIAL. Antes ese estado no tenía "ciclo"
            # configurado (solo se avisaba una vez al hacerse urgente); se
            # añade cotizacion_ciclo para que, con la reclamación automática
            # activada, se repita cada N días mientras siga sin cotización.
            cur.execute("""
                INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (clave) DO NOTHING
            """, ('cotizacion_ciclo', '3', 'numero',
                  'Pendiente Cotización — Ciclo repetición (días)', 'estado_cotizacion', 3))
            # ── v12.27.8 — Backup automático de cuenta EmailJS ────────────────
            # A raíz de haberse quedado sin cuota EmailJS a mitad de mes: se
            # guardan 2 juegos de credenciales (cuenta 1 = activa por
            # defecto, cuenta 2 = backup) y un contador de envíos. Al superar
            # el umbral (195 por defecto, 5 antes del límite gratuito de
            # 200/mes) el sistema cambia solo a la cuenta 2 sin necesidad de
            # desplegar nada — ver /api/emailjs/config y
            # /api/emailjs/registrar-envio. Los valores de la cuenta 1 se
            # inicializan con las credenciales ya en uso en producción para
            # no cortar nada en el primer deploy de esta versión.
            #
            # (2026-08-12) 3ª cuenta EmailJS, a petición del usuario: cuenta 1
            # (principal) → cuenta 2 (secundaria) → cuenta 3 (backup) → vuelta
            # a la cuenta 1, en ciclo indefinido — ver el nuevo
            # _EMAILJS_MAX_CUENTAS y la lógica cíclica de
            # /api/emailjs/registrar-envio. `ON CONFLICT DO NOTHING` en todas
            # estas filas: si ya existían de la v12.27.8 (2 cuentas), esta
            # migración solo añade las 3 nuevas claves de la cuenta 3 sin
            # tocar nada de lo ya configurado.
            #
            # (2026-09-01) 4ª cuenta EmailJS, misma petición de siempre
            # (más margen de backup): ciclo ahora 1 (principal) → 2
            # (secundaria) → 3 (terciaria) → 4 (backup) → vuelta a la 1.
            # Igual que la vez anterior, solo hace falta subir
            # _EMAILJS_MAX_CUENTAS (la lógica cíclica ya estaba escrita en
            # función de esa constante desde que se generalizó de 2 a 3) y
            # añadir aquí las 4 claves nuevas de la cuenta 4 — `ON CONFLICT
            # DO NOTHING` de nuevo, no toca las 3 cuentas ya configuradas.
            _emailjs_defaults = [
                ('emailjs_public_key_1',   'bxFzHypsIrNqcDh15',  'texto',  'Cuenta 1 (principal) — Public Key',    'emailjs', 1),
                ('emailjs_service_id_1',   'service_shvrzuv',    'texto',  'Cuenta 1 (principal) — Service ID',    'emailjs', 2),
                ('emailjs_template_id_1',  'template_1zrv4ze',   'texto',  'Cuenta 1 (principal) — Template ID',   'emailjs', 3),
                ('emailjs_public_key_2',   '',                   'texto',  'Cuenta 2 (secundaria) — Public Key',   'emailjs', 4),
                ('emailjs_service_id_2',   '',                   'texto',  'Cuenta 2 (secundaria) — Service ID',   'emailjs', 5),
                ('emailjs_template_id_2',  '',                   'texto',  'Cuenta 2 (secundaria) — Template ID',  'emailjs', 6),
                ('emailjs_public_key_3',   '',                   'texto',  'Cuenta 3 (terciaria) — Public Key',    'emailjs', 7),
                ('emailjs_service_id_3',   '',                   'texto',  'Cuenta 3 (terciaria) — Service ID',    'emailjs', 8),
                ('emailjs_template_id_3',  '',                   'texto',  'Cuenta 3 (terciaria) — Template ID',   'emailjs', 9),
                ('emailjs_cuenta_activa',  '1',                  'numero', 'Cuenta actualmente en uso (1, 2, 3 o 4)', 'emailjs', 10),
                ('emailjs_contador',       '0',                  'numero', 'Envíos contabilizados este ciclo',     'emailjs', 11),
                ('emailjs_umbral_cambio',  '195',                'numero', 'Cambiar de cuenta al llegar a',        'emailjs', 12),
                ('emailjs_cambio_automatico_en', '',             'texto',  'Último cambio automático (fecha)',     'emailjs', 13),
                # (2026-08-19) A petición del usuario: campo informativo (sin
                # ningún uso automático, solo para consulta desde el propio
                # panel) con la fecha en la que cada cuenta EmailJS recupera
                # su cupo mensual de 200 envíos — visible en el panel de cada
                # cuenta en EmailJS.com ("Resets on ..."), pero solo se ve
                # entrando a cada cuenta por separado; el admin la copia aquí
                # a mano para tenerlas todas controladas de un vistazo (desde
                # v12.30.92 esta fecha además se avanza sola +30 días en
                # cuanto se cumple, ver _job_avanzar_reinicio_emailjs).
                ('emailjs_reinicio_fecha_1', '',                 'fecha',  'Cuenta 1 (principal) — Reinicia cupo el',  'emailjs', 14),
                ('emailjs_reinicio_fecha_2', '',                 'fecha',  'Cuenta 2 (secundaria) — Reinicia cupo el', 'emailjs', 15),
                ('emailjs_reinicio_fecha_3', '',                 'fecha',  'Cuenta 3 (terciaria) — Reinicia cupo el',  'emailjs', 16),
                # (2026-09-01) Cuenta 4 (backup) — mismo patrón que las otras 3.
                ('emailjs_public_key_4',    '',                  'texto',  'Cuenta 4 (backup) — Public Key',       'emailjs', 17),
                ('emailjs_service_id_4',    '',                  'texto',  'Cuenta 4 (backup) — Service ID',       'emailjs', 18),
                ('emailjs_template_id_4',   '',                  'texto',  'Cuenta 4 (backup) — Template ID',      'emailjs', 19),
                ('emailjs_reinicio_fecha_4', '',                 'fecha',  'Cuenta 4 (backup) — Reinicia cupo el',     'emailjs', 20),
            ]
            for _clave, _valor, _tipo, _label, _grupo, _orden in _emailjs_defaults:
                cur.execute("""
                    INSERT INTO config_alertas (clave, valor, tipo, label, grupo, orden)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (clave) DO NOTHING
                """, (_clave, _valor, _tipo, _label, _grupo, _orden))
            # ── v11.9.0 — Cola de restauración de backups (Opción C) ──────────
            # El panel web inserta filas aquí; un agente local (restore_agent.py)
            # ejecutado en el PC con acceso a la carpeta de red las procesa.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS restore_queue (
                    id                  SERIAL PRIMARY KEY,
                    backup_nombre       TEXT NOT NULL,
                    backup_ruta         TEXT NOT NULL,
                    modo                TEXT NOT NULL DEFAULT 'pedidos',
                    estado              TEXT NOT NULL DEFAULT 'pendiente',
                    solicitado_por      TEXT,
                    solicitado_en       TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    iniciado_en         TIMESTAMPTZ,
                    completado_en       TIMESTAMPTZ,
                    resumen             JSONB,
                    error_msg           TEXT,
                    pre_restore_backup  TEXT
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_restore_queue_estado ON restore_queue(estado)"
            )
            # Columna añadida en v11.9.1 — backups existentes de v11.9.0 la reciben aquí
            cur.execute(
                "ALTER TABLE restore_queue ADD COLUMN IF NOT EXISTS pre_restore_backup TEXT"
            )
            # ── Fix v11.8.6 — Listado de backups vía caché del agente local ───
            # /api/admin/backup/listar intentaba leer Path(ruta) directamente
            # en el servidor (Render), que no tiene acceso a la red local de
            # la oficina — el mismo problema ya resuelto para /restaurar con
            # la cola restore_queue. Ahora restore_agent.py escanea la carpeta
            # de backups en cada ciclo y sincroniza el resultado aquí; el
            # panel web solo lee esta tabla, nunca toca el filesystem.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS backups_cache (
                    id                SERIAL PRIMARY KEY,
                    ruta              TEXT NOT NULL,
                    ruta_normalizada  TEXT NOT NULL,
                    nombre            TEXT NOT NULL,
                    fecha             TEXT NOT NULL,
                    fecha_raw         TIMESTAMP,
                    mb                NUMERIC NOT NULL DEFAULT 0,
                    adjuntos          INTEGER NOT NULL DEFAULT 0,
                    tiene_log         BOOLEAN NOT NULL DEFAULT FALSE,
                    log_contenido     TEXT,
                    valido            BOOLEAN NOT NULL DEFAULT FALSE,
                    tipo              TEXT NOT NULL DEFAULT 'diario',
                    actualizado_en    TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    UNIQUE (ruta_normalizada, nombre)
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_backups_cache_ruta ON backups_cache(ruta_normalizada)"
            )

            # ── v11.9.4 — Columna es_correo en pedido_adjuntos ──────────────
            # Antes, distinguir un correo (.eml/.msg) de un documento normal
            # se hacía mirando la extensión del nombre del archivo guardado
            # (nombre ILIKE '%.eml'), tanto en las validaciones de subida como
            # en las de cambio de estado. Funciona, pero es un acoplamiento
            # frágil: si el nombre se llega a guardar de otra forma en algún
            # punto futuro, la clasificación se rompe en silencio sin que
            # salte ningún error. Pasamos a una columna explícita.
            cur.execute(
                "ALTER TABLE pedido_adjuntos ADD COLUMN IF NOT EXISTS es_correo BOOLEAN"
            )
            # Backfill: rellenar la columna para adjuntos ya existentes,
            # usando la misma heurística de extensión que se usaba antes
            # (es la única información disponible para datos ya guardados).
            # A partir de aquí, todo adjunto nuevo se inserta con el valor
            # ya calculado en el momento de la subida, sin volver a inferir.
            cur.execute("""
                UPDATE pedido_adjuntos
                SET es_correo = (
                    LOWER(nombre) LIKE '%.eml' OR LOWER(nombre) LIKE '%.msg'
                )
                WHERE es_correo IS NULL
            """)
            cur.execute(
                "ALTER TABLE pedido_adjuntos ALTER COLUMN es_correo SET DEFAULT FALSE"
            )
            cur.execute(
                "ALTER TABLE pedido_adjuntos ALTER COLUMN es_correo SET NOT NULL"
            )
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_adjuntos_tipo_correo ON pedido_adjuntos(pedido_id, tipo, es_correo)"
            )
            # ── v12.0.5 — Tarifa acordada (pedido sin presupuesto) ──────────
            # Permite marcar un pedido como "tarifa acordada" para eximirlo
            # de la obligatoriedad de Nº Presupuesto + documento adjunto al
            # pasar a ENVIADO AL PROVEEDOR. Por defecto siempre desmarcado.
            cur.execute(
                "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS tarifa_acordada BOOLEAN NOT NULL DEFAULT FALSE"
            )
            # ── Fix egress — miniaturas para imagen_articulo (Jul 2026) ────
            # Las imágenes de artículo (hasta 2 MB) se mostraban a tamaño
            # completo como miniatura en la lista de adjuntos, disparando
            # el egress mensual del proyecto. Ahora se guarda una versión
            # reducida (datos_thumb) generada al subir el archivo; las ya
            # existentes se generan la primera vez que se piden (lazy) y
            # quedan cacheadas en esta misma columna para siempre.
            cur.execute(
                "ALTER TABLE pedido_adjuntos ADD COLUMN IF NOT EXISTS datos_thumb BYTEA"
            )
            cur.execute(
                "ALTER TABLE pedido_adjuntos ADD COLUMN IF NOT EXISTS thumb_mime_type TEXT"
            )
            # ── Migración de adjuntos cerrados a Supabase Storage (v12.8.0) ──
            # `datos` deja de ser NOT NULL: un adjunto migrado tiene
            # `storage_path` con la ruta en Storage y `datos = NULL` (libera
            # el TOAST). `datos_thumb` NO se toca — las miniaturas se quedan
            # siempre en Postgres, son pequeñas y así la vista previa sigue
            # siendo instantánea aunque el archivo original esté en Storage.
            cur.execute(
                "ALTER TABLE pedido_adjuntos ADD COLUMN IF NOT EXISTS storage_path TEXT"
            )
            cur.execute(
                "ALTER TABLE pedido_adjuntos ALTER COLUMN datos DROP NOT NULL"
            )
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_adjuntos_storage_path "
                "ON pedido_adjuntos(storage_path) WHERE storage_path IS NOT NULL"
            )
            # ── Alerta de egress a admins por Telegram (Jul 2026) ───────────
            # Tabla de acumulado diario de bytes servidos por la app, usada
            # para estimar cuánto egress llevamos consumido en el ciclo de
            # facturación actual y avisar a los admins antes de volver a
            # acercarnos al límite del plan Free de Supabase.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS egress_tracking (
                    fecha  DATE PRIMARY KEY,
                    bytes  BIGINT NOT NULL DEFAULT 0
                )
            """)
            # ── Seguimiento de tamaño de base de datos (Jul 2026) ───────────
            # A diferencia del egress, el tamaño de la BD solo crece — no hay
            # "caché" que compense. `pedido_adjuntos` es, con diferencia, la
            # mayor consumidora (archivos en TOAST). Snapshot diario para ver
            # la tendencia sin tener que entrar al dashboard de Supabase.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS db_size_tracking (
                    fecha           DATE PRIMARY KEY,
                    bytes_total     BIGINT NOT NULL DEFAULT 0,
                    bytes_adjuntos  BIGINT NOT NULL DEFAULT 0
                )
            """)
            # Historial de compactaciones (VACUUM FULL) de pedido_adjuntos —
            # v12.8.1. Solo se inserta una fila cuando de verdad se ejecuta
            # (ver _vacuum_full_adjuntos), no en cada intento.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS db_vacuum_log (
                    id            SERIAL PRIMARY KEY,
                    fecha         TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    mb_antes      NUMERIC,
                    mb_despues    NUMERIC,
                    mb_liberados  NUMERIC
                )
            """)
            # ── Seguridad de sesión: caducidad diaria + verificación por
            # email tras varios días de inactividad (Jul 2026) ─────────────
            # Los usuarios suelen dejar la app abierta todo el día en el
            # ordenador de la oficina, así que la sesión de Flask nunca
            # llegaba a expirar de forma natural. Ahora se obliga a volver
            # a introducir la contraseña cada vez que cambia el día, y si
            # han pasado varios días sin actividad, además se exige un
            # código enviado al email antes de completar el login.
            cur.execute(
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS ultimo_login TIMESTAMPTZ"
            )
            cur.execute("""
                CREATE TABLE IF NOT EXISTS login_verification_codes (
                    id          SERIAL PRIMARY KEY,
                    usuario_id  INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
                    codigo      TEXT NOT NULL,
                    expira_en   TIMESTAMPTZ NOT NULL,
                    usado       INTEGER NOT NULL DEFAULT 0,
                    creado_en   TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            # ── Configuración de Avisos (v12.4.0) ────────────────────────────
            # Sustituye la lógica hardcodeada de qué administradores reciben
            # cada tipo de alerta de sistema (antes: TIPOS_SUPERVISION_ADMIN +
            # "todos los admins con telegram_chat_id" para cualquier evento).
            # Ahora cada "evento/causa" tiene una lista configurable de
            # usuarios y, por usuario, qué canal(es) recibe. Se gestiona desde
            # Administrador → Configuración de Avisos, y se consulta en
            # tiempo real tanto aquí como desde main_agenda / otros módulos
            # vía GET /api/config-avisos/resolver.
            cur.execute("""
                CREATE TABLE IF NOT EXISTS eventos_aviso (
                    codigo       TEXT PRIMARY KEY,
                    nombre       TEXT NOT NULL,
                    descripcion  TEXT,
                    orden        INTEGER NOT NULL DEFAULT 0
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS config_avisos (
                    id            SERIAL PRIMARY KEY,
                    evento_codigo TEXT NOT NULL REFERENCES eventos_aviso(codigo) ON DELETE CASCADE,
                    usuario_id    INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
                    telegram      BOOLEAN NOT NULL DEFAULT FALSE,
                    email         BOOLEAN NOT NULL DEFAULT FALSE,
                    UNIQUE(evento_codigo, usuario_id)
                )
            """)
            # Cola de emails de sistema (avisos de evento, no ligados a un
            # pedido concreto) — se generan desde jobs sin navegador abierto
            # (APScheduler), así que no pueden enviarse vía EmailJS en el
            # momento. Quedan aquí pendientes y cualquier admin que abra la
            # app los envía en segundo plano (mismo patrón que el resto de
            # emails de la aplicación, que dependen de EmailJS en el navegador).
            cur.execute("""
                CREATE TABLE IF NOT EXISTS emails_sistema_pendientes (
                    id             SERIAL PRIMARY KEY,
                    evento_codigo  TEXT NOT NULL,
                    destinatario   TEXT NOT NULL,
                    asunto         TEXT NOT NULL,
                    cuerpo_html    TEXT,
                    cuerpo_text    TEXT,
                    enviado        BOOLEAN NOT NULL DEFAULT FALSE,
                    creado_en      TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    enviado_en     TIMESTAMPTZ,
                    visible_en     TIMESTAMPTZ NOT NULL DEFAULT NOW()
                )
            """)
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_emails_sistema_pendientes_estado "
                "ON emails_sistema_pendientes(enviado)"
            )
            # Migración v12.11.0: vincular una fila de la cola a una solicitud
            # de acceso concreta (para el badge del panel admin) y registrar
            # cuándo se mandó el último recordatorio por Telegram de "abre la
            # app para que se despache" (para no reenviarlo cada minuto).
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes "
                "ADD COLUMN IF NOT EXISTS solicitud_acceso_id INTEGER"
            )
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes "
                "ADD COLUMN IF NOT EXISTS recordado_en TIMESTAMPTZ"
            )
            # Migración v12.29.96: reserva atómica anti-duplicados. Antes
            # GET /api/emails-sistema-pendientes devolvía las filas
            # enviado=FALSE sin marcar nada, así que si dos pestañas/
            # sesiones (o un poll + una recarga de página) pedían la cola
            # casi a la vez, ambas veían la misma fila pendiente y ambas
            # la enviaban de verdad por EmailJS antes de que ninguna la
            # marcara como enviada — duplicados reales al destinatario
            # (reportado por el usuario: pedido 39909, 2 correos idénticos).
            # `en_proceso_desde` guarda cuándo se "reservó" una fila para
            # que la pida solo una sesión a la vez (ver claim atómico en
            # api_emails_sistema_pendientes); si esa sesión nunca confirma
            # el envío (fallo de EmailJS, pestaña cerrada a media faena...),
            # la reserva caduca sola a los 2 minutos y otra sesión puede
            # reintentarla.
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes "
                "ADD COLUMN IF NOT EXISTS en_proceso_desde TIMESTAMPTZ"
            )
            # Migración (2026-08-14): retraso de 5 min para los correos de
            # cambio de estado de pedido (evento_codigo 'cambio_estado_
            # proveedor' / 'cambio_estado_interno'), que a partir de ahora
            # también pasan por esta cola en vez de enviarse de inmediato
            # desde el navegador que hizo el cambio — ver
            # _encolar_email_pedido_retrasado(). Por defecto NOW(): el resto
            # de eventos de esta cola (techo urgente, familias repetidas,
            # solicitudes de acceso...) siguen siendo inmediatos, sin cambio
            # de comportamiento.
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes "
                "ADD COLUMN IF NOT EXISTS visible_en TIMESTAMPTZ NOT NULL DEFAULT NOW()"
            )
            # Migración (2026-08-19): freno de reintentos infinitos — hasta
            # ahora, si un correo encolado fallaba SIEMPRE al enviarse (p.
            # ej. porque EmailJS lo rechazaba por tamaño con 413, algo que
            # de hecho pasó con correos de "Comparar Pedidos + Albaranes"
            # generados ANTES de acotar su tamaño, ver v12.30.15/20), la
            # reserva de la fila caducaba cada 2 minutos y CUALQUIER sesión
            # abierta la reintentaba sin límite — cada intento, fallase o
            # no, descontaba cupo de EmailJS igualmente (reportado por
            # Víctor: el contador subió de 54 a 71 sin que llegara ningún
            # correo nuevo, porque una fila ya encolada de antes seguía
            # reintentándose sola). `intentos` cuenta los intentos de envío
            # de cada fila; a partir de MAX_INTENTOS_EMAIL_SISTEMA
            # (constante más abajo) deja de recogerse — se para sola, en
            # vez de sangrar cupo para siempre. `descartado_en` permite
            # además descartar una fila a mano desde el panel de admin
            # (ver /api/admin/emails-sistema-atascados).
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes "
                "ADD COLUMN IF NOT EXISTS intentos INTEGER NOT NULL DEFAULT 0"
            )
            cur.execute(
                "ALTER TABLE emails_sistema_pendientes "
                "ADD COLUMN IF NOT EXISTS descartado_en TIMESTAMPTZ"
            )
            cur.execute("SELECT COUNT(*) as n FROM eventos_aviso")
            _row_ev = cur.fetchone()
            _n_ev = _row_ev[0] if isinstance(_row_ev, tuple) else _row_ev['n']
            if _n_ev == 0:
                _eventos_default = [
                    ("cambio_estado_supervision", "Cambio de estado con alerta urgente",
                     "Copia de supervisión cuando un cambio de estado de pedido genera una alerta urgente.", 1),
                    ("pedido_urgente_admin", "Pedido crítico parado (job diario)",
                     "Copia de supervisión cuando el job diario detecta pedidos en nivel urgente sin resolver.", 2),
                    ("techo_urgente_admin", "Techo de gastos superado (100%)",
                     "Aviso cuando un hotel supera el techo mensual de gasto o de número de pedidos.", 3),
                    ("techo_nuevo_pedido_admin", "Nuevo pedido sujeto a techo",
                     "Aviso informativo cada vez que se crea un pedido que computa contra el techo de un hotel.", 4),
                    ("familia_repetida_admin", "Familias de artículos repetidas",
                     "Aviso cuando se detectan pedidos repetidos de la misma familia en un hotel/mes.", 5),
                    ("egress_alerta", "Consumo de egress (Supabase) elevado",
                     "Aviso del job diario cuando el consumo de egress se acerca al límite del plan.", 6),
                    ("health_check", "Fallo de integridad operativa",
                     "Aviso del job diario de salud del sistema cuando detecta una incidencia.", 7),
                    ("solicitud_acceso", "Nueva solicitud de acceso",
                     "Aviso cuando un usuario nuevo solicita acceso a la aplicación (Fase 1 / Fase 2).", 8),
                ]
                cur.executemany(
                    "INSERT INTO eventos_aviso (codigo, nombre, descripcion, orden) VALUES (%s,%s,%s,%s)",
                    _eventos_default
                )
            else:
                # El bloque de arriba solo siembra si la tabla está vacía
                # (primera instalación) — en un sistema ya en marcha, como
                # este, el texto del evento "egress_alerta" se queda con el
                # de cuando se creó. Se refresca aparte, aquí, porque desde
                # Jul 2026 este evento ya no es solo egress: el job de las
                # 08:30 avisa también de tamaño de BD en el mismo mensaje.
                cur.execute("""
                    UPDATE eventos_aviso
                    SET nombre = 'Consumo Supabase elevado (egress / tamaño BD)',
                        descripcion = 'Aviso del job diario (08:30) cuando el consumo de egress o el tamaño de la base de datos se acercan al límite del plan — un único mensaje si cualquiera de las dos supera el umbral.'
                    WHERE codigo = 'egress_alerta'
                """)

            # ── Configuración de Avisos v2 (v12.17.0) — panel unificado ──────
            # Amplía el modelo anterior (solo eventos globales de supervisión
            # admin) para cubrir también los avisos operativos ligados a un
            # hotel concreto (cambio de estado, alertas de pedido parado),
            # que antes salían de _get_compradores_hotel/_get_usuarios_hotel_
            # rol_telegram (tablas usuario_comprador_hoteles/usuario_hoteles),
            # sin ningún control desde el panel de admin.
            #
            # requiere_hotel=TRUE marca los eventos cuya lista de destinatarios
            # depende del hotel del pedido (se guarda una fila por hotel en
            # notificaciones_config); el resto son eventos globales (hotel_id
            # NULL), igual que antes.
            cur.execute(
                "ALTER TABLE eventos_aviso ADD COLUMN IF NOT EXISTS requiere_hotel BOOLEAN NOT NULL DEFAULT FALSE"
            )
            # Alta idempotente de los dos eventos operativos nuevos — con
            # ON CONFLICT DO NOTHING para no pisar el nombre/descripción si
            # un admin ya los hubiera editado a mano en una ejecución previa.
            cur.execute("""
                INSERT INTO eventos_aviso (codigo, nombre, descripcion, orden, requiere_hotel)
                VALUES
                    ('cambio_estado_pedido', 'Cambio de estado de pedido (aviso normal)',
                     'Aviso por Telegram/popup a los usuarios de un hotel cuando cambia el estado de uno de sus pedidos (enviado al proveedor, entrega, cancelación...). Antes fijo por código; ahora configurable por hotel.',
                     0, TRUE),
                    ('alerta_pedido_hotel', 'Alerta de pedido pendiente (aviso al hotel)',
                     'Aviso/urgente al hotel cuando uno de sus pedidos lleva demasiado tiempo sin avanzar — incluye el job diario y el botón «Re-notificar» de la vista Alertas. Antes fijo por código; ahora configurable por hotel.',
                     0, TRUE),
                    ('techo_mensual_comprador', 'Techo de gastos mensual (aviso al hotel)',
                     'Aviso/urgente al hotel cuando su techo mensual de gasto o de nº de pedidos llega al umbral configurado (job diario). Antes fijo por código; ahora configurable por hotel.',
                     0, TRUE),
                    ('techo_nuevo_pedido_comprador', 'Nuevo pedido sujeto a techo (aviso al hotel)',
                     'Aviso informativo al hotel, en el momento de crear un pedido que computa contra su techo de gastos. Antes fijo por código; ahora configurable por hotel.',
                     0, TRUE),
                    ('familia_repetida_comprador', 'Familia de artículos repetida (aviso al hotel)',
                     'Aviso al hotel cuando se detectan pedidos repetidos de la misma familia en su hotel/mes. Antes fijo por código; ahora configurable por hotel.',
                     0, TRUE)
                ON CONFLICT (codigo) DO NOTHING
            """)
            # Los eventos globales ya existentes (creados antes de esta
            # versión) se marcan explícitamente requiere_hotel=FALSE — el
            # DEFAULT ya cubre altas nuevas, este UPDATE es solo para dejar
            # constancia explícita en instalaciones que vinieran de antes.
            cur.execute("""
                UPDATE eventos_aviso SET requiere_hotel = FALSE
                WHERE codigo IN (
                    'cambio_estado_supervision','pedido_urgente_admin',
                    'techo_urgente_admin','techo_nuevo_pedido_admin',
                    'familia_repetida_admin','egress_alerta','health_check',
                    'solicitud_acceso'
                )
            """)

            cur.execute("""
                CREATE TABLE IF NOT EXISTS notificaciones_config (
                    id            SERIAL PRIMARY KEY,
                    evento_codigo TEXT NOT NULL REFERENCES eventos_aviso(codigo) ON DELETE CASCADE,
                    hotel_id      INTEGER REFERENCES hoteles(id) ON DELETE CASCADE,
                    usuario_id    INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,
                    telegram      BOOLEAN NOT NULL DEFAULT FALSE,
                    email         BOOLEAN NOT NULL DEFAULT FALSE,
                    popup         BOOLEAN NOT NULL DEFAULT FALSE
                )
            """)
            # Sin UNIQUE de tabla a propósito: en Postgres dos NULL nunca
            # violan un UNIQUE (cada NULL cuenta como distinto), lo que
            # rompería la deduplicación de los eventos globales (hotel_id
            # NULL). En vez de un índice de expresión, el guardado hace
            # DELETE + INSERT dentro de la misma transacción (ver
            # api_save_config_avisos) — evita el problema sin depender de
            # sintaxis de índice más frágil.
            cur.execute(
                "CREATE INDEX IF NOT EXISTS idx_notif_config_lookup "
                "ON notificaciones_config (evento_codigo, hotel_id)"
            )

            # ── Semilla única: migrar el modelo hardcodeado actual ───────────
            # Solo se ejecuta si la tabla está vacía (primer despliegue de esta
            # versión) — así el día 1 nadie deja de recibir nada: se copia tal
            # cual quién recibía qué antes, y a partir de ahora ya es editable
            # desde Administrador → Configuración de Avisos sin tocar código.
            cur.execute("SELECT COUNT(*) FROM notificaciones_config")
            if cur.fetchone()[0] == 0:
                log.info("[NOTIF-CONFIG] Sembrando notificaciones_config desde el modelo anterior…")
                # 'cambio_estado_pedido' — antes: compradores del hotel (Telegram)
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, popup)
                    SELECT 'cambio_estado_pedido', uch.hotel_id, uch.usuario_id, TRUE, TRUE
                    FROM usuario_comprador_hoteles uch
                    JOIN usuarios u ON u.id = uch.usuario_id AND u.activo = 1 AND u.rol = 'compras'
                """)
                # 'cambio_estado_pedido' — antes: usuarios rol hotel del hotel
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, popup)
                    SELECT 'cambio_estado_pedido', uh.hotel_id, uh.usuario_id, TRUE, TRUE
                    FROM usuario_hoteles uh
                    JOIN usuarios u ON u.id = uh.usuario_id AND u.activo = 1 AND u.rol = 'hotel'
                """)
                # 'alerta_pedido_hotel' — antes: compradores del hotel (Telegram)
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, popup)
                    SELECT 'alerta_pedido_hotel', uch.hotel_id, uch.usuario_id, TRUE, TRUE
                    FROM usuario_comprador_hoteles uch
                    JOIN usuarios u ON u.id = uch.usuario_id AND u.activo = 1 AND u.rol = 'compras'
                """)
                # Eventos globales existentes — migrar tal cual desde config_avisos
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, email, popup)
                    SELECT evento_codigo, NULL, usuario_id, telegram, email, telegram
                    FROM config_avisos
                """)
                db.commit()
                log.info("[NOTIF-CONFIG] Semilla completada.")

            # ── Semilla fase 2 (v12.17.1) — techo + familia repetida ─────────
            # Gate independiente del de arriba a propósito: si esta instalación
            # ya desplegó la v12.17.0 (fase 1), la tabla notificaciones_config
            # ya NO está vacía, así que el "if _n == 0" de arriba no volvería a
            # ejecutarse — y estos tres eventos nuevos se quedarían sin
            # sembrar. Se comprueba por código de evento, no por tabla vacía.
            cur.execute(
                "SELECT COUNT(*) FROM notificaciones_config WHERE evento_codigo IN "
                "('techo_mensual_comprador','techo_nuevo_pedido_comprador','familia_repetida_comprador')"
            )
            if cur.fetchone()[0] == 0:
                log.info("[NOTIF-CONFIG] Sembrando fase 2 (techo + familia repetida) desde el modelo anterior…")
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, popup)
                    SELECT 'techo_mensual_comprador', uch.hotel_id, uch.usuario_id, TRUE, TRUE
                    FROM usuario_comprador_hoteles uch
                    JOIN usuarios u ON u.id = uch.usuario_id AND u.activo = 1 AND u.rol = 'compras'
                """)
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, popup)
                    SELECT 'techo_nuevo_pedido_comprador', uch.hotel_id, uch.usuario_id, TRUE, TRUE
                    FROM usuario_comprador_hoteles uch
                    JOIN usuarios u ON u.id = uch.usuario_id AND u.activo = 1 AND u.rol = 'compras'
                """)
                cur.execute("""
                    INSERT INTO notificaciones_config (evento_codigo, hotel_id, usuario_id, telegram, popup)
                    SELECT 'familia_repetida_comprador', uch.hotel_id, uch.usuario_id, TRUE, TRUE
                    FROM usuario_comprador_hoteles uch
                    JOIN usuarios u ON u.id = uch.usuario_id AND u.activo = 1 AND u.rol = 'compras'
                """)
                db.commit()
                log.info("[NOTIF-CONFIG] Semilla fase 2 completada.")

            # ── Saneado (v12.17.2) — datos "contaminados" por la UI confusa ──
            # Antes de este arreglo, la matriz mezclaba eventos globales y por
            # hotel bajo el mismo selector sin separación visual, y era fácil
            # guardar sin querer un hotel_id real en un evento global (p.ej.
            # "Cambio de estado con alerta urgente"). Se borran esas filas mal
            # etiquetadas en cada arranque — es un DELETE con WHERE, así que no
            # pasa nada si ya estaba limpio (0 filas afectadas y listo).
            cur.execute("""
                DELETE FROM notificaciones_config nc
                USING eventos_aviso ea
                WHERE nc.evento_codigo = ea.codigo
                  AND ea.requiere_hotel = FALSE
                  AND nc.hotel_id IS NOT NULL
            """)
            if cur.rowcount:
                log.warning("[NOTIF-CONFIG] Saneado: %d fila(s) con hotel_id indebido en eventos globales, eliminadas.", cur.rowcount)
                db.commit()

            # ── Dashboard configurable por usuario (v12.16.2) ──────────────
            # Cada usuario puede ocultar/reordenar los widgets de su propio
            # Dashboard. Se guarda como JSON (lista de {id, visible}) en una
            # columna nueva; NULL = configuración por defecto (todo visible,
            # orden original), sin necesidad de sembrar nada al crear el
            # usuario.
            cur.execute(
                "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS dashboard_prefs TEXT"
            )
        db.close()
        log.info("Auto-migración OK")
    except Exception as e:
        log.warning(f"Auto-migración omitida: {e}")
        # v12.29.35 — DEBUG temporal: el mensaje corto (str(e)) no basta para
        # localizar en qué sentencia exacta falla _auto_migrate() (p.ej. el
        # caso real que motivó esto: "Auto-migración omitida: 0", un mensaje
        # inútil sin más contexto). log.exception() vuelca aquí el traceback
        # completo con número de línea, para diagnosticar sin tener que
        # adivinar. Quitar una vez identificada y corregida la causa.
        log.exception("Auto-migración — traceback completo del fallo:")

with app.app_context():
    _auto_migrate()

# ── Base de datos (psycopg2 / PostgreSQL) ─────────────────────────────────────
#
# Pool de conexiones — antes get_db() abría una conexión nueva con
# psycopg2.connect() en cada request y se cerraba en el teardown. Con el
# volumen actual (10 hoteles) ese handshake TCP/TLS+auth contra Supabase en
# cada petición era el principal cuello de botella. Ahora se reserva un
# pool de conexiones ya abiertas al arrancar la app y cada request
# simplemente toma una prestada (getconn) y la devuelve al terminar
# (putconn) en vez de cerrarla. El puerto de conexión (5432, directo) y el
# resto de parámetros (keepalives, application_name, cursor_factory) se
# mantienen exactamente igual que antes. (v12.7.0: el pool del chat ya no
# vive aquí — ver control_pedidos_chat.)
_db_pool = None

def _crear_pool(url, maxconn, app_name):
    return ThreadedConnectionPool(
        1, maxconn, url,
        cursor_factory=RealDictCursor,
        connect_timeout=10,
        keepalives=1,
        keepalives_idle=30,
        keepalives_interval=10,
        keepalives_count=3,
        # Identifica las conexiones de este pool en pg_stat_activity / Postgres
        # Logs de Supabase, para distinguirlas de un vistazo de cualquier otro
        # proceso que se conecte a la misma base de datos (ej. restore_agent.py,
        # que usa su propio application_name).
        application_name=app_name,
    )

def get_db():
    global _db_pool
    if "db" not in g:
        if not DATABASE_URL:
            raise RuntimeError(
                "DATABASE_URL no está configurada. "
                "Ve a Render → tu servicio → Environment y añade la variable DATABASE_URL "
                "con la URI de tu base de datos PostgreSQL (Supabase → Settings → Database → URI)."
            )
        if _db_pool is None:
            _db_pool = _crear_pool(DATABASE_URL, DB_POOL_MAXCONN, "control_pedidos_web")
            atexit.register(lambda: _db_pool.closeall())
        g.db = _db_pool.getconn()
        g.db.autocommit = False
    return g.db

def _devolver_conexion(pool, conn):
    """
    Devuelve una conexión al pool en vez de cerrarla. Si quedó con una
    transacción abierta sin commit (p.ej. porque la request terminó en una
    excepción no controlada), se hace rollback antes de devolverla, para que
    la siguiente request que la reciba del pool empiece siempre en un estado
    limpio. Si la conexión ya está rota (closed != 0), se le pide al pool que
    la descarte en vez de reciclarla.
    """
    try:
        rota = conn.closed != 0
        if not rota:
            conn.rollback()
    except Exception:
        rota = True
    try:
        pool.putconn(conn, close=rota)
    except Exception:
        try:
            conn.close()
        except Exception:
            pass

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db and _db_pool is not None:
        _devolver_conexion(_db_pool, db)

@app.after_request
def _track_egress(response):
    """
    Estimación interna de egress (Jul 2026) — acumula por día los bytes de
    cada respuesta que sirve esta app, MÁS los bytes leídos de Postgres en
    las consultas de esta misma petición (ver _track_db_bytes / query()),
    que es la parte que antes quedaba invisible en casos como un adjunto
    servido con 304 (cuerpo vacío hacia el navegador, pero lectura completa
    desde Supabase para comparar el ETag). No es exactamente el mismo
    número que el contador de Supabase (que también incluye overhead del
    propio protocolo Postgres, Auth, Storage, Realtime, etc.), pero al ser
    esta app el origen de prácticamente todo el tráfico del proyecto, sirve
    como aviso temprano fiable de la tendencia día a día. Nunca debe romper
    una respuesta real: cualquier fallo aquí se ignora en silencio.
    """
    try:
        nbytes = (response.content_length or 0) + g.get("egress_db_bytes", 0)
        if nbytes:
            execute(
                """INSERT INTO egress_tracking (fecha, bytes)
                   VALUES ((NOW() AT TIME ZONE 'Atlantic/Canary')::date, %s)
                   ON CONFLICT (fecha) DO UPDATE
                   SET bytes = egress_tracking.bytes + EXCLUDED.bytes""",
                (nbytes,)
            )
            get_db().commit()
    except Exception as e:
        log.debug(f"[EGRESS_TRACK] No se pudo registrar: {e}")
    return response

def query(sql, args=(), one=False):
    """SELECT helper — devuelve list[RealDictRow] o un RealDictRow."""
    with get_db().cursor() as cur:
        cur.execute(sql, args)
        rv = cur.fetchall()
    _track_db_bytes(sum(_tam_fila(r) for r in rv))
    return (rv[0] if rv else None) if one else rv

def execute(sql, args=()):
    """INSERT/UPDATE/DELETE helper — devuelve el cursor para leer RETURNING."""
    cur = get_db().cursor()
    cur.execute(sql, args)
    return cur

# ── Estimación de egress — tamaño de lo leído de Postgres (Jul 2026) ──────────
# `_track_egress` (más abajo) solo veía los bytes que Flask reenvía al
# navegador. Eso deja fuera casos reales de egress facturado por Supabase:
# p.ej. un adjunto ya cacheado en el navegador responde 304 (0 bytes al
# usuario), pero la fila con el archivo completo igualmente se leyó de
# Postgres para poder comparar el ETag — ese tráfico sí lo cobra Supabase y
# antes era completamente invisible para nuestra propia estimación.
# Instrumentamos `query()` (punto único por el que pasan todos los SELECT
# de la app) para sumar el tamaño aproximado de cada fila leída, acumulado
# en `g` durante la petición o el job en curso.
def _tam_valor(v) -> int:
    """Tamaño aproximado en bytes de un valor de columna."""
    if v is None:
        return 0
    if isinstance(v, (bytes, bytearray, memoryview)):
        return len(v)
    if isinstance(v, str):
        return len(v.encode("utf-8", errors="ignore"))
    try:
        return len(str(v).encode("utf-8", errors="ignore"))
    except Exception:
        return 0

def _tam_fila(row) -> int:
    try:
        return sum(_tam_valor(v) for v in row.values())
    except Exception:
        return 0

def _track_db_bytes(nbytes: int):
    """Acumula bytes leídos de Postgres en el contexto actual (request HTTP
    o job en background — ambos corren dentro de un app_context, donde `g`
    vive y se resetea solo al terminar cada uno)."""
    if not nbytes:
        return
    try:
        g.egress_db_bytes = g.get("egress_db_bytes", 0) + nbytes
    except Exception:
        pass

def _flush_egress_bytes():
    """Registra en egress_tracking los bytes de Postgres acumulados durante
    un job en segundo plano. Los jobs no pasan por `_track_egress` (no hay
    respuesta HTTP), así que cada job debe llamar a esto explícitamente al
    terminar, dentro de su propio app_context."""
    try:
        nbytes = g.get("egress_db_bytes", 0)
        if nbytes:
            execute(
                """INSERT INTO egress_tracking (fecha, bytes)
                   VALUES ((NOW() AT TIME ZONE 'Atlantic/Canary')::date, %s)
                   ON CONFLICT (fecha) DO UPDATE
                   SET bytes = egress_tracking.bytes + EXCLUDED.bytes""",
                (nbytes,)
            )
            get_db().commit()
    except Exception as e:
        log.debug(f"[EGRESS_TRACK] No se pudo registrar (job): {e}")

def row_to_dict(row):
    return dict(row) if row else None

def rows_to_list(rows):
    return [dict(r) for r in rows]

def _normalizar_fecha_entrega_especifica(p):
    """
    (fix v12.29.45) fecha_entrega_especifica es la única fecha de
    'pedidos' guardada como columna DATE real — todas las demás
    (fecha_solicitud, fecha_envio_visto_bueno, fecha_tramitacion...) son
    TEXT con formato 'YYYY-MM-DD' desde el principio. Por eso psycopg2
    la devuelve como datetime.date, y el serializador JSON por defecto
    de Flask convierte cualquier `date`/`datetime` a formato RFC 1123
    ('Wed, 10 Aug 2026 00:00:00 GMT') en vez de ISO. El <input
    type="date"> del frontend no acepta ese formato y se queda vacío al
    reabrir el pedido — parece que la fecha "no se grabó" aunque sí se
    guardó correctamente en BD. Se normaliza aquí a texto ISO antes de
    devolverla en cualquier respuesta JSON. Aplicar SIEMPRE que un dict
    de pedido con esta columna vaya a pasar por jsonify().
    """
    if p is None:
        return p
    v = p.get("fecha_entrega_especifica")
    if hasattr(v, "strftime"):
        p["fecha_entrega_especifica"] = v.strftime("%Y-%m-%d")
    return p

def _puede_ver_hotel_pruebas() -> bool:
    """
    (2026-08-03) True si la sesión actual puede ver/usar el hotel de
    pruebas ('PR'): el rol admin, o el usuario dedicado a estas pruebas
    (username 'usuario prueba'), sea cual sea su rol real. El resto de usuarios
    no deben verlo ni interactuar con sus pedidos en ningún sitio.
    """
    return (session.get("rol") == "admin"
            or session.get("username") == USERNAME_HOTEL_PRUEBAS)

def _es_hotel_pruebas_id(hotel_id) -> bool:
    """
    (2026-08-03) True si hotel_id corresponde al hotel de pruebas ('PR').
    Usado junto con _puede_ver_hotel_pruebas() para bloquear a compras/
    hotel (salvo el usuario 'Prueba') la creación o edición de pedidos
    sobre este hotel.
    """
    if not hotel_id:
        return False
    row = query("SELECT 1 FROM hoteles WHERE id=%s AND codigo=%s",
                (hotel_id, HOTEL_CODIGO_PRUEBAS), one=True)
    return bool(row)

# ── Helpers de Supabase Storage (v12.8.0) ────────────────────────────────────
# Llamadas directas a la API REST de Storage (no usamos supabase-py, para no
# añadir una dependencia grande solo para esto). Autenticación con la
# service_role key — bypassa RLS, así que el bucket puede (y debe) quedar
# privado; el control de acceso lo sigue haciendo esta app con @login_required,
# exactamente igual que ahora con los adjuntos en la base de datos.

def _storage_headers(content_type=None, upsert=False):
    h = {"Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}"}
    if content_type:
        h["Content-Type"] = content_type
    if upsert:
        h["x-upsert"] = "true"
    return h


def _storage_asegurar_bucket():
    """Crea el bucket si no existe todavía (idempotente). Se llama una vez
    al arrancar la app, junto a _auto_migrate(). Si falla (permisos,
    STORAGE_CONFIGURADO=False, etc.) solo se loguea — no debe impedir que
    arranque el resto de la aplicación."""
    if not STORAGE_CONFIGURADO:
        log.warning("[STORAGE] SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY no configurados — "
                    "la migración de adjuntos a Storage queda desactivada.")
        return
    try:
        r = requests.post(
            f"{SUPABASE_URL}/storage/v1/bucket",
            headers=_storage_headers("application/json"),
            json={"id": SUPABASE_STORAGE_BUCKET, "name": SUPABASE_STORAGE_BUCKET, "public": False},
            timeout=10,
        )
        if r.status_code in (200, 201):
            log.info("[STORAGE] Bucket '%s' creado.", SUPABASE_STORAGE_BUCKET)
        elif r.status_code == 400 and "already exists" in r.text.lower():
            pass  # ya existía, nada que hacer
        else:
            log.warning("[STORAGE] No se pudo verificar/crear el bucket '%s': %s %s",
                        SUPABASE_STORAGE_BUCKET, r.status_code, r.text[:200])
    except Exception as e:
        log.warning("[STORAGE] Error creando bucket: %s", e)


def _storage_subir(path: str, contenido: bytes, mime_type: str) -> bool:
    try:
        r = requests.post(
            f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_STORAGE_BUCKET}/{path}",
            headers=_storage_headers(mime_type or "application/octet-stream", upsert=True),
            data=contenido,
            timeout=30,
        )
        return r.status_code in (200, 201)
    except Exception as e:
        log.error("[STORAGE] Error subiendo '%s': %s", path, e)
        return False


def _storage_descargar(path: str):
    """Devuelve los bytes del objeto, o None si falla. El propio tráfico de
    esta descarga (Storage → esta app) SÍ cuenta como egress de Supabase,
    igual que antes contaba el SELECT de la columna `datos` — moverlo a
    Storage no elimina ese coste, solo el de tamaño de BD."""
    try:
        r = requests.get(
            f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_STORAGE_BUCKET}/{path}",
            headers=_storage_headers(),
            timeout=30,
        )
        if r.status_code == 200:
            return r.content
        log.error("[STORAGE] Error descargando '%s': %s", path, r.status_code)
        return None
    except Exception as e:
        log.error("[STORAGE] Error descargando '%s': %s", path, e)
        return None


def _storage_borrar(path: str) -> bool:
    try:
        r = requests.delete(
            f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_STORAGE_BUCKET}/{path}",
            headers=_storage_headers("application/json"),
            timeout=15,
        )
        return r.status_code in (200, 204)
    except Exception as e:
        log.error("[STORAGE] Error borrando '%s': %s", path, e)
        return False

_storage_asegurar_bucket()

def _fmt_importe_es(v) -> str:
    """Formato español ('1.234,56') a partir de un float — mismo patrón ya
    usado en el resto de la app (p.ej. la nota de cancelación del techo)."""
    return f"{float(v):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")


def format_albaran_display(albaran_str):
    """
    Convierte el string de albarán almacenado al formato legible para Excel/emails.
    Formato almacenado: "NUM1::FECHA1::IMPORTE1 | NUM2::FECHA2::IMPORTE2"
    Formato legible:   "NUM1 (FECHA1, 123,45 €) | NUM2 (FECHA2)"
    Retrocompatible con el formato antiguo "NUM1::FECHA1 | NUM2" y "NUM1 | NUM2"
    (2026-08-27: 3er segmento opcional, la base imponible del albarán — ver
    _parse_albaran_entries).
    """
    if not albaran_str:
        return albaran_str
    partes = []
    for entry in albaran_str.split('|'):
        entry = entry.strip()
        if not entry:
            continue
        if '::' in entry:
            campos = entry.split('::', 2)
            num   = campos[0].strip()
            fecha = campos[1].strip() if len(campos) > 1 else ''
            base_imp_txt = campos[2].strip() if len(campos) > 2 else ''
            detalle = fecha
            if base_imp_txt:
                try:
                    detalle = f"{detalle}, {_fmt_importe_es(float(base_imp_txt))} €" if detalle else f"{_fmt_importe_es(float(base_imp_txt))} €"
                except ValueError:
                    pass
            partes.append(f"{num} ({detalle})" if detalle else num)
        else:
            partes.append(entry)
    return ' | '.join(partes) if partes else albaran_str


def _parse_albaran_entries(albaran_str):
    """
    Parsea el campo entrada_albaran_num
    ("NUM::FECHA::IMPORTE | NUM::FECHA::IMPORTE | NUM") en una lista de
    entregas: [{"num": str, "fecha_iso": "YYYY-MM-DD"|None,
    "base_imponible": float|None}, ...]
    Retrocompatible con entradas antiguas sin fecha (solo "NUM") y sin base
    imponible (solo "NUM::FECHA").
    Conserva el orden cronológico en que fueron registradas.
    """
    if not albaran_str:
        return []
    entradas = []
    for entry in albaran_str.split('|'):
        entry = entry.strip()
        if not entry:
            continue
        if '::' in entry:
            campos = entry.split('::', 2)
            num   = campos[0].strip()
            fecha = campos[1].strip() if len(campos) > 1 else ''
            base_imponible = None
            if len(campos) > 2 and campos[2].strip():
                try:
                    base_imponible = float(campos[2].strip())
                except ValueError:
                    base_imponible = None
        else:
            num, fecha, base_imponible = entry, '', None
        entradas.append({"num": num or '—', "fecha_iso": fecha or None, "base_imponible": base_imponible})
    return entradas


def _construir_entrada_albaran_num(entradas):
    """
    Inverso de _parse_albaran_entries(): reconstruye el string combinado
    "NUM::FECHA::IMPORTE | ..." a partir de una lista de entregas — usado
    para reescribir UNA sola entrada (p.ej. rellenar su base imponible)
    sin tocar el resto. Solo añade el 3er segmento (base imponible) cuando
    hay valor, para no ensuciar entradas que no lo tienen.
    """
    partes = []
    for e in entradas:
        num = (e.get("num") or "").strip()
        if not num or num == '—':
            continue
        fecha = e.get("fecha_iso") or ""
        base_imp = e.get("base_imponible")
        if base_imp is not None:
            partes.append(f"{num}::{fecha}::{base_imp:.2f}")
        elif fecha:
            partes.append(f"{num}::{fecha}")
        else:
            partes.append(num)
    return ' | '.join(partes)


def _validar_base_imponible_entradas(entradas: list) -> bool:
    """
    (2026-08-28) A petición de Víctor: la Base imp. (€) de cada entrada de
    "Nº Entrada DALI / SAP" deja de ser opcional — obligatoria tanto en una
    entrada parcial como en la entrada marcada como final (total), para
    poder continuar. Se aplica sobre TODA la lista de entradas cada vez que
    se guarda un pedido en ENTREGA PARCIAL o ENTREGADO (ver update_pedido,
    tanto en la rama de rol Hotel como en la general) — no solo a la
    entrada nueva —, para que un pedido con alguna entrada antigua sin base
    imponible (de antes de este cambio) no se pueda seguir editando sin
    completarla también. Lista vacía = válida (nada que exigir todavía).
    """
    return all(
        (e.get("base_imponible") is not None and e["base_imponible"] > 0)
        for e in entradas
    )


def _fecha_es(fecha_val):
    """Convierte una fecha 'YYYY-MM-DD' (o similar) en 'DD/MM/YYYY'. None si no hay valor."""
    if not fecha_val:
        return None
    try:
        return datetime.strptime(str(fecha_val)[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return str(fecha_val)


def _nota_base_imponible_html() -> str:
    """Aviso reutilizable (HTML) recordando que los importes mostrados son base imponible (sin IGIC)."""
    return ('<p style="margin:10px 0 0;font-size:12px;color:#666">'
            'ℹ️ Los importes indicados son siempre <b>base imponible</b> (total sin IGIC).</p>')


def _nota_base_imponible_text() -> str:
    """Aviso reutilizable (texto plano) recordando que los importes mostrados son base imponible (sin IGIC)."""
    return "Nota: los importes indicados son siempre base imponible (total sin IGIC)."


def _resumen_entregas(pedido: dict, estado_nuevo: str = None) -> dict:
    """
    Construye un resumen de las entregas (albaranes) registradas en el pedido
    a partir de entrada_albaran_num, listo para insertar en correos/Telegram
    de cambio de estado.

    Cuando el estado de referencia es ENTREGADO, la última entrada registrada
    se marca como "es_final" (la entrega que cierra el pedido).

    (2026-09-02) A petición de Víctor: además del histórico de importes, el
    correo interno de ENTREGA PARCIAL/ENTREGADO debe poder indicar cuántos
    días han transcurrido entre la fecha de tramitación del pedido y cada
    entrega (parcial o final). Se añade "dias_desde_pedido" por entrada, y
    "total_pedido"/"total_pendiente" a nivel de resumen (importe que falta
    por recibir sobre el total del pedido, base imponible).

    Devuelve:
        {
          "entregas":        [{"num","fecha_iso","fecha_es","es_final","base_imponible",
                                "dias_desde_pedido"}, ...],
          "total":           int,
          "ultima_fecha_es": str|None,   # fecha de la entrega más reciente registrada
          "tiene_fechas":    bool,       # alguna entrega tiene fecha informada
          "total_recibido":  float|None, # suma de base_imponible de las entradas (None si ninguna la tiene)
          "total_pedido":    float|None, # total del pedido (base imponible), tal cual en `pedido`
          "total_pendiente": float|None, # total_pedido - total_recibido (None si falta algún dato)
          "dias_pedido_a_final": int|None,  # días entre tramitación y la entrega marcada es_final
        }
    """
    entradas = _parse_albaran_entries(pedido.get("entrada_albaran_num"))
    estado_ref = estado_nuevo or pedido.get("estado")

    _fecha_tram_iso = None
    if pedido.get("fecha_tramitacion"):
        try:
            _fecha_tram_iso = datetime.strptime(str(pedido["fecha_tramitacion"])[:10], "%Y-%m-%d").date()
        except Exception:
            _fecha_tram_iso = None

    out = []
    total_recibido = 0.0
    hay_importe = False
    dias_pedido_a_final = None
    for idx, e in enumerate(entradas):
        es_final = (estado_ref == "ENTREGADO" and idx == len(entradas) - 1)
        bi = e.get("base_imponible")
        if bi is not None:
            total_recibido += bi
            hay_importe = True
        dias_desde_pedido = None
        if _fecha_tram_iso and e["fecha_iso"]:
            try:
                _fe = datetime.strptime(str(e["fecha_iso"])[:10], "%Y-%m-%d").date()
                dias_desde_pedido = (_fe - _fecha_tram_iso).days
            except Exception:
                dias_desde_pedido = None
        if es_final:
            dias_pedido_a_final = dias_desde_pedido
        out.append({
            "num":               e["num"],
            "fecha_iso":         e["fecha_iso"],
            "fecha_es":          _fecha_es(e["fecha_iso"]),
            "es_final":          es_final,
            "base_imponible":    bi,
            "dias_desde_pedido": dias_desde_pedido,
        })
    fechas_validas = [e["fecha_es"] for e in out if e["fecha_es"]]
    _total_recibido_val = round(total_recibido, 2) if hay_importe else None
    _total_pedido_val = pedido.get("total_pedido")
    _total_pendiente_val = None
    if _total_pedido_val is not None and _total_recibido_val is not None:
        _total_pendiente_val = round(float(_total_pedido_val) - _total_recibido_val, 2)
    return {
        "entregas":            out,
        "total":               len(out),
        "ultima_fecha_es":     fechas_validas[-1] if fechas_validas else None,
        "tiene_fechas":        bool(fechas_validas),
        "total_recibido":      _total_recibido_val,
        "total_pedido":        _total_pedido_val,
        "total_pendiente":     _total_pendiente_val,
        "dias_pedido_a_final": dias_pedido_a_final,
    }


def _html_bloque_entregas(resumen: dict, estado_nuevo: str) -> str:
    """Tabla HTML con el histórico de entregas (albaranes + fechas + base imponible + días desde el pedido) para el correo interno."""
    if not resumen["entregas"]:
        return ""
    filas = []
    for i, e in enumerate(resumen["entregas"], 1):
        etiqueta = "Entrega final (TOTAL)" if e["es_final"] else f"Entrega parcial {i}"
        fecha_txt = e["fecha_es"] or "fecha no indicada"
        importe_txt = _fmt_importe_es(e["base_imponible"]) + " €" if e["base_imponible"] is not None else "—"
        # (2026-09-02) A petición de Víctor: indicar los días transcurridos
        # entre la fecha de tramitación del pedido y cada entrega (parcial
        # o final), para que quede constancia del plazo real de cada envío.
        dias_txt = f"{e['dias_desde_pedido']} día(s)" if e.get("dias_desde_pedido") is not None else "—"
        estilo = ' style="background:#e8f5e9;font-weight:600"' if e["es_final"] else ''
        filas.append(
            f'<tr{estilo}><td>{i}</td><td>{etiqueta}</td>'
            f'<td>{e["num"]}</td><td>{fecha_txt}</td><td>{importe_txt}</td><td>{dias_txt}</td></tr>'
        )
    titulo = ("Histórico de entregas registradas" if estado_nuevo == "ENTREGADO"
              else "Entregas parciales registradas hasta la fecha")
    plural = "s" if resumen["total"] != 1 else ""
    total_html = ""
    if resumen.get("total_recibido") is not None:
        total_html = (f'<p style="margin:6px 0 0"><b>Total recibido hasta la fecha (base imponible):</b> '
                       f'{_fmt_importe_es(resumen["total_recibido"])} €</p>')
    # (2026-09-02) Importe pendiente sobre el total del pedido — solo tiene
    # sentido mostrarlo cuando aún queda algo por recibir (no en ENTREGADO,
    # donde el propio estado ya indica que no queda nada pendiente).
    if estado_nuevo != "ENTREGADO" and resumen.get("total_pendiente") is not None:
        total_html += (f'<p style="margin:2px 0 0"><b>Pendiente de recibir sobre el total del pedido '
                        f'({_fmt_importe_es(resumen["total_pedido"])} €):</b> '
                        f'{_fmt_importe_es(resumen["total_pendiente"])} €</p>')
    return (
        f'<p style="margin:16px 0 6px"><b>{titulo}</b> ({resumen["total"]} entrada{plural}):</p>'
        f'<table border="1" cellpadding="6" style="border-collapse:collapse;font-family:sans-serif;font-size:13px">'
        f'<tr style="background:#f0f0f0"><th>#</th><th>Tipo</th><th>Nº Entrada DALI/SAP</th><th>Fecha</th>'
        f'<th>Base imp. (€)</th><th>Días desde pedido</th></tr>'
        + "".join(filas) + "</table>" + total_html + _nota_base_imponible_html()
    )


def _text_bloque_entregas(resumen: dict, estado_nuevo: str) -> str:
    """Bloque de texto plano con el histórico de entregas (con base imponible y días desde el pedido), para el correo interno (fallback texto)."""
    if not resumen["entregas"]:
        return ""
    titulo = ("Histórico de entregas registradas" if estado_nuevo == "ENTREGADO"
              else "Entregas parciales registradas hasta la fecha")
    lineas = [f"{titulo} ({resumen['total']}):"]
    for i, e in enumerate(resumen["entregas"], 1):
        etiqueta = "ENTREGA FINAL (TOTAL)" if e["es_final"] else f"Entrega parcial {i}"
        fecha_txt = e["fecha_es"] or "fecha no indicada"
        importe_txt = f" — {_fmt_importe_es(e['base_imponible'])} €" if e["base_imponible"] is not None else ""
        dias_txt = f" — {e['dias_desde_pedido']} día(s) desde el pedido" if e.get("dias_desde_pedido") is not None else ""
        lineas.append(f"  {i}. {etiqueta} — Nº {e['num']} — {fecha_txt}{importe_txt}{dias_txt}")
    if resumen.get("total_recibido") is not None:
        lineas.append(f"  Total recibido hasta la fecha (base imponible): {_fmt_importe_es(resumen['total_recibido'])} €")
    if estado_nuevo != "ENTREGADO" and resumen.get("total_pendiente") is not None:
        lineas.append(
            f"  Pendiente de recibir sobre el total del pedido ({_fmt_importe_es(resumen['total_pedido'])} €): "
            f"{_fmt_importe_es(resumen['total_pendiente'])} €"
        )
    lineas.append(_nota_base_imponible_text())
    return "\n".join(lineas)


def _telegram_bloque_entregas(resumen: dict, estado_nuevo: str) -> list:
    """Líneas (para añadir a un mensaje Markdown de Telegram) con el histórico de entregas (con base imponible)."""
    if not resumen["entregas"]:
        return []
    titulo = ("📦 *Histórico de entregas*" if estado_nuevo == "ENTREGADO"
              else "📦 *Entregas parciales hasta la fecha*")
    lineas = ["", f"{titulo} ({resumen['total']}):"]
    for i, e in enumerate(resumen["entregas"], 1):
        marca = "✅" if e["es_final"] else "▫️"
        etiqueta = "Entrega final (TOTAL)" if e["es_final"] else f"Parcial {i}"
        fecha_txt = e["fecha_es"] or "sin fecha"
        importe_txt = f" — {_fmt_importe_es(e['base_imponible'])} €" if e["base_imponible"] is not None else ""
        lineas.append(f"{marca} {etiqueta} — Nº {e['num']} — {fecha_txt}{importe_txt}")
    if resumen.get("total_recibido") is not None:
        lineas.append(f"💶 Total recibido hasta la fecha (base imponible): {_fmt_importe_es(resumen['total_recibido'])} €")
    lineas.append("_Importes en base imponible (sin IGIC)._")
    return lineas

# ── Autenticación ──────────────────────────────────────────────────────────────

def _hoy_canarias():
    import pytz
    return datetime.now(pytz.timezone("Atlantic/Canary")).date()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "No autenticado"}), 401
        # Caducidad diaria: si la sesión se creó un día distinto de hoy
        # (hora Canarias), se invalida y se obliga a volver a iniciar
        # sesión — evita que una pestaña olvidada abierta en la oficina
        # se quede autenticada indefinidamente.
        if session.get("login_date") != _hoy_canarias().isoformat():
            session.clear()
            return jsonify({"error": "Tu sesión ha caducado, vuelve a iniciar sesión", "sesion_caducada": True}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return jsonify({"error": "No autenticado"}), 401
        if session.get("login_date") != _hoy_canarias().isoformat():
            session.clear()
            return jsonify({"error": "Tu sesión ha caducado, vuelve a iniciar sesión", "sesion_caducada": True}), 401
        if session.get("rol") != "admin":
            return jsonify({"error": "Solo administradores"}), 403
        return f(*args, **kwargs)
    return decorated

def current_user_id():
    return session.get("user_id")

def _log_email(db, pedido_id, tipo, destinatario, asunto, enviado, error=None):
    with db.cursor() as cur:
        cur.execute(
            "INSERT INTO emails_log (pedido_id,tipo,destinatario,asunto,enviado,error) VALUES (%s,%s,%s,%s,%s,%s)",
            (pedido_id, tipo, destinatario, asunto, 1 if enviado else 0, error)
        )

def _encolar_email_pedido_retrasado(pedido_id: int, evento_codigo: str, destinatario: str,
                                     asunto: str, cuerpo_html: str, cuerpo_text: str,
                                     cc_emails: str = "", retraso_segundos: int = 300,
                                     marca_comunicado_ab: bool = False,
                                     marca_comunicado_jefe_dep: bool = False) -> None:
    """
    (2026-08-14) Encola un correo de cambio de estado en emails_sistema_pendientes
    con retraso (columna visible_en = NOW() + retraso_segundos), en vez de
    devolverlo para que el navegador que hizo el cambio lo envíe de inmediato
    vía EmailJS — mismo objetivo y mismo mecanismo que el retraso ya aplicado
    al popup (ver _encolar_bridge_notificacion): a petición del usuario
    (Víctor), varios cambios de estado seguidos sobre el mismo pedido (p. ej.
    un error corregido al momento) no deben disparar un correo por cada uno.

    Si ya hay un correo sin enviar y sin reservar (en_proceso_desde libre o
    caducado) para el mismo (pedido_id, evento_codigo), se SOBRESCRIBE con el
    contenido más reciente y se reinicia la espera, en vez de encolar uno
    nuevo — así solo se entrega el último cambio, pasados esos minutos sin
    más cambios sobre ese pedido.

    Reutiliza la cola/poller que ya existía para "emails de sistema" (avisos
    generados por jobs sin navegador abierto, p. ej. techo urgente o
    familias repetidas): cualquier admin o comprador con la app abierta la
    revisa cada 5 minutos (`_enviarEmailsSistemaPendientes` en
    templates/index.html) y envía los correos con visible_en ya cumplido,
    con reserva atómica (`en_proceso_desde`) para no duplicar si hay varias
    sesiones abiertas a la vez — ver GET /api/emails-sistema-pendientes.
    Ventaja adicional sobre el envío inmediato de antes: ya no depende de que
    quien hizo el cambio no cierre la pestaña — lo despacha cualquier sesión
    abierta, la que sea.

    marca_comunicado_ab / marca_comunicado_jefe_dep: (2026-08-31) si este
    correo concreto va a llevar en copia a A&B / al buzón del departamento
    — se guardan en la propia fila de la cola para que, cuando se confirme
    que el correo se ha enviado de verdad (ver api_marcar_email_sistema_
    enviado), se marquen solas las casillas "Comunicado A&B" / "Comunicado
    Jefe Dep." del pedido — nunca antes, y nunca a mano por el usuario.
    """
    try:
        db = get_db()
        cur = db.cursor()
        cur.execute(
            """UPDATE emails_sistema_pendientes
               SET destinatario=%s, asunto=%s, cuerpo_html=%s, cuerpo_text=%s,
                   cc_emails=%s, creado_en=NOW(),
                   marca_comunicado_ab=%s, marca_comunicado_jefe_dep=%s,
                   visible_en=NOW() + make_interval(secs => %s)
               WHERE pedido_id=%s AND evento_codigo=%s
                 AND enviado=FALSE
                 AND (en_proceso_desde IS NULL OR en_proceso_desde < NOW() - INTERVAL '2 minutes')
               RETURNING id""",
            (destinatario, asunto, cuerpo_html, cuerpo_text, cc_emails,
             marca_comunicado_ab, marca_comunicado_jefe_dep,
             retraso_segundos, pedido_id, evento_codigo)
        )
        if cur.fetchone() is not None:
            db.commit()
            return
        cur.execute(
            """INSERT INTO emails_sistema_pendientes
               (evento_codigo, destinatario, asunto, cuerpo_html, cuerpo_text,
                cc_emails, pedido_id, marca_comunicado_ab, marca_comunicado_jefe_dep,
                visible_en)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW() + make_interval(secs => %s))""",
            (evento_codigo, destinatario, asunto, cuerpo_html, cuerpo_text,
             cc_emails, pedido_id, marca_comunicado_ab, marca_comunicado_jefe_dep,
             retraso_segundos)
        )
        db.commit()
    except Exception as exc:
        log.warning("emails_pedido_retrasado: no se pudo encolar (%s, pedido %s) — %s",
                    evento_codigo, pedido_id, exc)


def _obtener_o_crear_token_adjunto(adjunto_id: int, dias_validez: int = 180) -> str:
    """
    (2026-08-28) Devuelve un token de descarga pública para este adjunto
    (ver /descargas/adjunto/<token> y la tabla adjunto_descarga_tokens),
    reutilizando uno todavía vigente si ya existe, o creando uno nuevo si
    no hay ninguno o el último ya caducó. Pensado para el enlace de
    descarga del PDF del pedido en el correo al proveedor — ver
    _enlaces_descarga_pedido_doc().

    No hay invalidación manual desde la app: un enlace ya enviado sigue
    funcionando hasta expira_en (180 días por defecto) aunque se genere
    uno nuevo después para otro envío del mismo adjunto.
    """
    fila = row_to_dict(query(
        "SELECT token FROM adjunto_descarga_tokens WHERE adjunto_id=%s AND expira_en > NOW() "
        "ORDER BY creado_en DESC LIMIT 1",
        (adjunto_id,), one=True
    ))
    if fila:
        return fila["token"]
    token = secrets.token_urlsafe(32)
    expira = datetime.now(timezone.utc) + timedelta(days=dias_validez)
    try:
        db = get_db()
        db.cursor().execute(
            "INSERT INTO adjunto_descarga_tokens (adjunto_id, token, expira_en) VALUES (%s,%s,%s)",
            (adjunto_id, token, expira)
        )
        db.commit()
    except Exception as e:
        log.warning(f"No se pudo crear token de descarga para adjunto {adjunto_id}: {e}")
        return None
    return token


def _enlaces_descarga_pedido_doc(pedido_id: int) -> list:
    """
    (2026-08-28) A petición de Víctor: en el correo "ENVIADO AL
    PROVEEDOR", en vez de adjuntar directamente el PDF del pedido (el que
    se sube en "Nº Pedido (DALI/SAP)" → "Adjuntar doc. / correo" —
    EmailJS en el plan actual, Free, no admite adjuntos, ver conversación
    del 28/08), se incluye un enlace de descarga temporal (ver
    /descargas/adjunto/<token>).

    Solo se enlazan los documentos realmente PDF (tipo 'pedido_doc' o el
    legacy 'pedido_pdf') que NO sean copias de correo (es_correo=TRUE son
    los .eml/.msg subidos como evidencia interna, no documentos a
    reenviar al proveedor). Si el pedido no tiene ningún PDF subido ahí,
    devuelve una lista vacía — el correo se sigue enviando igual, sin
    ningún enlace, nunca se bloquea por esto.

    El filtro de mime_type acepta tanto 'application/pdf' como
    'application/octet-stream' con nombre terminado en «.pdf»: en el
    apartado «pedido_doc» se admite subir un PDF cuyo navegador informe el
    mime genérico application/octet-stream (ver validación en
    upload_adjunto/MIME_SOLICITUD_DOC), y ese mismo valor es el que queda
    guardado tal cual en mime_type — sin este segundo caso se perderían
    enlaces de PDF genuinos subidos así.
    """
    filas = rows_to_list(query(
        """SELECT id, nombre FROM pedido_adjuntos
           WHERE pedido_id=%s AND tipo IN ('pedido_doc','pedido_pdf')
             AND NOT es_correo
             AND (mime_type='application/pdf'
                  OR (mime_type='application/octet-stream' AND nombre ILIKE '%%.pdf'))
           ORDER BY creado_en""",
        (pedido_id,)
    )) or []
    app_url = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    enlaces = []
    for f in filas:
        token = _obtener_o_crear_token_adjunto(f["id"])
        if token:
            enlaces.append({"nombre": f["nombre"], "url": f"{app_url}/descargas/adjunto/{token}"})
    return enlaces


def enviar_emails_estado(db, pedido_id: int, estado_nuevo: str, estado_antes: str = None,
                          usuario_nombre: str = "", usuario_id: int = None,
                          es_automatico: bool = False):
    """
    Construye los correos de notificación de cambio de estado (proveedor +
    internos) y los registra en _log_email.

    (2026-08-14) Ya NO se devuelven para envío inmediato desde el navegador
    que hizo el cambio: se encolan con 5 minutos de retraso en
    emails_sistema_pendientes vía _encolar_email_pedido_retrasado() — a
    petición del usuario, para que varios cambios de estado seguidos sobre
    el mismo pedido no disparen un correo por cada uno (mismo motivo y
    mismo mecanismo que el retraso ya aplicado al popup, ver
    _encolar_bridge_notificacion). El envío real lo sigue haciendo el
    navegador vía EmailJS (esta app no tiene SMTP propio), pero ahora lo
    hace el poller de "emails de sistema" que ya recorría esa cola cada 5
    minutos desde cualquier sesión abierta — no necesariamente la del
    usuario que hizo el cambio.

    usuario_nombre: quién realizó el cambio (o creó el pedido) — se incluye
    en el correo interno ("Realizado por:"), nunca en el correo al
    proveedor (es un dato interno, no debe salir fuera de la empresa).

    (2026-08-19) usuario_id / es_automatico — a petición de Víctor, el
    correo interno de cambio de estado ya NO se manda a las dos partes por
    igual en un cambio manual: se excluye de los destinatarios a la
    PERSONA CONCRETA que ha realizado el cambio (no a todo su rol/lado —
    si comparte hotel con más compañeros de compras o de hotel, esos
    siguen recibiendo el correo con normalidad), porque esa persona ya
    sabe lo que acaba de hacer. Quien hizo el cambio se sigue enterando
    por el popup/Telegram de "cambio_estado_pedido" (canal totalmente
    aparte, ver _telegram_cambio_estado — no se toca aquí), tal como
    pidió. (2026-08-19, ajuste: la primera versión excluía a todo el
    lado/rol de quien hacía el cambio — Víctor pidió que fuera solo la
    persona concreta.)
    - usuario_id: id de quien realizó el cambio; se consulta su email
      (y email2) y se quita de la lista de destinatarios del correo
      interno. Si no se indica, o no tiene email, se manda a todos los
      internos como antes (comportamiento más seguro por defecto que
      dejar a alguien sin avisar).
    - es_automatico=True: el cambio no lo ha decidido una persona en ese
      momento (p. ej. _aplicar_coincidencia_albaran(), al confirmar una
      coincidencia de "Comparar Pedidos + Albaranes") — se manda SIEMPRE
      a todos los internos, igual que antes, sin excluir a nadie.
      (2026-09-03) Además, con es_automatico=True el correo se encola con
      un retraso mínimo (2s) en vez de los 300s (5 min) de un cambio
      manual — ver _retraso_email_estado más abajo: el retraso largo
      existe para agrupar varias ediciones manuales SEGUIDAS sobre el
      mismo pedido en un único correo, algo que no aplica a un cambio
      automático (una única escritura determinista por pedido). Motivo
      del cambio: un cambio automático (hotel GY) se quedó en la cola sin
      enviarse porque nadie dejó la app abierta 5 minutos más tras pulsar
      "Aplicar" — a diferencia del correo de resumen de la comparación,
      que si despacha la cola de inmediato desde el propio navegador.

    Devuelve: [] siempre — se mantiene por compatibilidad con los callers
    (create_pedido / update_pedido), que incluyen el valor en su respuesta
    JSON como "emails_pendientes"; el frontend ya no tiene nada que enviar
    de inmediato con ese valor (ver _enviarEmailsPendientesEstado en
    templates/index.html, ahora en desuso — el envío real pasa por
    _enviarEmailsSistemaPendientes).
    """
    pendientes = []

    # (2026-09-03) A petición de Víctor, tras detectar que los dos correos
    # internos de un cambio automático (hotel GY, "Comparar Pedidos +
    # Albaranes") no llegaron: el retraso de 300s (5 min) con el que se
    # encolan estos correos (ver _encolar_email_pedido_retrasado) existe
    # para agrupar varios cambios de estado SEGUIDOS sobre el MISMO pedido
    # hechos a mano (varios guardados rápidos) en un único correo — no
    # tiene sentido aplicado a un cambio automático, que es una única
    # escritura determinista por pedido en el momento de pulsar "Aplicar".
    # Además, a diferencia del correo de resumen de la comparación (que sí
    # dispara un despacho inmediato desde el propio navegador nada más
    # encolarse, ver enviarResumenComparacionAlbaranes() en
    # templates/index.html), el botón "Aplicar" no lo hacía — así que,
    # combinado con el retraso de 5 min, si nadie dejaba la app abierta
    # con sesión de admin/compras al menos 5 minutos más tras pulsar
    # "Aplicar" (p. ej. por ser fin de jornada), el correo se quedaba en
    # la cola sin enviarse hasta que alguien volviera a abrir la app —
    # visible como pendiente en Admin → EmailJS → "Cola de correos de
    # sistema pendientes", nunca perdido, pero sin salir de verdad. Con
    # `es_automatico=True` el retraso baja a prácticamente cero (mismo
    # criterio que el correo de resumen, sin delay) — ver uso más abajo,
    # en el encolado del correo a proveedor y del correo interno.
    _retraso_email_estado = 2 if es_automatico else 300

    pedido = row_to_dict(query(
        """SELECT p.*, h.nombre as hotel_nombre, h.codigo as hotel_codigo,
                  d.nombre as departamento_nombre,
                  pr.nombre as proveedor_nombre,
                  (SELECT pc.email
                     FROM proveedor_contactos pc
                    WHERE pc.proveedor_id = pr.id AND pc.es_principal = 1
                      AND pc.email IS NOT NULL AND pc.email != ''
                      AND (EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id AND pch.hotel_id = p.hotel_id)
                           OR NOT EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch2 WHERE pch2.contacto_id = pc.id))
                    ORDER BY EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch3 WHERE pch3.contacto_id = pc.id AND pch3.hotel_id = p.hotel_id) DESC,
                             pc.orden, pc.id
                    LIMIT 1) as proveedor_email
           FROM pedidos p
           LEFT JOIN hoteles h ON p.hotel_id = h.id
           LEFT JOIN departamentos d ON p.departamento_id = d.id
           LEFT JOIN proveedores pr ON p.proveedor_id = pr.id
           WHERE p.id = %s""", (pedido_id,), one=True
    ))
    if not pedido:
        return pendientes

    # (2026-08-31) Definido aquí arriba (antes solo existía más abajo, junto
    # al párrafo introductorio) porque ahora también hace falta pronto, para
    # calcular _incluye_ab_email (ver más abajo, junto al correo de
    # departamento) — mismo criterio en los dos sitios: los hoteles con
    # departamento "RESTAURANTE & BARES" combinado son los que reciben aviso
    # de A&B en el correo ENVIADO AL PROVEEDOR.
    _DEPARTAMENTOS_AB = {"COCINA", "BARES", "RESTAURANTE", "RESTAURANTE & BARES"}
    _dept_nombre_i = (pedido.get('departamento_nombre') or '').strip()

    # Motivo real del cambio de estado (CANCELADO / DENEGADO POR DIRECCION
    # GENERAL): se guarda en historial_estados.nota en el momento de la
    # transición (ver update_pedido / denegar_expediente), NO en
    # pedido.observaciones — ese es un campo aparte de notas generales del
    # pedido y normalmente está vacío en una cancelación, por eso el correo
    # se quedaba sin motivo aunque el usuario sí lo hubiera indicado.
    _motivo_estado = None
    if estado_nuevo in ("CANCELADO", "DENEGADO POR DIRECCION GENERAL"):
        _fila_hist = row_to_dict(query(
            """SELECT nota FROM historial_estados
               WHERE pedido_id=%s AND estado_nuevo=%s
               ORDER BY creado_en DESC LIMIT 1""",
            (pedido_id, estado_nuevo), one=True
        ))
        _motivo_estado = ((_fila_hist or {}).get("nota") or "").strip() or None
        if not _motivo_estado:
            _motivo_estado = (pedido.get("observaciones") or "").strip() or None

    _proveedor_emails = _get_proveedor_emails_principales(pedido.get("proveedor_id"), pedido.get("hotel_id"))
    _usuarios_hotel   = _get_todos_usuarios_hotel(pedido.get("hotel_codigo",""))
    _emails_compradores = [e for u in _usuarios_hotel["compradores"] for e in _emails_usuario(u)]
    _emails_hotel_users = [e for u in _usuarios_hotel["hotel_users"]  for e in _emails_usuario(u)]

    # (2026-08-19) Se excluye del correo interno SOLO a la persona concreta
    # que ha realizado el cambio (no a todo su rol/lado — un comprador o
    # usuario hotel puede compartir hotel con más compañeros de su mismo
    # rol, y esos sí deben seguir recibiendo el correo). Esa persona ya
    # sabe lo que acaba de hacer y se entera por el popup/Telegram de
    # "cambio_estado_pedido" (canal aparte, no tocado aquí). Si es un
    # cambio automático, o no se sabe quién lo ha hecho, no se excluye a
    # nadie (más seguro que dejar a alguien sin avisar por error).
    #
    # (2026-09-03) A petición de Víctor: SOLO se excluye el email PRINCIPAL
    # del actor, nunca su email2. El email2 es un correo de control de esa
    # misma cuenta (recibe copia de los avisos de sus hoteles asignados sin
    # que tenga por qué coincidir con quien está operando el pedido), así
    # que debe seguir recibiendo siempre la info referente a esos hoteles
    # con independencia de quién haga el cambio — incluido el caso en que
    # el propio dueño del email2 sea quien lo hizo: entonces solo se le
    # quita el principal y el email2 sigue en la lista.
    _emails_actor = []
    if not es_automatico and usuario_id:
        _actor = row_to_dict(query("SELECT email FROM usuarios WHERE id=%s", (usuario_id,), one=True))
        _email_actor_principal = ((_actor or {}).get("email") or "").strip()
        if _email_actor_principal:
            _emails_actor = [_email_actor_principal]

    _todos_internos = list(dict.fromkeys(_emails_compradores + _emails_hotel_users))  # sin duplicados
    if _emails_actor:
        _todos_internos = [e for e in _todos_internos if e not in _emails_actor]

    # (2026-08-28) Copia al departamento solicitante del pedido — a petición
    # de Víctor (ver PENDIENTES.md): cada hotel puede tener registrado un
    # correo distinto para el mismo departamento (Administrador →
    # Departamentos, tabla departamento_hotel_email). Se añade en el mismo
    # correo con copia a todos, nunca aparte — y se omite en silencio, sin
    # avisos, si ese departamento de ese hotel todavía no tiene correo
    # configurado (comportamiento seguro por defecto, igual que el resto de
    # esta función cuando falta algún destinatario). No se excluye aunque
    # coincida con quien hizo el cambio: es el buzón de un departamento, no
    # una persona concreta a la que no haga falta avisar de lo que acaba de
    # hacer.
    # (2026-08-31) _incluye_jefe_depto_email — a petición de Víctor, sobre
    # el correo interno de "PEDIDO ENVIADO AL PROVEEDOR": "en todos los
    # casos que se ponga en copia al responsable del departamento también
    # se marque la correspondiente [casilla]" — a diferencia de A&B (solo
    # aplica a ciertos departamentos), aquí "todos los casos" es CUALQUIER
    # departamento, siempre que tenga correo configurado; pero sigue
    # limitado a ESE correo (ENVIADO AL PROVEEDOR), igual que A&B — no debe
    # marcarse por un correo de ENTREGADO/CANCELADO/etc. que también vaya
    # con copia al departamento. Se usa más abajo, al encolar el correo,
    # para que la casilla "Comunicado Jefe Dep." del pedido se marque sola
    # SOLO cuando el envío real lleve ese destinatario — "en caso de no
    # tener correo configurado un departamento entonces no se marcará".
    _incluye_jefe_depto_email = False
    if pedido.get("hotel_id") and pedido.get("departamento_id"):
        _depto_email_row = row_to_dict(query(
            "SELECT email, email2 FROM departamento_hotel_email WHERE hotel_id=%s AND departamento_id=%s",
            (pedido["hotel_id"], pedido["departamento_id"]), one=True
        ))
        _emails_depto = _emails_usuario(_depto_email_row)
        _incluye_jefe_depto_email = bool(_emails_depto) and estado_nuevo == "ENVIADO AL PROVEEDOR"
        for _e in _emails_depto:
            if _e not in _todos_internos:
                _todos_internos.append(_e)

    # (2026-08-31) _incluye_ab_email — mismo criterio que la frase de A&B en
    # el párrafo introductorio del correo ENVIADO AL PROVEEDOR (ver más
    # abajo, _intro_html/_intro_text): departamentos COCINA/BARES/
    # RESTAURANTE/RESTAURANTE & BARES. Se usa para marcar sola la casilla
    # "Comunicado A&B" del pedido cuando el correo que se envía de verdad es
    # justo ese, el que informa a A&B.
    _incluye_ab_email = (
        estado_nuevo == "ENVIADO AL PROVEEDOR"
        and _dept_nombre_i.upper() in _DEPARTAMENTOS_AB
    )

    # (2026-08-28) Contactos adicionales de notificación — a petición de
    # Víctor: además del comprador, rol hotel y correo de departamento
    # (arriba), permite poner en copia a contactos sueltos que no son
    # usuarios de la app (p. ej. "Chef Ejecutivo", "Director de Compras")
    # solo para determinadas combinaciones de departamento del pedido +
    # estado nuevo — configurado en Administrador → Notificaciones,
    # global para toda la cadena (no varía por hotel, a diferencia del
    # correo de departamento de arriba). Ver notificacion_contactos /
    # notificacion_contacto_reglas y /api/admin/notificaciones-contactos
    # más abajo en este archivo. Igual que el correo de departamento: se
    # añade al mismo correo con copia a todos, nunca aparte, y se omite en
    # silencio si no hay ninguna regla que aplique — nunca bloquea el
    # envío. No se excluye aunque coincida con quien hizo el cambio (son
    # buzones/roles, no la persona concreta que acaba de actuar).
    #
    # (2026-08-28, ampliado el mismo día) A petición de Víctor: además de
    # la regla normal por `estado_nuevo`, si este envío en concreto viene
    # de superar el techo de gastos del mes y pasar por autorización de
    # Dirección General, también se consulta la regla especial
    # ESTADO_NOTIF_EXCESO_TECHO_DG — independiente de si ese mismo
    # contacto también está marcado para "ENVIADO AL PROVEEDOR" (el envío
    # normal, sin exceso): ambas reglas son compatibles y se acumulan, sin
    # duplicar destinatarios. Ver aprobar_expediente(), único sitio que
    # produce exactamente esta transición de estado.
    if pedido.get("departamento_id"):
        _estados_regla_buscar = [estado_nuevo]
        if estado_nuevo == "ENVIADO AL PROVEEDOR" and estado_antes == "PENDIENTE Vº Bº DIRECCIÓN GENERAL":
            _estados_regla_buscar.append(ESTADO_NOTIF_EXCESO_TECHO_DG)
        _ph_estados = ",".join(["%s"] * len(_estados_regla_buscar))
        _contactos_regla = rows_to_list(query(
            f"""SELECT DISTINCT nc.email, nc.email2
               FROM notificacion_contacto_reglas ncr
               JOIN notificacion_contactos nc ON nc.id = ncr.contacto_id
               WHERE ncr.departamento_id=%s AND ncr.estado IN ({_ph_estados}) AND nc.activo=1""",
            (pedido["departamento_id"], *_estados_regla_buscar)
        )) or []
        for _c in _contactos_regla:
            for _e in _emails_usuario(_c):
                if _e not in _todos_internos:
                    _todos_internos.append(_e)

    # ── Correo al proveedor (solo ENVIADO AL PROVEEDOR) ───────────────────────
    # Para:  todos los contactos principales del proveedor
    # BCC:   ninguno — el correo interno de más abajo (ESTADOS_EMAIL_INTERNO)
    #        ya avisa a todos los compradores y usuarios hotel del cambio de
    #        estado, así que este correo va exclusivamente al proveedor y no
    #        duplica destinatarios ni información interna.
    if estado_nuevo in ESTADOS_EMAIL_PROVEEDOR and _proveedor_emails:
        _compradores_firma = _usuarios_hotel["compradores"]
        if not (_compradores_firma and _compradores_firma[0].get("email")):
            log.warning("[EMAIL] Pedido %s: no hay comprador con email asignado al hotel %s — email a proveedor omitido",
                        pedido_id, pedido.get("hotel_codigo",""))
        else:
            _email_comprador_firma  = _compradores_firma[0]["email"]
            _nombre_comprador_firma = _compradores_firma[0].get("nombre") or ""
            _movil_comprador_firma  = _compradores_firma[0].get("movil") or ""
            _firma_contacto_html = _firma_comprador_html(_nombre_comprador_firma, _email_comprador_firma, _movil_comprador_firma)
            _firma_contacto_text = _firma_comprador_text(_nombre_comprador_firma, _email_comprador_firma, _movil_comprador_firma)
            # (2026-08-27) A petición de Víctor, las comunicaciones al proveedor
            # también deben incluir el Total Pedido cuando esté disponible,
            # dejando siempre claro que se trata de base imponible (sin IGIC).
            _total_pedido_prov = pedido.get("total_pedido")
            _fila_total_pedido_html = (
                f'<br><strong>Total Pedido (base imponible):</strong> {_fmt_importe_es(_total_pedido_prov)} €'
                if _total_pedido_prov is not None else ''
            )
            _linea_total_pedido_text = (
                f"Total Pedido (base imponible): {_fmt_importe_es(_total_pedido_prov)} €\n"
                if _total_pedido_prov is not None else ''
            )
            _nota_igic_prov_html = _nota_base_imponible_html() if _total_pedido_prov is not None else ''
            _nota_igic_prov_text = ("\n" + _nota_base_imponible_text() + "\n") if _total_pedido_prov is not None else ''
            # (2026-08-28) A petición de Víctor: en vez de adjuntar el PDF del
            # pedido al correo (EmailJS, plan Free, no admite adjuntos — ver
            # conversación del 28/08), se incluye un enlace de descarga
            # temporal del documento (ver _enlaces_descarga_pedido_doc /
            # /descargas/adjunto/<token>). Si el pedido no tiene ningún PDF
            # subido en "Nº Pedido (DALI/SAP)" todavía, no se añade nada — el
            # correo se envía igual, sin bloquear por esto.
            _enlaces_doc_prov = _enlaces_descarga_pedido_doc(pedido_id)
            _bloque_doc_html = ""
            if _enlaces_doc_prov:
                _botones_doc = "".join(
                    f'<p style="margin:6px 0"><a href="{e["url"]}" '
                    f'style="display:inline-block;background:#1a3c6e;color:#fff;text-decoration:none;'
                    f'padding:9px 18px;border-radius:5px;font-size:13px;font-weight:700">'
                    f'📄 Descargar {e["nombre"]}</a></p>'
                    for e in _enlaces_doc_prov
                )
                _bloque_doc_html = f'<div style="margin:18px 0">{_botones_doc}</div>'
            _bloque_doc_text = ""
            if _enlaces_doc_prov:
                _bloque_doc_text = "\n" + "\n".join(
                    f"Documento del pedido ({e['nombre']}): {e['url']}" for e in _enlaces_doc_prov
                ) + "\n"
            body_html = f"""
            <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
              {_email_header_html("Princess Hotels &amp; Resorts", "Dpto. Central de Compras Princess en Canarias",
                                    color_fondo="#8B0000", color_subtitulo="#f5c6c6")}
              <div style="padding:24px">
                <p style="background:#fff7e6;border:1px solid #f0c36d;color:#7a5b00;padding:10px 14px;border-radius:4px;font-size:12.5px;margin:0 0 18px">
                  ⚠️ Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
                </p>
                <p>Estimado/a proveedor/a,</p>
                <p>
                  Recientemente habrá recibido, a través de nuestro sistema habitual de pedidos, el pedido que se detalla a continuación.
                  El presente correo tiene como finalidad <strong>confirmar su recepción</strong> y solicitarle que, a la mayor brevedad posible,
                  nos indique si ha recibido dicho pedido y nos facilite la <strong>fecha estimada de entrega en el hotel</strong>.
                </p>
                <p style="background:#f5f8ff;border-left:4px solid #1a3c6e;padding:12px 16px;border-radius:0 4px 4px 0;margin:18px 0">
                  <strong>Pedido Nº:</strong> {pedido.get('pedido_num','—')}<br>
                  <strong>Hotel:</strong> {pedido.get('hotel_nombre','—')}<br>
                  <strong>Departamento:</strong> {pedido.get('departamento_nombre','—')}{_fila_total_pedido_html}
                </p>
                {_nota_igic_prov_html}
                {_bloque_doc_html}
                <p>
                  Para confirmar la recepción del pedido y facilitar la fecha estimada de entrega, por favor responda
                  a la dirección de correo que figura en la firma de este mensaje:
                  <a href="mailto:{_email_comprador_firma}">{_email_comprador_firma}</a>
                </p>
                <p>Quedamos a su disposición para cualquier consulta.<br><br>
                   Atentamente,<br>
                   {_firma_contacto_html}
                </p>
                <p style="font-size:11.5px;color:#8a6d00;background:#fff7e6;border:1px solid #f0c36d;padding:8px 12px;border-radius:4px;margin-top:14px">
                  Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
                </p>
              </div>
            </div>
            """
            body_text = (
                f"Estimado/a proveedor/a,\n\n"
                f"Recientemente habrá recibido, a través de nuestro sistema habitual de pedidos, el pedido que se detalla a continuación.\n"
                f"El presente correo tiene como finalidad confirmar su recepción y solicitarle que, a la mayor brevedad posible,\n"
                f"nos indique si ha recibido dicho pedido y nos facilite la fecha estimada de entrega en el hotel.\n\n"
                f"Pedido Nº: {pedido.get('pedido_num','—')}\n"
                f"Hotel: {pedido.get('hotel_nombre','—')}\n"
                f"Departamento: {pedido.get('departamento_nombre','—')}\n"
                f"{_linea_total_pedido_text}"
                f"{_nota_igic_prov_text}"
                f"{_bloque_doc_text}\n"
                f"Para confirmar la recepción del pedido y facilitar la fecha estimada de entrega, por favor responda\n"
                f"a la dirección de correo que figura en la firma de este mensaje: {_email_comprador_firma}\n\n"
                f"Quedamos a su disposición para cualquier consulta.\n\n"
                f"Atentamente,\n{_firma_contacto_text}\n\n"
                f"Este correo es exclusivo para notificaciones automáticas. "
                f"Por favor, responda única y exclusivamente a la dirección que firma este comunicado."
            )
            _destino_proveedor = ", ".join(_proveedor_emails)
            _log_email(db, pedido_id, "proveedor", _destino_proveedor, subject := f"Pedido Nº {pedido.get('pedido_num','—')} — Princess Hotels & Resorts", False, "Pendiente de envío vía EmailJS (encolado con retraso)")
            _encolar_email_pedido_retrasado(
                pedido_id=pedido_id,
                evento_codigo="cambio_estado_proveedor",
                destinatario=_destino_proveedor,
                asunto=subject,
                cuerpo_html=body_html,
                cuerpo_text=body_text,
                retraso_segundos=_retraso_email_estado,
            )

    # ── Correo interno (ENVIADO AL PROVEEDOR, ENTREGA PARCIAL, ENTREGADO, CANCELADO) ──
    # Para:  primer comprador del hotel
    # BCC:   resto de compradores + usuarios hotel del mismo hotel
    # Nota:  para ENVIADO AL PROVEEDOR este correo interno se manda ADEMÁS
    #        del correo al proveedor de arriba — el de arriba es la
    #        comunicación externa al proveedor (sin BCC interno desde
    #        2026-08-14, ver comentario de arriba); este es el único aviso
    #        interno propiamente dicho, con datos que nunca deben salir
    #        fuera de la empresa (quién hizo el cambio), sin duplicar
    #        destinatarios ni correos entre ambos envíos.
    if estado_nuevo in ESTADOS_EMAIL_INTERNO and _todos_internos:
        _resumen_ent = _resumen_entregas(pedido, estado_nuevo)

        # Días transcurridos desde la tramitación, para contexto de seguimiento
        _dias_transcurridos = None
        try:
            if pedido.get("fecha_tramitacion"):
                _ft = datetime.strptime(str(pedido["fecha_tramitacion"])[:10], "%Y-%m-%d").date()
                _dias_transcurridos = (datetime.now(timezone.utc).date() - _ft).days
        except Exception:
            _dias_transcurridos = None

        _importe_txt    = f"{pedido.get('importe'):.2f} €" if pedido.get('importe') is not None else '—'
        # (2026-08-27) "Importe" (arriba) es el importe de techo de gastos —
        # un campo distinto de "Total Pedido" (columna 6 del PDF SAP / valor
        # manual del comprador). A petición de Víctor se muestran ambos, cada
        # uno con su etiqueta, y siempre indicando que son base imponible.
        _total_pedido_txt = f"{_fmt_importe_es(pedido.get('total_pedido'))} €" if pedido.get('total_pedido') is not None else '—'
        _fecha_tram_txt = _fecha_es(pedido.get('fecha_tramitacion')) or '—'
        _dias_txt       = f" ({_dias_transcurridos} día(s) desde tramitación)" if _dias_transcurridos is not None else ''

        # (2026-08-28, ampliado el mismo día) A petición de Víctor: cuando
        # este envío al proveedor viene de superar el techo de gastos del
        # mes y pasar por autorización de Dirección General (misma
        # transición exacta que detecta ESTADO_NOTIF_EXCESO_TECHO_DG más
        # abajo: estado_nuevo="ENVIADO AL PROVEEDOR" con
        # estado_antes="PENDIENTE Vº Bº DIRECCIÓN GENERAL"), el propio
        # correo interno de cambio de estado debe explicarlo con claridad
        # — familia, importe, disponible y exceso en el momento de la
        # solicitud, motivo concreto de la superación y quién/cuándo lo
        # autorizó — a TODOS los destinatarios ya definidos de este correo
        # (_todos_internos), sin excepción: es información relevante para
        # todos ellos, no solo para quien lo solicitó.
        _aviso_exceso_html = ""
        _aviso_exceso_text = ""
        if estado_nuevo == "ENVIADO AL PROVEEDOR" and estado_antes == "PENDIENTE Vº Bº DIRECCIÓN GENERAL":
            _exp = row_to_dict(query(
                """SELECT e.*, f.nombre AS familia_nombre, ur.nombre AS usuario_resuelve_nombre
                   FROM expediente_exceso e
                   LEFT JOIN familias f  ON e.familia_id         = f.id
                   LEFT JOIN usuarios ur ON e.usuario_resuelve_id = ur.id
                   WHERE e.pedido_id=%s AND e.resultado='aprobado'
                   ORDER BY e.fecha_resolucion DESC NULLS LAST, e.creado_en DESC LIMIT 1""",
                (pedido_id,), one=True
            ))
            if _exp:
                _fam_exceso_txt   = _exp.get("familia_nombre") or "sin especificar"
                _motivo_exceso_txt = (_exp.get("motivo_solicitud") or "").strip() or "Supera el techo de gastos configurado."
                _disponible_txt  = f"{_fmt_importe_es(_exp.get('disponible_en_solicitud'))} €" if _exp.get('disponible_en_solicitud') is not None else '—'
                _importe_exp_txt = f"{_fmt_importe_es(_exp.get('importe_pedido'))} €" if _exp.get('importe_pedido') is not None else '—'
                _exceso_txt      = f"{_fmt_importe_es(_exp.get('exceso'))} €" if _exp.get('exceso') is not None else '—'
                _resuelve_txt    = _exp.get("usuario_resuelve_nombre") or "Dirección General"
                _fecha_resol_txt = _fecha_es(_exp.get("fecha_resolucion")) or '—'
                _obs_dg          = (_exp.get("observaciones_direccion_general") or "").strip()

                _fila_obs_dg_html = f'<tr><td style="padding:3px 6px;color:#7a5b00"><b>Nota de Dirección General</b></td><td style="padding:3px 6px;color:#7a5b00">{_obs_dg}</td></tr>' if _obs_dg else ''
                _aviso_exceso_html = f"""
                <div style="background:#fff3cd;border:1px solid #ffc107;border-radius:6px;padding:14px 16px;margin:16px 0">
                  <p style="margin:0 0 8px;color:#7a5b00;font-size:13px">
                    📉 <strong>Este pedido superó el techo de gastos mensual del hotel</strong> y ha tenido que pasar
                    por autorización de Dirección General antes de poder tramitarse y enviarse al proveedor.
                  </p>
                  <table style="width:100%;border-collapse:collapse;font-size:12.5px">
                    <tr><td style="padding:3px 6px;color:#7a5b00"><b>Familia</b></td><td style="padding:3px 6px;color:#7a5b00">{_fam_exceso_txt}</td></tr>
                    <tr><td style="padding:3px 6px;color:#7a5b00"><b>Motivo de la superación</b></td><td style="padding:3px 6px;color:#7a5b00">{_motivo_exceso_txt}</td></tr>
                    <tr><td style="padding:3px 6px;color:#7a5b00"><b>Disponible en el momento de la solicitud</b></td><td style="padding:3px 6px;color:#7a5b00">{_disponible_txt}</td></tr>
                    <tr><td style="padding:3px 6px;color:#7a5b00"><b>Importe de este pedido</b></td><td style="padding:3px 6px;color:#7a5b00">{_importe_exp_txt}</td></tr>
                    <tr><td style="padding:3px 6px;color:#7a5b00"><b>Exceso sobre el techo</b></td><td style="padding:3px 6px;color:#b91c1c;font-weight:700">{_exceso_txt}</td></tr>
                    <tr><td style="padding:3px 6px;color:#7a5b00"><b>Autorizado por</b></td><td style="padding:3px 6px;color:#7a5b00">{_resuelve_txt} — {_fecha_resol_txt}</td></tr>
                    {_fila_obs_dg_html}
                  </table>
                </div>"""

                _linea_obs_dg_text = f"\n   Nota de Dirección General: {_obs_dg}" if _obs_dg else ""
                _aviso_exceso_text = (
                    "\n\n⚠️  EXCESO DE TECHO DE GASTOS — AUTORIZADO POR DIRECCIÓN GENERAL\n"
                    "   Este pedido superó el techo de gastos mensual del hotel y ha tenido que pasar\n"
                    "   por autorización de Dirección General antes de poder tramitarse y enviarse al proveedor.\n"
                    f"   Familia:                        {_fam_exceso_txt}\n"
                    f"   Motivo de la superación:        {_motivo_exceso_txt}\n"
                    f"   Disponible en la solicitud:     {_disponible_txt}\n"
                    f"   Importe de este pedido:         {_importe_exp_txt}\n"
                    f"   Exceso sobre el techo:          {_exceso_txt}\n"
                    f"   Autorizado por:                 {_resuelve_txt} — {_fecha_resol_txt}"
                    f"{_linea_obs_dg_text}"
                )
        # (2026-08-31, v12.30.63) Víctor: "Si el pedido supera el techo de
        # gastos establecido, indícalo explícitamente en el texto para que
        # no pase desapercibido" — el recuadro amarillo de detalle
        # (_aviso_exceso_html/_aviso_exceso_text, justo debajo) ya existía,
        # pero el propio párrafo introductorio no lo mencionaba. Esta
        # bandera se usa más abajo para añadir una frase de aviso corta en
        # el propio párrafo, sin repetir el detalle que ya da el recuadro.
        _hubo_exceso_techo = bool(_aviso_exceso_html)

        # (2026-08-31) A petición de Víctor: el correo AL PROVEEDOR ya tiene,
        # desde el 2026-08-28, un botón de descarga del PDF del pedido (ver
        # _enlaces_descarga_pedido_doc más arriba) — pero el correo INTERNO
        # de este mismo cambio de estado (el de más abajo, ESTADOS_EMAIL_INTERNO)
        # nunca lo tuvo, así que quien lo recibe (compradores/usuarios hotel)
        # no tiene forma de ver el PDF sin entrar a la app. Víctor: "no
        # habíamos modificado tanto el correo interno de comunicación estado
        # ENVIADO AL PROVEEDOR como el que se envía al mismo proveedor para
        # este asunto, para que adjúntense un botón y poder descargar el PDF
        # del pedido en destino?" — no, solo se había hecho para el correo al
        # proveedor (ver CHANGELOG v12.30.40); esto añade el mismo botón
        # aquí, con el mismo enlace público y temporal (sin login, ver
        # /descargas/adjunto/<token>), inicialmente solo para ENVIADO AL
        # PROVEEDOR — igual que el correo al proveedor.
        # (2026-09-02) A petición de Víctor: extender el mismo botón también
        # a ENTREGA PARCIAL y ENTREGADO (entrega total) — mismo pedido PDF,
        # mismo enlace público/temporal, solo cambia el texto introductorio
        # para no hablar de "enviado al proveedor" en esos dos estados. Se
        # mantiene sin botón para CANCELADO/DENEGADO POR DIRECCION GENERAL,
        # donde no aporta nada revisar el PDF del pedido.
        _bloque_doc_html_interno = ""
        _bloque_doc_text_interno = ""
        _ESTADOS_CON_BOTON_DOC = ("ENVIADO AL PROVEEDOR", "ENTREGA PARCIAL", "ENTREGADO")
        if estado_nuevo in _ESTADOS_CON_BOTON_DOC:
            _enlaces_doc_interno = _enlaces_descarga_pedido_doc(pedido_id)
            if _enlaces_doc_interno:
                _botones_doc_interno = "".join(
                    f'<p style="margin:6px 0"><a href="{e["url"]}" '
                    f'style="display:inline-block;background:#1a3c6e;color:#fff;text-decoration:none;'
                    f'padding:9px 18px;border-radius:5px;font-size:13px;font-weight:700">'
                    f'📄 Descargar {e["nombre"]}</a></p>'
                    for e in _enlaces_doc_interno
                )
                # (2026-08-31) A petición de Víctor: "en todo momento dar las
                # instrucciones pertinentes para que se puedan descargar el
                # pedido PDF con el botón al uso" — el botón aparecía solo,
                # sin ninguna frase que explique qué es o para qué sirve.
                if estado_nuevo == "ENVIADO AL PROVEEDOR":
                    _texto_intro_doc_html = (
                        'Puede descargar el documento del pedido tramitado y enviado al proveedor '
                        'pulsando el siguiente botón:'
                    )
                    _texto_intro_doc_text = (
                        "Puede descargar el documento del pedido tramitado y enviado al proveedor "
                        "en el siguiente enlace:"
                    )
                else:
                    _texto_intro_doc_html = (
                        'Puede descargar el documento del pedido pulsando el siguiente botón:'
                    )
                    _texto_intro_doc_text = (
                        "Puede descargar el documento del pedido en el siguiente enlace:"
                    )
                _bloque_doc_html_interno = (
                    f'<p style="margin:14px 0 4px">{_texto_intro_doc_html}</p>'
                    f'<div style="margin:4px 0 14px">{_botones_doc_interno}</div>'
                )
                _bloque_doc_text_interno = (
                    f"\n{_texto_intro_doc_text}\n"
                    + "\n".join(f"Documento del pedido ({e['nombre']}): {e['url']}" for e in _enlaces_doc_interno)
                    + "\n"
                )

        # (2026-08-28) A petición de Víctor: cuando el cambio es automático
        # (es_automatico=True — decidido por _aplicar_coincidencia_albaran()
        # al confirmar una coincidencia de "Comparar Pedidos + Albaranes",
        # nunca por una persona en ese momento concreto), "Realizado por" NO
        # debe mostrar el nombre de quien tenía la sesión abierta cuando se
        # disparó la comparación/confirmación — induce a pensar que esa
        # persona ha tramitado el pedido a mano, cuando en realidad lo ha
        # decidido el cruce automático de listados. Se sustituye por una
        # etiqueta de sistema con fecha y hora del cierre.
        if es_automatico:
            import pytz
            _usuario_txt = (
                "Cierre automático — comparación de listados "
                f"({datetime.now(pytz.timezone('Atlantic/Canary')).strftime('%d/%m/%Y %H:%M')})"
            )
        else:
            _usuario_txt = (usuario_nombre or '').strip()

        _ICONO_ESTADO = {
            "ENVIADO AL PROVEEDOR":            "📤",
            "ENTREGA PARCIAL":                 "📦",
            "ENTREGADO":                       "✅",
            "CANCELADO":                       "❌",
            "DENEGADO POR DIRECCION GENERAL":  "🚫",
        }
        _icono = _ICONO_ESTADO.get(estado_nuevo, "🔔")

        _INTRO_ESTADO = {
            "CANCELADO":            "El pedido ha sido <strong>CANCELADO</strong>.",
            "DENEGADO POR DIRECCION GENERAL": "El pedido ha sido <strong>DENEGADO POR DIRECCIÓN GENERAL</strong>.",
        }
        # (2026-09-02) A petición de Víctor: el párrafo introductorio de
        # ENTREGA PARCIAL debe indicar por qué importe es esta entrega
        # concreta y cuánto queda pendiente sobre el total del pedido; el de
        # ENTREGADO debe confirmar la entrega total y el nº de días
        # transcurridos desde la tramitación del pedido, contando también
        # las entregas parciales intermedias si las hubo. Los importes y
        # días salen de _resumen_ent (ver _resumen_entregas más arriba).
        _entregas_ent = _resumen_ent["entregas"]
        _monto_esta_entrega = None
        for _e in reversed(_entregas_ent):
            if _e["base_imponible"] is not None:
                _monto_esta_entrega = _e["base_imponible"]
                break
        if estado_nuevo == "ENTREGA PARCIAL":
            _intro_parcial_html = "Se ha registrado una <strong>entrega parcial</strong>"
            if _monto_esta_entrega is not None:
                _intro_parcial_html += f" por un total de <strong>{_fmt_importe_es(_monto_esta_entrega)} €</strong>"
            _intro_parcial_html += " en este pedido."
            if _resumen_ent.get("total_pendiente") is not None:
                _intro_parcial_html += (
                    f' Queda pendiente la entrega de un total de '
                    f'<strong>{_fmt_importe_es(_resumen_ent["total_pendiente"])} €</strong> '
                    f'sobre el pedido adjunto (total del pedido: {_fmt_importe_es(_resumen_ent["total_pedido"])} €).'
                )
            _intro_parcial_html += " A continuación se detalla el histórico de entregas recibidas hasta la fecha."
            _INTRO_ESTADO["ENTREGA PARCIAL"] = _intro_parcial_html
        elif estado_nuevo == "ENTREGADO":
            _intro_entregado_html = "Queda <strong>confirmada la entrega total</strong> del pedido."
            _n_parciales = sum(1 for _e in _entregas_ent if not _e["es_final"])
            if _resumen_ent.get("dias_pedido_a_final") is not None:
                _intro_entregado_html += (
                    f' Han transcurrido <strong>{_resumen_ent["dias_pedido_a_final"]} día(s)</strong> entre la '
                    f'fecha de tramitación del pedido y la entrega total'
                )
                if _n_parciales:
                    _intro_entregado_html += f", con {_n_parciales} entrega(s) parcial(es) intermedia(s)"
                _intro_entregado_html += "."
            _intro_entregado_html += (
                " A continuación se detalla el histórico completo de entregas, incluyendo la fecha de la entrega final."
            )
            _INTRO_ESTADO["ENTREGADO"] = _intro_entregado_html
        # (2026-08-31) A petición de Víctor: este correo interno de "ENVIADO
        # AL PROVEEDOR" cambia de enfoque — antes era un aviso genérico de
        # "cambio de estado" (con Estado anterior/Estado nuevo, ver más
        # abajo, ahora retirados solo para este caso); ahora se usa
        # específicamente para que el departamento que hizo el pedido (y,
        # si aplica, A&B) sepan que ya se tramitó y se envió al proveedor, y
        # que entran en espera de entrega. Víctor: "por la presente se
        # informa al responsable del Dpto. X que su pedido ha sido
        # tramitado correctamente al proveedor y entramos en el proceso de
        # espera para la entrega, que informaremos de cualquier otra
        # novedad" — y, para los departamentos de cocina/sala ("en los
        # casos de que el departamento sea COCINA, BARES, RESTAURANTE Y/O
        # RESTAURANTE & BARES"), añadir que también se comunica a A&B para
        # su control. Nombres exactos de departamento tomados de
        # models.py/SQL_STATEMENTS (semilla de la tabla departamentos).
        # (_DEPARTAMENTOS_AB / _dept_nombre_i ya están definidos más arriba,
        # justo tras cargar `pedido` — se reutilizan aquí.)
        _proveedor_nombre_i = pedido.get('proveedor_nombre') or '—'
        # (2026-08-31) Víctor pidió un texto más conciso y corporativo (4-5
        # líneas), que confirme la tramitación, nombre al proveedor en la
        # propia redacción (no solo en la tabla) y mantenga el aviso a A&B
        # para Cocina/Bares/Restaurante(s). El resto (cuadro de datos y
        # aviso de exceso de techo, ya en copia a Notificaciones
        # Adicionales) se deja tal cual: "El cuadro esta perfecto".
        # (2026-08-31, ajuste tras ver el resultado) Víctor: la coletilla
        # "Por la presente..." no convencía del todo — se sustituye por un
        # "Confirmamos que..." más directo. El aviso a A&B se deja en una
        # sola frase de mero trámite ("para su control interno"), sin
        # justificar por qué se le informa (antes explicaba "por tratarse
        # de un pedido de X", información que ya está en la fila
        # Departamento de la tabla).
        if estado_nuevo == "ENVIADO AL PROVEEDOR":
            _intro_html = (
                f"Confirmamos que el pedido ha sido tramitado y enviado correctamente al proveedor "
                f"<strong>{_proveedor_nombre_i}</strong>, quedando informado el departamento de "
                f"<strong>{_dept_nombre_i or '—'}</strong>."
            )
            if _dept_nombre_i.upper() in _DEPARTAMENTOS_AB:
                _intro_html += (
                    ' Se informa también al departamento de <strong>A&amp;B</strong> para su control interno.'
                )
            if _hubo_exceso_techo:
                _intro_html += (
                    ' <strong style="color:#7a5b00">Este pedido superó el techo de gastos mensual y fue '
                    'autorizado por Dirección General</strong> (detalle más abajo).'
                )
            _intro_html += ' Cualquier novedad sobre la entrega se comunicará en su momento.'
        else:
            _intro_html = _INTRO_ESTADO.get(estado_nuevo, "")
        _intro_html_block = f'<p style="margin:0 0 16px;line-height:1.55;color:#222">{_intro_html}</p>' if _intro_html else ""

        subject_i = f"[Control Pedidos] {pedido.get('hotel_codigo','')} · Pedido {pedido.get('pedido_num','—')} → {_icono} {estado_nuevo}"
        if estado_nuevo == "ENTREGADO" and _resumen_ent["ultima_fecha_es"]:
            subject_i += f" ({_resumen_ent['ultima_fecha_es']})"
        elif estado_nuevo == "ENTREGA PARCIAL" and _resumen_ent["ultima_fecha_es"]:
            subject_i += f" — última entrega {_resumen_ent['ultima_fecha_es']}"

        _fila_usuario_html = f'<tr><td><b>Realizado por</b></td><td>{_usuario_txt}</td></tr>' if _usuario_txt else ''
        # (2026-08-31) Víctor: "podemos incluir en el cuadro el apartado
        # observaciones que ya tenemos en pedidos? esto siempre puede dar
        # mas información relevante" — se añade como última fila, cuando
        # el pedido tiene observaciones. Se omite para CANCELADO/DENEGADO
        # POR DIRECCION GENERAL porque ese mismo campo (pedido.observaciones)
        # ya se muestra ahí, aparte de la tabla, como "Motivo de la
        # cancelación/denegación" (ver _motivo_estado más abajo) — mostrarlo
        # también en la tabla sería puro duplicado.
        _fila_obs_html = ''
        if pedido.get('observaciones') and estado_nuevo not in ("CANCELADO", "DENEGADO POR DIRECCION GENERAL"):
            _fila_obs_html = (
                f'<tr><td><b>Observaciones</b></td><td>{pedido["observaciones"].replace(chr(10), "<br>")}</td></tr>'
            )
        # (2026-08-31) Estado anterior/Estado nuevo ya no aportan nada en el
        # caso ENVIADO AL PROVEEDOR — el correo entero ya trata justo de
        # eso (ver _intro_html arriba), así que sobra repetirlo en la
        # tabla. Se mantienen sin cambios para el resto de estados
        # (ENTREGA PARCIAL/ENTREGADO/CANCELADO/DENEGADO), donde sí importa
        # dejar constancia de qué estado había antes.
        _filas_estado_html = (
            ''
            if estado_nuevo == "ENVIADO AL PROVEEDOR" else
            f'<tr><td><b>Estado anterior</b></td><td>{estado_antes or "—"}</td></tr>\n'
            f'          <tr><td><b>Estado nuevo</b></td><td><b>{estado_nuevo}</b></td></tr>'
        )
        body_html_i = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
          {_email_header_html("Princess Hotels &amp; Resorts", "Control de Pedidos — Aviso interno",
                                color_fondo="#1a3a6b", color_subtitulo="#a8c0e8")}
          <div style="padding:24px">
        <p style="font-size:15px;margin:0 0 16px"><strong>{_icono} {estado_nuevo}</strong> — Pedido {pedido.get('pedido_num','—')} · {pedido.get('hotel_codigo','')}</p>
        {_intro_html_block}
        {_aviso_exceso_html}
        <table border="1" cellpadding="8" style="border-collapse:collapse;font-family:sans-serif;font-size:13px;line-height:1.4;margin:4px 0 16px">
          <tr><td><b>Hotel</b></td><td>{pedido.get('hotel_nombre','')} ({pedido.get('hotel_codigo','')})</td></tr>
          <tr><td><b>Departamento</b></td><td>{pedido.get('departamento_nombre','')}</td></tr>
          <tr><td><b>Pedido Nº</b></td><td>{pedido.get('pedido_num','—')}</td></tr>
          <tr><td><b>Presupuesto Nº</b></td><td>{pedido.get('presupuesto_num') or '—'}</td></tr>
          <tr><td><b>Proveedor</b></td><td>{pedido.get('proveedor_nombre','—')}</td></tr>
          <tr><td><b>Importe (techo de gastos)</b></td><td>{_importe_txt}</td></tr>
          <tr><td><b>Total Pedido (base imponible)</b></td><td>{_total_pedido_txt}</td></tr>
          {_filas_estado_html}
          <tr><td><b>Fecha tramitación</b></td><td>{_fecha_tram_txt}{_dias_txt}</td></tr>
          {_fila_usuario_html}
          {_fila_obs_html}
        </table>
        {_bloque_doc_html_interno}
        {_nota_base_imponible_html() if not _resumen_ent["entregas"] else ''}
        {_html_bloque_entregas(_resumen_ent, estado_nuevo)}
        """
        if estado_nuevo in ("CANCELADO", "DENEGADO POR DIRECCION GENERAL") and _motivo_estado:
            _label_motivo = "Motivo de la denegación" if estado_nuevo == "DENEGADO POR DIRECCION GENERAL" else "Motivo de la cancelación"
            _motivo_html = _motivo_estado.replace("\n", "<br>")
            body_html_i += f'<p style="margin-top:14px"><b>{_label_motivo}:</b><br>{_motivo_html}</p>'
        # (2026-08-31) A petición de Víctor: "me gustaría que no fuera en
        # oculto, es interesante que todos los involucrados sepan quiénes
        # están informados" — este correo interno pasa de ir en Bcc a ir en
        # CC visible (ver cambio en _enviarEmailsSistemaPendientes, en
        # templates/index.html: para evento_codigo == "cambio_estado_
        # interno" ahora se manda como `cc` en vez de `bcc` a EmailJS — el
        # resto de correos de esta misma cola, p. ej. la reclamación
        # automática al proveedor, siguen en Bcc sin cambios). Como red de
        # seguridad — por si la plantilla de EmailJS todavía no tiene un
        # campo "cc" enlazado a un encabezado real de copia visible (hace
        # falta añadirlo a mano en las 3 cuentas, ver Admin → EmailJS) — se
        # lista también aquí, dentro del propio cuerpo, a quién se ha
        # enviado el aviso: así el objetivo de Víctor (transparencia sobre
        # quién está informado) se cumple aunque el encabezado CC de
        # EmailJS falle o tarde en configurarse.
        _destinatarios_aviso_txt = ", ".join(_todos_internos) if _todos_internos else "—"
        body_html_i += (
            '<p style="margin-top:14px;font-size:11.5px;color:#888">'
            f'<b>Aviso enviado también a:</b> {_destinatarios_aviso_txt}'
            '</p>'
        )
        body_html_i += '<p style="margin-top:8px;font-size:11.5px;color:#888">Aviso automático del sistema de Control de Pedidos — Princess Hotels &amp; Resorts.</p>'
        body_html_i += '</div></div>'

        _INTRO_ESTADO_TXT = {
            "CANCELADO":            "El pedido ha sido CANCELADO.",
            "DENEGADO POR DIRECCION GENERAL": "El pedido ha sido DENEGADO POR DIRECCIÓN GENERAL.",
        }
        # (2026-09-02) Mismo contenido dinámico que en _INTRO_ESTADO (HTML)
        # más arriba, en texto plano — ver comentario allí.
        if estado_nuevo == "ENTREGA PARCIAL":
            _intro_parcial_text = "Se ha registrado una entrega parcial"
            if _monto_esta_entrega is not None:
                _intro_parcial_text += f" por un total de {_fmt_importe_es(_monto_esta_entrega)} €"
            _intro_parcial_text += " en este pedido."
            if _resumen_ent.get("total_pendiente") is not None:
                _intro_parcial_text += (
                    f' Queda pendiente la entrega de un total de {_fmt_importe_es(_resumen_ent["total_pendiente"])} € '
                    f'sobre el pedido adjunto (total del pedido: {_fmt_importe_es(_resumen_ent["total_pedido"])} €).'
                )
            _intro_parcial_text += " A continuación se detalla el histórico de entregas recibidas hasta la fecha."
            _INTRO_ESTADO_TXT["ENTREGA PARCIAL"] = _intro_parcial_text
        elif estado_nuevo == "ENTREGADO":
            _intro_entregado_text = "Queda confirmada la entrega total del pedido."
            _n_parciales_txt = sum(1 for _e in _entregas_ent if not _e["es_final"])
            if _resumen_ent.get("dias_pedido_a_final") is not None:
                _intro_entregado_text += (
                    f' Han transcurrido {_resumen_ent["dias_pedido_a_final"]} día(s) entre la fecha de '
                    f'tramitación del pedido y la entrega total'
                )
                if _n_parciales_txt:
                    _intro_entregado_text += f", con {_n_parciales_txt} entrega(s) parcial(es) intermedia(s)"
                _intro_entregado_text += "."
            _intro_entregado_text += (
                " A continuación se detalla el histórico completo de entregas, incluyendo la fecha de la entrega final."
            )
            _INTRO_ESTADO_TXT["ENTREGADO"] = _intro_entregado_text
        if estado_nuevo == "ENVIADO AL PROVEEDOR":
            _intro_text = (
                f"Confirmamos que el pedido ha sido tramitado y enviado correctamente al proveedor "
                f"{_proveedor_nombre_i}, quedando informado el departamento de {_dept_nombre_i or '—'}."
            )
            if _dept_nombre_i.upper() in _DEPARTAMENTOS_AB:
                _intro_text += " Se informa también al departamento de A&B para su control interno."
            if _hubo_exceso_techo:
                _intro_text += (
                    " ATENCIÓN: este pedido superó el techo de gastos mensual y fue autorizado por "
                    "Dirección General (detalle más abajo)."
                )
            _intro_text += " Cualquier novedad sobre la entrega se comunicará en su momento."
        else:
            _intro_text = _INTRO_ESTADO_TXT.get(estado_nuevo, "")

        _SEP = "─" * 42
        _datos_lineas = [
            f"   Hotel:              {pedido.get('hotel_nombre','')} ({pedido.get('hotel_codigo','')})",
            f"   Departamento:       {pedido.get('departamento_nombre','')}",
            f"   Pedido Nº:          {pedido.get('pedido_num','—')}",
            f"   Presupuesto Nº:     {pedido.get('presupuesto_num') or '—'}",
            f"   Proveedor:          {pedido.get('proveedor_nombre','—')}",
            f"   Importe (techo):    {_importe_txt}",
            f"   Total Pedido (base imponible): {_total_pedido_txt}",
        ]
        # (2026-08-31) Igual que en la tabla HTML: Estado anterior/Estado
        # nuevo se omiten solo para ENVIADO AL PROVEEDOR (ver _filas_estado_html).
        if estado_nuevo != "ENVIADO AL PROVEEDOR":
            _datos_lineas.append(f"   Estado anterior:    {estado_antes or '—'}")
            _datos_lineas.append(f"   Estado nuevo:       {estado_nuevo}")
        _datos_lineas.append(f"   Fecha tramitación:  {_fecha_tram_txt}{_dias_txt}")
        if _usuario_txt:
            _datos_lineas.append(f"   Realizado por:      {_usuario_txt}")
        # (2026-08-31) Observaciones — mismo criterio que en la tabla HTML:
        # se omite para CANCELADO/DENEGADO POR DIRECCION GENERAL porque ya
        # se muestra aparte como "Motivo de la cancelación/denegación".
        if pedido.get('observaciones') and estado_nuevo not in ("CANCELADO", "DENEGADO POR DIRECCION GENERAL"):
            _datos_lineas.append(f"   Observaciones:      {pedido['observaciones']}")

        body_text_i = (
            f"{_icono} {estado_nuevo} — Pedido {pedido.get('pedido_num','—')} · {pedido.get('hotel_codigo','')}\n"
            f"{_SEP}\n\n"
            + (f"{_intro_text}\n\n" if _intro_text else "")
            + (f"{_aviso_exceso_text.strip()}\n\n" if _aviso_exceso_text else "")
            + "📋 Datos del pedido\n"
            + "\n".join(_datos_lineas)
            + _bloque_doc_text_interno
        )
        if not _resumen_ent["entregas"]:
            body_text_i += "\n\n" + _nota_base_imponible_text()
        _bloque_text_ent = _text_bloque_entregas(_resumen_ent, estado_nuevo)
        if _bloque_text_ent:
            body_text_i += "\n\n📦 " + _bloque_text_ent
        if estado_nuevo in ("CANCELADO", "DENEGADO POR DIRECCION GENERAL") and _motivo_estado:
            _label_motivo_txt = "Motivo de la denegación" if estado_nuevo == "DENEGADO POR DIRECCION GENERAL" else "Motivo de la cancelación"
            body_text_i += f"\n\n{_label_motivo_txt}:\n{_motivo_estado}"
        # (ver comentario equivalente en body_html_i, más arriba)
        body_text_i += f"\n\nAviso enviado también a: {_destinatarios_aviso_txt}"
        body_text_i += f"\n\n{_SEP}\nAviso automático del sistema de Control de Pedidos — Princess Hotels & Resorts."

        for dest in _todos_internos:
            _log_email(db, pedido_id, "interno", dest, subject_i, False, "Pendiente de envío vía EmailJS (encolado con retraso)")
        _encolar_email_pedido_retrasado(
            pedido_id=pedido_id,
            evento_codigo="cambio_estado_interno",
            destinatario=_todos_internos[0],
            asunto=subject_i,
            cuerpo_html=body_html_i,
            cuerpo_text=body_text_i,
            cc_emails=",".join(_todos_internos[1:]),
            marca_comunicado_ab=_incluye_ab_email,
            marca_comunicado_jefe_dep=_incluye_jefe_depto_email,
            retraso_segundos=_retraso_email_estado,
        )

    return pendientes

# ── Helper norden ──────────────────────────────────────────────────────────────

# La asignación ya no está hardcodeada: se gestiona desde el panel de admin
# en Usuarios → sección "Hoteles asignados (compras)".
# La función _get_compradores_hotel(hotel_codigo) sustituye al antiguo diccionario
# HOTEL_COMPRADOR y lee en tiempo real qué compradores tienen ese hotel asignado.

# ── Telegram Bot — alertas automáticas ─────────────────────────────────────────
TELEGRAM_BOT_TOKEN      = os.environ.get("TELEGRAM_BOT_TOKEN", "")
# ADMIN_TELEGRAM_CHAT_ID eliminado: los chat_id se gestionan desde el panel de admin (campo telegram_chat_id en usuario)

# ── Tipos de alerta que generan copia de supervisión a los administradores ──────
# "urgente"       → job diario con nivel urgente (pedidos críticos parados)
# "techo"         → alerta de techo de gastos (supervisión financiera)
# ── Configuración de Avisos (v12.4.0) ──────────────────────────────────────────
# Sustituye TIPOS_SUPERVISION_ADMIN y "todos los admins con Telegram/email" por
# una configuración editable en Administrador → Configuración de Avisos (tabla
# config_avisos). Cada evento/causa tiene su propia lista de destinatarios y,
# por destinatario, qué canal(es) recibe. Si nadie está configurado para un
# evento, simplemente no se envía nada — ya no hay un fallback "todos los
# admins reciben todo".

def _resolver_notificacion(evento_codigo: str, canal: str, hotel_id: int = None) -> list:
    """
    Única fuente de verdad para "quién recibe qué evento, por qué canal, en
    qué hotel" (v12.17.0) — sustituye tanto a la _destinatarios_evento()
    original (eventos globales de supervisión admin) como a las funciones
    hardcodeadas _get_compradores_hotel()/_get_usuarios_hotel_rol_telegram()
    (avisos operativos ligados a un pedido/hotel concreto). Todo se gestiona
    ahora desde Administrador → Configuración de Avisos, tabla
    notificaciones_config.

    canal: 'telegram' | 'email' | 'popup'
    hotel_id: None para eventos globales (requiere_hotel=FALSE en
              eventos_aviso); id de hotel para eventos operativos ligados a
              un pedido concreto (requiere_hotel=TRUE).
    Cada dict incluye: id, username, nombre, email, telegram_chat_id, rol.
    """
    columna = {"telegram": "telegram", "email": "email", "popup": "popup"}.get(canal, "telegram")
    try:
        if hotel_id is not None:
            rows = rows_to_list(query(
                f"""SELECT u.id, u.username, u.nombre, u.email, u.telegram_chat_id, u.rol
                    FROM notificaciones_config nc
                    JOIN usuarios u ON u.id = nc.usuario_id
                    WHERE nc.evento_codigo = %s AND nc.hotel_id = %s
                      AND nc.{columna} = TRUE AND u.activo = 1""",
                (evento_codigo, hotel_id)
            )) or []
        else:
            rows = rows_to_list(query(
                f"""SELECT u.id, u.username, u.nombre, u.email, u.telegram_chat_id, u.rol
                    FROM notificaciones_config nc
                    JOIN usuarios u ON u.id = nc.usuario_id
                    WHERE nc.evento_codigo = %s AND nc.hotel_id IS NULL
                      AND nc.{columna} = TRUE AND u.activo = 1""",
                (evento_codigo,)
            )) or []
        return rows
    except Exception as exc:
        log.error("[NOTIF-CONFIG] Error resolviendo evento=%s canal=%s hotel_id=%s: %s",
                  evento_codigo, canal, hotel_id, exc)
        return []


def _destinatarios_evento(evento_codigo: str, canal: str) -> list:
    """
    Compatibilidad: todos los eventos globales de supervisión admin siguen
    llamando a esta función tal cual — ahora es un envoltorio fino sobre
    _resolver_notificacion() con hotel_id=None.
    """
    return _resolver_notificacion(evento_codigo, canal, hotel_id=None)


def _destinatarios_evento_emails(evento_codigo: str) -> list:
    """Atajo: lista de emails (strings) configurados para el evento indicado."""
    return [d["email"] for d in _destinatarios_evento(evento_codigo, "email") if d.get("email")]


def _get_solo_admin_emails() -> list:
    """
    Compatibilidad: antes devolvía TODOS los admins con email en BD.
    Ahora devuelve los emails configurados para el evento 'solicitud_acceso'
    en Administrador → Configuración de Avisos.
    """
    return _destinatarios_evento_emails("solicitud_acceso")


def _html_a_texto_plano(html: str) -> str:
    """
    (2026-07-30) Conversión básica de HTML a texto plano — red de
    seguridad para cuando se encola un email de sistema sin una versión
    de texto explícita.

    (2026-07-31) La plantilla de EmailJS ya usa `{{{{message}}}}` (triple
    llave, sin escapar), así que el frontend ahora prioriza
    `message: p.cuerpo_html || p.cuerpo_text || ''` — el HTML SÍ se
    renderiza como email real. `cuerpo_text` deja de ser lo que ve el
    destinatario en el caso normal; queda como respaldo para el caso
    (raro) de que `cuerpo_html` también falte, y esta función sigue
    generándolo solo, aunque un caller se olvide de pasarlo explícito.

    (2026-07-31) FIX: los `body_html` de las plantillas son f-strings
    Python escritas en varias líneas con indentación — esos saltos de
    línea/espacios "de formato del código fuente" son ruido invisible en
    HTML (el navegador los ignora), pero antes esta función los dejaba
    intactos y ADEMÁS insertaba su propio '\\n' al convertir cada
    `<br>`/`</p>` — el resultado eran líneas en blanco dobles/triples
    entre casi cualquier par de líneas ("muy desorganizado", reportado
    por el usuario). Ahora se colapsa TODO el espacio en blanco crudo
    (incluidos saltos de línea) a un único espacio ANTES de insertar los
    saltos de línea con significado — igual que hace un navegador al
    renderizar HTML — así el único origen de saltos de línea en el
    resultado son las etiquetas que realmente los representan.
    """
    if not html:
        return ""
    # 1) Neutralizar el formato del código fuente: todo salto de línea /
    #    tabulación / espacios repetidos de la plantilla Python se colapsa
    #    a un único espacio, igual que haría un navegador con el HTML.
    texto = re.sub(r'\s+', ' ', html)
    # 2) Insertar los saltos de línea con significado real:
    #    - <br> → salto de línea simple
    #    - cierre de párrafo/bloque/título → salto de línea doble (línea en blanco)
    #    - cierre de fila/elemento de lista → salto de línea simple
    texto = re.sub(r'<br\s*/?>', '\n', texto, flags=re.IGNORECASE)
    texto = re.sub(r'</p>|</div>|</h[1-6]>', '\n\n', texto, flags=re.IGNORECASE)
    texto = re.sub(r'</tr>|</li>', '\n', texto, flags=re.IGNORECASE)
    texto = re.sub(r'<[^>]+>', '', texto)
    texto = _html_unescape(texto)
    # 3) Limpieza final: espacios sueltos alrededor de cada salto de línea,
    #    y como mucho una línea en blanco seguida entre bloques.
    texto = re.sub(r'[ \t]+', ' ', texto)
    texto = re.sub(r' *\n *', '\n', texto)
    texto = re.sub(r'\n{3,}', '\n\n', texto)
    return texto.strip()


def _encolar_email_sistema(evento_codigo: str, destinatarios_email: list,
                           asunto: str, cuerpo_html: str = None, cuerpo_text: str = None,
                           solicitud_acceso_id: int = None,
                           cc_emails: list = None, pedido_id: int = None) -> None:
    """
    Encola un email de sistema para cada destinatario. Estos avisos se
    generan desde jobs sin navegador abierto (APScheduler), y esta app no
    tiene SMTP propio en el backend — el envío real depende de EmailJS en
    el navegador (igual que el resto de emails de la aplicación). Por eso
    se encolan en emails_sistema_pendientes: el primer admin que abra la
    app los envía en segundo plano.

    solicitud_acceso_id (opcional): vincula la fila a una solicitud de acceso
    concreta, para que el panel admin pueda mostrar si su email de Fase 2
    ya se despachó.

    cc_emails (opcional, v12.19.0): lista de emails en copia (p.ej. los
    compradores del hotel cuando el destinatario es un proveedor), se
    guarda como string separado por comas y se envía como bcc en EmailJS.
    pedido_id (opcional, v12.19.0): vincula la fila a un pedido concreto
    (reclamaciones automáticas a proveedor) para trazabilidad.
    """
    if not destinatarios_email:
        return
    if not cuerpo_text and cuerpo_html:
        cuerpo_text = _html_a_texto_plano(cuerpo_html)
    cc_str = ",".join([e for e in (cc_emails or []) if e]) or None
    try:
        db = get_db()
        cur = db.cursor()
        for email in destinatarios_email:
            cur.execute(
                """INSERT INTO emails_sistema_pendientes
                   (evento_codigo, destinatario, asunto, cuerpo_html, cuerpo_text,
                    solicitud_acceso_id, cc_emails, pedido_id)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                (evento_codigo, email, asunto, cuerpo_html, cuerpo_text,
                 solicitud_acceso_id, cc_str, pedido_id)
            )
        db.commit()
    except Exception as exc:
        log.error("[CONFIG-AVISOS] Error encolando email sistema evento=%s: %s", evento_codigo, exc)


def _notificar_evento(evento_codigo: str, texto_telegram: str,
                      titulo_bridge: str = None, pedido_id_bridge: int = None,
                      nivel_bridge: str = "urgente", tipo_bridge: str = "supervision",
                      asunto_email: str = None, cuerpo_email_html: str = None,
                      cuerpo_email_text: str = None) -> None:
    """
    Punto único de envío para avisos de sistema/administración configurables.
    Resuelve destinatarios desde notificaciones_config y despacha por
    Telegram, popup (Agenda/bridge) y email — los tres canales, de forma
    independiente entre sí (v12.17.0: antes el popup viajaba siempre pegado
    a la lista de Telegram; ahora cada uno tiene su propio checkbox en el
    panel de admin).
    """
    destinatarios_tg = _destinatarios_evento(evento_codigo, "telegram")
    for dest in destinatarios_tg:
        chat_id  = dest.get("telegram_chat_id")
        username = dest.get("username", "?")
        if chat_id:
            res = _send_telegram(chat_id, texto_telegram)
            log.info("[AVISO-%s] Telegram → %s (%s): %s",
                     evento_codigo, username, chat_id, "OK" if res.get("ok") else res.get("error"))
        else:
            log.warning("[AVISO-%s] %s configurado para Telegram pero sin telegram_chat_id", evento_codigo, username)

    destinatarios_popup = _resolver_notificacion(evento_codigo, "popup", hotel_id=None)
    for dest in destinatarios_popup:
        _encolar_bridge_notificacion(
            usuario=dest.get("username", "?"),
            tipo=tipo_bridge,
            titulo=titulo_bridge or f"📋 Aviso — {evento_codigo}",
            mensaje=texto_telegram.replace("*", ""),
            nivel=nivel_bridge,
            pedido_id=pedido_id_bridge,
        )

    if asunto_email:
        destinatarios_email = _destinatarios_evento_emails(evento_codigo)
        _encolar_email_sistema(evento_codigo, destinatarios_email, asunto_email, cuerpo_email_html, cuerpo_email_text)


def _notify_solicitud_telegram(texto: str) -> None:
    """
    Notifica una nueva solicitud de acceso (Fase 1 / Fase 2) a los usuarios
    configurados para el evento 'solicitud_acceso'.
    """
    _notificar_evento(
        "solicitud_acceso", texto,
        titulo_bridge="📋 Nueva solicitud de acceso",
        nivel_bridge="aviso",
        tipo_bridge="solicitud_acceso",
    )


def _enviar_supervision_admins(texto: str, evento_codigo: str,
                               titulo_bridge: str = None,
                               pedido_id_bridge: int = None) -> None:
    """
    Envía copia de supervisión a los usuarios configurados para el
    evento_codigo indicado (Administrador → Configuración de Avisos).
    Antes: enviaba a "todos los admins con Telegram" filtrado por un
    conjunto fijo TIPOS_SUPERVISION_ADMIN = {"urgente"} en código.
    """
    _notificar_evento(
        evento_codigo, texto,
        titulo_bridge=titulo_bridge or "📋 [Supervisión Admin] — copia automática",
        pedido_id_bridge=pedido_id_bridge,
        nivel_bridge="urgente",
        tipo_bridge="supervision",
    )


def _emails_usuario(u: dict) -> list:
    """
    v12.25.8 — Devuelve [email] o [email, email2] de un dict de usuario
    (compradores u otros roles), sin duplicados ni vacíos.

    Uso: SOLO para listas de destinatarios (Para/CC/BCC) — nunca para la
    firma, que sigue usando exclusivamente el email principal (primero,
    obligatorio) vía _firma_comprador_html()/_firma_comprador_text().
    """
    if not u:
        return []
    out = []
    e1 = (u.get("email") or "").strip()
    e2 = (u.get("email2") or "").strip()
    if e1:
        out.append(e1)
    if e2 and e2 != e1:
        out.append(e2)
    return out


def _get_compradores_hotel(hotel_codigo: str) -> list:
    """
    Devuelve lista de dicts {username, nombre, email, movil, telegram_chat_id}
    de los usuarios con rol 'compras' que tienen asignado el hotel indicado.

    Sustituye al antiguo diccionario HOTEL_COMPRADOR hardcodeado.
    La asignación se gestiona desde admin: Usuarios → Hoteles asignados (compras).
    """
    if not hotel_codigo:
        return []
    hotel_codigo = hotel_codigo.upper()
    hotel_row = query("SELECT id FROM hoteles WHERE codigo=%s AND activo=1", (hotel_codigo,), one=True)
    if not hotel_row:
        return []
    hotel_id = hotel_row["id"]
    rows = rows_to_list(query(
        """SELECT u.id, u.username, u.nombre, u.email, u.email2, u.movil, u.telegram_chat_id
           FROM usuarios u
           JOIN usuario_comprador_hoteles uch ON uch.usuario_id = u.id
           WHERE uch.hotel_id = %s AND u.activo = 1 AND u.rol = 'compras'
           ORDER BY u.nombre""",
        (hotel_id,)
    ))
    return rows

def _get_usuarios_hotel_rol_telegram(hotel_codigo: str) -> list:
    """
    Devuelve lista de dicts {username, nombre, email, movil, telegram_chat_id}
    de los usuarios con rol 'hotel' asignados al hotel indicado (tabla
    usuario_hoteles), SIN filtrar por si tienen o no telegram_chat_id —
    eso se comprueba en el momento de enviar (igual que con los compradores).

    Equivalente, para el canal Telegram, a la parte "hotel_users" de
    _get_todos_usuarios_hotel() (que se usa para los correos). Se mantiene
    como función independiente porque el correo filtra por email NOT NULL
    y aquí no aplica ese filtro (lo relevante es el telegram_chat_id).
    """
    if not hotel_codigo:
        return []
    hotel_codigo = hotel_codigo.upper()
    hotel_row = query("SELECT id FROM hoteles WHERE codigo=%s AND activo=1", (hotel_codigo,), one=True)
    if not hotel_row:
        return []
    hotel_id = hotel_row["id"]
    rows = rows_to_list(query(
        """SELECT u.id, u.username, u.nombre, u.email, u.movil, u.telegram_chat_id
           FROM usuarios u
           JOIN usuario_hoteles uh ON uh.usuario_id = u.id
           WHERE uh.hotel_id = %s AND u.activo = 1 AND u.rol = 'hotel'
           ORDER BY u.nombre""",
        (hotel_id,)
    ))
    return rows

def _send_telegram(chat_id: str, text: str) -> dict:
    """Envía un mensaje de Telegram al chat_id indicado. Devuelve {ok, error, permanente}.

    (2026-08-06) `permanente=True` cuando el error indica que NUNCA va a
    tener éxito reintentando (p. ej. el usuario bloqueó el bot, borró la
    conversación, o desactivó su cuenta) — a diferencia de un fallo
    transitorio (timeout, 5xx, sin red), que sí merece reintentarse al día
    siguiente. Quien llama a esta función decide qué hacer con el flag;
    aquí solo se detecta y se etiqueta.

    (v12.32.03) Si Telegram rechaza el mensaje con 400 "can't parse
    entities" (típico cuando el texto interpolado — nombre de usuario,
    email, etc. — contiene un `*`, `_` o `` ` `` suelto que rompe el
    parseo de Markdown), se reintenta UNA vez enviando el mismo texto
    sin `parse_mode`, como texto plano. Así el aviso llega igual (sin
    negritas/cursivas en ese mensaje concreto) en vez de perderse en
    silencio — antes este caso quedaba marcado como fallo no permanente
    y dependía de un reintento al día siguiente que iba a fallar
    exactamente igual, una y otra vez, por ser un problema de formato
    y no de entrega.
    """
    import urllib.request, urllib.error

    def _post(payload_dict):
        payload = json.dumps(payload_dict).encode()
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return json.loads(resp.read().decode())

    try:
        result = _post({"chat_id": chat_id, "text": text, "parse_mode": "Markdown"})
        return {"ok": result.get("ok", False), "error": None, "permanente": False}
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors="replace")
        body_lower = body.lower()

        # Fallback: error de parseo de Markdown → reintentar como texto plano
        if e.code == 400 and "can't parse entities" in body_lower:
            try:
                result = _post({"chat_id": chat_id, "text": text})
                log.warning("Telegram: fallback a texto plano tras error de Markdown para chat_id %s (motivo: %s)",
                            chat_id, body[:200])
                return {"ok": result.get("ok", False), "error": None, "permanente": False}
            except Exception as e2:
                log.error("Telegram: fallback a texto plano también falló para chat_id %s: %s", chat_id, e2)
                # cae al tratamiento de error normal de abajo, usando la respuesta original

        log.error("Telegram HTTP %s para chat_id %s: %s", e.code, chat_id, body)
        # Errores 403/400 típicos de Telegram cuando NUNCA va a poder
        # entregarse reintentando: bot bloqueado, chat borrado/inexistente,
        # cuenta de usuario desactivada. Se detecta por texto porque
        # Telegram no da un código de error específico para esto, solo una
        # "description" en inglés.
        es_permanente = e.code in (400, 403) and any(frase in body_lower for frase in (
            "bot was blocked by the user",
            "user is deactivated",
            "chat not found",
            "chat_id is empty",
            "peer_id_invalid",
        ))
        return {"ok": False, "error": f"HTTP {e.code}: {body[:200]}", "permanente": es_permanente}
    except Exception as e:
        log.error("Telegram error para chat_id %s: %s", chat_id, e)
        return {"ok": False, "error": str(e), "permanente": False}


def _encolar_bridge_notificacion(usuario: str, tipo: str, titulo: str, mensaje: str,
                                  nivel: str = "aviso", pedido_id: int = None,
                                  retraso_segundos: int = 0) -> None:
    """
    Inserta una fila en bridge_notificaciones para que el bridge de main_agenda
    la recoja en la próxima consulta a /api/bridge/notificaciones.

    Esta función se llama SIEMPRE que se envía un Telegram a un comprador o admin,
    garantizando paridad total entre los avisos de Telegram y los de main_agenda.

    Parámetros:
        usuario           – username del destinatario (igual que en la tabla usuarios)
        tipo               – 'cambio_estado' | 'alerta_auto' | 'techo' | 'familia_repetida' | 'supervision'
        titulo             – línea resumen (se mostrará como título del popup)
        mensaje            – cuerpo completo del aviso
        nivel              – 'aviso' | 'urgente'
        pedido_id          – id del pedido (None para alertas de techo sin pedido concreto)
        retraso_segundos   – (2026-08-14) si > 0, el aviso no será "visible" para
            el bridge hasta pasados esos segundos (columna visible_en =
            NOW() + retraso_segundos) — /api/bridge/notificaciones solo
            devuelve avisos con visible_en <= NOW(). Además, si ya existe un
            aviso pendiente (no leído, todavía no visible) para el mismo
            (usuario, pedido_id, tipo), NO se inserta uno nuevo: se
            SOBRESCRIBE ese con el contenido más reciente y se reinicia la
            cuenta atrás. Con esto, varios cambios seguidos sobre el mismo
            pedido en esa ventana (p. ej. un error corregido al momento)
            solo generan un popup, con el contenido del último cambio —
            pedido explícitamente por el usuario (Víctor) para
            'cambio_estado_pedido', que hasta ahora enviaba un popup
            inmediato por cada cambio de estado sin ninguna protección.
            retraso_segundos=0 (default) mantiene el comportamiento de
            siempre: visible de inmediato, un aviso por llamada.
    """
    try:
        db = get_db()
        cur = db.cursor()
        if retraso_segundos > 0 and pedido_id is not None:
            # ¿Ya hay un aviso del mismo tipo para este pedido/usuario sin
            # leer? Si lo hay —esté ya visible o todavía en espera—, se
            # reemplaza su contenido y se reinicia la espera, en vez de
            # encolar uno más. Se comprueba también el ya-visible-pero-aún-
            # no-leído (no solo visible_en > NOW()) para cubrir el margen
            # entre que un aviso se hace visible y el siguiente poll del
            # bridge lo recoge — si en ese margen llega otro cambio del
            # mismo pedido, también se absorbe en vez de duplicarse.
            cur.execute(
                """UPDATE bridge_notificaciones
                   SET titulo=%s, mensaje=%s, nivel=%s,
                       creado_en=NOW(), visible_en=NOW() + make_interval(secs => %s)
                   WHERE usuario=%s AND tipo=%s AND pedido_id=%s
                     AND leido=FALSE
                   RETURNING id""",
                (titulo, mensaje, nivel, retraso_segundos, usuario.lower(), tipo, pedido_id)
            )
            if cur.fetchone() is not None:
                db.commit()
                return
            cur.execute(
                """INSERT INTO bridge_notificaciones
                   (usuario, tipo, pedido_id, titulo, mensaje, nivel, visible_en)
                   VALUES (%s, %s, %s, %s, %s, %s, NOW() + make_interval(secs => %s))""",
                (usuario.lower(), tipo, pedido_id, titulo, mensaje, nivel, retraso_segundos)
            )
        else:
            cur.execute(
                """INSERT INTO bridge_notificaciones
                   (usuario, tipo, pedido_id, titulo, mensaje, nivel)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (usuario.lower(), tipo, pedido_id, titulo, mensaje, nivel)
            )
        db.commit()
    except Exception as exc:
        log.warning("bridge_notif: no se pudo encolar para %s — %s", usuario, exc)


def _enviar_telegram_compradores(pedido: dict, dias: int, nivel: str) -> list:
    """
    Envía alerta automática (Telegram + popup) a los destinatarios configurados
    para el hotel del pedido — evento 'alerta_pedido_hotel' en Administrador →
    Configuración de Avisos (v12.17.0; antes fijo vía _get_compradores_hotel()).
    Devuelve lista de resultados de Telegram [{username, chat_id, ok, error}].
    """
    hotel_codigo   = (pedido.get("hotel_codigo") or "").upper()
    hotel_id       = pedido.get("hotel_id")
    # v12.17.0: destinatarios configurables por hotel desde Administrador →
    # Configuración de Avisos (evento 'alerta_pedido_hotel'), en vez del
    # antiguo _get_compradores_hotel() fijo. Telegram y popup son
    # independientes: alguien puede estar marcado para uno sin el otro.
    destinatarios_tg    = _resolver_notificacion("alerta_pedido_hotel", "telegram", hotel_id)
    destinatarios_popup = _resolver_notificacion("alerta_pedido_hotel", "popup", hotel_id)
    if not destinatarios_tg and not destinatarios_popup:
        log.warning("Telegram: sin destinatarios configurados para hotel %s (evento alerta_pedido_hotel)", hotel_codigo)
        return []

    # ── Construir mensaje compacto y limpio ──────────────────────────────────
    emoji     = "🔴" if nivel == "urgente" else "🟡"
    nivel_txt = "ALERTA URGENTE" if nivel == "urgente" else "AVISO"
    estado    = pedido.get("estado", "")

    estado_titulo = {
        "ENVIADO AL PROVEEDOR":               "Enviado al proveedor",
        "PENDIENTE FIRMA DIRECCION COMPRAS":  "Pendiente firma compras",
        "PENDIENTE DE FIRMA DIRECCION HOTEL": "Pendiente firma hotel",
        "ENTREGA PARCIAL":                    "Entrega parcial",
        "PENDIENTE COTIZACIÓN":               "Pendiente cotización",
    }.get(estado, estado.capitalize())

    hotel_cod  = pedido.get("hotel_codigo", "?")
    hotel_nom  = pedido.get("hotel_nombre", "")
    proveedor  = pedido.get("proveedor_nombre") or ""
    pedido_sap = pedido.get("pedido_num") or ""
    norden     = pedido.get("norden") or ""
    fecha_ref  = pedido.get("fecha_tramitacion") or pedido.get("fecha_solicitud") or ""

    def _fmt_fecha(f):
        if not f:
            return ""
        try:
            if hasattr(f, "strftime"):
                return f.strftime("%d/%m/%Y")
            parts = str(f)[:10].split("-")
            return "/".join(reversed(parts)) if len(parts) == 3 else str(f)[:10]
        except Exception:
            return str(f)[:10]

    lineas = [f"{emoji} *{nivel_txt} — {estado_titulo}*", ""]
    lineas.append(f"🏨 Hotel: *{hotel_cod}* — {hotel_nom}")
    if pedido_sap:
        lineas.append(f"📄 Pedido SAP: *{pedido_sap}*")
    elif norden:
        lineas.append(f"📄 Línea #: *{norden}*")
    if proveedor:
        lineas.append(f"🏢 Proveedor: {proveedor}")
    if fecha_ref:
        lineas.append(f"📅 Fecha origen: {_fmt_fecha(fecha_ref)}")
    lineas.append(f"⏳ Sin respuesta: *{dias} días*")
    lineas += ["", "— Control Pedidos Princess Canarias"]
    texto = "\n".join(lineas)

    # ── Construir título corto para el popup de agenda ────────────────────────
    # (2026-07-29) Antes usaba pedido.get("id") — el id interno de la base de
    # datos, que no tiene relación visible con el "Nº" (norden) ni con el
    # "Pedido DALI/SAP" que se ven en el panel. Un comprador no tiene forma
    # de saber a qué fila corresponde "Pedido #13537" sin ir a mirarlo en el
    # navegador. Se usa el mismo criterio que ya tenía el cuerpo del mensaje
    # de Telegram: SAP si existe, si no el nº de línea — el id interno solo
    # como último recurso si no hay ninguno de los dos.
    pid_pedido = pedido.get("id")
    if pedido_sap:
        identificador_bridge = f"SAP {pedido_sap}"
    elif norden:
        identificador_bridge = f"Nº{norden}"
    else:
        identificador_bridge = f"#{pid_pedido}"
    titulo_bridge = f"{emoji} [{nivel_txt}] Pedido {identificador_bridge} · {hotel_cod}"

    resultados = []
    for dest in destinatarios_tg:
        username = dest.get("username", "?")
        chat_id  = dest.get("telegram_chat_id")
        if not chat_id:
            log.warning("Telegram: sin telegram_chat_id para %s", username)
            resultados.append({"username": username, "chat_id": None, "ok": False, "error": "Sin telegram_chat_id"})
        else:
            res = _send_telegram(chat_id, texto)
            log.info("Telegram → %s (%s): %s", username, chat_id, "OK" if res["ok"] else res["error"])
            resultados.append({"username": username, "chat_id": chat_id, **res})

    # ── Popup (bridge agenda) — lista propia, independiente de Telegram ──────
    for dest in destinatarios_popup:
        _encolar_bridge_notificacion(
            usuario=dest.get("username", "?"),
            tipo="alerta_auto",
            titulo=titulo_bridge,
            mensaje=texto.replace("*", ""),  # quitar markdown de Telegram
            nivel=nivel,
            pedido_id=pid_pedido,
        )

    # ── Encolar en bridge agenda para ADMINS — SOLO eventos urgentes ─────────
    # Los admins solo deben recibir popup en su Agenda para: solicitudes de
    # acceso y eventos marcados como urgentes (quién concretamente los recibe
    # se configura en Administrador → Configuración de Avisos). Los avisos de
    # nivel normal del comprador NO se replican al admin, para no saturar su
    # Agenda con el día a día rutinario de cada comprador.
    if nivel == "urgente":
        _enviar_supervision_admins(
            texto, "pedido_urgente_admin",
            titulo_bridge=titulo_bridge,
            pedido_id_bridge=pid_pedido,
        )

    return resultados

# ── Telegram inmediato por cambio de estado en edición de pedido ───────────────

# Estados que activan Telegram inmediato al cambiar durante una edición.
# Se excluyen estados terminales sin acción pendiente (SERVIDO TOTAL, CANCELADO).
def _calcular_info_alerta(pedido: dict, estado_nuevo: str,
                          ignorar_si_modificacion_manual: bool = False) -> dict | None:
    """
    Dado un pedido y su nuevo estado, calcula si ese estado genera una condición
    de alerta según UMBRALES_ALERTAS, y devuelve un dict con:
        { "nivel": "aviso"|"urgente", "dias": int, "motivo": str }
    Si no genera alerta, devuelve None.

    Se usa para enriquecer el mensaje de Telegram con contexto de alerta
    cuando el cambio de estado recae en un estado vigilado.

    Parámetro ignorar_si_modificacion_manual:
    ─────────────────────────────────────────
    Cuando True, suprime SIEMPRE la alerta temporal aunque el estado esté en
    UMBRALES_ALERTAS y los días superen el umbral.

    Motivación: las alertas de tipo "N días desde tramitación" están pensadas
    para detectar pedidos *parados* sin acción. Cuando el operador acaba de
    cambiar el estado manualmente (p. ej. PENDIENTE → ENVIADO AL PROVEEDOR),
    añadir "⚠️ 15 días desde tramitación" es contradictorio — el usuario
    literalmente acaba de actuar. El job diario (_job_alertas_diarias) ya
    recoge esa alerta en su próxima ejecución si el pedido sigue sin avanzar.

    Por tanto:
    - ignorar_si_modificacion_manual=False (default) → comportamiento normal,
      usado por el job diario y consultas de diagnóstico.
    - ignorar_si_modificacion_manual=True → sin bloque de alerta, usado desde
      _telegram_cambio_estado (cambios manuales desde update_pedido).
    """
    # ── Guard: si es cambio manual, la alerta temporal es contradictoria ──────
    if ignorar_si_modificacion_manual:
        return None

    cfg = _build_umbrales().get(estado_nuevo)
    if not cfg:
        return None  # estado no vigilado (SERVIDO TOTAL, CANCELADO, etc.)

    fecha_ref_campo = cfg.get("fecha_ref", "fecha_tramitacion")
    dias = _dias_desde_fecha(pedido.get(fecha_ref_campo))
    if dias is None or dias < cfg["primera"]:
        return None  # aún dentro del plazo normal, no hay alerta

    nivel  = "urgente" if (cfg.get("urgente") and dias >= cfg["urgente"]) else "aviso"
    motivo = (
        f"{dias} días desde {'fecha de solicitud' if fecha_ref_campo == 'fecha_solicitud' else 'tramitación'} "
        f"(umbral: {cfg['primera']}d"
        + (f", urgente: {cfg['urgente']}d" if cfg.get("urgente") else "")
        + ")"
    )
    return {"nivel": nivel, "dias": dias, "motivo": motivo}


def _telegram_cambio_estado(db, pedido_id: int, estado_nuevo: str, estado_antes: str,
                             usuario_nombre: str = "",
                             es_cambio_manual: bool = True,
                             es_automatico: bool = False) -> None:
    """
    Envía Telegram inmediato en cambio de estado (PUT /api/pedidos/<pid>),
    alineado con la misma lógica que el correo interno (enviar_emails_estado):

    - Solo se dispara para los estados establecidos en ESTADOS_EMAIL_INTERNO
      ("ENVIADO AL PROVEEDOR", "ENTREGA PARCIAL", "ENTREGADO", "CANCELADO").
      Para el resto de estados (PENDIENTE...) no se envía nada, igual que
      ocurre con el correo.
    - Nunca se envía Telegram al proveedor (a diferencia del correo, que sí
      le escribe en ENVIADO AL PROVEEDOR) — el Telegram es exclusivamente
      un canal interno.
    - Destinatarios (v12.17.0): configurables por hotel desde Administrador →
      Configuración de Avisos (evento 'cambio_estado_pedido'), con Telegram y
      popup como toggles independientes por usuario. Antes era fijo por
      código (compradores del hotel + usuarios rol "hotel", igual conjunto
      que el BCC del correo interno) — ese seguía siendo el conjunto de
      partida al desplegar esta versión (se migró tal cual), pero ya es
      editable sin tocar código.
    - Si el nuevo estado genera una condición de alerta (UMBRALES_ALERTAS)
      Y es_cambio_manual=False, se añade al mensaje: nivel y motivo (días).
    - Con es_cambio_manual=True (default para cambios desde update_pedido):
      el bloque de alerta temporal se suprime. Motivo: las alertas de "N días
      desde tramitación" detectan pedidos parados sin acción; mostrarla justo
      cuando el operador acaba de actuar sería contradictorio. El job diario
      recoge la alerta en su próxima ejecución si el pedido sigue sin avanzar.
    - SIN protección _ya_notificado_hoy: los cambios manuales siempre llegan.
      El job diario usa su propia deduplicación (tipo='telegram_auto').
    - Registra en whatsapp_log con tipo='telegram_estado' para trazabilidad
      separada del job automático.
    """
    # ── Filtro de estados: igual conjunto que el correo interno ──────────────
    if estado_nuevo not in ESTADOS_EMAIL_INTERNO:
        log.debug("[ESTADO] Estado %s fuera de ESTADOS_EMAIL_INTERNO — sin Telegram", estado_nuevo)
        return

    try:
        pedido = row_to_dict(query(f"{PEDIDO_SELECT_ALERTA} WHERE p.id=%s", (pedido_id,), one=True))
        if not pedido:
            log.warning("[ESTADO] Pedido %s no encontrado para Telegram", pedido_id)
            return

        hotel_cod = (pedido.get("hotel_codigo") or "").upper()
        hotel_id  = pedido.get("hotel_id")
        # ── Destinatarios: configurables por hotel desde Administrador → ────
        # Configuración de Avisos (evento 'cambio_estado_pedido'), en vez del
        # antiguo _get_compradores_hotel()+_get_usuarios_hotel_rol_telegram()
        # fijo por código (v12.17.0). Telegram y popup son independientes.
        destinatarios_tg    = _resolver_notificacion("cambio_estado_pedido", "telegram", hotel_id)
        destinatarios_popup = _resolver_notificacion("cambio_estado_pedido", "popup", hotel_id)
        if not destinatarios_tg and not destinatarios_popup:
            log.warning("[ESTADO] Sin destinatarios configurados para %s (evento cambio_estado_pedido)", hotel_cod)
            return

        # ── Bloque base: siempre presente ─────────────────────────────────────
        num_pedido = pedido.get("pedido_num") or f"Nº Orden {pedido.get('norden', '?')}"
        _ICONO_ESTADO = {
            "ENTREGA PARCIAL": "📦 Entrega parcial registrada.",
            "ENTREGADO":        "✅ Pedido entregado en su totalidad.",
            "CANCELADO":        "❌ Pedido cancelado.",
        }
        lineas = [
            "🔔 *Cambio de estado*",
            f"Hotel: *{pedido.get('hotel_codigo', '?')}* — {pedido.get('hotel_nombre', '')}",
            f"Pedido: *{num_pedido}*",
        ]
        if pedido.get("presupuesto_num"):
            lineas.append(f"Presupuesto: {pedido.get('presupuesto_num')}")
        lineas.append(f"Proveedor: {pedido.get('proveedor_nombre', '—')}")
        if pedido.get("importe") is not None:
            lineas.append(f"Importe: {pedido.get('importe'):.2f} €")
        lineas.append(f"Estado: {estado_antes or '—'}  →  *{estado_nuevo}*")
        # (2026-08-28) Mismo criterio que "Realizado por" en el correo
        # interno (enviar_emails_estado) — si el cambio lo ha decidido el
        # cruce automático de listados, no mostrar el nombre de quien tenía
        # la sesión abierta en ese momento.
        if es_automatico:
            import pytz
            lineas.append(
                "Modificado por: Cierre automático — comparación de listados "
                f"({datetime.now(pytz.timezone('Atlantic/Canary')).strftime('%d/%m/%Y %H:%M')})"
            )
        elif usuario_nombre:
            lineas.append(f"Modificado por: {usuario_nombre}")
        _intro_tg = _ICONO_ESTADO.get(estado_nuevo)
        if _intro_tg:
            lineas += ["", _intro_tg]

        # ── Histórico de entregas (albaranes + fechas), parciales y/o total ───
        _resumen_ent_tg = _resumen_entregas(pedido, estado_nuevo)
        lineas += _telegram_bloque_entregas(_resumen_ent_tg, estado_nuevo)

        # ── Motivo de cancelación, si lo hay ──────────────────────────────────
        if estado_nuevo == "CANCELADO" and pedido.get("observaciones"):
            lineas += ["", f"📝 Motivo: {pedido.get('observaciones')}"]

        # ── Bloque de alerta: solo si el nuevo estado genera alerta ───────────
        # ignorar_si_modificacion_manual suprime alertas temporales contradictorias
        # cuando el operador acaba de actuar (ver docstring de _calcular_info_alerta)
        info_alerta = _calcular_info_alerta(pedido, estado_nuevo,
                                            ignorar_si_modificacion_manual=es_cambio_manual)
        if info_alerta:
            emoji_nivel = "🔴" if info_alerta["nivel"] == "urgente" else "⚠️"
            lineas += [
                "",
                f"{emoji_nivel} *Alerta {info_alerta['nivel'].upper()}*",
                f"Motivo: {info_alerta['motivo']}",
            ]

        lineas.append("— Control Pedidos Princess Canarias")
        texto = "\n".join(lineas)

        # ── Título corto para popup bridge ────────────────────────────────────
        nivel_estado = info_alerta["nivel"] if info_alerta else "aviso"
        titulo_bridge = f"🔔 Cambio estado pedido #{pedido_id} · {pedido.get('hotel_codigo', '?')}"

        # ── Envío Telegram — solo destinatarios marcados para este hotel ──────
        resultados = []
        for dest in destinatarios_tg:
            username = dest.get("username", "?")
            chat_id  = dest.get("telegram_chat_id")
            if not chat_id:
                log.warning("[ESTADO] Sin telegram_chat_id para %s", username)
                resultados.append({"username": username, "chat_id": None,
                                   "ok": False, "error": "Sin telegram_chat_id"})
            else:
                res = _send_telegram(chat_id, texto)
                log.info("[ESTADO] Telegram → %s (%s): %s",
                         username, chat_id, "OK" if res["ok"] else res["error"])
                resultados.append({"username": username, "chat_id": chat_id, **res})

        # ── Popup (bridge agenda) — lista propia, independiente de Telegram ───
        # (2026-08-14) retraso_segundos=300: antirrepetición pedida por el
        # usuario — si el mismo pedido cambia de estado varias veces en menos
        # de 5 minutos (p. ej. un error corregido al momento), no se manda un
        # popup por cada cambio; se espera 5 min desde el último y se entrega
        # solo ese, el definitivo. Ver _encolar_bridge_notificacion(). El
        # Telegram de arriba NO se retrasa — sigue siendo inmediato, el
        # usuario solo pidió cambiar el popup.
        for dest in destinatarios_popup:
            _encolar_bridge_notificacion(
                usuario=dest.get("username", "?"),
                tipo="cambio_estado",
                titulo=titulo_bridge,
                mensaje=texto.replace("*", ""),
                nivel=nivel_estado,
                pedido_id=pedido_id,
                retraso_segundos=300,
            )

        # ── Copia de supervisión a admins: solo si la alerta es urgente ────────
        # Cambio de estado normal o con alerta no urgente → solo al comprador.
        # Cambio de estado con alerta urgente → comprador + admins.
        if info_alerta and info_alerta.get("nivel") == "urgente":
            _enviar_supervision_admins(
                texto, "cambio_estado_supervision",
                titulo_bridge=titulo_bridge,
                pedido_id_bridge=pedido_id,
            )

        # ── Log en whatsapp_log (tipo separado del job diario) ─────────────────
        nota_log = f"Cambio estado: {estado_antes} → {estado_nuevo}"
        if info_alerta:
            nota_log += f" | Alerta {info_alerta['nivel']} ({info_alerta['dias']}d)"
        for r in resultados:
            _log_whatsapp(
                db, pedido_id, "telegram_estado",
                r.get("username", "?"),
                nota_log,
                r.get("ok", False) or r.get("permanente", False),
                r.get("error"),
            )
        db.commit()

    except Exception as exc:
        log.exception("[ESTADO] Error enviando Telegram cambio estado pedido %s: %s",
                      pedido_id, exc)


def _notificar_cambio_estado(db, pedido_id: int, estado_nuevo: str, estado_antes: str,
                              usuario_nombre: str = "", usuario_id: int = None,
                              es_automatico: bool = False) -> list:
    """
    Centraliza todas las notificaciones de un cambio de estado manual.

    Llama en orden a:
      1. enviar_emails_estado       → correo al proveedor y/o internos
      2. _telegram_cambio_estado    → mensaje Telegram inmediato (sin alerta temporal)
    Desde v12.32.03, el paso 2 se ejecuta siempre, aunque el paso 1 falle.

    Uso en update_pedido — tanto flujo normal como flujo hotel:

        if estado_nuevo != estado_antes:
            pendientes = _notificar_cambio_estado(db, pid, estado_nuevo, estado_antes,
                                     usuario_nombre=session.get("nombre", ""),
                                     usuario_id=uid)

    Devuelve la lista de correos pendientes de envío vía EmailJS (ver
    enviar_emails_estado), para que el caller la incluya en su respuesta JSON.

    Ventajas de centralizar aquí:
    - update_pedido no acumula lógica de negocio de notificaciones.
    - El flujo hotel pasa por aquí igual que el flujo normal: cualquier
      canal futuro (Teams, Slack, push, webhook) queda cubierto automáticamente
      para ambos flujos con un único cambio en este método.
    - es_cambio_manual=True queda encapsulado: el caller no necesita saber
      el detalle de la supresión de alertas contradictorias.

    usuario_id / es_automatico: ver enviar_emails_estado() — deciden a qué
    lado (comprador/hotel) se excluye del correo interno porque es quien ha
    hecho el cambio. No afecta al Telegram/popup de _telegram_cambio_estado,
    que sigue igual (canal aparte, no filtrado por quién hizo el cambio).

    (v12.32.03) El Telegram/popup se dispara SIEMPRE, incluso si falla la
    construcción/encolado del correo interno — antes, una excepción dentro
    de enviar_emails_estado() cortaba la ejecución de esta función antes de
    llegar a _telegram_cambio_estado(), dejando también sin Telegram un
    cambio de estado que sí se había aplicado en BD (ver incidencia GY,
    pedidos 40907/40908, v12.32.02: el correo fallaba por un TypeError
    Decimal/float en _resumen_entregas(), y como consecuencia el Telegram
    tampoco se llegaba a enviar). El error del correo se relanza al final
    para que el caller lo siga tratando exactamente igual que hasta ahora
    (aviso en rojo, coincidencia no contada como "aplicada" en
    comparar-listado-albaranes).
    """
    _email_exc = None
    pendientes = []
    try:
        pendientes = enviar_emails_estado(db, pedido_id, estado_nuevo, estado_antes,
                                           usuario_nombre=usuario_nombre, usuario_id=usuario_id,
                                           es_automatico=es_automatico)
    except Exception as exc:
        _email_exc = exc
        log.error("[NOTIFICAR-CAMBIO-ESTADO] Fallo construyendo/encolando el correo interno "
                  "del pedido %s (estado %s): %s — se continúa igualmente con Telegram/popup",
                  pedido_id, estado_nuevo, exc)

    _telegram_cambio_estado(db, pedido_id, estado_nuevo, estado_antes,
                             usuario_nombre=usuario_nombre,
                             es_cambio_manual=True,
                             es_automatico=es_automatico)

    if _email_exc is not None:
        raise _email_exc

    return pendientes


# ── Job diario: alertas por fecha (independiente del usuario) ──────────────────

def get_config() -> dict:
    """Carga la configuración de alertas desde BD. Cachea en g Flask."""
    try:
        from flask import g as _g
        if hasattr(_g, '_config_alertas'):
            return _g._config_alertas
    except RuntimeError:
        pass

    try:
        rows = query("SELECT clave, valor, tipo FROM config_alertas")
        cfg = {}
        for r in (rows or []):
            v = r["valor"]
            if r["tipo"] == "numero":
                try:
                    v = float(v) if "." in str(v) else int(v)
                except (ValueError, TypeError):
                    pass
            cfg[r["clave"]] = v
    except Exception as exc:
        log.error("[get_config] Error leyendo config_alertas, usando defaults: %s", exc)
        cfg = {}

    defaults = {
        "enviado_primera": 15, "enviado_urgente": 25, "enviado_ciclo": 10,
        "firma_compras_primera": 8, "firma_compras_urgente": 0, "firma_compras_ciclo": 8,
        "firma_hotel_primera": 5, "firma_hotel_urgente": 0, "firma_hotel_ciclo": 5,
        "entrega_parcial_primera": 10, "entrega_parcial_urgente": 0, "entrega_parcial_ciclo": 10,
        "cotizacion_primera": 2, "cotizacion_urgente": 3, "cotizacion_ciclo": 3,
        "dias_critico": 60,
        "activar_uso_plazo_entrega": 1,
        "plazo_aviso_dias_antes": 5,
        "plazo_urgente_ciclo": 2,
        "plazo_parcial_aviso_dias_antes": 3,
        "plazo_parcial_urgente_ciclo": 2,
        "activar_reclamacion_proveedor_auto": 0,
        "techo_max_pedido": 3000, "techo_max_mes": 6000,
        "techo_max_pedidos": 2, "techo_max_pedidos_familia": 1, "techo_max_mes_familia": 0, "techo_pct_amarillo": 60,
        "enviado_popup_repetir": 1, "enviado_popup_horas_critico": 1, "enviado_popup_horas_normal": 24,
        "firma_compras_popup_repetir": 1, "firma_compras_popup_horas_critico": 1, "firma_compras_popup_horas_normal": 24,
        "firma_hotel_popup_repetir": 1, "firma_hotel_popup_horas_critico": 1, "firma_hotel_popup_horas_normal": 24,
        "entrega_parcial_popup_repetir": 1, "entrega_parcial_popup_horas_critico": 1, "entrega_parcial_popup_horas_normal": 24,
        "cotizacion_popup_repetir": 1, "cotizacion_popup_horas_critico": 1, "cotizacion_popup_horas_normal": 24,
        "techo_urgente_admin_reenvio_dias": 2,
        "familia_repetida_admin_reenvio_dias": 2,
        "emailjs_public_key_1": "", "emailjs_service_id_1": "", "emailjs_template_id_1": "",
        "emailjs_public_key_2": "", "emailjs_service_id_2": "", "emailjs_template_id_2": "",
        "emailjs_public_key_3": "", "emailjs_service_id_3": "", "emailjs_template_id_3": "",
        "emailjs_cuenta_activa": 1, "emailjs_contador": 0, "emailjs_umbral_cambio": 195,
        "emailjs_cambio_automatico_en": "",
        "emailjs_reinicio_fecha_1": "", "emailjs_reinicio_fecha_2": "", "emailjs_reinicio_fecha_3": "",
    }
    for k, v in defaults.items():
        cfg.setdefault(k, v)

    try:
        from flask import g as _g
        _g._config_alertas = cfg
    except RuntimeError:
        pass
    return cfg


def _build_umbrales() -> dict:
    """UMBRALES_ALERTAS construido dinámicamente desde BD.

    (fix v12.29.45) "PENDIENTE FIRMA DIRECCION COMPRAS" y "PENDIENTE DE
    FIRMA DIRECCION HOTEL" son justo los dos estados que
    ESTADOS_SIN_TRAMITAR marca como "sin fecha_tramitacion todavía" —
    ese campo se rellena más adelante en el flujo. Sin `fecha_ref`
    explícito aquí, _clasificar_alertas() y el job de reclamaciones
    caían en el default "fecha_tramitacion", que para estos dos estados
    está vacío siempre → dias=None → la alerta nunca se disparaba, por
    muchos días que llevara el pedido esperando firma. Se usa
    fecha_solicitud como referencia, igual que ya hacía
    PENDIENTE COTIZACIÓN.
    """
    c = get_config()
    return {
        "ENVIADO AL PROVEEDOR": {
            "primera": c["enviado_primera"],
            "urgente": c["enviado_urgente"] or None,
            "ciclo":   c["enviado_ciclo"],
        },
        "PENDIENTE FIRMA DIRECCION COMPRAS": {
            "primera": c["firma_compras_primera"],
            "urgente": c["firma_compras_urgente"] or None,
            "ciclo":   c["firma_compras_ciclo"],
            "fecha_ref": "fecha_solicitud",
        },
        "PENDIENTE DE FIRMA DIRECCION HOTEL": {
            "primera": c["firma_hotel_primera"],
            "urgente": c["firma_hotel_urgente"] or None,
            "ciclo":   c["firma_hotel_ciclo"],
            "fecha_ref": "fecha_solicitud",
        },
        "ENTREGA PARCIAL": {
            "primera": c["entrega_parcial_primera"],
            "urgente": c["entrega_parcial_urgente"] or None,
            "ciclo":   c["entrega_parcial_ciclo"],
        },
        "PENDIENTE COTIZACIÓN": {
            "primera": c["cotizacion_primera"],
            "urgente": c["cotizacion_urgente"] or None,
            "ciclo":   c["cotizacion_ciclo"],
            "fecha_ref": "fecha_solicitud",
        },
    }


UMBRALES_ALERTAS = {
    "ENVIADO AL PROVEEDOR": {
        "primera": 15, "urgente": 25, "ciclo": 10,
    },
    "PENDIENTE FIRMA DIRECCION COMPRAS": {
        "primera": 8, "urgente": None, "ciclo": 8, "fecha_ref": "fecha_solicitud",
    },
    "PENDIENTE DE FIRMA DIRECCION HOTEL": {
        "primera": 5, "urgente": None, "ciclo": 5, "fecha_ref": "fecha_solicitud",
    },
    "ENTREGA PARCIAL": {
        "primera": 10, "urgente": None, "ciclo": 10,
    },
    "PENDIENTE COTIZACIÓN": {
        "primera": 2, "urgente": 3, "ciclo": None, "fecha_ref": "fecha_solicitud",
    },
}

def _dias_desde_fecha(fecha_str):
    """Calcula días transcurridos desde una fecha (string o date/datetime)."""
    if not fecha_str:
        return None
    try:
        # (2026-07-30) FIX CRÍTICO: antes usaba `_d` y `_dt`, dos nombres que
        # NUNCA se importaron a nivel de esta función (solo existían como
        # imports locales dentro de otras funciones sin relación). Cada
        # llamada lanzaba un NameError por dentro, silenciado por el
        # `except Exception: return None` de más abajo — así que esta
        # función devolvía SIEMPRE None, para cualquier pedido, sin
        # excepción. Como todo el job de alertas (incluida la reclamación
        # automática) depende de "dias" para decidir si hay que avisar,
        # esto rompía TODO el envío automático en silencio, no solo la
        # reclamación. Usa los nombres bien importados a nivel de módulo
        # (línea ~7: `from datetime import datetime, ..., date as _date`).
        if isinstance(fecha_str, datetime):
            f = fecha_str.date()
        elif isinstance(fecha_str, _date):
            f = fecha_str
        else:
            f = datetime.strptime(str(fecha_str)[:10], "%Y-%m-%d").date()
        return (_date.today() - f).days
    except Exception as exc:
        log.warning("[_dias_desde_fecha] No se pudo interpretar %r: %s", fecha_str, exc)
        return None

def _ya_notificado_hoy(pedido_id: int, tipo: str = "telegram_auto") -> bool:
    """
    Devuelve True si ya se INTENTÓ (con éxito o no) una notificación del
    tipo indicado para este pedido HOY.
    - tipo='telegram_auto'   → job diario de alertas por fecha
    - tipo='telegram_estado' → cambio de estado inmediato desde update_pedido
    Evita duplicar notificaciones si la misma acción se dispara varias veces
    — el job corre cada minuto, así que esto es lo que evita reintentar
    (con éxito o sin él) 1440 veces al día si algo va mal.

    (2026-08-06) A propósito NO exige enviado=1 aquí — ver _nunca_notificado()
    para el fix real de "un fallo no debe bloquear para siempre". Esta
    función es solo el freno de "no more de una vez por día", tenga éxito
    o no el intento; _nunca_notificado() es la que decide si, al empezar
    un día nuevo, hay que volver a intentarlo.
    """
    try:
        row = query(
            """SELECT COUNT(*) as n FROM whatsapp_log
               WHERE pedido_id=%s AND tipo=%s
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date""",
            (pedido_id, tipo), one=True
        )
        return (row["n"] if row else 0) > 0
    except Exception:
        return False


def _ya_reclamado_hoy_manual(pedido_id: int, tipo: str = "alerta_proveedor") -> bool:
    """
    (2026-07-30) Devuelve True si ya se envió una alerta MANUAL hoy para
    este pedido (botón "Notificar"/"Re-notificar" del panel), del tipo
    indicado en emails_log — una tabla distinta de whatsapp_log, que es la
    que consulta _ya_notificado_hoy.
    tipo='alerta_proveedor' (default): email manual al proveedor — evita
      que la reclamación automática al proveedor la duplique el mismo día.
    tipo='alerta_interno' (v12.23.8): email manual interno (p.ej. aviso de
      firma pendiente al comprador) — evita que el aviso automático
      equivalente la duplique el mismo día.

    (2026-08-06) A propósito NO exige enviado=1 — es solo el freno de "ya
    se intentó hoy", igual que _ya_notificado_hoy(). Ver _nunca_notificado()
    para el fix real del problema de fondo (un fallo no debe bloquear para
    siempre, solo por el resto del día en curso).
    """
    try:
        row = query(
            """SELECT COUNT(*) as n FROM emails_log
               WHERE pedido_id=%s AND tipo=%s
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date""",
            (pedido_id, tipo), one=True
        )
        return (row["n"] if row else 0) > 0
    except Exception:
        return False

# SQL inline para el job (no depende de PEDIDO_SELECT_ALERTA que se define más abajo)
_JOB_PEDIDO_SQL = """
    SELECT p.id, p.norden, p.pedido_num, p.presupuesto_num, p.estado,
           p.fecha_tramitacion, p.fecha_solicitud, p.observaciones,
           p.plazo_entrega_dias, p.fecha_entrega_especifica, p.hotel_id, p.proveedor_id,
           p.sujeto_techo, p.familia_id, p.importe,
           fam.nombre as familia_nombre,
           h.codigo as hotel_codigo, h.nombre as hotel_nombre,
           d.nombre as departamento_nombre,
           pr.nombre as proveedor_nombre,
           (SELECT pc.email
              FROM proveedor_contactos pc
             WHERE pc.proveedor_id = pr.id AND pc.es_principal = 1
               AND pc.email IS NOT NULL AND pc.email != \'\'
               AND (EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id AND pch.hotel_id = p.hotel_id)
                    OR NOT EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch2 WHERE pch2.contacto_id = pc.id))
             ORDER BY EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch3 WHERE pch3.contacto_id = pc.id AND pch3.hotel_id = p.hotel_id) DESC,
                      pc.orden, pc.id
             LIMIT 1) as proveedor_email
    FROM pedidos p
    LEFT JOIN hoteles h ON p.hotel_id = h.id
    LEFT JOIN departamentos d ON p.departamento_id = d.id
    LEFT JOIN proveedores pr ON p.proveedor_id = pr.id
    LEFT JOIN familias fam ON p.familia_id = fam.id
"""

def _nunca_notificado(pedido_id: int, tipo: str = "telegram_auto") -> bool:
    """
    Devuelve True si el pedido nunca ha recibido CON ÉXITO una notificación
    del tipo indicado.

    (2026-08-06) FIX real del problema reportado: antes contaba también los
    intentos fallidos (enviado=0), así que un pedido cuyo primer envío
    fallara (p. ej. sin destinatarios configurados para el hotel/evento, o
    un error puntual de Telegram) se consideraba "ya notificado" para
    siempre y jamás se volvía a intentar — aunque en la pantalla de
    Alertas siguiera apareciendo, correctamente, como "Sin notificar".
    Ahora exige enviado=1: un fallo dejará de bloquear el reintento.

    El reintento no es inmediato ni cada minuto — _ya_notificado_hoy()
    (sin este fix, a propósito) sigue frenando cualquier intento adicional
    el resto del mismo día en que ya se probó, tenga éxito o no. Es al
    empezar el día siguiente cuando esta función vuelve a decir "nunca se
    notificó con éxito" y el job lo reintenta — una vez al día mientras
    la causa de fondo no se arregle, no 1440 veces.
    """
    try:
        row = query(
            "SELECT COUNT(*) as n FROM whatsapp_log WHERE pedido_id=%s AND tipo=%s AND enviado=1",
            (pedido_id, tipo), one=True
        )
        return (row["n"] if row else 0) == 0
    except Exception:
        return True  # En caso de error, asumir que nunca se notificó

def _dias_ultima_notificacion(pedido_id: int, tipo: str = "telegram_auto"):
    """Devuelve cuántos días han pasado desde la última notificación de ese tipo (enviado=1)."""
    try:
        row = query(
            """SELECT DATE(MAX(creado_en)) as ultima FROM whatsapp_log
               WHERE pedido_id=%s AND tipo=%s AND enviado=1""",
            (pedido_id, tipo), one=True
        )
        if not row or not row["ultima"]:
            return None
        from datetime import date as _d
        ultima = row["ultima"]
        if hasattr(ultima, "date"):
            ultima = ultima.date()
        return (_d.today() - ultima).days
    except Exception:
        return None



# ── Helpers para lógica de plazo de entrega ──────────────────────────────────

def _resolver_fecha_entrega_prevista(pedido: dict):
    """
    (2026-08-01) Devuelve la fecha de entrega prevista (date) para un
    pedido, con esta prioridad:
      1. fecha_entrega_especifica, si el proveedor dio un día de entrega
         concreto — se usa tal cual, sin calcular nada.
      2. fecha_tramitacion + plazo_entrega_dias, si hay un plazo en días
         informado (comportamiento de siempre, sin cambios).
      3. None si no hay ninguno de los dos datos — mismo comportamiento
         que cuando no se rellena plazo_entrega_dias.

    Usa `datetime`/`_date` (importados a nivel de módulo, línea ~8) — NO
    `_d`/`_dt`, que es el mismo error ya corregido en _dias_desde_fecha
    (2026-07-30): esos nombres solo existían como imports locales dentro
    de otras funciones sin relación, así que un NameError silenciado por
    el `except Exception` hacía que la función devolviera siempre None.
    """
    fee = pedido.get("fecha_entrega_especifica")
    if fee:
        try:
            if isinstance(fee, datetime):
                return fee.date()
            if isinstance(fee, _date):
                return fee
            return datetime.strptime(str(fee)[:10], "%Y-%m-%d").date()
        except Exception:
            pass
    plazo = pedido.get("plazo_entrega_dias")
    if plazo:
        return _calcular_fecha_entrega_prevista(pedido.get("fecha_tramitacion"), plazo)
    return None


def _calcular_fecha_entrega_prevista(fecha_tramitacion, plazo_dias):
    """
    Devuelve (date) fecha_tramitacion + plazo_dias, o None si falta algún dato.

    (2026-08-01) FIX CRÍTICO — mismo bug que _dias_desde_fecha
    (2026-07-30), nunca corregido aquí: usaba `_d`/`_dt`, dos nombres que
    NUNCA se importaron a nivel de esta función. Como fecha_tramitacion
    se guarda como TEXT (ver models.py), siempre caía en la rama
    `else: _dt.strptime(...)` y lanzaba NameError — silenciado por el
    `except Exception: return None` de abajo. Resultado: esta función
    devolvía SIEMPRE None para cualquier pedido con plazo informado, así
    que toda la lógica de alertas basada en "Plazo entrega (días)"
    (_alertas_plazo_entrega, reclamación automática incluida) llevaba
    inactiva desde que existe la funcionalidad — nunca disparó ni un
    solo aviso por esta vía, sin que hubiera ningún error visible.
    """
    if not fecha_tramitacion or not plazo_dias:
        return None
    try:
        if isinstance(fecha_tramitacion, datetime):
            base = fecha_tramitacion.date()
        elif isinstance(fecha_tramitacion, _date):
            base = fecha_tramitacion
        else:
            base = datetime.strptime(str(fecha_tramitacion)[:10], "%Y-%m-%d").date()
        return base + timedelta(days=int(plazo_dias))
    except Exception:
        return None


def _alertas_plazo_entrega(pedido: dict, cfg_activado: bool):
    """
    Calcula si hoy debe generarse una alerta basada en la fecha de entrega
    prevista del pedido (ver _resolver_fecha_entrega_prevista): o bien una
    fecha de entrega específica indicada por el proveedor, o bien
    fecha_tramitacion + plazo_entrega_dias si no hay fecha específica.

    Estados soportados:
      - ENVIADO AL PROVEEDOR  → usa plazo_aviso_dias_antes / plazo_urgente_ciclo
      - ENTREGA PARCIAL       → usa plazo_parcial_aviso_dias_antes / plazo_parcial_urgente_ciclo

    Reglas (umbrales configurables desde Admin → Config Alertas):
      - primerAviso : EXACTAMENTE N días antes de fecha_entrega_prevista
      - silencio    : entre N-1 y 1 días antes
      - avisoEntrega: el mismo día de fecha_entrega_prevista  (delta == 0)
      - urgenteCada : cada M días a partir del día siguiente  (delta == M, 2M, 3M …)

    Devuelve None si no aplica, o dict {"nivel": "aviso"|"urgente", "motivo": str,
                                         "fecha_entrega_prevista": date}
    """
    if not cfg_activado:
        return None

    estado = pedido.get("estado")
    if estado == "ENVIADO AL PROVEEDOR":
        cfg_key_aviso  = "plazo_aviso_dias_antes"
        cfg_key_ciclo  = "plazo_urgente_ciclo"
        cfg_def_aviso  = 5
        cfg_def_ciclo  = 2
    elif estado == "ENTREGA PARCIAL":
        cfg_key_aviso  = "plazo_parcial_aviso_dias_antes"
        cfg_key_ciclo  = "plazo_parcial_urgente_ciclo"
        cfg_def_aviso  = 3
        cfg_def_ciclo  = 2
    else:
        return None

    fecha_entrega = _resolver_fecha_entrega_prevista(pedido)
    if not fecha_entrega:
        return None

    cfg = get_config()
    dias_aviso = int(cfg.get(cfg_key_aviso, cfg_def_aviso) or cfg_def_aviso)
    ciclo      = int(cfg.get(cfg_key_ciclo,  cfg_def_ciclo)  or cfg_def_ciclo)
    if ciclo < 1:
        ciclo = 1  # evitar división por cero

    from datetime import date as _d
    hoy   = _d.today()
    delta = (hoy - fecha_entrega).days  # negativo = antes, 0 = hoy, positivo = después

    fecha_str = fecha_entrega.strftime("%d/%m/%Y")

    # Primer aviso: únicamente el día exacto de N días antes
    if delta == -dias_aviso:
        return {
            "nivel":  "aviso",
            "motivo": f"Entrega prevista el {fecha_str} (faltan {dias_aviso} días)",
            "fecha_entrega_prevista": fecha_entrega,
        }

    # Silencio entre -(N-1) y -1 inclusive
    if -(dias_aviso - 1) <= delta <= -1:
        return None

    # Aviso el día exacto de la entrega
    if delta == 0:
        return {
            "nivel":  "urgente",
            "motivo": f"Hoy es la fecha de entrega prevista ({fecha_str})",
            "fecha_entrega_prevista": fecha_entrega,
        }

    # Urgente cada M días a partir del día siguiente (delta == M, 2M, 3M …)
    if delta > 0 and delta % ciclo == 0:
        return {
            "nivel":  "urgente",
            "motivo": f"Entrega prevista {fecha_str} superada hace {delta} día(s)",
            "fecha_entrega_prevista": fecha_entrega,
        }

    return None


def _debe_usar_logica_plazo(pedido: dict) -> bool:
    """True si el pedido tiene plazo o fecha de entrega específica informados
    Y la feature está activada en config."""
    cfg = get_config()
    activado = bool(int(cfg.get("activar_uso_plazo_entrega", 1) or 0))
    return activado and bool(pedido.get("plazo_entrega_dias") or pedido.get("fecha_entrega_especifica"))


def _job_alertas_diarias():
    """
    Job automático: calcula alertas por fecha y envía Telegram
    a los compradores responsables sin ninguna interacción del usuario.
    Se ejecuta cada 60 segundos en horario 07:00-16:00 hora Canarias.

    Lógica de envío (por pedido):
      1. Nunca ha recibido telegram_auto  → envía siempre (primer aviso)
      2. días >= get_config()['dias_critico'] → envía siempre (umbral crítico)
      3. Resto                            → solo envía si han pasado >= ciclo días
                                            desde la última notificación
    En todos los casos, _ya_notificado_hoy() evita duplicados dentro del mismo día.
    """
    with app.app_context():
        _job_alertas_diarias_inner()
        _flush_egress_bytes()

def _job_alertas_diarias_inner():
    # (2026-08-02) Fin de semana: no se envía nada — ni reclamación al
    # proveedor, ni Telegram, ni popup de main_agenda (los tres salen de
    # este mismo job, vía _enviar_telegram_compradores /
    # _encolar_reclamacion_proveedor_auto / _encolar_aviso_firma_pendiente_auto).
    # El contador de "días" (fecha_tramitacion → hoy) sigue siendo en días
    # naturales, sin tocar — solo se retrasa el ENVÍO al lunes. Como el ciclo
    # de reenvío (_dias_ultima_notificacion / _ya_notificado_hoy) se basa en
    # la fecha real del último aviso guardado en whatsapp_log, al no haber
    # ningún envío en sábado/domingo el recuento del ciclo sigue contando con
    # normalidad desde el último aviso real (viernes) hasta el lunes, sin
    # necesidad de ningún ajuste especial aquí.
    import pytz
    tz_canarias = pytz.timezone("Atlantic/Canary")
    ahora = datetime.now(tz_canarias)
    if ahora.weekday() >= 5:  # sábado=5, domingo=6
        log.debug("[SCHEDULER] Fin de semana — saltando job de alertas diarias (correo/telegram/popup)")
        return

    log.info("▶ [SCHEDULER] Inicio job alertas diarias — %s", _date.today())
    log.info("BUILD-MARKER v12.22.2 reclamacion-fix activo")
    try:
        alertas_raw = rows_to_list(query(
            _JOB_PEDIDO_SQL + """
            WHERE p.estado IN (
                \'ENVIADO AL PROVEEDOR\',
                \'PENDIENTE FIRMA DIRECCION COMPRAS\',
                \'PENDIENTE DE FIRMA DIRECCION HOTEL\',
                \'ENTREGA PARCIAL\',
                \'PENDIENTE COTIZACIÓN\'
            )
              AND (
                p.fecha_tramitacion IS NOT NULL
                OR (p.estado = \'PENDIENTE COTIZACIÓN\' AND p.fecha_solicitud IS NOT NULL)
              )
            ORDER BY p.fecha_tramitacion ASC
        """))
    except Exception as exc:
        log.error("[SCHEDULER] Error consultando pedidos: %s", exc)
        return

    log.info("RECLAMACION-DEBUG alertas_raw=%d filas devueltas por la consulta del job", len(alertas_raw))

    enviados = 0
    omitidos = 0
    cfg_activar_plazo = bool(int(get_config().get("activar_uso_plazo_entrega", 1) or 0))

    for p in alertas_raw:
        # ── Lógica por plazo de entrega (si el pedido la tiene y está activada) ─
        # v12.29.52 — FIX: un pedido con fecha_entrega_especifica o
        # plazo_entrega_dias informado debe evaluarse SIEMPRE por esta vía
        # mientras la tenga informada — nunca por la lógica estándar de
        # "días desde fecha_tramitacion", aunque hoy no le toque disparar
        # nada (p. ej. faltan más de N días para la entrega, o está en la
        # ventana de silencio). Antes se usaba `if info_plazo:` para decidir
        # esto, pero _alertas_plazo_entrega() devuelve None tanto si el
        # pedido no tiene plazo/fecha informados COMO si los tiene pero hoy
        # no es día de disparo — el código no distinguía ambos casos y en
        # el segundo caía por error a la lógica estándar (días desde
        # fecha_tramitacion), disparando alertas/reclamaciones automáticas
        # ignorando una fecha de entrega específica todavía lejana. Ahora
        # se usa _debe_usar_logica_plazo(p) —ya existía, pero no se llamaba
        # desde ningún sitio— para decidir de entrada si el pedido "vive"
        # en la vía de plazo; si es así, se evalúa _alertas_plazo_entrega()
        # y se continúa al siguiente pedido pase lo que pase (haya o no
        # alerta hoy), sin tocar nunca la lógica estándar.
        if _debe_usar_logica_plazo(p):
            info_plazo = _alertas_plazo_entrega(p, cfg_activar_plazo)
            if not info_plazo:
                # Tiene fecha/plazo informado pero hoy no toca ningún aviso
                # por esa vía (todavía faltan días, o está en la ventana de
                # silencio) — no caer a la lógica estándar.
                omitidos += 1
                continue
            # Usar lógica de plazo en lugar de la estándar para ENVIADO AL PROVEEDOR
            nivel  = info_plazo["nivel"]
            motivo = info_plazo["motivo"]
            dias   = _dias_desde_fecha(p.get("fecha_tramitacion")) or 0

            # ── 2026-07-30: reclamación automática, ANTES del gate de
            # "ya notificado hoy" de Telegram (misma razón que en el camino
            # estándar, un poco más abajo) — con su propia deduplicación
            # diaria, independiente de si el aviso interno ya salió hoy.
            cfg_reclamacion_auto = bool(int(get_config().get("activar_reclamacion_proveedor_auto", 0) or 0))
            if (cfg_reclamacion_auto and nivel == "urgente"
                    and not _ya_notificado_hoy(p["id"], "reclamacion_proveedor_auto")):
                try:
                    ok_reclamacion = _encolar_reclamacion_proveedor_auto(p, dias, nivel)
                    if ok_reclamacion:
                        db = get_db()
                        _log_whatsapp(
                            db, p["id"], "reclamacion_proveedor_auto", "sistema",
                            f"Reclamación automática encolada al proveedor — {motivo}",
                            True, None,
                        )
                        db.commit()
                except Exception as exc:
                    log.error("[SCHEDULER] Error encolando reclamación automática pedido %s: %s", p["id"], exc)

            if _ya_notificado_hoy(p["id"], "telegram_auto"):
                omitidos += 1
                continue

            # Para alertas de plazo: siempre enviar si corresponde (ciclo cada 2 días
            # ya está controlado por _alertas_plazo_entrega — solo devuelve algo
            # en días que toca). Dedup entre jobs del mismo día: _ya_notificado_hoy.
            debe_enviar = True
            log.info("[SCHEDULER-PLAZO] Pedido %s — %s (%s)", p["id"], nivel, motivo)

            resultados = _enviar_telegram_compradores(p, dias, nivel)
            try:
                db = get_db()
                for r in resultados:
                    _log_whatsapp(
                        db, p["id"], "telegram_auto",
                        r.get("username", "?"),
                        f"Alerta plazo entrega {nivel} — {motivo}",
                        r.get("ok", False) or r.get("permanente", False),
                        r.get("error"),
                    )
                db.commit()
            except Exception as exc:
                log.error("[SCHEDULER] Error guardando log (plazo) pedido %s: %s", p["id"], exc)

            enviados += 1
            continue
        # ── Lógica estándar (sin plazo informado o feature desactivada) ──────────

        cfg = _build_umbrales().get(p["estado"])
        log.info("RECLAMACION-DEBUG pedido=%s estado=%s cfg_encontrado=%s", p["id"], p.get("estado"), bool(cfg))
        if not cfg:
            continue

        fecha_ref_campo = cfg.get("fecha_ref", "fecha_tramitacion")
        dias = _dias_desde_fecha(p.get(fecha_ref_campo))
        log.info("RECLAMACION-DEBUG pedido=%s fecha_ref_campo=%s valor_campo=%s dias=%s umbral_primera=%s",
                  p["id"], fecha_ref_campo, p.get(fecha_ref_campo), dias, cfg.get("primera"))
        if dias is None or dias < cfg["primera"]:
            continue

        nivel = "urgente" if (cfg["urgente"] and dias >= cfg["urgente"]) else "aviso"

        # ── 2026-07-30: reclamación automática al proveedor — INDEPENDIENTE
        # del ciclo de reenvío de Telegram de más abajo, pero respetando el
        # MISMO "ciclo" configurado para ese estado en Config Alertas ───────
        # Antes este bloque estaba DESPUÉS de la lógica de "debe_enviar"
        # (primer aviso / umbral crítico / ciclo de N días), así que si el
        # ciclo de Telegram interno decía "todavía no toca reenviar" (p. ej.
        # se notificó hace 1 día y el ciclo es de 2), el `continue` de esa
        # rama saltaba TODO lo de después, incluida la reclamación — aunque
        # llevara semanas sin dispararse nunca por esa razón.
        #
        # 2026-07-30 (2ª vuelta): la primera versión de este bloque solo
        # evitaba mandar dos veces EL MISMO DÍA (_ya_notificado_hoy), pero
        # no respetaba ningún ciclo de varios días — así que, tal cual,
        # reclamaría TODOS los días mientras el pedido siguiera urgente, en
        # vez de cada N días como el aviso interno. A petición del usuario,
        # ahora reutiliza el mismo `cfg["ciclo"]` que ya se configura por
        # estado en Config Alertas — así se controla desde el mismo panel,
        # sin un ajuste aparte que mantener sincronizado a mano.
        cfg_reclamacion_auto = bool(int(get_config().get("activar_reclamacion_proveedor_auto", 0) or 0))
        debe_reclamar = False
        if cfg_reclamacion_auto and nivel == "urgente":
            if _nunca_notificado(p["id"], tipo="reclamacion_proveedor_auto"):
                debe_reclamar = True
            else:
                ciclo_reclamacion = cfg.get("ciclo")
                dias_desde_ultima_reclamacion = _dias_ultima_notificacion(
                    p["id"], tipo="reclamacion_proveedor_auto")
                if ciclo_reclamacion and dias_desde_ultima_reclamacion is not None:
                    debe_reclamar = dias_desde_ultima_reclamacion >= ciclo_reclamacion
            # Red de seguridad final: nunca dos veces el mismo día, aunque
            # el job se dispare más de una vez por algún motivo.
            if debe_reclamar and _ya_notificado_hoy(p["id"], "reclamacion_proveedor_auto"):
                debe_reclamar = False
            log.info(
                "RECLAMACION-DEBUG pedido=%s estado=%s dias=%s activo=%s ciclo=%s debe_reclamar=%s",
                p["id"], p.get("estado"), dias, cfg_reclamacion_auto, cfg.get("ciclo"), debe_reclamar,
            )
        if debe_reclamar:
            try:
                ok_reclamacion = _encolar_reclamacion_proveedor_auto(p, dias, nivel)
                log.info("RECLAMACION-DEBUG pedido=%s resultado_encolar=%s", p["id"], ok_reclamacion)
                if ok_reclamacion:
                    db = get_db()
                    _log_whatsapp(
                        db, p["id"], "reclamacion_proveedor_auto", "sistema",
                        f"Reclamación automática encolada al proveedor — {dias}d sin respuesta",
                        True, None,
                    )
                    db.commit()
            except Exception as exc:
                log.error("[SCHEDULER] Error encolando reclamación automática pedido %s: %s", p["id"], exc)

        # No enviar si ya se notificó hoy (evita duplicados por los 60 ciclos diarios)
        if _ya_notificado_hoy(p["id"], "telegram_auto"):
            omitidos += 1
            continue

        # ── Decisión de envío ────────────────────────────────────────────────
        # 1) Primer aviso: el pedido nunca recibió telegram_auto → enviar siempre
        # 2) Umbral crítico: >= 60 días → enviar siempre
        # 3) Resto: respetar ciclo — solo si han pasado >= ciclo días desde el último
        debe_enviar = False
        motivo_omision = ""

        if _nunca_notificado(p["id"]):
            debe_enviar = True
            log.info("[SCHEDULER] Pedido %s — primer aviso (%dd)", p["id"], dias)
        elif dias >= get_config()["dias_critico"]:
            debe_enviar = True
            log.info("[SCHEDULER] Pedido %s — umbral crítico (%dd >= %dd)", p["id"], dias, get_config()["dias_critico"])
        else:
            ciclo = cfg.get("ciclo")
            if ciclo:
                dias_desde_ultimo = _dias_ultima_notificacion(p["id"])
                if dias_desde_ultimo is None or dias_desde_ultimo >= ciclo:
                    debe_enviar = True
                    log.info("[SCHEDULER] Pedido %s — ciclo OK (%dd desde último)", p["id"], dias_desde_ultimo or 0)
                else:
                    motivo_omision = f"ciclo no cumplido ({dias_desde_ultimo}d < {ciclo}d)"
            else:
                motivo_omision = "sin ciclo, ya notificado anteriormente"

        if not debe_enviar:
            log.debug("[SCHEDULER] Pedido %s omitido — %s", p["id"], motivo_omision)
            omitidos += 1
            continue
        # ────────────────────────────────────────────────────────────────────

        resultados = _enviar_telegram_compradores(p, dias, nivel)

        # Registrar en whatsapp_log
        try:
            db = get_db()
            for r in resultados:
                _log_whatsapp(
                    db, p["id"], "telegram_auto",
                    r.get("username", "?"),
                    f"Alerta automática {nivel} — {dias}d sin respuesta",
                    r.get("ok", False) or r.get("permanente", False),
                    r.get("error"),
                )
            db.commit()
        except Exception as exc:
            log.error("[SCHEDULER] Error guardando log pedido %s: %s", p["id"], exc)

        # ── v12.23.8: aviso automático al comprador por email, en Pendiente
        # Firma Dirección Compras / Dirección Hotel — mismo disparo que el
        # Telegram de arriba (1ª alerta + ciclo), no depende de "urgente"
        # (a petición del usuario: esos dos estados tienen el umbral
        # urgente en 0=nunca por defecto, así que exigir "urgente" para el
        # email lo dejaría sin efecto). Bajo el mismo interruptor maestro
        # que el resto de avisos automáticos por email (Config Alertas →
        # "activar_reclamacion_proveedor_auto").
        if p["estado"] in ("PENDIENTE FIRMA DIRECCION COMPRAS", "PENDIENTE DE FIRMA DIRECCION HOTEL") \
                and bool(int(get_config().get("activar_reclamacion_proveedor_auto", 0) or 0)):
            try:
                ok_aviso = _encolar_aviso_firma_pendiente_auto(p, dias, p["estado"])
                if ok_aviso:
                    db = get_db()
                    _log_whatsapp(
                        db, p["id"], "aviso_firma_auto", "sistema",
                        f"Aviso de firma pendiente encolado al comprador — {dias}d sin firmar",
                        True, None,
                    )
                    db.commit()
            except Exception as exc:
                log.error("[SCHEDULER] Error encolando aviso de firma pendiente pedido %s: %s", p["id"], exc)

        enviados += 1

    log.info("✅ [SCHEDULER] Job finalizado — %d alertas enviadas, %d omitidas", enviados, omitidos)


# ── Job de alerta FAMILIA/PARTIDA REPETIDA — rojo inmediato + reenvío dinámico ──
#
# Disparo:  cuando un hotel repite familia dentro del mismo mes (Regla 2 de _check_techo).
# Primera alerta: Telegram rojo 🔴 al comprador del hotel + a todos los admins.
# Reenvíos:
#   - Comprador  → 1 mensaje por hotel y día (todas las familias agrupadas)
#   - Admins     → 1 mensaje por hotel cada 2 días (todas las familias agrupadas)
# Deduplicación: máx. 1 notificación por HOTEL (no por familia) y día natural.
#
# Tipos usados en whatsapp_log:
#   'familia_repetida_comprador'   → dedup diario a nivel hotel
#   'familia_repetida_admin'       → dedup/ciclo 2 días a nivel hotel
# ──────────────────────────────────────────────────────────────────────────────────


def _ya_notificado_familia_repetida_hotel_hoy(hotel_codigo: str, tipo: str) -> bool:
    """Devuelve True si ya se envió hoy una alerta de familia repetida para este hotel."""
    try:
        row = query(
            """SELECT COUNT(*) as n FROM whatsapp_log
               WHERE pedido_id IS NULL
                 AND tipo = %s
                 AND destinatario LIKE %s
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date""",
            (tipo, f"%{hotel_codigo}|famrep%"), one=True
        )
        return (row["n"] if row else 0) > 0
    except Exception:
        return False


def _dias_desde_ultimo_familia_repetida_admin(hotel_codigo: str) -> int | None:
    """
    Días transcurridos desde la última notificación 'familia_repetida_admin'
    para este hotel. None si nunca se envió.
    """
    try:
        row = query(
            """SELECT DATE(MAX(creado_en) AT TIME ZONE 'Atlantic/Canary') as ultima
               FROM whatsapp_log
               WHERE pedido_id IS NULL
                 AND tipo = 'familia_repetida_admin'
                 AND destinatario LIKE %s""",
            (f"%{hotel_codigo}|famrep%",), one=True
        )
        if not row or row["ultima"] is None:
            return None
        import pytz
        hoy = datetime.now(pytz.timezone("Atlantic/Canary")).date()
        return (hoy - row["ultima"]).days
    except Exception:
        return None


def _log_familia_repetida_hotel(hotel_codigo: str, tipo: str,
                                 destinatario: str, mensaje: str,
                                 enviado: bool, error=None) -> None:
    """Registra en whatsapp_log el envío de alerta de familia repetida (a nivel hotel)."""
    try:
        db = get_db()
        db.cursor().execute(
            "INSERT INTO whatsapp_log (pedido_id,tipo,destinatario,mensaje,enviado,error) "
            "VALUES (NULL,%s,%s,%s,%s,%s)",
            (tipo, f"{hotel_codigo}|famrep|{destinatario}",
             mensaje, 1 if enviado else 0, error)
        )
        db.commit()
    except Exception as exc:
        log.error("[FAM-REP] Error guardando log %s %s: %s", hotel_codigo, destinatario, exc)


def _job_familia_repetida() -> None:
    """
    Job que detecta hoteles con familia/partida repetida en el mes actual
    y dispara UNA alerta agrupada al comprador (diario) y a los admins (cada 2 días).
    Un único mensaje por hotel lista todas las familias repetidas.
    """
    with app.app_context():
        _job_familia_repetida_inner()
        _flush_egress_bytes()


def _job_familia_repetida_inner() -> None:
    """Lógica interna del job de familia repetida."""
    import pytz
    tz_canarias = pytz.timezone("Atlantic/Canary")
    ahora = datetime.now(tz_canarias)

    # Solo en horario laboral (lun-vie 07:00-16:59)
    if ahora.weekday() >= 5 or not (7 <= ahora.hour <= 16):
        log.debug("[FAM-REP] Fuera de horario o día no laborable — saltando")
        return

    year, month = ahora.year, ahora.month
    mes_txt = ahora.strftime("%B %Y")

    log.info("▶ [FAM-REP] Revisando familias repetidas — %s", ahora.strftime("%Y-%m-%d %H:%M"))

    try:
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        ))
    except Exception as exc:
        log.error("[FAM-REP] Error consultando hoteles: %s", exc)
        return

    # v12.28.0 — antes "repetida" era fijo a >1 pedido de la misma familia en
    # el mes; ahora se compara contra el límite configurable
    # techo_max_pedidos_familia (Config alertas → Techo de gastos).
    # (2026-08-01 — rediseño Techo de Gastos, Fase 3) Mismo límite que usa
    # _check_techo() en el momento de pasar a ENVIADO AL PROVEEDOR (ya no
    # al crear/editar el pedido) — y mismo cambio de filtro que los otros
    # 2 jobs de techo: por mes_consumo_techo, no por fecha de creación.
    #
    # v12.29.48 — FIX falso positivo "familia repetida" con el 1er pedido:
    # este job comparaba COUNT(*) >= max_pedidos_familia contando TODOS los
    # pedidos de la familia (sin excluir ninguno). Con el valor por defecto
    # techo_max_pedidos_familia=1, el primer y único pedido de una familia
    # ya cumplía "COUNT(*) >= 1" y se etiquetaba como "repetida" sin que
    # existiera ningún duplicado real. _check_techo() (línea ~7801, el que
    # SÍ bloquea el envío al proveedor) no tiene este problema porque
    # excluye el propio pedido del recuento antes de comparar — por tanto
    # solo bloquea cuando YA había otro pedido antes de este. Este job debe
    # comportarse igual: solo alertar cuando el total de pedidos supere
    # (no solo alcance) el máximo configurado, es decir, cuando de verdad
    # haya más de uno. Cambiado ">=" por ">".
    max_pedidos_familia = get_config()["techo_max_pedidos_familia"]
    _mes_techo_famrep = ahora.strftime("%Y-%m")

    enviados = 0

    for hotel in hoteles:
        hotel_id     = hotel["id"]
        hotel_codigo = (hotel["codigo"] or "").upper()
        hotel_nombre = hotel["nombre"] or ""

        # ── Detectar familias que superan el máximo de pedidos/mes/familia ─
        try:
            familias_repetidas = rows_to_list(query("""
                SELECT p.familia_id, f.nombre as familia_nombre,
                       COUNT(*) as num_pedidos
                FROM pedidos p
                LEFT JOIN familias f ON p.familia_id = f.id
                WHERE p.hotel_id = %s
                  AND p.sujeto_techo = 1
                  AND p.mes_consumo_techo = %s
                  AND p.familia_id IS NOT NULL
                GROUP BY p.familia_id, f.nombre
                HAVING COUNT(*) > %s
                ORDER BY f.nombre
            """, (hotel_id, _mes_techo_famrep, max_pedidos_familia)))
        except Exception as exc:
            log.error("[FAM-REP] Error consultando pedidos hotel %s: %s", hotel_codigo, exc)
            continue

        if not familias_repetidas:
            continue

        # ── Construir UN único mensaje con todas las familias repetidas ────
        reenvio_adm = _dias_desde_ultimo_familia_repetida_admin(hotel_codigo)

        reenvio_txt = (
            f"⏱ Reenvío automático — sin resolver desde hace {reenvio_adm}d\n"
            if reenvio_adm is not None else
            "🔔 Primera alerta — familias repetidas detectadas\n"
        )

        familias_lista = "\n".join(
            "  • {} ({} pedidos)".format(
                f["familia_nombre"] or "ID {}".format(f["familia_id"]),
                f["num_pedidos"]
            )
            for f in familias_repetidas
        )

        texto = (
            "🔴 *ALERTA — Familia/Partida REPETIDA en el mes*\n"
            "\n"
            f"🏨 Hotel: *{hotel_codigo}* — {hotel_nombre}\n"
            "\n"
            f"📂 Familias repetidas ({len(familias_repetidas)}):\n"
            f"{familias_lista}\n"
            "\n"
            f"📅 Mes: {mes_txt}\n"
            f"{reenvio_txt}"
            "— Control Pedidos Princess Canarias"
        )

        titulo_bridge = (
            f"🔴 [FAMILIA REPETIDA] {hotel_codigo} — "
            f"{len(familias_repetidas)} familia(s)"
        )

        # ── Notificar al COMPRADOR (1 vez por hotel y día) ────────────────
        skip_comp = _ya_notificado_familia_repetida_hotel_hoy(
            hotel_codigo, "familia_repetida_comprador"
        )
        if not skip_comp:
            # v12.17.1 (fase 2): destinatarios configurables por hotel desde
            # Administrador → Configuración de Avisos (evento
            # 'familia_repetida_comprador'), en vez del antiguo
            # _get_compradores_hotel() fijo. Telegram y popup, listas propias.
            destinatarios_tg    = _resolver_notificacion("familia_repetida_comprador", "telegram", hotel_id)
            destinatarios_popup = _resolver_notificacion("familia_repetida_comprador", "popup", hotel_id)
            if not destinatarios_tg and not destinatarios_popup:
                log.warning("[FAM-REP] Sin destinatarios configurados para hotel %s (evento familia_repetida_comprador)", hotel_codigo)
            else:
                # (2026-08-28) FIX — Víctor reportó que el popup de familia
                # repetida le llegaba "cada pocos minutos, continuamente".
                # Causa: el registro de dedup (_log_familia_repetida_hotel,
                # de quien depende _ya_notificado_familia_repetida_hotel_hoy
                # más arriba) solo se escribía DENTRO del bucle de Telegram
                # — si un hotel tiene comprador(es) con el popup activado
                # pero sin Telegram configurado (destinatarios_tg vacío,
                # caso real de este hotel), esa escritura nunca llegaba a
                # ejecutarse, así que "ya notificado hoy" nunca se cumplía:
                # el job (cada 60s, 07:00-16:59 laborables) volvía a encolar
                # un popup nuevo en cada pasada, sin fin. Se registra aquí
                # el dedup a nivel hotel ANTES de enviar, específicamente
                # para cubrir el caso sin destinatarios de Telegram — si sí
                # los hay, el bucle de abajo ya lo registra por su cuenta.
                if not destinatarios_tg:
                    _log_familia_repetida_hotel(
                        hotel_codigo, "familia_repetida_comprador",
                        "solo-popup",
                        f"Familia repetida x{len(familias_repetidas)} — {mes_txt}",
                        True
                    )
                for dest in destinatarios_tg:
                    username = dest.get("username", "?")
                    chat_id  = dest.get("telegram_chat_id")
                    # ⚠️ LOG PRIMERO — garantiza dedup aunque falle el envío Telegram
                    _log_familia_repetida_hotel(
                        hotel_codigo, "familia_repetida_comprador",
                        username,
                        f"Familia repetida x{len(familias_repetidas)} — {mes_txt}",
                        False
                    )
                    if chat_id:
                        res = _send_telegram(chat_id, texto)
                        ok  = res.get("ok", False)
                        log.info("[FAM-REP] → comprador %s hotel %s (%d familias): %s",
                                 username, hotel_codigo, len(familias_repetidas),
                                 "OK" if ok else res.get("error"))
                        if ok:
                            try:
                                db = get_db()
                                db.cursor().execute(
                                    """UPDATE whatsapp_log SET enviado=1
                                       WHERE ctid = (
                                           SELECT ctid FROM whatsapp_log
                                           WHERE tipo='familia_repetida_comprador'
                                             AND destinatario=%s AND enviado=0
                                           ORDER BY creado_en DESC LIMIT 1
                                       )""",
                                    (f"{hotel_codigo}|famrep|{username}",)
                                )
                                db.commit()
                            except Exception as _elog:
                                log.warning("[FAM-REP] No se pudo actualizar log enviado comprador %s: %s", username, _elog)
                    else:
                        log.warning("[FAM-REP] Sin telegram_chat_id para comprador %s", username)

                # ── Popup (bridge agenda) — lista propia, independiente de Telegram ──
                for dest in destinatarios_popup:
                    _encolar_bridge_notificacion(
                        usuario=dest.get("username", "?"),
                        tipo="familia_repetida",
                        titulo=titulo_bridge,
                        mensaje=texto.replace("*", ""),
                        nivel="urgente",
                        pedido_id=None,
                    )
        else:
            log.debug("[FAM-REP] Comprador hotel %s — ya notificado hoy, omitiendo", hotel_codigo)

        # ── Notificar a ADMINS (1 vez por hotel cada 2 días) ──────────────
        skip_adm_hoy = _ya_notificado_familia_repetida_hotel_hoy(
            hotel_codigo, "familia_repetida_admin"
        )
        if skip_adm_hoy:
            log.debug("[FAM-REP] Admin hotel %s — ya notificado hoy, omitiendo", hotel_codigo)
        elif reenvio_adm is not None and reenvio_adm < get_config().get("familia_repetida_admin_reenvio_dias", 2):
            log.debug("[FAM-REP] Admin hotel %s — último aviso hace %d día(s), esperando %dd",
                      hotel_codigo, reenvio_adm, get_config().get("familia_repetida_admin_reenvio_dias", 2))
        else:
            admins = _destinatarios_evento("familia_repetida_admin", "telegram")
            if not admins:
                log.warning("[FAM-REP] Sin admins con Telegram configurado")
            else:
                for adm in admins:
                    username = adm.get("username", "?")
                    chat_id  = adm.get("telegram_chat_id")
                    # ⚠️ LOG PRIMERO — garantiza dedup aunque falle el envío Telegram
                    _log_familia_repetida_hotel(
                        hotel_codigo, "familia_repetida_admin",
                        username,
                        f"Familia repetida x{len(familias_repetidas)} — {mes_txt}",
                        False
                    )
                    if chat_id:
                        res = _send_telegram(chat_id, texto)
                        ok  = res.get("ok", False)
                        log.info("[FAM-REP] → admin %s hotel %s (%d familias): %s",
                                 username, hotel_codigo, len(familias_repetidas),
                                 "OK" if ok else res.get("error"))
                        if ok:
                            try:
                                db = get_db()
                                db.cursor().execute(
                                    """UPDATE whatsapp_log SET enviado=1
                                       WHERE ctid = (
                                           SELECT ctid FROM whatsapp_log
                                           WHERE tipo='familia_repetida_admin'
                                             AND destinatario=%s AND enviado=0
                                           ORDER BY creado_en DESC LIMIT 1
                                       )""",
                                    (f"{hotel_codigo}|famrep|{username}",)
                                )
                                db.commit()
                            except Exception as _elog:
                                log.warning("[FAM-REP] No se pudo actualizar log enviado admin %s: %s", username, _elog)
                    else:
                        log.warning("[FAM-REP] Sin telegram_chat_id para admin %s", username)
                    _encolar_bridge_notificacion(
                        usuario=username,
                        tipo="familia_repetida",
                        titulo=titulo_bridge,
                        mensaje=texto.replace("*", ""),
                        nivel="urgente",
                        pedido_id=None,
                    )

        enviados += 1

    log.info("✅ [FAM-REP] Fin revisión — %d hoteles con familias repetidas notificados", enviados)


# ── Job de techo URGENTE — cada 60 s, laborables 07:00-17:00, reenvío c/2 días ─

def _techo_urgente_es_horario_valido() -> bool:
    """
    Devuelve True si ahora mismo cumple las tres condiciones de envío:
      1. Día laborable (lunes=0 … viernes=4)
      2. Hora local (Atlantic/Canary) entre 07:00 y 16:59 inclusive
      3. El mes actual no ha cambiado respecto al mes del techo (siempre True aquí;
         la comprobación de mes se hace al calcular el semáforo, que usa CURRENT MONTH)
    """
    import pytz
    tz_canarias = pytz.timezone("Atlantic/Canary")
    ahora = datetime.now(tz_canarias)
    if ahora.weekday() >= 5:          # sábado=5, domingo=6
        return False
    if not (7 <= ahora.hour <= 16):   # 07:00–16:59; a las 17:00 ya no entra
        return False
    return True


def _ya_notificado_techo_urgente_hoy(hotel_codigo: str) -> bool:
    """
    Devuelve True si ya se envió hoy una alerta de techo URGENTE a admins
    para este hotel (tipo 'telegram_techo_urgente_admin').
    """
    try:
        row = query(
            """SELECT COUNT(*) as n FROM whatsapp_log
               WHERE pedido_id IS NULL
                 AND tipo = 'telegram_techo_urgente_admin'
                 AND destinatario LIKE %s
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date""",
            (f"%{hotel_codigo}%",), one=True
        )
        return (row["n"] if row else 0) > 0
    except Exception as exc:
        log.error("[_ya_notificado_techo_urgente_hoy] Error consultando log para hotel %s: %s", hotel_codigo, exc)
        return False


def _dias_desde_ultimo_techo_urgente_admin(hotel_codigo: str) -> int | None:
    """
    Devuelve los días naturales transcurridos desde la última notificación
    de techo URGENTE a admins para este hotel, o None si nunca se envió.
    """
    try:
        row = query(
            """SELECT DATE(MAX(creado_en) AT TIME ZONE 'Atlantic/Canary') as ultima
               FROM whatsapp_log
               WHERE pedido_id IS NULL
                 AND tipo = 'telegram_techo_urgente_admin'
                 AND destinatario LIKE %s""",
            (f"%{hotel_codigo}%",), one=True
        )
        if not row or row["ultima"] is None:
            return None
        import pytz
        from datetime import date as _d
        hoy = datetime.now(pytz.timezone("Atlantic/Canary")).date()
        return (hoy - row["ultima"]).days
    except Exception as exc:
        log.error("[_dias_desde_ultimo_techo_urgente_admin] Error consultando log para hotel %s: %s", hotel_codigo, exc)
        return None


def _log_techo_urgente_admin(hotel_codigo: str, destinatario: str,
                              mensaje: str, enviado: bool, error=None) -> None:
    """Registra en whatsapp_log el envío de alerta de techo URGENTE a admins."""
    try:
        db = get_db()
        db.cursor().execute(
            "INSERT INTO whatsapp_log (pedido_id,tipo,destinatario,mensaje,enviado,error) "
            "VALUES (NULL,'telegram_techo_urgente_admin',%s,%s,%s,%s)",
            (destinatario, mensaje, 1 if enviado else 0, error)
        )
        db.commit()
    except Exception as exc:
        log.error("[TECHO-URG] Error guardando log %s %s: %s", hotel_codigo, destinatario, exc)


def _job_techo_urgente_admins() -> None:
    """
    Job que se ejecuta cada 60 segundos.

    Notifica a los administradores por Telegram cuando un hotel tiene su techo
    mensual en estado URGENTE (semáforo rojo), con las siguientes reglas:

    • Solo en días laborables (lun–vie).
    • Solo entre las 07:00 y las 16:59 (hora Canarias).
    • Primer envío: el mismo día en que el hotel entra en URGENTE.
    • Reenvíos: cada 2 días naturales desde el último aviso a admins,
      siempre que el hotel siga en rojo Y no haya cambiado de mes.
    • Deduplicación diaria: como máximo 1 notificación por hotel y día.
    """
    with app.app_context():
        _job_techo_urgente_admins_inner()
        _flush_egress_bytes()


def _job_techo_urgente_admins_inner() -> None:
    """Lógica interna del job de techo urgente a admins."""

    if not _techo_urgente_es_horario_valido():
        log.debug("[TECHO-URG] Fuera de horario o día no laborable — saltando")
        return

    import pytz
    tz_canarias = pytz.timezone("Atlantic/Canary")
    ahora = datetime.now(tz_canarias)
    year, month = ahora.year, ahora.month

    log.info("▶ [TECHO-URG] Revisando techos URGENTES — %s", ahora.strftime("%Y-%m-%d %H:%M"))

    try:
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        ))
    except Exception as exc:
        log.error("[TECHO-URG] Error consultando hoteles: %s", exc)
        return

    cfg = get_config()
    enviados = 0

    for hotel in hoteles:
        hotel_id     = hotel["id"]
        hotel_codigo = (hotel["codigo"] or "").upper()
        hotel_nombre = hotel["nombre"] or ""

        # ── 1. Calcular semáforo del mes actual ───────────────────────────
        # (2026-08-01 — rediseño Techo de Gastos, Fase 3) Filtro cambiado de
        # EXTRACT(YEAR/MONTH FROM p.creado_en) a mes_consumo_techo — solo
        # cuentan pedidos que YA han consumido techo de verdad (pasaron por
        # ENVIADO AL PROVEEDOR), igual que en _check_techo().
        _mes_techo_job = ahora.strftime("%Y-%m")
        try:
            pedidos = rows_to_list(query("""
                SELECT p.importe, p.familia_id, f.nombre as familia_nombre
                FROM pedidos p
                LEFT JOIN familias f ON p.familia_id = f.id
                WHERE p.hotel_id = %s
                  AND p.sujeto_techo = 1
                  AND p.mes_consumo_techo = %s
            """, (hotel_id, _mes_techo_job)))
        except Exception as exc:
            log.error("[TECHO-URG] Error consultando pedidos hotel %s: %s", hotel_codigo, exc)
            continue

        acumulado   = sum(float(p["importe"] or 0) for p in pedidos)
        num_pedidos = len(pedidos)

        # Semáforo urgente — solo dispara si es genuinamente ROJO:
        #   ROJO → acumulado >= techo_max_mes (100%)  O  num_pedidos >= techo_max_pedidos
        #   El job mensual ya cubre el amarillo (60% o 1 pedido) al comprador.
        #   Este job urgente solo notifica a admins cuando el techo está realmente superado.
        es_rojo = (
            acumulado >= cfg["techo_max_mes"]
            or num_pedidos > cfg["techo_max_pedidos"]
        )
        if not es_rojo:
            log.debug(
                "[TECHO-URG] Hotel %s — %.1f %% del techo (%d pedidos), no urgente",
                hotel_codigo,
                acumulado / cfg["techo_max_mes"] * 100 if cfg["techo_max_mes"] else 0,
                num_pedidos
            )
            continue

        # ── 2. Deduplicación diaria (máx. 1 aviso/hotel/día) ─────────────
        if _ya_notificado_techo_urgente_hoy(hotel_codigo):
            log.debug("[TECHO-URG] Hotel %s — ya notificado hoy, omitiendo", hotel_codigo)
            continue

        # ── 3. Regla de reenvío cada 2 días ──────────────────────────────
        dias_desde_ultimo = _dias_desde_ultimo_techo_urgente_admin(hotel_codigo)
        _reenvio_dias_techo = get_config().get("techo_urgente_admin_reenvio_dias", 2)
        if dias_desde_ultimo is not None and dias_desde_ultimo < _reenvio_dias_techo:
            log.debug(
                "[TECHO-URG] Hotel %s — último aviso hace %d día(s), esperando %dd",
                hotel_codigo, dias_desde_ultimo, _reenvio_dias_techo
            )
            continue

        # ── 4. Construir y enviar mensaje ─────────────────────────────────
        mes_txt = ahora.strftime("%B %Y")
        pct     = int(acumulado / cfg["techo_max_mes"] * 100) if cfg["techo_max_mes"] else 0

        familias_lista = "\n".join(
            f"• {f}" for f in sorted({
                p["familia_nombre"] for p in pedidos if p.get("familia_nombre")
            })
        ) or "—"

        motivo = []
        if acumulado >= cfg["techo_max_mes"]:
            motivo.append(f"gasto {acumulado:,.2f} € ≥ límite {cfg['techo_max_mes']:,.0f} € (100 %)")
        if num_pedidos >= cfg["techo_max_pedidos"]:
            motivo.append(f"{num_pedidos} pedidos ≥ máximo {cfg['techo_max_pedidos']}")

        reenvio_txt = (
            f"⏱ Reenvío automático — {dias_desde_ultimo}d sin resolver\n"
            if dias_desde_ultimo is not None else
            "🔔 Primera alerta de techo URGENTE\n"
        )

        texto = (
            "🔴 *URGENTE — Techo mensual SUPERADO*\n"
            "\n"
            f"🏨 Hotel: *{hotel_codigo}* — {hotel_nombre}\n"
            "\n"
            f"💰 Acumulado: *{acumulado:,.2f} €* ({pct} % del límite)\n"
            f"📦 Pedidos sujetos: {num_pedidos} / {cfg['techo_max_pedidos']}\n"
            f"⚠️ Motivo: {' | '.join(motivo)}\n"
            "\n"
            f"📂 Familias:\n{familias_lista}\n"
            "\n"
            f"📅 Mes: {mes_txt}\n"
            f"{reenvio_txt}"
            "— Control Pedidos Princess Canarias"
        )

        admins = _destinatarios_evento("techo_urgente_admin", "telegram")
        if not admins:
            log.warning("[TECHO-URG] Sin destinatarios configurados en Configuración de Avisos")
            continue

        for adm in admins:
            username = adm.get("username", "?")
            chat_id  = adm.get("telegram_chat_id")
            if chat_id:
                res = _send_telegram(chat_id, texto)
                ok  = res.get("ok", False)
                log.info(
                    "[TECHO-URG] → admin %s hotel %s: %s",
                    username, hotel_codigo, "OK" if ok else res.get("error")
                )
                _log_techo_urgente_admin(
                    hotel_codigo,
                    f"{username}|{hotel_codigo}",
                    f"Techo URGENTE admin — {acumulado:,.2f} € — {mes_txt}",
                    ok, res.get("error")
                )
            # ── Encolar en bridge agenda para este admin ─────────────────────
            _encolar_bridge_notificacion(
                usuario=username,
                tipo="techo",
                titulo=f"💰 [TECHO URGENTE] Hotel {hotel_codigo} — {mes_txt}",
                mensaje=texto.replace("*", ""),
                nivel="urgente",
                pedido_id=None,
            )

        enviados += 1
        log.info(
            "[TECHO-URG] ✅ Hotel %s notificado a admins — %.2f € / %d pedidos",
            hotel_codigo, acumulado, num_pedidos
        )

    log.info("✅ [TECHO-URG] Fin revisión — %d hoteles urgentes notificados", enviados)

    # ── 5. Alerta específica: pedidos enviados sin autorización previa ──────
    # (2026-08-01 — rediseño Techo de Gastos, Fase 3, punto 5) Si por
    # cualquier vía un pedido llegó a ENVIADO AL PROVEEDOR con
    # no_autorizado_previo=TRUE, ya queda constancia permanente en
    # historial_estados (ver update_pedido) — esto es SOLO para que un
    # admin lo vea cuanto antes, sin tener que ir a buscarlo. Deduplicado
    # por pedido (una sola vez, whatsapp_log con pedido_id=el del pedido).
    try:
        _pedidos_no_autorizados = rows_to_list(query("""
            SELECT p.id, p.pedido_num, p.importe, h.codigo as hotel_codigo, h.nombre as hotel_nombre
            FROM pedidos p
            JOIN hoteles h ON p.hotel_id = h.id
            WHERE p.no_autorizado_previo = TRUE
              AND NOT EXISTS (
                  SELECT 1 FROM whatsapp_log w
                  WHERE w.pedido_id = p.id AND w.tipo = 'telegram_no_autorizado_previo'
              )
        """))
    except Exception as exc:
        log.error("[TECHO-URG] Error consultando pedidos no_autorizado_previo: %s", exc)
        _pedidos_no_autorizados = []

    if _pedidos_no_autorizados:
        admins_integridad = _destinatarios_evento("techo_urgente_admin", "telegram")
        for p_na in _pedidos_no_autorizados:
            texto_na = (
                "⚠️ *Pedido enviado sin autorización previa de Dirección General*\n\n"
                f"🏨 Hotel: *{p_na['hotel_codigo']}* — {p_na['hotel_nombre']}\n"
                f"📄 Pedido Nº: {p_na.get('pedido_num') or p_na['id']}\n"
                f"💰 Importe: {float(p_na.get('importe') or 0):,.2f} €\n\n"
                "Revísese manualmente — no debería haber llegado a ENVIADO AL PROVEEDOR "
                "sin pasar por el circuito de autorización.\n"
                "— Control Pedidos Princess Canarias"
            )
            for adm in admins_integridad or []:
                chat_id = adm.get("telegram_chat_id")
                if chat_id:
                    _send_telegram(chat_id, texto_na)
            db_na = get_db()
            db_na.cursor().execute(
                "INSERT INTO whatsapp_log (pedido_id,tipo,destinatario,mensaje,enviado,error) "
                "VALUES (%s,'telegram_no_autorizado_previo',%s,%s,1,NULL)",
                (p_na["id"], f"admins|{p_na['hotel_codigo']}", texto_na)
            )
            db_na.commit()
            log.warning("[TECHO-URG] ⚠️ Pedido %s sin autorización previa — alertado a admins",
                        p_na.get("pedido_num") or p_na["id"])


# ── Job de alertas de techo mensual ───────────────────────────────────────────

def _ya_notificado_techo_mes_hoy(hotel_codigo: str, semaforo: str) -> bool:
    """
    Devuelve True si ya se envió hoy una alerta de techo mensual para este hotel
    con el mismo nivel de semáforo (rojo/amarillo).
    Usa whatsapp_log con pedido_id=NULL y tipo='telegram_techo_mes_<semaforo>'.
    """
    tipo = f"telegram_techo_mes_{semaforo}"
    try:
        row = query(
            """SELECT COUNT(*) as n FROM whatsapp_log
               WHERE pedido_id IS NULL
                 AND tipo = %s
                 AND destinatario LIKE %s
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date""",
            (tipo, f"%{hotel_codigo}%"), one=True
        )
        return (row["n"] if row else 0) > 0
    except Exception:
        return False


def _log_whatsapp_techo_mes(hotel_codigo: str, semaforo: str, destinatario: str,
                             mensaje: str, enviado: bool, error=None) -> None:
    """Registra en whatsapp_log una notificación de techo mensual (sin pedido_id)."""
    tipo = f"telegram_techo_mes_{semaforo}"
    try:
        db = get_db()
        db.cursor().execute(
            "INSERT INTO whatsapp_log (pedido_id,tipo,destinatario,mensaje,enviado,error) "
            "VALUES (NULL,%s,%s,%s,%s,%s)",
            (tipo, destinatario, mensaje, 1 if enviado else 0, error)
        )
        db.commit()
    except Exception as exc:
        log.error("[TECHO-MES] Error guardando log %s %s: %s", hotel_codigo, destinatario, exc)


def _job_alertas_techo_mensual() -> None:
    """
    Job diario que notifica por Telegram el estado del techo de gastos mensual por hotel.

    Lógica:
      - semáforo ROJO  (techo superado o nº pedidos >= máximo):
          → alerta URGENTE al comprador del hotel  +  copia supervisión a admins
      - semáforo AMARILLO (>= 75 % del techo o nº pedidos == máximo - 1):
          → aviso al comprador del hotel  (sin copia a admins)
      - semáforo VERDE  → sin notificación

    Deduplicación: solo envía una vez por hotel y nivel en el mismo día natural.
    """
    with app.app_context():
        _job_alertas_techo_mensual_inner()
        _flush_egress_bytes()


def _job_alertas_techo_mensual_inner() -> None:
    """Lógica interna del job — llamada siempre dentro de app.app_context()."""
    # (2026-08-02) Mismo criterio que _job_alertas_diarias_inner(): este job
    # es automático (no reacciona a ninguna acción real de un usuario, solo
    # escanea hoteles cada día a las 08:00), igual que techo urgente y
    # familia repetida — que ya tenían este guardián y este job se había
    # quedado fuera por inconsistencia. En fin de semana no se envía nada
    # (ni Telegram ni popup); el lunes se retoma con normalidad. El
    # acumulado del mes y el semáforo se siguen calculando en vivo con datos
    # reales, así que el lunes refleja el estado correcto sin necesidad de
    # ningún ajuste adicional.
    import pytz
    tz_canarias = pytz.timezone("Atlantic/Canary")
    ahora = datetime.now(tz_canarias)
    if ahora.weekday() >= 5:  # sábado=5, domingo=6
        log.debug("[TECHO-MES] Fin de semana — saltando job de techo mensual")
        return

    from datetime import date as _date_local
    hoy   = _date_local.today()
    year  = hoy.year
    month = hoy.month

    log.info("▶ [TECHO-MES] Inicio job techo mensual — %s", hoy)

    try:
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        ))
    except Exception as exc:
        log.error("[TECHO-MES] Error consultando hoteles: %s", exc)
        return

    enviados = 0
    omitidos = 0

    for hotel in hoteles:
        hotel_id     = hotel["id"]
        hotel_codigo = (hotel["codigo"] or "").upper()
        hotel_nombre = hotel["nombre"] or ""

        # ── Calcular acumulado del mes ────────────────────────────────────────
        # (2026-08-01 — rediseño Techo de Gastos, Fase 3) Filtro cambiado de
        # EXTRACT(YEAR/MONTH FROM p.creado_en) a mes_consumo_techo, igual
        # que en _check_techo() y en el job urgente.
        _mes_techo_mensual = hoy.strftime("%Y-%m")
        try:
            pedidos = rows_to_list(query("""
                SELECT p.importe, p.familia_id, f.nombre as familia_nombre
                FROM pedidos p
                LEFT JOIN familias f ON p.familia_id = f.id
                WHERE p.hotel_id = %s
                  AND p.sujeto_techo = 1
                  AND p.mes_consumo_techo = %s
            """, (hotel_id, _mes_techo_mensual)))
        except Exception as exc:
            log.error("[TECHO-MES] Error consultando pedidos hotel %s: %s", hotel_codigo, exc)
            continue

        acumulado   = sum(float(p["importe"] or 0) for p in pedidos)
        num_pedidos = len(pedidos)

        # Semáforo:
        #   ROJO     → acumulado >= techo_max_mes  O  num_pedidos > techo_max_pedidos (techo realmente superado)
        #   AMARILLO → acumulado >= techo_max_mes * pct_amarillo/100  O  num_pedidos >= techo_max_pedidos (límite alcanzado)
        #   VERDE    → sin actividad sujeta al techo (sin notificación)
        umbral_amarillo = get_config()["techo_max_mes"] * get_config()["techo_pct_amarillo"] / 100
        if acumulado >= get_config()["techo_max_mes"] or num_pedidos > get_config()["techo_max_pedidos"]:
            semaforo = "rojo"
        elif acumulado >= umbral_amarillo or num_pedidos >= get_config()["techo_max_pedidos"]:
            semaforo = "amarillo"
        else:
            omitidos += 1
            log.debug("[TECHO-MES] Hotel %s — verde, sin notificación", hotel_codigo)
            continue

        # ── Deduplicación diaria por hotel + nivel ────────────────────────────
        if _ya_notificado_techo_mes_hoy(hotel_codigo, semaforo):
            omitidos += 1
            log.info("[TECHO-MES] Hotel %s — YA NOTIFICADO HOY semaforo=%s, omitiendo", hotel_codigo, semaforo)
            continue

        log.info("[TECHO-MES] Hotel %s — semaforo=%s acumulado=%.2f pedidos=%d -> enviando",
                 hotel_codigo, semaforo, acumulado, num_pedidos)
        # v12.17.1 (fase 2): destinatarios configurables por hotel (evento
        # 'techo_mensual_comprador'), en vez del antiguo _get_compradores_hotel().
        destinatarios_tg    = _resolver_notificacion("techo_mensual_comprador", "telegram", hotel_id)
        destinatarios_popup = _resolver_notificacion("techo_mensual_comprador", "popup", hotel_id)
        if not destinatarios_tg and not destinatarios_popup:
            log.warning("[TECHO-MES] Sin destinatarios configurados para hotel %s (evento techo_mensual_comprador)", hotel_codigo)
            continue

        # (2026-08-28) FIX — mismo bug que en _job_familia_repetida_inner
        # (ver esa función para el diagnóstico completo): el dedup diario
        # (_ya_notificado_techo_mes_hoy) dependía por completo de que
        # _log_whatsapp_techo_mes() se ejecutara dentro del bucle de
        # Telegram, y ahí solo se escribía cuando el destinatario tenía
        # chat_id configurado — si NINGÚN comprador de este hotel tiene
        # Telegram (o tiene el evento activado pero sin chat_id), el dedup
        # nunca se registraba y el job (diario, pero re-evaluado en cada
        # arranque/ejecución) podía volver a encolar el popup una y otra
        # vez. Se registra aquí, una sola vez por hotel, en cuanto se sabe
        # que hay al menos un destinatario (Telegram o popup) — antes de
        # depender de si el envío por Telegram llega a completarse.
        _log_whatsapp_techo_mes(
            hotel_codigo, semaforo, f"solo-popup|{hotel_codigo}",
            f"Techo mensual {semaforo} — {acumulado:,.2f} € — {hoy.strftime('%B %Y')}",
            True
        )

        # ── Construir mensaje ─────────────────────────────────────────────────
        mes_txt      = hoy.strftime("%B %Y")
        pct          = int(acumulado / get_config()["techo_max_mes"] * 100) if get_config()["techo_max_mes"] else 0
        familias_txt = ", ".join({
            p["familia_nombre"] for p in pedidos if p.get("familia_nombre")
        }) or "—"

        if semaforo == "rojo":
            emoji     = "🔴"
            nivel_txt = "URGENTE — Techo mensual superado"
        else:
            emoji     = "🟡"
            nivel_txt = f"AVISO — Techo mensual al {pct} %"

        familias_lista = "\n".join(
            f"• {f}" for f in sorted({p["familia_nombre"] for p in pedidos if p.get("familia_nombre")})
        ) or "—"

        _techo_mes     = get_config()["techo_max_mes"]
        _techo_pedidos = get_config()["techo_max_pedidos"]
        texto = (
            f"{emoji} *{nivel_txt}*\n"
            f"\n"
            f"🏨 Hotel: *{hotel_codigo}* — {hotel_nombre}\n"
            f"\n"
            f"💰 Acumulado actual: *{acumulado:,.2f} €*\n"
            f"📊 Límite configurado: {_techo_mes:,.0f} €\n"
            f"📦 Pedidos sujetos: {num_pedidos} / {_techo_pedidos}\n"
            f"\n"
            f"📂 Familias:\n{familias_lista}\n"
            f"\n"
            f"📅 Mes: {mes_txt}\n"
            "— Control Pedidos Princess Canarias"
        )

        # ── Enviar a compradores ──────────────────────────────────────────────
        nivel_techo = "urgente" if semaforo == "rojo" else "aviso"
        for dest in destinatarios_tg:
            username = dest.get("username", "?")
            chat_id  = dest.get("telegram_chat_id")
            if chat_id:
                res = _send_telegram(chat_id, texto)
                ok  = res.get("ok", False)
                log.info("[TECHO-MES] → %s (%s): %s", username, hotel_codigo,
                         "OK" if ok else res.get("error"))
                _log_whatsapp_techo_mes(
                    hotel_codigo, semaforo,
                    f"{username}|{hotel_codigo}",
                    f"Techo mensual {semaforo} — {acumulado:,.2f} € — {mes_txt}",
                    ok, res.get("error")
                )
            else:
                log.warning("[TECHO-MES] Sin telegram_chat_id para %s", username)

        # ── Popup (bridge agenda) — lista propia, independiente de Telegram ──
        for dest in destinatarios_popup:
            _encolar_bridge_notificacion(
                usuario=dest.get("username", "?"),
                tipo="techo",
                titulo=f"{emoji} [{nivel_txt}] Hotel {hotel_codigo} — {mes_txt}",
                mensaje=texto.replace("*", ""),
                nivel=nivel_techo,
                pedido_id=None,
            )

        # ── Copia a admins: gestionada exclusivamente por _job_techo_urgente_admins ──
        # No se envía copia aquí para evitar duplicado. El job _job_techo_urgente_admins
        # es el canal oficial hacia admins (con reenvío cada 2 días, horario laboral
        # y deduplicación diaria).

        enviados += 1

    log.info("✅ [TECHO-MES] Job finalizado — %d hoteles notificados, %d omitidos",
             enviados, omitidos)


def _telegram_alerta_techo(pedido_id: int, hotel_codigo: str, importe: float, familia_nombre: str):
    """
    Envía Telegram inmediato cuando se crea un pedido sujeto al techo de gastos.
    Se dispara en el momento del INSERT, sin esperar al job diario.
    """
    try:
        pedido = row_to_dict(query(f"{PEDIDO_SELECT_ALERTA} WHERE p.id=%s", (pedido_id,), one=True))
        if not pedido:
            return

        hotel_cod = (hotel_codigo or pedido.get("hotel_codigo") or "").upper()
        hotel_id  = pedido.get("hotel_id")
        # v12.17.1 (fase 2): destinatarios configurables por hotel (evento
        # 'techo_nuevo_pedido_comprador'), en vez del antiguo _get_compradores_hotel().
        destinatarios_tg    = _resolver_notificacion("techo_nuevo_pedido_comprador", "telegram", hotel_id)
        destinatarios_popup = _resolver_notificacion("techo_nuevo_pedido_comprador", "popup", hotel_id)
        if not destinatarios_tg and not destinatarios_popup:
            log.warning("[TECHO] Sin destinatarios configurados para hotel %s (evento techo_nuevo_pedido_comprador)", hotel_cod)
            return

        mes_txt = _date.today().strftime("%B %Y")
        pedido_sap = pedido.get("pedido_num") or ""
        norden_val = pedido.get("norden") or ""
        ref_line   = f"📄 Pedido SAP: *{pedido_sap}*" if pedido_sap else f"📄 Línea #: *{norden_val}*"

        texto = (
            "🏦 *Nuevo pedido sujeto a techo de gastos*\n"
            f"\n"
            f"🏨 Hotel: *{pedido.get('hotel_codigo','?')}* — {pedido.get('hotel_nombre','')}\n"
            f"{ref_line}\n"
            f"📂 Familia: {familia_nombre or '—'}\n"
            f"💰 Importe: *{importe:,.2f} €*\n"
            f"📅 Mes: {mes_txt}\n"
            f"\n"
            "⚠️ Este pedido computa en el techo de gastos mensual.\n"
            "— Control Pedidos Princess Canarias"
        )

        resultados = []
        for dest in destinatarios_tg:
            username = dest.get("username", "?")
            chat_id  = dest.get("telegram_chat_id")
            if chat_id:
                res = _send_telegram(chat_id, texto)
                log.info("[TECHO] Telegram → %s (%s): %s", username, chat_id, "OK" if res["ok"] else res["error"])
                resultados.append({"username": username, "chat_id": chat_id, **res})

        # ── Popup (bridge agenda) — lista propia, independiente de Telegram ──
        for dest in destinatarios_popup:
            _encolar_bridge_notificacion(
                usuario=dest.get("username", "?"),
                tipo="techo",
                titulo=f"🏦 Nuevo pedido sujeto a techo · Hotel {hotel_cod}",
                mensaje=texto.replace("*", ""),
                nivel="aviso",
                pedido_id=pedido_id,
            )

        # ── Copia de supervisión a admins: creación de pedido sujeto a techo es siempre urgente ──
        _enviar_supervision_admins(
            texto, "techo_nuevo_pedido_admin",
            titulo_bridge=f"🏦 [Supervisión] Nuevo pedido techo · Hotel {hotel_cod}",
            pedido_id_bridge=pedido_id,
        )

        # Registrar en log
        db = get_db()
        for r in resultados:
            _log_whatsapp(
                db, pedido_id, "telegram_techo",
                r.get("username", "?"),
                f"Alerta techo gastos — {importe:,.2f} € — {familia_nombre}",
                r.get("ok", False) or r.get("permanente", False),
                r.get("error"),
            )
        db.commit()

    except Exception as exc:
        log.error("[TECHO] Error enviando telegram techo pedido %s: %s", pedido_id, exc)


def _get_proveedor_emails_principales(proveedor_id, hotel_id=None) -> list:
    """Devuelve la lista de emails de los contactos marcados como
    principales (es_principal=1) para un proveedor, en el orden definido
    por `orden`. Un proveedor puede tener varios contactos marcados a la
    vez con la estrella dorada — todos reciben las notificaciones como
    destinatario directo ("Para:"), no en copia.

    v12.27.4 — Correos específicos por hotel: si se pasa hotel_id y el
    proveedor tiene contactos principales asignados específicamente a ese
    hotel (proveedor_contacto_hoteles), se usan SOLO esos — para que, si
    un proveedor sirve a varios hoteles con interlocutores distintos, la
    reclamación llegue al que corresponde y no a todos a la vez. Si no
    hay ninguno asignado a ese hotel (o no se pasa hotel_id), se cae al
    comportamiento de siempre: los contactos "generales" — principales
    sin ningún hotel asignado en absoluto.
    """
    if not proveedor_id:
        return []
    if hotel_id:
        rows_hotel = query(
            """SELECT DISTINCT pc.email, pc.orden, pc.id FROM proveedor_contactos pc
               JOIN proveedor_contacto_hoteles pch ON pch.contacto_id = pc.id
               WHERE pc.proveedor_id=%s AND pc.es_principal=1
                 AND pc.email IS NOT NULL AND pc.email != ''
                 AND pch.hotel_id=%s
               ORDER BY pc.orden, pc.id""",
            (proveedor_id, hotel_id)
        ) or []
        if rows_hotel:
            return [r["email"] for r in rows_hotel]
    rows = query(
        """SELECT pc.email FROM proveedor_contactos pc
           WHERE pc.proveedor_id=%s AND pc.es_principal=1
             AND pc.email IS NOT NULL AND pc.email != ''
             AND NOT EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id)
           ORDER BY pc.orden, pc.id""",
        (proveedor_id,)
    ) or []
    return [r["email"] for r in rows]


def _get_todos_usuarios_hotel(hotel_codigo: str) -> dict:
    """
    Devuelve todos los usuarios activos asignados a un hotel, separados por rol:
      - "compradores": rol='compras' en usuario_comprador_hoteles
      - "hotel_users": rol='hotel'   en usuario_hoteles
    Cada lista contiene dicts con {id, username, nombre, email} para
    "hotel_users", y {id, username, nombre, email, email2, movil} para
    "compradores" (movil se usa para incluirlo en la firma de los correos
    al proveedor; email2, si existe, se añade como destinatario extra en
    BCC pero nunca en la firma — ver _emails_usuario()).
    Uso: determinar destinatarios de correos internos de cambio de estado,
    incluyendo tanto el comprador responsable como el usuario del hotel.
    """
    if not hotel_codigo:
        return {"compradores": [], "hotel_users": []}
    hotel_codigo = hotel_codigo.upper()
    hotel_row = query("SELECT id FROM hoteles WHERE codigo=%s AND activo=1", (hotel_codigo,), one=True)
    if not hotel_row:
        return {"compradores": [], "hotel_users": []}
    hotel_id = hotel_row["id"]

    compradores = rows_to_list(query(
        """SELECT u.id, u.username, u.nombre, u.email, u.email2, u.movil
           FROM usuarios u
           JOIN usuario_comprador_hoteles uch ON uch.usuario_id = u.id
           WHERE uch.hotel_id = %s AND u.activo = 1 AND u.rol = 'compras'
             AND u.email IS NOT NULL AND TRIM(u.email) != ''
           ORDER BY u.nombre""",
        (hotel_id,)
    )) or []

    hotel_users = rows_to_list(query(
        """SELECT u.id, u.username, u.nombre, u.email
           FROM usuarios u
           JOIN usuario_hoteles uh ON uh.usuario_id = u.id
           WHERE uh.hotel_id = %s AND u.activo = 1 AND u.rol = 'hotel'
             AND u.email IS NOT NULL AND TRIM(u.email) != ''
           ORDER BY u.nombre""",
        (hotel_id,)
    )) or []

    return {"compradores": compradores, "hotel_users": hotel_users}


def _get_compradores_cc(hotel_codigo: str):
    """Devuelve lista de dicts {email, nombre, movil} de los compradores responsables del hotel.
    Usa _get_compradores_hotel() para obtener los compradores dinámicamente desde BD."""
    return _get_compradores_hotel(hotel_codigo)

# ── Firma estándar de correo saliente (v12.24.0) ───────────────────────────────
# Mismo formato ya usado en el resto de correspondencia de Compras: nombre,
# departamento, dirección fija del departamento, teléfono con prefijo (+34) y
# email como enlace. Solo nombre/teléfono/email cambian según el comprador
# que firma; departamento y dirección son fijos.

def _formatear_movil_firma(movil: str) -> str:
    """
    (2026-08-10) Normaliza el móvil para la firma de los correos.
    Algunos usuarios lo guardan ya con el prefijo +34 — el propio
    placeholder del campo lo sugiere ("+34 600 000 000") — y anteponer
    "(+34)" sin más duplicaba el prefijo: "(+34) +34681111792". Se quita
    cualquier +34/0034/34 inicial (con o sin espacio después) antes de
    anteponer el "(+34)" fijo de la firma, así el resultado es siempre
    limpio, se haya guardado el número como se haya guardado.
    """
    if not movil:
        return ""
    return re.sub(r'^\s*(\+34|0034|34)\s*', '', movil.strip())

def _firma_comprador_html(nombre: str, email: str, movil: str) -> str:
    nombre_html = f"<strong>{nombre}</strong><br>" if nombre else ""
    tel_html    = f"(+34) {_formatear_movil_firma(movil)}<br>" if movil else ""
    email_html  = f'<a href="mailto:{email}">{email}</a>' if email else ""
    return (
        f"{nombre_html}"
        "Dpto. Central de Compras Canarias<br><br>"
        "Av. Touroperador Tui, s/n<br>"
        "35100 - Maspalomas (Gran Canaria)<br>"
        f"{tel_html}"
        f"{email_html}"
    )

def _firma_comprador_text(nombre: str, email: str, movil: str) -> str:
    lineas = []
    if nombre:
        lineas.append(nombre)
    lineas.append("Dpto. Central de Compras Canarias")
    lineas.append("")
    lineas.append("Av. Touroperador Tui, s/n")
    lineas.append("35100 - Maspalomas (Gran Canaria)")
    if movil:
        lineas.append(f"(+34) {_formatear_movil_firma(movil)}")
    if email:
        lineas.append(email)
    return "\n".join(lineas)


def _email_header_html(titulo: str, subtitulo: str, color_fondo: str = "#0f2044",
                        color_titulo: str = "#ffffff", color_subtitulo: str = "#b9c3dc") -> str:
    """
    (2026-07-31) Cabecera estándar y ÚNICA para TODOS los emails de la app
    (proveedores, internos, cambio de estado, pendiente de firma,
    cotizaciones, verificación de acceso, etc.): barra superior con
    título/subtítulo a la izquierda y el logo de Princess a la derecha.

    Para cambiar el logo, el texto por defecto o los colores de TODOS los
    emails a la vez, basta con tocar esta única función — todas las
    plantillas la usan en vez de repetir su propia cabecera.
    """
    app_url = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    return f"""
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:{color_fondo};">
      <tr>
        <td style="padding:16px 0 16px 24px;vertical-align:middle;" valign="middle">
          <h2 style="margin:0;color:{color_titulo};font-size:18px;">{titulo}</h2>
          <p style="margin:4px 0 0;color:{color_subtitulo};font-size:13px;">{subtitulo}</p>
        </td>
        <td style="padding:12px 24px 12px 16px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
          <img src="{app_url}/static/logo-sidebar-email.png" alt="Princess Hotels &amp; Resorts"
               width="60" height="56"
               style="height:56px;width:60px;display:block;margin-left:auto;">
        </td>
      </tr>
    </table>
    """


def _email_html_simple(nombre: str, parrafos: list, boton: dict = None,
                        pie_extra: str = None, titulo: str = "Control de Pedidos",
                        subtitulo: str = "Princess Canarias") -> str:
    """
    (2026-07-31) Helper reutilizable para emails cortos tipo "código /
    enlace" (verificación de login, reset de contraseña, avisos simples),
    para no repetir el mismo bloque de estilos inline en cada f-string.
    Devuelve HTML listo para el {{{{message}}}} de EmailJS (triple llave,
    sin escapar) — ver plantilla template_1zrv4ze.

    nombre:    para el saludo ("Hola <strong>nombre</strong>,"). Si viene
               vacío, se omite el saludo.
    parrafos:  lista de strings; cada uno se envuelve en su propio <p>.
               El contenido ya puede traer HTML propio (p.ej. <strong>,
               <br>, el bloque del código en monoespaciado).
    boton:     opcional, {"texto": ..., "url": ...} — genera el mismo
               botón rojo que ya usábamos en el reset de contraseña.
    pie_extra: opcional, línea adicional en gris antes de la firma fija.
    titulo/subtitulo: texto de la cabecera — ver _email_header_html().

    (2026-07-31) Antes esta función devolvía solo párrafos sueltos, sin
    cabecera ni logo — quedaba fuera del rollout de logo de
    v12.27.19-22 pese a ser la plantilla del código de verificación de
    login y (de forma indirecta) la única fuente de estilo del email de
    reset de contraseña. Ahora envuelve el cuerpo con la misma cabecera
    de marca (_email_header_html) que usan las otras 7 plantillas.
    """
    partes = []
    if nombre:
        partes.append(f"<p>Hola <strong>{nombre}</strong>,</p>")
    for p in parrafos:
        partes.append(f"<p>{p}</p>")
    if boton:
        partes.append(
            f'<p><a href="{boton["url"]}" style="background:#8B0000;color:#fff;'
            f'padding:10px 20px;border-radius:4px;text-decoration:none;'
            f'display:inline-block;">{boton["texto"]}</a></p>'
        )
    if pie_extra:
        partes.append(f'<p style="color:#666;font-size:12px;">{pie_extra}</p>')
    cuerpo = "\n    ".join(partes)
    return f"""
    <div style="font-family:sans-serif;max-width:560px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      {_email_header_html(titulo, subtitulo)}
      <div style="padding:22px 24px;font-size:14px;color:#333;line-height:1.6;">
        {cuerpo}
        <p style="color:#666;font-size:12px;margin-top:16px">Control de Pedidos · Princess Canarias</p>
      </div>
    </div>
    """


# ── Plantillas de email por tipo de alerta (v9.5) ─────────────────────────────

def _email_template_enviado_proveedor(pedido: dict, dias: int, urgente: bool, comprador_email: str = "",
                                       comprador_nombre: str = "", comprador_movil: str = "") -> tuple:
    """Pedido enviado al proveedor sin acuse de recibo tras varios días."""
    nivel = "URGENTE" if urgente else "Recordatorio"
    _firma_contacto = _firma_comprador_html(comprador_nombre, comprador_email, comprador_movil)
    subject = f"[{nivel}] Seguimiento pedido Nº {pedido.get('pedido_num','—')} — Princess Hotels & Resorts"
    # (2026-08-27) A petición de Víctor: incluir Total Pedido (si está
    # disponible) e indicar siempre que se trata de base imponible (sin IGIC).
    _tp_ep = pedido.get("total_pedido")
    _fila_tp_ep = f'<br><strong>Total Pedido:</strong> {_fmt_importe_es(_tp_ep)} €' if _tp_ep is not None else ''
    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Dpto. Central de Compras Princess en Canarias",
                            color_fondo="#8B0000", color_subtitulo="#f5c6c6")}
      <div style="padding:24px">
        <p style="background:#fff7e6;border:1px solid #f0c36d;color:#7a5b00;padding:10px 14px;border-radius:4px;font-size:12.5px;margin:0 0 18px">
          ⚠️ Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
        </p>
        <p>Estimado/a proveedor/a,</p>
        <p>Nos ponemos en contacto con usted en relación al pedido que figura a continuación,
           el cual fue tramitado hace <strong>{dias} días</strong> y aún no hemos recibido confirmación de entrega.</p>
        <p style="margin:16px 0;line-height:2;font-size:14px">
          <strong>Pedido Nº:</strong> {pedido.get('pedido_num','—')}<br>
          <strong>Hotel:</strong> {pedido.get('hotel_nombre','—')}<br>
          <strong>Departamento:</strong> {pedido.get('departamento_nombre','—')}<br>
          <strong>Estado actual:</strong> <span style="color:#8B0000">{pedido.get('estado','ENVIADO AL PROVEEDOR')}</span><br>
          <strong>Días transcurridos:</strong> <span style="color:{'#dc2626' if urgente else '#b45309'};font-weight:bold">{dias} días</span>{_fila_tp_ep}{('<br><strong>Observaciones:</strong> ' + pedido['observaciones']) if pedido.get('observaciones') else ''}
        </p>
        {_nota_base_imponible_html() if _tp_ep is not None else ''}
        <p>Le rogamos que nos confirme el estado actual del pedido y la fecha estimada de entrega
           a la mayor brevedad posible.</p>
        {'<p style="color:#dc2626;font-weight:bold;border:1px solid #fca5a5;background:#fee2e2;padding:10px;border-radius:4px">⚠️ ATENCIÓN: Esta es una solicitud urgente. Por favor, responda en el día de hoy.</p>' if urgente else ''}
        <p>Muchas gracias por su colaboración.</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Atentamente,<br>
           {_firma_contacto}</p>
        <p style="font-size:11.5px;color:#8a6d00;background:#fff7e6;border:1px solid #f0c36d;padding:8px 12px;border-radius:4px;margin-top:14px">
          Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
        </p>
      </div>
    </div>
    """
    return subject, body

def _email_template_pendiente_firma(pedido: dict, dias: int, tipo: str) -> tuple:
    """Pedido pendiente de firma (dirección compras o dirección hotel)."""
    if tipo == "PENDIENTE FIRMA DIRECCION COMPRAS":
        dest_label = "Dirección de Compras"
        accion = "firma por parte de Dirección de Compras"
    else:
        dest_label = "Dirección del Hotel"
        accion = "firma por parte de la Dirección del Hotel"
    subject = f"[Recordatorio] Pedido Nº {pedido.get('pedido_num','—')} pendiente de {accion}"

    # (2026-08-03) A petición del usuario: si el pedido está marcado "sujeto
    # al techo de gasto mensual", la persona que gestione la firma debe
    # saberlo — puede requerir una atención distinta (revisar margen
    # disponible, adjuntar justificación, etc.). Se avisa aquí con un aviso
    # destacado, y si Compras ya adjuntó algún listado de apoyo a la
    # solicitud de firma (tipo 'firma_techo_doc'), se menciona también para
    # que quien firme sepa que debe consultarlo en la ficha del pedido.
    aviso_techo_html = ""
    if pedido.get("sujeto_techo"):
        try:
            _n_adj = query(
                "SELECT COUNT(*) as n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo='firma_techo_doc'",
                (pedido.get("id"),), one=True
            )
            _n_adjuntos = (_n_adj or {}).get("n", 0) or 0
        except Exception:
            _n_adjuntos = 0
        _importe_txt = f"{float(pedido['importe']):,.2f} €".replace(",", "@").replace(".", ",").replace("@", ".") if pedido.get("importe") is not None else "—"
        aviso_techo_html = f"""
        <p style="background:#fff3cd;border:1px solid #ffc107;color:#7a5b00;padding:12px 14px;border-radius:4px;font-size:13px;margin:0 0 16px">
          📉 <strong>Este pedido está sujeto al Techo de Gastos mensual del hotel</strong> — familia
          <strong>{pedido.get('familia_nombre') or 'sin especificar'}</strong>, importe <strong>{_importe_txt}</strong>.
          Tenlo en cuenta al gestionar esta firma.
          {f'<br>📎 Compras adjuntó {_n_adjuntos} documento(s) de apoyo a esta solicitud — consúltelo en la ficha del pedido antes de firmar.' if _n_adjuntos else ''}
        </p>"""

    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Control de Pedidos — Aviso interno",
                            color_fondo="#1a3a6b", color_subtitulo="#a8c0e8")}
      <div style="padding:24px">
        {aviso_techo_html}
        <p>Se le notifica que el siguiente pedido lleva <strong>{dias} días</strong>
           pendiente de {accion}:</p>
        <table style="width:100%;border-collapse:collapse;margin:16px 0;font-size:14px">
          <tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold;width:40%">Pedido Nº</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('pedido_num','—')}</td></tr>
          <tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Nº de Orden</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('norden','—')}</td></tr>
          <tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Hotel</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('hotel_nombre','—')}</td></tr>
          <tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Departamento</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('departamento_nombre','—')}</td></tr>
          <tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Proveedor</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('proveedor_nombre','—')}</td></tr>
          <tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Días en espera</td>
              <td style="padding:8px 12px;border:1px solid #ddd;color:#b45309;font-weight:bold">{dias} días</td></tr>
          {f'<tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Total Pedido (base imponible)</td><td style="padding:8px 12px;border:1px solid #ddd">{_fmt_importe_es(pedido.get("total_pedido"))} €</td></tr>' if pedido.get("total_pedido") is not None else ''}
          {f'<tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Observaciones</td><td style="padding:8px 12px;border:1px solid #ddd">{pedido["observaciones"]}</td></tr>' if pedido.get("observaciones") else ''}
        </table>
        {_nota_base_imponible_html() if pedido.get("total_pedido") is not None else ''}
        <p>Por favor, gestione con {dest_label} la revisión y firma del pedido
           a la mayor brevedad posible para no retrasar el proceso de compra.</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Mensaje automático generado por el sistema de Control de Pedidos.<br>
           <strong>Princess Hotels &amp; Resorts</strong></p>
      </div>
    </div>
    """
    return subject, body

def _email_template_entrega_parcial(pedido: dict, dias: int, comprador_email: str = "",
                                     comprador_nombre: str = "", comprador_movil: str = "") -> tuple:
    """Pedido con entrega parcial sin cierre."""
    subject = f"[Seguimiento] Pedido Nº {pedido.get('pedido_num','—')} — Entrega parcial pendiente de completar"
    _firma_contacto = _firma_comprador_html(comprador_nombre, comprador_email, comprador_movil)
    # (2026-08-27) A petición de Víctor: incluir Total Pedido y el detalle de
    # entregas parciales registradas hasta la fecha (con su base imponible),
    # indicando siempre que los importes son base imponible (sin IGIC).
    _tp_pp = pedido.get("total_pedido")
    _fila_tp_pp = f'<br><strong>Total Pedido:</strong> {_fmt_importe_es(_tp_pp)} €' if _tp_pp is not None else ''
    _resumen_ent_pp = _resumen_entregas(pedido)
    _bloque_ent_pp = _html_bloque_entregas(_resumen_ent_pp, pedido.get("estado"))
    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Dpto. Central de Compras Princess en Canarias",
                            color_fondo="#8B0000", color_subtitulo="#f5c6c6")}
      <div style="padding:24px">
        <p style="background:#fff7e6;border:1px solid #f0c36d;color:#7a5b00;padding:10px 14px;border-radius:4px;font-size:12.5px;margin:0 0 18px">
          ⚠️ Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
        </p>
        <p>Estimado/a proveedor/a,</p>
        <p>Le contactamos en relación al pedido indicado, cuya entrega se registró de forma
           <strong>parcial</strong> hace <strong>{dias} días</strong> y aún está pendiente de completarse.</p>
        <p style="margin:16px 0;line-height:2;font-size:14px">
          <strong>Pedido Nº:</strong> {pedido.get('pedido_num','—')}<br>
          <strong>Hotel:</strong> {pedido.get('hotel_nombre','—')}<br>
          <strong>Departamento:</strong> {pedido.get('departamento_nombre','—')}<br>
          <strong>Estado actual:</strong> <span style="color:#8B0000">ENTREGA PARCIAL</span><br>
          <strong>Días transcurridos:</strong> <span style="color:#b45309;font-weight:bold">{dias} días</span>{_fila_tp_pp}{('<br><strong>Observaciones:</strong> ' + pedido['observaciones']) if pedido.get('observaciones') else ''}
        </p>
        {_bloque_ent_pp if _bloque_ent_pp else (_nota_base_imponible_html() if _tp_pp is not None else '')}
        <p>Le rogamos que nos informe sobre la fecha prevista para completar la entrega pendiente.</p>
        <p>Muchas gracias.</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Atentamente,<br>
           {_firma_contacto}</p>
        <p style="font-size:11.5px;color:#8a6d00;background:#fff7e6;border:1px solid #f0c36d;padding:8px 12px;border-radius:4px;margin-top:14px">
          Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
        </p>
      </div>
    </div>
    """
    return subject, body

def _email_template_pendiente_cotizacion(pedido: dict, dias: int, urgente: bool, comprador_email: str = "",
                                          comprador_nombre: str = "", comprador_movil: str = "") -> tuple:
    """Pedido pendiente de cotización del proveedor."""
    nivel = "URGENTE" if urgente else "Solicitud de cotización"
    subject = f"[{nivel}] Cotización solicitada — {pedido.get('hotel_nombre','Princess Hotels')} — Princess Hotels & Resorts"
    _firma_contacto = _firma_comprador_html(comprador_nombre, comprador_email, comprador_movil)
    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Dpto. Central de Compras Princess en Canarias",
                            color_fondo="#8B0000", color_subtitulo="#f5c6c6")}
      <div style="padding:24px">
        <p style="background:#fff7e6;border:1px solid #f0c36d;color:#7a5b00;padding:10px 14px;border-radius:4px;font-size:12.5px;margin:0 0 18px">
          ⚠️ Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
        </p>
        <p>Estimado/a proveedor/a,</p>
        <p>Le recordamos que hace <strong>{dias} días</strong> se le solicitó cotización
           para el siguiente pedido y aún estamos a la espera de su propuesta económica.</p>
        <p style="margin:16px 0;line-height:2;font-size:14px">
          <strong>Pedido Nº:</strong> {pedido.get('pedido_num','—')}<br>
          <strong>Hotel:</strong> {pedido.get('hotel_nombre','—')}<br>
          <strong>Departamento:</strong> {pedido.get('departamento_nombre','—')}<br>
          <strong>Estado actual:</strong> <span style="color:#8B0000">PENDIENTE COTIZACIÓN</span><br>
          <strong>Días transcurridos:</strong> <span style="color:{'#dc2626' if urgente else '#b45309'};font-weight:bold">{dias} días</span>{('<br><strong>Observaciones:</strong> ' + pedido['observaciones']) if pedido.get('observaciones') else ''}
        </p>
        {'<p style="color:#dc2626;font-weight:bold;border:1px solid #fca5a5;background:#fee2e2;padding:10px;border-radius:4px">⚠️ URGENTE: Necesitamos su cotización hoy para no retrasar la tramitación del pedido.</p>' if urgente else '<p>Le agradecemos que nos envíe su mejor oferta a la mayor brevedad posible.</p>'}
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Atentamente,<br>
           {_firma_contacto}</p>
        <p style="font-size:11.5px;color:#8a6d00;background:#fff7e6;border:1px solid #f0c36d;padding:8px 12px;border-radius:4px;margin-top:14px">
          Este correo es exclusivo para notificaciones automáticas. Por favor, responda única y exclusivamente a la dirección que firma este comunicado.
        </p>
      </div>
    </div>
    """
    return subject, body

def _build_alerta_email(pedido: dict, dias: int, nivel: str) -> tuple:
    """Selecciona la plantilla correcta según el estado del pedido y devuelve (subject, body, es_proveedor).
    Devuelve (None, None, False) si no hay comprador con email asignado al hotel."""
    estado    = pedido.get("estado", "")
    urgente   = nivel == "urgente"
    # Obtener email del comprador responsable del hotel para incluir en la firma
    _compradores = _get_compradores_cc(pedido.get("hotel_codigo",""))
    if not (_compradores and _compradores[0].get("email")):
        log.warning("[ALERTA EMAIL] Pedido %s: no hay comprador con email asignado al hotel %s — email de alerta omitido",
                    pedido.get("id"), pedido.get("hotel_codigo",""))
        return None, None, False
    _comprador_email  = _compradores[0]["email"]
    _comprador_nombre = _compradores[0].get("nombre") or ""
    _comprador_movil  = _compradores[0].get("movil") or ""
    if estado == "ENVIADO AL PROVEEDOR":
        s, b = _email_template_enviado_proveedor(pedido, dias, urgente, _comprador_email,
                                                   _comprador_nombre, _comprador_movil)
        return s, b, True
    elif estado in ("PENDIENTE FIRMA DIRECCION COMPRAS", "PENDIENTE DE FIRMA DIRECCION HOTEL"):
        s, b = _email_template_pendiente_firma(pedido, dias, estado)
        return s, b, False
    elif estado == "ENTREGA PARCIAL":
        s, b = _email_template_entrega_parcial(pedido, dias, _comprador_email,
                                                 _comprador_nombre, _comprador_movil)
        return s, b, True
    elif estado == "PENDIENTE COTIZACIÓN":
        s, b = _email_template_pendiente_cotizacion(pedido, dias, urgente, _comprador_email,
                                                       _comprador_nombre, _comprador_movil)
        return s, b, True
    return None, None, False


def _email_template_cotizacion_sin_proveedor(pedido: dict, dias: int) -> tuple:
    """
    Aviso interno al comprador del hotel cuando un pedido lleva días en
    PENDIENTE COTIZACIÓN pero todavía no tiene proveedor asignado — no hay
    a quién reclamar, así que se avisa a quien puede resolverlo.
    """
    fecha_sol = pedido.get("fecha_solicitud")
    if fecha_sol and hasattr(fecha_sol, "strftime"):
        fecha_str = fecha_sol.strftime("%d/%m/%Y")
    elif fecha_sol:
        parts = str(fecha_sol)[:10].split("-")
        fecha_str = "/".join(reversed(parts)) if len(parts) == 3 else str(fecha_sol)[:10]
    else:
        fecha_str = "—"

    subject = f"[Recordatorio] Cotización pendiente sin proveedor asignado — Pedido Nº {pedido.get('pedido_num','—')}"
    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Control de Pedidos — Aviso interno",
                            color_fondo="#1a3a6b", color_subtitulo="#a8c0e8")}
      <div style="padding:24px">
        <p>Te recordamos que tenemos pendiente la cotización referente a la solicitud
           con fecha <strong>{fecha_str}</strong>, que lleva <strong>{dias} días</strong> sin resolver.</p>
        <table style="width:100%;border-collapse:collapse;margin:16px 0;font-size:14px">
          <tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold;width:40%">Pedido Nº</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('pedido_num','—')}</td></tr>
          <tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Nº de Orden</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('norden','—')}</td></tr>
          <tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Hotel</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('hotel_nombre','—')}</td></tr>
          <tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Departamento</td>
              <td style="padding:8px 12px;border:1px solid #ddd">{pedido.get('departamento_nombre','—')}</td></tr>
          <tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Proveedor</td>
              <td style="padding:8px 12px;border:1px solid #ddd;color:#b45309;font-weight:bold">Sin proveedor asignado hasta la fecha</td></tr>
          <tr><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Días en espera</td>
              <td style="padding:8px 12px;border:1px solid #ddd;color:#b45309;font-weight:bold">{dias} días</td></tr>
          {f'<tr style="background:#f5f5f5"><td style="padding:8px 12px;border:1px solid #ddd;font-weight:bold">Observaciones</td><td style="padding:8px 12px;border:1px solid #ddd">{pedido["observaciones"]}</td></tr>' if pedido.get("observaciones") else ''}
        </table>
        <p>Por favor, asigna un proveedor y solicita la cotización a la mayor brevedad
           posible para no retrasar la tramitación del pedido.</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Mensaje automático generado por el sistema de Control de Pedidos.<br>
           <strong>Princess Hotels &amp; Resorts</strong></p>
      </div>
    </div>
    """
    return subject, body


def _encolar_aviso_cotizacion_sin_proveedor(pedido: dict, dias: int) -> bool:
    """
    Sustituye a la reclamación automática al proveedor cuando un pedido en
    PENDIENTE COTIZACIÓN aún no tiene proveedor asignado: en vez de omitir
    en silencio, envía un único email interno (todos los compradores del
    hotel juntos en "Para:") avisando de la cotización pendiente y de que
    no hay proveedor asignado. Comparte tipo/dedup ('reclamacion_proveedor_auto')
    con la reclamación normal, así que respeta el mismo ciclo/umbral crítico
    configurado en Config Alertas para PENDIENTE COTIZACIÓN.
    """
    compradores = _get_compradores_cc(pedido.get("hotel_codigo", ""))
    destinos = [e for c in compradores for e in _emails_usuario(c)]
    if not destinos:
        log.warning("[RECLAMACION-AUTO] Pedido %s: PENDIENTE COTIZACIÓN sin proveedor y sin comprador con email — aviso omitido",
                    pedido.get("id"))
        return False

    subject, body_html = _email_template_cotizacion_sin_proveedor(pedido, dias)
    destino = ", ".join(destinos)
    _encolar_email_sistema(
        "reclamacion_proveedor_auto", [destino], subject, body_html,
        pedido_id=pedido.get("id"),
    )
    log.info("[RECLAMACION-AUTO] Pedido %s — sin proveedor asignado, aviso interno encolado (1 envío) a comprador(es) %s",
              pedido.get("id"), destino)
    return True


def _encolar_aviso_firma_pendiente_auto(pedido: dict, dias: int, estado: str) -> bool:
    """
    v12.23.8 — Aviso automático al comprador cuando un pedido lleva días
    pendiente de firma (Dirección de Compras o Dirección del Hotel).

    A diferencia de la reclamación al proveedor, este aviso es interno y
    va SIEMPRE al/los comprador(es) del hotel — no hay proveedor implicado
    en estos dos estados. Se dispara con el mismo criterio que ya usa el
    Telegram automático para estos estados (1ª alerta + repetición por
    ciclo, configurado en Config Alertas), sin depender del nivel "urgente"
    (que para estos dos estados está desactivado por defecto, 0 = nunca) —
    lo decide el caller (_job_alertas_diarias_inner) pasando por aquí
    exactamente cuando también envía el Telegram automático.

    Reutiliza la plantilla _email_template_pendiente_firma() ya existente
    (antes solo se usaba para la propuesta manual desde el panel).

    Devuelve True si se encoló correctamente, False si se omitió (ya se
    mandó a mano hoy, o no hay comprador con email para este hotel).
    """
    if _ya_reclamado_hoy_manual(pedido.get("id"), tipo="alerta_interno"):
        log.info("[AVISO-FIRMA-AUTO] Pedido %s: ya se envió un aviso manual interno hoy — se omite el automático",
                  pedido.get("id"))
        return False

    compradores = _get_compradores_cc(pedido.get("hotel_codigo", ""))
    destinos = [e for c in compradores for e in _emails_usuario(c)]
    if not destinos:
        log.warning("[AVISO-FIRMA-AUTO] Pedido %s: sin comprador con email en el hotel %s — aviso omitido",
                    pedido.get("id"), pedido.get("hotel_codigo", ""))
        return False

    subject, body_html = _email_template_pendiente_firma(pedido, dias, estado)
    destino = ", ".join(destinos)
    _encolar_email_sistema(
        "aviso_firma_pendiente_auto", [destino], subject, body_html,
        pedido_id=pedido.get("id"),
    )
    log.info("[AVISO-FIRMA-AUTO] Pedido %s — aviso de firma pendiente encolado (1 envío) a comprador(es) %s",
              pedido.get("id"), destino)
    return True


def _encolar_reclamacion_proveedor_auto(pedido: dict, dias: int, nivel: str) -> bool:
    """
    v12.19.0 — Reclamación automática por email al proveedor.
    v12.23.6 — extendida también a PENDIENTE COTIZACIÓN.

    Se llama desde el job diario de alertas cuando:
      - el plazo de entrega informado por el proveedor ya venció (nivel ==
        'urgente' en _alertas_plazo_entrega) y el pedido sigue en ENVIADO AL
        PROVEEDOR o ENTREGA PARCIAL, o
      - el pedido lleva sin cotizar más de lo configurado en Config Alertas
        (estado PENDIENTE COTIZACIÓN, nivel 'urgente' vía _build_umbrales).
    Reutiliza exactamente la misma plantilla que el envío
    manual (_build_alerta_email) y la encola en emails_sistema_pendientes
    con los compradores del hotel en copia (bcc), porque esta app no tiene
    SMTP propio — el despacho real lo hace el navegador de un admin vía
    EmailJS (mismo patrón que el resto de la cola de sistema).

    Devuelve True si se encoló correctamente, False si se omitió (sin
    comprador con email para la firma, sin destinatario alguno, etc.).
    Caso particular: si es PENDIENTE COTIZACIÓN y el pedido aún no tiene
    proveedor asignado, no reclama a nadie externo — en su lugar avisa
    por email al/los comprador(es) del hotel (ver
    _encolar_aviso_cotizacion_sin_proveedor).
    """
    if pedido.get("estado") not in ("ENVIADO AL PROVEEDOR", "ENTREGA PARCIAL", "PENDIENTE COTIZACIÓN"):
        log.info("RECLAMACION-DEBUG pedido=%s omitido: estado=%s no es ENVIADO AL PROVEEDOR/ENTREGA PARCIAL/PENDIENTE COTIZACIÓN",
                  pedido.get("id"), pedido.get("estado"))
        return False

    # (2026-07-30) Si un comprador ya mandó una reclamación manual al
    # proveedor hoy (botón "Re-notificar"), no duplicamos con la automática.
    if _ya_reclamado_hoy_manual(pedido.get("id")):
        log.info("[RECLAMACION-AUTO] Pedido %s: ya se envió una reclamación manual hoy — se omite la automática",
                  pedido.get("id"))
        return False

    subject, body_html, es_proveedor = _build_alerta_email(pedido, dias, nivel)
    if not subject or not es_proveedor:
        log.info("RECLAMACION-DEBUG pedido=%s omitido: subject=%s es_proveedor=%s (probable falta de comprador con email en el hotel, o estado no soportado por _build_alerta_email)",
                  pedido.get("id"), bool(subject), es_proveedor)
        return False

    proveedor_emails = _get_proveedor_emails_principales(pedido.get("proveedor_id"), pedido.get("hotel_id"))
    if not proveedor_emails and pedido.get("proveedor_email"):
        proveedor_emails = [pedido["proveedor_email"]]
    if not proveedor_emails:
        # (2026-07-31) PENDIENTE COTIZACIÓN sin proveedor asignado todavía:
        # no hay a quién reclamar, así que en vez de omitir en silencio se
        # avisa por email únicamente al/los comprador(es) del hotel, para
        # que gestionen la asignación de proveedor. El resto de estados
        # (ENVIADO AL PROVEEDOR / ENTREGA PARCIAL) siempre tienen proveedor
        # ya asignado en ese punto del flujo, así que solo aplica aquí.
        if pedido.get("estado") == "PENDIENTE COTIZACIÓN":
            return _encolar_aviso_cotizacion_sin_proveedor(pedido, dias)
        log.warning("[RECLAMACION-AUTO] Pedido %s: proveedor sin email — reclamación automática omitida",
                    pedido.get("id"))
        return False

    compradores = _get_compradores_cc(pedido.get("hotel_codigo", ""))
    cc_emails = [e for c in compradores for e in _emails_usuario(c)]

    # (2026-07-30) FIX: antes se pasaba `proveedor_emails` (una lista) como
    # `destinatarios_email` a `_encolar_email_sistema()`, que encola UNA fila
    # — y por tanto UN envío independiente — POR CADA elemento de esa lista.
    # Si un proveedor tenía 2 o 3 contactos marcados "principal", eso
    # generaba 2 o 3 reclamaciones separadas para el mismo pedido en la
    # misma tanda (reportado por el usuario: "el 40130 3 veces..."). Se
    # quiere un único envío, con todos los contactos principales juntos en
    # el "Para:" — mismo patrón ya usado en el aviso al cambiar de estado
    # (`_destino_proveedor = ", ".join(_proveedor_emails)`, más arriba en
    # este archivo): EmailJS sí admite varias direcciones separadas por
    # comas en un solo campo "to_email".
    destino_proveedor = ", ".join(proveedor_emails)

    _encolar_email_sistema(
        "reclamacion_proveedor_auto", [destino_proveedor], subject, body_html,
        cc_emails=cc_emails, pedido_id=pedido.get("id"),
    )
    log.info("[RECLAMACION-AUTO] Pedido %s — reclamación encolada (1 envío) a %s (cc: %s)",
              pedido.get("id"), destino_proveedor, cc_emails)
    return True


def _whatsapp_text(pedido: dict, dias: int, nivel: str) -> str:
    """Genera el texto de WhatsApp/Telegram (plano, sin HTML) para notificación al comprador."""
    emoji      = "🔴" if nivel == "urgente" else "🟡"
    nivel_txt  = "ALERTA URGENTE" if nivel == "urgente" else "AVISO"
    hotel_cod  = pedido.get("hotel_codigo", "—")
    hotel_nom  = pedido.get("hotel_nombre", "")
    pedido_sap = pedido.get("pedido_num") or ""
    norden     = pedido.get("norden") or ""
    proveedor  = pedido.get("proveedor_nombre") or ""
    estado     = pedido.get("estado", "—")

    lineas = [f"{emoji} *{nivel_txt}*", ""]
    lineas.append(f"🏨 Hotel: *{hotel_cod}* — {hotel_nom}")
    if pedido_sap:
        lineas.append(f"📄 Pedido SAP: *{pedido_sap}*")
    elif norden:
        lineas.append(f"📄 Línea #: *{norden}*")
    if proveedor:
        lineas.append(f"🏢 Proveedor: {proveedor}")
    lineas.append(f"📋 Estado: {estado}")
    lineas.append(f"⏳ Días transcurridos: *{dias}*")
    lineas += ["", "— Control Pedidos Princess Canarias"]
    return "\n".join(lineas)

def _log_whatsapp(db, pedido_id, tipo, destinatario, mensaje, enviado, error=None):
    with db.cursor() as cur:
        cur.execute(
            "INSERT INTO whatsapp_log (pedido_id,tipo,destinatario,mensaje,enviado,error) VALUES (%s,%s,%s,%s,%s,%s)",
            (pedido_id, tipo, destinatario, mensaje, 1 if enviado else 0, error)
        )

# ── API: preparar propuesta de email por alerta ───────────────────────────────

PEDIDO_SELECT_ALERTA = """
    SELECT p.id, p.norden, p.pedido_num, p.presupuesto_num, p.estado,
           p.fecha_tramitacion, p.fecha_solicitud, p.observaciones,
           p.proveedor_id, p.hotel_id,
           p.sujeto_techo, p.familia_id, p.importe,
           fam.nombre as familia_nombre,
           h.codigo as hotel_codigo, h.nombre as hotel_nombre,
           d.nombre as departamento_nombre,
           pr.nombre as proveedor_nombre,
           (SELECT pc.email
              FROM proveedor_contactos pc
             WHERE pc.proveedor_id = pr.id AND pc.es_principal = 1
               AND pc.email IS NOT NULL AND pc.email != ''
               AND (EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id AND pch.hotel_id = p.hotel_id)
                    OR NOT EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch2 WHERE pch2.contacto_id = pc.id))
             ORDER BY EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch3 WHERE pch3.contacto_id = pc.id AND pch3.hotel_id = p.hotel_id) DESC,
                      pc.orden, pc.id
             LIMIT 1) as proveedor_email
    FROM pedidos p
    LEFT JOIN hoteles h ON p.hotel_id = h.id
    LEFT JOIN departamentos d ON p.departamento_id = d.id
    LEFT JOIN proveedores pr ON p.proveedor_id = pr.id
    LEFT JOIN familias fam ON p.familia_id = fam.id
"""

@app.route("/api/alertas/<int:pedido_id>/email-preview", methods=["GET"])
@login_required
def alerta_email_preview(pedido_id):
    """Devuelve la propuesta de email (destinatarios + cuerpo) para una alerta concreta."""
    dias_str = request.args.get("dias", "0")
    nivel    = request.args.get("nivel", "aviso")
    try:
        dias = int(dias_str)
    except Exception:
        dias = 0

    pedido = row_to_dict(query(f"{PEDIDO_SELECT_ALERTA} WHERE p.id=%s", (pedido_id,), one=True))
    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    subject, body_html, es_proveedor = _build_alerta_email(pedido, dias, nivel)
    if not subject:
        return jsonify({"error": "No hay plantilla para este estado"}), 400

    hotel_codigo = pedido.get("hotel_codigo", "")
    compradores  = _get_compradores_cc(hotel_codigo)

    # Destinatario principal
    if es_proveedor:
        _proveedor_emails = _get_proveedor_emails_principales(pedido.get("proveedor_id"), pedido.get("hotel_id"))
        to_email   = ", ".join(_proveedor_emails)
        to_nombre  = pedido.get("proveedor_nombre") or ""
    else:
        # Email interno: comprador responsable como destinatario (con su
        # email2 también, si tiene uno asignado — ver _emails_usuario())
        to_email  = ", ".join(_emails_usuario(compradores[0])) if compradores else ""
        to_nombre = compradores[0]["nombre"] if compradores else ""

    _to_emails_set = {e.strip() for e in to_email.split(",") if e.strip()}
    cc_emails = [e for c in compradores for e in _emails_usuario(c) if e not in _to_emails_set]

    # WhatsApp text para compradores (legacy manual)
    wa_text = _whatsapp_text(pedido, dias, nivel)
    wa_recipients = [{"nombre": c["nombre"], "movil": c.get("movil","")} for c in compradores if c.get("movil")]

    # Telegram — destinatarios configurados para este hotel (v12.17.0,
    # evento 'alerta_pedido_hotel') — antes mostraba _get_compradores_hotel()
    # a secas, lo que podía no coincidir con lo que realmente se envía ahora.
    _compradores_telegram = _resolver_notificacion("alerta_pedido_hotel", "telegram", pedido.get("hotel_id"))
    telegram_recipients = [
        {"username": c.get("username"), "chat_id": c.get("telegram_chat_id"), "nombre": c.get("nombre", c.get("username"))}
        for c in _compradores_telegram if c.get("telegram_chat_id")
    ]

    return jsonify({
        "pedido_id":     pedido_id,
        "estado":        pedido.get("estado"),
        "hotel_codigo":  hotel_codigo,
        "hotel_nombre":  pedido.get("hotel_nombre"),
        "pedido_num":    pedido.get("pedido_num"),
        "proveedor_nombre": pedido.get("proveedor_nombre"),
        "es_proveedor":  es_proveedor,
        "to_email":      to_email,
        "to_nombre":     to_nombre,
        "cc_emails":     cc_emails,
        "compradores":   compradores,
        "subject":       subject,
        "body_html":     body_html,
        "wa_text":       wa_text,
        "wa_recipients": wa_recipients,
        "telegram_recipients": telegram_recipients,
        "dias":            dias,
        "nivel":           nivel,
    })

@app.route("/api/alertas/<int:pedido_id>/enviar-email", methods=["POST"])
@login_required
def alerta_enviar_email(pedido_id):
    """Envía el email de alerta al destinatario/s indicados y lo registra en emails_log."""
    data      = request.get_json(silent=True) or {}
    to_email  = (data.get("to_email") or "").strip()
    subject   = (data.get("subject") or "").strip()
    body_html = data.get("body_html") or ""
    body_text = data.get("body_text") or ""
    dias      = int(data.get("dias", 0))
    nivel     = data.get("nivel", "aviso")
    es_proveedor = data.get("es_proveedor", False)

    if not to_email or not subject:
        return jsonify({"error": "Faltan destinatario o asunto"}), 400

    # ── Recalcular CC en backend para no depender del frontend ────────────────
    # El frontend puede no enviar cc_emails o enviarlos incompletos.
    # Siempre recalculamos los compradores asignados al hotel del pedido.
    pedido_data = row_to_dict(query(f"{PEDIDO_SELECT_ALERTA} WHERE p.id=%s", (pedido_id,), one=True))
    if pedido_data:
        hotel_codigo = pedido_data.get("hotel_codigo", "")
        compradores_hotel = _get_compradores_cc(hotel_codigo)
        # CC = todos los emails (principal + email2) de los compradores del
        # hotel, excepto los que ya estén en el TO principal (el TO puede
        # traer varias direcciones separadas por coma si el destinatario
        # tiene email2)
        _to_emails_set = {e.strip() for e in to_email.split(",") if e.strip()}
        cc_emails_backend = [
            e for c in compradores_hotel for e in _emails_usuario(c)
            if e not in _to_emails_set
        ]
    else:
        cc_emails_backend = []

    # Combinar con cualquier CC extra que venga del frontend (sin duplicados)
    cc_frontend = [e.strip() for e in (data.get("cc_emails") or []) if e.strip()]
    cc_emails = list(dict.fromkeys(cc_emails_backend + [e for e in cc_frontend if e not in cc_emails_backend]))

    log.info("Alerta email pedido %s → TO: %s | CC/BCC: %s", pedido_id, to_email, cc_emails)

    db = get_db()
    resultados = []

    # Registro en log — el envío real lo hace el frontend vía EmailJS
    tipo_log = "alerta_proveedor" if es_proveedor else "alerta_interno"
    _log_email(db, pedido_id, tipo_log, to_email, subject, False, "Pendiente de envío vía EmailJS")
    resultados.append({"email": to_email, "ok": True, "error": None, "mode": "emailjs_pending"})

    # ── Telegram automático — se dispara siempre al enviar la alerta ──────────
    # pedido_data ya fue cargado arriba para calcular los CC
    if not pedido_data:
        pedido_data = row_to_dict(query(f"{PEDIDO_SELECT_ALERTA} WHERE p.id=%s", (pedido_id,), one=True))
    telegram_resultados = []
    if pedido_data:
        telegram_resultados = _enviar_telegram_compradores(pedido_data, dias, nivel)
        for tr in telegram_resultados:
            _log_whatsapp(db, pedido_id, "telegram_auto",
                          tr.get("username", "?"),
                          f"Alerta {nivel} — {pedido_data.get('hotel_codigo')} · Pedido {pedido_data.get('pedido_num')}",
                          tr["ok"], tr.get("error"))

    db.commit()
    todos_ok = all(r["ok"] for r in resultados)
    primer_error = next((r["error"] for r in resultados if not r["ok"]), None)
    return jsonify({
        "ok": todos_ok,
        "resultados": resultados,
        "error": primer_error,
        "telegram": telegram_resultados,
        # Datos para que el frontend envíe vía EmailJS
        "email_pendiente": {
            "to_email":  to_email,
            "cc_emails": cc_emails,
            "subject":   subject,
            "body_html": body_html,
            "body_text": body_text,
        },
    })

@app.route("/api/alertas/<int:pedido_id>/log-whatsapp", methods=["POST"])
@login_required
def alerta_log_whatsapp(pedido_id):
    """Registra en BD que se inició un envío de WhatsApp (el envío real es client-side via wa.me)."""
    data         = request.get_json(silent=True) or {}
    destinatario = (data.get("destinatario") or "").strip()
    mensaje      = (data.get("mensaje") or "").strip()
    if not destinatario:
        return jsonify({"error": "Destinatario requerido"}), 400
    db = get_db()
    _log_whatsapp(db, pedido_id, "alerta_comprador", destinatario, mensaje, True)
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/alertas/<int:pedido_id>/enviar-telegram", methods=["POST"])
@login_required
def alerta_enviar_telegram(pedido_id):
    """Envía alerta automática por Telegram a los compradores del hotel."""
    data  = request.get_json(silent=True) or {}
    dias  = int(data.get("dias", 0))
    nivel = data.get("nivel", "aviso")

    pedido = row_to_dict(query(f"{PEDIDO_SELECT_ALERTA} WHERE p.id=%s", (pedido_id,), one=True))
    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    resultados = _enviar_telegram_compradores(pedido, dias, nivel)

    db = get_db()
    for tr in resultados:
        _log_whatsapp(db, pedido_id, "telegram_auto",
                      tr.get("username", "?"),
                      f"Alerta {nivel} — {pedido.get('hotel_codigo')} · Pedido {pedido.get('pedido_num')}",
                      tr["ok"], tr.get("error"))
    db.commit()

    todos_ok = all(r["ok"] for r in resultados)
    return jsonify({"ok": todos_ok, "resultados": resultados})



def _next_norden(db):
    year = datetime.now().year
    with db.cursor() as cur:
        cur.execute(
            "SELECT COALESCE(MAX(norden), 0) as mx FROM pedidos WHERE EXTRACT(YEAR FROM creado_en) = %s",
            (year,)
        )
        row = cur.fetchone()
    return (row["mx"] or 0) + 1

# ── Rutas estáticas ────────────────────────────────────────────────────────────

def _index_html_bytes_and_hash():
    """
    Lee templates/index.html una sola vez y devuelve (bytes, hash).
    El hash es el mismo cálculo (MD5, 12 caracteres) que ya usaba
    /api/version, así que ambos endpoints quedan siempre consistentes.
    """
    tpl_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates", "index.html")
    with open(tpl_path, "rb") as f:
        content = f.read()
    return content, hashlib.md5(content).hexdigest()[:12]

@app.route("/")
def index():
    """
    ── Fix egress (Jul 2026) ──────────────────────────────────────────
    index.html pesa varios cientos de KB y se servía sin ninguna cabecera
    de caché: cada apertura de la app, cada refresco de pestaña y cada
    comprobación del navegador volvía a descargar el archivo entero,
    convirtiéndolo en el mayor origen de egress del proyecto.
    Ahora se sirve con ETag (el mismo hash MD5 de /api/version) y
    Cache-Control: no-cache, así que el navegador siempre revalida con
    una petición condicional ligera (cabeceras, sin cuerpo) y solo
    vuelve a descargar el archivo completo cuando de verdad cambió
    tras un despliegue.
    """
    try:
        content, version_hash = _index_html_bytes_and_hash()
    except Exception:
        return send_from_directory("templates", "index.html")

    if request.headers.get("If-None-Match") == version_hash:
        return Response(status=304)

    resp = Response(content, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["ETag"] = version_hash
    return resp

@app.route("/api/version")
def app_version():
    """
    Devuelve un hash MD5 del contenido real de index.html.
    Cualquier cambio en el archivo, por pequeño que sea, produce un hash diferente.
    """
    try:
        _, version_hash = _index_html_bytes_and_hash()
    except Exception:
        version_hash = "unknown"
    return jsonify({"version": version_hash})

def _resumen_ultima_version_changelog(contenido: str) -> str:
    """
    (2026-09-03) Extrae de la entrada MÁS RECIENTE (la primera) de
    CHANGELOG.md solo la cabecera de versión ("vX.Y.Z — fecha") y el
    título-resumen de una línea que la sigue (el emoji + frase corta),
    sin ninguno de los párrafos de detalle (Petición/Diagnóstico/
    Cambio/Verificación/Entrega...) que sí ve un admin. Ver app_changelog().
    """
    lineas = contenido.splitlines()
    version_linea, titulo = None, None
    for i, linea in enumerate(lineas):
        if linea.startswith("# "):
            version_linea = linea[2:].strip()
            for siguiente in lineas[i + 1:]:
                if siguiente.strip():
                    titulo = siguiente.strip()
                    break
            break
    if version_linea and titulo:
        return f"{version_linea}\n\n{titulo}"
    return version_linea or "Hay una nueva versión disponible."


@app.route("/api/changelog")
def app_changelog():
    """
    Devuelve las notas de versión para el modal de "nueva versión
    detectada" del cliente (ver _mostrarModalNuevaVersion en
    templates/index.html).

    (2026-09-03) A petición de Víctor — "el exceso de información aturde
    al usuario, vamos a limitar la pantalla... solo mensaje de nueva
    actualización (para forzar la misma) con un título resumen pero sin
    entrar en detalles... solo mostrar todo a los administradores" —:
    el changelog COMPLETO (CHANGELOG.md entero, cada entrada con
    petición/diagnóstico/cambio/verificación) solo se sirve cuando
    `session.get("rol") == "admin"`, como `{"changelog": "..."}`. El
    resto de roles (compras/hotel) recibe únicamente el título-resumen de
    una línea de la última entrada, como `{"resumen": "..."}` — basta
    para saber que hay una actualización y forzar la recarga, sin los
    párrafos de detalle que antes veía todo el mundo por igual. El
    frontend decide qué bloque del modal mostrar según cuál de las dos
    claves recibe la respuesta.
    """
    try:
        changelog_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "CHANGELOG.md")
        with open(changelog_path, "r", encoding="utf-8") as f:
            contenido = f.read()
    except FileNotFoundError:
        contenido = "_No hay notas de versión disponibles._"
    except Exception as e:
        contenido = f"_Error al leer el changelog: {e}_"

    if session.get("rol") == "admin":
        return jsonify({"changelog": contenido})
    return jsonify({"resumen": _resumen_ultima_version_changelog(contenido)})

@app.route("/static/<path:filename>")
def static_files(filename):
    """
    ── Fix egress (Jul 2026) ──────────────────────────────────────────
    Los ficheros de /static (logos, etc.) son estáticos y cambian solo
    de forma manual junto con un despliegue. send_from_directory ya
    gestiona automáticamente ETag/Last-Modified y responde 304 a
    peticiones condicionales; con max_age además añadimos Cache-Control,
    así el navegador ni siquiera necesita revalidar durante 7 días.
    """
    return send_from_directory("static", filename, max_age=604800)

# ── API Auth ───────────────────────────────────────────────────────────────────

DIAS_VERIFICACION_EMAIL = 3  # a partir de cuántos días sin login se exige código por email


def _es_hash_password(valor):
    """
    Distingue un hash de werkzeug (pbkdf2:... / scrypt:...) de una
    contraseña todavía en texto plano heredada de antes de la v12.29.37.
    """
    return isinstance(valor, str) and valor.startswith(("pbkdf2:", "scrypt:"))


def _verifica_y_migra_password(user, password_recibida):
    """
    Verifica la contraseña recibida contra la almacenada, con migración
    transparente al hash (v12.29.37 — corrección de seguridad: las
    contraseñas se guardaban en texto plano).

    - Si la contraseña guardada YA es un hash → se compara con
      check_password_hash, como cualquier login normal.
    - Si la contraseña guardada SIGUE en texto plano (cuenta antigua
      todavía no migrada) → se compara tal cual como se hacía antes; si
      coincide, se rehashea y se sobreescribe en la BD en ese mismo
      instante, así el usuario queda migrado sin darse cuenta y sin
      necesidad de resetear nada.

    Devuelve True/False. No lanza excepciones por credenciales
    incorrectas — solo por fallos reales de BD.
    """
    guardada = user.get("password") or ""

    if _es_hash_password(guardada):
        return check_password_hash(guardada, password_recibida)

    # Contraseña heredada en texto plano: comparación legacy
    if guardada == password_recibida:
        nuevo_hash = generate_password_hash(password_recibida)
        try:
            execute("UPDATE usuarios SET password=%s WHERE id=%s", (nuevo_hash, user["id"]))
            get_db().commit()
        except Exception:
            # Si la migración en caliente falla por lo que sea, no bloqueamos
            # el login por eso — se reintentará en el próximo login.
            log.exception("No se pudo migrar a hash la contraseña del usuario id=%s", user["id"])
        return True

    return False


@app.route("/api/login", methods=["POST"])
def login():
    body     = request.get_json(silent=True) or {}
    username = body.get("username", "").strip().lower()
    password = body.get("password", "").strip()

    user = query(
        "SELECT * FROM usuarios WHERE username=%s AND activo=1",
        (username,), one=True
    )
    if not user or not _verifica_y_migra_password(user, password):
        return jsonify({"error": "Usuario o contraseña incorrectos"}), 401

    # ── Verificación por email tras varios días de inactividad ──────────────
    # No afecta al uso diario normal (la sesión ya caduca cada día y exige
    # contraseña de nuevo); esto es una capa extra solo para el caso de
    # cuentas que llevan tiempo sin usarse (vacaciones, bajas, etc.).
    import pytz
    dias_inactivo = None
    if user.get("ultimo_login"):
        dias_inactivo = (_hoy_canarias() - user["ultimo_login"].astimezone(
            pytz.timezone("Atlantic/Canary")).date()).days

    requiere_verificacion = (dias_inactivo is None) or (dias_inactivo >= DIAS_VERIFICACION_EMAIL)

    if requiere_verificacion and user.get("email"):
        import secrets
        codigo = f"{secrets.randbelow(1_000_000):06d}"
        # datetime.now(timezone.utc) en vez de utcnow(): el valor queda
        # explícitamente marcado como UTC (tz-aware) al insertarlo en la
        # columna TIMESTAMPTZ, sin depender de que la sesión de Postgres
        # tenga el timezone en UTC por defecto — ventana de 10 min muy
        # ajustada como para arriesgarse a un desfase de interpretación.
        expira = datetime.now(timezone.utc) + timedelta(minutes=10)
        db = get_db()
        execute("UPDATE login_verification_codes SET usado=1 WHERE usuario_id=%s AND usado=0", (user["id"],))
        execute(
            "INSERT INTO login_verification_codes (usuario_id, codigo, expira_en) VALUES (%s,%s,%s)",
            (user["id"], codigo, expira)
        )
        db.commit()
        log.info("LOGIN — verificación por email requerida para '%s' (código no logueado)", username)

        subject = "Código de verificación – Control de Pedidos"
        mensaje = (
            f"Hola {user['nombre']},\n\n"
            f"Detectamos que hace tiempo que no accedes a Control de Pedidos. "
            f"Por seguridad, confirma que eres tú introduciendo este código:\n\n"
            f"    {codigo}\n\n"
            f"Válido durante 10 minutos.\n"
            f"Si no has sido tú, ignora este mensaje y avisa al administrador.\n\n"
            f"Control de Pedidos · Princess Canarias"
        )
        body_html = _email_html_simple(
            nombre=user["nombre"],
            parrafos=[
                "Detectamos que hace tiempo que no accedes a Control de Pedidos. "
                "Por seguridad, confirma que eres tú introduciendo este código:",
                f'<span style="font-size:22px;font-weight:700;letter-spacing:4px;'
                f'display:inline-block;padding:8px 16px;background:#f5f5f5;'
                f'border-radius:4px;">{codigo}</span>',
                "Válido durante <strong>10 minutos</strong>.<br>"
                "Si no has sido tú, ignora este mensaje y avisa al administrador.",
            ],
        )
        # (2026-08-11) Este email lo envía el navegador vía EmailJS ANTES de
        # que exista sesión (_completar_login() aún no se ha llamado) — sin
        # esta marca, /api/emailjs/registrar-envio lo rechazaría con 401 y
        # el contador de envíos no se enteraría de un email que sí se envía
        # de verdad. Ver _permite_registrar_envio_no_autenticado().
        session["pdte_registrar_envio_email"] = True
        return jsonify({
            "ok": True,
            "requiere_verificacion": True,
            "username": user["username"],
            "email":    user.get("email", ""),
            "nombre":   user.get("nombre", user.get("username", "")),
            "subject":  subject,
            "message":  mensaje,
            "body_html": body_html,
        })

    if requiere_verificacion and not user.get("email"):
        # Sin email registrado no podemos verificar — dejamos entrar
        # directamente para no bloquear al usuario, pero queda logueado.
        log.warning("LOGIN — '%s' requeriría verificación por email pero no tiene email registrado", username)

    return _completar_login(user)


def _completar_login(user):
    """Fija la sesión y actualiza ultimo_login. Compartido entre el login
    normal y la confirmación de código de verificación."""
    hoy = _hoy_canarias()
    session.clear()
    session["user_id"]    = user["id"]
    session["username"]   = user["username"]
    session["nombre"]     = user["nombre"]
    session["rol"]        = user["rol"]
    session["login_date"] = hoy.isoformat()
    hoteles_ids = []
    if user["rol"] == "hotel":
        rows = query("SELECT hotel_id FROM usuario_hoteles WHERE usuario_id=%s", (user["id"],))
        hoteles_ids = [r["hotel_id"] for r in rows]
    session["hoteles_ids"] = hoteles_ids

    db = get_db()
    execute("UPDATE usuarios SET ultimo_login=NOW() WHERE id=%s", (user["id"],))
    db.commit()

    return jsonify({"ok": True, "id": user["id"], "username": user["username"],
                    "nombre": user["nombre"], "rol": user["rol"], "hoteles_ids": hoteles_ids})


def _generar_token_sso_dali(payload: dict, ttl_segundos: int = 100) -> str:
    """
    Token de un solo uso para el acceso automático a la app DALI:
    `<payload-b64url>.<hmac-sha256-hex>`, firmado con el secreto compartido
    DALI_SSO_SECRET (debe ser idéntico en el servicio de Render de DALI —
    ver backend/src/controllers/authController.js de ese repo). No cifra
    el contenido (no lleva nada más sensible que email/nombre/rol), solo
    garantiza que lo emitió este backend, que no ha caducado y que no se
    ha reutilizado (jti, comprobado en el lado de DALI).

    (2026-08-22) Antes eran 60s. DALI vive en el plan gratuito de Render,
    que duerme tras 15 min sin tráfico y tarda ~60s (a veces más) en
    despertar — con 60s de TTL, un acceso a DALI justo tras un rato sin
    uso caducaba el token casi siempre antes de que DALI llegara a
    verificarlo, cayendo al login manual con un aviso de "enlace
    caducado" (ver HISTORIAL_CAMBIOS.md). Subido a 100s para cubrir un
    cold-start normal con margen de sobra, coordinado con el margen de
    aceptación del lado de DALI (SSO_MARGEN_RELOJ_SEGUNDOS en
    authController.js, bajado de 90s a 20s en el mismo cambio — ese 90s
    era un parche temporal mientras este TTL seguía en 60s; con el TTL
    ya arreglado aquí, el margen de DALI vuelve a ser solo margen real
    de reloj/latencia, no un sustituto del TTL). Ventana total efectiva:
    ~120s (100s aquí + 20s en DALI).
    """
    if not DALI_SSO_SECRET:
        raise RuntimeError("DALI_SSO_SECRET no está configurada.")
    cuerpo = dict(payload)
    cuerpo["exp"] = int(datetime.now(timezone.utc).timestamp()) + ttl_segundos
    cuerpo["jti"] = secrets.token_hex(16)
    datos_b64 = base64.urlsafe_b64encode(
        json.dumps(cuerpo, separators=(",", ":")).encode()
    ).decode().rstrip("=")
    firma = hmac.new(DALI_SSO_SECRET.encode(), datos_b64.encode(), hashlib.sha256).hexdigest()
    return f"{datos_b64}.{firma}"


@app.route("/api/dali/sso", methods=["GET"])
@login_required
def dali_sso():
    """
    Menú lateral "Catálogo DALI": genera la URL de acceso automático a la
    app DALI para el usuario de la sesión actual, con su rol ya mapeado
    (ver DALI_ROL_MAP). El frontend abre esa URL en una pestaña nueva
    (ver abrirDali() en templates/index.html); DALI se encarga de
    verificar el token y abrir sesión sin pedir contraseña.
    """
    if not DALI_SSO_SECRET:
        return jsonify({"error": "El acceso a DALI no está configurado todavía (falta DALI_SSO_SECRET)."}), 503

    rol_dali = DALI_ROL_MAP.get(session.get("rol"))
    if not rol_dali:
        return jsonify({"error": "Tu perfil no tiene acceso al catálogo DALI."}), 403

    user = query("SELECT nombre, email FROM usuarios WHERE id=%s", (session["user_id"],), one=True)
    email = (user.get("email") or "").strip().lower() if user else ""
    if not email:
        return jsonify({
            "error": "Tu usuario no tiene un email configurado. Pide a un administrador que te lo "
                     "añada en Usuarios antes de entrar a DALI (DALI identifica a cada usuario por su email)."
        }), 400

    token = _generar_token_sso_dali({"email": email, "nombre": user["nombre"], "rol": rol_dali})
    return jsonify({"url": f"{DALI_FRONTEND_URL}/?dali_token={token}"})


@app.route("/api/login/verificar-codigo", methods=["POST"])
def verificar_codigo_login():
    body     = request.get_json(silent=True) or {}
    username = body.get("username", "").strip().lower()
    codigo   = (body.get("codigo") or "").strip()

    user = query("SELECT * FROM usuarios WHERE username=%s AND activo=1", (username,), one=True)
    if not user:
        return jsonify({"error": "Usuario no válido"}), 401

    row = query(
        """SELECT * FROM login_verification_codes
           WHERE usuario_id=%s AND codigo=%s AND usado=0 AND expira_en > NOW()
           ORDER BY id DESC LIMIT 1""",
        (user["id"], codigo), one=True
    )
    if not row:
        # El código no es válido — antes se devolvía siempre el mismo mensaje
        # genérico, lo que impedía distinguir si de verdad había caducado por
        # tiempo o si había sido invalidado por una petición de login
        # posterior (p.ej. doble clic, o un "Reenviar código" que el propio
        # usuario no recuerda haber pulsado). Se diagnostica el motivo real
        # para poder confirmarlo en logs y en el mensaje mostrado.
        fila_cualquiera = query(
            """SELECT * FROM login_verification_codes
               WHERE usuario_id=%s AND codigo=%s
               ORDER BY id DESC LIMIT 1""",
            (user["id"], codigo), one=True
        )
        if fila_cualquiera is None:
            log.warning("LOGIN-CODIGO — '%s' introdujo un código que no existe para su cuenta", username)
            return jsonify({"error": "Código incorrecto"}), 401
        if fila_cualquiera["usado"]:
            log.warning(
                "LOGIN-CODIGO — '%s' introdujo un código ya superado por uno más reciente "
                "(id=%s, creado_en=%s) — probable doble solicitud de login",
                username, fila_cualquiera["id"], fila_cualquiera.get("creado_en")
            )
            return jsonify({"error": "Este código ya no es válido: se generó uno más reciente. Usa el último código recibido por email."}), 401
        log.warning(
            "LOGIN-CODIGO — '%s' introdujo un código realmente caducado por tiempo (id=%s, expira_en=%s)",
            username, fila_cualquiera["id"], fila_cualquiera.get("expira_en")
        )
        return jsonify({"error": "El código ha caducado (han pasado más de 10 minutos). Pulsa \"Reenviar código\"."}), 401

    db = get_db()
    execute("UPDATE login_verification_codes SET usado=1 WHERE id=%s", (row["id"],))
    db.commit()

    return _completar_login(user)


@app.route("/api/bridge/login", methods=["POST"])
def bridge_login():
    """
    Login específico para el bridge de main_agenda (cuenta de servicio en
    segundo plano, sin ninguna persona delante de la pantalla en el momento
    exacto de conectar).

    Usa las mismas credenciales que /api/login (usuario/contraseña reales
    de la tabla `usuarios` — NO hay debilitamiento de la autenticación en
    sí), pero se salta a propósito el paso de verificación por email tras
    varios días de inactividad: ese paso pide un código de un solo uso que
    nadie va a poder introducir en un proceso desatendido, así que sin este
    endpoint separado el bridge quedaría atrapado en un bucle de login
    fallido cada vez que le tocara esa verificación (algo que iba a pasar
    con toda seguridad, dado que el bridge sondea automáticamente día a
    día sin depender de que nadie escriba la contraseña a mano).

    La caducidad diaria de sesión SÍ se aplica igual que a cualquier otra
    cuenta — el bridge ya vuelve a llamar a este endpoint solo, de forma
    automática, en cuanto recibe un 401 (ver _get() en pedidos_agenda_bridge.py),
    así que no hace falta ninguna excepción para eso.
    """
    body     = request.get_json(silent=True) or {}
    username = body.get("username", "").strip().lower()
    password = body.get("password", "").strip()

    user = query(
        "SELECT * FROM usuarios WHERE username=%s AND activo=1",
        (username,), one=True
    )
    if not user or not _verifica_y_migra_password(user, password):
        return jsonify({"error": "Usuario o contraseña incorrectos"}), 401

    return _completar_login(user)

@app.route("/api/bridge/existe", methods=["GET"])
def bridge_existe_usuario():
    """
    Comprueba si un nombre de usuario (login) ya existe y está activo en
    Control de Pedidos. Usado por main_agenda para decidir, en la ventana
    'Mi Usuario' de Administración, si debe pedir la contraseña existente
    (validándola contra la plataforma) o mostrar el formulario de alta de
    un usuario nuevo.

    Solo confirma existencia por USERNAME — nunca revela ni acepta
    contraseña en este endpoint, para no debilitar la superficie de
    ataque de enumeración de cuentas más de lo estrictamente necesario
    (el username aquí es el usuario de Windows, ya semi-público en la
    empresa, no un dato sensible por sí solo).
    """
    usuario = (request.args.get("usuario_windows") or request.args.get("username") or "").strip().lower()
    if not usuario:
        return jsonify({"error": "Falta usuario_windows"}), 400
    row = query(
        "SELECT id FROM usuarios WHERE username=%s AND activo=1",
        (usuario,), one=True
    )
    return jsonify({"existe": bool(row)})


@app.route("/api/bridge/solicitar-alta", methods=["POST"])
def bridge_solicitar_alta():
    """
    Registra la solicitud de alta de un usuario nuevo desde main_agenda
    (ventana Admin → Mi Usuario, cuando el usuario de Windows todavía no
    tiene cuenta en Control de Pedidos).

    Deliberadamente NO recibe ni almacena aquí la contraseña elegida por
    el usuario: la contraseña la guarda main_agenda solo localmente
    (cifrada, con hash+salt) hasta que un administrador cree la cuenta
    real en Control de Pedidos (Admin → Usuarios) con esas mismas
    credenciales, comunicadas por el usuario directamente al admin. Este
    endpoint únicamente avisa de que hay una solicitud pendiente.
    """
    body            = request.get_json(silent=True) or {}
    usuario_windows = (body.get("usuario_windows") or "").strip().lower()
    nombre          = (body.get("nombre") or "").strip()

    if not usuario_windows:
        return jsonify({"error": "Falta usuario_windows"}), 400
    if not nombre:
        return jsonify({"error": "Falta nombre"}), 400

    ya_existe = query(
        "SELECT id FROM usuarios WHERE username=%s AND activo=1",
        (usuario_windows,), one=True
    )
    if ya_existe:
        return jsonify({"error": "Ese usuario ya existe en Control de Pedidos"}), 409

    _notify_solicitud_telegram(
        f"\U0001F195 *[Alta de usuario — OrganizadorPrincess]*\n\n"
        f"\U0001F4BB Usuario Windows: `{usuario_windows}`\n"
        f"\U0001F464 Nombre a mostrar: *{nombre}*\n\n"
        f"Este usuario ha configurado su acceso al panel Admin de "
        f"OrganizadorPrincess pero todavía no tiene cuenta en Control de "
        f"Pedidos.\nCréale la cuenta en *Admin \u2192 Usuarios* con el mismo "
        f"nombre de usuario (`{usuario_windows}`) y la contraseña que él "
        f"te indique; en cuanto exista, su acceso quedará validado "
        f"automáticamente."
    )
    log.info("Bridge: solicitud de alta registrada para '%s' (%s)", usuario_windows, nombre)
    return jsonify({"ok": True})


@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

# ── Restablecimiento de contraseña ────────────────────────────────────────────

@app.route("/api/password-reset/solicitar", methods=["POST"])
def solicitar_reset_password():
    """El usuario introduce su username o email y recibe un enlace de reset."""
    body    = request.get_json(silent=True) or {}
    usuario = (body.get("usuario") or "").strip().lower()
    if not usuario:
        return jsonify({"error": "Indica tu usuario o email"}), 400

    user = query(
        "SELECT * FROM usuarios WHERE (username=%s OR email=%s) AND activo=1",
        (usuario, usuario), one=True
    )
    # Siempre respuesta OK para no revelar si el usuario existe
    if not user or not user.get("email"):
        return jsonify({"ok": True, "msg": "Si el usuario existe, recibirás un correo."})

    # Generar token seguro con 2 h de validez
    token    = secrets.token_urlsafe(32)
    expira   = datetime.utcnow() + timedelta(hours=2)
    db       = get_db()
    # Invalidar tokens anteriores del mismo usuario
    execute("UPDATE password_reset_tokens SET usado=1 WHERE usuario_id=%s AND usado=0", (user["id"],))
    execute(
        "INSERT INTO password_reset_tokens (usuario_id, token, expira_en) VALUES (%s,%s,%s)",
        (user["id"], token, expira)
    )
    db.commit()

    # Construir enlace (se puede configurar la URL base con env var APP_URL)
    app_url  = os.environ.get("APP_URL", request.host_url.rstrip("/"))
    link     = f"{app_url}/?reset_token={token}"

    subject  = "Restablecimiento de contraseña – Control de Pedidos"
    # (2026-07-31) Antes: párrafos sueltos sin cabecera ni logo — el único
    # email que quedó fuera del rollout de logo junto con el código de
    # verificación de login. Ahora usa el mismo _email_html_simple() que ya
    # lleva la cabecera de marca (_email_header_html) del resto de la app.
    body_html = _email_html_simple(
        nombre=user["nombre"],
        parrafos=["Hemos recibido una solicitud para restablecer tu contraseña."],
        boton={"texto": "Restablecer contraseña", "url": link},
        pie_extra="Este enlace es válido durante <strong>2 horas</strong>. Si no lo solicitaste, ignora este mensaje.",
    )
    # Siempre loguear el enlace en el servidor
    log.info("PASSWORD RESET solicitado por '%s' (id=%s) — enlace: %s",
             user["username"], user["id"], link)

    # El envío real lo hace el frontend vía EmailJS
    log.info("PASSWORD RESET — datos pendientes de envío vía EmailJS a '%s' (%s)", user["username"], user.get("email"))
    # (2026-08-11) Igual que en el código de verificación de login: este
    # email lo envía un navegador SIN sesión iniciada (nadie ha hecho
    # login todavía — está pidiendo restablecer su contraseña). Marca de
    # un solo uso para que /api/emailjs/registrar-envio no lo rechace con
    # 401 y el contador cuente también estos envíos reales.
    session["pdte_registrar_envio_email"] = True
    return jsonify({
        "ok":        True,
        "sin_email": True,
        "link":      link,
        "email":     user.get("email", ""),
        "nombre":    user.get("nombre", user.get("username", "")),
        "subject":   subject,
        "body_html": body_html,
        "msg":       "Email pendiente de envío vía EmailJS.",
    })


@app.route("/api/password-reset/validar/<token>", methods=["GET"])
def validar_reset_token(token):
    """Comprueba si el token es válido y no ha caducado."""
    row = query(
        """SELECT prt.*, u.nombre FROM password_reset_tokens prt
           JOIN usuarios u ON u.id = prt.usuario_id
           WHERE prt.token=%s AND prt.usado=0 AND prt.expira_en > NOW()""",
        (token,), one=True
    )
    if not row:
        return jsonify({"valido": False, "error": "El enlace no es válido o ha caducado"}), 400
    return jsonify({"valido": True, "nombre": row["nombre"]})


@app.route("/api/password-reset/cambiar", methods=["POST"])
def cambiar_password_con_token():
    """El usuario envía el token + nueva contraseña elegida por él."""
    body     = request.get_json(silent=True) or {}
    token    = (body.get("token") or "").strip()
    nueva    = (body.get("nueva_password") or "").strip()
    if not token or not nueva:
        return jsonify({"error": "Datos incompletos"}), 400
    if len(nueva) < 6:
        return jsonify({"error": "La contraseña debe tener al menos 6 caracteres"}), 400

    row = query(
        "SELECT * FROM password_reset_tokens WHERE token=%s AND usado=0 AND expira_en > NOW()",
        (token,), one=True
    )
    if not row:
        return jsonify({"error": "El enlace no es válido o ha caducado"}), 400

    db = get_db()
    execute("UPDATE usuarios SET password=%s WHERE id=%s", (generate_password_hash(nueva), row["usuario_id"]))
    execute("UPDATE password_reset_tokens SET usado=1 WHERE token=%s", (token,))
    db.commit()
    return jsonify({"ok": True, "msg": "Contraseña actualizada correctamente"})

# ══════════════════════════════════════════════════════════════════════════════
#  SOLICITUD DE ACCESO EN 2 FASES (v10.5)
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/api/solicitar-usuario/detectar-windows", methods=["GET"])
def detectar_usuario_windows():
    """
    Intenta detectar el usuario Windows del cliente.
    En entornos web públicos siempre devuelve None;
    útil solo en intranet con autenticación Windows integrada (NTLM/Kerberos).
    """
    import os as _os
    usuario = _os.environ.get("USERNAME") or _os.environ.get("USER") or ""
    if usuario and usuario not in ("root", "www-data", "nobody", "daemon"):
        return jsonify({"usuario_windows": usuario})
    return jsonify({"usuario_windows": None})


# ─── FASE 1: recibir datos básicos, guardar en BD, notificar admin ────────────

@app.route("/api/solicitar-usuario", methods=["POST"])
def solicitar_usuario_fase1():
    """
    FASE 1 — El usuario rellena nombre, apellidos, email y hotel(es).
    Se guarda la solicitud, se genera el token de verificación y se pasa
    directo a 'fase2_pendiente'.

    v12.11.0: el email de Fase 2 ya NO se manda desde el navegador de quien
    solicita el acceso (poco fiable — suele ser un equipo nuevo/no
    homologado, con antivirus o red corporativa más restrictiva) sino que
    se ENCOLA en emails_sistema_pendientes, igual que el resto de avisos
    de sistema: lo despacha el navegador del primer admin que abra la
    aplicación (mismo mecanismo — y misma fiabilidad — que cuando este
    paso lo hacía un admin a mano). El aviso de Telegram al admin dice
    explícitamente que tiene que abrir la app para que se complete.
    """
    import re as _re

    body      = request.get_json(silent=True) or {}
    nombre    = (body.get("nombre") or "").strip()
    apellidos = (body.get("apellidos") or "").strip()
    email_sol = (body.get("email") or "").strip()
    movil_sol = (body.get("movil") or "").strip()
    hoteles   = (body.get("hoteles") or body.get("hotel") or "").strip()

    if not nombre:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if not apellidos:
        return jsonify({"error": "Los apellidos son obligatorios"}), 400
    if not email_sol:
        return jsonify({"error": "El correo electrónico es obligatorio"}), 400
    if not _re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_sol):
        return jsonify({"error": "El formato del correo electrónico no es válido"}), 400
    if not movil_sol:
        return jsonify({"error": "El teléfono móvil de empresa es obligatorio"}), 400
    if not hoteles:
        return jsonify({"error": "Debes seleccionar al menos un hotel"}), 400

    nombre_completo = f"{nombre} {apellidos}"
    ip_cliente = request.remote_addr or ""

    # El token de verificación se genera ya en Fase 1 para que, en cuanto
    # se despache el email (por un admin), el enlace funcione directamente.
    import secrets as _sec
    token     = _sec.token_urlsafe(32)
    expira_en = datetime.utcnow() + timedelta(hours=72)

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            INSERT INTO solicitudes_acceso
                (nombre, apellidos, email, hoteles, movil, estado, ip_solicitante,
                 token, token_expira)
            VALUES (%s, %s, %s, %s, %s, 'fase2_pendiente', %s, %s, %s)
            RETURNING id
        """, (nombre, apellidos, email_sol, hoteles, movil_sol, ip_cliente, token, expira_en))
        sol_id = cur.fetchone()["id"]
    db.commit()

    sol = query("SELECT * FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)
    email_fase2 = _construir_email_fase2(sol)

    app_url   = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    url_admin = f"{app_url}/admin/solicitudes#{sol_id}" if app_url else ""
    asunto    = f"[FASE 1] Nueva solicitud de acceso — {nombre_completo}"

    body_html = f"""
    <div style="font-family:sans-serif;max-width:620px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#0f2044;">
        <tr>
          <td style="padding:24px 0 24px 28px;vertical-align:middle;" valign="middle">
            <h2 style="margin:0;color:#c9a84c;font-size:18px;">
              📋 Nueva solicitud de acceso — Fase 1
            </h2>
            <p style="margin:6px 0 0;color:#b9c3dc;font-size:13px;">
              Control de Pedidos · Princess Canarias
            </p>
          </td>
          <td style="padding:14px 28px 14px 16px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
            <img src="{app_url}/static/logo-sidebar-email-64.png" alt="Princess Hotels & Resorts"
                 width="69" height="64"
                 style="height:64px;width:69px;display:block;margin-left:auto;">
          </td>
        </tr>
      </table>
      <div style="padding:24px 28px;">
        <p style="margin:0 0 16px;font-size:14px;color:#333;">
          Se ha recibido una nueva solicitud. El email con el enlace de
          verificación (Fase 2) está <strong>en cola</strong> — se enviará
          al usuario automáticamente en cuanto abras la aplicación (o ya se
          habrá enviado, si otro admin la tiene abierta ahora mismo).
        </p>
        <table border="0" cellpadding="0" cellspacing="0"
               style="width:100%;font-size:14px;border-collapse:collapse;">
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 0;color:#888;width:160px;">Nombre completo</td>
            <td style="padding:10px 0;font-weight:600;">{nombre_completo}</td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 0;color:#888;">Correo electrónico</td>
            <td style="padding:10px 0;">
              <a href="mailto:{email_sol}" style="color:#0f2044;">{email_sol}</a>
            </td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 0;color:#888;">Móvil empresa</td>
            <td style="padding:10px 0;">{movil_sol}</td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 0;color:#888;">Hotel(es)</td>
            <td style="padding:10px 0;">{hoteles}</td>
          </tr>
          <tr>
            <td style="padding:10px 0;color:#888;">ID solicitud</td>
            <td style="padding:10px 0;font-family:monospace;">#{sol_id}</td>
          </tr>
        </table>
        {f'<div style="margin-top:24px;text-align:center;"><a href="{url_admin}" style="display:inline-block;padding:12px 28px;background:#c9a84c;color:#0f2044;border-radius:7px;text-decoration:none;font-weight:700;font-size:14px;">➜ Ver solicitud</a></div>' if url_admin else ''}
      </div>
      <div style="padding:14px 28px;background:#f0f0f0;font-size:11px;color:#aaa;">
        Mensaje automático · Control Pedidos Princess Canarias
      </div>
    </div>
    """

    body_text = (
        f"NUEVA SOLICITUD DE ACCESO (FASE 1)\n"
        f"{'='*44}\n"
        f"Nombre        : {nombre_completo}\n"
        f"Email         : {email_sol}\n"
        f"Móvil empresa : {movil_sol}\n"
        f"Hotel(es)     : {hoteles}\n"
        f"ID solicitud  : #{sol_id}\n"
        f"{'='*44}\n"
        f"El email de verificación (Fase 2) está en cola — se enviará al "
        f"usuario en cuanto abras la aplicación."
    )

    destinatarios_admin = _get_solo_admin_emails()

    # Telegram SIEMPRE, antes de devolver la respuesta al frontend — y
    # explícito sobre que hace falta abrir la app para que Fase 2 se
    # despache (el email ya no sale del navegador del solicitante).
    _notify_solicitud_telegram(
        f"\U0001F514 *[Nueva solicitud de acceso]*\n\n"
        f"\U0001F464 *{nombre} {apellidos}*\n"
        f"\U0001F4E7 {email_sol}\n"
        f"\U0001F3E8 {hoteles}\n"
        f"\U0001F4CB Solicitud `#{sol_id}`\n\n"
        f"\u26A0\uFE0F *Abre la aplicaci\u00f3n* para que el email de verificaci\u00f3n "
        f"(Fase 2) se env\u00ede autom\u00e1ticamente al usuario."
        + (f"\n\U0001F517 {url_admin}" if url_admin else "")
    )

    # Se encolan ambos emails — el de Fase 2 al usuario y el aviso a los
    # admins — para que los despache el navegador de un admin ya logado,
    # en vez de depender del navegador (sin homologar) de quien solicita
    # el acceso.
    _encolar_email_sistema(
        "solicitud_acceso_fase2", email_fase2["destinatarios"], email_fase2["asunto"],
        email_fase2["body_html"], email_fase2["body_text"], solicitud_acceso_id=sol_id,
    )
    if destinatarios_admin:
        _encolar_email_sistema("solicitud_acceso", destinatarios_admin, asunto, body_html, body_text)
    else:
        log.warning("[SOL_FASE1] Sin emails admin configurados. Sol #%s", sol_id)

    return jsonify({"ok": True, "sol_id": sol_id, "encolado": True})


@app.route("/api/solicitar-usuario/directo", methods=["POST"])
def solicitar_usuario_directo():
    """
    Alta en un solo paso (v12.20.2, desde el Organizador de escritorio).

    Fusiona fase 1 + fase 2 en una sola llamada: el escritorio YA conoce
    el usuario de Windows (no hace falta el rodeo del .bat/email para
    detectarlo), así que no tiene sentido generar un token de un solo
    uso ni esperar a que el usuario abra un email para completar nada —
    esa comprobación por email solo aportaba valor cuando el origen era
    un navegador sin autenticar. Aquí el origen es la propia app interna
    instalada en el equipo del solicitante.

    La solicitud se crea directamente en estado 'completada' —
    exactamente el mismo estado en el que queda una solicitud tras pasar
    por las dos fases web — así que cae en la misma cola de aprobación
    del panel admin (GET /api/admin/solicitudes-acceso) sin ningún
    cambio ahí. El admin sigue sin intervenir hasta que la vea y la
    apruebe con /api/admin/solicitudes-acceso/<id>/aprobar, que es
    donde se genera y envía la contraseña — no aquí.
    """
    import re as _re

    body            = request.get_json(silent=True) or {}
    nombre          = (body.get("nombre") or "").strip()
    apellidos       = (body.get("apellidos") or "").strip()
    email_sol       = (body.get("email") or "").strip()
    movil_sol       = (body.get("movil") or "").strip()
    hoteles         = (body.get("hoteles") or body.get("hotel") or "").strip()
    usuario_windows = (body.get("usuario_windows") or "").strip().upper()

    if not nombre:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if not apellidos:
        return jsonify({"error": "Los apellidos son obligatorios"}), 400
    if not email_sol:
        return jsonify({"error": "El correo electrónico es obligatorio"}), 400
    if not _re.match(r'^[^\s@]+@[^\s@]+\.[^\s@]+$', email_sol):
        return jsonify({"error": "El formato del correo electrónico no es válido"}), 400
    if not movil_sol:
        return jsonify({"error": "El teléfono móvil de empresa es obligatorio"}), 400
    if not hoteles:
        return jsonify({"error": "Debes seleccionar al menos un hotel"}), 400
    if not usuario_windows:
        return jsonify({"error": "Falta el usuario de Windows"}), 400

    # Mismo check que la fase 2 web — evita duplicar una cuenta que ya
    # existe y está activa (o avisar si existe pero desactivada).
    usuario_existente = query(
        "SELECT id, activo FROM usuarios WHERE LOWER(username)=LOWER(%s)",
        (usuario_windows,), one=True
    )
    if usuario_existente:
        if usuario_existente["activo"]:
            return jsonify({
                "error": f"El usuario Windows '{usuario_windows}' ya tiene una cuenta activa en el sistema.",
                "ya_existe": True
            }), 409
        else:
            return jsonify({
                "error": f"El usuario Windows '{usuario_windows}' existe en el sistema pero está desactivado. "
                         f"Contacta con el administrador para reactivar tu cuenta.",
                "ya_existe": True,
                "desactivado": True
            }), 409

    nombre_completo = f"{nombre} {apellidos}"
    ip_cliente = request.remote_addr or ""

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            INSERT INTO solicitudes_acceso
                (nombre, apellidos, email, hoteles, movil, usuario_windows,
                 estado, ip_solicitante, completado_en)
            VALUES (%s, %s, %s, %s, %s, %s, 'completada', %s, NOW())
            RETURNING id
        """, (nombre, apellidos, email_sol, hoteles, movil_sol, usuario_windows, ip_cliente))
        sol_id = cur.fetchone()["id"]
    db.commit()

    app_url   = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    url_admin = f"{app_url}/admin/solicitudes#{sol_id}" if app_url else ""
    asunto    = f"[Alta desde Organizador] {nombre_completo} / {usuario_windows}"

    body_html = f"""
    <div style="font-family:sans-serif;max-width:620px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#065f46;">
        <tr>
          <td style="padding:24px 0 24px 28px;vertical-align:middle;" valign="middle">
            <h2 style="margin:0;color:#6ee7b7;font-size:18px;">
              ✅ Solicitud de acceso — lista para aprobar
            </h2>
            <p style="margin:6px 0 0;color:#b9c3dc;font-size:13px;">
              Recibida desde el Organizador de Escritorio (alta en un solo paso)
            </p>
          </td>
          <td style="padding:14px 28px 14px 16px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
            <img src="{app_url}/static/logo-sidebar-email-64.png" alt="Princess Hotels & Resorts"
                 width="69" height="64"
                 style="height:64px;width:69px;display:block;margin-left:auto;">
          </td>
        </tr>
      </table>
      <div style="padding:24px 28px;">
        <table border="0" cellpadding="0" cellspacing="0"
               style="width:100%;font-size:14px;border-collapse:collapse;">
          <tr style="background:#f0fdf4;border-bottom:1px solid #d1fae5;">
            <td style="padding:12px 14px;color:#065f46;font-weight:700;width:170px;">Usuario Windows</td>
            <td style="padding:12px 14px;font-family:monospace;font-size:16px;font-weight:700;color:#0f2044;">
              {usuario_windows}
            </td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Nombre completo</td>
            <td style="padding:10px 14px;font-weight:600;">{nombre_completo}</td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Correo electrónico</td>
            <td style="padding:10px 14px;"><a href="mailto:{email_sol}" style="color:#0f2044;">{email_sol}</a></td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Móvil empresa</td>
            <td style="padding:10px 14px;">{movil_sol}</td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Hotel(es)</td>
            <td style="padding:10px 14px;">{hoteles}</td>
          </tr>
          <tr>
            <td style="padding:10px 14px;color:#888;">ID solicitud</td>
            <td style="padding:10px 14px;font-family:monospace;">#{sol_id}</td>
          </tr>
        </table>
        {f'<div style="margin-top:24px;text-align:center;"><a href="{url_admin}" style="display:inline-block;padding:12px 28px;background:#c9a84c;color:#0f2044;border-radius:7px;text-decoration:none;font-weight:700;font-size:14px;">➜ Ver y aprobar</a></div>' if url_admin else ''}
      </div>
      <div style="padding:14px 28px;background:#f0f0f0;font-size:11px;color:#aaa;">
        Mensaje automático · Control Pedidos Princess Canarias
      </div>
    </div>
    """

    body_text = (
        f"SOLICITUD DE ACCESO — LISTA PARA APROBAR (desde el Organizador)\n"
        f"{'='*44}\n"
        f"Usuario Windows : {usuario_windows}\n"
        f"Nombre          : {nombre_completo}\n"
        f"Email           : {email_sol}\n"
        f"Móvil empresa   : {movil_sol}\n"
        f"Hotel(es)       : {hoteles}\n"
        f"ID solicitud    : #{sol_id}\n"
        f"{'='*44}\n"
        f"Apruébala desde el panel para crear la cuenta y enviar la contraseña."
    )

    # Telegram siempre, igual que en las dos fases web.
    _notify_solicitud_telegram(
        f"\U00002705 *[Alta desde Organizador] Lista para aprobar*\n\n"
        f"\U0001F464 *{nombre_completo}*\n"
        f"\U0001F5A5 Usuario Windows: `{usuario_windows}`\n"
        f"\U0001F4E7 {email_sol}\n"
        f"\U0001F3E8 {hoteles}\n"
        f"\U0001F4CB Solicitud `#{sol_id}` — lista para aprobar."
        + (f"\n\U0001F517 {url_admin}" if url_admin else "")
    )

    # Igual que fase 1: se encola (no EmailJS en vivo, porque aquí no hay
    # navegador del solicitante) — lo despacha el primer admin que abra
    # la app, mismo mecanismo fiable que ya usa el resto de avisos.
    destinatarios_admin = _get_solo_admin_emails()
    if destinatarios_admin:
        _encolar_email_sistema("solicitud_acceso", destinatarios_admin, asunto, body_html, body_text,
                                solicitud_acceso_id=sol_id)
    else:
        log.warning("[SOL_DIRECTO] Sin emails admin configurados. Sol #%s", sol_id)

    return jsonify({"ok": True, "sol_id": sol_id})


# ─── ADMIN: listar solicitudes de acceso ──────────────────────────────────────

@app.route("/api/admin/solicitudes-acceso", methods=["GET"])
def admin_listar_solicitudes():
    """Devuelve todas las solicitudes de acceso (solo admins)."""
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permisos"}), 403
    rows = query("""
        SELECT s.id, s.nombre, s.apellidos, s.email, s.hoteles, s.usuario_windows,
               s.estado, s.creado_en, s.completado_en,
               s.fase2_email_estado, s.fase2_email_detalle, s.fase2_email_en,
               e.enviado AS cola_enviado, e.enviado_en AS cola_enviado_en,
               e.creado_en AS cola_creado_en
        FROM solicitudes_acceso s
        LEFT JOIN emails_sistema_pendientes e
               ON e.solicitud_acceso_id = s.id AND e.evento_codigo = 'solicitud_acceso_fase2'
        ORDER BY s.creado_en DESC
        LIMIT 200
    """)
    return jsonify(rows)


# ─── ADMIN: generar y descargar el .bat para Fase 2 ──────────────────────────

@app.route("/api/admin/solicitudes-acceso/<int:sol_id>/generar-bat", methods=["POST", "GET"])
def admin_generar_bat(sol_id):
    """
    Genera un token único, actualiza estado a 'fase2_pendiente' y devuelve
    un archivo .bat. Al ejecutarlo, Windows resuelve %USERNAME% y abre el
    navegador con token + usuario ya detectado automáticamente.
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permisos"}), 403

    sol = query("SELECT * FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)
    if not sol:
        return jsonify({"error": "Solicitud no encontrada"}), 404
    if sol["estado"] not in ("fase1_pendiente", "fase2_pendiente"):
        return jsonify({"error": f"La solicitud ya está en estado '{sol['estado']}' — no se puede regenerar el .bat."}), 409

    # Reutiliza el token vigente si ya existe (p.ej. el que se envió
    # automáticamente en Fase 1) para no invalidar el email ya enviado al
    # usuario. Solo se genera uno nuevo si no hay token o si ha caducado.
    token_vigente = sol.get("token") and sol.get("token_expira") and sol["token_expira"] > datetime.utcnow()
    if token_vigente:
        token, expira_en = sol["token"], sol["token_expira"]
    else:
        import secrets as _sec
        token     = _sec.token_urlsafe(32)
        expira_en = datetime.utcnow() + timedelta(hours=72)
        db = get_db()
        with db.cursor() as cur:
            cur.execute("""
                UPDATE solicitudes_acceso
                SET token=%(token)s, token_expira=%(expira)s, estado='fase2_pendiente'
                WHERE id=%(id)s
            """, {"token": token, "expira": expira_en, "id": sol_id})
        db.commit()

    app_url  = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    nombre_c = f"{sol['nombre']} {sol['apellidos']}"

    # %USERNAME% la resuelve Windows al ejecutar el .bat — clave del truco
    bat_content = (
        f"@echo off\r\n"
        f":: Control de Pedidos Princess - Verificacion de acceso\r\n"
        f":: Solicitud de: {nombre_c}\r\n"
        f":: Archivo de un solo uso - expira en 72 horas\r\n"
        f"::\r\n"
        f":: Instrucciones: haz doble clic en este archivo.\r\n"
        f":: Se abrira el navegador con tu usuario Windows detectado automaticamente.\r\n"
        f"@echo Abriendo verificacion de acceso, por favor espera...\r\n"
        f"set TOKEN={token}\r\n"
        f"set URL={app_url}/?token=%TOKEN%^&wu=%USERNAME%\r\n"
        f"start \"\" \"%URL%\"\r\n"
        f"exit\r\n"
    )

    from flask import Response
    nombre_archivo = (
        f"verificar_acceso_"
        f"{sol['nombre'].lower().replace(' ','_')}_"
        f"{sol['apellidos'].split()[0].lower()}.bat"
    )
    return Response(
        bat_content,
        mimetype="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{nombre_archivo}"'}
    )


# ─── Construcción del email de Fase 2 (verificación) ──────────────────────────
# Reutilizado tanto por el envío automático (Fase 1) como por el reenvío
# manual desde el panel de admin. Requiere que `sol` ya tenga token asignado.

def _construir_email_fase2(sol: dict) -> dict:
    app_url   = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    url_token = f"{app_url}/?token={sol['token']}"
    nombre_c  = f"{sol['nombre']} {sol['apellidos']}"
    asunto    = "Verificación de acceso — Control de Pedidos Princess"

    # v12.11.0: se ha quitado toda mención a un archivo .bat adjunto — nunca
    # se adjuntaba de verdad (EmailJS, tal y como se usa aquí, no manda
    # adjuntos) y el texto probablemente disparaba filtros anti-phishing
    # corporativos (mención a "doble clic en un .bat" + enlace de
    # verificación urgente es una combinación clásica que Microsoft 365 /
    # Mimecast suelen poner en cuarentena en silencio). Ahora es solo un
    # enlace, igual que el resto de emails transaccionales de la app.
    body_html = f"""
    <div style="font-family:sans-serif;max-width:620px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#0f2044;">
        <tr>
          <td style="padding:24px 0 24px 28px;vertical-align:middle;" valign="middle">
            <h2 style="margin:0;color:#c9a84c;font-size:18px;">Verifica tu acceso al sistema</h2>
            <p style="margin:6px 0 0;color:#b9c3dc;font-size:13px;">
              Control de Pedidos · Princess Canarias
            </p>
          </td>
          <td style="padding:14px 28px 14px 16px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
            <img src="{app_url}/static/logo-sidebar-email-64.png" alt="Princess Hotels & Resorts"
                 width="69" height="64"
                 style="height:64px;width:69px;display:block;margin-left:auto;">
          </td>
        </tr>
      </table>
      <div style="padding:28px;">
        <p style="margin:0 0 12px;font-size:15px;color:#333;">
          Hola, <strong>{nombre_c}</strong>
        </p>
        <p style="margin:0 0 20px;font-size:14px;color:#555;line-height:1.6;">
          Hemos recibido tu solicitud de acceso a Control de Pedidos. Para
          continuar, confirma tu usuario de Windows en el siguiente enlace:
        </p>
        <div style="text-align:center;margin-bottom:20px;">
          <a href="{url_token}"
             style="display:inline-block;padding:12px 28px;background:#1a3a6b;
                    color:#fff;border-radius:7px;text-decoration:none;
                    font-weight:600;font-size:14px;">Continuar verificación →</a>
        </div>
        <p style="margin:0;font-size:12px;color:#aaa;line-height:1.5;">
          Enlace personal e intransferible. Caduca en <strong>72 horas</strong>.
          Si tienes problemas contacta con el departamento de informática.
        </p>
      </div>
      <div style="padding:14px 28px;background:#f0f0f0;font-size:11px;color:#aaa;">
        Mensaje automático · Control Pedidos Princess Canarias
      </div>
    </div>
    """

    body_text = (
        f"Hola {nombre_c},\n\n"
        f"Hemos recibido tu solicitud de acceso a Control de Pedidos.\n"
        f"Para continuar, abre este enlace y confirma tu usuario de Windows:\n\n"
        f"{url_token}\n\n"
        f"El enlace caduca en 72 horas.\n\nControl Pedidos Princess Canarias"
    )

    return {
        "destinatarios": [sol["email"]],
        "asunto":        asunto,
        "body_html":     body_html,
        "body_text":     body_text,
        "url_token":     url_token,
    }


# ─── ADMIN: reenviar Fase 2 por email al usuario (fallback manual) ────────────
# Desde v12.4 el email de Fase 2 se envía automáticamente al terminar la
# Fase 1 (ver solicitar_usuario_fase1). Este endpoint queda como reenvío
# manual por si el envío automático falla o el enlace caducó.

@app.route("/api/admin/solicitudes-acceso/<int:sol_id>/enviar-fase2", methods=["POST"])
def admin_enviar_fase2(sol_id):
    """
    Regenera el token (refrescando la caducidad de 72h) y devuelve el
    email de Fase 2 para que el frontend lo reenvíe al usuario vía EmailJS.
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permisos"}), 403

    sol = query("SELECT * FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)
    if not sol:
        return jsonify({"error": "Solicitud no encontrada"}), 404
    if sol["estado"] not in ("fase1_pendiente", "fase2_pendiente"):
        return jsonify({"error": f"La solicitud ya está en estado '{sol['estado']}' — no se puede reenviar."}), 409

    # Siempre se regenera el token al reenviar manualmente, para refrescar
    # la caducidad de 72h (por ejemplo si el enlace original había expirado).
    import secrets as _sec
    token     = _sec.token_urlsafe(32)
    expira_en = datetime.utcnow() + timedelta(hours=72)
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            UPDATE solicitudes_acceso
            SET token=%(t)s, token_expira=%(e)s, estado='fase2_pendiente'
            WHERE id=%(id)s
        """, {"t": token, "e": expira_en, "id": sol_id})
    db.commit()
    sol = query("SELECT * FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)

    email_fase2 = _construir_email_fase2(sol)
    log.info("[SOL_FASE2] Reenvío manual pendiente de EmailJS a %s", sol["email"])

    return jsonify({
        "ok":            True,
        "sin_email":     True,
        "destinatarios": email_fase2["destinatarios"],
        "asunto":        email_fase2["asunto"],
        "body_html":     email_fase2["body_html"],
        "body_text":     email_fase2["body_text"],
        "url_token":     email_fase2["url_token"],
    })


# ─── Registro del resultado real del envío EmailJS (v12.11.0) ────────────────
#
# El envío del email de Fase 2 se hace 100% desde el navegador (EmailJS),
# tanto en el alta automática como en el reenvío manual del admin. Hasta
# ahora, si EmailJS aceptaba la llamada pero el email se perdía después
# (filtro anti-phishing corporativo, cuota agotada, etc.), nadie se
# enteraba: el panel admin no tenía forma de distinguir "se envió de
# verdad" de "se perdió en algún sitio". Este endpoint deja constancia de
# lo que el navegador vio realmente (éxito o error de emailjs.send), para
# verlo en el panel en vez de asumirlo a ciegas.
#
# Sin @login_required a propósito: el alta de Fase 1 la hace un usuario
# que todavía no tiene sesión. No expone ni acepta datos sensibles, solo
# un estado de envío referido a una solicitud que ya existe.

@app.route("/api/solicitudes-acceso/<int:sol_id>/registrar-envio-fase2", methods=["POST"])
def registrar_envio_fase2(sol_id):
    body    = request.get_json(silent=True) or {}
    ok      = bool(body.get("ok"))
    detalle = (body.get("detalle") or "")[:300]

    existe = query("SELECT id FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)
    if not existe:
        return jsonify({"error": "No encontrada"}), 404

    execute("""
        UPDATE solicitudes_acceso
        SET fase2_email_estado=%s, fase2_email_detalle=%s, fase2_email_en=NOW()
        WHERE id=%s
    """, ("enviado" if ok else "error", detalle, sol_id))
    get_db().commit()
    return jsonify({"ok": True})


# ─── FASE 2: el usuario llega con token + wu, completa la solicitud ───────────

@app.route("/api/solicitar-usuario/completar-fase2", methods=["POST"])
def solicitar_usuario_fase2():
    """
    FASE 2 — El usuario ejecutó el .bat o abrió el enlace.
    Valida token + usuario_windows, marca como completada y notifica admins.
    """
    import re as _re

    body            = request.get_json(silent=True) or {}
    token           = (body.get("token") or "").strip()
    usuario_windows = (body.get("usuario_windows") or "").strip().upper()

    if not token:
        return jsonify({"error": "Token no proporcionado"}), 400
    if not usuario_windows:
        return jsonify({"error": "El usuario Windows es obligatorio"}), 400

    sol = query("SELECT * FROM solicitudes_acceso WHERE token=%s", (token,), one=True)
    if not sol:
        return jsonify({"error": "Enlace no válido o ya utilizado"}), 404

    if sol["token_expira"]:
        expira = sol["token_expira"]
        if hasattr(expira, "tzinfo") and expira.tzinfo:
            from datetime import timezone
            now = datetime.now(timezone.utc)
        else:
            now = datetime.utcnow()
        if now > expira:
            return jsonify({
                "error": "Este enlace ha caducado (72 horas). "
                         "Contacta con el administrador para generar uno nuevo."
            }), 410

    if sol["estado"] == "completada":
        return jsonify({"error": "Esta solicitud ya fue completada anteriormente."}), 409

    usuario_existente = query(
        "SELECT id, activo FROM usuarios WHERE LOWER(username)=LOWER(%s)",
        (usuario_windows,), one=True
    )
    if usuario_existente:
        if usuario_existente["activo"]:
            return jsonify({
                "error": f"El usuario Windows '{usuario_windows}' ya tiene una cuenta activa en el sistema. "
                         f"Puedes iniciar sesión directamente o recuperar tu contraseña si la olvidaste.",
                "ya_existe": True
            }), 409
        else:
            return jsonify({
                "error": f"El usuario Windows '{usuario_windows}' existe en el sistema pero está desactivado. "
                         f"Contacta con el administrador para reactivar tu cuenta.",
                "ya_existe": True,
                "desactivado": True
            }), 409

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            UPDATE solicitudes_acceso
            SET usuario_windows=%(uw)s, estado='completada',
                completado_en=NOW(), token=NULL
            WHERE id=%(id)s
        """, {"uw": usuario_windows, "id": sol["id"]})
    db.commit()

    nombre_c = f"{sol['nombre']} {sol['apellidos']}"
    asunto   = f"[FASE 2 COMPLETADA] Alta usuario — {nombre_c} / {usuario_windows}"
    app_url  = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")

    body_html = f"""
    <div style="font-family:sans-serif;max-width:620px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#065f46;">
        <tr>
          <td style="padding:24px 0 24px 28px;vertical-align:middle;" valign="middle">
            <h2 style="margin:0;color:#6ee7b7;font-size:18px;">
              ✅ Solicitud completa — Crear cuenta de usuario
            </h2>
            <p style="margin:6px 0 0;color:#b9c3dc;font-size:13px;">
              Control de Pedidos · Princess Canarias
            </p>
          </td>
          <td style="padding:14px 28px 14px 16px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
            <img src="{app_url}/static/logo-sidebar-email-64.png" alt="Princess Hotels & Resorts"
                 width="69" height="64"
                 style="height:64px;width:69px;display:block;margin-left:auto;">
          </td>
        </tr>
      </table>
      <div style="padding:24px 28px;">
        <p style="margin:0 0 16px;font-size:14px;color:#333;">
          El usuario ha completado la verificación.
          <strong>Ya podéis crear la cuenta</strong> con los siguientes datos:
        </p>
        <table border="0" cellpadding="0" cellspacing="0"
               style="width:100%;font-size:14px;border-collapse:collapse;">
          <tr style="background:#f0fdf4;border-bottom:1px solid #d1fae5;">
            <td style="padding:12px 14px;color:#065f46;font-weight:700;width:170px;">Usuario Windows</td>
            <td style="padding:12px 14px;font-family:monospace;font-size:16px;font-weight:700;color:#0f2044;">
              {usuario_windows}
            </td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Nombre completo</td>
            <td style="padding:10px 14px;font-weight:600;">{nombre_c}</td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Correo electrónico</td>
            <td style="padding:10px 14px;">
              <a href="mailto:{sol['email']}" style="color:#0f2044;">{sol['email']}</a>
            </td>
          </tr>
          <tr style="border-bottom:1px solid #eee;">
            <td style="padding:10px 14px;color:#888;">Hotel(es)</td>
            <td style="padding:10px 14px;">{sol['hoteles']}</td>
          </tr>
          <tr>
            <td style="padding:10px 14px;color:#888;">ID solicitud</td>
            <td style="padding:10px 14px;font-family:monospace;">#{sol['id']}</td>
          </tr>
        </table>
      </div>
      <div style="padding:14px 28px;background:#f0f0f0;font-size:11px;color:#aaa;">
        Mensaje automático · Control Pedidos Princess Canarias
      </div>
    </div>
    """

    body_text = (
        f"SOLICITUD COMPLETADA — CREAR CUENTA\n"
        f"{'='*44}\n"
        f"Usuario Windows : {usuario_windows}\n"
        f"Nombre          : {nombre_c}\n"
        f"Email           : {sol['email']}\n"
        f"Hotel(es)       : {sol['hoteles']}\n"
        f"ID solicitud    : #{sol['id']}\n"
        f"{'='*44}\n"
        f"Crea la cuenta en el sistema con los datos anteriores."
    )

    destinatarios = _get_solo_admin_emails()
    app_url       = os.environ.get("APP_URL", "").rstrip("/")
    url_admin     = f"{app_url}/admin/solicitudes#{sol['id']}" if app_url else ""

    if not destinatarios:
        log.warning("[SOL_FASE2] Sin emails admin. Sol #%s", sol["id"])
        return jsonify({"ok": True,
                        "msg": "¡Verificación completada! Los administradores podrán verla en el panel."})

    # Telegram SIEMPRE, antes de devolver la respuesta al frontend
    _notify_solicitud_telegram(
        f"\u2705 *[FASE 2 COMPLETADA] Alta pendiente de aprobar*\n\n"
        f"\U0001F464 *{nombre_c}*\n"
        f"\U0001F5A5 Usuario Windows: `{usuario_windows}`\n"
        f"\U0001F4E7 {sol['email']}\n"
        f"\U0001F3E8 {sol['hoteles']}\n"
        f"\U0001F4CB Solicitud `#{sol['id']}` — lista para aprobar."
        + (f"\n\U0001F517 {url_admin}" if url_admin else "")
    )

    # Email via EmailJS en el frontend (sin_email=True siempre)
    # (2026-08-11) Igual que en el código de verificación de login y en la
    # recuperación de contraseña: quien completa esta Fase 2 es un usuario
    # NUEVO, sin cuenta todavía — no hay sesión posible. Marca de un solo
    # uso para que /api/emailjs/registrar-envio no rechace este envío real
    # con 401 y el contador lo cuente también.
    session["pdte_registrar_envio_email"] = True
    return jsonify({
        "ok":            True,
        "sin_email":     True,
        "destinatarios": destinatarios,
        "asunto":        asunto,
        "body_html":     body_html,
        "body_text":     body_text,
        "url_admin":     url_admin,
        "reply_to":      sol["email"],
        "msg": "¡Verificación completada! Los administradores han recibido todos los datos para crear tu cuenta."
    })


# ─── ADMIN: aprobar solicitud → crear cuenta automáticamente ─────────────────

@app.route("/api/admin/solicitudes-acceso/<int:sol_id>/aprobar", methods=["POST"])
def admin_aprobar_solicitud(sol_id):
    """
    Aprueba una solicitud en estado 'completada':
      1. Crea el usuario con usuario_windows como username.
      2. Asigna los hoteles por nombre (mapeo nombre → id).
      3. Genera contraseña temporal.
      4. Envía email de bienvenida al solicitante con sus credenciales.
      5. Marca la solicitud como 'aprobada'.
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permisos"}), 403

    sol = query("SELECT * FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)
    if not sol:
        return jsonify({"error": "Solicitud no encontrada"}), 404
    if sol["estado"] != "completada":
        return jsonify({"error": f"La solicitud está en estado '{sol['estado']}', debe estar 'completada' para aprobarla"}), 409

    username = (sol["usuario_windows"] or "").strip().lower()
    if not username:
        return jsonify({"error": "No hay usuario Windows registrado en esta solicitud"}), 400

    # Comprobar que el username no existe ya
    existing = query("SELECT id FROM usuarios WHERE username=%s", (username,), one=True)
    if existing:
        return jsonify({"error": f"Ya existe un usuario con el username '{username}'"}), 409

    # Generar contraseña temporal legible: Princess + 4 dígitos
    import random as _rnd
    password_temp = "Princess" + str(_rnd.randint(1000, 9999))

    nombre_c = f"{sol['nombre']} {sol['apellidos']}"
    db = get_db()

    # Crear usuario (v11.6.7: se incluye movil de la solicitud y rol 'compras' por defecto)
    # v12.29.37: se guarda el hash de password_temp, no el texto plano; el
    # valor sin hashear (password_temp) se sigue usando tal cual para el
    # email de bienvenida y la respuesta JSON de este endpoint, más abajo.
    cur = execute(
        "INSERT INTO usuarios (username, nombre, email, password, movil, rol, activo) VALUES (%s,%s,%s,%s,%s,'compras',1) RETURNING id",
        (username, nombre_c, sol["email"], generate_password_hash(password_temp), sol.get("movil") or None)
    )
    new_uid = cur.fetchone()["id"]

    # Mapear hoteles texto → IDs (comparación flexible, ignorando mayúsculas y & vs and)
    def _normalizar(s):
        return s.lower().replace("&", "and").replace("  ", " ").strip()

    todos_hoteles = rows_to_list(query("SELECT id, nombre FROM hoteles WHERE activo=1"))
    hoteles_texto = [h.strip() for h in (sol["hoteles"] or "").split(",")]
    hotel_ids_asignados = []
    hoteles_no_encontrados = []
    for ht in hoteles_texto:
        if not ht:
            continue
        hn = _normalizar(ht)
        match = next((h for h in todos_hoteles if _normalizar(h["nombre"]) == hn), None)
        # Búsqueda parcial como fallback
        if not match:
            match = next((h for h in todos_hoteles if hn in _normalizar(h["nombre"]) or _normalizar(h["nombre"]) in hn), None)
        if match:
            hotel_ids_asignados.append(match["id"])
        else:
            hoteles_no_encontrados.append(ht)

    for hid in hotel_ids_asignados:
        execute(
            "INSERT INTO usuario_hoteles (usuario_id, hotel_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
            (new_uid, hid)
        )

    # Marcar solicitud como aprobada
    execute(
        "UPDATE solicitudes_acceso SET estado='aprobada' WHERE id=%s",
        (sol_id,)
    )
    db.commit()

    # Email de bienvenida al nuevo usuario
    app_url   = os.environ.get("APP_URL", "https://control-pedidos-princess.onrender.com").rstrip("/")
    asunto_u  = "✅ Tu acceso ha sido aprobado — Control de Pedidos Princess"
    hoteles_lista = sol["hoteles"] or "—"
    aviso_hoteles = (
        f"<p style='margin:0 0 10px;font-size:12px;color:#b45309;'>"
        f"⚠️ Los siguientes hoteles no se pudieron asignar automáticamente y requerirán ajuste manual: "
        f"<strong>{', '.join(hoteles_no_encontrados)}</strong></p>"
    ) if hoteles_no_encontrados else ""

    body_html_u = f"""
    <div style="font-family:sans-serif;max-width:620px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#0f2044;">
        <tr>
          <td style="padding:24px 0 24px 28px;vertical-align:middle;" valign="middle">
            <h2 style="margin:0;color:#c9a84c;font-size:18px;">🎉 ¡Tu cuenta ha sido creada!</h2>
            <p style="margin:6px 0 0;color:#b9c3dc;font-size:13px;">
              Control de Pedidos · Princess Canarias
            </p>
          </td>
          <td style="padding:14px 28px 14px 16px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
            <img src="{app_url}/static/logo-sidebar-email-64.png" alt="Princess Hotels & Resorts"
                 width="69" height="64"
                 style="height:64px;width:69px;display:block;margin-left:auto;">
          </td>
        </tr>
      </table>
      <div style="padding:28px;">
        <p style="margin:0 0 16px;font-size:15px;color:#333;">
          Hola, <strong>{nombre_c}</strong>
        </p>
        <p style="margin:0 0 20px;font-size:14px;color:#555;line-height:1.6;">
          Tu solicitud de acceso ha sido aprobada. Ya puedes acceder al sistema
          con las siguientes credenciales:
        </p>
        <div style="background:#fff;border:2px solid #c9a84c;border-radius:8px;
                    padding:20px 24px;margin-bottom:20px;">
          <table border="0" cellpadding="0" cellspacing="0" style="width:100%;font-size:14px;">
            <tr>
              <td style="padding:8px 0;color:#888;width:140px;">Usuario</td>
              <td style="padding:8px 0;font-family:monospace;font-size:16px;
                         font-weight:700;color:#0f2044;">{username}</td>
            </tr>
            <tr>
              <td style="padding:8px 0;color:#888;">Contraseña temporal</td>
              <td style="padding:8px 0;font-family:monospace;font-size:16px;
                         font-weight:700;color:#c9a84c;">{password_temp}</td>
            </tr>
            <tr>
              <td style="padding:8px 0;color:#888;">Hoteles asignados</td>
              <td style="padding:8px 0;font-size:13px;color:#333;">{hoteles_lista}</td>
            </tr>
          </table>
        </div>
        {aviso_hoteles}
        <div style="text-align:center;margin-bottom:20px;">
          <a href="{app_url}"
             style="display:inline-block;padding:12px 28px;background:#0f2044;
                    color:#c9a84c;border-radius:7px;text-decoration:none;
                    font-weight:700;font-size:14px;">Acceder al sistema →</a>
        </div>
        <p style="margin:0;font-size:12px;color:#aaa;line-height:1.5;">
          Por seguridad, te recomendamos cambiar la contraseña en tu primer acceso.<br>
          Si tienes cualquier problema contacta con el departamento de compras.
        </p>
      </div>
      <div style="padding:14px 28px;background:#f0f0f0;font-size:11px;color:#aaa;">
        Mensaje automático · Control Pedidos Princess Canarias
      </div>
    </div>
    """

    body_text_u = (
        f"Hola {nombre_c},\n\n"
        f"Tu solicitud de acceso ha sido aprobada.\n\n"
        f"Usuario         : {username}\n"
        f"Contraseña temp.: {password_temp}\n"
        f"Hoteles         : {hoteles_lista}\n\n"
        f"Accede en: {app_url}\n\n"
        f"Te recomendamos cambiar la contraseña en tu primer acceso.\n\n"
        f"Control Pedidos Princess Canarias"
    )

    # El envío real lo hace el frontend vía EmailJS — siempre pendiente
    res_u = {"ok": False}

    # Email de confirmación a los admins
    asunto_a = f"[APROBADA] Alta usuario {username} — {nombre_c}"
    body_html_a = f"""
    <div style="font-family:sans-serif;max-width:580px;margin:0 auto;
                background:#f9f9f9;border-radius:10px;overflow:hidden;
                border:1px solid #e0e0e0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#065f46;">
        <tr>
          <td style="padding:20px 0 20px 24px;vertical-align:middle;" valign="middle">
            <h2 style="margin:0;color:#6ee7b7;font-size:16px;">✅ Cuenta creada automáticamente</h2>
          </td>
          <td style="padding:10px 24px 10px 14px;vertical-align:middle;text-align:right;width:1%;white-space:nowrap;" valign="middle" align="right">
            <img src="{app_url}/static/logo-sidebar-email.png" alt="Princess Hotels & Resorts"
                 width="43" height="40"
                 style="height:40px;width:43px;display:block;margin-left:auto;">
          </td>
        </tr>
      </table>
      <div style="padding:20px 24px;font-size:14px;color:#333;">
        <p>La solicitud #{sol_id} de <strong>{nombre_c}</strong> ha sido aprobada.</p>
        <table border="0" cellpadding="0" cellspacing="0" style="font-size:13px;width:100%;">
          <tr><td style="color:#888;padding:5px 0;width:130px;">Username</td><td style="font-family:monospace;font-weight:700;">{username}</td></tr>
          <tr><td style="color:#888;padding:5px 0;">Email</td><td>{sol['email']}</td></tr>
          <tr><td style="color:#888;padding:5px 0;">Hoteles</td><td>{hoteles_lista}</td></tr>
          {'<tr><td style="color:#b45309;padding:5px 0;">⚠️ Sin asignar</td><td style="color:#b45309;">' + ", ".join(hoteles_no_encontrados) + "</td></tr>" if hoteles_no_encontrados else ""}
        </table>
        <p style="margin:14px 0 0;font-size:12px;color:#aaa;">
          El usuario ha recibido su email de bienvenida con credenciales.
        </p>
      </div>
    </div>
    """
    body_text_a = (
        f"Cuenta creada automáticamente\n\n"
        f"La solicitud #{sol_id} de {nombre_c} ha sido aprobada.\n\n"
        f"Username: {username}\n"
        f"Email   : {sol['email']}\n"
        f"Hoteles : {hoteles_lista}\n"
        + (f"⚠️ Hoteles sin asignar: {', '.join(hoteles_no_encontrados)}\n" if hoteles_no_encontrados else "")
        + f"\nEl usuario ha recibido su email de bienvenida con credenciales."
    )
    destinatarios = _get_solo_admin_emails()
    # Los emails a admins también se envían desde el frontend vía EmailJS
    admins_email_enviado = False

    return jsonify({
        "ok":       True,
        "uid":      new_uid,
        "username": username,
        "password": password_temp,
        "hoteles_asignados":      hotel_ids_asignados,
        "hoteles_no_encontrados": hoteles_no_encontrados,
        "email_enviado":          res_u.get("ok", False),
        # Datos para que el frontend envíe vía EmailJS:
        "email_usuario_pendiente": (not res_u.get("ok", False)) and {
            "to_email":  sol["email"],
            "asunto":    asunto_u,
            "body_html": body_html_u,
            "body_text": body_text_u,
        } or None,
        "email_admins_pendiente": (not admins_email_enviado and destinatarios) and {
            "destinatarios": destinatarios,
            "asunto":        asunto_a,
            "body_html":     body_html_a,
            "body_text":     body_text_a,
        } or None,
        "abrir_edicion":          True,
        "msg": f"Cuenta creada para {nombre_c} ({username})."
    })


# ─── ADMIN: rechazar solicitud ────────────────────────────────────────────────

@app.route("/api/admin/solicitudes-acceso/<int:sol_id>/rechazar", methods=["POST"])
def admin_rechazar_solicitud(sol_id):
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permisos"}), 403
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            "UPDATE solicitudes_acceso SET estado='rechazada' WHERE id=%s", (sol_id,)
        )
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/admin/solicitudes-acceso/<int:sol_id>", methods=["DELETE"])
def admin_borrar_solicitud(sol_id):
    """Elimina una solicitud de acceso del histórico (solo admins).
    No afecta a la cuenta de usuario ya creada, si la solicitud estaba aprobada."""
    if session.get("rol") != "admin":
        return jsonify({"error": "Sin permisos"}), 403
    sol = query("SELECT id FROM solicitudes_acceso WHERE id=%s", (sol_id,), one=True)
    if not sol:
        return jsonify({"error": "Solicitud no encontrada"}), 404
    db = get_db()
    execute("DELETE FROM solicitudes_acceso WHERE id=%s", (sol_id,))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/me")
def me():
    if "user_id" not in session:
        return jsonify({"logged": False})
    if session.get("login_date") != _hoy_canarias().isoformat():
        session.clear()
        return jsonify({"logged": False, "sesion_caducada": True})
    return jsonify({"logged": True, "id": session["user_id"], "username": session["username"],
                    "nombre": session["nombre"], "rol": session["rol"],
                    "hoteles_ids": session.get("hoteles_ids", [])})

# ── API Maestros ───────────────────────────────────────────────────────────────

@app.route("/api/maestros")
@login_required
def get_maestros():
    _rol = session.get("rol")
    _ver_pruebas = _puede_ver_hotel_pruebas()
    if _rol == "hotel":
        hoteles_ids = session.get("hoteles_ids", [])
        if not hoteles_ids:
            hoteles = []
        elif _ver_pruebas:
            placeholders = ",".join(["%s"] * len(hoteles_ids))
            hoteles = rows_to_list(query(
                f"SELECT * FROM hoteles WHERE activo=1 AND id IN ({placeholders}) ORDER BY codigo",
                tuple(hoteles_ids)
            ))
        else:
            placeholders = ",".join(["%s"] * len(hoteles_ids))
            hoteles = rows_to_list(query(
                f"SELECT * FROM hoteles WHERE activo=1 AND id IN ({placeholders}) "
                f"AND codigo <> %s ORDER BY codigo",
                tuple(hoteles_ids) + (HOTEL_CODIGO_PRUEBAS,)
            ))
    elif _ver_pruebas:
        hoteles = rows_to_list(query("SELECT * FROM hoteles WHERE activo=1 ORDER BY codigo"))
    else:
        # compras (u otro rol no-admin, salvo el usuario 'Prueba'): mismo
        # listado salvo el hotel de pruebas.
        hoteles = rows_to_list(query(
            "SELECT * FROM hoteles WHERE activo=1 AND codigo <> %s ORDER BY codigo",
            (HOTEL_CODIGO_PRUEBAS,)
        ))
    departamentos = rows_to_list(query("SELECT * FROM departamentos WHERE activo=1 ORDER BY nombre"))
    familias      = rows_to_list(query("SELECT * FROM familias WHERE activo=1 ORDER BY nombre"))
    return jsonify({
        "hoteles":       hoteles,
        "departamentos": departamentos,
        "proveedores":   [],
        "estados":       ESTADOS_VALIDOS,
        "familias":      familias,
    })

# ── API Familias ───────────────────────────────────────────────────────────────

@app.route("/api/familias", methods=["GET"])
@login_required
def get_familias():
    rows = rows_to_list(query("SELECT * FROM familias WHERE activo=1 ORDER BY nombre"))
    return jsonify(rows)

@app.route("/api/familias", methods=["POST"])
@admin_required
def create_familia():
    data   = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400
    db  = get_db()
    cur = execute("INSERT INTO familias (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING RETURNING id", (nombre,))
    row = cur.fetchone()
    if not row:
        return jsonify({"error": "Ya existe una familia con ese nombre"}), 409
    db.commit()
    return jsonify({"ok": True, "id": row["id"], "nombre": nombre}), 201

@app.route("/api/familias/<int:fid>", methods=["PUT"])
@admin_required
def update_familia(fid):
    data   = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400
    db = get_db()
    execute("UPDATE familias SET nombre=%s WHERE id=%s", (nombre, fid))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/familias/<int:fid>", methods=["DELETE"])
@admin_required
def delete_familia(fid):
    db  = get_db()
    cnt = query("SELECT COUNT(*) as n FROM pedidos WHERE familia_id=%s AND sujeto_techo=1", (fid,), one=True)["n"]
    if cnt > 0:
        return jsonify({"error": f"No se puede eliminar: tiene {cnt} pedido(s) asociado(s)"}), 409
    execute("UPDATE familias SET activo=0 WHERE id=%s", (fid,))
    db.commit()
    return jsonify({"ok": True})

# ── API Usuarios (gestión admin) ───────────────────────────────────────────────

@app.route("/api/usuarios", methods=["GET"])
@admin_required
def get_usuarios():
    rows = rows_to_list(query(
        "SELECT id, username, nombre, email, email2, movil, rol, activo, creado_en, telegram_chat_id FROM usuarios ORDER BY nombre"
    ))
    return jsonify(rows)

@app.route("/api/usuarios", methods=["POST"])
@admin_required
def create_usuario():
    data     = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip().lower()
    nombre   = (data.get("nombre")   or "").strip()
    password = (data.get("password") or "").strip()
    email    = (data.get("email")    or "").strip()
    email2   = (data.get("email2")   or "").strip()
    if not username or not nombre or not password:
        return jsonify({"error": "username, nombre y contraseña son obligatorios"}), 400
    # (2026-07-31) El email NO es obligatorio: un admin puede dejarlo vacío
    # a propósito para anular el envío de correos a ese usuario sin tener
    # que desactivar la cuenta entera. La falta de email en compradores/
    # admins activos ya se detecta y avisa en Admin → Integridad, no hace
    # falta bloquear el guardado aquí.
    existing = query("SELECT id FROM usuarios WHERE username=%s", (username,), one=True)
    if existing:
        return jsonify({"error": "Ya existe un usuario con ese username"}), 409
    db  = get_db()
    rol = data.get("rol", "user")
    if rol not in ("admin", "user", "hotel", "compras"):
        rol = "user"
    cur = execute(
        "INSERT INTO usuarios (username, nombre, email, email2, movil, password, rol, activo, telegram_chat_id) VALUES (%s,%s,%s,%s,%s,%s,%s,1,%s) RETURNING id",
        (username, nombre, email, email2 or None, data.get("movil",""), generate_password_hash(password), rol,
         (data.get("telegram_chat_id") or "").strip() or None)
    )
    new_id = cur.fetchone()["id"]
    db.commit()
    return jsonify({"ok": True, "id": new_id}), 201

@app.route("/api/usuarios/<int:uid>", methods=["PUT"])
@admin_required
def update_usuario(uid):
    data = request.get_json(silent=True) or {}
    db   = get_db()
    # No permitir que el admin se quite el rol a sí mismo
    if uid == current_user_id() and data.get("rol") in ("user", "hotel", "compras"):
        return jsonify({"error": "No puedes quitarte el rol de administrador a ti mismo"}), 400
    # Construir UPDATE dinámico solo con campos enviados
    fields, args = [], []
    if "nombre" in data:
        fields.append("nombre=%s"); args.append(data["nombre"].strip())
    if "email" in data:
        # (2026-07-31) Se permite vaciarlo a propósito — es la forma de que
        # un admin anule el envío de correos a este usuario sin desactivar
        # la cuenta. Se detecta y avisa en Admin → Integridad si el usuario
        # sigue activo con rol comprador/admin.
        fields.append("email=%s"); args.append((data["email"] or "").strip())
    if "email2" in data:
        fields.append("email2=%s"); args.append((data["email2"] or "").strip() or None)
    if "movil" in data:
        fields.append("movil=%s"); args.append((data["movil"] or "").strip())
    if "rol" in data and data["rol"] in ("admin", "user", "hotel", "compras"):
        fields.append("rol=%s"); args.append(data["rol"])
    if "activo" in data:
        # ── Protección: no desactivar comprador si deja hoteles huérfanos ────
        desactivando = (not data["activo"]) or (data["activo"] == 0)
        if desactivando:
            usuario_actual = query("SELECT rol, activo FROM usuarios WHERE id=%s", (uid,), one=True)
            if usuario_actual and usuario_actual["rol"] == "compras" and usuario_actual["activo"] == 1:
                huerfanos = rows_to_list(query("""
                    SELECT h.codigo FROM hoteles h
                    JOIN usuario_comprador_hoteles uch ON uch.hotel_id = h.id
                    WHERE uch.usuario_id = %s
                      AND h.activo = 1
                      AND NOT EXISTS (
                          SELECT 1 FROM usuario_comprador_hoteles uch2
                          JOIN usuarios u2 ON u2.id = uch2.usuario_id
                          WHERE uch2.hotel_id = h.id
                            AND uch2.usuario_id != %s
                            AND u2.activo = 1
                            AND u2.rol = 'compras'
                      )
                """, (uid, uid)))
                if huerfanos:
                    codigos = ", ".join(r["codigo"] for r in huerfanos)
                    return jsonify({
                        "error": f"⚠️ No se puede desactivar: los hoteles {codigos} quedarían sin comprador asignado. "
                                 f"Reasígnalos a otro comprador antes de desactivar este usuario."
                    }), 409
        fields.append("activo=%s"); args.append(1 if data["activo"] else 0)
    if "password" in data and data["password"].strip():
        fields.append("password=%s"); args.append(generate_password_hash(data["password"].strip()))
    if "telegram_chat_id" in data:
        fields.append("telegram_chat_id=%s"); args.append((data["telegram_chat_id"] or "").strip() or None)
    if not fields:
        return jsonify({"error": "Nada que actualizar"}), 400
    args.append(uid)
    execute(f"UPDATE usuarios SET {', '.join(fields)} WHERE id=%s", args)
    db.commit()
    # Si cambié mi propio nombre, actualizar sesión
    if uid == current_user_id() and "nombre" in data:
        session["nombre"] = data["nombre"].strip()
    return jsonify({"ok": True})

# ── API Hoteles de usuario (rol hotel) ────────────────────────────────────────

@app.route("/api/usuarios/<int:uid>/hoteles", methods=["GET"])
@admin_required
def get_usuario_hoteles(uid):
    rows = rows_to_list(query(
        "SELECT hotel_id FROM usuario_hoteles WHERE usuario_id=%s", (uid,)
    ))
    return jsonify([r["hotel_id"] for r in rows])

@app.route("/api/usuarios/<int:uid>/hoteles", methods=["PUT"])
@admin_required
def set_usuario_hoteles(uid):
    data = request.get_json(silent=True) or {}
    hotel_ids = data.get("hotel_ids", [])
    db = get_db()
    execute("DELETE FROM usuario_hoteles WHERE usuario_id=%s", (uid,))
    for hid in hotel_ids:
        execute("INSERT INTO usuario_hoteles (usuario_id, hotel_id) VALUES (%s,%s) ON CONFLICT DO NOTHING", (uid, hid))
    db.commit()
    return jsonify({"ok": True})

# ── API Hoteles de usuario compras (rol compras) ──────────────────────────────
# Permite asignar/desasignar hoteles a compradores desde el panel de admin,
# sustituyendo el diccionario HOTEL_COMPRADOR hardcodeado.

@app.route("/api/usuarios/<int:uid>/hoteles-compras", methods=["GET"])
@admin_required
def get_usuario_comprador_hoteles(uid):
    """Devuelve los hotel_id asignados a un usuario compras."""
    rows = rows_to_list(query(
        "SELECT hotel_id FROM usuario_comprador_hoteles WHERE usuario_id=%s", (uid,)
    ))
    return jsonify([r["hotel_id"] for r in rows])


@app.route("/api/usuarios/<int:uid>/hoteles-compras", methods=["PUT"])
@admin_required
def set_usuario_comprador_hoteles(uid):
    """
    Reemplaza completamente los hoteles asignados a un usuario compras.

    Modelo: 1 hotel → 1 comprador.
    Si algún hotel ya tiene otro comprador asignado, devuelve 409 con la lista
    de conflictos para que el frontend muestre la confirmación de reasignación.
    Si se envía forzar=true, los hoteles en conflicto se reasignan automáticamente
    (se eliminan del comprador anterior).
    """
    data      = request.get_json(silent=True) or {}
    hotel_ids = data.get("hotel_ids", [])
    forzar    = bool(data.get("forzar", False))
    db        = get_db()

    # ── Protección: hoteles que este comprador va a PERDER → ¿quedarán huérfanos? ──
    # Se calcula antes de cualquier DELETE. Un hotel queda huérfano si:
    #   - está actualmente asignado a este comprador
    #   - NO está en la nueva lista hotel_ids
    #   - NO tiene otro comprador alternativo activo
    hoteles_actuales = [
        r["hotel_id"] for r in rows_to_list(
            query("SELECT hotel_id FROM usuario_comprador_hoteles WHERE usuario_id=%s", (uid,))
        )
    ]
    hoteles_a_perder = [hid for hid in hoteles_actuales if hid not in hotel_ids]
    huerfanos_por_vaciado = []
    for hid in hoteles_a_perder:
        otro_comprador = query(
            """SELECT u.id FROM usuario_comprador_hoteles uch
               JOIN usuarios u ON u.id = uch.usuario_id
               WHERE uch.hotel_id = %s AND uch.usuario_id != %s
                 AND u.activo = 1 AND u.rol = 'compras'
               LIMIT 1""",
            (hid, uid), one=True
        )
        if not otro_comprador:
            hotel = query("SELECT codigo, nombre FROM hoteles WHERE id=%s AND activo=1", (hid,), one=True)
            if hotel:
                huerfanos_por_vaciado.append({
                    "hotel_id":     hid,
                    "hotel_codigo": hotel["codigo"],
                    "hotel_nombre": hotel["nombre"],
                })
    if huerfanos_por_vaciado and not forzar:
        codigos = ", ".join(h["hotel_codigo"] for h in huerfanos_por_vaciado)
        return jsonify({
            "ok": False,
            "error": f"⚠️ Los hoteles {codigos} se quedarían sin comprador asignado "
                     f"(todavía no se los has dado a nadie más).",
            "huerfanos": huerfanos_por_vaciado,
        }), 409
    elif huerfanos_por_vaciado:
        # forzar=true: el admin ha confirmado que quiere dejarlos sin comprador
        # de momento (p. ej. porque va a asignárselos a otro usuario a
        # continuación, en un segundo guardado). Se deja constancia en el log.
        log.warning(
            "Hoteles-compras: %s quedan SIN comprador (forzado por admin, usuario_id=%s)",
            [h["hotel_codigo"] for h in huerfanos_por_vaciado], uid
        )

    # ── Detectar conflictos: hoteles ya asignados a otro comprador ───────────
    conflictos = []
    for hid in hotel_ids:
        otro = query(
            """SELECT u.id, u.nombre
               FROM usuario_comprador_hoteles uch
               JOIN usuarios u ON u.id = uch.usuario_id
               WHERE uch.hotel_id = %s AND uch.usuario_id != %s
               LIMIT 1""",
            (hid, uid), one=True
        )
        if otro:
            hotel = query("SELECT codigo, nombre FROM hoteles WHERE id=%s", (hid,), one=True)
            conflictos.append({
                "hotel_id":               hid,
                "hotel_codigo":           hotel["codigo"]  if hotel else str(hid),
                "hotel_nombre":           hotel["nombre"]  if hotel else "",
                "comprador_actual_id":    otro["id"],
                "comprador_actual_nombre": otro["nombre"],
            })

    # Si hay conflictos y no se ha confirmado la reasignación, devolver 409
    if conflictos and not forzar:
        return jsonify({"ok": False, "conflictos": conflictos}), 409

    # ── Reasignación: quitar estos hoteles de cualquier comprador anterior ────
    for hid in hotel_ids:
        execute("DELETE FROM usuario_comprador_hoteles WHERE hotel_id=%s", (hid,))

    # ── Borrar asignaciones previas de este comprador y aplicar las nuevas ────
    execute("DELETE FROM usuario_comprador_hoteles WHERE usuario_id=%s", (uid,))
    for hid in hotel_ids:
        execute(
            "INSERT INTO usuario_comprador_hoteles (usuario_id, hotel_id) VALUES (%s,%s)",
            (uid, hid)
        )
    db.commit()
    reasignados = len(conflictos) if forzar else 0
    log.info(
        "Hoteles-compras actualizados: usuario_id=%s hoteles=%s reasignados=%s",
        uid, hotel_ids, reasignados
    )
    return jsonify({"ok": True, "reasignados": reasignados})


@app.route("/api/usuarios/hoteles-asignados")
@admin_required
def get_usuarios_hoteles_asignados():
    """
    (2026-09-01, repaso "agilizar y limpiar", Etapa 4) Versión "todos de
    una vez" de GET /api/usuarios/<id>/hoteles y GET
    /api/usuarios/<id>/hoteles-compras — esos dos endpoints se mantienen
    tal cual (los sigue usando el modal de edición de un usuario
    concreto), pero loadUsuarios() los llamaba una vez POR CADA usuario
    con rol hotel/compras para pintar la tabla entera — con 40 usuarios,
    eran ~40 peticiones solo para esto. Aquí se hacen las 2 consultas
    (una por tabla) una única vez y se devuelven ya agrupadas por
    usuario_id, sin tocar el modelo de datos ni los endpoints existentes.
    """
    hoteles_rows = rows_to_list(query(
        "SELECT usuario_id, hotel_id FROM usuario_hoteles"
    ))
    compras_rows = rows_to_list(query(
        "SELECT usuario_id, hotel_id FROM usuario_comprador_hoteles"
    ))
    hoteles = {}
    for r in hoteles_rows:
        hoteles.setdefault(r["usuario_id"], []).append(r["hotel_id"])
    compras = {}
    for r in compras_rows:
        compras.setdefault(r["usuario_id"], []).append(r["hotel_id"])
    return jsonify({"hoteles": hoteles, "compras": compras})


@app.route("/api/compradores-por-hotel")
@admin_required
def get_compradores_por_hotel():
    """
    Devuelve un resumen de todos los hoteles con sus compradores asignados.
    Incluye campo sin_comprador=True para los hoteles que no tienen comprador,
    y un resumen de integridad global al final.
    Útil para que admin visualice la distribución actual y detecte huérfanos.
    """
    hoteles = rows_to_list(query("SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"))
    resultado = []
    huerfanos = 0
    for hotel in hoteles:
        compradores = rows_to_list(query(
            """SELECT u.id, u.username, u.nombre, u.email, u.movil, u.telegram_chat_id
               FROM usuarios u
               JOIN usuario_comprador_hoteles uch ON uch.usuario_id = u.id
               WHERE uch.hotel_id = %s AND u.activo = 1 AND u.rol = 'compras'
               ORDER BY u.nombre""",
            (hotel["id"],)
        ))
        sin_comprador = len(compradores) == 0
        if sin_comprador:
            huerfanos += 1
        resultado.append({
            "hotel_id":      hotel["id"],
            "hotel_codigo":  hotel["codigo"],
            "hotel_nombre":  hotel["nombre"],
            "compradores":   compradores,
            "sin_comprador": sin_comprador,
        })
    return jsonify({
        "hoteles":   resultado,
        "integridad": {
            "total_hoteles":       len(resultado),
            "hoteles_sin_comprador": huerfanos,
            "ok":                  huerfanos == 0,
        },
    })


@app.route("/api/usuarios/<int:uid>", methods=["DELETE"])
@admin_required
def delete_usuario(uid):
    # No puede eliminarse a sí mismo
    if uid == current_user_id():
        return jsonify({"error": "No puedes eliminar tu propio usuario"}), 400
    # Verificar que existe y obtener nombre
    user = query("SELECT id, username, nombre FROM usuarios WHERE id=%s", (uid,), one=True)
    if not user:
        return jsonify({"error": "Usuario no encontrado"}), 404
    nombre = user["nombre"]
    db = get_db()
    # ── Protección: no eliminar comprador si deja hoteles huérfanos ──────────
    if user_row := query("SELECT rol FROM usuarios WHERE id=%s", (uid,), one=True):
        if user_row["rol"] == "compras":
            huerfanos = rows_to_list(query("""
                SELECT h.codigo FROM hoteles h
                JOIN usuario_comprador_hoteles uch ON uch.hotel_id = h.id
                WHERE uch.usuario_id = %s
                  AND h.activo = 1
                  AND NOT EXISTS (
                      SELECT 1 FROM usuario_comprador_hoteles uch2
                      JOIN usuarios u2 ON u2.id = uch2.usuario_id
                      WHERE uch2.hotel_id = h.id
                        AND uch2.usuario_id != %s
                        AND u2.activo = 1
                        AND u2.rol = 'compras'
                  )
            """, (uid, uid)))
            if huerfanos:
                codigos = ", ".join(r["codigo"] for r in huerfanos)
                return jsonify({
                    "error": f"⚠️ No se puede eliminar: los hoteles {codigos} quedarían sin comprador asignado. "
                             f"Reasígnalos a otro comprador antes de eliminar este usuario."
                }), 409
    # ── Congelar nombre en pedidos antes de que la FK quede NULL ─────────────
    execute("""
        UPDATE pedidos SET creado_por_nombre = %s
        WHERE creado_por_id = %s AND (creado_por_nombre IS NULL OR creado_por_nombre = '')
    """, (nombre, uid))
    execute("""
        UPDATE pedidos SET modificado_por_nombre = %s
        WHERE modificado_por_id = %s AND (modificado_por_nombre IS NULL OR modificado_por_nombre = '')
    """, (nombre, uid))
    execute("""
        UPDATE historial_estados SET usuario_nombre = %s
        WHERE usuario_id = %s AND (usuario_nombre IS NULL OR usuario_nombre = '')
    """, (nombre, uid))
    # ── Eliminar usuario (usuario_hoteles y password_reset_tokens en CASCADE) ─
    execute("DELETE FROM usuarios WHERE id=%s", (uid,))
    db.commit()
    return jsonify({"ok": True})

# ── API Proveedores ────────────────────────────────────────────────────────────

def _prov_with_contactos(rows):
    """Añade lista de contactos a cada proveedor (con los hoteles
    específicos asignados a cada uno, si tiene)."""
    result = rows_to_list(rows)
    if not result:
        return result
    ids = [p["id"] for p in result]
    placeholders = ",".join(["%s"] * len(ids))
    contactos_rows = rows_to_list(query(
        f"""SELECT pc.id, pc.proveedor_id, pc.nombre, pc.telefono, pc.movil, pc.email, pc.es_principal,
                   COALESCE(
                       (SELECT array_agg(pch.hotel_id ORDER BY pch.hotel_id)
                        FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id),
                       ARRAY[]::INTEGER[]
                   ) AS hotel_ids
            FROM proveedor_contactos pc
            WHERE pc.proveedor_id IN ({placeholders})
            ORDER BY pc.proveedor_id, pc.es_principal DESC, pc.orden, pc.id""",
        tuple(ids)
    ))
    # Agrupar por proveedor_id
    from collections import defaultdict
    cmap = defaultdict(list)
    for c in contactos_rows:
        cmap[c["proveedor_id"]].append({
            "nombre":       c["nombre"] or "",
            "telefono":     c["telefono"] or "",
            "movil":        c["movil"] or "",
            "email":        c["email"] or "",
            "es_principal": bool(c["es_principal"]),
            "hotel_ids":    list(c["hotel_ids"] or []),
        })
    for p in result:
        p["contactos"] = cmap.get(p["id"], [])
        # Campos de compatibilidad: usar contacto principal (o primero si no hay)
        principal = next((c for c in p["contactos"] if c.get("es_principal")), p["contactos"][0] if p["contactos"] else {})
        p["contacto"]       = principal.get("nombre", "")
        p["email"]          = principal.get("email", "")
        p["telefono"]       = principal.get("telefono", "")
        p["movil_principal"] = principal.get("movil", "")
    return result

@app.route("/api/proveedores", methods=["GET"])
@login_required
def get_proveedores():
    # (2026-08-31, auditoría de rendimiento — Víctor: "la ficha proveedores
    # se atasca un poco") Antes este endpoint devolvía SIEMPRE la tabla
    # `proveedores` entera (sin paginar), y el frontend reconstruía toda la
    # tabla en cada carga de la vista y en cada tecla del buscador — cuantos
    # más proveedores se dan de alta con el tiempo, más pesada se pone cada
    # carga, a diferencia de /api/pedidos, que ya está paginado desde hace
    # tiempo. Mismo patrón de paginación que get_pedidos() (page/page_size,
    # devuelve total/pages), aplicado aquí ahora. El buscador (q) usa ILIKE
    # con comodín al principio en los 3 campos — ahora sí puede apoyarse en
    # los índices pg_trgm creados en _auto_migrate().
    #
    # OJO al tocar la forma de la respuesta: antes era un array plano;
    # ahora es {proveedores,total,page,page_size,pages}. Los dos
    # consumidores del frontend (loadProveedores() y el autocompletado
    # buscarProveedor() en el modal de pedido) se actualizaron a la vez en
    # esta misma entrega — si se añade un tercer consumidor, recordar que
    # ya no es un array.
    q = request.args.get("q", "").strip()
    try:
        page      = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(100, int(request.args.get("page_size", 30))))
    except ValueError:
        page, page_size = 1, 30

    wheres, args = ["activo=1"], []
    if q:
        # (2026-08-31) Víctor: "debe dejar buscar por nombre, codigo sap y
        # codigo dali" — antes solo buscaba por nombre. Un único cuadro de
        # búsqueda, coincide con cualquiera de los tres campos (OR), igual
        # de flexible (ILIKE + comodines) en los tres.
        wheres.append("(nombre ILIKE %s OR codigo ILIKE %s OR codigo_dali ILIKE %s)")
        args += [f"%{q}%", f"%{q}%", f"%{q}%"]
    where_sql = " AND ".join(wheres)

    total = query(f"SELECT COUNT(*) as total FROM proveedores WHERE {where_sql}", args, one=True)["total"]
    rows = query(
        f"""SELECT id,codigo,codigo_dali,nombre,observaciones,sujeto_seguimiento
            FROM proveedores WHERE {where_sql}
            ORDER BY nombre
            LIMIT %s OFFSET %s""",
        args + [page_size, (page - 1) * page_size]
    )
    result = _prov_with_contactos(rows)
    # Rol hotel: solo consulta — se eliminan observaciones de la respuesta
    if session.get("rol") == "hotel":
        for p in result:
            p.pop("observaciones", None)
    return jsonify({
        "proveedores": result,
        "total":       total,
        "page":        page,
        "page_size":   page_size,
        "pages":       max(1, (total + page_size - 1) // page_size),
    })


def _buscar_proveedor_duplicado(campo: str, valor: str, excluir_id: int = None) -> dict:
    """
    (2026-08-31) Usado por create_proveedor/update_proveedor para el aviso
    de código SAP/DALI duplicado que pidió Víctor: "deberá indicar qué
    código está duplicado, nombre asociado, etc. para poder localizarlo y
    arreglarlo". Devuelve {id, nombre} del proveedor activo que ya tiene
    ese valor exacto en la columna `campo` ('codigo' o 'codigo_dali' — SOLO
    se llama con estos dos literales fijos desde este archivo, nunca con
    algo que venga del usuario, así que interpolarlo en el SQL es seguro),
    o None si no hay ninguno. `excluir_id` se usa al editar, para no
    comparar un proveedor consigo mismo.
    """
    sql = f"SELECT id, nombre FROM proveedores WHERE activo=1 AND {campo}=%s"
    args = [valor]
    if excluir_id is not None:
        sql += " AND id!=%s"
        args.append(excluir_id)
    return row_to_dict(query(sql, tuple(args), one=True))


@app.route("/api/proveedores", methods=["POST"])
@login_required
def create_proveedor():
    # (2026-08-31) Víctor: "creo que solo es admin la creacion y
    # modificacion del nombre y codigo, los compradores pueden editar
    # contactos". Antes (2026-08-10) compras también podía dar de alta
    # proveedores nuevos con nombre y código SAP propios — decisión
    # explícita de aquella entrega. Víctor la revisa ahora y pide que la
    # creación (que implica fijar nombre/código) quede solo para admin;
    # compras deja de poder crear proveedores nuevos, pero conserva la
    # edición de contactos/observaciones en los ya existentes (ver
    # update_proveedor más abajo).
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido — solo un administrador puede crear proveedores nuevos"}), 403
    data   = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    codigo = (data.get("codigo") or "").strip()
    codigo_dali = (data.get("codigo_dali") or "").strip()
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400
    if not codigo:
        return jsonify({"error": "El código SAP es obligatorio"}), 400
    # (2026-08-31) A petición de Víctor: "tanto el codigo SAP como el DALI
    # son obligatorios al crear un proveedor" — antes codigo_dali era
    # opcional (NULL permitido) desde que se añadió la columna (v12.30.56).
    if not codigo_dali:
        return jsonify({"error": "El código DALI es obligatorio"}), 400
    # Anti-duplicado: mismo nombre (insensible a mayúsculas)
    dup_nombre = query(
        "SELECT id FROM proveedores WHERE activo=1 AND LOWER(nombre)=LOWER(%s)", (nombre,)
    )
    if rows_to_list(dup_nombre):
        return jsonify({"error": f"Ya existe un proveedor con el nombre '{nombre}'"}), 409
    # (2026-08-31) A petición de Víctor: "en caso de duplicidad de alguno de
    # los dos códigos ahora está realizando error silencioso (...) deberá
    # indicar qué código está duplicado, nombre asociado, etc. para poder
    # localizarlo y arreglarlo" — el 409 de código SAP duplicado ya existía,
    # pero (a) el frontend no lo llegaba a mostrar nunca (ver saveProveedor,
    # templates/index.html — faltaba el try/catch alrededor de api(), así
    # que la excepción se perdía sin avisar) y (b) el mensaje no decía CON
    # QUÉ proveedor chocaba. Además, código DALI no tenía ningún chequeo de
    # duplicado — se podían crear dos proveedores con el mismo código DALI
    # sin ningún aviso. _buscar_proveedor_duplicado() (definida más abajo,
    # junto a update_proveedor) da nombre + ID del proveedor que ya tiene
    # ese código, para poder localizarlo y arreglarlo tal como pidió.
    dup_codigo = _buscar_proveedor_duplicado("codigo", codigo)
    if dup_codigo:
        return jsonify({"error": (
            f"El código SAP '{codigo}' ya está en uso por el proveedor «{dup_codigo['nombre']}» "
            f"(ID {dup_codigo['id']}) — corrige uno de los dos códigos."
        )}), 409
    dup_dali = _buscar_proveedor_duplicado("codigo_dali", codigo_dali)
    if dup_dali:
        return jsonify({"error": (
            f"El código DALI '{codigo_dali}' ya está en uso por el proveedor «{dup_dali['nombre']}» "
            f"(ID {dup_dali['id']}) — corrige uno de los dos códigos."
        )}), 409
    db  = get_db()
    # sujeto_seguimiento: solo admin puede llegar aquí ahora, así que se
    # toma directamente del payload (ya no hace falta la rama "compras
    # siempre FALSE" de antes, porque compras ya no crea proveedores).
    sujeto_seg = bool(data.get("sujeto_seguimiento", False))
    cur = execute(
        "INSERT INTO proveedores (codigo,codigo_dali,nombre,observaciones,sujeto_seguimiento) VALUES (%s,%s,%s,%s,%s) RETURNING id",
        (codigo, codigo_dali, nombre, data.get("observaciones",""), sujeto_seg)
    )
    new_id = cur.fetchone()["id"]
    # Insertar contactos
    contactos = data.get("contactos", [])
    for i, c in enumerate(contactos):
        nombre_c    = (c.get("nombre") or "").strip() or None
        tel_c       = (c.get("telefono") or "").strip() or None
        movil_c     = (c.get("movil") or "").strip() or None
        email_c     = (c.get("email") or "").strip() or None
        principal_c = 1 if c.get("es_principal") else 0
        if nombre_c or tel_c or movil_c or email_c:
            cur_c = execute(
                "INSERT INTO proveedor_contactos (proveedor_id,nombre,telefono,movil,email,es_principal,orden) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                (new_id, nombre_c, tel_c, movil_c, email_c, principal_c, i)
            )
            contacto_id = cur_c.fetchone()["id"]
            for hid in (c.get("hotel_ids") or []):
                execute(
                    "INSERT INTO proveedor_contacto_hoteles (contacto_id, hotel_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (contacto_id, hid)
                )
    db.commit()
    return jsonify({"ok": True, "id": new_id, "nombre": nombre}), 201

@app.route("/api/proveedores/<int:pid>", methods=["PUT"])
@login_required
def update_proveedor(pid):
    if session.get("rol") not in ("admin", "compras"):
        return jsonify({"error": "Acceso restringido"}), 403
    data = request.get_json(silent=True) or {}
    # (2026-08-31) Víctor: "creo que solo es admin la creacion y
    # modificacion del nombre y codigo, los compradores pueden editar
    # contactos ( esto ultimo verificalo porque creo que les da error o no
    # hace nada cuando intentan guardar los cambios". Confirmado: antes,
    # nombre/codigo eran obligatorios en TODO PUT sin importar el rol, pero
    # el frontend nunca enviaba "codigo" para compras (el campo ni se
    # muestra editable para ese rol) — así que cualquier guardado de
    # compras (aunque solo tocara contactos) fallaba siempre con "El código
    # SAP es obligatorio". Se corrige con el mismo patrón ya usado para
    # sujeto_seguimiento: nombre/codigo/codigo_dali solo se toman del
    # payload si el rol es admin; si no, se conservan los valores ya
    # guardados en BD (nunca se exige ni se sobrescriben con lo que traiga
    # el payload de un no-admin, y ya no hace falta que el payload los
    # incluya siquiera).
    if session.get("rol") == "admin":
        nombre = (data.get("nombre") or "").strip()
        codigo = (data.get("codigo") or "").strip()
        codigo_dali = (data.get("codigo_dali") or "").strip()
        if not nombre:
            return jsonify({"error": "Nombre requerido"}), 400
        if not codigo:
            return jsonify({"error": "El código SAP es obligatorio"}), 400
        # (2026-08-31) Mismo requisito que en create_proveedor — Víctor:
        # "tanto el codigo SAP como el DALI son obligatorios". Se aplica
        # también al editar (solo admin llega aquí) para que una ficha
        # antigua sin código DALI no se pueda volver a guardar sin
        # rellenarlo — de lo contrario el hueco se perpetúa indefinidamente.
        if not codigo_dali:
            return jsonify({"error": "El código DALI es obligatorio"}), 400
    else:
        _actual = query("SELECT nombre,codigo,codigo_dali FROM proveedores WHERE id=%s", (pid,), one=True)
        if not _actual:
            return jsonify({"error": "Proveedor no encontrado"}), 404
        nombre = _actual["nombre"]
        codigo = _actual["codigo"]
        codigo_dali = _actual["codigo_dali"]
    # Anti-duplicado: nombre en uso por otro proveedor
    dup_nombre = query(
        "SELECT id FROM proveedores WHERE activo=1 AND LOWER(nombre)=LOWER(%s) AND id!=%s", (nombre, pid)
    )
    if rows_to_list(dup_nombre):
        return jsonify({"error": f"Ya existe otro proveedor con el nombre '{nombre}'"}), 409
    # (2026-08-31) Mismos avisos detallados que en create_proveedor — ver
    # comentario extenso ahí y _buscar_proveedor_duplicado() (definida justo
    # antes de create_proveedor). Solo se comprueba cuando el admin está
    # editando codigo/codigo_dali de verdad (los conserva sin más si no es
    # admin, así que comparar consigo mismo con id!=pid ya lo cubre bien en
    # ambos casos).
    dup_codigo = _buscar_proveedor_duplicado("codigo", codigo, excluir_id=pid)
    if dup_codigo:
        return jsonify({"error": (
            f"El código SAP '{codigo}' ya está en uso por el proveedor «{dup_codigo['nombre']}» "
            f"(ID {dup_codigo['id']}) — corrige uno de los dos códigos."
        )}), 409
    dup_dali = _buscar_proveedor_duplicado("codigo_dali", codigo_dali, excluir_id=pid) if codigo_dali else None
    if dup_dali:
        return jsonify({"error": (
            f"El código DALI '{codigo_dali}' ya está en uso por el proveedor «{dup_dali['nombre']}» "
            f"(ID {dup_dali['id']}) — corrige uno de los dos códigos."
        )}), 409
    db   = get_db()
    # (2026-08-10) sujeto_seguimiento solo lo puede cambiar un admin —
    # compras puede seguir editando la ficha del proveedor con normalidad
    # (contactos, observaciones...), pero no marcar/desmarcar este campo.
    # Si no es admin, se conserva el valor que ya tuviera guardado en vez
    # de aceptar lo que traiga el payload (y sobre todo, en vez de
    # resetearlo a FALSE sin querer solo por editar otra cosa).
    if session.get("rol") == "admin":
        sujeto_seg = bool(data.get("sujeto_seguimiento", False))
    else:
        _actual2 = query("SELECT sujeto_seguimiento FROM proveedores WHERE id=%s", (pid,), one=True)
        sujeto_seg = bool(_actual2["sujeto_seguimiento"]) if _actual2 else False
    execute(
        "UPDATE proveedores SET codigo=%s,codigo_dali=%s,nombre=%s,observaciones=%s,sujeto_seguimiento=%s WHERE id=%s",
        (codigo, codigo_dali, nombre, data.get("observaciones",""), sujeto_seg, pid)
    )
    # Reemplazar contactos (el DELETE cascada también borra sus filas en
    # proveedor_contacto_hoteles vía ON DELETE CASCADE)
    execute("DELETE FROM proveedor_contactos WHERE proveedor_id=%s", (pid,))
    contactos = data.get("contactos", [])
    for i, c in enumerate(contactos):
        nombre_c    = (c.get("nombre") or "").strip() or None
        tel_c       = (c.get("telefono") or "").strip() or None
        movil_c     = (c.get("movil") or "").strip() or None
        email_c     = (c.get("email") or "").strip() or None
        principal_c = 1 if c.get("es_principal") else 0
        if nombre_c or tel_c or movil_c or email_c:
            cur_c = execute(
                "INSERT INTO proveedor_contactos (proveedor_id,nombre,telefono,movil,email,es_principal,orden) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                (pid, nombre_c, tel_c, movil_c, email_c, principal_c, i)
            )
            contacto_id = cur_c.fetchone()["id"]
            for hid in (c.get("hotel_ids") or []):
                execute(
                    "INSERT INTO proveedor_contacto_hoteles (contacto_id, hotel_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
                    (contacto_id, hid)
                )
    db.commit()
    return jsonify({"ok": True})

# (2026-08-10) Estado en memoria de los jobs de "Comparar listado PDF" —
# ver comparar_listado_pdf()/_ejecutar_comparacion_pdf_bg() más abajo para
# el porqué (evitar timeouts de proxy con PDFs grandes). Vive en memoria
# del propio proceso, no en BD — asumible porque son resultados de
# consulta efímeros (se limpian solos a los 30 min) y no algo que
# necesite sobrevivir a un reinicio del servidor.
_PDF_JOBS = {}
_PDF_JOBS_LOCK = threading.Lock()

def _normalizar_pedido_num(s):
    """
    (2026-08-06) Normaliza un Nº de pedido para comparar el listado de SAP
    contra lo que haya en la app: quita ceros a la izquierda y espacios, y
    pasa a mayúsculas — '00040159' y '40159' deben considerarse el mismo
    pedido aunque alguien lo haya tecleado en la app sin los ceros.
    """
    if not s:
        return ""
    s = str(s).strip().upper()
    m = re.match(r'^0*(\d+)$', s)
    return m.group(1) if m else s

def _normalizar_num_albaran(s):
    """
    (2026-08-19) Normaliza un número de albarán/registro DALI para poder
    comparar aunque tenga o no ceros a la izquierda ('81970' y '00081970'
    deben considerarse el MISMO albarán) — mismo criterio que
    _normalizar_pedido_num(), con nombre propio para dejar claro dónde se
    usa: al detectar si un albarán ya está registrado en el pedido
    (comparación "Comparar Pedidos + Albaranes", ver ya_registrado en
    _comparar_listado_albaranes_logica() y el guard de duplicados en
    _aplicar_coincidencia_albaran()). Antes de este fix se comparaba con
    un simple "in" de texto, que no reconocía '00081970' como el mismo
    albarán que '81970' ya registrado: se añadía una entrada duplicada
    con los ceros a la izquierda y, como esa nueva entrada tampoco casaba
    por texto exacto en la siguiente comparación, el pedido volvía a salir
    como pendiente indefinidamente.
    """
    return _normalizar_pedido_num(s)

def _normalizar_nombre_proveedor(s):
    """
    (2026-08-06) Normaliza el nombre de un proveedor para intentar
    emparejar el texto libre del listado de SAP ('LANDE CANARIAS SL') con
    el nombre guardado en el catálogo de proveedores de la app — quita
    acentos, puntuación y las formas societarias más comunes al final,
    para que pequeñas diferencias de formato ('Pastelería' vs
    'Pasteleria') no impidan el emparejamiento.
    """
    if not s:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", s.upper().strip())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[.,]', '', s)
    s = re.sub(r'\s+', ' ', s)
    for suf in (' SLL', ' SLU', ' SCOOP', ' SL', ' SA', ' CB'):
        if s.endswith(suf):
            s = s[: -len(suf)].strip()
            break
    return s

def _match_proveedor_catalogo(nombre_norm: str, cat_por_nombre: dict):
    """
    (2026-08-06, hoisted a nivel de módulo 2026-08-15 para reutilizarlo
    también en la comparación de listado de albaranes) Empareja un nombre
    de proveedor ya normalizado (_normalizar_nombre_proveedor) contra el
    catálogo de proveedores de la app — primero coincidencia exacta,
    luego parcial (el nombre de SAP/DALI a veces viene truncado o con
    alguna palabra de más/menos respecto al catálogo).

    cat_por_nombre: {nombre_normalizado: fila_proveedor} — construir con
    {_normalizar_nombre_proveedor(p["nombre"]): p for p in proveedores_cat}
    """
    if not nombre_norm:
        return None
    if nombre_norm in cat_por_nombre:
        return cat_por_nombre[nombre_norm]
    for nombre_cat, prov in cat_por_nombre.items():
        if nombre_cat and (nombre_cat in nombre_norm or nombre_norm in nombre_cat):
            return prov
    return None

def _normalizar_texto_generico(s: str) -> str:
    """
    (2026-08-15) Normaliza texto genérico (mayúsculas, sin acentos,
    espacios/saltos de línea colapsados a uno solo) — usado para
    identificar el departamento al principio del bloque "departamento +
    proveedor" del listado de albaranes de DALI (ver
    _match_departamento_prefijo). A diferencia de
    _normalizar_nombre_proveedor, no quita formas societarias (no aplica
    a nombres de departamento).
    """
    if not s:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", s.upper())
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _match_departamento_prefijo(blob: str, deptos_norm: list):
    """
    (2026-08-15) En el listado de albaranes de DALI, el nombre del
    departamento y el del proveedor quedan pegados sin separador fiable
    en el texto que extrae pypdf (p. ej. "COCINAPASTELERIA FLOYPE SLL" o,
    partido en varias líneas de ajuste de página, "BAR SALON\n(Discoteca,
    Princess)FRUCAPE S.L"). Como los nombres de departamento son un
    catálogo cerrado (tabla departamentos, igual que se hace con
    proveedores más abajo), se identifica buscando cuál de ellos es
    PREFIJO del bloque completo, en vez de intentar adivinar dónde corta
    con un separador — verificado 265/265 líneas contra un listado real.

    blob: texto crudo "departamento+proveedor", puede traer saltos de
    línea de ajuste de página.
    deptos_norm: [(id, nombre, nombre_normalizado), ...] de la tabla
    departamentos, ordenada de nombre normalizado MÁS LARGO a más corto
    (para que un nombre específico no quede enmascarado por otro más
    corto que sea prefijo suyo, p. ej. "BAR SALON" dentro de un nombre
    más largo que también empezara así).

    Devuelve (departamento_id_o_None, nombre_o_None, resto_normalizado)
    — "resto" es ya el texto del proveedor (normalizado), listo para
    _match_proveedor_catalogo(). Si no se identifica ningún departamento
    del catálogo, devuelve (None, None, blob_normalizado_completo) para
    intentar igualmente el proveedor contra el bloque entero — mejor un
    intento imperfecto que perder la fila entera por no reconocer el
    departamento.
    """
    blob_norm = _normalizar_texto_generico(blob)
    for dep_id, dep_nombre, dep_norm in deptos_norm:
        if dep_norm and blob_norm.startswith(dep_norm):
            return dep_id, dep_nombre, blob_norm[len(dep_norm):].strip()
    return None, None, blob_norm

@app.route("/api/pedidos/comparar-listado-pdf", methods=["POST"])
@login_required
def comparar_listado_pdf():
    """
    (2026-08-06) Verificación de listados PDF de SAP contra los pedidos ya
    registrados en la app — pensado para un repaso semanal por hotel.
    Restringido solo a rol admin (2026-08-10, a petición del usuario).

    (2026-08-10) FIX: con un PDF real de 178 páginas, el usuario recibía
    "Unexpected token '<' ... is not valid JSON" — el proceso completo
    (leer+extraer texto de las 178 páginas, ~8s solo en la extracción, más
    lo que tarde el cold-start del servicio gratuito de Render) tardaba
    más que el timeout de algún punto intermedio entre el navegador y el
    servidor (el proxy de Cloudflare Worker delante de la app), que
    devolvía su propia página de error HTML en vez de dejar pasar la
    respuesta JSON — de ahí el "<html>" en el mensaje de error.
    Solucionado haciendo el endpoint asíncrono: esta petición SOLO valida
    y arranca el trabajo pesado en un hilo aparte, devolviendo
    inmediatamente un job_id (habrá respondido en milisegundos, muy por
    debajo de cualquier timeout). El resultado real se consulta aparte,
    vía polling, con GET /api/pedidos/comparar-listado-pdf/<job_id>.

    (2026-08-11) Se lee el "Listado de Pedidos" SIMPLIFICADO que exporta
    SAP (uno por hotel) — una tabla de una línea por pedido, sin el
    detalle de artículos del listado completo (mucho más ligero: unas
    pocas páginas en vez de cientos). Cada línea tiene el formato fijo:
    "NNNNNNNN Pedido DD/MM/AAAA HH:MM:SS DD/MM/AAAA DD/MM/AAAA PROVEEDOR
    IMPORTE_BASE IMPORTE_RECIBIDO IMPORTE_PENDIENTE Abierto|Cerrado ..."
    es decir: Nº pedido, fecha y hora de realización, fecha de pedido,
    fecha de entrega indicada, proveedor, importe (base imponible),
    importe recibido, un importe pendiente que no se usa aquí, y el
    estado del pedido en SAP. Se extraen todos los números de pedido del
    PDF (con pypdf, sin necesidad de ningún binario del sistema — más
    portable que pdftotext/poppler en un despliegue de Render estándar) y
    se comparan contra pedido_num en esta app para ese hotel, para
    detectar compras que se hicieron en SAP pero nunca se dieron de alta
    aquí para su seguimiento.

    Además, comparando el importe (base imponible) contra el importe
    recibido de cada línea se deduce el estado real de entrega —
    "No entregado" si lo recibido es cero, "Entregado" si coincide con el
    importe del pedido, y "Entrega parcial" en cualquier otro caso — sin
    necesidad de abrir el listado completo con el detalle de artículos.

    El filtro de proveedores "sujeto_seguimiento" (Admin → Proveedores)
    es opt-in (2026-08-10): por defecto NINGÚN proveedor está sujeto a
    seguimiento hasta que un admin lo marque explícitamente en su ficha
    — solo entonces sus pedidos entran en la comparación; el resto ni
    cuentan como encontrados ni como no encontrados, simplemente no se
    evalúan. Hasta que se marquen los proveedores que interesan, esta
    comparación devolverá pocos o ningún pedido evaluado — es el
    comportamiento esperado, no un fallo.

    POST /api/pedidos/comparar-listado-pdf
    form-data: hotel_id, file (el PDF)
    → 202 {"ok": true, "job_id": "..."}
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403

    hotel_id_raw = request.form.get("hotel_id")
    if not hotel_id_raw:
        return jsonify({"error": "Falta indicar el hotel"}), 400
    try:
        hotel_id = int(hotel_id_raw)
    except ValueError:
        return jsonify({"error": "Hotel no válido"}), 400
    if "file" not in request.files:
        return jsonify({"error": "No se ha adjuntado ningún archivo"}), 400
    archivo = request.files["file"]
    if not archivo.filename or not archivo.filename.lower().endswith(".pdf"):
        return jsonify({"error": "El archivo debe ser un PDF"}), 400

    pdf_bytes = archivo.read()
    if not pdf_bytes:
        return jsonify({"error": "El archivo está vacío"}), 400

    import time as _time_pdf
    job_id = secrets.token_hex(16)
    with _PDF_JOBS_LOCK:
        # Limpieza de jobs viejos (>30 min) para no acumular memoria indefinidamente
        limite = _time_pdf.time() - 1800
        for jid in [j for j, v in _PDF_JOBS.items() if v.get("creado_en", 0) < limite]:
            del _PDF_JOBS[jid]
        _PDF_JOBS[job_id] = {
            "status": "processing", "creado_en": _time_pdf.time(),
            "hotel_id": hotel_id, "usuario_id": session.get("user_id"),
        }

    hilo = threading.Thread(
        target=_ejecutar_comparacion_pdf_bg,
        args=(job_id, hotel_id, pdf_bytes),
        daemon=True,
    )
    hilo.start()
    return jsonify({"ok": True, "job_id": job_id}), 202

@app.route("/api/pedidos/comparar-listado-pdf/<job_id>", methods=["GET"])
@login_required
def comparar_listado_pdf_estado(job_id):
    """Consulta del resultado de un job lanzado por comparar_listado_pdf()."""
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403
    with _PDF_JOBS_LOCK:
        job = dict(_PDF_JOBS.get(job_id) or {})
    if not job:
        return jsonify({"error": "El job no existe o ha caducado — vuelve a subir el PDF"}), 404
    job.pop("creado_en", None)
    return jsonify(job)

@app.route("/api/pedidos/comparar-listado-pdf/<job_id>/enviar-resumen", methods=["POST"])
@login_required
def comparar_listado_pdf_enviar_resumen(job_id):
    """
    (2026-08-10) A petición del usuario: envía un correo interno con el
    resumen de pedidos detectados en el listado de SAP/DALI que NO están
    dados de alta en Control de Pedidos, al comprador responsable del
    hotel, con copia al administrador que hizo la consulta.

    Acción explícita (botón "Enviar resumen por correo" en el resultado),
    no automática — para no reenviar sin querer cada vez que un admin
    vuelve a comparar el mismo listado mientras revisa el resultado.

    Como el resto de correos de esta app, se encola vía
    _encolar_email_sistema() — el envío real depende de que alguien
    tenga la aplicación abierta en el navegador (arquitectura ya
    establecida, sin SMTP propio en el servidor).
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403

    with _PDF_JOBS_LOCK:
        job = dict(_PDF_JOBS.get(job_id) or {})
    if not job:
        return jsonify({"error": "El job no existe o ha caducado — vuelve a subir el PDF"}), 404
    if job.get("status") != "done":
        return jsonify({"error": "La comparación todavía no ha terminado"}), 400

    resultado = job.get("resultado") or {}
    pedidos_no_encontrados = [p for p in resultado.get("pedidos", []) if not p.get("encontrado")]
    # (2026-08-11) A petición del usuario: el correo al comprador solo debe
    # listar pedidos cuyo proveedor se ha identificado con certeza contra
    # el catálogo — un pedido con proveedor NO identificado (nombre
    # truncado o distinto en SAP respecto al catálogo, `proveedor_pdf` tal
    # cual vino del PDF) es menos fiable y podría confundir al comprador;
    # esos quedan solo para revisión visual del admin en la propia
    # pantalla (con su aviso ⚠️), no se envían por correo.
    pedidos_faltantes = [p for p in pedidos_no_encontrados if p.get("proveedor_identificado")]
    no_identificados = len(pedidos_no_encontrados) - len(pedidos_faltantes)
    if not pedidos_faltantes:
        aviso_no_id = (
            f" ({no_identificados} pendiente(s) más de proveedor no identificado — "
            "revísalos en pantalla)" if no_identificados else ""
        )
        return jsonify({"error": f"No hay pedidos con proveedor identificado que reportar — no se envía nada{aviso_no_id}"}), 400

    hotel = query("SELECT codigo, nombre FROM hoteles WHERE id=%s", (job.get("hotel_id"),), one=True)
    if not hotel:
        return jsonify({"error": "Hotel no encontrado"}), 404

    compradores = _get_compradores_hotel(hotel["codigo"])
    destinatarios = [c["email"] for c in compradores if c.get("email")]
    if not destinatarios:
        return jsonify({
            "error": f"No hay ningún comprador con email asignado al hotel {hotel['codigo']} "
                     "— asígnalo en Admin → Usuarios → Hoteles asignados (Compras)"
        }), 400

    admin_row = query("SELECT nombre, email FROM usuarios WHERE id=%s", (job.get("usuario_id"),), one=True) or {}
    admin_nombre = admin_row.get("nombre") or session.get("nombre") or "Administrador"
    admin_email  = admin_row.get("email")

    subject, body = _email_resumen_pdf_sap(
        hotel["nombre"], hotel["codigo"], pedidos_faltantes,
        resultado.get("total_pdf", 0), resultado.get("excluidos_seguimiento", 0), admin_nombre,
        no_identificados,
    )
    _encolar_email_sistema(
        "resumen_listado_pdf_sap", destinatarios, subject, cuerpo_html=body,
        cc_emails=[admin_email] if admin_email else None,
    )
    return jsonify({"ok": True, "destinatarios": destinatarios, "cc": admin_email})

def _email_resumen_pdf_sap(hotel_nombre: str, hotel_codigo: str, pedidos_faltantes: list,
                            total_pdf: int, excluidos: int, admin_nombre: str,
                            no_identificados: int = 0) -> tuple:
    """
    (2026-08-10) Resumen de "Comparar listado PDF" — pedidos que figuran
    en el listado de SAP/DALI pero no están dados de alta en Control de
    Pedidos. Mismo patrón visual que el resto de emails internos de la
    app (_email_header_html + tabla + pie).

    (2026-08-11) `pedidos_faltantes` ya viene filtrado por el llamador
    para incluir solo proveedor identificado — `no_identificados` es
    solo el recuento de los que se quedaron fuera por eso, para que el
    correo avise de que hay más pendientes de revisar en pantalla, sin
    listarlos (no son fiables al 100%: el nombre del proveedor no se
    pudo emparejar con el catálogo).
    """
    subject = f"[Control de Pedidos] {len(pedidos_faltantes)} pedido(s) de {hotel_codigo} en SAP sin registrar en la app"

    # Con un listado grande podrían ser cientos de filas — se acota la
    # tabla del correo para que no sea kilométrica; el detalle completo
    # ya se ha visto en pantalla al comparar.
    LIMITE_FILAS = 100
    mostrar = pedidos_faltantes[:LIMITE_FILAS]
    resto = len(pedidos_faltantes) - len(mostrar)

    def _color_estado_aparente(estado):
        if estado == "SIN ENTREGAR":
            return "#8B0000"
        if estado == "ENTREGA PARCIAL":
            return "#856404"
        return "#155724"

    filas = "".join(f"""
        <tr style="{'background:#f5f5f5' if i % 2 else ''}">
          <td style="padding:8px 12px;border:1px solid #ddd">{p['pedido_num_sap']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p['proveedor_pdf']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p.get('fecha_pedido') or p.get('fecha', '')}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p.get('entrega_estado', '')}</td>
          <td style="padding:8px 12px;border:1px solid #ddd;color:{_color_estado_aparente(p.get('estado_aparente', ''))};font-weight:700">
            {p.get('estado_aparente', '')}
          </td>
        </tr>""" for i, p in enumerate(mostrar))

    aviso_resto = (
        f'<p style="font-size:12px;color:#888;font-style:italic">'
        f'…y {resto} pedido(s) más — consulta el listado completo en la aplicación.</p>'
        if resto > 0 else ""
    )

    body = f"""
    <div style="font-family:Arial,sans-serif;max-width:650px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Control de Pedidos — Aviso interno",
                            color_fondo="#1a3a6b", color_subtitulo="#a8c0e8")}
      <div style="padding:24px">
        <p>Se ha comparado el listado de pedidos exportado de SAP/DALI para el hotel
           <strong>{hotel_nombre}</strong> ({hotel_codigo}) contra los pedidos ya
           registrados en Control de Pedidos.</p>
        <p>Se han detectado <strong>{len(pedidos_faltantes)}</strong> pedido(s) que
           figuran en SAP/DALI pero <strong>no están dados de alta en la aplicación</strong>
           todavía, de un total de {total_pdf} pedidos en el listado
           ({excluidos} de proveedores no sujetos a seguimiento, no evaluados).</p>
        <table style="width:100%;border-collapse:collapse;margin:16px 0;font-size:13px">
          <tr style="background:#1a3a6b;color:#fff">
            <th style="padding:8px 12px;text-align:left">Nº Pedido SAP</th>
            <th style="padding:8px 12px;text-align:left">Proveedor</th>
            <th style="padding:8px 12px;text-align:left">Fecha</th>
            <th style="padding:8px 12px;text-align:left">Entrega</th>
            <th style="padding:8px 12px;text-align:left">Estado aparente</th>
          </tr>
          {filas}
        </table>
        <p style="font-size:12px;color:#888;font-style:italic">"Estado aparente" se calcula
           directamente sobre el importe recibido y el importe pendiente que trae el listado de
           SAP: <strong>SIN ENTREGAR</strong> si todavía no se ha recibido nada,
           <strong>ENTREGA PARCIAL</strong> si ya se ha recibido algo pero queda importe
           pendiente, y <strong>ENTREGA COMPLETA</strong> si no queda nada pendiente — es una
           lectura automática, no una verificación: pendiente de confirmación final por el
           comprador y el hotel.</p>
        {aviso_resto}
        {f'<p style="font-size:12px;color:#856404;background:#fff3cd;padding:8px 12px;border-radius:4px">'
          f'⚠️ Hay además <strong>{no_identificados}</strong> pedido(s) sin dar de alta cuyo proveedor '
          f'no se ha podido identificar con certeza en el catálogo — no se incluyen aquí por fiabilidad; '
          f'revísalos en pantalla, en "Comparar listado PDF" (mostrar solo los que faltan).</p>'
          if no_identificados else ''}
        <p>Por favor, revise estos pedidos y dé de alta en Control de Pedidos los que
           corresponda, para que queden dentro del seguimiento habitual.</p>
        <p style="font-size:12px;color:#888">Consulta generada por {admin_nombre} desde
           "Comparar listado PDF".</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Mensaje automático generado por el sistema de Control de Pedidos.<br>
           <strong>Princess Hotels &amp; Resorts</strong></p>
      </div>
    </div>
    """
    return subject, body

def _ejecutar_comparacion_pdf_bg(job_id, hotel_id, pdf_bytes):
    """
    (2026-08-10) Cuerpo real de la comparación — corre en un hilo aparte,
    fuera del ciclo petición/respuesta original, por eso necesita su
    propio contexto de aplicación (with app.app_context()) para poder
    usar query()/get_db(), que dependen de Flask g (con ámbito de
    petición, no accesible desde un hilo nuevo sin esto). Mismo patrón
    usado en init_db.py para lo mismo.
    """
    import time as _time_pdf
    with app.app_context():
        try:
            resultado = _comparar_listado_pdf_logica(hotel_id, pdf_bytes)
            with _PDF_JOBS_LOCK:
                if job_id in _PDF_JOBS:
                    _PDF_JOBS[job_id] = {**_PDF_JOBS[job_id], "status": "done", "resultado": resultado}
        except Exception as exc:
            log.error("[COMPARAR-PDF] Error en job %s: %s", job_id, exc)
            with _PDF_JOBS_LOCK:
                if job_id in _PDF_JOBS:
                    _PDF_JOBS[job_id] = {**_PDF_JOBS[job_id], "status": "error", "error": str(exc)}

def _parse_importe_es(s: str):
    """
    (2026-08-11) Convierte un importe con formato español ('2.852,10',
    '0,00', '-401,84') a float. Los listados de SAP siempre traen dos
    decimales y '.' como separador de miles, ',' como decimal.
    """
    if s is None:
        return 0.0
    s = str(s).strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def _entrega_estado(importe_base: float, importe_recibido: float) -> str:
    """
    (2026-08-11) Deriva el estado real de entrega de un pedido a partir
    de los importes del listado simplificado de SAP — columna 6 (base
    imponible) vs. columna 7 (importe recibido).

    (2026-08-14) Ajuste a petición del usuario — antes "Entregado" exigía
    columna 7 == columna 6 exacto; ahora es columna 7 >= columna 6 (cubre
    el caso de que el importe recibido informado en SAP supere ligeramente
    a la base, p. ej. por ajustes/recargos, que antes caía por error en
    "Entrega parcial"):
      - columna 7 == 0             -> "No entregado"
      - columna 7 >= columna 6     -> "Entregado"
      - 0 < columna 7 < columna 6  -> "Entrega parcial"
    Se compara redondeando a 2 decimales para evitar falsos negativos
    por errores de coma flotante. Un importe recibido negativo (dato
    anómalo) se trata también como "No entregado", por seguridad.
    """
    recibido = round(importe_recibido, 2)
    base = round(importe_base, 2)
    if recibido <= 0:
        return "No entregado"
    if recibido >= base:
        return "Entregado"
    return "Entrega parcial"

# (2026-08-19) Orden de "avance" de los estados de entrega de un pedido —
# única fuente de verdad, usada tanto por _aplicar_coincidencia_albaran()
# (para no retroceder el estado de un pedido) como por
# _comparar_listado_albaranes_logica() (para no proponer/mostrar como
# pendiente una "entrega parcial" según SAP cuando el pedido ya está en un
# estado más avanzado en la app — ver nota en esa función, caso pedido
# 42644: SAP seguía marcando un importe pendiente pequeño mientras el
# pedido ya se había dado por ENTREGADO en la app, y la comparativa lo
# volvía a mostrar como "ENTREGADO → ENTREGA PARCIAL" en cada comparación,
# como si aplicar la coincidencia fuera a retroceder el pedido — cuando en
# realidad _aplicar_coincidencia_albaran() ya lo protegía y no hacía nada).
_ORDEN_ENTREGA_ESTADOS = {"ENVIADO AL PROVEEDOR": 0, "PENDIENTE COTIZACIÓN": 0, "ENTREGA PARCIAL": 1, "ENTREGADO": 2}

def _estado_aparente_entrega(importe_recibido: float, importe_pendiente: float) -> str:
    """
    (2026-08-11) "Estado aparente" de entrega, a petición del usuario —
    distinto de `_entrega_estado()` (que compara base vs. recibido): este
    se calcula directamente sobre columnas "en bruto" del listado
    simplificado de SAP (importe recibido e importe pendiente, tal cual
    las trae SAP — el pendiente no siempre coincide con "base - recibido"
    calculado a mano, por eso se usa como dato aparte en vez de sustituir
    al anterior):
      - pendiente <= 0                 -> "ENTREGA COMPLETA"
      - pendiente > 0 y recibido <= 0  -> "SIN ENTREGAR"    (nada recibido todavía)
      - pendiente > 0 y recibido > 0   -> "ENTREGA PARCIAL" (ha llegado algo, falta el resto)

    (2026-08-14) Ajuste a petición del usuario: antes cualquier pedido
    con algo pendiente salía como "ENTREGA PARCIAL", incluso si no se
    había recibido NADA todavía (recibido = 0) — confuso, porque
    "parcial" da a entender que ya llegó una parte cuando en realidad no
    ha llegado nada. Ahora ese caso se distingue como "SIN ENTREGAR", y
    "ENTREGA PARCIAL" queda reservado a cuando de verdad se ha recibido
    algo pero todavía queda importe pendiente.

    Se llama "aparente" a propósito: es una lectura directa del PDF, no
    una verificación — pensado para que el comprador y el hotel lo
    revisen y confirmen, no como dato definitivo por sí solo.
    """
    pendiente = round(importe_pendiente, 2)
    recibido = round(importe_recibido, 2)
    if pendiente <= 0:
        return "ENTREGA COMPLETA"
    if recibido <= 0:
        return "SIN ENTREGAR"
    return "ENTREGA PARCIAL"

# (2026-08-11) Listado de Pedidos SIMPLIFICADO de SAP: una línea por
# pedido, sin el detalle de artículos. Aunque en el PDF renderizado las
# columnas se ven en el orden "Nº pedido, fecha/hora, fecha pedido,
# fecha entrega, proveedor, base, recibido, pendiente, estado", el texto
# que realmente devuelve pypdf (extract_text, por flujo del PDF, no por
# posición visual) viene en OTRO orden, verificado contra un listado
# real de 221 pedidos:
# "NNNNNNNN Pedido DD/MM/AAAA HH:MM:SS IMPORTE_BASE PROVEEDOR
#  DD/MM/AAAA(pedido) DD/MM/AAAA(entrega) Abierto|Cerrado IMPORTE_RECIBIDO IMPORTE_PENDIENTE ..."
#
# (2026-08-11) FIX: entre ciertos pares de columnas contiguas (importe
# base→proveedor, proveedor→fecha de pedido, estado→importe recibido)
# el PDF no tiene un espacio real entre celdas — solo separación visual
# por posición X. pypdf 3.x rellenaba ese hueco con un espacio al
# extraer el texto; pypdf ≥4 (lo que instala este proyecto, sin techo de
# versión en requirements.txt: "pypdf>=4.0") ya NO lo hace, así que el
# texto sale pegado ("2.852,10PILSA HOSTELERIA..."). Por eso los
# separadores van con \s* (cero o más) y no \s+ — sigue funcionando
# igual si hay espacio, y no rompe si no lo hay. Verificado contra el
# mismo listado real con pypdf 3.17.4 y con pypdf 6.15.0 (221/221 en
# ambos casos).
_NUM_ES = r'-?\d{1,3}(?:\.\d{3})*,\d{2}'
_PATRON_LISTADO_SIMPLIFICADO = re.compile(
    r'(\d{6,})\s*Pedido\s*'                       # Nº pedido
    r'(\d{2}/\d{2}/\d{4})\s*(\d{1,2}:\d{2}:\d{2})\s*'  # fecha y hora de realización
    r'(' + _NUM_ES + r')\s*'                      # importe (base imponible)
    r'(.+?)\s*'                                   # proveedor
    r'(\d{2}/\d{2}/\d{4})\s*'                     # fecha de pedido
    r'(\d{2}/\d{2}/\d{4})\s*'                     # fecha de entrega indicada
    r'(Abierto|Cerrado)\s*'                       # estado en SAP
    r'(' + _NUM_ES + r')\s*'                      # importe recibido
    r'(' + _NUM_ES + r')'                         # importe pendiente (8ª columna — usada en _estado_aparente_entrega)
)

def _comparar_listado_pdf_logica(hotel_id: int, pdf_bytes: bytes) -> dict:
    """Lógica pura de extracción+comparación — separada de la vista Flask
    para poder llamarla igual desde una petición normal o desde un hilo
    en segundo plano. Lanza excepción con el mensaje de error si algo
    falla (el PDF no es legible, no se reconoce ningún pedido, etc.).

    (2026-08-27) ÚNICA excepción a la filosofía "solo lectura, nunca
    escribe sola" del resto de "Comparar listado PDF" (ver
    _comparar_listado_albaranes_logica): de paso, para cada pedido
    localizado en la app, guarda tres cosas sin pedir confirmación, a
    petición explícita de Víctor — ninguna de las tres dispara
    notificaciones ni cambia el estado del pedido, por eso no hay nada
    que el usuario tenga que revisar antes de aplicar (a diferencia de un
    cambio de estado o un nuevo albarán):
      1. El importe base del PDF (6ª columna) en `pedidos.total_pedido`.
      2. La base imponible de la ÚLTIMA entrada/albarán ya registrado en
         el pedido (si tiene alguna), en el 3er segmento de
         `entrada_albaran_num` — calculada como el importe recibido
         acumulado del PDF (7ª columna) menos la suma de las bases
         imponibles de las entradas anteriores del mismo pedido.
      3. (2026-08-28) `pedidos.fecha_tramitacion`, con la "fecha de
         pedido" del PDF (misma columna que ya se usa como
         "fecha_pedido" más abajo) — SOLO si el pedido todavía no tiene
         ninguna fecha de tramitación guardada. A diferencia de los dos
         puntos anteriores (que se sobrescriben si el valor calculado
         cambia), esta NUNCA se sobrescribe una vez tiene un valor,
         porque a diferencia de Total Pedido/base imponible (que son
         puramente automáticos desde que existe el PDF oficial, v12.30.42
         y v12.30.44) la fecha de tramitación sigue siendo un campo
         normal, editable a mano en cualquier momento — mismo criterio
         que ya se usa al leer el PDF oficial individual de "Nº Pedido"
         (ver _procesarFechasPdfPedidoOficial en templates/index.html):
         solo se rellena si está vacía, nunca se pregunta ni se pisa un
         valor ya introducido, porque aquí no hay ningún usuario delante
         para preguntarle cuál de las dos fechas es la correcta — es una
         comparación de listado, en background. Petición de Víctor:
         extender a esta comparación masiva el mismo auto-relleno de
         Fecha tramitación ya implementado para el PDF oficial individual
         "para los pedidos antiguos que nunca tuvieron el PDF oficial
         individual adjuntado".
    """
    try:
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(pdf_bytes))
        texto = ""
        for pagina in reader.pages:
            texto += (pagina.extract_text() or "") + "\n"
    except Exception as exc:
        log.error("[COMPARAR-PDF] Error leyendo el PDF: %s", exc)
        raise RuntimeError(f"No se pudo leer el PDF: {exc}")

    encontrados_pdf = _PATRON_LISTADO_SIMPLIFICADO.findall(texto)
    if not encontrados_pdf:
        raise RuntimeError(
            "No se ha reconocido ningún pedido en el PDF — "
            "¿es el \"Listado de Pedidos\" simplificado de SAP, con el formato habitual "
            "(Nº pedido, fechas, proveedor, importes y estado en una sola línea por pedido)?"
        )

    # ── Catálogo de proveedores, para el filtro de seguimiento ───────────────
    proveedores_cat = rows_to_list(query(
        "SELECT id, nombre, sujeto_seguimiento FROM proveedores WHERE activo=1"
    ))
    cat_por_nombre = {_normalizar_nombre_proveedor(p["nombre"]): p for p in proveedores_cat if p["nombre"]}

    # (2026-08-15) _match_proveedor() vivía aquí como función anidada;
    # ahora es _match_proveedor_catalogo() a nivel de módulo, para poder
    # reutilizarla también en _comparar_listado_albaranes_logica() sin
    # duplicar la lógica de emparejamiento.
    def _match_proveedor(nombre_norm):
        return _match_proveedor_catalogo(nombre_norm, cat_por_nombre)

    # ── Pedidos ya registrados en la app para este hotel ──────────────────────
    pedidos_app = rows_to_list(query(
        "SELECT id, norden, pedido_num, estado, total_pedido, entrada_albaran_num, fecha_tramitacion FROM pedidos "
        "WHERE hotel_id=%s AND pedido_num IS NOT NULL AND pedido_num != ''",
        (hotel_id,)
    ))
    app_por_num = {}
    for p in pedidos_app:
        app_por_num.setdefault(_normalizar_pedido_num(p["pedido_num"]), p)

    # (2026-08-27) Total Pedido real — petición de Víctor: al comparar el
    # listado de SAP, el importe base (6ª columna del PDF) se guarda solo
    # como `pedidos.total_pedido` para cada pedido localizado, sin ningún
    # paso de confirmación (a diferencia del resto de "Comparar listado
    # PDF", que solo propone) — es un campo puramente informativo, no
    # dispara notificaciones ni cambia estado, así que no hay nada que
    # confirmar. Solo se escribe cuando el valor cambia de verdad, para no
    # generar escrituras de más en cada comparación repetida del mismo PDF.
    _total_pedido_actualizados = []

    # (2026-08-27) Base imponible de la ENTRADA (albarán) — segunda parte
    # de la misma petición: "del mismo modo" que Total Pedido, al comparar
    # este PDF se rellena sola la base imponible del ALBARÁN más reciente
    # ya registrado en el pedido (celda nueva junto a cada entrada, ver
    # templates/index.html). SAP solo da un importe recibido ACUMULADO
    # (columna 7) por pedido, no desglosado por albarán, así que el valor
    # que corresponde a la última entrada = columna 7 menos la suma de las
    # bases imponibles YA registradas en las entradas anteriores de ese
    # mismo pedido (las entregas parciales previas) — si no hay ninguna
    # anterior, el valor es directamente la columna 7 completa. Se aplica
    # igual tanto si SAP marca el pedido como "Entrega parcial" como
    # "Entregado" (a petición explícita de Víctor: en ambos casos se usa
    # la columna 7, que en una entrega total coincide con el total del
    # pedido). Solo se toca si hay AL MENOS UNA entrada ya creada en el
    # pedido — no se inventa ninguna entrada nueva aquí, esto solo rellena
    # la base imponible de la que ya exista más reciente.
    _base_imponible_entrada_actualizados = []

    # (2026-08-28) Fecha tramitación — tercera escritura silenciosa, a
    # petición de Víctor: "para los pedidos antiguos que nunca tuvieron el
    # PDF oficial individual adjuntado" (ver nota más arriba en el
    # docstring). Se usa la "fecha de pedido" del PDF (misma columna que
    # ya alimenta `fecha_pedido` más abajo, en el resultado por fila) —
    # SOLO cuando el pedido no tiene AÚN ninguna fecha de tramitación
    # guardada. Nunca se sobrescribe un valor ya existente (a diferencia
    # de Total Pedido/base imponible arriba), porque este campo sigue
    # siendo editable a mano y aquí no hay nadie a quien preguntar cuál de
    # las dos fechas es la correcta si hubiera un conflicto.
    _fecha_tramitacion_actualizados = []

    resultado = []
    vistos = set()
    for (num_sap, fecha_hora_fecha, fecha_hora_hora, importe_base_txt, proveedor_raw,
         fecha_pedido, fecha_entrega, estado_sap, importe_recibido_txt,
         importe_pendiente_txt) in encontrados_pdf:
        if num_sap in vistos:
            continue
        vistos.add(num_sap)

        nombre_prov = proveedor_raw.strip()
        prov_match = _match_proveedor(_normalizar_nombre_proveedor(nombre_prov))

        if prov_match and not prov_match["sujeto_seguimiento"]:
            continue  # proveedor excluido a propósito (p.ej. alimentación/bebida)

        importe_base      = _parse_importe_es(importe_base_txt)
        importe_recibido  = _parse_importe_es(importe_recibido_txt)
        importe_pendiente = _parse_importe_es(importe_pendiente_txt)

        pedido_app = app_por_num.get(_normalizar_pedido_num(num_sap))

        if pedido_app and importe_base is not None:
            _total_actual = pedido_app.get("total_pedido")
            _total_actual_f = float(_total_actual) if _total_actual is not None else None
            if _total_actual_f is None or round(_total_actual_f, 2) != round(importe_base, 2):
                _total_pedido_actualizados.append((pedido_app["id"], importe_base))

        if pedido_app and importe_recibido is not None:
            _entradas_pedido = _parse_albaran_entries(pedido_app.get("entrada_albaran_num"))
            if _entradas_pedido:
                _suma_anteriores = sum(
                    e["base_imponible"] for e in _entradas_pedido[:-1] if e.get("base_imponible") is not None
                )
                _valor_ultima = round(importe_recibido - _suma_anteriores, 2)
                if _valor_ultima >= 0:
                    _ultima_actual = _entradas_pedido[-1].get("base_imponible")
                    if _ultima_actual is None or round(_ultima_actual, 2) != _valor_ultima:
                        _entradas_pedido[-1]["base_imponible"] = _valor_ultima
                        _base_imponible_entrada_actualizados.append((
                            pedido_app["id"], _construir_entrada_albaran_num(_entradas_pedido)
                        ))
                # Si sale negativo (datos ya inconsistentes — p.ej. entradas
                # anteriores con una base imponible mayor de lo que SAP
                # acumula ahora), no se escribe nada: mejor dejarlo en blanco
                # para revisión manual que guardar un importe sin sentido.

        if pedido_app and not pedido_app.get("fecha_tramitacion") and fecha_pedido:
            _fecha_tramitacion_iso = _parsear_fecha_es_a_iso(fecha_pedido)
            if _fecha_tramitacion_iso:
                _fecha_tramitacion_actualizados.append((pedido_app["id"], _fecha_tramitacion_iso))

        resultado.append({
            "pedido_num_sap":         num_sap,
            "fecha":                  fecha_pedido,
            "hora":                   fecha_hora_hora,
            "fecha_realizacion":      fecha_hora_fecha,
            "fecha_pedido":           fecha_pedido,
            "fecha_entrega":          fecha_entrega,
            "proveedor_pdf":          nombre_prov,
            "proveedor_id":           prov_match["id"] if prov_match else None,
            "proveedor_identificado": bool(prov_match),
            "importe_base":           importe_base,
            "importe_recibido":       importe_recibido,
            "importe_pendiente":      importe_pendiente,
            "estado_sap":             estado_sap,
            "entrega_estado":         _entrega_estado(importe_base, importe_recibido),
            "estado_aparente":        _estado_aparente_entrega(importe_recibido, importe_pendiente),
            "encontrado":             bool(pedido_app),
            "pedido_id":              pedido_app["id"] if pedido_app else None,
            "norden":                 pedido_app["norden"] if pedido_app else None,
            "estado_app":             pedido_app["estado"] if pedido_app else None,
        })

    # Escritura del Total Pedido, de la base imponible de la última
    # entrada y de la fecha de tramitación — las tres silenciosas (sin
    # confirmación, ver notas arriba), cada una solo de las filas cuyo
    # valor realmente cambia (o, en el caso de fecha_tramitacion, solo de
    # las que estaban vacías).
    if _total_pedido_actualizados:
        for _pid_tp, _importe_tp in _total_pedido_actualizados:
            execute("UPDATE pedidos SET total_pedido=%s WHERE id=%s", (_importe_tp, _pid_tp))
        get_db().commit()
    if _base_imponible_entrada_actualizados:
        for _pid_be, _entrada_str in _base_imponible_entrada_actualizados:
            execute("UPDATE pedidos SET entrada_albaran_num=%s WHERE id=%s", (_entrada_str, _pid_be))
        get_db().commit()
    if _fecha_tramitacion_actualizados:
        for _pid_ft, _fecha_ft in _fecha_tramitacion_actualizados:
            execute("UPDATE pedidos SET fecha_tramitacion=%s WHERE id=%s", (_fecha_ft, _pid_ft))
        get_db().commit()

    total_evaluados = len(resultado)
    no_encontrados  = sum(1 for r in resultado if not r["encontrado"])
    no_entregados   = sum(1 for r in resultado if r["entrega_estado"] == "No entregado")
    parciales       = sum(1 for r in resultado if r["entrega_estado"] == "Entrega parcial")
    entregados      = sum(1 for r in resultado if r["entrega_estado"] == "Entregado")
    return {
        "ok": True,
        "total_pdf":             len(encontrados_pdf),
        "total_evaluados":       total_evaluados,
        "excluidos_seguimiento": len(encontrados_pdf) - total_evaluados,
        "encontrados":           total_evaluados - no_encontrados,
        "no_encontrados":        no_encontrados,
        "entregados":            entregados,
        "no_entregados":         no_entregados,
        "entregas_parciales":    parciales,
        "total_pedido_actualizados": len(_total_pedido_actualizados),
        "base_imponible_entrada_actualizados": len(_base_imponible_entrada_actualizados),
        "fecha_tramitacion_actualizados": len(_fecha_tramitacion_actualizados),
        "pedidos":               resultado,
    }

# (2026-08-15) Listado de Albaranes registrados en DALI — segundo PDF de
# "Comparar listado PDF" (ahora "Pedidos + Albaranes"), a petición del
# usuario: cruzar el listado de pedidos de SAP (arriba) con el listado de
# albaranes que se van registrando en DALI en base a esos pedidos, para
# proponer el registro automático de la entrega en Control de Pedidos.
#
# Verificado 265/265 líneas contra un listado real. Igual que con el
# listado de pedidos, el texto que extrae pypdf no sigue el orden visual
# de columnas sino el orden real del flujo del PDF:
#   Nº registro DALI (8 dígitos) · "Albarán: " · el mismo Nº de registro
#   otra vez · " - " · Nº de albarán del PROVEEDOR (texto libre: puede
#   llevar '/', '.', espacios, y partirse en varias líneas por ajuste de
#   página) · "(EURO)" · importe (con 2 O 4 decimales — SAP a veces trae
#   4, _parse_importe_es ya los soporta a ambos por igual) · fecha de
#   registro en DALI (DD/MM/AAAA) · hora (HH:MM:SS) · departamento del
#   hotel donde se registra la mercancía, pegado sin separador al nombre
#   del proveedor (ambos texto libre, cualquiera de los dos puede partirse
#   en 2-3 líneas) — ver _match_departamento_prefijo() para cómo se
#   separan sin necesitar un separador explícito.
_PATRON_LISTADO_ALBARANES = re.compile(
    r'(\d{8})\s*Albar[aá]n:\s*(\d{8})\s*-\s*'    # nº registro DALI (aparece dos veces)
    r'([^\(]+?)\s*'                               # nº de albarán del proveedor
    r'\(EURO\)\s*'
    r'(' + _NUM_ES + r'\d{0,2})\s*'               # importe (2 o 4 decimales)
    r'(\d{2}/\d{2}/\d{4})\s*'                     # fecha de registro en DALI
    r'(\d{1,2}:\d{2}:\d{2})\s*'                   # hora de registro
    r'(.+?)(?=\d{8}\s*Albar[aá]n:|$)',            # departamento + proveedor, pegados
    re.S
)

def _parsear_fecha_es_a_iso(fecha_ddmmyyyy):
    """'DD/MM/AAAA' -> 'AAAA-MM-DD' (ISO, formato de columna DATE), o None
    si no se puede parsear — misma idea que _fecha_es() pero a la inversa."""
    if not fecha_ddmmyyyy:
        return None
    try:
        return datetime.strptime(str(fecha_ddmmyyyy).strip(), "%d/%m/%Y").strftime("%Y-%m-%d")
    except Exception:
        return None

def _comparar_listado_albaranes_logica(hotel_id: int, pdf1_bytes: bytes, pdf2_bytes: bytes) -> dict:
    """
    (2026-08-15) Ampliación de "Comparar listado PDF" a petición del
    usuario (Víctor): además del "Listado de Pedidos" simplificado de SAP
    (PDF 1, igual que _comparar_listado_pdf_logica), lee un segundo PDF —
    el "Listado de Albaranes" registrados en DALI en base a esos pedidos —
    y cruza ambos para proponer el registro automático de la entrega
    (fecha de tramitación, número de entrada del albarán y estado) en los
    pedidos de esta app que ya están dados de alta.

    Criterio de coincidencia (decisión del usuario, preguntado
    explícitamente): mismo proveedor Y mismo importe entre el importe YA
    RECIBIDO de un pedido del PDF 1 que SAP ya muestra como "Entregado" o
    "Entrega parcial", y el importe de un albarán del PDF 2 — se comparó
    contra el importe recibido (no el importe base/total del pedido)
    porque en una entrega parcial el importe base puede ser mayor que lo
    que trae un albarán suelto y daría falsos negativos.

    Solo se evalúan proveedores marcados como sujetos a seguimiento (Admin
    → Proveedores) en AMBOS PDF — mismo filtro que ya usa
    _comparar_listado_pdf_logica, aplicado aquí también a los albaranes.

    Empates (mismo proveedor+importe con más de un candidato en cualquiera
    de los dos PDF): a petición del usuario, NINGUNA de las combinaciones
    posibles se autorregistra — todas quedan en "pendientes_ambiguos" para
    decisión manual, en vez de arriesgarse a emparejar mal.

    Esta función es de solo lectura — NO escribe nada en la base de
    datos, solo propone. La escritura real (fecha_tramitacion si estaba
    vacía, nuevo albarán en entrada_albaran_num, cambio de estado con su
    notificación) la hace _aplicar_coincidencia_albaran(), llamada desde
    el endpoint .../aplicar solo cuando el usuario confirma explícitamente
    qué coincidencias aplicar — decisión de diseño explícita: revisar y
    confirmar antes de aplicar, nunca automático al comparar.

    RELLENO AUTOMÁTICO DE BASE IMPONIBLE EN ENTRADAS YA REGISTRADAS
    (2026-08-28, a petición de Víctor, ampliado el mismo día — ver
    CHANGELOG v12.30.44): igual que _comparar_listado_pdf_logica() rellena
    sola la base imponible de la ÚLTIMA entrada de un pedido a partir del
    importe acumulado de SAP, aquí se hace un barrido de TODAS las
    entradas (parciales o totales) de TODOS los pedidos dados de alta de
    este hotel — no solo las que forman parte de alguna "coincidencia"
    propuesta arriba — buscando las que ya tienen un número de entrada
    (Nº Entrada DALI/SAP) pero les falta la base imponible. Petición
    literal: "cuando estas entradas parciales o totales ya estan
    registradas pero no se relleno la celda total sin igic, la
    aplicacion deberia comprobar si tiene o no valor esta celda y
    rellenarla en caso de que este vacia". Para cada una de esas
    entradas, si su número normalizado coincide con el de un albarán del
    "Listado de Albaranes" (PDF 2) recién subido, se rellena con el
    importe de ese albarán — solo ese campo (nunca el número de entrada,
    la fecha ni el estado, que siempre requieren confirmación explícita
    vía _aplicar_coincidencia_albaran()) y solo si estaba vacío, nunca se
    sobrescribe un valor ya introducido. Si el número de registro
    aparece más de una vez en el PDF 2 (caso raro, duplicado), esa fila
    se salta por seguridad — ante la duda, no se inventa. Esto sustituye
    y amplía la excepción anterior (limitada a las coincidencias ya "sin
    cambios pendientes"), que quedaba corta: una entrada antigua de un
    pedido que no aparece entre las coincidencias del PDF 1 recién
    subido (p. ej. porque SAP ya no lo lista como pendiente, o es de una
    entrega parcial anterior) tampoco se tocaba antes, y ahora sí.

    Devuelve dict con: coincidencias, pendientes_ambiguos,
    pendientes_sin_albaran (pedido Entregado/Parcial en SAP sin albarán
    DALI con ese importe), pendientes_sin_pedido (albarán DALI sin
    pareja de importe exacto en el PDF 1 recién subido — cada elemento
    puede llevar opcionalmente "posible_pedido_hint", ver nota
    2026-08-19 más abajo, con un pedido ya dado de alta en la app que
    podría corresponder a ese albarán, a falta de verificación manual),
    total_pdf1, total_pdf2, excluidos_pdf1, excluidos_pdf2 (proveedor no
    identificado o no sujeto a seguimiento, en cada PDF).

    Lanza RuntimeError si algún PDF no se puede leer o no se reconoce
    ningún registro en él.
    """
    import io
    from pypdf import PdfReader
    from collections import defaultdict

    def _leer_texto(pdf_bytes, etiqueta):
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            texto = ""
            for pagina in reader.pages:
                texto += (pagina.extract_text() or "") + "\n"
            return texto
        except Exception as exc:
            log.error("[COMPARAR-ALBARANES] Error leyendo PDF (%s): %s", etiqueta, exc)
            raise RuntimeError(f"No se pudo leer el PDF de {etiqueta}: {exc}")

    texto1 = _leer_texto(pdf1_bytes, "pedidos")
    texto2 = _leer_texto(pdf2_bytes, "albaranes")

    encontrados_pdf1 = _PATRON_LISTADO_SIMPLIFICADO.findall(texto1)
    if not encontrados_pdf1:
        raise RuntimeError(
            "No se ha reconocido ningún pedido en el primer PDF — "
            "¿es el \"Listado de Pedidos\" simplificado de SAP?"
        )
    encontrados_pdf2 = _PATRON_LISTADO_ALBARANES.findall(texto2)
    if not encontrados_pdf2:
        raise RuntimeError(
            "No se ha reconocido ningún albarán en el segundo PDF — "
            "¿es el \"Listado de Albaranes\" registrados en DALI?"
        )

    # ── Catálogos: proveedores (sujeto_seguimiento) y departamentos ──────────
    proveedores_cat = rows_to_list(query(
        "SELECT id, nombre, sujeto_seguimiento FROM proveedores WHERE activo=1"
    ))
    cat_por_nombre = {_normalizar_nombre_proveedor(p["nombre"]): p for p in proveedores_cat if p["nombre"]}

    deptos_cat = rows_to_list(query("SELECT id, nombre FROM departamentos WHERE activo=1"))
    deptos_norm = sorted(
        (
            (d["id"], d["nombre"], _normalizar_texto_generico(d["nombre"]))
            for d in deptos_cat if d["nombre"]
        ),
        key=lambda x: -len(x[2])
    )

    # ── Pedidos ya registrados en la app para este hotel ──────────────────────
    pedidos_app = rows_to_list(query(
        """SELECT id, norden, pedido_num, estado, fecha_tramitacion, entrada_albaran_num
           FROM pedidos WHERE hotel_id=%s AND pedido_num IS NOT NULL AND pedido_num != ''""",
        (hotel_id,)
    ))
    app_por_num = {}
    for p in pedidos_app:
        app_por_num.setdefault(_normalizar_pedido_num(p["pedido_num"]), p)

    ESTADOS_EXCLUIDOS_AUTO = {"CANCELADO", "DENEGADO POR DIRECCION GENERAL"}

    # ── PDF1: pedidos ya dados de alta, que SAP ya muestra Entregado/Parcial ──
    excluidos_pdf1 = 0
    candidatos_pdf1 = []
    vistos1 = set()
    for (num_sap, fecha_hora_fecha, fecha_hora_hora, importe_base_txt, proveedor_raw,
         fecha_pedido, fecha_entrega, estado_sap, importe_recibido_txt,
         importe_pendiente_txt) in encontrados_pdf1:
        if num_sap in vistos1:
            continue
        vistos1.add(num_sap)
        prov_match = _match_proveedor_catalogo(_normalizar_nombre_proveedor(proveedor_raw.strip()), cat_por_nombre)
        if not prov_match or not prov_match["sujeto_seguimiento"]:
            excluidos_pdf1 += 1
            continue
        importe_base     = _parse_importe_es(importe_base_txt)
        importe_recibido = _parse_importe_es(importe_recibido_txt)
        entrega_estado = _entrega_estado(importe_base, importe_recibido)
        if entrega_estado not in ("Entregado", "Entrega parcial"):
            continue
        pedido_app = app_por_num.get(_normalizar_pedido_num(num_sap))
        if not pedido_app or pedido_app["estado"] in ESTADOS_EXCLUIDOS_AUTO:
            continue  # sin pedido dado de alta, o en un estado que no se debe tocar automáticamente
        candidatos_pdf1.append({
            "pedido_num_sap":             num_sap,
            "pedido_id":                  pedido_app["id"],
            "norden":                     pedido_app["norden"],
            "estado_app_actual":          pedido_app["estado"],
            "fecha_tramitacion_actual":   pedido_app.get("fecha_tramitacion"),
            "entrada_albaran_num_actual": pedido_app.get("entrada_albaran_num"),
            "fecha_pedido_es":            fecha_pedido,
            "proveedor_id":               prov_match["id"],
            "proveedor_nombre":           prov_match["nombre"],
            "importe_base":               importe_base,
            "importe_recibido":           importe_recibido,
            "entrega_estado":             entrega_estado,
            "estado_objetivo":            "ENTREGADO" if entrega_estado == "Entregado" else "ENTREGA PARCIAL",
            "clave":                      (prov_match["id"], round(importe_recibido, 2)),
        })

    # ── PDF2: albaranes de proveedores sujetos a seguimiento ─────────────────
    excluidos_pdf2 = 0
    candidatos_pdf2 = []
    vistos2 = set()
    for (registro1, registro2, albaran_prov_raw, importe_txt, fecha_registro,
         hora_registro, resto_raw) in encontrados_pdf2:
        registro = registro1.strip()
        if registro in vistos2:
            continue
        vistos2.add(registro)
        dep_id, dep_nombre, resto_norm = _match_departamento_prefijo(resto_raw, deptos_norm)
        prov_match = _match_proveedor_catalogo(_normalizar_nombre_proveedor(resto_norm), cat_por_nombre)
        if not prov_match or not prov_match["sujeto_seguimiento"]:
            excluidos_pdf2 += 1
            continue
        importe = _parse_importe_es(importe_txt)
        candidatos_pdf2.append({
            "registro_dali":       registro,
            "albaran_proveedor":   re.sub(r'\s+', ' ', albaran_prov_raw.strip()),
            "proveedor_id":        prov_match["id"],
            "proveedor_nombre":    prov_match["nombre"],
            "proveedor_pdf":       resto_norm,
            "departamento_id":     dep_id,
            "departamento_nombre": dep_nombre,
            "fecha_registro_es":   fecha_registro,
            "hora_registro":       hora_registro,
            "importe":             importe,
            "clave":               (prov_match["id"], round(importe, 2)),
        })

    # ── Cruce por (proveedor_id, importe) ──────────────────────────────────────
    grupo1 = defaultdict(list)
    for c in candidatos_pdf1:
        grupo1[c["clave"]].append(c)
    grupo2 = defaultdict(list)
    for c in candidatos_pdf2:
        grupo2[c["clave"]].append(c)

    coincidencias          = []
    pendientes_ambiguos    = []
    pendientes_sin_albaran = []
    pendientes_sin_pedido  = []

    for clave in set(grupo1) | set(grupo2):
        lado1 = grupo1.get(clave, [])
        lado2 = grupo2.get(clave, [])
        if len(lado1) == 1 and len(lado2) == 1:
            p1, p2 = lado1[0], lado2[0]
            fecha_tram_iso     = _parsear_fecha_es_a_iso(p1["fecha_pedido_es"])
            fecha_registro_iso = _parsear_fecha_es_a_iso(p2["fecha_registro_es"])
            # (2026-08-19) Comparación por número normalizado (ignora ceros a
            # la izquierda) en vez de un simple "in" de texto — ver
            # _normalizar_num_albaran() para el motivo del cambio.
            _reg_dali_norm = _normalizar_num_albaran(p2["registro_dali"])
            _entradas_p1 = _parse_albaran_entries(p1["entrada_albaran_num_actual"])
            ya_registrado = bool(
                p2["registro_dali"] and p1["entrada_albaran_num_actual"]
                and any(
                    _normalizar_num_albaran(_e["num"]) == _reg_dali_norm
                    for _e in _entradas_p1
                )
            )
            # (2026-08-19) El pedido puede estar YA en un estado más avanzado
            # en la app que el que propone esta línea de SAP (p. ej. la app
            # ya lo tiene ENTREGADO, pero SAP todavía muestra un importe
            # pendiente pequeño y por tanto "Entrega parcial" según
            # _entrega_estado) — _aplicar_coincidencia_albaran() ya protege
            # este caso y NO retrocede el estado, pero antes de este fix la
            # comparativa lo seguía mostrando cada vez como
            # "ENTREGADO → ENTREGA PARCIAL", como si aplicar fuera a
            # retroceder el pedido, cuando en realidad no iba a cambiar
            # nada. Ver _ORDEN_ENTREGA_ESTADOS.
            _estado_ya_avanzado = (
                _ORDEN_ENTREGA_ESTADOS.get(p1["estado_app_actual"], 0)
                > _ORDEN_ENTREGA_ESTADOS.get(p1["estado_objetivo"], 0)
            )
            _sin_cambio_estado = p1["estado_app_actual"] == p1["estado_objetivo"] or _estado_ya_avanzado
            coincidencias.append({
                "clave":                      f'{clave[0]}_{str(clave[1]).replace(".", "_")}',
                "pedido_id":                  p1["pedido_id"],
                "pedido_num_sap":             p1["pedido_num_sap"],
                "norden":                     p1["norden"],
                "estado_app_actual":          p1["estado_app_actual"],
                "estado_objetivo":            p1["estado_objetivo"],
                "proveedor_nombre":           p1["proveedor_nombre"],
                "importe":                    clave[1],
                "fecha_tramitacion_actual":   p1["fecha_tramitacion_actual"],
                "fecha_tramitacion_pdf1_iso": fecha_tram_iso,
                "fecha_tramitacion_pdf1_es":  p1["fecha_pedido_es"],
                "registro_dali":              p2["registro_dali"],
                "albaran_proveedor":          p2["albaran_proveedor"],
                "fecha_registro_dali_iso":    fecha_registro_iso,
                "fecha_registro_dali_es":     p2["fecha_registro_es"],
                "departamento_nombre":        p2["departamento_nombre"],
                "ya_registrado":              ya_registrado,
                "ya_en_estado_objetivo":      p1["estado_app_actual"] == p1["estado_objetivo"],
                "estado_ya_avanzado":         _estado_ya_avanzado,
                "sin_cambios_pendientes":     ya_registrado and _sin_cambio_estado
                                               and p1["fecha_tramitacion_actual"],
            })
        elif len(lado1) == 1 and len(lado2) == 0:
            pendientes_sin_albaran.append(lado1[0])
        elif len(lado1) == 0 and len(lado2) == 1:
            pendientes_sin_pedido.append(lado2[0])
        else:
            pendientes_ambiguos.append({
                "clave":            f'{clave[0]}_{str(clave[1]).replace(".", "_")}',
                "proveedor_nombre": (lado1[0]["proveedor_nombre"] if lado1 else lado2[0]["proveedor_nombre"]),
                "importe":          clave[1],
                "pedidos":          lado1,
                "albaranes":        lado2,
            })

    # ── Relleno automático de Base imp. (€) en entradas ya registradas ──────
    # (2026-08-28) Ver nota en el docstring de esta función ("RELLENO
    # AUTOMÁTICO DE BASE IMPONIBLE EN ENTRADAS YA REGISTRADAS"). Barrido
    # de TODAS las entradas de TODOS los pedidos de este hotel (no solo
    # los que aparecen en `coincidencias` de arriba): si una entrada ya
    # tiene número (Nº Entrada DALI/SAP) pero le falta la base imponible,
    # y ese número coincide con un albarán del PDF 2 recién subido, se
    # rellena con el importe de ese albarán. Un número de registro
    # duplicado en el PDF 2 se descarta por seguridad (no se puede saber
    # cuál de los dos importes es el correcto).
    _pdf2_por_registro = {}
    _pdf2_registro_ambiguo = set()
    for _c2 in candidatos_pdf2:
        _rn2 = _normalizar_num_albaran(_c2["registro_dali"])
        if _rn2 in _pdf2_por_registro:
            _pdf2_registro_ambiguo.add(_rn2)
        else:
            _pdf2_por_registro[_rn2] = _c2

    _base_imponible_albaranes_actualizados = []
    for _p in pedidos_app:
        _entradas_pedido = _parse_albaran_entries(_p.get("entrada_albaran_num"))
        if not _entradas_pedido:
            continue
        _cambiado_base_imponible = False
        for _e in _entradas_pedido:
            if _e.get("base_imponible") is not None or not _e.get("num"):
                continue
            _rn = _normalizar_num_albaran(_e["num"])
            if _rn in _pdf2_registro_ambiguo or _rn not in _pdf2_por_registro:
                continue
            _e["base_imponible"] = round(float(_pdf2_por_registro[_rn]["importe"]), 2)
            _cambiado_base_imponible = True
        if _cambiado_base_imponible:
            _base_imponible_albaranes_actualizados.append((
                _p["id"], _construir_entrada_albaran_num(_entradas_pedido)
            ))

    # (2026-08-19) NOTA sobre "pendientes_sin_pedido" y su
    # "posible_pedido_hint" — dos intentos:
    #
    # Intento 1 (v12.30.10, descartado): antes de dar por perdido un
    # albarán sin pareja en PDF1, se comprobaba contra los pedidos YA
    # DADOS DE ALTA en la base de datos de la app usando la MISMA clave
    # (proveedor_id, importe) que el cruce con PDF1. Caso real detectado
    # (Víctor): pedido Nº618 (SISCOCAN, 2.774,39 €), tramitado 02/06/2026
    # y ya ENTREGADO en la app, cuyo albarán en DALI se registró mucho
    # después (10/08/2026) — muy fuera del rango del PDF de SAP
    # comparado ese día. Ese intento SEGUÍA fallando en producción porque
    # `pedidos.importe` es un importe introducido a mano al dar de alta
    # el pedido (estimación/presupuesto, usado para el techo de gastos
    # mensual) — NO el importe realmente recibido según SAP, que solo se
    # conoce al leer un PDF de SAP recién subido. No hay ninguna garantía
    # de que coincida con el importe del albarán de DALI, así que
    # comparar por importe contra la BD daba 0 candidatos y el caso
    # seguía cayendo en "pendientes_sin_pedido" sin más explicación.
    #
    # Intento 2 (este, v12.30.11): en vez de exigir importe exacto contra
    # la BD, se usa un criterio más flojo pero fiable — mismo proveedor Y
    # que el pedido de la app NO aparezca entre los `vistos1` (todos los
    # números de pedido SAP vistos en el PDF 1, tenga o no tenga ese
    # pedido seguimiento/estado de entrega), es decir, que quede fuera
    # del rango de fechas que cubre el PDF de SAP recién subido — como el
    # caso real del pedido Nº618. Si para un proveedor hay EXACTAMENTE UN
    # pedido de la app en esa situación, se adjunta como
    # "posible_pedido_hint" al elemento de "pendientes_sin_pedido"
    # correspondiente — sin sacarlo de la lista de pendientes (el importe
    # no se ha podido verificar, así que no se puede dar por resuelto
    # automáticamente) y sin aplicar ningún cambio; solo es una pista
    # para que la persona que revisa lo pendiente no tenga que buscar el
    # pedido a mano. Si hay 0 o más de 1 candidato para ese proveedor, no
    # se adjunta ninguna pista — mismo criterio de "ante la duda, no
    # inventar" que el resto de esta función.
    vistos1_normalizados = {_normalizar_pedido_num(n) for n in vistos1}
    pedidos_entregados_bd = rows_to_list(query(
        """SELECT id, norden, pedido_num, proveedor_id, importe, estado
           FROM pedidos
           WHERE hotel_id=%s AND estado IN ('ENTREGADO', 'ENTREGA PARCIAL')
             AND proveedor_id IS NOT NULL
             AND pedido_num IS NOT NULL AND pedido_num != ''""",
        (hotel_id,)
    ))
    candidatos_hint_por_proveedor = defaultdict(list)
    for p in pedidos_entregados_bd:
        if _normalizar_pedido_num(p["pedido_num"]) in vistos1_normalizados:
            continue  # ya cubierto por el PDF1 recién subido, no es el caso que buscamos
        candidatos_hint_por_proveedor[p["proveedor_id"]].append(p)

    for a in pendientes_sin_pedido:
        candidatos_hint = candidatos_hint_por_proveedor.get(a["proveedor_id"], [])
        if len(candidatos_hint) == 1:
            p = candidatos_hint[0]
            a["posible_pedido_hint"] = {
                "pedido_id":         p["id"],
                "norden":            p["norden"],
                "pedido_num_sap":    p["pedido_num"],
                "estado_app_actual": p["estado"],
                "importe_pedido":    float(p["importe"]) if p["importe"] is not None else None,
            }
        else:
            a["posible_pedido_hint"] = None

    # (2026-08-15) A petición del usuario: el resultado de esta comparación
    # de dos PDF debe ser la UNIÓN de lo que ya hace la comparación de un
    # solo PDF (auditoría completa: qué pedidos de SAP no están dados de
    # alta en la app, o están dados de alta pero sin ese estado de entrega)
    # más lo que aporta el cruce con los albaranes de DALI — para que
    # tanto la pantalla como el correo final lo muestren junto, en vez de
    # como dos comparaciones independientes. Se reutiliza tal cual
    # _comparar_listado_pdf_logica() sobre el mismo PDF 1 (relee y
    # reanaliza el mismo texto ya leído arriba — coste asumible, es un
    # job en segundo plano) para no duplicar esa lógica.
    # Escritura de la base imponible silenciosa — ver nota "RELLENO
    # AUTOMÁTICO DE BASE IMPONIBLE EN ENTRADAS YA REGISTRADAS" en el
    # docstring. Importante: esto se escribe ANTES de
    # llamar a _comparar_listado_pdf_logica() (justo debajo), porque esa
    # función vuelve a leer los pedidos de la BD desde cero — así, si
    # ambos mecanismos tocan la base imponible de la MISMA entrada (la
    # última del pedido), el cálculo de _comparar_listado_pdf_logica()
    # (columna 7 de SAP menos entradas anteriores, más fiable al ser el
    # importe realmente recibido acumulado) se ejecuta después, sobre el
    # dato ya fresco, y prevalece sin perder ni pisar nada. Si escribiera
    # aquí después de auditoria_pdf1, correría el riesgo de reconstruir la
    # entrada a partir de una foto ya desactualizada y pisar ese valor más
    # fiable.
    if _base_imponible_albaranes_actualizados:
        for _pid_ba, _entrada_str in _base_imponible_albaranes_actualizados:
            execute("UPDATE pedidos SET entrada_albaran_num=%s WHERE id=%s", (_entrada_str, _pid_ba))
        get_db().commit()

    auditoria_pdf1 = _comparar_listado_pdf_logica(hotel_id, pdf1_bytes)

    return {
        "ok": True,
        "coincidencias":          coincidencias,
        "pendientes_ambiguos":    pendientes_ambiguos,
        "pendientes_sin_albaran": pendientes_sin_albaran,
        "pendientes_sin_pedido":  pendientes_sin_pedido,
        "total_pdf1":             len(encontrados_pdf1),
        "total_pdf2":             len(encontrados_pdf2),
        "excluidos_pdf1":         excluidos_pdf1,
        "excluidos_pdf2":         excluidos_pdf2,
        "auditoria_pdf1":         auditoria_pdf1,
        "base_imponible_albaranes_actualizados": len(_base_imponible_albaranes_actualizados),
    }

def _aplicar_coincidencia_albaran(db, coincidencia: dict, usuario_id: int, usuario_nombre: str) -> dict:
    """
    (2026-08-15) Aplica UNA coincidencia propuesta por
    _comparar_listado_albaranes_logica(): actualiza el pedido ya dado de
    alta con lo detectado en los dos PDF, y dispara la misma notificación
    de cambio de estado que un cambio manual (email retrasado 5 min +
    Telegram inmediato + popup retrasado 5 min, ver
    _notificar_cambio_estado) cuando el estado realmente cambia.

    Solo se llama desde el endpoint .../aplicar, y solo para las
    coincidencias que el usuario ha confirmado explícitamente — nunca de
    forma automática al comparar (decisión de diseño explícita).

    Campos que toca, y solo si hace falta (idempotente — volver a aplicar
    la misma coincidencia dos veces no duplica nada ni reenvía avisos si
    ya estaba todo al día):
      - fecha_tramitacion: SOLO si el pedido no tenía ninguna ya guardada
        (decisión del usuario: no se sobrescribe una fecha existente).
      - entrada_albaran_num: añade "REGISTRO_DALI::FECHA_ISO" — se salta
        si ese registro_dali ya está presente (ya aplicado antes).
      - estado: pasa a ENTREGADO o ENTREGA PARCIAL según lo que
        determinó el PDF 1 — se salta si el pedido ya está en ese estado.

    Devuelve {"aplicado": bool, "cambios": [...], "motivo_sin_cambios": str|None}.
    """
    pedido_id = coincidencia["pedido_id"]
    pedido_actual = row_to_dict(query(
        "SELECT id, estado, fecha_tramitacion, entrada_albaran_num FROM pedidos WHERE id=%s",
        (pedido_id,), one=True
    ))
    if not pedido_actual:
        return {"aplicado": False, "cambios": [], "motivo_sin_cambios": "El pedido ya no existe"}

    cambios = []
    estado_antes = pedido_actual["estado"]

    # Nunca tocar un pedido que mientras tanto se canceló/denegó, o que ya
    # avanzó por delante del estado que proponemos (p. ej. si alguien ya
    # lo marcó ENTREGADO a mano entre que se comparó y se aplicó).
    if estado_antes in ("CANCELADO", "DENEGADO POR DIRECCION GENERAL"):
        return {"aplicado": False, "cambios": [], "motivo_sin_cambios": f"El pedido está {estado_antes}"}

    nueva_fecha_tramitacion = pedido_actual["fecha_tramitacion"]
    if not nueva_fecha_tramitacion and coincidencia.get("fecha_tramitacion_pdf1_iso"):
        nueva_fecha_tramitacion = coincidencia["fecha_tramitacion_pdf1_iso"]
        cambios.append(f"fecha de tramitación → {coincidencia.get('fecha_tramitacion_pdf1_es', nueva_fecha_tramitacion)}")

    nuevo_albaran = pedido_actual["entrada_albaran_num"]
    registro_dali = coincidencia.get("registro_dali")
    _importe_coincidencia = coincidencia.get("importe")
    _entradas_actuales = _parse_albaran_entries(nuevo_albaran)
    # (2026-08-19) Comparación por número normalizado (ignora ceros a la
    # izquierda), no por texto exacto — antes, si el pedido ya tenía
    # registrado p.ej. "81970" y el PDF de DALI traía "00081970", el "in"
    # de texto no lo reconocía como el mismo albarán y se añadía una
    # entrada duplicada. Ver _normalizar_num_albaran().
    ya_presente = bool(registro_dali) and any(
        _normalizar_num_albaran(_e["num"]) == _normalizar_num_albaran(registro_dali)
        for _e in _entradas_actuales
    )
    if registro_dali and not ya_presente:
        # (2026-08-28) A petición de Víctor: la entrada nueva se guarda ya
        # con su base imponible (columna "Importe" de la tabla de
        # coincidencias, el mismo importe recibido con el que se emparejó
        # el albarán) — antes solo se guardaba número + fecha, y la celda
        # "Base imp. (€)" se quedaba vacía aunque el dato ya estaba
        # disponible en ese mismo momento.
        entrada_nueva = _serializar_entrada_albaran(
            registro_dali, coincidencia.get("fecha_registro_dali_iso"), _importe_coincidencia
        )
        nuevo_albaran = f"{nuevo_albaran} | {entrada_nueva}" if nuevo_albaran else entrada_nueva
        cambios.append(f"nueva entrada de albarán {registro_dali}")
    elif registro_dali and ya_presente and _importe_coincidencia is not None:
        # (2026-08-28) La entrada ya estaba registrada (de antes de este
        # cambio, o registrada a mano) pero sin base imponible — se
        # rellena ahora la celda vacía con el importe del albarán, sin
        # tocar el resto de la entrada (número, fecha) ni duplicar nada.
        _cambiado_importe = False
        for _e in _entradas_actuales:
            if (_normalizar_num_albaran(_e["num"]) == _normalizar_num_albaran(registro_dali)
                    and _e.get("base_imponible") is None):
                _e["base_imponible"] = round(float(_importe_coincidencia), 2)
                _cambiado_importe = True
        if _cambiado_importe:
            nuevo_albaran = _construir_entrada_albaran_num(_entradas_actuales)
            cambios.append(f"base imponible del albarán {registro_dali} → {_fmt_importe_es(_importe_coincidencia)} €")

    estado_nuevo = coincidencia["estado_objetivo"]
    if estado_antes == estado_nuevo:
        estado_nuevo = estado_antes  # sin cambio de estado
    elif _ORDEN_ENTREGA_ESTADOS.get(estado_antes, 0) > _ORDEN_ENTREGA_ESTADOS.get(estado_nuevo, 0):
        # El pedido ya está en un estado de entrega más avanzado que el que
        # propone esta coincidencia (p. ej. ya ENTREGADO y esto solo
        # confirma una entrega parcial anterior) — no retroceder el estado.
        estado_nuevo = estado_antes
    else:
        cambios.append(f"estado → {estado_nuevo}")

    if not cambios:
        return {"aplicado": False, "cambios": [], "motivo_sin_cambios": "Ya estaba todo al día"}

    execute(
        """UPDATE pedidos SET
               fecha_tramitacion=%s, entrada_albaran_num=%s, estado=%s,
               modificado_por_id=%s, modificado_por_nombre=%s, modificado_en=NOW()
           WHERE id=%s""",
        (nueva_fecha_tramitacion, nuevo_albaran, estado_nuevo, usuario_id, usuario_nombre, pedido_id)
    )
    if estado_nuevo != estado_antes:
        # (2026-08-19) En la trazabilidad (Historial de estados) este cambio
        # NO debe figurar con el nombre de quien pulsó "Aplicar" en la
        # comparativa — no es una edición manual suya del pedido, es el
        # propio sistema aplicando lo detectado en los PDF. Se guarda un
        # nombre descriptivo fijo en vez del usuario real, para que se
        # distinga a simple vista de un cambio de estado hecho a mano.
        # Petición de Víctor: "LOS EJECUTADOS AUTOMATICAMENTE DEBERIAN
        # SALIR ASI DEFINIDOS Y NO CON NOMBRE DE USUARIO".
        _nombre_automatico = "Automática — listado comparativo pedidos y albaranes"
        execute(
            "INSERT INTO historial_estados (pedido_id,estado_antes,estado_nuevo,usuario_id,usuario_nombre,nota) "
            "VALUES (%s,%s,%s,%s,%s,%s)",
            (pedido_id, estado_antes, estado_nuevo, usuario_id, _nombre_automatico,
             f"Registro automático — comparación de listados PDF (pedidos + albaranes), "
             f"albarán DALI {registro_dali or '—'}")
        )
    db.commit()

    if estado_nuevo != estado_antes:
        _notificar_cambio_estado(db, pedido_id, estado_nuevo, estado_antes, usuario_nombre=usuario_nombre,
                                  usuario_id=usuario_id, es_automatico=True)

    return {"aplicado": True, "cambios": cambios, "motivo_sin_cambios": None}

def _serializar_entrada_albaran(num: str, fecha_iso: str = None, base_imponible=None) -> str:
    """Igual formato que _serializeAlbaranEntry() del frontend: 'NUM::FECHA::IMPORTE',
    'NUM::FECHA' o 'NUM' — el 3er segmento (base imponible) es opcional (2026-08-28)."""
    num = (num or "").strip()
    if not num:
        return ""
    if base_imponible is not None:
        return f"{num}::{fecha_iso or ''}::{float(base_imponible):.2f}"
    return f"{num}::{fecha_iso}" if fecha_iso else num

def _ejecutar_comparacion_albaranes_bg(job_id, hotel_id, pdf1_bytes, pdf2_bytes):
    """Igual patrón que _ejecutar_comparacion_pdf_bg() — corre en un hilo
    aparte con su propio contexto de aplicación, para no bloquear la
    petición HTTP original con la lectura de dos PDF."""
    import time as _time_pdf
    with app.app_context():
        try:
            resultado = _comparar_listado_albaranes_logica(hotel_id, pdf1_bytes, pdf2_bytes)
            with _PDF_JOBS_LOCK:
                if job_id in _PDF_JOBS:
                    _PDF_JOBS[job_id] = {**_PDF_JOBS[job_id], "status": "done", "resultado": resultado}
        except Exception as exc:
            log.error("[COMPARAR-ALBARANES] Error en job %s: %s", job_id, exc)
            with _PDF_JOBS_LOCK:
                if job_id in _PDF_JOBS:
                    _PDF_JOBS[job_id] = {**_PDF_JOBS[job_id], "status": "error", "error": str(exc)}

@app.route("/api/pedidos/comparar-listado-albaranes", methods=["POST"])
@login_required
def comparar_listado_albaranes():
    """
    (2026-08-15) Ampliación de "Comparar listado PDF": ahora acepta un
    segundo PDF — el "Listado de Albaranes" registrados en DALI en base a
    los pedidos — y propone el registro automático de la entrega
    (fecha de tramitación, número de albarán, estado) en los pedidos que
    ya están dados de alta y que SAP muestra como Entregado/Entrega
    parcial. Ver _comparar_listado_albaranes_logica() para el criterio de
    coincidencia completo. Restringido a admin, igual que el resto de
    "Comparar listado PDF".

    Mismo patrón asíncrono que /api/pedidos/comparar-listado-pdf (job_id +
    polling) — necesario aquí todavía más, porque ahora se leen DOS PDF.

    POST /api/pedidos/comparar-listado-albaranes
    form-data: hotel_id, file (PDF 1 — listado de pedidos SAP),
               file2 (PDF 2 — listado de albaranes DALI)
    → 202 {"ok": true, "job_id": "..."}
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403

    hotel_id_raw = request.form.get("hotel_id")
    if not hotel_id_raw:
        return jsonify({"error": "Falta indicar el hotel"}), 400
    try:
        hotel_id = int(hotel_id_raw)
    except ValueError:
        return jsonify({"error": "Hotel no válido"}), 400
    if "file" not in request.files or "file2" not in request.files:
        return jsonify({"error": "Faltan uno o los dos PDF (listado de pedidos y listado de albaranes)"}), 400
    archivo1 = request.files["file"]
    archivo2 = request.files["file2"]
    if not archivo1.filename or not archivo1.filename.lower().endswith(".pdf"):
        return jsonify({"error": "El primer archivo (listado de pedidos) debe ser un PDF"}), 400
    if not archivo2.filename or not archivo2.filename.lower().endswith(".pdf"):
        return jsonify({"error": "El segundo archivo (listado de albaranes) debe ser un PDF"}), 400

    pdf1_bytes = archivo1.read()
    pdf2_bytes = archivo2.read()
    if not pdf1_bytes or not pdf2_bytes:
        return jsonify({"error": "Alguno de los dos archivos está vacío"}), 400

    import time as _time_pdf
    job_id = secrets.token_hex(16)
    with _PDF_JOBS_LOCK:
        limite = _time_pdf.time() - 1800
        for jid in [j for j, v in _PDF_JOBS.items() if v.get("creado_en", 0) < limite]:
            del _PDF_JOBS[jid]
        _PDF_JOBS[job_id] = {
            "status": "processing", "creado_en": _time_pdf.time(),
            "hotel_id": hotel_id, "usuario_id": session.get("user_id"),
        }

    hilo = threading.Thread(
        target=_ejecutar_comparacion_albaranes_bg,
        args=(job_id, hotel_id, pdf1_bytes, pdf2_bytes),
        daemon=True,
    )
    hilo.start()
    return jsonify({"ok": True, "job_id": job_id}), 202

@app.route("/api/pedidos/comparar-listado-albaranes/<job_id>", methods=["GET"])
@login_required
def comparar_listado_albaranes_estado(job_id):
    """Consulta del resultado de un job lanzado por comparar_listado_albaranes()."""
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403
    with _PDF_JOBS_LOCK:
        job = dict(_PDF_JOBS.get(job_id) or {})
    if not job:
        return jsonify({"error": "El job no existe o ha caducado — vuelve a subir los PDF"}), 404
    job.pop("creado_en", None)
    return jsonify(job)

@app.route("/api/pedidos/comparar-listado-albaranes/<job_id>/aplicar", methods=["POST"])
@login_required
def comparar_listado_albaranes_aplicar(job_id):
    """
    (2026-08-15) Aplica las coincidencias que el usuario ha confirmado
    explícitamente desde la pantalla de resultado — decisión de diseño:
    revisar y confirmar antes de aplicar, nunca automático al comparar.

    POST /api/pedidos/comparar-listado-albaranes/<job_id>/aplicar
    body JSON: {"claves": ["12_45_30", ...]}  — claves de coincidencias.claves
               (o {"todas": true} para aplicar todas las propuestas)
    → {"ok": true, "aplicadas": [...], "sin_cambios": [...], "errores": [...]}
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403
    with _PDF_JOBS_LOCK:
        job = dict(_PDF_JOBS.get(job_id) or {})
    if not job:
        return jsonify({"error": "El job no existe o ha caducado — vuelve a subir los PDF"}), 404
    if job.get("status") != "done":
        return jsonify({"error": "La comparación todavía no ha terminado"}), 400

    body = request.get_json(silent=True) or {}
    coincidencias = (job.get("resultado") or {}).get("coincidencias", [])
    if body.get("todas"):
        seleccionadas = coincidencias
    else:
        claves_pedidas = set(body.get("claves") or [])
        if not claves_pedidas:
            return jsonify({"error": "No se ha indicado ninguna coincidencia a aplicar"}), 400
        seleccionadas = [c for c in coincidencias if c["clave"] in claves_pedidas]
        if not seleccionadas:
            return jsonify({"error": "Ninguna de las coincidencias indicadas existe en este resultado"}), 400

    db = get_db()
    uid = current_user_id()
    usuario_nombre = session.get("nombre", "")
    aplicadas, sin_cambios, errores = [], [], []
    for c in seleccionadas:
        try:
            res = _aplicar_coincidencia_albaran(db, c, uid, usuario_nombre)
            etiqueta = f"Pedido {c.get('pedido_num_sap')} — {c.get('proveedor_nombre')}"
            if res["aplicado"]:
                aplicadas.append({"clave": c["clave"], "descripcion": etiqueta, "cambios": res["cambios"]})
            else:
                sin_cambios.append({"clave": c["clave"], "descripcion": etiqueta, "motivo": res["motivo_sin_cambios"]})
        except Exception as exc:
            log.error("[COMPARAR-ALBARANES] Error aplicando coincidencia %s: %s", c.get("clave"), exc)
            errores.append({"clave": c.get("clave"), "error": str(exc)})

    # Guarda lo aplicado en el propio job, para que enviar-resumen() pueda
    # distinguir "realizado ahora" de "seguía pendiente" sin que el
    # frontend tenga que reenviar la lista.
    with _PDF_JOBS_LOCK:
        if job_id in _PDF_JOBS:
            previamente = _PDF_JOBS[job_id].get("aplicadas_acumuladas", [])
            _PDF_JOBS[job_id]["aplicadas_acumuladas"] = previamente + aplicadas

    return jsonify({"ok": True, "aplicadas": aplicadas, "sin_cambios": sin_cambios, "errores": errores})

@app.route("/api/pedidos/comparar-listado-albaranes/<job_id>/enviar-resumen", methods=["POST"])
@login_required
def comparar_listado_albaranes_enviar_resumen(job_id):
    """
    (2026-08-15) Envía un correo interno con el resumen de la comparación
    de listados (pedidos + albaranes), a petición del usuario: los
    registros realizados automáticamente (aplicados en esta sesión de
    comparación, ver .../aplicar) y los que han quedado pendientes de
    realizar (ambiguos, sin albarán, sin pedido, o proveedor no
    identificado en cualquiera de los dos PDF).

    Mismo patrón que .../comparar-listado-pdf/<job_id>/enviar-resumen:
    se encola vía _encolar_email_sistema (sin SMTP propio, lo despacha
    cualquier sesión admin/compras con la app abierta).
    """
    if session.get("rol") != "admin":
        return jsonify({"error": "Acceso restringido a administradores"}), 403
    with _PDF_JOBS_LOCK:
        job = dict(_PDF_JOBS.get(job_id) or {})
    if not job:
        return jsonify({"error": "El job no existe o ha caducado — vuelve a subir los PDF"}), 404
    if job.get("status") != "done":
        return jsonify({"error": "La comparación todavía no ha terminado"}), 400

    resultado = job.get("resultado") or {}
    aplicadas = job.get("aplicadas_acumuladas", [])

    # (2026-08-27) A petición de Víctor: las coincidencias que el sistema
    # SÍ detectó (entrega parcial/total lista para registrar) pero que NO
    # se han aplicado — porque el administrador canceló el aviso
    # automático al terminar la comparación, o simplemente no llegó a
    # pulsar "Aplicar" — deben seguir apareciendo en el correo de resumen,
    # en la sección de pendientes, para que quede constancia de que hay
    # algo por hacer aunque no se haya hecho automáticamente. Antes
    # desaparecían del correo sin más en cuanto no se aplicaban.
    _claves_aplicadas = {a["clave"] for a in aplicadas}
    coincidencias_no_aplicadas = [
        c for c in resultado.get("coincidencias", [])
        if c["clave"] not in _claves_aplicadas and not c.get("sin_cambios_pendientes")
    ]

    # (2026-08-15) A petición del usuario: el correo de esta comparación
    # (pedidos + albaranes) debe ser la UNIÓN de lo que ya envía la
    # comparación de un solo PDF (pedidos de SAP sin dar de alta en la
    # app) más lo que aporta el cruce con los albaranes — un único correo
    # al comprador/admin, no dos independientes. Mismo filtro que usa
    # .../comparar-listado-pdf/<job_id>/enviar-resumen: solo pedidos con
    # proveedor identificado con certeza (el resto queda solo para
    # revisión visual en pantalla).
    auditoria_pdf1 = resultado.get("auditoria_pdf1") or {}
    pedidos_no_encontrados_audit = [p for p in auditoria_pdf1.get("pedidos", []) if not p.get("encontrado")]
    pedidos_faltantes_audit = [p for p in pedidos_no_encontrados_audit if p.get("proveedor_identificado")]
    no_identificados_audit = len(pedidos_no_encontrados_audit) - len(pedidos_faltantes_audit)

    total_pendientes_alb = (
        len(resultado.get("pendientes_ambiguos", []))
        + len(resultado.get("pendientes_sin_albaran", []))
        + len(resultado.get("pendientes_sin_pedido", []))
        + len(coincidencias_no_aplicadas)
    )
    if not pedidos_faltantes_audit and not aplicadas and not total_pendientes_alb:
        return jsonify({"error": "No hay nada que reportar — no se envía nada"}), 400

    hotel = query("SELECT codigo, nombre FROM hoteles WHERE id=%s", (job.get("hotel_id"),), one=True)
    if not hotel:
        return jsonify({"error": "Hotel no encontrado"}), 404

    compradores = _get_compradores_hotel(hotel["codigo"])
    destinatarios = [c["email"] for c in compradores if c.get("email")]
    if not destinatarios:
        return jsonify({
            "error": f"No hay ningún comprador con email asignado al hotel {hotel['codigo']} "
                     "— asígnalo en Admin → Usuarios → Hoteles asignados (Compras)"
        }), 400

    admin_row = query("SELECT nombre, email FROM usuarios WHERE id=%s", (job.get("usuario_id"),), one=True) or {}
    admin_nombre = admin_row.get("nombre") or session.get("nombre") or "Administrador"
    admin_email  = admin_row.get("email")

    subject, body = _email_resumen_comparacion_albaranes(
        hotel["nombre"], hotel["codigo"], resultado, aplicadas, admin_nombre,
        pedidos_faltantes=pedidos_faltantes_audit,
        total_pdf1_audit=auditoria_pdf1.get("total_pdf", 0),
        excluidos_pdf1_audit=auditoria_pdf1.get("excluidos_seguimiento", 0),
        no_identificados_audit=no_identificados_audit,
        coincidencias_no_aplicadas=coincidencias_no_aplicadas,
    )
    _encolar_email_sistema(
        "resumen_comparacion_albaranes", destinatarios, subject, cuerpo_html=body,
        cc_emails=[admin_email] if admin_email else None,
    )
    return jsonify({"ok": True, "destinatarios": destinatarios, "cc": admin_email})

def _motivo_sin_pedido(a):
    """
    (2026-08-19) Texto de motivo para un elemento de "pendientes_sin_pedido"
    (albarán de DALI sin pareja de importe exacto en el PDF de SAP
    comparado), para el CORREO de resumen — ver nota de tamaño más abajo.
    El JS de templates/index.html tiene su propia versión, más larga/
    explicativa, para la pantalla (sin límite de tamaño ahí).

    A petición de Víctor (2026-08-19): el texto debe dejar claro que el
    pedido antiguo detectado NO se puede verificar con la información
    disponible (el PDF de SAP recién subido no llega a esa fecha, y el
    importe de la app es solo una estimación, no el importe realmente
    recibido) y qué hacer para resolverlo: adjuntar un listado de SAP que
    cubra esa fecha, o comprobarlo a mano.

    (2026-08-19) FIX tamaño de correo: la primera redacción de este
    texto (~540 caracteres) multiplicada por decenas de "pendientes" en
    hoteles con mucho volumen (caso real: 79 pendientes) hacía que el
    correo completo superase el límite de tamaño por petición de
    EmailJS — la petición se enviaba, EmailJS la contaba contra el cupo,
    pero la rechazaba con HTTP 413 (Payload Too Large) sin llegar a
    entregarse, y sin ningún aviso visible salvo la discrepancia entre
    el cupo consumido en EmailJS y el contador propio de la app
    (confirmado viendo el Network del navegador: varias peticiones
    `send` a EmailJS con status 413). Se acorta a menos de la mitad
    mantenimiento la información esencial (pedido candidato, estado,
    importe de la app, y las dos acciones a tomar) — ver también el
    límite de filas añadido en _email_resumen_comparacion_albaranes.
    """
    hint = a.get("posible_pedido_hint")
    if not hint:
        return "Sin ningún pedido Entregado/Parcial de ese proveedor que pueda corresponderle"
    # (2026-08-19) A petición de Víctor: identificar el pedido candidato por
    # su número de pedido DALI/SAP (`pedido_num_sap`, el mismo que aparece
    # en los listados de SAP y de DALI que maneja el usuario) en vez de por
    # el "Nº" lineal interno de la app (`norden`) — más intuitivo y fácil
    # de verificar contra esos listados.
    ref_pedido = hint.get("pedido_num_sap") or "—"
    importe_pedido = hint.get("importe_pedido")
    importe_pedido_txt = f", importe app: {importe_pedido:.2f} €" if importe_pedido is not None else ""
    return (
        f"Posible pedido de fecha anterior ya en la app: Pedido {ref_pedido} "
        f"({hint.get('estado_app_actual', 'ENTREGADO')}{importe_pedido_txt}) — sin confirmar, no está en "
        f"este PDF de SAP. Sube un listado que cubra esa fecha o verifícalo a mano."
    )

def _email_resumen_comparacion_albaranes(hotel_nombre, hotel_codigo, resultado, aplicadas, admin_nombre,
                                          pedidos_faltantes=None, total_pdf1_audit=0,
                                          excluidos_pdf1_audit=0, no_identificados_audit=0,
                                          coincidencias_no_aplicadas=None):
    """
    (2026-08-15) Cuerpo del correo de resumen de "Comparar listado PDF"
    (pedidos + albaranes) — misma factura visual que el resto de correos
    internos de la app. Es la UNIÓN de las dos comparaciones que se hacen
    en esta pantalla (a petición del usuario, para enviar un único correo
    al comprador y al admin en vez de dos independientes):
      1. La auditoría del PDF 1 solo (`pedidos_faltantes`, igual que en
         .../comparar-listado-pdf/<job_id>/enviar-resumen): pedidos que
         figuran en SAP pero no están dados de alta en la app todavía.
      2. El cruce con los albaranes de DALI: los registros realizados
         automáticamente (aplicados en esta sesión) y los que quedan
         pendientes de realizar — ambiguos (varios candidatos, requieren
         decisión manual), sin albarán (SAP dice entregado/parcial pero
         no se encontró el albarán en DALI), sin pedido (hay albarán en
         DALI pero ningún pedido dado de alta con ese importe recibido) y
         (2026-08-27) detectadas pero no aplicadas — coincidencias que sí
         se pudieron emparejar con seguridad, pero que el administrador no
         llegó a confirmar (canceló el aviso automático, o no aplicó a
         mano) — sin este añadido desaparecían del correo sin dejar
         rastro de que seguían pendientes.
    """
    pedidos_faltantes = pedidos_faltantes or []
    coincidencias_no_aplicadas = coincidencias_no_aplicadas or []
    pend_ambiguos      = resultado.get("pendientes_ambiguos", [])
    pend_sin_albaran   = resultado.get("pendientes_sin_albaran", [])
    pend_sin_pedido    = resultado.get("pendientes_sin_pedido", [])
    total_pendientes = (
        len(pend_ambiguos) + len(pend_sin_albaran) + len(pend_sin_pedido) + len(coincidencias_no_aplicadas)
    )

    subject = (
        f"[Control de Pedidos] Comparación pedidos+albaranes {hotel_codigo}: "
        f"{len(pedidos_faltantes)} sin dar de alta, {len(aplicadas)} registrado(s), "
        f"{total_pendientes} pendiente(s)"
    )

    def _fila_faltante(p):
        return f"""<tr>
          <td style="padding:8px 12px;border:1px solid #ddd">{p['pedido_num_sap']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p['proveedor_pdf']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p.get('fecha_pedido') or p.get('fecha', '')}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p.get('entrega_estado', '')}</td>
        </tr>"""

    def _fila_aplicada(a):
        return f"""<tr>
          <td style="padding:8px 12px;border:1px solid #ddd">{a['descripcion']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{'; '.join(a['cambios'])}</td>
        </tr>"""

    def _fila_sin_albaran(p):
        return f"""<tr>
          <td style="padding:8px 12px;border:1px solid #ddd">{p['pedido_num_sap']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p['proveedor_nombre']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{p['importe_recibido']:.2f} €</td>
          <td style="padding:8px 12px;border:1px solid #ddd">Sin albarán DALI con ese importe recibido</td>
        </tr>"""

    def _fila_sin_pedido(a):
        motivo = _motivo_sin_pedido(a)
        return f"""<tr>
          <td style="padding:8px 12px;border:1px solid #ddd">Albarán DALI {a['registro_dali']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{a['proveedor_nombre']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{a['importe']:.2f} €</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{motivo}</td>
        </tr>"""

    def _fila_ambiguo(g):
        return f"""<tr>
          <td style="padding:8px 12px;border:1px solid #ddd">{g['proveedor_nombre']}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{g['importe']:.2f} €</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{len(g['pedidos'])} pedido(s) / {len(g['albaranes'])} albarán(es) posibles</td>
          <td style="padding:8px 12px;border:1px solid #ddd">Varios candidatos — requiere decisión manual</td>
        </tr>"""

    def _fila_no_aplicada(c):
        return f"""<tr>
          <td style="padding:8px 12px;border:1px solid #ddd">Pedido {c.get('pedido_num_sap')}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{c.get('proveedor_nombre','')}</td>
          <td style="padding:8px 12px;border:1px solid #ddd">{(c.get('importe') or 0):.2f} €</td>
          <td style="padding:8px 12px;border:1px solid #ddd">Detectado — {c.get('estado_app_actual','—')} →
            {c.get('estado_objetivo','')}, albarán {c.get('registro_dali','—')} — no se aplicó
            (el administrador no lo confirmó al comparar); aplícalo en pantalla</td>
        </tr>"""

    # Filas ya renderizadas una sola vez (independiente del recorte que se
    # aplique después) — la lista de "pendientes" es la unión de las 4
    # categorías, igual que ya hacía la pantalla (más las no aplicadas,
    # ver comentario de coincidencias_no_aplicadas arriba).
    filas_faltantes_todas  = [_fila_faltante(p) for p in pedidos_faltantes]
    filas_aplicadas_todas  = [_fila_aplicada(a) for a in aplicadas]
    filas_pendientes_todas = (
        [_fila_sin_albaran(p) for p in pend_sin_albaran]
        + [_fila_sin_pedido(a) for a in pend_sin_pedido]
        + [_fila_ambiguo(g) for g in pend_ambiguos]
        + [_fila_no_aplicada(c) for c in coincidencias_no_aplicadas]
    )

    aviso_no_identificados_audit = (
        f'<p style="font-size:12px;color:#856404;background:#fff3cd;padding:8px 12px;border-radius:4px">'
        f'⚠️ Hay además <strong>{no_identificados_audit}</strong> pedido(s) sin dar de alta cuyo proveedor '
        f'no se ha podido identificar con certeza en el catálogo — no se incluyen aquí por fiabilidad; '
        f'revísalos en pantalla, en "Comparar listado PDF".</p>'
        if no_identificados_audit else ''
    )

    def _bloque(titulo_html, filas_todas, cap, cols_header, celda_extra_vacio):
        mostrar = filas_todas[:cap]
        resto = len(filas_todas) - len(mostrar)
        aviso_resto = (
            f'<p style="font-size:12px;color:#888;font-style:italic">'
            f'…y {resto} más — consulta el listado completo en la aplicación '
            f'(este correo se acorta para no superar el límite de tamaño de envío).</p>'
            if resto > 0 else ""
        )
        tabla = f"""
        <table style="width:100%;border-collapse:collapse;margin:16px 0;font-size:13px">
          {cols_header}
          {"".join(mostrar)}
        </table>
        {aviso_resto}""" if filas_todas else celda_extra_vacio
        return tabla

    def _construir_body(cap_falt, cap_apl, cap_pend):
        bloque_faltantes = _bloque(
            None, filas_faltantes_todas, cap_falt,
            """<tr style="background:#1a3a6b;color:#fff">
                 <th style="padding:8px 12px;text-align:left">Nº Pedido SAP</th>
                 <th style="padding:8px 12px;text-align:left">Proveedor</th>
                 <th style="padding:8px 12px;text-align:left">Fecha</th>
                 <th style="padding:8px 12px;text-align:left">Entrega</th>
               </tr>""",
            '<p style="color:#155724;font-weight:600">🎉 Todos los pedidos del listado de SAP están dados de alta en la app.</p>'
        )
        bloque_aplicadas = _bloque(
            None, filas_aplicadas_todas, cap_apl,
            """<tr style="background:#155724;color:#fff">
                 <th style="padding:8px 12px;text-align:left">Registro</th>
                 <th style="padding:8px 12px;text-align:left">Cambios aplicados</th>
               </tr>""",
            '<p style="color:#888;font-style:italic">Ningún registro se ha aplicado en esta sesión.</p>'
        )
        bloque_pendientes = _bloque(
            None, filas_pendientes_todas, cap_pend,
            """<tr style="background:#856404;color:#fff">
                 <th style="padding:8px 12px;text-align:left">Referencia</th>
                 <th style="padding:8px 12px;text-align:left">Proveedor</th>
                 <th style="padding:8px 12px;text-align:left">Importe</th>
                 <th style="padding:8px 12px;text-align:left">Motivo</th>
               </tr>""",
            '<p style="color:#155724;font-weight:600">🎉 No ha quedado ningún registro pendiente.</p>'
        )
        return f"""
    <div style="font-family:Arial,sans-serif;max-width:700px;margin:0 auto;border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;">
      {_email_header_html("Princess Hotels &amp; Resorts", "Control de Pedidos — Aviso interno",
                            color_fondo="#1a3a6b", color_subtitulo="#a8c0e8")}
      <div style="padding:24px">
        <p>Se ha comparado el listado de pedidos de SAP con el listado de albaranes
           registrados en DALI para el hotel <strong>{hotel_nombre}</strong> ({hotel_codigo}).</p>
        {_nota_base_imponible_html()}

        <h3 style="color:#1a3a6b;margin-bottom:6px">📋 Pedidos de SAP sin dar de alta en la app ({len(pedidos_faltantes)})</h3>
        <p style="font-size:13px;color:#555">De un total de {total_pdf1_audit} pedidos en el listado de SAP
           ({excluidos_pdf1_audit} de proveedores no sujetos a seguimiento, no evaluados).</p>
        {bloque_faltantes}
        {aviso_no_identificados_audit}
        {'<p>Por favor, revise estos pedidos y dé de alta en Control de Pedidos los que '
         'corresponda — la creación de pedidos no registrados no se hace automáticamente, '
         'debe realizarla manualmente el comprador.</p>' if pedidos_faltantes else ''}

        <h3 style="color:#155724;margin-bottom:6px;margin-top:22px">✅ Registrados automáticamente ({len(aplicadas)})</h3>
        {bloque_aplicadas}
        <h3 style="color:#856404;margin-bottom:6px">⏳ Pendientes de realizar ({total_pendientes})</h3>
        {bloque_pendientes}
        <p style="font-size:12px;color:#888">Consulta generada por {admin_nombre} desde
           "Comparar listado PDF" (Pedidos + Albaranes).</p>
        <hr style="border:none;border-top:1px solid #eee;margin:20px 0">
        <p style="font-size:12px;color:#666">Mensaje automático generado por el sistema de Control de Pedidos.<br>
           <strong>Princess Hotels &amp; Resorts</strong></p>
      </div>
    </div>
    """

    # (2026-08-19) FIX tamaño de correo, definitivo — hasta ahora cada una
    # de las 3 tablas (faltantes / aplicadas / pendientes) se acotaba por
    # separado con un límite de filas fijo (primero solo "pendientes" en
    # v12.30.15, luego también "faltantes" y "aplicadas" aquí mismo) — pero
    # un límite fijo por tabla no evita que la SUMA de las tres, cuando las
    # tres son grandes a la vez, siga superando el límite real de tamaño
    # de EmailJS: fue justo lo que le pasó a Víctor (el correo acabó
    # llegando, pero después de que ~10 intentos anteriores fallasen y
    # descontasen cupo igualmente). En vez de adivinar un límite fijo que
    # sirva para todos los casos, se prueban unos niveles de recorte cada
    # vez más agresivos y se usa el primero cuyo resultado quede por
    # debajo de un margen de seguridad conocido (el caso real de 79 filas
    # "pendientes" solo, recortado a 50, dio 24.002 caracteres y SÍ llegó;
    # el mismo caso sin recortar, 36.445 caracteres, dio 413 — el margen
    # se fija bastante por debajo de ese límite conocido). La pantalla NO
    # tiene este límite en ningún caso, ahí se ve siempre el listado
    # completo sin recortar.
    MARGEN_SEGURO_CHARS = 22000
    NIVELES_RECORTE = [(50, 50, 50), (30, 30, 25), (15, 15, 12), (6, 6, 5)]
    body = None
    for cap_falt, cap_apl, cap_pend in NIVELES_RECORTE:
        body = _construir_body(cap_falt, cap_apl, cap_pend)
        if len(body) <= MARGEN_SEGURO_CHARS:
            break
    return subject, body

@app.route("/api/proveedores/<int:pid>", methods=["DELETE"])
@admin_required
def delete_proveedor(pid):
    db = get_db()
    row = query("SELECT COUNT(*) as cnt FROM pedidos WHERE proveedor_id=%s", (pid,))
    cnt = rows_to_list(row)[0]["cnt"] if row else 0
    if cnt > 0:
        return jsonify({"error": f"No se puede eliminar: tiene {cnt} pedido{'s' if cnt!=1 else ''} asociado{'s' if cnt!=1 else ''}"}), 409
    execute("DELETE FROM proveedor_contactos WHERE proveedor_id=%s", (pid,))
    execute("DELETE FROM proveedores WHERE id=%s", (pid,))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/proveedores/exportar", methods=["GET"])
@login_required
def exportar_proveedores():
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403
    try:
        import openpyxl, io
        from datetime import datetime as dt
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file

        provs = _prov_with_contactos(query(
            "SELECT id,codigo,nombre,observaciones FROM proveedores WHERE activo=1 ORDER BY nombre"
        ))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proveedores"

        # Cabeceras: CODIGO · PROVEEDOR · PRINCIPAL · CONTACTO · TELEFONO · MOVIL · EMAIL · OBSERVACIONES
        headers = ["CODIGO", "PROVEEDOR", "PRINCIPAL", "CONTACTO", "TELEFONO", "MOVIL", "EMAIL", "OBSERVACIONES"]
        col_widths = [14, 42, 10, 25, 18, 18, 35, 38]

        hdr_fill_prov = PatternFill("solid", fgColor="1B2A4A")
        hdr_fill_ctc  = PatternFill("solid", fgColor="2E5090")
        hdr_font      = Font(bold=True, color="FFFFFF")

        ctc_cols = {3, 4, 5, 6, 7}  # columnas de contacto (1-based)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = hdr_fill_ctc if col_idx in ctc_cols else hdr_fill_prov
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

        # Estilos de filas
        from openpyxl.styles import Border, Side
        thin = Side(style="thin", color="D0D7E3")
        border = Border(bottom=thin)
        fill_principal = PatternFill("solid", fgColor="FFF8E7")   # dorado claro → principal
        fill_alt       = PatternFill("solid", fgColor="F5F7FA")   # gris claro → proveedor sin color
        fill_ctc_alt   = PatternFill("solid", fgColor="EEF2FA")   # azul muy claro → contacto secundario

        r_idx = 2
        for p in provs:
            contactos = p.get("contactos", [{}])
            if not contactos:
                contactos = [{}]
            for ci, c in enumerate(contactos):
                es_principal = c.get("es_principal", ci == 0)
                principal_val = "★" if es_principal else ""

                # Color de fila
                if es_principal:
                    row_fill = fill_principal
                elif ci > 0:
                    row_fill = fill_ctc_alt
                else:
                    row_fill = None

                vals = [
                    p.get("codigo") or ""      if ci == 0 else "",
                    p.get("nombre") or ""      if ci == 0 else "",
                    principal_val,
                    c.get("nombre") or "",
                    c.get("telefono") or "",
                    c.get("movil") or "",
                    c.get("email") or "",
                    p.get("observaciones") or "" if ci == 0 else "",
                ]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=r_idx, column=col_idx, value=val)
                    cell.border = border
                    if col_idx == 3:  # PRINCIPAL col
                        cell.alignment = Alignment(horizontal="center")
                        cell.font = Font(bold=True, color="B8860B")
                    if row_fill:
                        cell.fill = row_fill
                r_idx += 1

        # Nota de instrucciones en la parte inferior
        ws.cell(row=r_idx + 1, column=1, value="INSTRUCCIONES DE IMPORTACIÓN:").font = Font(bold=True, color="1B2A4A")
        instrucciones = [
            "• CODIGO: código SAP (obligatorio). Identifica al proveedor — si ya existe se actualiza, si no existe se crea.",
            "• PRINCIPAL: Escribe ★ o 1 o SI en la fila del contacto que recibirá emails/WhatsApp automáticos. Solo uno por proveedor.",
            "• Varios contactos del mismo proveedor: repite CODIGO y PROVEEDOR en filas adicionales, deja OBSERVACIONES vacío.",
            "• TELEFONO: teléfono fijo.  MOVIL: móvil/WhatsApp (se usará para alertas automáticas).",
            "• Para eliminar todos los contactos de un proveedor: deja CONTACTO, TELEFONO, MOVIL y EMAIL vacíos.",
        ]
        for i, txt in enumerate(instrucciones, r_idx + 2):
            cell = ws.cell(row=i, column=1, value=txt)
            cell.font = Font(italic=True, color="555555", size=9)
            ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=len(headers))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"PROVEEDORES_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _parse_excel_proveedores(archivo):
    """
    Lee un .xlsx de proveedores y devuelve (prov_order, prov_data).
    prov_data[key] = {codigo, nombre, observaciones, contactos: [(nombre,tel,movil,email,es_principal), ...]}
    Hace todo el trabajo en memoria — sin tocar la BD — para minimizar el tiempo de conexión.
    """
    import openpyxl
    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    ws = wb.active
    raw_headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip().upper() if h is not None else "" for h in raw_headers]

    PRINCIPAL_SI = {"★", "1", "SI", "SÍ", "S", "YES", "Y", "TRUE"}

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx].value
            if v is None:
                return None
            s = str(v).strip()
            return s if s and s.lower() not in ("none", "nan") else None
        except (ValueError, IndexError):
            return None

    prov_data  = {}   # key → dict
    prov_order = []   # mantener orden de aparición

    last_key = None  # último proveedor visto — para filas de contacto extra sin CODIGO/PROVEEDOR

    for row in ws.iter_rows(min_row=2):
        nombre = col(row, "PROVEEDOR")
        codigo = col(row, "CODIGO") or ""

        if nombre:
            # Fila con proveedor identificado → crear/localizar entrada
            key = codigo or nombre
            if key not in prov_data:
                obs = col(row, "OBSERVACIONES") or ""
                prov_data[key] = {
                    "codigo":        codigo,
                    "nombre":        nombre,
                    "observaciones": obs,
                    "contactos":     [],
                }
                prov_order.append(key)
            last_key = key
        else:
            # Fila sin PROVEEDOR → contacto adicional del último proveedor visto
            key = last_key

        if key is None:
            continue  # fila suelta sin proveedor de referencia, ignorar

        c_nombre    = col(row, "CONTACTO")  or ""
        c_tel       = col(row, "TELEFONO")  or ""
        c_movil     = col(row, "MOVIL")     or ""
        c_email     = col(row, "EMAIL")     or ""
        c_principal = col(row, "PRINCIPAL") or ""
        es_principal = str(c_principal).strip().upper() in PRINCIPAL_SI

        if c_nombre or c_tel or c_movil or c_email:
            prov_data[key]["contactos"].append(
                (c_nombre, c_tel, c_movil, c_email, es_principal)
            )

    wb.close()

    # Garantizar que cada proveedor con contactos tenga exactamente uno principal
    for key in prov_order:
        ctcs = prov_data[key]["contactos"]
        if ctcs and not any(c[4] for c in ctcs):
            ctcs[0] = (ctcs[0][0], ctcs[0][1], ctcs[0][2], ctcs[0][3], True)

    return prov_order, prov_data

@app.route("/api/proveedores/importar", methods=["POST"])
@admin_required
def importar_proveedores():
    """
    Importación incremental: actualiza existentes (por código SAP), inserta nuevos.
    Usa bulk operations para evitar timeouts con listas grandes (>500 proveedores).
    Total de round-trips a la BD: ~5, independientemente del tamaño del Excel.
    """
    try:
        if "archivo" not in request.files:
            return jsonify({"ok": False, "error": "No se recibió ningún archivo"}), 400
        archivo = request.files["archivo"]
        if not archivo.filename.endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "error": "El archivo debe ser .xlsx"}), 400

        # ── 1. Parsear Excel completamente en memoria (sin BD) ──────────────
        prov_order, prov_data = _parse_excel_proveedores(archivo)

        # ── 2. Una sola query para saber qué proveedores ya existen ─────────
        from psycopg2.extras import execute_values
        db  = get_db()
        existentes = {r["codigo"]: r["id"] for r in rows_to_list(
            query("SELECT id, codigo FROM proveedores WHERE codigo IS NOT NULL AND codigo != ''")
        )}

        to_update = []   # (nombre, obs, id)
        to_insert = []   # (codigo, nombre, obs)

        for key in prov_order:
            p = prov_data[key]
            codigo = p["codigo"]
            if codigo and codigo in existentes:
                to_update.append((p["nombre"], p["observaciones"], existentes[codigo]))
            else:
                to_insert.append((codigo or None, p["nombre"], p["observaciones"]))

        # ── 3. Bulk UPDATE de proveedores existentes ─────────────────────────
        actualizados = 0
        with db.cursor() as cur:
            if to_update:
                execute_values(
                    cur,
                    """UPDATE proveedores AS p SET nombre=v.nombre, observaciones=v.obs
                       FROM (VALUES %s) AS v(nombre, obs, id)
                       WHERE p.id = v.id::int""",
                    to_update,
                    template="(%s, %s, %s)"
                )
                actualizados = len(to_update)

        # ── 4. Bulk INSERT de proveedores nuevos → recuperar sus IDs ─────────
        insertados = 0
        nuevos_ids = {}   # codigo_o_nombre → id
        if to_insert:
            with db.cursor() as cur:
                execute_values(
                    cur,
                    """INSERT INTO proveedores (codigo, nombre, observaciones)
                       VALUES %s
                       ON CONFLICT DO NOTHING
                       RETURNING id, codigo, nombre""",
                    to_insert,
                    template="(%s, %s, %s)",
                    fetch=True
                )
                rows = cur.fetchall()
                for row in rows:
                    k = row["codigo"] or row["nombre"]
                    nuevos_ids[k] = row["id"]
                    insertados += 1

        # Mapear todos los keys a su ID final
        key_to_id = {}
        for key in prov_order:
            p = prov_data[key]
            codigo = p["codigo"]
            if codigo and codigo in existentes:
                key_to_id[key] = existentes[codigo]
            else:
                kid = nuevos_ids.get(codigo) or nuevos_ids.get(p["nombre"])
                if kid:
                    key_to_id[key] = kid

        # ── 5. Reemplazar contactos: DELETE existentes + bulk INSERT nuevos ──
        ids_con_datos = list(key_to_id.values())
        if ids_con_datos:
            with db.cursor() as cur:
                # DELETE en un solo IN (una query)
                cur.execute(
                    "DELETE FROM proveedor_contactos WHERE proveedor_id = ANY(%s)",
                    (ids_con_datos,)
                )

                # Construir todas las filas de contactos a insertar
                contactos_rows = []
                for key in prov_order:
                    pid = key_to_id.get(key)
                    if pid is None:
                        continue
                    for orden, (cn, ct, cm, ce, ep) in enumerate(prov_data[key]["contactos"]):
                        contactos_rows.append((
                            pid,
                            cn or None,
                            ct or None,
                            cm or None,
                            ce or None,
                            1 if ep else 0,
                            orden
                        ))

                # INSERT en una sola query bulk
                if contactos_rows:
                    execute_values(
                        cur,
                        """INSERT INTO proveedor_contactos
                           (proveedor_id, nombre, telefono, movil, email, es_principal, orden)
                           VALUES %s""",
                        contactos_rows,
                        template="(%s, %s, %s, %s, %s, %s, %s)",
                        page_size=500
                    )

        db.commit()
        return jsonify({"ok": True, "insertados": insertados, "actualizados": actualizados, "errores": []})

    except Exception as e:
        import traceback
        log.error(f"importar_proveedores error: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/proveedores/importar/reset", methods=["POST"])
@login_required
def importar_proveedores_reset():
    """Solo admin: borra todos los proveedores e importa desde el Excel.
    Usa bulk operations para evitar timeouts con listas grandes.
    Total de round-trips a la BD: ~4, independientemente del tamaño del Excel.
    """
    if session.get("rol") != "admin":
        return jsonify({"ok": False, "error": "Acceso restringido a administradores"}), 403
    try:
        if "archivo" not in request.files:
            return jsonify({"ok": False, "error": "No se recibió ningún archivo"}), 400
        archivo = request.files["archivo"]
        if not archivo.filename.endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "error": "El archivo debe ser .xlsx"}), 400

        # ── 1. Parsear Excel completamente en memoria ────────────────────────
        prov_order, prov_data = _parse_excel_proveedores(archivo)

        from psycopg2.extras import execute_values
        db = get_db()

        with db.cursor() as cur:
            # ── 2. Limpiar todo de una vez (3 queries) ───────────────────────
            cur.execute("UPDATE pedidos SET proveedor_id = NULL WHERE proveedor_id IS NOT NULL")
            cur.execute("DELETE FROM proveedor_contactos")
            cur.execute("DELETE FROM proveedores")

            # ── 3. Bulk INSERT proveedores → recuperar IDs en un paso ────────
            prov_rows = [
                (prov_data[k]["codigo"] or None, prov_data[k]["nombre"], prov_data[k]["observaciones"])
                for k in prov_order
            ]
            execute_values(
                cur,
                "INSERT INTO proveedores (codigo, nombre, observaciones) VALUES %s RETURNING id, codigo, nombre",
                prov_rows,
                template="(%s, %s, %s)",
                page_size=500,
                fetch=True
            )
            inserted_rows = cur.fetchall()

            # Mapear codigo/nombre → id preservando el orden
            key_to_id = {}
            for row in inserted_rows:
                k = row["codigo"] or row["nombre"]
                key_to_id[k] = row["id"]

            # ── 4. Bulk INSERT de todos los contactos ────────────────────────
            contactos_rows = []
            for key in prov_order:
                pid = key_to_id.get(key)
                if pid is None:
                    continue
                for orden, (cn, ct, cm, ce, ep) in enumerate(prov_data[key]["contactos"]):
                    contactos_rows.append((
                        pid,
                        cn or None,
                        ct or None,
                        cm or None,
                        ce or None,
                        1 if ep else 0,
                        orden
                    ))

            if contactos_rows:
                execute_values(
                    cur,
                    """INSERT INTO proveedor_contactos
                       (proveedor_id, nombre, telefono, movil, email, es_principal, orden)
                       VALUES %s""",
                    contactos_rows,
                    template="(%s, %s, %s, %s, %s, %s, %s)",
                    page_size=500
                )

        db.commit()
        insertados = len(prov_order)
        return jsonify({"ok": True, "insertados": insertados, "actualizados": 0, "errores": []})

    except Exception as e:
        import traceback
        log.error(f"importar_proveedores_reset error: {traceback.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Validación techo de gastos ─────────────────────────────────────────────────
# Nota: los valores de techo (techo_max_pedido, techo_max_mes, techo_max_pedidos)
# se leen siempre desde get_config() → BD. No se asignan aquí para no llamar
# a get_config() fuera de contexto Flask (rompe el arranque en Render).

def _check_techo(hotel_id, familia_id, importe, mes_str, excluir_pedido_id=None):
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 2) Comprueba las reglas del
    techo de gastos para un pedido que va a pasar a ENVIADO AL PROVEEDOR.

    IMPORTANTE — cambio de comportamiento respecto a versiones anteriores:
    esta función YA NO se llama al crear/editar un pedido (eso quedaba
    bloqueado con un confirm() de JS saltable por cualquiera, sin rastro —
    ver _forzar_techo, eliminado). Se llama ÚNICAMENTE desde update_pedido(),
    en el momento exacto en que el pedido va a pasar a ENVIADO AL
    PROVEEDOR — es el momento real de "consumo" del techo, no la creación.
    Los motivos que devuelve ya NO bloquean el guardado: son el detonante
    para abrir un expediente_exceso y desviar el pedido a
    PENDIENTE Vº Bº DIRECCIÓN GENERAL en vez de enviarlo directamente.

    mes_str: 'YYYY-MM' — el mes en el que el pedido va a CONSUMIR techo
    (normalmente el mes actual), no el mes en que se creó.
    excluir_pedido_id: por seguridad, excluye el propio pedido del conteo
    (no debería tener mes_consumo_techo relleno todavía en circunstancias
    normales, pero evita autocontarse en un reintento).

    Devuelve lista de strings con los motivos detectados (vacía = dentro
    de los límites, no requiere autorización de Dirección General).
    """
    cfg     = get_config()
    errores = []

    if importe and float(importe) > cfg["techo_max_pedido"]:
        lim = cfg["techo_max_pedido"]
        errores.append(
            f"⚠️ El importe {float(importe):,.2f} € supera el límite individual de {lim:,.0f} € por pedido."
        )

    if not familia_id:
        return errores   # sin familia no hay más que comprobar

    excl_clause = "AND p.id != %s" if excluir_pedido_id else ""
    excl_args   = (excluir_pedido_id,) if excluir_pedido_id else ()

    # Pedidos que YA consumen techo este hotel/mes (mes_consumo_techo se
    # rellena solo al pasar a ENVIADO AL PROVEEDOR — ya no se cuentan
    # pedidos por su fecha de creación, sino por si de verdad han llegado a
    # consumir presupuesto).
    base_args = (hotel_id, mes_str) + excl_args
    sql = (
        "SELECT p.id, p.familia_id, f.nombre as familia_nombre, "
        "COALESCE(p.importe, 0) as importe "
        "FROM pedidos p "
        "LEFT JOIN familias f ON p.familia_id = f.id "
        "WHERE p.hotel_id = %s "
        "  AND p.sujeto_techo = 1 "
        "  AND p.mes_consumo_techo = %s "
        + ("  " + excl_clause if excl_clause else "")
    )
    pedidos_mes = rows_to_list(query(sql, base_args))

    # (Rediseño 2026-08-01) Se elimina la antigua Regla 1 ("máximo N pedidos
    # sujetos al techo por hotel/mes", agregado sin distinguir familia) —
    # decisión de negocio del rediseño: solo queda vigente el límite por
    # hotel + familia (Regla siguiente).

    # Regla: máximo N pedidos por hotel/mes Y familia (configurable vía
    # techo_max_pedidos_familia).
    max_pedidos_familia = cfg["techo_max_pedidos_familia"]
    pedidos_familia = [p for p in pedidos_mes if p["familia_id"] == int(familia_id)]
    nuevo_importe   = float(importe) if importe else 0.0
    fname = None
    if len(pedidos_familia) >= max_pedidos_familia or cfg["techo_max_mes_familia"]:
        familia_row = query("SELECT nombre FROM familias WHERE id=%s", (familia_id,), one=True)
        fname = familia_row["nombre"] if familia_row else "ID {}".format(familia_id)
    if len(pedidos_familia) >= max_pedidos_familia:
        errores.append(
            f"🚫 Ya hay {len(pedidos_familia)} pedido(s) de la familia \u00ab{fname}\u00bb este mes "
            f"para este hotel (máximo {max_pedidos_familia} por hotel/mes/familia)."
        )

    # Regla: acumulado mensual no puede superar el techo mensual
    techo_mes     = cfg["techo_max_mes"]
    acumulado     = sum(float(p["importe"]) for p in pedidos_mes)
    if acumulado + nuevo_importe > techo_mes:
        errores.append(
            f"⚠️ El acumulado del mes sería {acumulado + nuevo_importe:,.2f} € "
            f"(actual {acumulado:,.2f} € + nuevo {nuevo_importe:,.2f} €), "
            f"superando el techo mensual de {techo_mes:,.0f} €."
        )

    # Regla: acumulado mensual de la FAMILIA no puede superar el techo
    # mensual por hotel/mes/familia (techo_max_mes_familia). 0 = sin límite
    # específico por familia (solo aplica la regla anterior, sobre el total
    # del hotel).
    techo_mes_familia = cfg["techo_max_mes_familia"]
    if techo_mes_familia:
        acumulado_familia = sum(float(p["importe"]) for p in pedidos_familia)
        if acumulado_familia + nuevo_importe > techo_mes_familia:
            errores.append(
                f"⚠️ El acumulado de la familia \u00ab{fname}\u00bb este mes sería "
                f"{acumulado_familia + nuevo_importe:,.2f} € "
                f"(actual {acumulado_familia:,.2f} € + nuevo {nuevo_importe:,.2f} €), "
                f"superando el techo de {techo_mes_familia:,.0f} € por hotel/mes/familia."
            )

    return errores


def _techo_snapshot(hotel_id, mes_str):
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 2) Fotografía del consumo
    de techo del hotel en el mes indicado, en el momento exacto de la
    llamada. Se usa para congelar consumido_en_solicitud/
    disponible_en_solicitud en expediente_exceso al crear la fila — nunca
    se recalcula después (punto 10 del rediseño: el informe de Dirección
    General debe reflejar la situación que había EN ESE MOMENTO, aunque
    otros pedidos consuman techo más tarde).

    Devuelve (consumido, disponible), ambos float, a nivel de hotel/mes
    (sobre techo_max_mes) — no desglosado por familia; el detalle por
    familia vive en las columnas propias de cada fila de expediente_exceso
    (consumo_previo/exceso, calculadas por el caller con los datos que ya
    tiene de _check_techo).
    """
    cfg = get_config()
    row = query(
        "SELECT COALESCE(SUM(importe),0) as total FROM pedidos "
        "WHERE hotel_id=%s AND sujeto_techo=1 AND mes_consumo_techo=%s",
        (hotel_id, mes_str), one=True
    )
    consumido  = float(row["total"] or 0)
    techo_mes  = float(cfg.get("techo_max_mes", 0) or 0)
    disponible = techo_mes - consumido
    return consumido, disponible

@app.route("/api/techo/resumen")
@login_required
def techo_resumen():
    """Devuelve el resumen del techo de gastos del mes actual por hotel.

    Versión optimizada: 2 queries fijas (hoteles + pedidos del mes en un solo
    SELECT) en lugar del patrón N+1 anterior (1 query por hotel).
    get_config() se lee una sola vez y sus valores se reutilizan para todos
    los hoteles.

    (2026-08-01 — rediseño Techo de Gastos, Fase 4)
    - Filtro de pedidos cambiado de EXTRACT(YEAR/MONTH FROM p.creado_en) a
      mes_consumo_techo — ya solo cuenta pedidos que de verdad han
      consumido techo (pasaron por ENVIADO AL PROVEEDOR).
    - Nuevos bloques por hotel: "pendientes_dg" (expedientes en
      expediente_exceso con resultado='pendiente' este mes) y
      "excesos_autorizados" (resultado='aprobado'), Sección 8 del
      documento de diseño.
    - Nuevo indicador "compromiso_potencial" = consumido + pendiente_dg
      (punto 9) — no modifica el cálculo del techo, solo visibilidad.
    - Semáforo: se mantienen los mismos umbrales rojo/amarillo/verde de
      siempre (reutilizados de _job_alertas_techo_mensual, sin cambios),
      añadiendo el nuevo caso "azul" cuando el hotel tiene al menos un
      expediente pendiente de resolver este mes (punto 12).
    """
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403
    from datetime import date
    hoy   = date.today()
    year  = hoy.year
    month = hoy.month
    mes_str = f"{year}-{month:02d}"

    # ── 1. Configuración: una sola lectura ───────────────────────────────────
    cfg              = get_config()
    techo_max_mes    = cfg["techo_max_mes"]
    techo_max_pedido = cfg["techo_max_pedido"]
    techo_max_ped_n  = cfg["techo_max_pedidos"]   # max numero de pedidos
    techo_max_ped_fam = cfg["techo_max_pedidos_familia"]   # max numero de pedidos por familia
    techo_max_mes_fam = cfg["techo_max_mes_familia"]        # max importe (€) por hotel/mes/familia (0=sin límite)
    pct_amarillo     = cfg["techo_pct_amarillo"]
    umbral_amarillo  = techo_max_mes * pct_amarillo / 100

    # ── 2. Hoteles activos: una query ────────────────────────────────────────
    if _puede_ver_hotel_pruebas():
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        ))
    else:
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 AND codigo <> %s ORDER BY codigo",
            (HOTEL_CODIGO_PRUEBAS,)
        ))
    if not hoteles:
        return jsonify({"mes": mes_str, "hoteles": []})

    hotel_ids = [h["id"] for h in hoteles]
    ph        = ",".join(["%s"] * len(hotel_ids))

    # ── 3. Pedidos del mes: una sola query para todos los hoteles ────────────
    pedidos_mes = rows_to_list(query(f"""
        SELECT p.id, p.hotel_id, p.importe, p.familia_id,
               f.nombre  AS familia_nombre,
               p.pedido_num, p.estado, p.norden,
               pr.nombre AS proveedor_nombre,
               p.observaciones
        FROM pedidos p
        LEFT JOIN familias    f  ON p.familia_id    = f.id
        LEFT JOIN proveedores pr ON p.proveedor_id  = pr.id
        WHERE p.hotel_id IN ({ph})
          AND p.sujeto_techo = 1
          AND p.mes_consumo_techo = %s
        ORDER BY p.hotel_id, p.creado_en
    """, hotel_ids + [mes_str]))

    # ── 3b. Expedientes del mes (Sección 8): pendientes + aprobados ──────────
    expedientes_mes = rows_to_list(query(f"""
        SELECT e.id, e.pedido_id, e.hotel_id, e.familia_id, e.importe_pedido,
               e.exceso, e.motivo_solicitud, e.resultado, e.creado_en,
               e.fecha_resolucion, e.observaciones_direccion_general,
               f.nombre AS familia_nombre, p.pedido_num
        FROM expediente_exceso e
        LEFT JOIN familias f ON e.familia_id = f.id
        LEFT JOIN pedidos  p ON e.pedido_id   = p.id
        WHERE e.hotel_id IN ({ph})
          AND e.mes = %s
          AND e.resultado IN ('pendiente', 'aprobado')
        ORDER BY e.hotel_id, e.creado_en DESC
    """, hotel_ids + [mes_str]))

    # ── 4. Agrupar pedidos y expedientes por hotel en memoria ────────────────
    from collections import defaultdict
    pedidos_por_hotel: dict = defaultdict(list)
    for p in pedidos_mes:
        pedidos_por_hotel[p["hotel_id"]].append(p)

    expedientes_por_hotel: dict = defaultdict(list)
    for e in expedientes_mes:
        expedientes_por_hotel[e["hotel_id"]].append(e)

    # ── 5. Construir resultado ────────────────────────────────────────────────
    resultado = []
    for hotel in hoteles:
        pedidos         = pedidos_por_hotel[hotel["id"]]
        acumulado       = sum(float(p["importe"] or 0) for p in pedidos)
        num_pedidos     = len(pedidos)
        familias_usadas = [p["familia_nombre"] for p in pedidos if p["familia_nombre"]]

        # v12.28.0 — conteo por familia, para poder avisar en el frontend
        # cuando una familia concreta está en (o supera) su propio límite
        # mensual por hotel (techo_max_pedidos_familia), no solo el total.
        familias_conteo: dict = {}
        for p in pedidos:
            fn = p["familia_nombre"]
            if fn:
                familias_conteo[fn] = familias_conteo.get(fn, 0) + 1

        # v12.29.0 — importe acumulado por familia, para avisar cuando una
        # familia concreta esté en (o supere) su propio techo de importe
        # mensual (techo_max_mes_familia), independiente del total del hotel.
        familias_importe: dict = {}
        for p in pedidos:
            fn = p["familia_nombre"]
            if fn:
                familias_importe[fn] = familias_importe.get(fn, 0) + float(p["importe"] or 0)

        # (2026-08-01) Expedientes pendientes/aprobados de este hotel/mes
        _exps_hotel   = expedientes_por_hotel[hotel["id"]]
        _pendientes   = [e for e in _exps_hotel if e["resultado"] == "pendiente"]
        _aprobados    = [e for e in _exps_hotel if e["resultado"] == "aprobado"]
        pendiente_dg_importe     = sum(float(e["importe_pedido"] or 0) for e in _pendientes)
        exceso_autorizado_importe = sum(float(e["importe_pedido"] or 0) for e in _aprobados)
        compromiso_potencial      = acumulado + pendiente_dg_importe

        # Semaforo:
        #   ROJO     -> acumulado >= techo_max_mes  O  num_pedidos > techo_max_ped_n
        #   AMARILLO -> acumulado >= umbral_amarillo O  num_pedidos >= techo_max_ped_n
        #   AZUL (2026-08-01, punto 12) -> hay al menos un expediente pendiente
        #   de resolver este mes — se superpone a rojo/amarillo/verde porque
        #   es la situación que más necesita atención (alguien está esperando
        #   una decisión de Dirección General).
        if acumulado >= techo_max_mes or num_pedidos > techo_max_ped_n:
            semaforo = "rojo"
        elif acumulado >= umbral_amarillo or num_pedidos >= techo_max_ped_n:
            semaforo = "amarillo"
        else:
            semaforo = "verde"
        if _pendientes:
            semaforo = "azul"

        resultado.append({
            "hotel_id":        hotel["id"],
            "hotel_codigo":    hotel["codigo"],
            "hotel_nombre":    hotel["nombre"],
            "num_pedidos":     num_pedidos,
            "max_pedidos":     techo_max_ped_n,
            "acumulado":       acumulado,
            "techo_mes":       techo_max_mes,
            "techo_pedido":    techo_max_pedido,
            "familias_usadas": familias_usadas,
            "familias_conteo": familias_conteo,
            "max_pedidos_familia": techo_max_ped_fam,
            "familias_importe": familias_importe,
            "max_importe_familia": techo_max_mes_fam,
            "semaforo":        semaforo,
            "pedidos":         pedidos,
            "pendientes_dg":         _pendientes,
            "pendientes_dg_importe": pendiente_dg_importe,
            "excesos_autorizados":         _aprobados,
            "excesos_autorizados_importe": exceso_autorizado_importe,
            "compromiso_potencial":  compromiso_potencial,
        })

    return jsonify({"mes": mes_str, "hoteles": resultado})

@app.route("/api/techo/resumen-historico")
@login_required
def techo_resumen_historico():
    """Devuelve el techo de gastos de un mes/año concreto (histórico).

    (2026-08-01 — rediseño Techo de Gastos, Fase 4) Simplificado: antes
    calculaba la "fecha de envío" con un COALESCE(historial_estados,
    fecha_tramitacion, creado_en) + DATE_TRUNC, y encima exigía
    p.estado = 'ENVIADO AL PROVEEDOR' — lo que EXCLUÍA incorrectamente
    cualquier pedido que ya hubiera avanzado a ENTREGA PARCIAL/ENTREGADO
    desde entonces (un pedido de hace 3 meses ya entregado desaparecía del
    histórico de ese mes). Ahora se usa directamente mes_consumo_techo, que
    ya captura el mes real de consumo independientemente del estado actual
    del pedido — más simple y más correcto. Los pedidos CANCELADO quedan
    excluidos automáticamente porque cancelar limpia mes_consumo_techo
    (ver update_pedido()), no hace falta filtrarlo aparte.

    También incluye ahora los mismos bloques de expedientes
    (pendientes_dg / excesos_autorizados) que el resumen del mes actual.
    """
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403

    try:
        year  = int(request.args.get("year",  0))
        month = int(request.args.get("month", 0))
        if not (2020 <= year <= 2099 and 1 <= month <= 12):
            return jsonify({"error": "Parámetros year/month inválidos"}), 400
    except (TypeError, ValueError):
        return jsonify({"error": "Parámetros year/month inválidos"}), 400

    mes_str = f"{year}-{month:02d}"

    # ── 1. Configuración: una sola lectura ───────────────────────────────────
    cfg              = get_config()
    techo_max_mes    = cfg["techo_max_mes"]
    techo_max_pedido = cfg["techo_max_pedido"]
    techo_max_ped_n  = cfg["techo_max_pedidos"]
    techo_max_ped_fam = cfg["techo_max_pedidos_familia"]
    techo_max_mes_fam = cfg["techo_max_mes_familia"]
    pct_amarillo     = cfg["techo_pct_amarillo"]
    umbral_amarillo  = techo_max_mes * pct_amarillo / 100

    # ── 2. Hoteles activos: una query ────────────────────────────────────────
    if _puede_ver_hotel_pruebas():
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        ))
    else:
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 AND codigo <> %s ORDER BY codigo",
            (HOTEL_CODIGO_PRUEBAS,)
        ))
    if not hoteles:
        return jsonify({"mes": mes_str, "hoteles": [], "historico": True})

    hotel_ids = [h["id"] for h in hoteles]
    ph        = ",".join(["%s"] * len(hotel_ids))

    # ── 3. Pedidos del mes histórico: mes_consumo_techo directamente ─────────
    pedidos_mes = rows_to_list(query(f"""
        SELECT p.id, p.hotel_id, p.importe, p.familia_id,
               f.nombre  AS familia_nombre,
               p.pedido_num, p.estado, p.norden,
               pr.nombre AS proveedor_nombre,
               p.observaciones
        FROM pedidos p
        LEFT JOIN familias    f  ON p.familia_id   = f.id
        LEFT JOIN proveedores pr ON p.proveedor_id = pr.id
        WHERE p.hotel_id IN ({ph})
          AND p.sujeto_techo = 1
          AND p.mes_consumo_techo = %s
        ORDER BY p.hotel_id, p.creado_en
    """, hotel_ids + [mes_str]))

    # ── 3b. Expedientes de ese mes histórico ─────────────────────────────────
    expedientes_mes = rows_to_list(query(f"""
        SELECT e.id, e.pedido_id, e.hotel_id, e.familia_id, e.importe_pedido,
               e.exceso, e.motivo_solicitud, e.resultado, e.creado_en,
               e.fecha_resolucion, e.observaciones_direccion_general,
               f.nombre AS familia_nombre, p.pedido_num
        FROM expediente_exceso e
        LEFT JOIN familias f ON e.familia_id = f.id
        LEFT JOIN pedidos  p ON e.pedido_id   = p.id
        WHERE e.hotel_id IN ({ph})
          AND e.mes = %s
          AND e.resultado IN ('pendiente', 'aprobado')
        ORDER BY e.hotel_id, e.creado_en DESC
    """, hotel_ids + [mes_str]))

    # ── 4. Agrupar por hotel en memoria ──────────────────────────────────────
    from collections import defaultdict
    pedidos_por_hotel: dict = defaultdict(list)
    for p in pedidos_mes:
        pedidos_por_hotel[p["hotel_id"]].append(p)

    expedientes_por_hotel: dict = defaultdict(list)
    for e in expedientes_mes:
        expedientes_por_hotel[e["hotel_id"]].append(e)

    # ── 5. Construir resultado ────────────────────────────────────────────────
    resultado = []
    for hotel in hoteles:
        pedidos         = pedidos_por_hotel[hotel["id"]]
        acumulado       = sum(float(p["importe"] or 0) for p in pedidos)
        num_pedidos     = len(pedidos)
        familias_usadas = [p["familia_nombre"] for p in pedidos if p["familia_nombre"]]

        familias_conteo: dict = {}
        for p in pedidos:
            fn = p["familia_nombre"]
            if fn:
                familias_conteo[fn] = familias_conteo.get(fn, 0) + 1

        familias_importe: dict = {}
        for p in pedidos:
            fn = p["familia_nombre"]
            if fn:
                familias_importe[fn] = familias_importe.get(fn, 0) + float(p["importe"] or 0)

        _exps_hotel   = expedientes_por_hotel[hotel["id"]]
        _pendientes   = [e for e in _exps_hotel if e["resultado"] == "pendiente"]
        _aprobados    = [e for e in _exps_hotel if e["resultado"] == "aprobado"]
        pendiente_dg_importe      = sum(float(e["importe_pedido"] or 0) for e in _pendientes)
        exceso_autorizado_importe = sum(float(e["importe_pedido"] or 0) for e in _aprobados)
        compromiso_potencial      = acumulado + pendiente_dg_importe

        if acumulado >= techo_max_mes or num_pedidos > techo_max_ped_n:
            semaforo = "rojo"
        elif acumulado >= umbral_amarillo or num_pedidos >= techo_max_ped_n:
            semaforo = "amarillo"
        else:
            semaforo = "verde"
        if _pendientes:
            semaforo = "azul"

        resultado.append({
            "hotel_id":        hotel["id"],
            "hotel_codigo":    hotel["codigo"],
            "hotel_nombre":    hotel["nombre"],
            "num_pedidos":     num_pedidos,
            "max_pedidos":     techo_max_ped_n,
            "acumulado":       acumulado,
            "techo_mes":       techo_max_mes,
            "techo_pedido":    techo_max_pedido,
            "familias_usadas": familias_usadas,
            "familias_conteo": familias_conteo,
            "max_pedidos_familia": techo_max_ped_fam,
            "familias_importe": familias_importe,
            "max_importe_familia": techo_max_mes_fam,
            "semaforo":        semaforo,
            "pendientes_dg":         _pendientes,
            "pendientes_dg_importe": pendiente_dg_importe,
            "excesos_autorizados":         _aprobados,
            "excesos_autorizados_importe": exceso_autorizado_importe,
            "compromiso_potencial":  compromiso_potencial,
            "pedidos":         pedidos,
        })

    return jsonify({"mes": f"{year}-{month:02d}", "hoteles": resultado, "historico": True})


@app.route("/api/expedientes")
@login_required
def listar_expedientes():
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 4) Histórico completo de
    expedientes de exceso de techo (Sección 9 del documento de diseño) —
    nunca se borra, solo se consulta. Filtros opcionales por querystring:
    hotel_id, familia_id, resultado (pendiente/aprobado/denegado), mes
    (YYYY-MM). Sin filtros, devuelve todos (ordenados del más reciente al
    más antiguo) — el frontend (Fase 6) decidirá si pagina.
    """
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403

    where   = ["1=1"]
    args: list = []

    hotel_id = request.args.get("hotel_id")
    if hotel_id:
        where.append("e.hotel_id = %s")
        args.append(hotel_id)

    familia_id = request.args.get("familia_id")
    if familia_id:
        where.append("e.familia_id = %s")
        args.append(familia_id)

    resultado_f = request.args.get("resultado")
    if resultado_f in ("pendiente", "aprobado", "denegado"):
        where.append("e.resultado = %s")
        args.append(resultado_f)

    mes_f = request.args.get("mes")
    if mes_f:
        where.append("e.mes = %s")
        args.append(mes_f)

    expedientes = rows_to_list(query(f"""
        SELECT e.*, 
               h.codigo AS hotel_codigo, h.nombre AS hotel_nombre,
               f.nombre AS familia_nombre,
               p.pedido_num, p.norden, p.estado AS pedido_estado,
               us.nombre AS usuario_solicitante_nombre,
               ur.nombre AS usuario_resuelve_nombre
        FROM expediente_exceso e
        LEFT JOIN hoteles   h  ON e.hotel_id              = h.id
        LEFT JOIN familias  f  ON e.familia_id            = f.id
        LEFT JOIN pedidos   p  ON e.pedido_id              = p.id
        LEFT JOIN usuarios  us ON e.usuario_solicitante_id = us.id
        LEFT JOIN usuarios  ur ON e.usuario_resuelve_id    = ur.id
        WHERE {' AND '.join(where)}
        ORDER BY e.creado_en DESC
    """, tuple(args)))

    return jsonify({"expedientes": expedientes})


@app.route("/api/expedientes/exportar")
@login_required
def exportar_expedientes_excel():
    """
    (2026-09-01, repaso "agilizar y limpiar", Etapa 3) Excel profesional con
    el histórico completo de expedientes de exceso de techo — a petición de
    Víctor, para poder consultar este histórico "en cualquier momento" sin
    depender de una pantalla que nunca se llegó a construir (Fase 6). Mismos
    filtros opcionales que listar_expedientes() (hotel_id/familia_id/
    resultado/mes) por si algún día hace falta un export acotado, pero sin
    filtros — el uso normal del botón — exporta el histórico entero, que es
    justo lo que se pidió.

    Mismo patrón que exportar_excel() (Excel de Pedidos): cabecera con el
    color de marca de la app (#1a3a6b), filas coloreadas según su estado
    (mismo criterio semáforo verde/amarillo/rojo que ya usa Techo de Gastos
    en pantalla), formato de moneda en los importes, fila de totales al
    final y auto-filtro para poder acotar por hotel/familia/resultado ya
    dentro del propio Excel.
    """
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403

    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from flask import send_file

        where   = ["1=1"]
        args: list = []

        hotel_id = request.args.get("hotel_id")
        if hotel_id:
            where.append("e.hotel_id = %s")
            args.append(hotel_id)

        familia_id = request.args.get("familia_id")
        if familia_id:
            where.append("e.familia_id = %s")
            args.append(familia_id)

        resultado_f = request.args.get("resultado")
        if resultado_f in ("pendiente", "aprobado", "denegado"):
            where.append("e.resultado = %s")
            args.append(resultado_f)

        mes_f = request.args.get("mes")
        if mes_f:
            where.append("e.mes = %s")
            args.append(mes_f)

        expedientes = rows_to_list(query(f"""
            SELECT e.*,
                   h.codigo AS hotel_codigo, h.nombre AS hotel_nombre,
                   f.nombre AS familia_nombre,
                   p.pedido_num, p.norden, p.estado AS pedido_estado,
                   us.nombre AS usuario_solicitante_nombre,
                   ur.nombre AS usuario_resuelve_nombre
            FROM expediente_exceso e
            LEFT JOIN hoteles   h  ON e.hotel_id              = h.id
            LEFT JOIN familias  f  ON e.familia_id            = f.id
            LEFT JOIN pedidos   p  ON e.pedido_id              = p.id
            LEFT JOIN usuarios  us ON e.usuario_solicitante_id = us.id
            LEFT JOIN usuarios  ur ON e.usuario_resuelve_id    = ur.id
            WHERE {' AND '.join(where)}
            ORDER BY e.creado_en DESC
        """, tuple(args)))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EXPEDIENTES TECHO"

        HEADERS = [
            "MES", "HOTEL", "FAMILIA", "Nº PEDIDO", "ESTADO PEDIDO",
            "IMPORTE PEDIDO", "CONSUMO PREVIO", "EXCESO",
            "CONSUMIDO EN SOLICITUD", "DISPONIBLE EN SOLICITUD",
            "MOTIVO SOLICITUD", "SOLICITADO POR", "RESULTADO",
            "RESUELTO POR", "FECHA RESOLUCIÓN",
            "OBSERVACIONES DIRECCIÓN GENERAL", "CREADO EN",
        ]
        ws.append(HEADERS)
        header_fill = PatternFill("solid", fgColor="1a3a6b")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

        # Mismo criterio semáforo que ya usa Techo de Gastos en pantalla
        # (verde=aprobado, amarillo=pendiente, rojo=denegado) — y misma
        # paleta que ya usa exportar_excel() para ENTREGADO/CANCELADO, para
        # que los dos Excel de la app se vean coherentes entre sí.
        RESULTADO_COLORES = {
            "aprobado":  "d4edda",
            "pendiente": "fff3cd",
            "denegado":  "f8d7da",
        }
        MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
                    "Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

        def strip_tz(val):
            if hasattr(val, "tzinfo") and val.tzinfo is not None:
                return val.replace(tzinfo=None)
            return val

        def mes_legible(mes_str):
            # mes_str viene como "YYYY-MM" (ver expediente_exceso.mes) —
            # se traduce a "Agosto 2026" para que el Excel sea legible sin
            # tener que descifrar el formato de la base de datos.
            if not mes_str or "-" not in str(mes_str):
                return mes_str or "—"
            try:
                anio, mes_n = str(mes_str).split("-")
                return f"{MESES_ES[int(mes_n) - 1]} {anio}"
            except (ValueError, IndexError):
                return mes_str

        EUR = '#,##0.00 "€"'
        FECHA = "DD/MM/YYYY"

        totales = {"importe_pedido": 0, "exceso": 0}
        conteo_resultado = {"aprobado": 0, "pendiente": 0, "denegado": 0}

        for e in expedientes:
            resultado = (e.get("resultado") or "pendiente").lower()
            conteo_resultado[resultado] = conteo_resultado.get(resultado, 0) + 1
            totales["importe_pedido"] += float(e.get("importe_pedido") or 0)
            totales["exceso"]         += float(e.get("exceso") or 0)

            hotel_txt = f"{e.get('hotel_codigo') or ''} — {e.get('hotel_nombre') or ''}".strip(" —")
            fila = [
                mes_legible(e.get("mes")),
                hotel_txt or "—",
                e.get("familia_nombre") or "—",
                e.get("pedido_num") or (f"#{e.get('norden')}" if e.get("norden") else "—"),
                e.get("pedido_estado") or "—",
                float(e.get("importe_pedido") or 0),
                float(e.get("consumo_previo") or 0),
                float(e.get("exceso") or 0),
                float(e.get("consumido_en_solicitud") or 0),
                float(e.get("disponible_en_solicitud") or 0),
                e.get("motivo_solicitud") or "",
                e.get("usuario_solicitante_nombre") or "—",
                resultado.upper(),
                e.get("usuario_resuelve_nombre") or "—",
                strip_tz(e.get("fecha_resolucion")),
                e.get("observaciones_direccion_general") or "",
                strip_tz(e.get("creado_en")),
            ]
            ws.append(fila)
            row = ws.max_row
            fill = PatternFill("solid", fgColor=RESULTADO_COLORES.get(resultado, "FFFFFF"))
            for cell in ws[row]:
                cell.fill = fill
            for col in (6, 7, 8, 9, 10):
                ws.cell(row=row, column=col).number_format = EUR
            for col in (15, 17):
                ws.cell(row=row, column=col).number_format = FECHA
            ws.cell(row=row, column=11).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=16).alignment = Alignment(wrap_text=True, vertical="top")

        # Fila de totales — separada con un borde superior para que no se
        # confunda con una fila de datos más.
        if expedientes:
            fila_totales = ws.max_row + 1
            ws.cell(row=fila_totales, column=1, value=f"TOTAL ({len(expedientes)} expedientes: "
                    f"{conteo_resultado.get('aprobado',0)} aprobados, "
                    f"{conteo_resultado.get('pendiente',0)} pendientes, "
                    f"{conteo_resultado.get('denegado',0)} denegados)")
            ws.cell(row=fila_totales, column=6, value=totales["importe_pedido"]).number_format = EUR
            ws.cell(row=fila_totales, column=8, value=totales["exceso"]).number_format = EUR
            top_border = Border(top=Side(style="thin"))
            bold = Font(bold=True)
            for cell in ws[fila_totales]:
                cell.border = top_border
                cell.font = bold

        COL_WIDTHS = [16, 26, 20, 14, 20, 16, 16, 14, 18, 18, 34, 20, 12, 20, 16, 34, 16]
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        if expedientes:
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}{ws.max_row - 1}"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"EXPEDIENTES_TECHO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({"error": "openpyxl no instalado"}), 500


@app.route("/api/expedientes/pedido/<int:pedido_id>")
@login_required
def historico_expedientes_pedido(pedido_id):
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 4, punto 11) Histórico
    cronológico de TODOS los intentos de un pedido concreto — cada
    reintento tras una denegación es una fila independiente en
    expediente_exceso (nunca se sobrescribe, ver Fase 1/2), así que basta
    con listarlas todas por pedido_id, ordenadas por creado_en, para
    reconstruir la cronología completa (solicitud → denegación → nueva
    solicitud → aprobación, etc.) sin ninguna tabla ni cálculo adicional.
    """
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403

    expedientes = rows_to_list(query("""
        SELECT e.*,
               f.nombre AS familia_nombre,
               us.nombre AS usuario_solicitante_nombre,
               ur.nombre AS usuario_resuelve_nombre
        FROM expediente_exceso e
        LEFT JOIN familias f  ON e.familia_id            = f.id
        LEFT JOIN usuarios us ON e.usuario_solicitante_id = us.id
        LEFT JOIN usuarios ur ON e.usuario_resuelve_id    = ur.id
        WHERE e.pedido_id = %s
        ORDER BY e.creado_en ASC
    """, (pedido_id,)))

    return jsonify({"pedido_id": pedido_id, "expedientes": expedientes})


@app.route("/api/expedientes/<int:eid>/informe")
@login_required
def informe_expediente(eid):
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 5) Datos completos para el
    informe imprimible de un expediente (Sección 10 del documento de
    diseño). Todo lo que necesita el frontend en una sola llamada:

    - El propio expediente, CON su fotografía presupuestaria congelada
      (consumido_en_solicitud / disponible_en_solicitud / importe_pedido /
      consumo_previo / exceso) — estos valores NUNCA se recalculan aquí
      (punto 10): el informe siempre refleja la situación que había en el
      momento exacto de la solicitud, aunque otros pedidos hayan consumido
      techo después. Si se quisiera la situación EN VIVO, para eso está
      /api/techo/resumen, no este endpoint.
    - Datos del pedido asociado (hotel, departamento, proveedor, importe).
    - Histórico cronológico de TODOS los intentos de este mismo pedido
      (reintentos tras denegación) — mismo dato que
      /api/expedientes/pedido/<id>, incluido aquí también para no
      necesitar una segunda llamada al montar el informe.
    - Histórico de excesos anteriores del mismo hotel + familia (otros
      pedidos, ya resueltos) — contexto para quien tenga que decidir:
      "¿esto ya ha pasado antes en este hotel/familia?".
    """
    if session.get("rol") == "hotel":
        return jsonify({"error": "Sin permisos"}), 403

    expediente = row_to_dict(query("""
        SELECT e.*, h.codigo AS hotel_codigo, h.nombre AS hotel_nombre,
               f.nombre AS familia_nombre,
               us.nombre AS usuario_solicitante_nombre,
               ur.nombre AS usuario_resuelve_nombre
        FROM expediente_exceso e
        LEFT JOIN hoteles  h  ON e.hotel_id              = h.id
        LEFT JOIN familias f  ON e.familia_id            = f.id
        LEFT JOIN usuarios us ON e.usuario_solicitante_id = us.id
        LEFT JOIN usuarios ur ON e.usuario_resuelve_id    = ur.id
        WHERE e.id = %s
    """, (eid,), one=True))
    if not expediente:
        return jsonify({"error": "Expediente no encontrado"}), 404

    pedido = row_to_dict(query(f"{PEDIDO_SELECT} WHERE p.id=%s", (expediente["pedido_id"],), one=True))

    historico_pedido = rows_to_list(query("""
        SELECT e.*, us.nombre AS usuario_solicitante_nombre, ur.nombre AS usuario_resuelve_nombre
        FROM expediente_exceso e
        LEFT JOIN usuarios us ON e.usuario_solicitante_id = us.id
        LEFT JOIN usuarios ur ON e.usuario_resuelve_id    = ur.id
        WHERE e.pedido_id = %s
        ORDER BY e.creado_en ASC
    """, (expediente["pedido_id"],)))

    # Excesos anteriores del mismo hotel+familia, ya resueltos (contexto) —
    # familia_id puede ser NULL, así que la comparación se arma aparte.
    if expediente.get("familia_id"):
        fam_clause = "e.familia_id = %s"
        fam_args   = (expediente["familia_id"],)
    else:
        fam_clause = "e.familia_id IS NULL"
        fam_args   = ()
    historico_hotel_familia = rows_to_list(query(f"""
        SELECT e.*, p.pedido_num,
               us.nombre AS usuario_solicitante_nombre, ur.nombre AS usuario_resuelve_nombre
        FROM expediente_exceso e
        LEFT JOIN pedidos  p  ON e.pedido_id              = p.id
        LEFT JOIN usuarios us ON e.usuario_solicitante_id = us.id
        LEFT JOIN usuarios ur ON e.usuario_resuelve_id    = ur.id
        WHERE e.hotel_id = %s AND {fam_clause}
          AND e.id != %s
          AND e.resultado != 'pendiente'
        ORDER BY e.creado_en DESC
        LIMIT 20
    """, (expediente["hotel_id"],) + fam_args + (eid,)))

    return jsonify({
        "expediente":              expediente,
        "pedido":                  pedido,
        "historico_pedido":        historico_pedido,
        "historico_hotel_familia": historico_hotel_familia,
    })


# ── API Pedidos ────────────────────────────────────────────────────────────────

# ── Lógica de clasificación de alertas — fuente única de verdad ──────────────
#
# Tres consumidores usaban copias idénticas de esta lógica:
#   • /api/stats        (bloque rol=hotel)
#   • /api/stats        (bloque resto de roles)
#   • /api/bridge/alertas
#
# Extraída aquí para que cualquier cambio de umbral o regla se aplique
# en los tres sitios sin riesgo de desincronización.
# ─────────────────────────────────────────────────────────────────────────────

from datetime import date as _date_alerta, datetime as _dt_alerta


def _dias_desde_alerta(fecha_str) -> int | None:
    """Días transcurridos desde fecha_str hasta hoy. None si no parseable."""
    if not fecha_str:
        return None
    try:
        if hasattr(fecha_str, "date"):
            f = fecha_str.date()
        elif isinstance(fecha_str, _date_alerta):
            f = fecha_str
        else:
            f = _dt_alerta.strptime(str(fecha_str)[:10], "%Y-%m-%d").date()
        return (_date_alerta.today() - f).days
    except Exception:
        return None


# Umbrales por estado (días desde fecha de referencia).
# "primera": días mínimos para emitir aviso.
# "urgente": días para escalar a urgente (None = nunca).
# "ciclo":   cada cuántos días se reavisa (no usado en clasificación actual).
# "fecha_ref": campo de fecha a usar (default "fecha_tramitacion").
#
# NOTA v12.5.0 — FIX de inconsistencia: hasta ahora esta clasificación usaba
# un dict fijo en código (_UMBRALES_ALERTAS), mientras que el resto de la app
# (cambio de estado, job diario del scheduler) ya usaba _build_umbrales(),
# que lee de la tabla config_alertas (editable desde el panel Admin). Esto
# provocaba que cambiar un umbral en Admin no afectase a los popups de
# Agenda. Se elimina el dict fijo y se usa _build_umbrales() también aquí,
# para que Admin sea la única fuente de verdad en todos los canales.

# Mapeo estado de pedido → prefijo de claves de config_alertas para la
# repetición de popups en Agenda (grupo "popup_repeticion").
_ESTADO_POPUP_PREFIX: dict = {
    "ENVIADO AL PROVEEDOR":               "enviado",
    "PENDIENTE FIRMA DIRECCION COMPRAS":  "firma_compras",
    "PENDIENTE DE FIRMA DIRECCION HOTEL": "firma_hotel",
    "ENTREGA PARCIAL":                    "entrega_parcial",
    "PENDIENTE COTIZACIÓN":               "cotizacion",
}


def _aplicar_config_popup(p: dict) -> None:
    """Añade al pedido los campos de repetición de popup (leídos de Admin):
        popup_repetir        (bool)  — si False, el popup se muestra 1 sola vez
        popup_horas_critico  (int)   — cada cuántas horas se repite si 🔴 urgente
        popup_horas_normal   (int)   — cada cuántas horas se repite si 🟡 aviso
    Consumido por pedidos_agenda_bridge.py (Organizador Princess).
    """
    prefix = _ESTADO_POPUP_PREFIX.get(p.get("estado"))
    c = get_config()
    if prefix:
        p["popup_repetir"]       = bool(int(c.get(f"{prefix}_popup_repetir", 1) or 0))
        p["popup_horas_critico"] = int(c.get(f"{prefix}_popup_horas_critico", 1) or 1)
        p["popup_horas_normal"]  = int(c.get(f"{prefix}_popup_horas_normal", 24) or 24)
    else:
        p["popup_repetir"]       = True
        p["popup_horas_critico"] = 1
        p["popup_horas_normal"]  = 24


def _resumen_ultima_notificacion(p: dict) -> dict:
    """
    A partir de ultima_notif_email / ultima_notif_telegram (subconsultas de
    PEDIDO_SELECT_STATS) calcula un resumen único para mostrar en el panel
    de Alertas: cuándo se notificó por última vez esta alerta y por qué canal.

    Devuelve:
        {
          "fecha":   str ISO o None,   # la más reciente entre email y Telegram
          "canales": list[str],        # ["Email"], ["Telegram"] o ["Email","Telegram"]
          "dias":    int|None,         # días transcurridos desde esa notificación
        }
    """
    ultima_email = p.get("ultima_notif_email")
    ultima_tg    = p.get("ultima_notif_telegram")
    candidatas   = [v for v in (ultima_email, ultima_tg) if v]

    # ── Reclamación automática al proveedor — indicador aparte (2026-07-30) ──
    # Se muestra separado del resto (no mezclado en "canales") para que se
    # vea a simple vista si YA se reclamó automáticamente hoy o hace poco,
    # y así evitar que alguien mande una reclamación manual duplicando una
    # que el sistema ya envió solo.
    reclamacion_auto = None
    ultima_reclamacion = p.get("ultima_reclamacion_auto")
    if ultima_reclamacion:
        try:
            fecha_ref_r = ultima_reclamacion.date() if hasattr(ultima_reclamacion, "date") else ultima_reclamacion
            dias_r = (_date_alerta.today() - fecha_ref_r).days
        except Exception:
            dias_r = None
        try:
            fecha_iso_r = ultima_reclamacion.isoformat()
        except Exception:
            fecha_iso_r = str(ultima_reclamacion)
        reclamacion_auto = {"fecha": fecha_iso_r, "dias": dias_r}

    if not candidatas:
        return {"fecha": None, "canales": [], "dias": None, "reclamacion_auto": reclamacion_auto}

    fecha_max = max(candidatas)
    canales = []
    if ultima_email:
        canales.append("Email")
    if ultima_tg:
        canales.append("Telegram")

    try:
        fecha_ref = fecha_max.date() if hasattr(fecha_max, "date") else fecha_max
        dias = (_date_alerta.today() - fecha_ref).days
    except Exception:
        dias = None

    try:
        fecha_iso = fecha_max.isoformat()
    except Exception:
        fecha_iso = str(fecha_max)

    return {"fecha": fecha_iso, "canales": canales, "dias": dias, "reclamacion_auto": reclamacion_auto}


def _clasificar_alertas(pedidos_raw: list, cfg_activar_plazo: bool) -> list:
    """Clasifica una lista de pedidos y devuelve solo los que generan alerta.

    Para cada pedido:
      1. Si tiene plazo_entrega_dias y cfg_activar_plazo=True, aplica la
         lógica de _alertas_plazo_entrega (fecha de entrega esperada).
      2. Si no, aplica la lógica estándar de _UMBRALES_ALERTAS.

    Añade a cada pedido:
      • dias_tramitacion  (int)
      • nivel_alerta      ("aviso" | "urgente")
      • fecha_entrega_prevista (str ISO o None)
      • ultima_notificacion    (dict — ver _resumen_ultima_notificacion)
      • popup_repetir, popup_horas_critico, popup_horas_normal
                          (ver _aplicar_config_popup — config editable desde Admin)

    Devuelve la lista ordenada: urgentes primero, luego por días descendente.
    """
    alertas: list = []
    for p in pedidos_raw:
        _normalizar_fecha_entrega_especifica(p)
        # ── Lógica plazo de entrega ──────────────────────────────────────
        # v12.29.52 — mismo fix que en _job_alertas_diarias_inner: un
        # pedido con fecha_entrega_especifica/plazo_entrega_dias informado
        # se evalúa SIEMPRE por esta vía (con _debe_usar_logica_plazo), y
        # si hoy no le toca ningún aviso por esta vía, sencillamente no
        # genera alerta — nunca cae a la lógica estándar de "días desde
        # fecha_tramitacion", que ignoraría una fecha de entrega concreta
        # todavía lejana y mostraría el pedido como urgente sin serlo.
        if _debe_usar_logica_plazo(p):
            info_plazo = _alertas_plazo_entrega(p, cfg_activar_plazo)
            if not info_plazo:
                continue
            dias = _dias_desde_alerta(p.get("fecha_tramitacion")) or 0
            p["dias_tramitacion"]      = dias
            p["nivel_alerta"]          = info_plazo["nivel"]
            fep = info_plazo["fecha_entrega_prevista"]
            p["fecha_entrega_prevista"] = fep.strftime("%Y-%m-%d") if fep else None
            p["ultima_notificacion"]   = _resumen_ultima_notificacion(p)
            _aplicar_config_popup(p)
            alertas.append(p)
            continue
        # ── Lógica estándar ─────────────────────────────────────────────
        cfg = _build_umbrales().get(p["estado"])
        if not cfg:
            continue
        fecha_ref_campo = cfg.get("fecha_ref", "fecha_tramitacion")
        dias = _dias_desde_alerta(p.get(fecha_ref_campo))
        if dias is None or dias < cfg["primera"]:
            continue
        nivel = "urgente" if (cfg["urgente"] and dias >= cfg["urgente"]) else "aviso"
        p["dias_tramitacion"]      = dias
        p["nivel_alerta"]          = nivel
        p["fecha_entrega_prevista"] = None
        p["ultima_notificacion"]   = _resumen_ultima_notificacion(p)
        _aplicar_config_popup(p)
        alertas.append(p)

    alertas.sort(key=lambda x: (0 if x["nivel_alerta"] == "urgente" else 1,
                                 -x["dias_tramitacion"]))
    return alertas


# ── Selector reducido para /api/stats (alertas del dashboard) ────────────────
# Solo los campos que loadAlertas() y updateAlertBadge() consumen.
# Sin subconsultas a proveedor_contactos — esos datos solo hacen falta al
# abrir el modal de email/telegram de una alerta concreta (usa PEDIDO_SELECT_ALERTA).
PEDIDO_SELECT_STATS = """
    SELECT p.id, p.norden, p.pedido_num, p.estado,
           p.fecha_tramitacion, p.fecha_solicitud,
           p.plazo_entrega_dias, p.fecha_entrega_especifica, p.observaciones, p.importe,
           h.codigo  as hotel_codigo,
           h.nombre  as hotel_nombre,
           d.nombre  as departamento_nombre,
           pr.nombre as proveedor_nombre,
           f.nombre  as familia_nombre,
           EXISTS (
               SELECT 1 FROM pedido_adjuntos pa WHERE pa.pedido_id = p.id
           ) AS has_adjuntos,
           -- (2026-08-06) FIX: antes solo miraba emails_log (envíos MANUALES,
           -- botón "Notificar"/"Re-notificar"). Todos los correos automáticos
           -- (reclamación al proveedor, aviso de firma pendiente, aviso de
           -- cotización sin proveedor...) se encolan y despachan vía
           -- emails_sistema_pendientes — una tabla distinta que esta
           -- subconsulta nunca miraba. Resultado: un pedido que solo había
           -- recibido avisos automáticos (nunca un clic manual) se quedaba
           -- marcado "Sin notificar" para siempre en el panel de Alertas,
           -- aunque el correo hubiera salido de verdad — confirmado con un
           -- correo real recibido por el usuario que la pantalla no reflejaba.
           -- GREATEST() combina ambas fuentes; enviado_en (no creado_en) es
           -- el momento real de envío en emails_sistema_pendientes, no el de
           -- encolado.
           GREATEST(
               (SELECT MAX(el.creado_en) FROM emails_log el
                  WHERE el.pedido_id = p.id
                    AND el.tipo IN ('alerta_proveedor','alerta_interno')),
               (SELECT MAX(esp.enviado_en) FROM emails_sistema_pendientes esp
                  WHERE esp.pedido_id = p.id AND esp.enviado = TRUE)
           ) AS ultima_notif_email,
           (SELECT MAX(wl.creado_en) FROM whatsapp_log wl
              WHERE wl.pedido_id = p.id
                AND wl.tipo = 'telegram_auto' AND wl.enviado = 1) AS ultima_notif_telegram,
           (SELECT MAX(wl2.creado_en) FROM whatsapp_log wl2
              WHERE wl2.pedido_id = p.id
                AND wl2.tipo = 'reclamacion_proveedor_auto' AND wl2.enviado = 1) AS ultima_reclamacion_auto
    FROM pedidos p
    LEFT JOIN hoteles       h  ON p.hotel_id        = h.id
    LEFT JOIN departamentos d  ON p.departamento_id = d.id
    LEFT JOIN proveedores   pr ON p.proveedor_id    = pr.id
    LEFT JOIN familias      f  ON p.familia_id      = f.id
"""

PEDIDO_SELECT = """
    SELECT p.*,
           h.codigo  as hotel_codigo,
           h.nombre  as hotel_nombre,
           d.nombre  as departamento_nombre,
           pr.nombre as proveedor_nombre,
           (SELECT pc.email
              FROM proveedor_contactos pc
             WHERE pc.proveedor_id = pr.id AND pc.es_principal = 1
               AND pc.email IS NOT NULL AND pc.email != ''
               AND (EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch WHERE pch.contacto_id = pc.id AND pch.hotel_id = p.hotel_id)
                    OR NOT EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch2 WHERE pch2.contacto_id = pc.id))
             ORDER BY EXISTS (SELECT 1 FROM proveedor_contacto_hoteles pch3 WHERE pch3.contacto_id = pc.id AND pch3.hotel_id = p.hotel_id) DESC,
                      pc.orden, pc.id
             LIMIT 1) as proveedor_email,
           (SELECT telefono FROM proveedor_contactos WHERE proveedor_id=pr.id AND telefono IS NOT NULL AND telefono!='' ORDER BY orden,id LIMIT 1) as proveedor_telefono,
           (SELECT nombre FROM proveedor_contactos WHERE proveedor_id=pr.id ORDER BY orden,id LIMIT 1) as proveedor_contacto,
           COALESCE(p.creado_por_nombre,    u1.nombre) as creado_por_nombre,
           COALESCE(p.modificado_por_nombre, u2.nombre) as modificado_por_nombre,
           f.nombre  as familia_nombre,
           EXISTS (
               SELECT 1 FROM pedido_adjuntos pa WHERE pa.pedido_id = p.id
           ) AS has_adjuntos
    FROM pedidos p
    LEFT JOIN hoteles       h  ON p.hotel_id          = h.id
    LEFT JOIN departamentos d  ON p.departamento_id   = d.id
    LEFT JOIN proveedores   pr ON p.proveedor_id      = pr.id
    LEFT JOIN usuarios      u1 ON p.creado_por_id     = u1.id
    LEFT JOIN usuarios      u2 ON p.modificado_por_id = u2.id
    LEFT JOIN familias      f  ON p.familia_id        = f.id
"""

@app.route("/api/pedidos")
@login_required
def get_pedidos():
    wheres, args = [], []

    # Restricción por rol hotel: solo ve sus hoteles asignados
    if session.get("rol") == "hotel":
        hoteles_ids = session.get("hoteles_ids", [])
        if not hoteles_ids:
            return jsonify({"pedidos": [], "total": 0, "page": 1, "page_size": 20, "pages": 1})
        placeholders = ",".join(["%s"] * len(hoteles_ids))
        wheres.append(f"p.hotel_id IN ({placeholders})")
        args += hoteles_ids

    # Hotel de pruebas ('PR'): invisible salvo para admin o el usuario
    # dedicado a estas pruebas.
    if not _puede_ver_hotel_pruebas():
        wheres.append("(h.codigo IS NULL OR h.codigo <> %s)")
        args.append(HOTEL_CODIGO_PRUEBAS)

    q           = request.args.get("q", "").strip()
    hotel       = request.args.get("hotel_id", "")
    estado      = request.args.get("estado", "")
    depto       = request.args.get("departamento_id", "")
    alerta      = request.args.get("alerta", "")
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()

    if q:
        wheres.append("(p.pedido_num ILIKE %s OR pr.nombre ILIKE %s OR p.observaciones ILIKE %s OR h.codigo ILIKE %s)")
        args += [f"%{q}%"] * 4
    if hotel:
        wheres.append("p.hotel_id = %s"); args.append(hotel)
    if estado:
        wheres.append("p.estado = %s"); args.append(estado)
    if depto:
        wheres.append("p.departamento_id = %s"); args.append(depto)
    if fecha_desde:
        wheres.append("p.fecha_solicitud >= %s"); args.append(fecha_desde)
    if fecha_hasta:
        wheres.append("p.fecha_solicitud <= %s"); args.append(fecha_hasta)
    if alerta == "1":
        # Filtro rápido: pedidos con fecha_tramitacion y estado activo
        # (el cálculo exacto de días y nivel se hace en /api/stats)
        wheres.append("""
            p.estado IN ('ENVIADO AL PROVEEDOR','PENDIENTE FIRMA DIRECCION COMPRAS',
                         'PENDIENTE DE FIRMA DIRECCION HOTEL','ENTREGA PARCIAL')
            AND p.fecha_tramitacion IS NOT NULL
        """)

    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""

    order_map = {
        "fecha_desc": "p.creado_en DESC",
        "fecha_asc":  "p.creado_en ASC",
        "estado":     "p.estado, p.creado_en DESC",
        "hotel":      "h.codigo, p.creado_en DESC",
    }
    order = order_map.get(request.args.get("orden", ""), "p.norden DESC")

    try:
        page      = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(100, int(request.args.get("page_size", 20))))
    except ValueError:
        page, page_size = 1, 20

    count_sql = f"""SELECT COUNT(*) as total FROM pedidos p
                    LEFT JOIN hoteles h ON p.hotel_id=h.id
                    LEFT JOIN proveedores pr ON p.proveedor_id=pr.id {where_sql}"""
    total = query(count_sql, args, one=True)["total"]

    sql     = f"{PEDIDO_SELECT} {where_sql} ORDER BY {order} LIMIT %s OFFSET %s"
    pedidos = rows_to_list(query(sql, args + [page_size, (page - 1) * page_size]))
    for _p in pedidos:
        _normalizar_fecha_entrega_especifica(_p)
    if session.get("rol") == "hotel":
        for p in pedidos:
            p["importe"] = None

    return jsonify({
        "pedidos":   pedidos,
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "pages":     max(1, (total + page_size - 1) // page_size),
    })

@app.route("/api/pedidos/<int:pid>")
@login_required
def get_pedido(pid):
    p = row_to_dict(query(f"{PEDIDO_SELECT} WHERE p.id=%s", (pid,), one=True))
    if not p:
        return jsonify({"error": "No encontrado"}), 404
    _normalizar_fecha_entrega_especifica(p)
    if not _puede_ver_hotel_pruebas() and p.get("hotel_codigo") == HOTEL_CODIGO_PRUEBAS:
        return jsonify({"error": "Sin acceso a este pedido"}), 403
    if session.get("rol") == "hotel":
        hoteles_ids = session.get("hoteles_ids", [])
        if p.get("hotel_id") not in hoteles_ids:
            return jsonify({"error": "Sin acceso a este pedido"}), 403
        p["importe"] = None
    historial = rows_to_list(query(
        """SELECT h.*, COALESCE(h.usuario_nombre, u.nombre) as usuario_nombre
           FROM historial_estados h LEFT JOIN usuarios u ON h.usuario_id=u.id
           WHERE h.pedido_id=%s ORDER BY h.creado_en DESC""", (pid,)
    ))
    return jsonify({"pedido": p, "historial": historial})

@app.route("/api/pedidos", methods=["POST"])
@login_required
def create_pedido():
    data   = request.get_json(silent=True) or {}
    if not _puede_ver_hotel_pruebas() and _es_hotel_pruebas_id(data.get("hotel_id")):
        return jsonify({"error": "Hotel no disponible"}), 403
    db     = get_db()
    uid    = current_user_id()
    norden = _next_norden(db)
    estado = data.get("estado", "PENDIENTE FIRMA DIRECCION COMPRAS")

    sujeto_techo = 1 if data.get("sujeto_techo") else 0
    familia_id   = data.get("familia_id") or None
    importe      = data.get("importe") or None

    # (2026-08-01 — rediseño Techo de Gastos) Ya NO se comprueba el techo al
    # crear un pedido — el momento real de "consumo" es al pasar a ENVIADO
    # AL PROVEEDOR (ver update_pedido()), no al crear/editar. _forzar_techo
    # queda eliminado, sustituido por el circuito de autorización real.

    cur = execute("""
        INSERT INTO pedidos (
            norden, hotel_id, departamento_id,
            fecha_solicitud, fecha_envio_visto_bueno, fecha_tramitacion,
            pedido_num, presupuesto_num, entrada_albaran_num,
            tarifa_acordada,
            estado, comunicado_ab, comunicado_jefe_dep,
            parte_rotura, parte_ampliacion,
            proveedor_id, observaciones,
            familia_id, importe, sujeto_techo,
            plazo_entrega_dias, fecha_entrega_especifica,
            total_pedido,
            creado_por_id, modificado_por_id,
            creado_por_nombre, modificado_por_nombre
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    """, (
        norden,
        data.get("hotel_id"), data.get("departamento_id"),
        data.get("fecha_solicitud"), data.get("fecha_envio_visto_bueno"),
        data.get("fecha_tramitacion"),
        # (2026-08-28) pedido_num y total_pedido NUNCA se aceptan del
        # cliente al crear: solo se rellenan al subir el PDF de pedido
        # oficial (ver upload_adjunto/_parsear_pdf_pedido_oficial), y eso
        # requiere un pedido_id que todavía no existe en este punto — por
        # eso siempre nacen a NULL, sea lo que sea lo que mande el formulario.
        None, data.get("presupuesto_num"),
        data.get("entrada_albaran_num"),
        bool(data.get("tarifa_acordada")),
        estado,
        1 if data.get("comunicado_ab") else 0,
        1 if data.get("comunicado_jefe_dep") else 0,
        1 if data.get("parte_rotura") else 0,
        1 if data.get("parte_ampliacion") else 0,
        data.get("proveedor_id"), data.get("observaciones"),
        familia_id, importe, sujeto_techo,
        data.get("plazo_entrega_dias") or None,
        data.get("fecha_entrega_especifica") or None,
        None,
        uid, uid,
        session.get("nombre"), session.get("nombre"),
    ))
    pedido_id = cur.fetchone()["id"]

    execute(
        "INSERT INTO historial_estados (pedido_id,estado_nuevo,usuario_id,usuario_nombre,nota) VALUES (%s,%s,%s,%s,%s)",
        (pedido_id, estado, uid, session.get("nombre"), "Pedido creado")
    )
    db.commit()

    _pendientes_email = enviar_emails_estado(db, pedido_id, estado, usuario_nombre=session.get("nombre", ""), usuario_id=uid)

    # ── Telegram inmediato si el pedido está sujeto al techo de gastos ────────
    if sujeto_techo:
        nombre_familia = None
        if familia_id:
            row_f = query("SELECT nombre FROM familias WHERE id=%s", (familia_id,), one=True)
            nombre_familia = row_f["nombre"] if row_f else None
        hotel_codigo = query("SELECT codigo FROM hoteles WHERE id=%s", (data.get("hotel_id"),), one=True)
        hotel_cod = hotel_codigo["codigo"] if hotel_codigo else ""
        _telegram_alerta_techo(pedido_id, hotel_cod, float(importe or 0), nombre_familia or "—")

    return jsonify({
        "ok": True, "id": pedido_id, "norden": norden, "emails_pendientes": _pendientes_email,
        "estado_final": estado,
        "requiere_autorizacion_dg": estado == "PENDIENTE Vº Bº DIRECCIÓN GENERAL",
    }), 201

@app.route("/api/pedidos/<int:pid>", methods=["PUT"])
@login_required
def update_pedido(pid):
    data = request.get_json(silent=True) or {}
    db   = get_db()
    uid  = current_user_id()

    pedido_actual = row_to_dict(query("SELECT * FROM pedidos WHERE id=%s", (pid,), one=True))
    if not pedido_actual:
        return jsonify({"error": "No encontrado"}), 404

    # Hotel de pruebas ('PR'): fuera del alcance salvo admin o el usuario
    # dedicado a estas pruebas, tanto si el pedido ya pertenece a ese hotel
    # como si se intenta reasignar a él.
    if not _puede_ver_hotel_pruebas() and (
        _es_hotel_pruebas_id(pedido_actual["hotel_id"])
        or _es_hotel_pruebas_id(data.get("hotel_id"))
    ):
        return jsonify({"error": "Sin acceso a este pedido"}), 403

    # ── Restricción rol hotel: solo puede modificar entrada_albaran_num, sin CANCELADO ──
    if session.get("rol") == "hotel":
        hoteles_ids = session.get("hoteles_ids", [])
        if pedido_actual["hotel_id"] not in hoteles_ids:
            return jsonify({"error": "Sin acceso a este pedido"}), 403
        # Solo permitir campos de albarán; ignorar todo lo demás
        albaran_val = data.get("entrada_albaran_num", pedido_actual["entrada_albaran_num"])
        # Determinar estado: SERVIDO PARCIAL / TOTAL según albarán, pero nunca CANCELADO
        estado_solicitado = data.get("estado", pedido_actual["estado"])
        if estado_solicitado == "CANCELADO":
            return jsonify({"error": "El usuario Hotel no puede cancelar pedidos"}), 403
        # (2026-08-28) Base imp. (€) obligatoria en cada entrada — ver
        # _validar_base_imponible_entradas().
        if estado_solicitado in ("ENTREGA PARCIAL", "ENTREGADO") and not _validar_base_imponible_entradas(_parse_albaran_entries(albaran_val)):
            return jsonify({
                "ok": False,
                "error": "La Base imp. (€) es obligatoria en cada entrada de «Nº Entrada DALI / SAP» — tanto "
                         "en una entrada parcial como en la entrada final (total) — para poder continuar."
            }), 422
        execute("""
            UPDATE pedidos SET
                entrada_albaran_num=%s, estado=%s,
                modificado_por_id=%s, modificado_por_nombre=%s, modificado_en=NOW()
            WHERE id=%s
        """, (albaran_val, estado_solicitado, uid, session.get("nombre"), pid))
        estado_antes = pedido_actual["estado"]
        if estado_solicitado != estado_antes:
            execute(
                "INSERT INTO historial_estados (pedido_id,estado_antes,estado_nuevo,usuario_id,usuario_nombre,nota) VALUES (%s,%s,%s,%s,%s,%s)",
                (pid, estado_antes, estado_solicitado, uid, session.get("nombre"), data.get("nota_historial", ""))
            )
        db.commit()
        _pendientes_email = []
        if estado_solicitado != estado_antes:
            _pendientes_email = _notificar_cambio_estado(
                db,
                pid,
                estado_solicitado,
                estado_antes,
                usuario_nombre=session.get("nombre", ""),
                usuario_id=uid,
            )
        return jsonify({"ok": True, "id": pid, "emails_pendientes": _pendientes_email})
    # ── Fin restricción hotel ──────────────────────────────────────────────────

    estado_antes = pedido_actual["estado"]
    estado_nuevo = data.get("estado", estado_antes)

    sujeto_techo = data.get("sujeto_techo", pedido_actual.get("sujeto_techo", 0))
    sujeto_techo = 1 if sujeto_techo else 0
    familia_id   = data.get("familia_id", pedido_actual.get("familia_id"))
    importe      = data.get("importe", pedido_actual.get("importe"))

    # (2026-08-01 — rediseño Techo de Gastos) Ya NO se comprueba el techo en
    # cualquier edición — el momento real de "consumo" es al pasar a
    # ENVIADO AL PROVEEDOR, comprobado más abajo dentro de esa misma
    # validación. _forzar_techo queda eliminado.

    # ── Validación obligatoria para ENVIADO AL PROVEEDOR ─────────────────────
    if estado_nuevo == "ENVIADO AL PROVEEDOR" and estado_antes != "ENVIADO AL PROVEEDOR":
        errores_envio = []

        # 0a. Proveedor asignado obligatorio
        proveedor_id_val = data.get("proveedor_id", pedido_actual.get("proveedor_id"))
        hotel_id_val = data.get("hotel_id", pedido_actual.get("hotel_id"))
        if not proveedor_id_val:
            errores_envio.append(
                "No se puede pasar a ENVIADO AL PROVEEDOR porque el pedido no tiene proveedor asignado. "
                "Asigne un proveedor antes de cambiar el estado."
            )
        else:
            # 0b. El proveedor debe tener al menos un contacto principal con email
            emails_proveedor = _get_proveedor_emails_principales(proveedor_id_val, hotel_id_val)
            if not emails_proveedor:
                # Obtener el nombre del proveedor para dar un mensaje más claro
                prov_row = query("SELECT nombre FROM proveedores WHERE id=%s", (proveedor_id_val,), one=True)
                prov_nombre = (prov_row["nombre"] if prov_row else f"ID {proveedor_id_val}")
                errores_envio.append(
                    f"El proveedor «{prov_nombre}» no tiene ningún correo electrónico configurado en su ficha "
                    f"(contacto principal con email). Acceda a la ficha del proveedor, añada un email al contacto "
                    f"principal y vuelva a cambiar el estado."
                )

        if errores_envio:
            return jsonify({"ok": False, "error": " | ".join(errores_envio), "errores": errores_envio}), 422

        # 1. Nº Pedido (DALI/SAP) y Total Pedido obligatorios — (2026-08-28)
        #    ya no se pueden escribir a mano: solo llegan a tener valor si
        #    se ha subido y leído correctamente el PDF de pedido oficial
        #    PRINCESS (ver punto 2 y _parsear_pdf_pedido_oficial). Se leen
        #    de pedido_actual, nunca de `data`: un envío manual de estos
        #    campos en el JSON no tiene ningún efecto (ver update_pedido()
        #    más abajo, donde el UPDATE tampoco los toma de `data`).
        pedido_num_val = pedido_actual.get("pedido_num") or ""
        if not pedido_num_val.strip() or pedido_actual.get("total_pedido") is None:
            errores_envio.append(
                "Debe adjuntar el PDF del pedido oficial PRINCESS en la sección «Nº Pedido (DALI/SAP)» "
                "para pasar a ENVIADO AL PROVEEDOR — la aplicación rellena sola el Nº de Pedido y el Total "
                "Pedido al leerlo."
            )

        # 2. Adjunto pedido_doc: exactamente 1 documento (el PDF de pedido
        #    oficial, obligatorio — ya no admite correo .eml/.msg en este
        #    apartado, ver upload_adjunto()).
        adjuntos_pedido = rows_to_list(query(
            "SELECT id, nombre FROM pedido_adjuntos WHERE pedido_id=%s AND tipo='pedido_doc'",
            (pid,)
        ))
        if len(adjuntos_pedido) == 0:
            errores_envio.append(
                "Debe adjuntar el PDF del pedido oficial PRINCESS en la sección «Nº Pedido (DALI/SAP)»."
            )
        elif len(adjuntos_pedido) > 1:
            errores_envio.append("Solo se permite un documento en la sección «Nº Pedido (DALI/SAP)» (actualmente hay %d)." % len(adjuntos_pedido))

        # 3. Nº Presupuesto obligatorio (salvo pedidos con tarifa acordada,
        #    que por definición no requieren presupuesto)
        tarifa_acordada_val = data.get("tarifa_acordada", pedido_actual.get("tarifa_acordada", False))
        if not tarifa_acordada_val:
            presupuesto_num_val = data.get("presupuesto_num", pedido_actual.get("presupuesto_num") or "")
            if not (presupuesto_num_val or "").strip():
                errores_envio.append("El campo «Nº Presupuesto» es obligatorio para pasar a ENVIADO AL PROVEEDOR.")

            # 4. Adjunto presupuesto_doc: mínimo 1 documento (puede haber también correos)
            adjuntos_presupuesto = rows_to_list(query(
                "SELECT id, nombre, es_correo FROM pedido_adjuntos WHERE pedido_id=%s AND tipo='presupuesto_doc'",
                (pid,)
            ))
            docs_presupuesto = [a for a in adjuntos_presupuesto if not a["es_correo"]]
            if len(adjuntos_presupuesto) == 0:
                errores_envio.append("Debe adjuntar al menos un documento (PDF/Word) en la sección «Nº Presupuesto».")
            elif len(docs_presupuesto) == 0:
                errores_envio.append("Debe adjuntar al menos un documento (PDF/Word) en «Nº Presupuesto» (solo correo electrónico no es suficiente).")

        if errores_envio:
            return jsonify({"ok": False, "error": " | ".join(errores_envio), "errores": errores_envio}), 422

        # ── Circuito de autorización — Techo de Gastos (rediseño Fase 2) ──
        # El pedido ya está "listo para enviar" (proveedor, docs, nº pedido
        # OK, comprobado arriba). Si está sujeto a techo y lo supera, se
        # desvía a PENDIENTE Vº Bº DIRECCIÓN GENERAL en vez de proceder —
        # salvo que ya exista un expediente aprobado para este pedido este
        # mismo mes (reintento tras aprobación). Es el único punto por el
        # que pasa cualquier vía que intente este cambio de estado, así que
        # también cumple de forma natural el "chequeo de integridad" del
        # punto 5 del rediseño — no hace falta un endpoint aparte.
        if sujeto_techo:
            _mes_techo = _date.today().strftime("%Y-%m")
            _hotel_id_techo = data.get("hotel_id", pedido_actual["hotel_id"])
            _expediente_aprobado = query(
                "SELECT id FROM expediente_exceso WHERE pedido_id=%s AND mes=%s AND resultado='aprobado' "
                "ORDER BY creado_en DESC LIMIT 1",
                (pid, _mes_techo), one=True
            )
            # (2026-08-27) FIX duplicados: si el pedido YA tiene un
            # expediente_exceso pendiente (típicamente porque está en
            # PENDIENTE Vº Bº DIRECCIÓN GENERAL esperando que alguien lo
            # resuelva en Techo de Gastos) y alguien insiste en volver a
            # intentar el cambio a ENVIADO AL PROVEEDOR desde la ficha del
            # pedido — sin darse cuenta de que ya hay un apunte esperando
            # su Vº Bº —, este bloque volvía a llamar a _check_techo() (que
            # por supuesto sigue devolviendo los mismos motivos, nada ha
            # cambiado) y creaba OTRO expediente_exceso pendiente más,
            # multiplicando el apunte tantas veces como reintentos —
            # reportado por Víctor. Ahora se corta aquí, ANTES de volver a
            # comprobar el techo: nunca se crea un segundo expediente
            # pendiente para el mismo pedido, y se devuelve un error
            # específico (`expediente_pendiente_id` + `hotel_codigo`) para
            # que el frontend lleve directamente a Víctor al apunte ya
            # existente en Techo de Gastos en vez de dejar que reintente a
            # ciegas (ver resolverExpedienteTechoDesdeError en
            # templates/index.html).
            _expediente_pendiente = query(
                "SELECT id FROM expediente_exceso WHERE pedido_id=%s AND resultado='pendiente' "
                "ORDER BY creado_en DESC LIMIT 1",
                (pid,), one=True
            )
            if _expediente_pendiente and not _expediente_aprobado:
                _hotel_row_techo = query("SELECT codigo FROM hoteles WHERE id=%s", (_hotel_id_techo,), one=True)
                return jsonify({
                    "ok": False,
                    "error": (
                        "Este pedido ya tiene un apunte pendiente de aprobación en Techo de Gastos — "
                        "acéptalo o recházalo allí antes de volver a intentar el cambio a ENVIADO AL PROVEEDOR "
                        "(no se crea un segundo apunte)."
                    ),
                    "expediente_pendiente_id": _expediente_pendiente["id"],
                    "hotel_codigo": _hotel_row_techo["codigo"] if _hotel_row_techo else None,
                    "requiere_autorizacion_dg": True,
                }), 422  # misma familia que el resto de validaciones de "ENVIADO AL PROVEEDOR" de aquí arriba — así api() en el frontend lo trata como dato normal (data.error), no como excepción (ver esValidacionDeNegocio en templates/index.html)
            if not _expediente_aprobado:
                _motivos_techo = _check_techo(_hotel_id_techo, familia_id, importe, _mes_techo, excluir_pedido_id=pid)
                if _motivos_techo:
                    _cfg_techo = get_config()
                    _consumido_snap, _disponible_snap = _techo_snapshot(_hotel_id_techo, _mes_techo)
                    _importe_f = float(importe) if importe else 0.0
                    _techo_mes_cfg = float(_cfg_techo.get("techo_max_mes", 0) or 0)
                    execute("""
                        INSERT INTO expediente_exceso (
                            pedido_id, hotel_id, familia_id, mes, importe_pedido, consumo_previo, exceso,
                            motivo_solicitud, usuario_solicitante_id, resultado,
                            consumido_en_solicitud, disponible_en_solicitud
                        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'pendiente',%s,%s)
                    """, (
                        pid, _hotel_id_techo, familia_id, _mes_techo,
                        _importe_f, _consumido_snap,
                        max(0.0, (_consumido_snap + _importe_f) - _techo_mes_cfg),
                        " · ".join(_motivos_techo), uid,
                        _consumido_snap, _disponible_snap,
                    ))
                    estado_nuevo = "PENDIENTE Vº Bº DIRECCIÓN GENERAL"
                    _nota_sistema_techo = "Requiere autorización de Dirección General: " + " · ".join(_motivos_techo)
                    _nota_usuario_techo = (data.get("nota_historial") or "").strip()
                    data["nota_historial"] = (
                        (_nota_usuario_techo + " — " if _nota_usuario_techo else "") + _nota_sistema_techo
                    )
    # ── Fin validación ENVIADO AL PROVEEDOR ──────────────────────────────────

    # ── Validación: Base imp. (€) obligatoria en cada entrada de «Nº Entrada
    #    DALI / SAP» (parcial o final) — ver _validar_base_imponible_entradas() ──
    if estado_nuevo in ("ENTREGA PARCIAL", "ENTREGADO"):
        _albaran_val_validar = data.get("entrada_albaran_num", pedido_actual["entrada_albaran_num"])
        if not _validar_base_imponible_entradas(_parse_albaran_entries(_albaran_val_validar)):
            return jsonify({
                "ok": False,
                "error": "La Base imp. (€) es obligatoria en cada entrada de «Nº Entrada DALI / SAP» — tanto "
                         "en una entrada parcial como en la entrada final (total) — para poder continuar."
            }), 422

    ESTADOS_SIN_TRAMITAR = {
        "PENDIENTE FIRMA DIRECCION COMPRAS",
        "PENDIENTE DE FIRMA DIRECCION HOTEL",
    }
    fecha_sol_nueva  = data.get("fecha_solicitud")
    fecha_sol_actual = pedido_actual.get("fecha_solicitud")
    if (
        fecha_sol_nueva
        and not fecha_sol_actual
        and estado_nuevo in ESTADOS_SIN_TRAMITAR
        and "estado" not in data
    ):
        estado_nuevo = "PENDIENTE COTIZACIÓN"

    # ── mes_consumo_techo (rediseño Techo de Gastos, Fase 2) ─────────────────
    # Se rellena SOLO en el instante en que el pedido pasa de verdad a
    # ENVIADO AL PROVEEDOR (momento real de consumo del techo) — no antes,
    # ni siquiera si sujeto_techo está marcado desde el principio. Si se
    # cancela un pedido que ya lo tenía relleno, se libera (vuelve a NULL,
    # deja de contar en el cálculo activo del mes), pero queda constancia
    # completa en historial_estados — nada se borra (punto 7 del rediseño).
    _mes_consumo_techo_val = pedido_actual.get("mes_consumo_techo")
    if estado_nuevo == "ENVIADO AL PROVEEDOR" and estado_antes != "ENVIADO AL PROVEEDOR":
        _mes_consumo_techo_val = _date.today().strftime("%Y-%m")
    elif estado_nuevo == "CANCELADO" and pedido_actual.get("mes_consumo_techo"):
        _quien_aprobo = query(
            "SELECT usuario_nombre FROM historial_estados WHERE pedido_id=%s AND estado_nuevo='ENVIADO AL PROVEEDOR' "
            "ORDER BY creado_en DESC LIMIT 1", (pid,), one=True
        )
        _nota_liberacion = (
            f"Techo liberado al cancelar — Pedido Nº {pedido_actual.get('pedido_num') or pid}, "
            f"importe {float(pedido_actual.get('importe') or 0):,.2f} €, "
            f"visto bueno original de {(_quien_aprobo['usuario_nombre'] if _quien_aprobo else 'desconocido')}, "
            f"consumía el mes {pedido_actual['mes_consumo_techo']}. "
            f"Cancelado por {session.get('nombre')}."
        )
        _nota_usuario_cancel = (data.get("nota_historial") or "").strip()
        data["nota_historial"] = (_nota_usuario_cancel + " — " if _nota_usuario_cancel else "") + _nota_liberacion
        _mes_consumo_techo_val = None

    execute("""
        UPDATE pedidos SET
            hotel_id=%s, departamento_id=%s,
            fecha_solicitud=%s, fecha_envio_visto_bueno=%s, fecha_tramitacion=%s,
            pedido_num=%s, presupuesto_num=%s, entrada_albaran_num=%s,
            tarifa_acordada=%s,
            estado=%s,
            comunicado_ab=%s, comunicado_jefe_dep=%s,
            parte_rotura=%s, parte_ampliacion=%s,
            proveedor_id=%s, observaciones=%s,
            familia_id=%s, importe=%s, sujeto_techo=%s, mes_consumo_techo=%s,
            plazo_entrega_dias=%s, fecha_entrega_especifica=%s,
            total_pedido=%s,
            modificado_por_id=%s, modificado_por_nombre=%s, modificado_en=NOW()
        WHERE id=%s
    """, (
        data.get("hotel_id",            pedido_actual["hotel_id"]),
        data.get("departamento_id",      pedido_actual["departamento_id"]),
        data.get("fecha_solicitud",      pedido_actual["fecha_solicitud"]),
        data.get("fecha_envio_visto_bueno", pedido_actual["fecha_envio_visto_bueno"]),
        data.get("fecha_tramitacion",    pedido_actual["fecha_tramitacion"]),
        # (2026-08-28) pedido_num NUNCA se toma de `data`: solo cambia vía
        # el UPDATE dedicado de upload_adjunto() al leer el PDF de pedido
        # oficial (ver validación más arriba y _parsear_pdf_pedido_oficial).
        pedido_actual["pedido_num"],
        data.get("presupuesto_num",      pedido_actual["presupuesto_num"]),
        data.get("entrada_albaran_num",  pedido_actual["entrada_albaran_num"]),
        bool(data.get("tarifa_acordada", pedido_actual.get("tarifa_acordada", False))),
        estado_nuevo,
        1 if data.get("comunicado_ab",       pedido_actual["comunicado_ab"]) else 0,
        1 if data.get("comunicado_jefe_dep", pedido_actual["comunicado_jefe_dep"]) else 0,
        1 if data.get("parte_rotura",        pedido_actual["parte_rotura"]) else 0,
        1 if data.get("parte_ampliacion",    pedido_actual["parte_ampliacion"]) else 0,
        data.get("proveedor_id",  pedido_actual["proveedor_id"]),
        data.get("observaciones", pedido_actual["observaciones"]),
        familia_id, importe, sujeto_techo, _mes_consumo_techo_val,
        data.get("plazo_entrega_dias", pedido_actual.get("plazo_entrega_dias")) or None,
        data.get("fecha_entrega_especifica", pedido_actual.get("fecha_entrega_especifica")) or None,
        # (2026-08-28) total_pedido tampoco se toma de `data`, mismo motivo
        # que pedido_num justo arriba — salvo que "Comparar listado PDF
        # (SAP)" lo actualice después por su propia vía (_comparar_listado_
        # pdf_logica), que sigue funcionando igual.
        pedido_actual.get("total_pedido"),
        uid, session.get("nombre"), pid,
    ))

    if estado_nuevo != estado_antes:
        execute(
            "INSERT INTO historial_estados (pedido_id,estado_antes,estado_nuevo,usuario_id,usuario_nombre,nota) VALUES (%s,%s,%s,%s,%s,%s)",
            (pid, estado_antes, estado_nuevo, uid, session.get("nombre"), data.get("nota_historial", ""))
        )

    db.commit()

    _pendientes_email = []
    if estado_nuevo != estado_antes:
        _pendientes_email = _notificar_cambio_estado(db, pid, estado_nuevo, estado_antes,
                                 usuario_nombre=session.get("nombre", ""), usuario_id=uid)

    return jsonify({
        "ok": True,
        "emails_pendientes": _pendientes_email,
        "estado_final": estado_nuevo,
        "requiere_autorizacion_dg": estado_nuevo == "PENDIENTE Vº Bº DIRECCIÓN GENERAL",
    })


@app.route("/api/expedientes/<int:eid>/aprobar", methods=["POST"])
@login_required
def aprobar_expediente(eid):
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 2) Aprueba un expediente de
    exceso pendiente — el pedido pasa a ENVIADO AL PROVEEDOR (aquí, no
    antes, se rellena mes_consumo_techo) y sigue el flujo normal de
    notificación (email al proveedor, avisos internos) vía
    _notificar_cambio_estado(), igual que cualquier otro cambio de estado.

    Alcance de esta fase: no se repiten aquí TODAS las validaciones de
    "listo para enviar" (nº pedido, adjuntos...) que ya se comprobaron
    cuando el pedido entró en el circuito — solo se revalida que el
    proveedor siga teniendo un contacto con email, por ser crítico para el
    envío. Si algo más cambió entretanto (p.ej. se borró el adjunto
    obligatorio), el email de confirmación al proveedor podría salir
    incompleto; limitación conocida de esta fase, a revisar si hace falta.
    """
    data = request.get_json(silent=True) or {}
    exp = row_to_dict(query("SELECT * FROM expediente_exceso WHERE id=%s", (eid,), one=True))
    if not exp:
        return jsonify({"error": "Expediente no encontrado"}), 404
    if exp["resultado"] != "pendiente":
        return jsonify({"error": f"Este expediente ya está resuelto ({exp['resultado']})"}), 409

    pedido = row_to_dict(query("SELECT * FROM pedidos WHERE id=%s", (exp["pedido_id"],), one=True))
    if not pedido:
        return jsonify({"error": "El pedido de este expediente ya no existe"}), 404
    if pedido["estado"] != "PENDIENTE Vº Bº DIRECCIÓN GENERAL":
        return jsonify({"error": f"El pedido ya no está pendiente de Dirección General (estado actual: {pedido['estado']})"}), 409

    if not pedido.get("proveedor_id"):
        return jsonify({"error": "El pedido ya no tiene proveedor asignado — revíselo antes de aprobar."}), 422
    if not _get_proveedor_emails_principales(pedido["proveedor_id"], pedido["hotel_id"]):
        return jsonify({"error": "El proveedor de este pedido ya no tiene ningún contacto con email — revíselo antes de aprobar."}), 422

    uid  = current_user_id()
    db   = get_db()
    nota = (data.get("nota_historial") or data.get("observaciones") or "").strip()

    execute("""
        UPDATE expediente_exceso SET resultado='aprobado', usuario_resuelve_id=%s,
               fecha_resolucion=NOW(), observaciones_direccion_general=%s
        WHERE id=%s
    """, (uid, nota or None, eid))

    estado_antes = pedido["estado"]
    mes_consumo  = _date.today().strftime("%Y-%m")
    execute("""
        UPDATE pedidos SET estado='ENVIADO AL PROVEEDOR', mes_consumo_techo=%s,
               modificado_por_id=%s, modificado_por_nombre=%s, modificado_en=NOW()
        WHERE id=%s
    """, (mes_consumo, uid, session.get("nombre"), exp["pedido_id"]))

    _nota_hist = "Autorizado por Dirección General" + (f": {nota}" if nota else "")
    execute(
        "INSERT INTO historial_estados (pedido_id,estado_antes,estado_nuevo,usuario_id,usuario_nombre,nota) VALUES (%s,%s,%s,%s,%s,%s)",
        (exp["pedido_id"], estado_antes, "ENVIADO AL PROVEEDOR", uid, session.get("nombre"), _nota_hist)
    )
    db.commit()

    _pendientes_email = _notificar_cambio_estado(
        db, exp["pedido_id"], "ENVIADO AL PROVEEDOR", estado_antes,
        usuario_nombre=session.get("nombre", ""), usuario_id=uid,
    )
    # (2026-08-27) pedido_id incluido en la respuesta para que el frontend,
    # cuando el usuario aprueba este expediente navegando directamente a
    # Techo de Gastos (sin venir de un intento bloqueado en la ficha del
    # pedido), pueda volver a abrir esa ficha y confirmarle ahí que el
    # pedido ya se ha enviado al proveedor — ver resolverExpedienteTecho()
    # en templates/index.html.
    return jsonify({"ok": True, "emails_pendientes": _pendientes_email, "pedido_id": exp["pedido_id"]})


@app.route("/api/expedientes/<int:eid>/denegar", methods=["POST"])
@login_required
def denegar_expediente(eid):
    """
    (2026-08-01 — rediseño Techo de Gastos, Fase 2) Deniega un expediente
    pendiente — el pedido pasa a DENEGADO POR DIRECCION GENERAL (estado
    reabrible: se puede reeditar y reintentar más adelante, lo que abrirá
    un nuevo expediente independiente — cada intento es una fila propia,
    nunca se sobrescribe). Motivo obligatorio (punto 8 del rediseño); la
    fecha de resolución no es un campo manual, se toma de
    fecha_resolucion/creado_en automáticamente.
    """
    data = request.get_json(silent=True) or {}
    nota = (data.get("nota_historial") or data.get("observaciones") or "").strip()
    if not nota:
        return jsonify({"error": "El motivo de la denegación es obligatorio"}), 400

    exp = row_to_dict(query("SELECT * FROM expediente_exceso WHERE id=%s", (eid,), one=True))
    if not exp:
        return jsonify({"error": "Expediente no encontrado"}), 404
    if exp["resultado"] != "pendiente":
        return jsonify({"error": f"Este expediente ya está resuelto ({exp['resultado']})"}), 409

    pedido = row_to_dict(query("SELECT * FROM pedidos WHERE id=%s", (exp["pedido_id"],), one=True))
    if not pedido:
        return jsonify({"error": "El pedido de este expediente ya no existe"}), 404

    uid = current_user_id()
    db  = get_db()

    execute("""
        UPDATE expediente_exceso SET resultado='denegado', usuario_resuelve_id=%s,
               fecha_resolucion=NOW(), observaciones_direccion_general=%s
        WHERE id=%s
    """, (uid, nota, eid))

    estado_antes = pedido["estado"]
    execute("""
        UPDATE pedidos SET estado='DENEGADO POR DIRECCION GENERAL',
               modificado_por_id=%s, modificado_por_nombre=%s, modificado_en=NOW()
        WHERE id=%s
    """, (uid, session.get("nombre"), exp["pedido_id"]))
    execute(
        "INSERT INTO historial_estados (pedido_id,estado_antes,estado_nuevo,usuario_id,usuario_nombre,nota) VALUES (%s,%s,%s,%s,%s,%s)",
        (exp["pedido_id"], estado_antes, "DENEGADO POR DIRECCION GENERAL", uid, session.get("nombre"),
         f"Denegado por Dirección General: {nota}")
    )
    db.commit()

    _pendientes_email = _notificar_cambio_estado(
        db, exp["pedido_id"], "DENEGADO POR DIRECCION GENERAL", estado_antes,
        usuario_nombre=session.get("nombre", ""), usuario_id=uid,
    )
    return jsonify({"ok": True, "emails_pendientes": _pendientes_email})


@app.route("/api/pedidos/<int:pid>", methods=["DELETE"])
@admin_required
def delete_pedido(pid):
    data   = request.get_json(silent=True) or {}
    motivo = (data.get("motivo") or "").strip()
    if not motivo:
        return jsonify({"error": "Debes indicar el motivo de la eliminación"}), 400

    db  = get_db()
    uid = current_user_id()

    # ── 1. Capturar datos completos del pedido antes de borrar ───────────────
    pedido = row_to_dict(query(f"{PEDIDO_SELECT} WHERE p.id=%s", (pid,), one=True))
    if not pedido:
        return jsonify({"error": "Pedido no encontrado"}), 404

    admin_nombre = session.get("nombre", session.get("username", "Desconocido"))

    # ── 2. Guardar registro histórico en pedidos_eliminados ──────────────────
    execute("""
        INSERT INTO pedidos_eliminados (
            pedido_id, norden, hotel_nombre, departamento_nombre,
            proveedor_nombre, proveedor_email, estado,
            fecha_solicitud, pedido_num, presupuesto_num,
            entrada_albaran_num, observaciones, creado_por_nombre,
            motivo_eliminacion, eliminado_por_id, eliminado_por_nombre
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        pid,
        pedido.get("norden"),
        pedido.get("hotel_nombre"),
        pedido.get("departamento_nombre"),
        pedido.get("proveedor_nombre"),
        pedido.get("proveedor_email"),
        pedido.get("estado"),
        pedido.get("fecha_solicitud"),
        pedido.get("pedido_num"),
        pedido.get("presupuesto_num"),
        pedido.get("entrada_albaran_num"),
        pedido.get("observaciones"),
        pedido.get("creado_por_nombre"),
        motivo,
        uid,
        admin_nombre,
    ))

    # ── 3. Eliminar el pedido (CASCADE borra adjuntos e historial) ───────────
    execute("DELETE FROM pedidos WHERE id=%s", (pid,))
    db.commit()
    return jsonify({"ok": True, "norden": pedido.get("norden")})

# ── Registro de pedidos eliminados ────────────────────────────────────────────

@app.route("/api/pedidos_eliminados")
@login_required
def get_pedidos_eliminados():
    if session.get("rol") not in ("admin", "compras"):
        return jsonify({"error": "Acceso restringido"}), 403

    # (2026-09-01, repaso "agilizar y limpiar", Etapa 2) Antes devolvía la
    # tabla `pedidos_eliminados` entera de golpe — un registro que solo
    # crece (nunca se purga, es el histórico de bajas) y no tenía filtros
    # ni límite. Mismo patrón de paginación ya probado en /api/pedidos y
    # /api/proveedores: {registros,total,page,page_size,pages}. Mantiene
    # el nombre de clave "registros" (no "items") para no romper nada que
    # ya dependa de esa forma de respuesta.
    try:
        page      = max(1, int(request.args.get("page", 1)))
        page_size = max(1, min(100, int(request.args.get("page_size", 30))))
    except ValueError:
        page, page_size = 1, 30

    total = query("SELECT COUNT(*) as total FROM pedidos_eliminados", one=True)["total"]
    registros = rows_to_list(query(
        "SELECT * FROM pedidos_eliminados ORDER BY eliminado_en DESC LIMIT %s OFFSET %s",
        (page_size, (page - 1) * page_size)
    ))
    return jsonify({
        "registros":  registros,
        "total":      total,
        "page":       page,
        "page_size":  page_size,
        "pages":      max(1, (total + page_size - 1) // page_size),
    })

# ── API Stats ──────────────────────────────────────────────────────────────────

@app.route("/api/stats")
@login_required
def get_stats():
    # ── Restricción rol hotel: solo sus hoteles asignados, alertas filtradas ──
    if session.get("rol") == "hotel":
        hoteles_ids = session.get("hoteles_ids", [])
        if not hoteles_ids:
            return jsonify({"total": 0, "by_estado": [], "by_hotel": [],
                            "alertas": [], "num_alertas": 0})
        placeholders = ",".join(["%s"] * len(hoteles_ids))
        total = query(
            f"SELECT COUNT(*) as n FROM pedidos WHERE hotel_id IN ({placeholders})",
            hoteles_ids, one=True)["n"]
        by_estado = rows_to_list(query(
            f"SELECT estado, COUNT(*) as total FROM pedidos WHERE hotel_id IN ({placeholders}) GROUP BY estado ORDER BY total DESC",
            hoteles_ids))
        by_hotel = rows_to_list(query(
            f"""SELECT h.codigo, h.nombre, COUNT(p.id) as total
                FROM hoteles h LEFT JOIN pedidos p ON p.hotel_id=h.id
                WHERE h.id IN ({placeholders})
                GROUP BY h.id, h.codigo, h.nombre ORDER BY total DESC""",
            hoteles_ids))
        # Calcular alertas reales para los hoteles visibles del usuario hotel
        alertas_raw_h = rows_to_list(query(f"""
            {PEDIDO_SELECT_STATS}
            WHERE p.estado IN (
                'ENVIADO AL PROVEEDOR',
                'PENDIENTE FIRMA DIRECCION COMPRAS',
                'PENDIENTE DE FIRMA DIRECCION HOTEL',
                'ENTREGA PARCIAL',
                'PENDIENTE COTIZACIÓN'
            )
              AND p.hotel_id IN ({placeholders})
              AND (
                p.fecha_tramitacion IS NOT NULL
                OR (p.estado = 'PENDIENTE COTIZACIÓN' AND p.fecha_solicitud IS NOT NULL)
              )
            ORDER BY p.fecha_tramitacion ASC
        """, hoteles_ids))
        cfg_activar_plazo_h = bool(int(get_config().get("activar_uso_plazo_entrega", 1) or 0))
        alertas_h = _clasificar_alertas(alertas_raw_h, cfg_activar_plazo_h)
        return jsonify({"total": total, "by_estado": by_estado,
                        "by_hotel": by_hotel, "alertas": alertas_h,
                        "num_alertas": len(alertas_h)})
    # ── Resto de roles (admin / compras) ─────────────────────────────────────
    # Hotel de pruebas ('PR'): invisible para compras, visible solo a admin.
    _excluir_pruebas = not _puede_ver_hotel_pruebas()
    _filtro_pruebas_pedidos = (
        " WHERE p.hotel_id NOT IN (SELECT id FROM hoteles WHERE codigo=%s)"
        if _excluir_pruebas else ""
    )
    _filtro_pruebas_args = (HOTEL_CODIGO_PRUEBAS,) if _excluir_pruebas else ()

    # total se deriva de by_estado: evita un COUNT(*) redundante sobre la tabla.
    by_estado = rows_to_list(query(
        f"SELECT estado, COUNT(*) as total FROM pedidos p{_filtro_pruebas_pedidos} "
        f"GROUP BY estado ORDER BY total DESC",
        _filtro_pruebas_args
    ))
    total = sum(r["total"] for r in by_estado)
    _filtro_pruebas_hotel = " AND h.codigo <> %s" if _excluir_pruebas else ""
    by_hotel  = rows_to_list(query(
        f"""SELECT h.codigo, h.nombre, COUNT(p.id) as total
           FROM hoteles h LEFT JOIN pedidos p ON p.hotel_id=h.id
           WHERE 1=1{_filtro_pruebas_hotel}
           GROUP BY h.id, h.codigo, h.nombre ORDER BY total DESC""",
        _filtro_pruebas_args
    ))
    # ── Alertas: clasificadas por _clasificar_alertas (fuente única) ──────────
    _filtro_pruebas_alertas = " AND h.codigo <> %s" if _excluir_pruebas else ""
    alertas_raw = rows_to_list(query(f"""
        {PEDIDO_SELECT_STATS}
        WHERE p.estado IN (
            'ENVIADO AL PROVEEDOR',
            'PENDIENTE FIRMA DIRECCION COMPRAS',
            'PENDIENTE DE FIRMA DIRECCION HOTEL',
            'ENTREGA PARCIAL',
            'PENDIENTE COTIZACIÓN'
        )
          AND (
            p.fecha_tramitacion IS NOT NULL
            OR (p.estado = 'PENDIENTE COTIZACIÓN' AND p.fecha_solicitud IS NOT NULL)
          ){_filtro_pruebas_alertas}
        ORDER BY p.fecha_tramitacion ASC
    """, _filtro_pruebas_args))
    cfg_activar_plazo = bool(int(get_config().get("activar_uso_plazo_entrega", 1) or 0))
    alertas = _clasificar_alertas(alertas_raw, cfg_activar_plazo)

    return jsonify({
        "total": total, "by_estado": by_estado,
        "by_hotel": by_hotel, "alertas": alertas,
        "num_alertas": len(alertas),
    })


# ── Dashboard v13 (Nivel 1) — resumen ejecutivo ───────────────────────────────
#
# Endpoint APARTE de /api/stats a propósito: /api/stats se llama desde medio
# programa (badge del sidebar, vista Alertas, modal de impresión, tras
# guardar/eliminar un pedido…), así que cualquier query que añadamos ahí se
# paga en todos esos sitios. Este endpoint solo se dispara al abrir la vista
# Dashboard, con su propia caché de 30s en el frontend — mismo patrón que
# _fetchStats/_fetchTecho.
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/api/dashboard/resumen")
@login_required
def get_dashboard_resumen():
    """
    Datos adicionales del Dashboard: variación mensual de pedidos e importe,
    pendientes activos + tiempo medio de espera, alertas por nivel y por
    hotel, actividad de hoy (entregas/envíos registrados) y últimos pedidos.
    GET /api/dashboard/resumen
    """
    vacio = {
        "variacion": {"pedidos_mes_actual": 0, "pedidos_mes_anterior": 0, "pct": None},
        "entregas_variacion": {"mes_actual": 0, "mes_anterior": 0, "pct": None},
        "series": {"dias": 14, "pedidos": [], "entregas": []},
        "importe": {"mes_actual": 0.0, "mes_anterior": 0.0, "pct": None},
        "pendientes": {"total": 0, "tiempo_medio_dias": None},
        "alertas_por_nivel": {"urgente": 0, "aviso": 0},
        "actividad_hoy": {"entregados": 0, "enviados": 0},
        "ultimos_pedidos": [],
        "by_hotel": [],
        "timeline": [],
        "ranking_proveedores": [],
        "sla_aprobacion_dias": None,
        "necesita_atencion": None,
    }

    rol = session.get("rol", "")
    hoteles_ids = session.get("hoteles_ids", []) if rol == "hotel" else None
    if rol == "hotel" and not hoteles_ids:
        return jsonify(vacio)

    # Hotel de pruebas ('PR'): fuera del dashboard salvo admin o el usuario
    # dedicado a estas pruebas.
    _pr_id = None
    if not _puede_ver_hotel_pruebas():
        _pr_row = query("SELECT id FROM hoteles WHERE codigo=%s", (HOTEL_CODIGO_PRUEBAS,), one=True)
        _pr_id = _pr_row["id"] if _pr_row else None

    _filtro_partes = []
    params_p = ()
    if hoteles_ids:
        placeholders = ",".join(["%s"] * len(hoteles_ids))
        _filtro_partes.append(f"p.hotel_id IN ({placeholders})")
        params_p += tuple(hoteles_ids)
    if _pr_id:
        _filtro_partes.append("p.hotel_id <> %s")
        params_p += (_pr_id,)
    filtro_p = ("AND " + " AND ".join(_filtro_partes)) if _filtro_partes else ""

    hoy = _date.today()
    primer_dia_mes = hoy.replace(day=1)
    primer_dia_mes_ant = (primer_dia_mes - timedelta(days=1)).replace(day=1)

    # ── Variación mensual de pedidos e importe ─────────────────────────────
    fila = query(f"""
        SELECT
            COUNT(*) FILTER (WHERE p.creado_en >= %s) AS mes_actual,
            COUNT(*) FILTER (WHERE p.creado_en >= %s AND p.creado_en < %s) AS mes_anterior,
            COALESCE(SUM(p.importe) FILTER (WHERE p.creado_en >= %s), 0) AS importe_mes_actual,
            COALESCE(SUM(p.importe) FILTER (WHERE p.creado_en >= %s AND p.creado_en < %s), 0) AS importe_mes_anterior
        FROM pedidos p
        WHERE 1=1 {filtro_p}
    """, (primer_dia_mes, primer_dia_mes_ant, primer_dia_mes,
          primer_dia_mes, primer_dia_mes_ant, primer_dia_mes) + params_p, one=True)

    def _pct(actual, anterior):
        if not anterior:
            return None
        return round((actual - anterior) / anterior * 100, 1)

    variacion = {
        "pedidos_mes_actual": fila["mes_actual"],
        "pedidos_mes_anterior": fila["mes_anterior"],
        "pct": _pct(fila["mes_actual"], fila["mes_anterior"]),
    }
    importe_actual = float(fila["importe_mes_actual"])
    importe_anterior = float(fila["importe_mes_anterior"])
    importe = {
        "mes_actual": importe_actual, "mes_anterior": importe_anterior,
        "pct": _pct(importe_actual, importe_anterior),
    }
    # El rol hotel no ve importes económicos en ningún otro sitio de la app
    # (p-importe se oculta en el modal de pedido) — se mantiene la misma
    # regla aquí, a nivel de API, no solo de UI.
    if rol == "hotel":
        importe = {"mes_actual": None, "mes_anterior": None, "pct": None}

    # ── Pendientes activos + alertas (misma base que /api/stats) ──────────
    activos_raw = rows_to_list(query(f"""
        {PEDIDO_SELECT_STATS}
        WHERE p.estado IN (
            'ENVIADO AL PROVEEDOR', 'PENDIENTE FIRMA DIRECCION COMPRAS',
            'PENDIENTE DE FIRMA DIRECCION HOTEL', 'ENTREGA PARCIAL',
            'PENDIENTE COTIZACIÓN'
        ) {filtro_p}
    """, params_p))

    dias_list = [d for d in (
        _dias_desde_alerta(p.get("fecha_tramitacion") or p.get("fecha_solicitud"))
        for p in activos_raw
    ) if d is not None]
    tiempo_medio = round(sum(dias_list) / len(dias_list), 1) if dias_list else None

    cfg_activar_plazo = bool(int(get_config().get("activar_uso_plazo_entrega", 1) or 0))
    alertables = [p for p in activos_raw if p.get("fecha_tramitacion") or
                  (p.get("estado") == "PENDIENTE COTIZACIÓN" and p.get("fecha_solicitud"))]
    alertas = _clasificar_alertas(alertables, cfg_activar_plazo)

    alertas_por_nivel = {"urgente": 0, "aviso": 0}
    alertas_por_hotel = {}
    for a in alertas:
        nivel = a.get("nivel_alerta")
        if nivel in alertas_por_nivel:
            alertas_por_nivel[nivel] += 1
        cod = a.get("hotel_codigo")
        if cod:
            alertas_por_hotel[cod] = alertas_por_hotel.get(cod, 0) + 1

    # ── Actividad de hoy: transiciones de estado registradas hoy ──────────
    fila_hoy = query(f"""
        SELECT
            COUNT(*) FILTER (WHERE he.estado_nuevo = 'ENTREGADO') AS entregados,
            COUNT(*) FILTER (WHERE he.estado_nuevo = 'ENVIADO AL PROVEEDOR') AS enviados
        FROM historial_estados he
        JOIN pedidos p ON p.id = he.pedido_id
        WHERE he.creado_en::date = CURRENT_DATE {filtro_p}
    """, params_p, one=True)

    # ── Variación mensual de entregas (mismo patrón que "variación" arriba,
    #    pero contando transiciones a ENTREGADO en vez de altas de pedido) ──
    fila_entregas = query(f"""
        SELECT
            COUNT(*) FILTER (WHERE he.creado_en >= %s) AS mes_actual,
            COUNT(*) FILTER (WHERE he.creado_en >= %s AND he.creado_en < %s) AS mes_anterior
        FROM historial_estados he
        JOIN pedidos p ON p.id = he.pedido_id
        WHERE he.estado_nuevo = 'ENTREGADO' {filtro_p}
    """, (primer_dia_mes, primer_dia_mes_ant, primer_dia_mes) + params_p, one=True)
    entregas_variacion = {
        "mes_actual": fila_entregas["mes_actual"],
        "mes_anterior": fila_entregas["mes_anterior"],
        "pct": _pct(fila_entregas["mes_actual"], fila_entregas["mes_anterior"]),
    }

    # ── Series diarias (últimos 14 días) — base de los sparklines del
    #    dashboard. Se rellenan los días sin movimiento con 0 mediante
    #    generate_series, para que la serie siempre tenga longitud fija. ────
    dias_serie = 14
    fecha_ini_serie = hoy - timedelta(days=dias_serie - 1)

    serie_pedidos = rows_to_list(query(f"""
        SELECT gs::date AS dia, COALESCE(t.total, 0) AS total
        FROM generate_series(%s::date, %s::date, interval '1 day') gs
        LEFT JOIN (
            SELECT p.creado_en::date AS dia, COUNT(*) AS total
            FROM pedidos p
            WHERE p.creado_en::date BETWEEN %s AND %s {filtro_p}
            GROUP BY p.creado_en::date
        ) t ON t.dia = gs::date
        ORDER BY gs
    """, (fecha_ini_serie, hoy, fecha_ini_serie, hoy) + params_p))

    serie_entregas = rows_to_list(query(f"""
        SELECT gs::date AS dia, COALESCE(t.total, 0) AS total
        FROM generate_series(%s::date, %s::date, interval '1 day') gs
        LEFT JOIN (
            SELECT he.creado_en::date AS dia, COUNT(*) AS total
            FROM historial_estados he
            JOIN pedidos p ON p.id = he.pedido_id
            WHERE he.estado_nuevo = 'ENTREGADO' AND he.creado_en::date BETWEEN %s AND %s {filtro_p}
            GROUP BY he.creado_en::date
        ) t ON t.dia = gs::date
        ORDER BY gs
    """, (fecha_ini_serie, hoy, fecha_ini_serie, hoy) + params_p))

    series = {
        "dias": dias_serie,
        "pedidos": [r["total"] for r in serie_pedidos],
        "entregas": [r["total"] for r in serie_entregas],
    }

    # ── Últimos pedidos ─────────────────────────────────────────────────
    ultimos = rows_to_list(query(f"""
        SELECT p.id, p.norden, p.pedido_num, p.estado, p.creado_en,
               h.codigo AS hotel_codigo, h.nombre AS hotel_nombre,
               pr.nombre AS proveedor_nombre
        FROM pedidos p
        LEFT JOIN hoteles h ON p.hotel_id = h.id
        LEFT JOIN proveedores pr ON p.proveedor_id = pr.id
        WHERE 1=1 {filtro_p}
        ORDER BY p.creado_en DESC
        LIMIT 6
    """, params_p))
    for u in ultimos:
        u["creado_en"] = u["creado_en"].isoformat() if u.get("creado_en") else None

    # ── Por hotel: total + entregados + % cumplimiento + alertas ──────────
    _filtro_h_partes = []
    params_h = ()
    if hoteles_ids:
        placeholders = ",".join(["%s"] * len(hoteles_ids))
        _filtro_h_partes.append(f"h.id IN ({placeholders})")
        params_h += tuple(hoteles_ids)
    if _pr_id:
        _filtro_h_partes.append("h.id <> %s")
        params_h += (_pr_id,)
    filtro_h = ("WHERE " + " AND ".join(_filtro_h_partes)) if _filtro_h_partes else ""
    by_hotel_raw = rows_to_list(query(f"""
        SELECT h.codigo, h.nombre, COUNT(p.id) AS total,
               COUNT(p.id) FILTER (WHERE p.estado = 'ENTREGADO') AS entregados
        FROM hoteles h
        LEFT JOIN pedidos p ON p.hotel_id = h.id
        {filtro_h}
        GROUP BY h.id, h.codigo, h.nombre
        ORDER BY total DESC
    """, params_h))
    by_hotel = []
    for h in by_hotel_raw:
        total = h["total"]
        entregados = h["entregados"]
        pct = round(entregados / total * 100, 1) if total else None
        by_hotel.append({
            "codigo": h["codigo"], "nombre": h["nombre"],
            "total": total, "entregados": entregados,
            "pct_cumplimiento": pct,
            "alertas": alertas_por_hotel.get(h["codigo"], 0),
        })

    # ── Necesita atención: la alerta más crítica (alertas ya viene ordenada
    #    urgentes primero y por días descendente, ver _clasificar_alertas) ──
    necesita_atencion = None
    if alertas:
        top = alertas[0]
        necesita_atencion = {
            "id": top.get("id"), "norden": top.get("norden"),
            "pedido_num": top.get("pedido_num"),
            "hotel_codigo": top.get("hotel_codigo"),
            "proveedor_nombre": top.get("proveedor_nombre"),
            "estado": top.get("estado"),
            "dias_tramitacion": top.get("dias_tramitacion"),
            "nivel_alerta": top.get("nivel_alerta"),
        }

    # ── Línea temporal: últimos eventos de historial_estados ───────────────
    # (2026-08-28) A petición de Víctor: mostrar el Nº Pedido (DALI/SAP,
    # `pedido_num`) en vez del "Nº" lineal interno de la app (`norden`) —
    # mismo criterio ya aplicado en otros sitios (ver comentario de
    # _texto_pedido_candidato_...() más arriba, línea ~9924). Se sigue
    # pidiendo `norden` como reserva para el pedido raro que aún no tenga
    # Nº Pedido (DALI/SAP) asignado (ver render en templates/index.html).
    timeline = rows_to_list(query(f"""
        SELECT he.estado_antes, he.estado_nuevo, he.usuario_nombre, he.creado_en,
               p.id AS pedido_id, p.norden, p.pedido_num, h.codigo AS hotel_codigo
        FROM historial_estados he
        JOIN pedidos p ON p.id = he.pedido_id
        LEFT JOIN hoteles h ON p.hotel_id = h.id
        WHERE 1=1 {filtro_p}
        ORDER BY he.creado_en DESC
        LIMIT 15
    """, params_p))
    for t in timeline:
        t["creado_en"] = t["creado_en"].isoformat() if t.get("creado_en") else None

    # ── Ranking de proveedores: pedidos, % cumplimiento, incidencias ──────
    # "Incidencias" = pedidos de ese proveedor que están hoy en la lista de
    # alertas (no hay tabla de reclamaciones real todavía).
    ranking_raw = rows_to_list(query(f"""
        SELECT pr.nombre, COUNT(p.id) AS total,
               COUNT(p.id) FILTER (WHERE p.estado = 'ENTREGADO') AS entregados
        FROM proveedores pr
        JOIN pedidos p ON p.proveedor_id = pr.id
        WHERE 1=1 {filtro_p}
        GROUP BY pr.id, pr.nombre
        HAVING COUNT(p.id) > 0
        ORDER BY total DESC
        LIMIT 8
    """, params_p))
    incidencias_por_proveedor = {}
    for a in alertas:
        pn = a.get("proveedor_nombre")
        if pn:
            incidencias_por_proveedor[pn] = incidencias_por_proveedor.get(pn, 0) + 1
    ranking_proveedores = []
    for r in ranking_raw:
        total = r["total"]
        entregados = r["entregados"]
        ranking_proveedores.append({
            "nombre": r["nombre"], "total": total, "entregados": entregados,
            "pct_cumplimiento": round(entregados / total * 100, 1) if total else None,
            "incidencias": incidencias_por_proveedor.get(r["nombre"], 0),
        })

    # ── SLA de aprobación: días entre "pendiente de firma" y "enviado al
    #    proveedor", promediado sobre los últimos 90 días ───────────────────
    filtro_sla = filtro_p.replace("p.hotel_id", "pe.hotel_id") if filtro_p else ""
    fila_sla = query(f"""
        WITH aprob AS (
            SELECT pedido_id, MIN(creado_en) AS t_inicio
            FROM historial_estados
            WHERE estado_nuevo IN ('PENDIENTE FIRMA DIRECCION COMPRAS',
                                    'PENDIENTE DE FIRMA DIRECCION HOTEL')
            GROUP BY pedido_id
        ),
        envio AS (
            SELECT pedido_id, MIN(creado_en) AS t_envio
            FROM historial_estados
            WHERE estado_nuevo = 'ENVIADO AL PROVEEDOR'
            GROUP BY pedido_id
        )
        SELECT AVG(EXTRACT(EPOCH FROM (e.t_envio - a.t_inicio)) / 86400.0) AS sla_dias,
               COUNT(*) AS n
        FROM aprob a
        JOIN envio e ON e.pedido_id = a.pedido_id
        JOIN pedidos pe ON pe.id = a.pedido_id
        WHERE e.t_envio > a.t_inicio
          AND a.t_inicio >= NOW() - INTERVAL '90 days'
          {filtro_sla}
    """, params_p, one=True)
    sla_dias = round(float(fila_sla["sla_dias"]), 1) if fila_sla and fila_sla["sla_dias"] is not None else None

    return jsonify({
        "variacion": variacion,
        "entregas_variacion": entregas_variacion,
        "series": series,
        "importe": importe,
        "pendientes": {"total": len(activos_raw), "tiempo_medio_dias": tiempo_medio},
        "alertas_por_nivel": alertas_por_nivel,
        "actividad_hoy": {
            "entregados": fila_hoy["entregados"] if fila_hoy else 0,
            "enviados": fila_hoy["enviados"] if fila_hoy else 0,
        },
        "ultimos_pedidos": ultimos,
        "by_hotel": by_hotel,
        "timeline": timeline,
        "ranking_proveedores": ranking_proveedores,
        "sla_aprobacion_dias": sla_dias,
        "necesita_atencion": necesita_atencion,
    })


# ── Dashboard configurable por usuario (v12.16.2) ──────────────────────────
# Widgets disponibles: mismos ids que data-widget en el HTML. Se valida la
# lista recibida contra este catálogo para no guardar basura si el frontend
# cambia de versión o llega una petición manipulada.
_DASHBOARD_WIDGETS_VALIDOS = {
    "insights", "actividad", "quicklinks", "chart-estado", "chart-hotel",
    "timeline", "ranking", "hoteles", "ultimos",
}

@app.route("/api/dashboard/prefs")
@login_required
def get_dashboard_prefs():
    """GET /api/dashboard/prefs — preferencias de widgets del usuario logado."""
    row = query("SELECT dashboard_prefs FROM usuarios WHERE id=%s", (session["user_id"],), one=True)
    prefs = None
    if row and row.get("dashboard_prefs"):
        try:
            prefs = json.loads(row["dashboard_prefs"])
        except Exception:
            prefs = None
    return jsonify({"prefs": prefs})


@app.route("/api/dashboard/prefs", methods=["PUT"])
@login_required
def set_dashboard_prefs():
    """
    PUT /api/dashboard/prefs
    Body: {"prefs": [{"id": "insights", "visible": true}, ...]} o
          {"prefs": null} para volver a la configuración por defecto.
    """
    data = request.get_json(force=True) or {}
    prefs = data.get("prefs")

    if prefs is not None:
        if not isinstance(prefs, list):
            return jsonify({"error": "Formato de preferencias inválido"}), 400
        limpio = []
        vistos = set()
        for item in prefs:
            wid = (item or {}).get("id")
            if wid not in _DASHBOARD_WIDGETS_VALIDOS or wid in vistos:
                continue
            vistos.add(wid)
            limpio.append({"id": wid, "visible": bool(item.get("visible", True))})
        prefs = limpio

    execute(
        "UPDATE usuarios SET dashboard_prefs=%s WHERE id=%s",
        (json.dumps(prefs) if prefs is not None else None, session["user_id"]),
    )
    return jsonify({"ok": True})


# ── API Bridge Agenda — alertas filtradas por usuario (v10.3) ─────────────────
#
# Endpoint consumido por pedidos_agenda_bridge.py en cada instancia de
# main_agenda. Devuelve SOLO las alertas que corresponden al usuario logado:
#
#   rol='compras' → alertas de los hoteles asignados en usuario_comprador_hoteles
#   rol='admin'   → todas las alertas (supervisión global)
#   rol='hotel'   → alertas de los hoteles asignados en usuario_hoteles (lectura)
#
# Misma lógica de umbrales y niveles que /api/stats pero filtrada.
# ─────────────────────────────────────────────────────────────────────────────

def _filtrar_popups_no_vistos(usuario: str, alertas: list) -> list:
    """
    v12.29.47 (PRUEBA) — Popup de main_agenda: entrega única persistida.

    Antes, este endpoint devolvía siempre TODAS las alertas activas del
    usuario, y era pedidos_agenda_bridge.py (en memoria, vía
    _estado_popups) quien decidía si tocaba repetir el popup según un
    intervalo en horas. Al reiniciarse Organizador Princess ese
    historial se perdía → reaparición de popups ya vistos. Se sospecha
    que esta es la causa del comprador de INSIRE viendo el mismo aviso
    "continuamente" (2026-08-04).

    Ahora el dedup se hace aquí, contra bridge_popup_visto (persistida
    en BD, no en el cliente):
      - Si (usuario, pedido_id, nivel) ya está en bridge_popup_visto,
        NO se devuelve — ya se entregó una vez, para siempre.
      - Si el pedido escala de nivel (aviso→urgente), la clave cambia
        de nivel → se trata como aviso nuevo y sí se entrega.
      - Las que SÍ se devuelven se marcan como vistas en la misma
        llamada (igual que /api/bridge/notificaciones marca leido=TRUE
        al servir) — si Organizador Princess está cerrada en este
        momento y nunca llega a recibir esta respuesta, ese popup
        concreto se entregará en la próxima conexión, pero no antes;
        una vez entregado no se repite ni con reintentos ni al
        reiniciar la app.
      - Si un pedido deja de ser alertable (se resuelve o cambia de
        estado), se borra su fila: si vuelve a alertar más adelante se
        trata como un aviso nuevo.

    Devuelve la sublista de `alertas` pendiente de entregar (posible
    lista vacía). No modifica `alertas` en sí — ese valor completo se
    sigue usando para el resumen/saludo diario (total de activas).
    """
    if not alertas:
        try:
            db = get_db()
            db.cursor().execute(
                "DELETE FROM bridge_popup_visto WHERE usuario=%s", (usuario,)
            )
            db.commit()
        except Exception as exc:
            log.warning("bridge_popup_visto: no se pudo limpiar (%s) — %s", usuario, exc)
        return []

    pedido_ids_activos = list({a["id"] for a in alertas})

    try:
        vistos_rows = rows_to_list(query(
            "SELECT pedido_id, nivel FROM bridge_popup_visto WHERE usuario=%s",
            (usuario,)
        ))
    except Exception as exc:
        log.warning("bridge_popup_visto: no se pudo leer vistos (%s) — %s", usuario, exc)
        vistos_rows = []

    vistos = {(r["pedido_id"], r["nivel"]) for r in vistos_rows}
    nuevas = [a for a in alertas if (a["id"], a["nivel_alerta"]) not in vistos]

    try:
        db  = get_db()
        cur = db.cursor()
        placeholders = ",".join(["%s"] * len(pedido_ids_activos))
        cur.execute(
            f"""DELETE FROM bridge_popup_visto
                WHERE usuario=%s AND pedido_id NOT IN ({placeholders})""",
            (usuario, *pedido_ids_activos)
        )
        for a in nuevas:
            cur.execute(
                """INSERT INTO bridge_popup_visto (usuario, pedido_id, nivel)
                   VALUES (%s, %s, %s)
                   ON CONFLICT (usuario, pedido_id, nivel) DO NOTHING""",
                (usuario, a["id"], a["nivel_alerta"])
            )
        db.commit()
    except Exception as exc:
        log.warning("bridge_popup_visto: no se pudo actualizar (%s) — %s", usuario, exc)

    return nuevas


@app.route("/api/bridge/alertas")
@login_required
def bridge_alertas_usuario():
    """
    Devuelve las alertas activas filtradas por el usuario de la sesión.
    Usado por pedidos_agenda_bridge.py para mostrar popups personalizados
    en main_agenda sin mezclar avisos entre compradores.
    """
    from datetime import date as _date, datetime as _dt

    rol      = session.get("rol", "")
    user_id  = session.get("user_id")

    # ── Determinar qué hotel_ids aplican al usuario ───────────────────────────
    if rol == "admin":
        # El admin recibe sus avisos exclusivamente por la cola push
        # (/api/bridge/notificaciones): techo urgente y supervisión de pedidos urgentes.
        # El polling de todos los pedidos en estado alertable es el canal del comprador,
        # no del supervisor. Devolver vacío evita que la agenda del admin muestre
        # el seguimiento diario de cada pedido de todos los hoteles.
        return jsonify({"alertas": [], "num_alertas": 0,
                        "usuario": session.get("username"), "rol": rol})
    elif rol == "compras":
        # Hoteles asignados al comprador en usuario_comprador_hoteles
        rows = rows_to_list(query(
            "SELECT hotel_id FROM usuario_comprador_hoteles WHERE usuario_id=%s",
            (user_id,)
        ))
        hotel_ids = [r["hotel_id"] for r in rows]
        if not hotel_ids:
            return jsonify({"alertas": [], "num_alertas": 0,
                            "usuario": session.get("username"), "rol": rol})
        placeholders = ",".join(["%s"] * len(hotel_ids))
        filtro_hotel_sql = f"AND p.hotel_id IN ({placeholders})"
        filtro_args      = hotel_ids
    elif rol == "hotel":
        # Hoteles asignados al usuario hotel en usuario_hoteles (lectura)
        rows = rows_to_list(query(
            "SELECT hotel_id FROM usuario_hoteles WHERE usuario_id=%s",
            (user_id,)
        ))
        hotel_ids = [r["hotel_id"] for r in rows]
        if not hotel_ids:
            return jsonify({"alertas": [], "num_alertas": 0,
                            "usuario": session.get("username"), "rol": rol})
        placeholders = ",".join(["%s"] * len(hotel_ids))
        filtro_hotel_sql = f"AND p.hotel_id IN ({placeholders})"
        filtro_args      = hotel_ids
    else:
        # Rol desconocido: sin alertas
        return jsonify({"alertas": [], "num_alertas": 0,
                        "usuario": session.get("username"), "rol": rol})

    # ── Consulta de pedidos en estados alertables ─────────────────────────────
    # Usa PEDIDO_SELECT_STATS (sin subconsultas de proveedor_contactos) y
    # _clasificar_alertas (fuente única de verdad para umbrales y niveles).
    alertas_raw = rows_to_list(query(f"""
        {PEDIDO_SELECT_STATS}
        WHERE p.estado IN (
            'ENVIADO AL PROVEEDOR',
            'PENDIENTE FIRMA DIRECCION COMPRAS',
            'PENDIENTE DE FIRMA DIRECCION HOTEL',
            'ENTREGA PARCIAL',
            'PENDIENTE COTIZACIÓN'
        )
          AND (
            p.fecha_tramitacion IS NOT NULL
            OR (p.estado = 'PENDIENTE COTIZACIÓN' AND p.fecha_solicitud IS NOT NULL)
          )
          {filtro_hotel_sql}
        ORDER BY p.fecha_tramitacion ASC
    """, filtro_args))
    cfg_activar_plazo_bridge = bool(int(get_config().get("activar_uso_plazo_entrega", 1) or 0))
    alertas = _clasificar_alertas(alertas_raw, cfg_activar_plazo_bridge)

    # v12.29.47 (PRUEBA): "alertas" pasa a contener SOLO lo pendiente de
    # entregar como popup (ver _filtrar_popups_no_vistos). Los totales sin
    # filtrar se mandan aparte en total_activas/urgentes_activas/
    # normales_activas, para que el resumen del saludo diario de Organizador
    # Princess (get_resumen_alertas()) siga reflejando TODAS las alertas
    # activas y no solo las nuevas de este ciclo.
    usuario = session.get("username", "").lower()
    nuevas  = _filtrar_popups_no_vistos(usuario, alertas)

    return jsonify({
        "alertas":           nuevas,
        "num_alertas":       len(nuevas),
        "total_activas":     len(alertas),
        "urgentes_activas":  sum(1 for a in alertas if a.get("nivel_alerta") == "urgente"),
        "normales_activas":  sum(1 for a in alertas if a.get("nivel_alerta") != "urgente"),
        "usuario":           session.get("username"),
        "nombre":            session.get("nombre"),
        "rol":               rol,
    })


# ── API Bridge: cola de notificaciones push (v10.7.7) ────────────────────────

@app.route("/api/bridge/notificaciones", methods=["GET"])
@login_required
def bridge_notificaciones_usuario():
    """
    Devuelve las notificaciones pendientes (no leídas) para el usuario de la sesión
    y las marca como leídas en la misma transacción.

    Garantiza paridad total con Telegram: cada vez que se envía un Telegram a un
    comprador o admin, se encola una fila en bridge_notificaciones para que
    main_agenda la reciba como popup — inmediato salvo para 'cambio_estado'
    (2026-08-14), que se retiene 5 minutos desde el último cambio del mismo
    pedido antes de quedar visible aquí (columna visible_en, ver
    _encolar_bridge_notificacion) para no generar un popup por cada cambio
    de estado si hay varios seguidos sobre el mismo pedido.

    Respuesta:
    {
        "notificaciones": [
            {
                "id": 42,
                "tipo": "cambio_estado",      -- 'cambio_estado'|'alerta_auto'|'techo'|'familia_repetida'|'supervision'
                "pedido_id": 123,             -- puede ser null
                "titulo": "...",
                "mensaje": "...",
                "nivel": "urgente",           -- 'aviso'|'urgente'
                "creado_en": "2026-05-25T..."
            }, ...
        ],
        "total": 3,
        "usuario": "comprador1",
        "rol": "compras"
    }
    """
    usuario  = session.get("username", "").lower()
    rol      = session.get("rol", "")

    try:
        rows = rows_to_list(query(
            """SELECT id, tipo, pedido_id, titulo, mensaje, nivel, creado_en
               FROM bridge_notificaciones
               WHERE usuario = %s AND leido = FALSE AND visible_en <= NOW()
               ORDER BY creado_en ASC""",
            (usuario,)
        ))
    except Exception as exc:
        log.warning("bridge_notif GET: error leyendo notificaciones — %s", exc)
        return jsonify({"notificaciones": [], "total": 0, "usuario": usuario, "rol": rol})

    if rows:
        ids = [r["id"] for r in rows]
        placeholders = ",".join(["%s"] * len(ids))
        try:
            db = get_db()
            db.cursor().execute(
                f"UPDATE bridge_notificaciones SET leido=TRUE WHERE id IN ({placeholders})",
                ids
            )
            db.commit()
        except Exception as exc:
            log.warning("bridge_notif: no se pudo marcar como leído — %s", exc)

    # Serializar timestamps a ISO string
    notifs = []
    for r in rows:
        r = dict(r)
        if hasattr(r.get("creado_en"), "isoformat"):
            r["creado_en"] = r["creado_en"].isoformat()
        notifs.append(r)

    return jsonify({
        "notificaciones": notifs,
        "total":          len(notifs),
        "usuario":        usuario,
        "rol":            rol,
    })


# ── API Reset completo (admin only) ───────────────────────────────────────────

@app.route("/api/importar/backup", methods=["GET"])
@admin_required
def exportar_backup_previo():
    """Genera y devuelve un Excel con todos los pedidos actuales (backup previo al reset)."""
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment

        pedidos = rows_to_list(query(f"{PEDIDO_SELECT} ORDER BY p.norden ASC"))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BACKUP PEDIDOS"

        HEADERS = [
            "Nº ORDEN", "HOTEL", "DEPARTAMENTO", "FECHA SOLICITUD",
            "FECHA ENVÍO Vº Bº", "PEDIDO Nº", "FECHA TRAMITACIÓN",
            "Nº PRESUPUESTO", "ESTADO", "Nº ENTRADA DALI / SAP",
            "COMUNICADO A&B", "COMUNICADO JEFE DEP.",
            "PARTE ROTURA", "PARTE AMPLIACIÓN",
            "PROVEEDOR", "EMAIL PROVEEDOR", "TELÉFONO", "CONTACTO",
            "OBSERVACIONES", "CREADO POR", "CREADO EN",
        ]
        ws.append(HEADERS)
        header_fill = PatternFill("solid", fgColor="8B0000")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        def strip_tz(val):
            if hasattr(val, "tzinfo") and val.tzinfo is not None:
                return val.replace(tzinfo=None)
            return val

        for p in pedidos:
            ws.append([
                p.get("norden"), p.get("hotel_codigo"), p.get("departamento_nombre"),
                strip_tz(p.get("fecha_solicitud")), strip_tz(p.get("fecha_envio_visto_bueno")),
                p.get("pedido_num"), strip_tz(p.get("fecha_tramitacion")),
                p.get("presupuesto_num"), p.get("estado"),
                format_albaran_display(p.get("entrada_albaran_num")),
                "SÍ" if p.get("comunicado_ab") else "NO",
                "SÍ" if p.get("comunicado_jefe_dep") else "NO",
                "SÍ" if p.get("parte_rotura") else "NO",
                "SÍ" if p.get("parte_ampliacion") else "NO",
                p.get("proveedor_nombre"), p.get("proveedor_email"),
                p.get("proveedor_telefono"), p.get("proveedor_contacto"),
                p.get("observaciones"), p.get("creado_por_nombre"), strip_tz(p.get("creado_en")),
            ])

        COL_WIDTHS = [8,8,22,14,14,16,14,18,32,16,12,14,12,12,28,28,14,16,40,18,18]
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from flask import send_file
        filename = f"BACKUP_PEDIDOS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/importar/reset", methods=["POST"])
@admin_required
def reset_e_importar():
    """
    Borra TODOS los pedidos (incluidos adjuntos PDF/imágenes via CASCADE)
    y el historial. Luego importa el Excel recibido desde cero.
    Solo accesible para administradores.
    """
    from datetime import datetime as dt
    try:
        import openpyxl

        if "archivo" not in request.files:
            return jsonify({"ok": False, "error": "No se recibió ningún archivo"}), 400

        archivo = request.files["archivo"]
        if not archivo.filename.endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "error": "El archivo debe ser .xlsx"}), 400

        db  = get_db()
        uid = current_user_id()

        # ── 1. Borrado total (CASCADE elimina adjuntos, historial, eliminados) ──
        with db.cursor() as cur_del:
            cur_del.execute("DELETE FROM pedidos")                # CASCADE → adjuntos + historial
            # pedidos_eliminados puede no existir si no se ejecutó la migración
            cur_del.execute("""
                DO $$ BEGIN
                    IF EXISTS (
                        SELECT 1 FROM information_schema.tables
                        WHERE table_schema = 'public'
                          AND table_name   = 'pedidos_eliminados'
                    ) THEN
                        DELETE FROM pedidos_eliminados;
                    END IF;
                END $$;
            """)

        log.info("RESET: todos los pedidos eliminados por admin user_id=%s", uid)

        # ── 2. Leer Excel y construir filas (misma lógica que /api/importar) ──
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]

        def col_raw(row, name):
            try:
                idx = headers.index(name)
                return row[idx].value
            except (ValueError, IndexError):
                return None

        def col(row, name):
            v = col_raw(row, name)
            return str(v).strip() if v is not None else None

        def parse_date(val):
            if val is None:
                return None
            if hasattr(val, 'strftime'):
                return val.strftime("%Y-%m-%d")
            try:
                n = int(float(str(val)))
                if 30000 < n < 60000:
                    from openpyxl.utils.datetime import from_excel
                    return from_excel(n).strftime("%Y-%m-%d")
            except Exception:
                pass
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return dt.strptime(str(val).strip(), fmt).strftime("%Y-%m-%d")
                except Exception:
                    pass
            log.warning("[reset_e_importar] Fecha no reconocida en Excel, valor descartado: %r", val)
            return None

        def bool_val(val):
            if not val:
                return 0
            return 1 if str(val).strip().upper() in ("SÍ", "SI", "S", "1", "TRUE", "YES") else 0

        hoteles_cache     = {r["codigo"]: r["id"] for r in rows_to_list(query("SELECT id, codigo FROM hoteles WHERE activo=1"))}
        deptos_cache      = {r["nombre"].upper(): r["id"] for r in rows_to_list(query("SELECT id, nombre FROM departamentos WHERE activo=1"))}
        proveedores_cache = {r["nombre"].upper(): r["id"] for r in rows_to_list(query("SELECT id, nombre FROM proveedores WHERE activo=1"))}

        errores = []
        filas_validas = []

        # Numeración correlativa desde 1 (reset completo)
        year = datetime.now().year

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            hotel_codigo = col(row, "HOTEL")
            if not hotel_codigo:
                continue

            hotel_id = hoteles_cache.get(str(hotel_codigo).upper())
            if not hotel_id:
                errores.append(f"Fila {i}: hotel '{hotel_codigo}' no encontrado")
                continue

            depto_nombre = col(row, "DEPARTAMENTO")
            depto_id = deptos_cache.get(str(depto_nombre).upper()) if depto_nombre else None

            prov_nombre = col(row, "PROVEEDOR")
            prov_id = proveedores_cache.get(str(prov_nombre).upper()) if prov_nombre else None

            estado_raw = col(row, "ESTADO")
            estado = estado_raw if estado_raw in ESTADOS_VALIDOS else "PENDIENTE FIRMA DIRECCION COMPRAS"

            # norden siempre correlativo desde 1, independiente del Excel
            filas_validas.append({
                "norden":    len(filas_validas) + 1,
                "hotel_id":  hotel_id, "depto_id": depto_id,
                "fecha_sol": parse_date(col_raw(row, "FECHA SOLICITUD")),
                "fecha_env": parse_date(col_raw(row, "FECHA ENVÍO Vº Bº")),
                "fecha_tra": parse_date(col_raw(row, "FECHA TRAMITACIÓN")),
                "pedido_num":  col(row, "PEDIDO Nº"),
                "presup_num":  col(row, "Nº PRESUPUESTO"),
                "albaran_num": col(row, "Nº ENTRADA DALI / SAP"),
                "estado":    estado,
                "com_ab":    bool_val(col(row, "COMUNICADO A&B")),
                "com_jefe":  bool_val(col(row, "COMUNICADO JEFE DEP.")),
                "p_rotura":  bool_val(col(row, "PARTE ROTURA")),
                "p_amplia":  bool_val(col(row, "PARTE AMPLIACIÓN")),
                "prov_id":   prov_id,
                "obs":       col(row, "OBSERVACIONES"),
            })

        # ── 3. Bulk insert ─────────────────────────────────────────────────────
        insertados = 0
        if filas_validas:
            from psycopg2.extras import execute_values
            _nombre = session.get("nombre")
            with db.cursor() as cur_i:
                pedido_rows = [
                    (f["norden"], f["hotel_id"], f["depto_id"],
                     f["fecha_sol"], f["fecha_env"], f["fecha_tra"],
                     f["pedido_num"], f["presup_num"], f["albaran_num"],
                     f["estado"], f["com_ab"], f["com_jefe"],
                     f["p_rotura"], f["p_amplia"], f["prov_id"],
                     f["obs"], uid, uid, _nombre, _nombre)
                    for f in filas_validas
                ]
                ids = execute_values(cur_i, """
                    INSERT INTO pedidos (
                        norden, hotel_id, departamento_id,
                        fecha_solicitud, fecha_envio_visto_bueno, fecha_tramitacion,
                        pedido_num, presupuesto_num, entrada_albaran_num,
                        estado, comunicado_ab, comunicado_jefe_dep,
                        parte_rotura, parte_ampliacion,
                        proveedor_id, observaciones,
                        creado_por_id, modificado_por_id,
                        creado_por_nombre, modificado_por_nombre
                    ) VALUES %s RETURNING id
                """, pedido_rows, fetch=True)

                insertados = len(ids)

                historial_rows = [
                    (ids[idx]["id"], filas_validas[idx]["estado"], uid, _nombre, "Importado desde Excel (reset completo)")
                    for idx in range(len(ids))
                ]
                execute_values(cur_i, """
                    INSERT INTO historial_estados (pedido_id, estado_nuevo, usuario_id, usuario_nombre, nota)
                    VALUES %s
                """, historial_rows)

        db.commit()
        log.info("RESET completado: %d pedidos importados por admin user_id=%s", insertados, uid)
        return jsonify({"ok": True, "insertados": insertados, "errores": errores})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── API Importar Excel ─────────────────────────────────────────────────────────

@app.route("/api/importar", methods=["POST"])
@login_required
def importar_excel():
    try:
        import openpyxl
        from datetime import datetime as dt

        if "archivo" not in request.files:
            return jsonify({"ok": False, "error": "No se recibió ningún archivo"}), 400

        archivo = request.files["archivo"]
        if not archivo.filename.endswith((".xlsx", ".xls")):
            return jsonify({"ok": False, "error": "El archivo debe ser .xlsx"}), 400

        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active

        # Leer cabeceras de la primera fila
        headers = [str(c.value).strip().upper() if c.value else "" for c in ws[1]]

        def col_raw(row, name):
            try:
                idx = headers.index(name)
                return row[idx].value
            except (ValueError, IndexError):
                return None

        def col(row, name):
            v = col_raw(row, name)
            return str(v).strip() if v is not None else None

        def parse_date(val):
            if val is None:
                return None
            if hasattr(val, 'strftime'):
                return val.strftime("%Y-%m-%d")
            try:
                n = int(float(str(val)))
                if 30000 < n < 60000:
                    from openpyxl.utils.datetime import from_excel
                    return from_excel(n).strftime("%Y-%m-%d")
            except Exception:
                pass
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return dt.strptime(str(val).strip(), fmt).strftime("%Y-%m-%d")
                except Exception:
                    pass
            log.warning("[importar_excel] Fecha no reconocida en Excel, valor descartado: %r", val)
            return None

        def bool_val(val):
            if not val:
                return 0
            return 1 if str(val).strip().upper() in ("SÍ", "SI", "S", "1", "TRUE", "YES") else 0

        db = get_db()
        uid = current_user_id()

        # Cachés para no consultar la BD en cada fila
        if _puede_ver_hotel_pruebas():
            hoteles_cache = {r["codigo"]: r["id"] for r in rows_to_list(query("SELECT id, codigo FROM hoteles WHERE activo=1"))}
        else:
            # Hotel de pruebas ('PR') excluido: solo admin o el usuario
            # dedicado a estas pruebas pueden crear pedidos en él, ni
            # siquiera vía importación masiva.
            hoteles_cache = {r["codigo"]: r["id"] for r in rows_to_list(query(
                "SELECT id, codigo FROM hoteles WHERE activo=1 AND codigo <> %s", (HOTEL_CODIGO_PRUEBAS,)
            ))}
        deptos_cache       = {r["nombre"].upper(): r["id"] for r in rows_to_list(query("SELECT id, nombre FROM departamentos WHERE activo=1"))}
        proveedores_cache  = {r["nombre"].upper(): r["id"] for r in rows_to_list(query("SELECT id, nombre FROM proveedores WHERE activo=1"))}

        errores = []
        filas_validas = []

        # 1. Obtener norden base en UNA sola query
        year = datetime.now().year
        with db.cursor() as cur_n:
            cur_n.execute(
                "SELECT COALESCE(MAX(norden), 0) as mx FROM pedidos WHERE EXTRACT(YEAR FROM creado_en) = %s",
                (year,)
            )
            base_norden = (cur_n.fetchone()["mx"] or 0) + 1

        # 2. Procesar todas las filas en memoria (sin queries)
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            hotel_codigo = col(row, "HOTEL")
            if not hotel_codigo:
                continue

            hotel_id = hoteles_cache.get(str(hotel_codigo).upper())
            if not hotel_id:
                errores.append(f"Fila {i}: hotel '{hotel_codigo}' no encontrado")
                continue

            depto_nombre = col(row, "DEPARTAMENTO")
            depto_id = deptos_cache.get(str(depto_nombre).upper()) if depto_nombre else None

            prov_nombre = col(row, "PROVEEDOR")
            prov_id = proveedores_cache.get(str(prov_nombre).upper()) if prov_nombre else None

            estado_raw = col(row, "ESTADO")
            estado = estado_raw if estado_raw in ESTADOS_VALIDOS else "PENDIENTE FIRMA DIRECCION COMPRAS"

            filas_validas.append({
                "norden": base_norden + len(filas_validas),
                "hotel_id": hotel_id, "depto_id": depto_id,
                "fecha_sol": parse_date(col_raw(row, "FECHA SOLICITUD")),
                "fecha_env": parse_date(col_raw(row, "FECHA ENVÍO Vº Bº")),
                "fecha_tra": parse_date(col_raw(row, "FECHA TRAMITACIÓN")),
                "pedido_num": col(row, "PEDIDO Nº"),
                "presup_num": col(row, "Nº PRESUPUESTO"),
                "albaran_num": col(row, "Nº ENTRADA DALI / SAP"),
                "estado": estado,
                "com_ab": bool_val(col(row, "COMUNICADO A&B")),
                "com_jefe": bool_val(col(row, "COMUNICADO JEFE DEP.")),
                "p_rotura": bool_val(col(row, "PARTE ROTURA")),
                "p_amplia": bool_val(col(row, "PARTE AMPLIACIÓN")),
                "prov_id": prov_id,
                "obs": col(row, "OBSERVACIONES"),
            })

        # 3. Bulk insert en 2 queries únicas
        insertados = 0
        if filas_validas:
            from psycopg2.extras import execute_values
            _nombre = session.get("nombre")
            with db.cursor() as cur_i:
                pedido_rows = [
                    (f["norden"], f["hotel_id"], f["depto_id"],
                     f["fecha_sol"], f["fecha_env"], f["fecha_tra"],
                     f["pedido_num"], f["presup_num"], f["albaran_num"],
                     f["estado"], f["com_ab"], f["com_jefe"],
                     f["p_rotura"], f["p_amplia"], f["prov_id"],
                     f["obs"], uid, uid, _nombre, _nombre)
                    for f in filas_validas
                ]
                ids = execute_values(cur_i, """
                    INSERT INTO pedidos (
                        norden, hotel_id, departamento_id,
                        fecha_solicitud, fecha_envio_visto_bueno, fecha_tramitacion,
                        pedido_num, presupuesto_num, entrada_albaran_num,
                        estado, comunicado_ab, comunicado_jefe_dep,
                        parte_rotura, parte_ampliacion,
                        proveedor_id, observaciones,
                        creado_por_id, modificado_por_id,
                        creado_por_nombre, modificado_por_nombre
                    ) VALUES %s RETURNING id
                """, pedido_rows, fetch=True)

                insertados = len(ids)

                historial_rows = [
                    (ids[idx]["id"], filas_validas[idx]["estado"], uid, _nombre, "Importado desde Excel")
                    for idx in range(len(ids))
                ]
                execute_values(cur_i, """
                    INSERT INTO historial_estados (pedido_id, estado_nuevo, usuario_id, usuario_nombre, nota)
                    VALUES %s
                """, historial_rows)

        db.commit()
        return jsonify({"ok": True, "insertados": insertados, "errores": errores})

    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── API Exportar Excel ─────────────────────────────────────────────────────────

@app.route("/api/exportar")
@login_required
def exportar_excel():
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file

        rol = session.get("rol", "")
        if rol == "hotel":
            hoteles_ids = session.get("hoteles_ids", [])
            if not hoteles_ids:
                pedidos = []
            else:
                placeholders = ",".join(["%s"] * len(hoteles_ids))
                pedidos = rows_to_list(query(
                    f"{PEDIDO_SELECT} WHERE p.hotel_id IN ({placeholders}) ORDER BY p.creado_en DESC",
                    tuple(hoteles_ids)
                ))
        elif _puede_ver_hotel_pruebas():
            pedidos = rows_to_list(query(f"{PEDIDO_SELECT} ORDER BY p.creado_en DESC"))
        else:
            # compras: mismo listado que admin salvo el hotel de pruebas
            pedidos = rows_to_list(query(
                f"{PEDIDO_SELECT} WHERE h.codigo IS NULL OR h.codigo <> %s ORDER BY p.creado_en DESC",
                (HOTEL_CODIGO_PRUEBAS,)
            ))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CONTROL PEDIDOS"

        HEADERS = [
            "Nº ORDEN", "HOTEL", "DEPARTAMENTO", "FECHA SOLICITUD",
            "FECHA ENVÍO Vº Bº", "PEDIDO Nº", "FECHA TRAMITACIÓN",
            "Nº PRESUPUESTO", "ESTADO", "Nº ENTRADA DALI / SAP",
            "COMUNICADO A&B", "COMUNICADO JEFE DEP.",
            "PARTE ROTURA", "PARTE AMPLIACIÓN",
            "PROVEEDOR", "EMAIL PROVEEDOR", "TELÉFONO", "CONTACTO",
            "OBSERVACIONES", "CREADO POR", "CREADO EN",
        ]
        ws.append(HEADERS)
        header_fill = PatternFill("solid", fgColor="1a3a6b")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        ESTADO_COLORES = {
            "ENTREGADO":                         "d4edda",
            "ENVIADO AL PROVEEDOR":              "cce5ff",
            "ENTREGA PARCIAL":                   "fff3cd",
            "PENDIENTE FIRMA DIRECCION COMPRAS": "ffeeba",
            "PENDIENTE DE FIRMA DIRECCION HOTEL":"ffe8a1",
            "CANCELADO":                           "f8d7da",
        }

        def strip_tz(val):
            """Elimina tzinfo de datetimes para compatibilidad con openpyxl/Excel."""
            if hasattr(val, "tzinfo") and val.tzinfo is not None:
                return val.replace(tzinfo=None)
            return val

        for p in pedidos:
            ws.append([
                p.get("norden"), p.get("hotel_codigo"), p.get("departamento_nombre"),
                strip_tz(p.get("fecha_solicitud")), strip_tz(p.get("fecha_envio_visto_bueno")),
                p.get("pedido_num"), strip_tz(p.get("fecha_tramitacion")),
                p.get("presupuesto_num"), p.get("estado"),
                format_albaran_display(p.get("entrada_albaran_num")),
                "SÍ" if p.get("comunicado_ab") else "NO",
                "SÍ" if p.get("comunicado_jefe_dep") else "NO",
                "SÍ" if p.get("parte_rotura") else "NO",
                "SÍ" if p.get("parte_ampliacion") else "NO",
                p.get("proveedor_nombre"), p.get("proveedor_email"),
                p.get("proveedor_telefono"), p.get("proveedor_contacto"),
                p.get("observaciones"), p.get("creado_por_nombre"), strip_tz(p.get("creado_en")),
            ])
            color = ESTADO_COLORES.get(p.get("estado", ""), "FFFFFF")
            fill  = PatternFill("solid", fgColor=color)
            for cell in ws[ws.max_row]:
                cell.fill = fill

        COL_WIDTHS = [8,8,22,14,14,16,14,18,32,16,12,14,12,12,28,28,14,16,40,18,18]
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"CONTROL_PEDIDOS_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({"error": "openpyxl no instalado"}), 500

# ── Adjuntos (PDFs e imágenes de artículos) ───────────────────────────────────

TIPOS_ADJUNTO_VALIDOS = {
    "presupuesto_pdf", "pedido_pdf", "imagen_articulo",
    "pedido_doc",       # PDF/Word/correo vinculado a Nº Pedido DALI/SAP
    "presupuesto_doc",  # PDF/Word/correo vinculado a Nº Presupuesto
    "solicitud_doc",    # Excel/PDF/Word + correo vinculado a Fecha Solicitud
    "vb_eml",           # Correo .eml/.msg vinculado a Fecha Envio Vº Bº
    "tramit_eml",       # Correo .eml/.msg vinculado a Fecha Tramitacion
    "firma_techo_doc",  # (2026-08-03) Excel/PDF/Word + correo — listado de
                        # apoyo adjunto a la solicitud de firma cuando el
                        # pedido está sujeto a techo de gastos
}
MIME_PERMITIDOS = {
    "application/pdf",
    "image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "message/rfc822",
    "application/vnd.ms-outlook",
    "application/octet-stream",
}
EXT_CORREO = {".eml", ".msg"}
EXT_DOC    = {".xlsx", ".xls", ".docx", ".doc", ".pdf"}
MIME_SOLICITUD_DOC = {
    "application/pdf",
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "message/rfc822", "application/vnd.ms-outlook", "application/octet-stream",
}
MIME_CORREO = {"message/rfc822", "application/vnd.ms-outlook", "application/octet-stream"}
MAX_ADJUNTO_BYTES = 20 * 1024 * 1024  # 20 MB por archivo (límite absoluto de respaldo)

# ── Límites de peso ajustados por tipo de contenido ──────────────────────────
# Los PDF/Word de gestión normal (albaranes, presupuestos, solicitudes) no
# deberían pesar más que un escaneado de pocas páginas. Los correos .eml/.msg
# llevan adjuntos incrustados, por lo que su límite es algo mayor pero sigue
# acotado para no arrastrar archivos grandes dentro del correo.
MAX_BYTES_DOCUMENTO = 5 * 1024 * 1024   # 5 MB — PDF / Word / Excel
MAX_BYTES_CORREO    = 3 * 1024 * 1024   # 3 MB — .eml / .msg
MAX_BYTES_IMAGEN    = 2 * 1024 * 1024   # 2 MB — imagen_articulo

# ── Límites de cantidad por apartado ──────────────────────────────────────────
# En los apartados que aceptan documento + correo, se cuentan por separado.
MAX_DOCUMENTOS_POR_APARTADO = 3
MAX_CORREOS_POR_APARTADO    = 1

# ── Fix egress — miniaturas para imagen_articulo (Jul 2026) ─────────────────
THUMB_MAX_ANCHO = 240   # px — de sobra para la miniatura de la lista de adjuntos
THUMB_JPEG_CALIDAD = 70


def _generar_thumbnail(datos_originales, mime_type):
    """Genera una miniatura JPEG reducida a partir de los bytes de una imagen.

    Devuelve (bytes_thumb, mime_thumb) o (None, None) si la imagen no se
    puede procesar (por ejemplo, un formato no soportado por Pillow) — en
    ese caso el llamador debe servir la imagen original como respaldo.
    """
    try:
        from PIL import Image
        import io as _io
        img = Image.open(_io.BytesIO(datos_originales))
        img.load()
        # Aplanar transparencia (RGBA/P) sobre fondo blanco: el destino es
        # JPEG, que no soporta canal alfa.
        if img.mode in ("RGBA", "LA", "P"):
            fondo = Image.new("RGB", img.size, (255, 255, 255))
            img_rgba = img.convert("RGBA")
            fondo.paste(img_rgba, mask=img_rgba.split()[-1])
            img = fondo
        elif img.mode != "RGB":
            img = img.convert("RGB")
        if img.width > THUMB_MAX_ANCHO:
            ratio = THUMB_MAX_ANCHO / float(img.width)
            nuevo_alto = max(1, int(img.height * ratio))
            img = img.resize((THUMB_MAX_ANCHO, nuevo_alto), Image.LANCZOS)
        buf = _io.BytesIO()
        img.save(buf, format="JPEG", quality=THUMB_JPEG_CALIDAD, optimize=True)
        return buf.getvalue(), "image/jpeg"
    except Exception as e:
        log.warning(f"No se pudo generar thumbnail: {e}")
        return None, None

# ── Lectura del PDF de pedido oficial PRINCESS (2026-08-28) ────────────────────
# A petición de Víctor: el apartado "Nº Pedido (DALI/SAP)" del formulario pasa
# a admitir ÚNICAMENTE el PDF de pedido oficial que genera SAP/DALI (siempre
# el mismo formato, con nombre de archivo libre) — y, al subirlo, la app debe
# leer y rellenar sola dos celdas que dejan de ser editables a mano:
#   - "Nº Pedido (DALI/SAP)"   ← la línea "PEDIDO 00016287" del PDF (sin ceros
#                                 a la izquierda, igual que _normalizar_pedido_num).
#   - "Total Pedido (€)"       ← la SUMA de la columna "Importe" de cada línea
#                                 de artículo — NUNCA el "Total Pedido..." que
#                                 el propio PDF trae al pie, porque ese valor no
#                                 incluye los descuentos aplicados y no es fiable
#                                 (ver petición de Víctor, 28/08).
# Mismo enfoque que _comparar_listado_pdf_logica() (pypdf.extract_text() +
# regex tolerante, ver comentario junto a _PATRON_LISTADO_SIMPLIFICADO): el
# orden del texto que devuelve pypdf NO seguía el orden visual de las columnas
# (comprobado con este mismo PDF de ejemplo — "Cantidad Precio Importe" salen
# como tres números seguidos, pero intercalados con cabeceras y pies de página
# en otro orden), así que en vez de intentar reconstruir la tabla completa,
# se buscan directamente los tríos "Cantidad(,dddd) Precio(,dd) Importe(,dd)"
# consecutivos — el único sitio del documento donde aparecen tres importes
# seguidos con ese patrón de decimales es una línea de artículo real.
_PATRON_PEDIDO_NUM_OFICIAL = re.compile(r'\bPEDIDO\s+(\d+)\b')
_PATRON_IMPORTE_LINEA_OFICIAL = re.compile(
    r'(\d{1,3}(?:\.\d{3})*,\d{2,4})\s+'   # Cantidad
    r'(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+'   # Precio
    r'(-?\d{1,3}(?:\.\d{3})*,\d{2})'      # Importe (el que nos interesa sumar)
)
# (2026-08-28) A petición de Víctor: además de Nº de Pedido y Total, se lee
# también "Fecha Pedido" y "Fecha Entrega" del mismo PDF oficial — ver
# _parsear_pdf_pedido_oficial() y el uso que hace de estos dos campos
# upload_adjunto() (comprobación/auto-relleno de Fecha tramitación y Fecha
# de entrega específica). Ambas etiquetas van seguidas directamente de la
# fecha en el texto extraído ("Fecha Pedido 21/08/2026", "Fecha Entrega
# 21/09/2026" — confirmado con el PDF real de ejemplo), así que un patrón
# simple basta; a diferencia del Nº de Pedido y el Total, estos dos campos
# son opcionales — si no se reconocen, no se rechaza el PDF (solo no hay
# fecha que proponer).
_PATRON_FECHA_PEDIDO_OFICIAL = re.compile(r'\bFecha\s+Pedido\s+(\d{2}/\d{2}/\d{4})\b')
_PATRON_FECHA_ENTREGA_OFICIAL = re.compile(r'\bFecha\s+Entrega\s+(\d{2}/\d{2}/\d{4})\b')

def _parsear_pdf_pedido_oficial(pdf_bytes: bytes) -> dict:
    """
    Lee un PDF de pedido oficial PRINCESS (SAP/DALI) y devuelve
    {"pedido_num": "16287", "total_pedido": 4614.60,
    "fecha_pedido_iso": "2026-08-21"|None, "fecha_entrega_iso": "2026-09-21"|None}.

    Las dos fechas son opcionales (ver comentario junto a los patrones de
    arriba) — el Nº de Pedido y el Total siguen siendo los únicos campos
    que, si faltan, hacen rechazar el PDF entero.

    Lanza ValueError con un mensaje pensado para mostrarse tal cual al
    usuario (ver upload_adjunto) si el PDF no se puede leer o no tiene la
    estructura esperada — nunca devuelve un resultado parcial o adivinado:
    o se reconoce el documento con garantías, o se rechaza con un mensaje
    claro pidiendo el PDF oficial correcto.
    """
    try:
        from pypdf import PdfReader
        import io
        reader = PdfReader(io.BytesIO(pdf_bytes))
        texto = ""
        for pagina in reader.pages:
            texto += (pagina.extract_text() or "") + "\n"
    except Exception as exc:
        log.warning(f"[PEDIDO-DOC] Error leyendo el PDF adjuntado: {exc}")
        raise ValueError(
            "No se ha podido leer el PDF adjuntado. Adjunte únicamente el PDF del pedido "
            "oficial PRINCESS (el que genera SAP/DALI) en este apartado."
        )

    m_pedido = _PATRON_PEDIDO_NUM_OFICIAL.search(texto)
    lineas_importe = _PATRON_IMPORTE_LINEA_OFICIAL.findall(texto)
    if not m_pedido or not lineas_importe:
        raise ValueError(
            "El PDF adjuntado no tiene el formato del pedido oficial PRINCESS — "
            "adjunte únicamente el PDF del pedido oficial (el que genera SAP/DALI, "
            "con el Nº de Pedido y las líneas de artículos con su importe) en este apartado."
        )

    pedido_num = _normalizar_pedido_num(m_pedido.group(1))
    total_pedido = round(sum(_parse_importe_es(l[2]) for l in lineas_importe), 2)

    m_fecha_pedido  = _PATRON_FECHA_PEDIDO_OFICIAL.search(texto)
    m_fecha_entrega = _PATRON_FECHA_ENTREGA_OFICIAL.search(texto)
    fecha_pedido_iso  = _parsear_fecha_es_a_iso(m_fecha_pedido.group(1))  if m_fecha_pedido  else None
    fecha_entrega_iso = _parsear_fecha_es_a_iso(m_fecha_entrega.group(1)) if m_fecha_entrega else None

    return {
        "pedido_num": pedido_num,
        "total_pedido": total_pedido,
        "fecha_pedido_iso": fecha_pedido_iso,
        "fecha_entrega_iso": fecha_entrega_iso,
    }

@app.route("/api/pedidos/<int:pid>/adjuntos", methods=["GET"])
@login_required
def get_adjuntos(pid):
    rows = query(
        "SELECT id, tipo, nombre, mime_type, es_correo, creado_en FROM pedido_adjuntos WHERE pedido_id=%s ORDER BY tipo, creado_en",
        (pid,)
    )
    return jsonify({"ok": True, "adjuntos": rows_to_list(rows)})


@app.route("/api/pedidos/<int:pid>/adjuntos", methods=["POST"])
@login_required
def upload_adjunto(pid):
    # Verificar que el pedido existe
    pedido = query("SELECT id FROM pedidos WHERE id=%s", (pid,), one=True)
    if not pedido:
        return jsonify({"ok": False, "error": "Pedido no encontrado"}), 404

    tipo = request.form.get("tipo", "")
    if tipo not in TIPOS_ADJUNTO_VALIDOS:
        return jsonify({"ok": False, "error": f"Tipo inválido. Valores: {', '.join(TIPOS_ADJUNTO_VALIDOS)}"}), 400

    if "archivo" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió ningún archivo"}), 400

    archivo = request.files["archivo"]
    if not archivo.filename:
        return jsonify({"ok": False, "error": "Nombre de archivo vacío"}), 400

    datos = archivo.read()
    if len(datos) > MAX_ADJUNTO_BYTES:
        return jsonify({"ok": False, "error": "El archivo supera el límite de 20 MB"}), 400

    mime = archivo.mimetype or "application/octet-stream"
    ext  = os.path.splitext(archivo.filename.lower())[1]  # ej. ".eml", ".xlsx"
    es_correo = ext in EXT_CORREO or mime in MIME_CORREO

    if tipo in ("presupuesto_pdf", "pedido_pdf"):
        if mime != "application/pdf":
            return jsonify({"ok": False, "error": "Solo se aceptan archivos PDF en este apartado"}), 400
        if len(datos) > MAX_BYTES_DOCUMENTO:
            return jsonify({"ok": False, "error": f"El PDF supera el límite de {MAX_BYTES_DOCUMENTO // (1024*1024)} MB para este apartado"}), 400

    elif tipo == "imagen_articulo":
        if len(datos) > MAX_BYTES_IMAGEN:
            return jsonify({"ok": False, "error": f"La imagen supera el límite de {MAX_BYTES_IMAGEN // (1024*1024)} MB para este apartado"}), 400

    elif tipo == "pedido_doc":
        # (2026-08-28) A petición de Víctor: este apartado deja de admitir
        # Word y correo .eml/.msg — SOLO se admite el PDF del pedido oficial
        # PRINCESS (SAP/DALI), con el mismo formato siempre. Se comprueba
        # tanto el formato del archivo como, más abajo, que su CONTENIDO
        # sea realmente ese pedido oficial (ver _parsear_pdf_pedido_oficial):
        # nunca se guarda un archivo que no se haya podido leer como tal.
        es_pdf_pedido_doc = mime == "application/pdf" or (mime == "application/octet-stream" and ext == ".pdf")
        if not es_pdf_pedido_doc:
            return jsonify({
                "ok": False,
                "error": "Este apartado solo admite el PDF del pedido oficial PRINCESS (el que genera SAP/DALI) — "
                         "adjunte únicamente ese documento en este punto."
            }), 400
        if len(datos) > MAX_BYTES_DOCUMENTO:
            return jsonify({"ok": False, "error": f"El PDF supera el límite de {MAX_BYTES_DOCUMENTO // (1024*1024)} MB para este apartado"}), 400

        # Máximo 1 documento en esta sección — para sustituirlo hay que
        # eliminar antes el que ya está.
        n_docs_existentes = query(
            "SELECT COUNT(*) AS n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo='pedido_doc'",
            (pid,), one=True
        )["n"]
        if n_docs_existentes >= 1:
            return jsonify({"ok": False, "error": "Ya existe un documento adjunto en «Nº Pedido (DALI/SAP)». Elimínelo antes de subir uno nuevo."}), 400

        # Lectura obligatoria del PDF — si no se reconoce como pedido
        # oficial, se rechaza el archivo entero (no se llega a guardar ni
        # el adjunto ni ningún dato a medias).
        try:
            _datos_pedido_pdf = _parsear_pdf_pedido_oficial(datos)
        except ValueError as _exc_pdf:
            return jsonify({"ok": False, "error": str(_exc_pdf)}), 400

    elif tipo in ("presupuesto_doc", "solicitud_doc", "firma_techo_doc"):
        etiqueta = "PDF, Word o correo (.eml/.msg)" if tipo == "presupuesto_doc" else "Excel, Word, PDF o correo (.eml/.msg)"
        if mime not in MIME_SOLICITUD_DOC:
            return jsonify({"ok": False, "error": f"Formato no permitido. Use {etiqueta}"}), 400
        if mime == "application/octet-stream" and ext not in EXT_CORREO | EXT_DOC:
            return jsonify({"ok": False, "error": "Extensión de archivo no reconocida"}), 400

        if es_correo:
            if len(datos) > MAX_BYTES_CORREO:
                return jsonify({"ok": False, "error": f"El correo supera el límite de {MAX_BYTES_CORREO // (1024*1024)} MB para este apartado"}), 400
            n_correos = query(
                "SELECT COUNT(*) as n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo=%s AND es_correo",
                (pid, tipo), one=True
            )
            if n_correos and n_correos["n"] >= MAX_CORREOS_POR_APARTADO:
                return jsonify({"ok": False, "error": f"Ya existe un correo adjunto en este apartado. Máximo {MAX_CORREOS_POR_APARTADO}. Elimínelo antes de subir uno nuevo."}), 400
        else:
            if len(datos) > MAX_BYTES_DOCUMENTO:
                return jsonify({"ok": False, "error": f"El documento supera el límite de {MAX_BYTES_DOCUMENTO // (1024*1024)} MB para este apartado"}), 400
            n_docs = query(
                "SELECT COUNT(*) as n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo=%s AND NOT es_correo",
                (pid, tipo), one=True
            )
            if n_docs and n_docs["n"] >= MAX_DOCUMENTOS_POR_APARTADO:
                return jsonify({"ok": False, "error": f"Máximo {MAX_DOCUMENTOS_POR_APARTADO} documentos en este apartado. Elimine alguno antes de subir uno nuevo."}), 400

    elif tipo == "vb_eml":
        # v12.15.0: además del correo .eml/.msg, se acepta un PDF — cubre el
        # caso de que el correo se adjunte impreso/escaneado en PDF en vez
        # del archivo de correo original. Documento y correo son slots
        # independientes (como en pedido_doc): pueden coexistir uno de cada.
        es_pdf = mime == "application/pdf" or (mime == "application/octet-stream" and ext == ".pdf")
        if not (mime in MIME_CORREO or es_pdf):
            return jsonify({"ok": False, "error": "Solo se aceptan correos electrónicos (.eml, .msg) o PDF en este apartado"}), 400
        if mime == "application/octet-stream" and ext not in (EXT_CORREO | {".pdf"}):
            return jsonify({"ok": False, "error": "Solo se aceptan archivos .eml, .msg o PDF"}), 400

        if es_pdf:
            if len(datos) > MAX_BYTES_DOCUMENTO:
                return jsonify({"ok": False, "error": f"El PDF supera el límite de {MAX_BYTES_DOCUMENTO // (1024*1024)} MB para este apartado"}), 400
            n_docs = query(
                "SELECT COUNT(*) as n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo=%s AND NOT es_correo",
                (pid, tipo), one=True
            )
            if n_docs and n_docs["n"] >= MAX_DOCUMENTOS_POR_APARTADO:
                return jsonify({"ok": False, "error": f"Máximo {MAX_DOCUMENTOS_POR_APARTADO} PDF en este apartado. Elimine alguno antes de subir uno nuevo."}), 400
        else:
            if len(datos) > MAX_BYTES_CORREO:
                return jsonify({"ok": False, "error": f"El correo supera el límite de {MAX_BYTES_CORREO // (1024*1024)} MB para este apartado"}), 400
            existentes = query(
                "SELECT COUNT(*) as n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo=%s AND es_correo",
                (pid, tipo), one=True
            )
            if existentes and existentes["n"] >= MAX_CORREOS_POR_APARTADO:
                return jsonify({"ok": False, "error": f"Ya existe un correo adjunto en este apartado. Máximo {MAX_CORREOS_POR_APARTADO}. Elimínelo antes de subir uno nuevo."}), 400

    elif tipo == "tramit_eml":
        # (2026-08-28) A petición de Víctor, a partir de dos capturas de
        # este apartado: deja de admitir PDF — SOLO se admite el correo
        # electrónico (.eml/.msg) de envío del pedido al proveedor, para no
        # confundirlo con el PDF del pedido oficial (que tiene su propio
        # apartado obligatorio, «Nº Pedido (DALI/SAP)», con lectura
        # automática, ver _parsear_pdf_pedido_oficial). Se mantiene el
        # límite de un único correo por apartado (MAX_CORREOS_POR_APARTADO)
        # — ya se aplicaba antes de este cambio, no es nuevo.
        if mime not in MIME_CORREO:
            return jsonify({"ok": False, "error": "Solo se acepta el correo electrónico (.eml, .msg) de envío del pedido al proveedor en este apartado"}), 400
        if mime == "application/octet-stream" and ext not in EXT_CORREO:
            return jsonify({"ok": False, "error": "Solo se aceptan archivos .eml o .msg"}), 400
        if len(datos) > MAX_BYTES_CORREO:
            return jsonify({"ok": False, "error": f"El correo supera el límite de {MAX_BYTES_CORREO // (1024*1024)} MB para este apartado"}), 400
        existentes = query(
            "SELECT COUNT(*) as n FROM pedido_adjuntos WHERE pedido_id=%s AND tipo=%s AND es_correo",
            (pid, tipo), one=True
        )
        if existentes and existentes["n"] >= MAX_CORREOS_POR_APARTADO:
            return jsonify({"ok": False, "error": f"Ya existe un correo adjunto en este apartado. Máximo {MAX_CORREOS_POR_APARTADO}. Elimínelo antes de subir uno nuevo."}), 400

    else:
        if mime not in MIME_PERMITIDOS:
            return jsonify({"ok": False, "error": f"Tipo de archivo no permitido: {mime}"}), 400

    uid = current_user_id()
    db  = get_db()

    thumb_datos, thumb_mime = (None, None)
    if tipo == "imagen_articulo":
        thumb_datos, thumb_mime = _generar_thumbnail(datos, mime)

    cur = execute(
        "INSERT INTO pedido_adjuntos (pedido_id, tipo, nombre, mime_type, datos, es_correo, subido_por_id, datos_thumb, thumb_mime_type) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
        (pid, tipo, archivo.filename, mime, psycopg2.Binary(datos), es_correo, uid,
         psycopg2.Binary(thumb_datos) if thumb_datos else None, thumb_mime)
    )
    adjunto_id = cur.fetchone()["id"]
    db.commit()

    # (2026-08-28) Al subir el PDF de pedido oficial, se rellenan solas
    # "Nº Pedido (DALI/SAP)" y "Total Pedido (€)" con lo leído del PDF —
    # ver _parsear_pdf_pedido_oficial más arriba. Estas dos celdas ya no
    # se pueden escribir a mano (ver create_pedido/update_pedido): esta es
    # la ÚNICA vía por la que cambian de valor.
    respuesta = {"ok": True, "id": adjunto_id}
    if tipo == "pedido_doc":
        execute(
            "UPDATE pedidos SET pedido_num=%s, total_pedido=%s WHERE id=%s",
            (_datos_pedido_pdf["pedido_num"], _datos_pedido_pdf["total_pedido"], pid)
        )
        db.commit()
        respuesta["pedido_num"] = _datos_pedido_pdf["pedido_num"]
        respuesta["total_pedido"] = _datos_pedido_pdf["total_pedido"]
        # (2026-08-28) A petición de Víctor: "Fecha Pedido" y "Fecha
        # Entrega" del PDF oficial NO se escriben aquí en la base de datos
        # — a diferencia de pedido_num/total_pedido, «Fecha tramitación» y
        # «Fecha de entrega específica» siguen siendo campos normales,
        # editables a mano en cualquier momento, así que la decisión de
        # usarlas o no (y, si hay conflicto, cuál de las dos fechas dejar)
        # se hace en el frontend nada más recibir esta respuesta — ver
        # subirAdjuntos() en templates/index.html (rama 'pedido_doc').
        respuesta["fecha_pedido_iso"] = _datos_pedido_pdf.get("fecha_pedido_iso")
        respuesta["fecha_entrega_iso"] = _datos_pedido_pdf.get("fecha_entrega_iso")
    return jsonify(respuesta), 201


def _servir_adjunto_response(aid: int):
    """
    Construye la respuesta Flask que sirve el contenido de un adjunto de
    pedido — misma lógica de ETag/caché y de origen transparente (base de
    datos o Supabase Storage, ver _storage_descargar) que usaba
    exclusivamente download_adjunto() hasta ahora. Extraída a helper
    (2026-08-28) para poder reutilizarla también desde el enlace público
    de descarga por token (ver descargar_adjunto_publico más abajo), que
    no requiere sesión — el resto del comportamiento es idéntico.

    Devuelve la Response (o el tuple (jsonify(...), 404/502) de error) tal
    cual debe devolverla la vista Flask que llama a esta función. Devuelve
    None solo cuando el adjunto no existe — cada caller decide cómo
    responder ese caso (404 JSON para la vista con sesión, texto plano
    para el enlace público).
    """
    from flask import Response

    # ── Fix egress (Jul 2026, parte 2) ──────────────────────────────────
    # Un adjunto es inmutable una vez subido (no hay ruta de edición, solo
    # alta con id nuevo o borrado), así que puede cachearse indefinidamente
    # sin riesgo de servir contenido obsoleto. El fix anterior añadió estas
    # cabeceras de caché, pero la consulta que trae `datos` (el archivo
    # completo, hasta 2MB) se seguía ejecutando ANTES de comprobar el ETag
    # — es decir, aunque el navegador ya tuviera el adjunto en caché y la
    # respuesta final fuera un 304 sin cuerpo, la app ya había descargado
    # el archivo entero desde Supabase. Eso consume egress de base de datos
    # de forma invisible para el navegador y para cualquier métrica de
    # bytes de respuesta HTTP. Comprobamos el ETag primero, con una
    # consulta ligera que NO trae `datos`, y solo si hace falta servir el
    # contenido de verdad hacemos la consulta pesada.
    etag = f'"{aid}"'
    if request.headers.get("If-None-Match") == etag:
        existe = query("SELECT id FROM pedido_adjuntos WHERE id=%s", (aid,), one=True)
        if existe:
            return Response(status=304)
        # Si no existe (fue borrado), seguimos abajo para devolver el 404 real.

    row = query("SELECT nombre, mime_type, datos, es_correo, storage_path FROM pedido_adjuntos WHERE id=%s", (aid,), one=True)
    if not row:
        return None
    # Los correos (.eml/.msg) se sirven como attachment para que el SO
    # los abra con el gestor de correo predeterminado.
    # El resto (PDF, imagenes, Word) se sirven inline para previsualizacion.
    disposition = "attachment" if row["es_correo"] else "inline"

    # ── Adjuntos de pedidos cerrados migrados a Storage (v12.8.0) ────────
    # `datos` es NULL cuando el archivo vive en Supabase Storage en vez de
    # en la base de datos — mismo endpoint, mismo comportamiento desde el
    # punto de vista del navegador, solo cambia de dónde sale el byte.
    if row["storage_path"]:
        contenido = _storage_descargar(row["storage_path"])
        if contenido is None:
            return jsonify({"ok": False, "error": "No se pudo recuperar el adjunto desde Storage"}), 502
    else:
        contenido = bytes(row["datos"])

    resp = Response(
        contenido,
        mimetype=row["mime_type"],
        headers={"Content-Disposition": f'{disposition}; filename="{row["nombre"]}"'}
    )
    resp.headers["Cache-Control"] = "private, max-age=31536000, immutable"
    resp.headers["ETag"] = etag
    return resp


@app.route("/api/adjuntos/<int:aid>", methods=["GET"])
@login_required
def download_adjunto(aid):
    resp = _servir_adjunto_response(aid)
    if resp is None:
        return jsonify({"ok": False, "error": "Adjunto no encontrado"}), 404
    return resp


@app.route("/descargas/adjunto/<token>", methods=["GET"])
def descargar_adjunto_publico(token):
    """
    Descarga pública (SIN login) de un único adjunto de pedido, mediante
    un token temporal — ver adjunto_descarga_tokens / _obtener_o_crear_
    token_adjunto(). Pensado para el enlace que se incluye en el correo
    "ENVIADO AL PROVEEDOR" (ver enviar_emails_estado) con el PDF del
    pedido, ya que el proveedor no tiene cuenta en la app y EmailJS (plan
    Free actual) no admite adjuntar el archivo directamente al correo —
    a petición de Víctor, 2026-08-28: "se me ocurre si en vez de adjuntar
    el archivo se ponga un enlace para descargar de Supabase pulsando en
    él".

    El token da acceso ÚNICAMENTE al archivo con el que se generó — nunca
    a ningún otro adjunto ni a ninguna otra parte de la app — y deja de
    funcionar solo al caducar (por defecto 180 días desde que se generó,
    ver _obtener_o_crear_token_adjunto). No hay revocación manual: si
    hiciera falta invalidar un enlace ya enviado, basta con borrar la fila
    correspondiente de adjunto_descarga_tokens en Supabase.
    """
    fila = row_to_dict(query(
        "SELECT adjunto_id FROM adjunto_descarga_tokens WHERE token=%s AND expira_en > NOW()",
        (token,), one=True
    ))
    if not fila:
        return ("Este enlace de descarga no es válido o ha caducado. "
                "Póngase en contacto con Princess Hotels & Resorts para solicitar el documento de nuevo."), 404
    resp = _servir_adjunto_response(fila["adjunto_id"])
    if resp is None:
        return "El documento ya no está disponible.", 404
    return resp


@app.route("/api/adjuntos/<int:aid>/thumb", methods=["GET"])
@login_required
def download_adjunto_thumb(aid):
    from flask import Response
    """Sirve una miniatura reducida de un adjunto de tipo imagen_articulo.

    Si la miniatura ya está generada (subidas posteriores a Jul 2026), se
    sirve directamente. Si no existe todavía (adjuntos subidos antes de
    este fix), se genera una vez a partir de la imagen original, se guarda
    en `datos_thumb` para no repetir el trabajo, y se sirve. Si por algún
    motivo no se puede generar (formato no soportado), se cae de vuelta a
    servir la imagen original — nunca se rompe la vista para el usuario.
    """
    # Paso 1: consulta ligera — NO trae la columna `datos` (imagen original,
    # hasta 2MB). Antes se traía siempre, aunque la miniatura ya existiera,
    # lo cual generaba tráfico innecesario entre Render y Supabase en CADA
    # vista (esto cuenta como egress de base de datos, aparte del tráfico
    # HTTP hacia el navegador, y era la causa real de que el egress
    # siguiera alto pese a que las miniaturas sí se estaban sirviendo).
    row = query(
        "SELECT nombre, mime_type, datos_thumb, thumb_mime_type "
        "FROM pedido_adjuntos WHERE id=%s",
        (aid,), one=True
    )
    if not row:
        return jsonify({"ok": False, "error": "Adjunto no encontrado"}), 404

    etag = f'"{aid}-thumb"'
    if request.headers.get("If-None-Match") == etag:
        return Response(status=304)

    thumb_bytes = row["datos_thumb"]
    thumb_mime  = row["thumb_mime_type"]

    if thumb_bytes is None:
        # Backfill perezoso — imagen subida antes de existir esta columna.
        # Solo aquí (la primera vez que se pide esta imagen concreta) se
        # trae la columna `datos` completa (o se descarga de Storage si el
        # adjunto ya fue migrado — v12.8.0); en peticiones posteriores el
        # SELECT de arriba ya la evita por completo.
        original = query(
            "SELECT datos, mime_type, storage_path FROM pedido_adjuntos WHERE id=%s",
            (aid,), one=True
        )
        if original["storage_path"]:
            datos_originales = _storage_descargar(original["storage_path"])
        else:
            datos_originales = bytes(original["datos"]) if original["datos"] is not None else None

        if datos_originales is None:
            return jsonify({"ok": False, "error": "No se pudo recuperar el adjunto original"}), 502

        thumb_bytes, thumb_mime = _generar_thumbnail(datos_originales, original["mime_type"])
        if thumb_bytes is not None:
            execute(
                "UPDATE pedido_adjuntos SET datos_thumb=%s, thumb_mime_type=%s WHERE id=%s",
                (psycopg2.Binary(thumb_bytes), thumb_mime, aid)
            )
            get_db().commit()
        else:
            # No se pudo generar miniatura (formato no soportado por Pillow):
            # servimos la original tal cual para no dejar la vista rota.
            thumb_bytes, thumb_mime = datos_originales, original["mime_type"]

    resp = Response(
        bytes(thumb_bytes),
        mimetype=thumb_mime,
        headers={"Content-Disposition": f'inline; filename="{row["nombre"]}"'}
    )
    resp.headers["Cache-Control"] = "private, max-age=31536000, immutable"
    resp.headers["ETag"] = etag
    return resp


@app.route("/api/adjuntos/<int:aid>", methods=["DELETE"])
@login_required
def delete_adjunto(aid):
    db  = get_db()
    row = query("SELECT id, storage_path FROM pedido_adjuntos WHERE id=%s", (aid,), one=True)
    if not row:
        return jsonify({"ok": False, "error": "Adjunto no encontrado"}), 404
    execute("DELETE FROM pedido_adjuntos WHERE id=%s", (aid,))
    db.commit()
    # Borrar la fila primero: si el borrado en Storage falla, no queremos
    # dejar el adjunto "medio borrado" — el archivo huérfano en Storage no
    # es grave (no vuelve a ser accesible desde la app) y puede limpiarse
    # más adelante; peor sería una fila inconsistente en la BD.
    if row["storage_path"]:
        _storage_borrar(row["storage_path"])
    return jsonify({"ok": True})


# ── Ping endpoint (usado por el workflow de anti-letargo, ver .github/workflows/keep-alive-princess.yml) ──

@app.route("/ping")
def ping():
    return "OK", 200

# ── Error handlers globales (siempre JSON para rutas /api/) ───────────────────

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Ruta no encontrada"}), 404
    return send_from_directory("templates", "index.html")

@app.errorhandler(500)
def server_error(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"Error interno del servidor: {str(e)}"}), 500
    return jsonify({"ok": False, "error": str(e)}), 500

@app.errorhandler(Exception)
def unhandled_exception(e):
    import traceback
    app.logger.error("Excepción no capturada:\n" + traceback.format_exc())
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"Error inesperado: {str(e)}"}), 500
    return jsonify({"ok": False, "error": str(e)}), 500

# ── Health Monitoring — validación de integridad operativa ────────────────────
# _validar_integridad_operativa() detecta configuraciones incompletas que
# causarían fallos silenciosos en alertas, emails y Telegram:
#   · hoteles activos sin comprador asignado
#   · compradores (rol='compras') sin ningún hotel asignado
#   · usuarios activos con rol compras/hotel sin email
#   · usuarios activos con rol compras sin telegram_chat_id
#   · emails vacíos en usuarios admin
# Devuelve un dict con todos los problemas encontrados y un flag global ok=True/False.

def _validar_integridad_operativa() -> dict:
    """
    Analiza la configuración de usuarios, hoteles y compradores y detecta
    huecos que provocarían fallos silenciosos en alertas, Telegram y emails.
    Usa queries agregadas (sin bucles N+1) para evitar cuelgues con BDs lentas.
    Devuelve:
      {
        "ok": bool,
        "timestamp": "ISO-8601",
        "problemas": {
          "hoteles_sin_comprador":    [...],
          "compradores_sin_hoteles":  [...],
          "compradores_sin_telegram": [...],
          "compradores_sin_movil":    [...],
          "compradores_sin_email":    [...],
          "admins_sin_email":         [...],
          "hoteles_duplicados":       [...],   # hoteles con > 1 comprador (violación uq)
          "emailjs":                  [...],   # estado del contador/backup/cambio automático EmailJS
        },
        "resumen": {
          "total_hoteles_activos": int,
          "total_compradores":     int,
          "total_problemas":       int,
        }
      }
    """

    problemas: dict = {
        "hoteles_sin_comprador":    [],
        "compradores_sin_hoteles":  [],
        "compradores_sin_telegram": [],
        "compradores_sin_movil":    [],
        "compradores_sin_email":    [],
        "admins_sin_email":         [],
        "hoteles_duplicados":       [],
        "emailjs":                  [],
    }

    try:
        db = get_db()
        # Aplicar timeout de statement para evitar cuelgues indefinidos
        with db.cursor() as _cur:
            _cur.execute("SET LOCAL statement_timeout = '15s'")

        # ── Totales para el resumen ──────────────────────────────────────────
        total_hoteles_activos = (query(
            "SELECT COUNT(*) AS n FROM hoteles WHERE activo=1", one=True
        ) or {}).get("n", 0)

        total_compradores = (query(
            "SELECT COUNT(*) AS n FROM usuarios WHERE rol='compras' AND activo=1", one=True
        ) or {}).get("n", 0)

        # ── Hoteles activos sin ningún comprador activo asignado ─────────────
        sin_comprador = rows_to_list(query(
            """SELECT h.id AS hotel_id, h.codigo AS hotel_codigo, h.nombre AS hotel_nombre
               FROM hoteles h
               WHERE h.activo = 1
                 AND NOT EXISTS (
                     SELECT 1 FROM usuario_comprador_hoteles uch
                     JOIN usuarios u ON u.id = uch.usuario_id
                     WHERE uch.hotel_id = h.id AND u.activo = 1 AND u.rol = 'compras'
                 )
               ORDER BY h.codigo"""
        ))
        problemas["hoteles_sin_comprador"] = sin_comprador

        # ── Compradores activos sin ningún hotel asignado ────────────────────
        sin_hoteles = rows_to_list(query(
            """SELECT u.id AS usuario_id, u.username, u.nombre
               FROM usuarios u
               WHERE u.rol = 'compras' AND u.activo = 1
                 AND NOT EXISTS (
                     SELECT 1 FROM usuario_comprador_hoteles uch
                     WHERE uch.usuario_id = u.id
                 )
               ORDER BY u.nombre"""
        ))
        problemas["compradores_sin_hoteles"] = sin_hoteles

        # ── Compradores sin telegram_chat_id ─────────────────────────────────
        sin_telegram = rows_to_list(query(
            """SELECT id AS usuario_id, username, nombre
               FROM usuarios
               WHERE rol = 'compras' AND activo = 1
                 AND (telegram_chat_id IS NULL OR TRIM(telegram_chat_id) = '')
               ORDER BY nombre"""
        ))
        problemas["compradores_sin_telegram"] = sin_telegram

        # ── Compradores sin móvil ─────────────────────────────────────────────
        sin_movil = rows_to_list(query(
            """SELECT id AS usuario_id, username, nombre
               FROM usuarios
               WHERE rol = 'compras' AND activo = 1
                 AND (movil IS NULL OR TRIM(movil) = '')
               ORDER BY nombre"""
        ))
        problemas["compradores_sin_movil"] = sin_movil

        # ── Compradores sin email ─────────────────────────────────────────────
        sin_email_comp = rows_to_list(query(
            """SELECT id AS usuario_id, username, nombre
               FROM usuarios
               WHERE rol = 'compras' AND activo = 1
                 AND (email IS NULL OR TRIM(email) = '')
               ORDER BY nombre"""
        ))
        problemas["compradores_sin_email"] = sin_email_comp

        # ── Admins sin email ──────────────────────────────────────────────────
        sin_email_admin = rows_to_list(query(
            """SELECT id AS usuario_id, username, nombre
               FROM usuarios
               WHERE rol = 'admin' AND activo = 1
                 AND (email IS NULL OR TRIM(email) = '')
               ORDER BY nombre"""
        ))
        problemas["admins_sin_email"] = sin_email_admin

        # ── Hoteles con más de un comprador activo (viola uq_comprador_hotel) ─
        duplicados = rows_to_list(query(
            """SELECT h.codigo AS hotel_codigo, h.nombre AS hotel_nombre,
                      COUNT(uch.usuario_id) AS n_compradores
               FROM hoteles h
               JOIN usuario_comprador_hoteles uch ON uch.hotel_id = h.id
               JOIN usuarios u ON u.id = uch.usuario_id AND u.activo = 1 AND u.rol = 'compras'
               WHERE h.activo = 1
               GROUP BY h.id, h.codigo, h.nombre
               HAVING COUNT(uch.usuario_id) > 1
               ORDER BY h.codigo"""
        ))
        problemas["hoteles_duplicados"] = duplicados

        # ── EmailJS: contador, backup y cambio automático (v12.27.8/.10, v12.29.94: ciclo de 3 cuentas; v12.30.93: ampliado a 4) ─
        try:
            _c = get_config()
            _contador  = int(_c.get("emailjs_contador", 0) or 0)
            _umbral    = int(_c.get("emailjs_umbral_cambio", 195) or 195)
            _activa    = _emailjs_cuenta_valida(_c.get("emailjs_cuenta_activa", 1))
            _cambio_en = (_c.get("emailjs_cambio_automatico_en") or "").strip()
            _siguiente = _emailjs_siguiente_cuenta(_activa)

            def _emailjs_cuenta_completa(_n):
                return bool(
                    (_c.get(f"emailjs_public_key_{_n}") or "").strip()
                    and (_c.get(f"emailjs_service_id_{_n}") or "").strip()
                    and (_c.get(f"emailjs_template_id_{_n}") or "").strip()
                )

            _otras_cuentas = [n for n in range(1, _EMAILJS_MAX_CUENTAS + 1) if n != _activa]
            _hay_backup_completo = any(_emailjs_cuenta_completa(n) for n in _otras_cuentas)
            _siguiente_completa  = _emailjs_cuenta_completa(_siguiente)

            if _cambio_en and _contador < max(_umbral - 20, 0):
                # Cambió recientemente (contador bajo tras el reset) — informativo
                problemas["emailjs"].append({
                    "tipo": "cambio_automatico_realizado",
                    "mensaje": (f"El sistema cambió automáticamente a la cuenta EmailJS {_activa} "
                                f"el {_cambio_en} tras alcanzar el umbral de envíos ({_umbral}). "
                                f"Comprueba que la cuenta {_siguiente} (la siguiente del ciclo) siga teniendo "
                                f"cuota disponible para el próximo cambio.")
                })
            elif _contador >= _umbral and not _hay_backup_completo:
                problemas["emailjs"].append({
                    "tipo": "umbral_alcanzado_sin_backup",
                    "mensaje": (f"Van {_contador} envíos contabilizados en la cuenta {_activa} (umbral: {_umbral}) "
                                f"y ninguna de las otras cuentas ({', '.join(str(n) for n in _otras_cuentas)}) tiene "
                                f"las 3 credenciales completas — el cambio automático no se pudo realizar. "
                                f"Rellénalas en Admin → Parámetros de alertas → EmailJS cuanto antes.")
                })
            elif not _siguiente_completa and _contador >= max(_umbral - 20, 0):
                problemas["emailjs"].append({
                    "tipo": "cerca_del_umbral_sin_backup",
                    "mensaje": (f"Van {_contador} envíos de {_umbral} en la cuenta {_activa} antes del cambio "
                                f"automático, y la cuenta {_siguiente} (siguiente del ciclo) todavía no está "
                                f"configurada." + (f" La cuenta {[n for n in _otras_cuentas if _emailjs_cuenta_completa(n)][0]} sí lo está y se usaría como backup." if _hay_backup_completo else ""))
                })
        except Exception as _exc_ej:
            log.error("[INTEGRIDAD] Error comprobando estado EmailJS: %s", _exc_ej)

    except Exception as exc:
        log.error("[INTEGRIDAD] Error validando integridad: %s", exc)
        return {
            "ok": False,
            "timestamp": datetime.utcnow().isoformat(),
            "error": str(exc),
            "problemas": problemas,
            "resumen": {"total_hoteles_activos": 0, "total_compradores": 0, "total_problemas": -1},
        }

    total_problemas = sum(len(v) for v in problemas.values())
    return {
        "ok": total_problemas == 0,
        "timestamp": datetime.utcnow().isoformat(),
        "problemas": problemas,
        "resumen": {
            "total_hoteles_activos": int(total_hoteles_activos),
            "total_compradores":     int(total_compradores),
            "total_problemas":       total_problemas,
        },
    }


# ── Alerta de consumo a admins por Telegram (Jul 2026) ───────────────────────
# Egress y tamaño de BD se avisan juntos, en un único mensaje/popup, para no
# duplicar avisos separados sobre el mismo tema (cuota de Supabase). Ambos
# comparten el mismo umbral de aviso (50%) y el mismo destinatario
# configurado en Config Avisos, evento "egress_alerta".
EGRESS_LIMITE_GB          = 5.0   # cuota del plan Free de Supabase (egress sin caché)
EGRESS_UMBRAL_AVISO_PCT   = 50    # avisar al llegar a este % del límite
EGRESS_CICLO_DIA_INICIO   = 23    # el ciclo de facturación de Supabase renueva el día 23 de cada mes

DB_SIZE_LIMITE_MB         = 512   # cuota del plan Free de Supabase (tamaño de BD)
DB_SIZE_UMBRAL_AVISO_PCT  = 50    # mismo umbral que egress, por consistencia


def _egress_ciclo_actual_inicio() -> "_date":
    """Fecha de inicio del ciclo de facturación actual de Supabase (día 23)."""
    hoy = _date.today()
    if hoy.day >= EGRESS_CICLO_DIA_INICIO:
        return hoy.replace(day=EGRESS_CICLO_DIA_INICIO)
    # Aún no llegamos al día 23 de este mes → el ciclo empezó el 23 del mes anterior
    mes_ant = hoy.month - 1 or 12
    anio_ant = hoy.year - 1 if hoy.month == 1 else hoy.year
    return hoy.replace(year=anio_ant, month=mes_ant, day=EGRESS_CICLO_DIA_INICIO)


def _egress_bytes_ciclo_actual() -> int:
    inicio = _egress_ciclo_actual_inicio()
    row = query(
        "SELECT COALESCE(SUM(bytes),0) as total FROM egress_tracking WHERE fecha >= %s",
        (inicio,), one=True
    )
    return int(row["total"]) if row else 0


def _db_size_bytes_actual() -> int:
    """Tamaño real de la BD ahora mismo (en vivo, no del último snapshot
    diario) — coherente con cómo egress también se recalcula en vivo en
    cada ejecución del job, en vez de depender de un valor cacheado."""
    row = query("SELECT pg_database_size(current_database()) AS bytes_total", one=True)
    return int(row["bytes_total"]) if row else 0


def _ya_alertado_consumo_hoy() -> bool:
    row = query(
        """SELECT COUNT(*) as n FROM whatsapp_log
           WHERE pedido_id IS NULL AND tipo='consumo_alerta'
             AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                 (NOW() AT TIME ZONE 'Atlantic/Canary')::date""",
        one=True
    )
    return (row["n"] if row else 0) > 0


def _job_alerta_consumo(force: bool = False):
    """
    Job diario (08:30 hora Canarias): un único mensaje combinado con egress
    Y tamaño de base de datos — dos métricas distintas pero ambas cuotas de
    Supabase, así que un solo aviso evita duplicar ruido sobre el mismo
    tema. Avisa por Telegram + popup bridge a los admins si CUALQUIERA de
    las dos se acerca o supera su límite del plan Free. Máximo un aviso al
    día. Si force=True, avisa siempre (para probar el canal desde el botón
    de admin), aunque ninguna haya superado el umbral.

    Nota (Jul 2026): al ejecutarse a primera hora, egress refleja lo
    acumulado HASTA AYER (tabla egress_tracking, por día) — si el umbral
    se cruza a media tarde, el aviso no llega hasta la mañana siguiente.
    Tamaño de BD sí se consulta en vivo en el momento del job (no depende
    del snapshot diario de las 08:10), así que esa parte del mensaje es
    siempre el dato actual real.
    """
    with app.app_context():
        _job_alerta_consumo_inner(force)
        _flush_egress_bytes()


def _job_alerta_consumo_inner(force: bool = False):
    log.info("▶ [CONSUMO] Inicio job alerta de egress + tamaño BD — %s", _date.today())
    if not force and _ya_alertado_consumo_hoy():
        log.info("[CONSUMO] Ya se avisó hoy — se omite.")
        return

    # Egress (acumulado del ciclo de facturación, por día)
    egress_bytes = _egress_bytes_ciclo_actual()
    egress_gb = egress_bytes / (1024 ** 3)
    egress_pct = (egress_gb / EGRESS_LIMITE_GB * 100) if EGRESS_LIMITE_GB else 0
    inicio_ciclo = _egress_ciclo_actual_inicio()

    # Tamaño de BD (en vivo, ahora mismo)
    db_bytes = _db_size_bytes_actual()
    db_mb = db_bytes / (1024 ** 2)
    db_pct = (db_mb / DB_SIZE_LIMITE_MB * 100) if DB_SIZE_LIMITE_MB else 0

    egress_supera = egress_pct >= EGRESS_UMBRAL_AVISO_PCT
    db_supera     = db_pct >= DB_SIZE_UMBRAL_AVISO_PCT

    if not force and not egress_supera and not db_supera:
        log.info("[CONSUMO] Egress %.2f GB (%.0f%%), BD %.1f MB (%.0f%%) — ambos por debajo del umbral, no se avisa.",
                  egress_gb, egress_pct, db_mb, db_pct)
        return

    admins = _destinatarios_evento("egress_alerta", "telegram")
    if not admins:
        log.warning("[CONSUMO] Umbral superado pero no hay destinatarios configurados en Configuración de Avisos.")
        return

    pct_max = max(egress_pct, db_pct)
    estado = "🔴 *LÍMITE SUPERADO*" if pct_max >= 100 else "🟠 *Acercándose al límite*"
    marca_egress = " ⚠️" if egress_supera else ""
    marca_db     = " ⚠️" if db_supera else ""
    texto = (
        f"{estado} — Consumo Supabase (control-pedidos-princess)\n\n"
        f"📶 Egress{marca_egress}: estimación desde el {inicio_ciclo.strftime('%d/%m/%Y')}: "
        f"*{egress_gb:.2f} GB* de {EGRESS_LIMITE_GB:.0f} GB ({egress_pct:.0f}%)\n"
        f"🗄️ Tamaño BD{marca_db}: *{db_mb:.0f} MB* de {DB_SIZE_LIMITE_MB} MB ({db_pct:.0f}%)\n\n"
        f"_Nota: el egress es una estimación basada en lo que sirve la app; "
        f"puede no coincidir exactamente con el contador de Supabase, que "
        f"también incluye tráfico interno del proyecto. El tamaño de BD sí "
        f"es el dato real, en vivo._\n\n"
        f"Revisa Supabase → Organization → Usage para el dato oficial de egress. "
        f"El ciclo de egress renueva el día {EGRESS_CICLO_DIA_INICIO}."
    )
    titulo_bridge = ("🔴 [Consumo] Límite superado — Supabase" if pct_max >= 100
                      else "🟠 [Consumo] Acercándose al límite — Supabase")
    for adm in admins:
        username = adm.get("username", "admin")
        chat_id  = adm.get("telegram_chat_id")
        if chat_id:
            res = _send_telegram(chat_id, texto)
            log.info("[CONSUMO] -> %s: %s", username, res)
        _encolar_bridge_notificacion(
            usuario=username,
            tipo="consumo",
            titulo=titulo_bridge,
            mensaje=texto.replace("*", ""),
            nivel="urgente" if pct_max >= 100 else "aviso",
            pedido_id=None,
        )

    # Registrar el aviso para no repetirlo hoy (reutiliza whatsapp_log,
    # igual que el resto de dedupes de notificaciones de la app).
    try:
        execute(
            "INSERT INTO whatsapp_log (pedido_id, tipo, destinatario, mensaje, enviado) "
            "VALUES (NULL, 'consumo_alerta', 'admins', %s, 1)",
            (texto,)
        )
        get_db().commit()
    except Exception as e:
        log.error("[CONSUMO] Error registrando dedupe: %s", e)


def _job_db_size_tracking() -> None:
    """
    Job diario (08:00 hora Canarias, junto al resto de jobs matutinos):
    snapshot del tamaño real de la base de datos (pg_database_size) y de
    pedido_adjuntos en concreto (con diferencia la mayor consumidora, al
    guardar los archivos como bytea/TOAST) — visible en Admin → Integridad.

    A diferencia del egress, aquí no hay "caché" que compense: el tamaño
    de la BD solo crece. Este tracking es puramente informativo por ahora
    (sin alerta automática todavía) — sirve para ver la tendencia sin
    tener que entrar al dashboard de Supabase cada vez.
    """
    with app.app_context():
        try:
            fila = query("""
                SELECT
                    pg_database_size(current_database()) AS bytes_total,
                    COALESCE(pg_total_relation_size('pedido_adjuntos'), 0) AS bytes_adjuntos
            """, one=True)
            execute("""
                INSERT INTO db_size_tracking (fecha, bytes_total, bytes_adjuntos)
                VALUES ((NOW() AT TIME ZONE 'Atlantic/Canary')::date, %s, %s)
                ON CONFLICT (fecha) DO UPDATE
                SET bytes_total = EXCLUDED.bytes_total,
                    bytes_adjuntos = EXCLUDED.bytes_adjuntos
            """, (fila["bytes_total"], fila["bytes_adjuntos"]))
            get_db().commit()
            log.info("[DB-SIZE] Snapshot: %.1f MB totales (%.1f MB en adjuntos)",
                      fila["bytes_total"] / 1024 / 1024, fila["bytes_adjuntos"] / 1024 / 1024)
        except Exception as e:
            get_db().rollback()
            log.error("[DB-SIZE] Error registrando snapshot: %s", e)
        _flush_egress_bytes()


def _job_migrar_adjuntos_storage(force: bool = False, limite: int = 50) -> dict:
    """
    Job diario (03:00 hora Canarias — horario de madrugada, sin prisa, sin
    interferir con nada): migra a Supabase Storage los adjuntos de pedidos
    ya cerrados (ENTREGADO/CANCELADO) que todavía viven en la base de datos.

    Hasta `limite` adjuntos por ejecución, pero **de uno en uno** (una fila
    por SELECT, no un fetchall() de todo el lote): así la memoria en uso es
    la de un único adjunto (hasta MAX_ADJUNTO_BYTES) en vez de la suma de
    hasta `limite` adjuntos retenidos a la vez, que en el peor caso teórico
    (varios adjuntos grandes seguidos) podía suponer varios cientos de MB
    de golpe. Cada fila se marca migrada (`storage_path` + `datos=NULL`)
    inmediatamente tras subirse, así que si el job se interrumpe a mitad,
    lo ya migrado no se repite en la siguiente ejecución — retoma donde lo
    dejó. Las filas que fallan en esta misma ejecución se excluyen de las
    siguientes vueltas del bucle (si no, al no quedar marcadas, volverían a
    salir en el siguiente SELECT y el bucle no avanzaría nunca).

    Devuelve un resumen (dict) — útil tanto para el log del job automático
    como para la respuesta del endpoint de disparo manual desde Admin.
    """
    resumen = {"migrados": 0, "fallidos": 0, "omitidos_sin_storage": False}

    if not STORAGE_CONFIGURADO:
        log.warning("[STORAGE-MIGRA] SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY no configurados — se omite.")
        resumen["omitidos_sin_storage"] = True
        return resumen

    fallidos_ids = []

    for _ in range(limite):
        adj = query("""
            SELECT pa.id, pa.pedido_id, pa.nombre, pa.mime_type, pa.datos
            FROM pedido_adjuntos pa
            JOIN pedidos p ON p.id = pa.pedido_id
            WHERE p.estado = ANY(%s)
              AND pa.storage_path IS NULL
              AND pa.datos IS NOT NULL
              AND pa.id != ALL(%s)
            ORDER BY pa.id
            LIMIT 1
        """, (list(ESTADOS_CERRADOS), fallidos_ids or [0]), one=True)

        if not adj:
            break

        path = f"pedidos/{adj['pedido_id']}/{adj['id']}_{_slugify_nombre_archivo(adj['nombre'])}"
        ok = _storage_subir(path, bytes(adj["datos"]), adj["mime_type"])
        if ok:
            try:
                execute(
                    "UPDATE pedido_adjuntos SET storage_path=%s, datos=NULL WHERE id=%s",
                    (path, adj["id"])
                )
                get_db().commit()
                resumen["migrados"] += 1
            except Exception as e:
                get_db().rollback()
                log.error("[STORAGE-MIGRA] Subido pero fallo al actualizar fila id=%s: %s", adj["id"], e)
                resumen["fallidos"] += 1
                fallidos_ids.append(adj["id"])
        else:
            resumen["fallidos"] += 1
            fallidos_ids.append(adj["id"])

    if resumen["migrados"] == 0 and resumen["fallidos"] == 0:
        log.info("[STORAGE-MIGRA] Nada pendiente de migrar.")
        return resumen

    log.info("[STORAGE-MIGRA] %d migrado(s), %d fallido(s) (lote de hasta %d).",
              resumen["migrados"], resumen["fallidos"], limite)
    return resumen


def _slugify_nombre_archivo(nombre: str) -> str:
    """Nombre de archivo seguro para usar como parte de una ruta de Storage
    (sin espacios ni caracteres que puedan dar problemas en una URL)."""
    import re
    base = re.sub(r"[^a-zA-Z0-9._-]+", "_", nombre or "archivo")
    return base[:150]  # límite razonable de longitud


def _vacuum_full_adjuntos() -> dict:
    """
    VACUUM FULL sobre pedido_adjuntos — solo se llama cuando esa noche se
    migró al menos un adjunto de verdad (ver _job_migrar_adjuntos_storage_diario).
    Poner `datos=NULL` libera el espacio LÓGICAMENTE, pero Postgres no
    encoge el archivo físico en disco por sí solo; sin este VACUUM FULL el
    tamaño reportado de la tabla no bajaría nunca, aunque el conteo de
    "migrados" siguiera subiendo cada noche.

    VACUUM FULL no puede ejecutarse dentro de una transacción normal, así
    que usa una conexión propia con autocommit — no la del pool compartido
    (para no dejarla en un estado raro si algo falla a mitad). También
    toma un ACCESS EXCLUSIVE lock sobre la tabla mientras dura — por eso
    solo se dispara a las 03:00, en horario de mínimo tráfico.
    """
    resultado = {"mb_antes": None, "mb_despues": None, "mb_liberados": None, "error": None}
    conn = None
    try:
        mb_antes = query(
            "SELECT pg_total_relation_size('pedido_adjuntos') / 1024.0 / 1024.0 AS mb", one=True
        )["mb"]
        resultado["mb_antes"] = float(mb_antes)

        conn = psycopg2.connect(DATABASE_URL, application_name="control_pedidos_vacuum")
        conn.autocommit = True
        with conn.cursor() as cur:
            cur.execute("VACUUM FULL pedido_adjuntos")
        conn.close()
        conn = None

        mb_despues = query(
            "SELECT pg_total_relation_size('pedido_adjuntos') / 1024.0 / 1024.0 AS mb", one=True
        )["mb"]
        resultado["mb_despues"] = float(mb_despues)
        resultado["mb_liberados"] = round(resultado["mb_antes"] - resultado["mb_despues"], 1)

        execute(
            """INSERT INTO db_vacuum_log (fecha, mb_antes, mb_despues, mb_liberados)
               VALUES (NOW(), %s, %s, %s)""",
            (resultado["mb_antes"], resultado["mb_despues"], resultado["mb_liberados"])
        )
        get_db().commit()

        log.info("[VACUUM] pedido_adjuntos: %.1f MB -> %.1f MB (%.1f MB liberados)",
                  resultado["mb_antes"], resultado["mb_despues"], resultado["mb_liberados"])
    except Exception as e:
        resultado["error"] = str(e)
        log.error("[VACUUM] Error compactando pedido_adjuntos: %s", e)
        if conn:
            try:
                conn.close()
            except Exception:
                pass
    return resultado


def _job_migrar_adjuntos_storage_diario() -> None:
    with app.app_context():
        resumen = {"migrados": 0}
        try:
            resumen = _job_migrar_adjuntos_storage()
        except Exception as e:
            log.error("[STORAGE-MIGRA] Error en job diario: %s", e)
        _flush_egress_bytes()

        # Compactar solo si esta noche se migró algo de verdad — si no,
        # el VACUUM FULL no liberaría nada y solo bloquearía la tabla sin
        # motivo (ver docstring de _vacuum_full_adjuntos).
        if resumen.get("migrados", 0) > 0:
            _vacuum_full_adjuntos()
        else:
            log.info("[VACUUM] Sin adjuntos migrados esta noche — se omite la compactación.")


def _job_recordar_emails_sistema_pendientes():
    """
    v12.11.0: desde que el email de Fase 2 de solicitudes de acceso se
    despacha desde el navegador de un admin (emails_sistema_pendientes) en
    vez del navegador de quien solicita el acceso, existe la posibilidad de
    que se quede en cola simplemente porque ningún admin ha abierto la app
    todavía. Este job avisa por Telegram con un recordatorio explícito,
    sin repetirlo cada minuto (solo si no se recordó en los últimos 30').
    Corre cada 10 minutos, 07:00–21:00 — no tiene sentido despertar a nadie
    de madrugada por esto, ya lo verán al abrir la app en horario normal.

    (2026-08-20) Excluye las filas descartadas a mano (`descartado_en`) y
    las que ya agotaron sus reintentos (`intentos >= MAX_INTENTOS_EMAIL_SISTEMA`,
    ver /api/admin/emails-sistema-atascados): antes esta consulta las seguía
    contando como "pendientes" para siempre, así que abrir la app y descartar
    una fila NO hacía que este recordatorio dejara de avisar por ella —
    reportado por Víctor, le llegó este aviso justo después de descartar los
    4 correos atascados desde el panel de admin. Ninguna de las dos deja de
    reportarse solo con "abrir la aplicación" (una está descartada a
    propósito, la otra ya no se reintenta sola), así que tampoco tiene
    sentido seguir avisando con ese texto para ellas.
    """
    with app.app_context():
        try:
            pendientes = query("""
                SELECT id, evento_codigo, creado_en
                FROM emails_sistema_pendientes
                WHERE enviado = FALSE
                  AND descartado_en IS NULL
                  AND intentos < %s
                  AND creado_en < NOW() - INTERVAL '10 minutes'
                  AND (recordado_en IS NULL OR recordado_en < NOW() - INTERVAL '30 minutes')
            """, (MAX_INTENTOS_EMAIL_SISTEMA,))
            if not pendientes:
                return
            n = len(pendientes)
            eventos = ", ".join(sorted({p["evento_codigo"] for p in pendientes}))
            plural = "es" if n != 1 else ""
            # (2026-08-20) Antes se usaba _notify_solicitud_telegram(), que lleva
            # el t\u00edtulo de popup "\ud83d\udccb Nueva solicitud de acceso" fijo \u2014 confuso
            # para este aviso, que no tiene nada que ver con solicitudes de
            # acceso (reportado por V\u00edctor: le lleg\u00f3 ese t\u00edtulo justo tras
            # descartar unos correos atascados, sin relaci\u00f3n aparente con el
            # texto del aviso). Se llama a _notificar_evento() directamente,
            # con el mismo evento_codigo "solicitud_acceso" (mismos
            # destinatarios configurados, sin cambios) pero un t\u00edtulo de
            # popup correcto para este caso.
            _notificar_evento(
                "solicitud_acceso",
                f"\u23F0 *Recordatorio*\n\n"
                f"Hay *{n}* email{plural} de sistema en cola sin enviar "
                f"({eventos}), esperando a que alguien abra la aplicaci\u00f3n "
                f"para despacharlos autom\u00e1ticamente.\n\n"
                f"\U0001F449 Abre Control de Pedidos para completarlo.",
                titulo_bridge="\u23F0 Correos de sistema en cola",
                nivel_bridge="aviso",
                tipo_bridge="solicitud_acceso",
            )
            ids = [p["id"] for p in pendientes]
            execute(
                "UPDATE emails_sistema_pendientes SET recordado_en = NOW() "
                "WHERE id = ANY(%s)",
                (ids,),
            )
            get_db().commit()
        except Exception as exc:
            log.error("[RECORDATORIO EMAILS SISTEMA] Error: %s", exc)


def _job_purgar_emails_sistema_descartados():
    """
    (2026-08-31) Víctor, sobre el panel "Cola de correos de sistema
    pendientes" (admin → EmailJS y Cola de Correo): "esto, una vez
    descartado no tiene sentido seguir llenado la pantalla, podemos poner
    otro botón para reactivar y que a los 2 días cauque y se elimine el
    envío descartado". Antes, una fila descartada a mano (descartado_en,
    ver api_descartar_email_sistema) se quedaba en la tabla para siempre
    como constancia — con el tiempo, se iba acumulando sin ningún valor
    real y ensuciando el panel.

    Job diario (no hace falta más frecuencia — 2 días de margen de sobra
    para reactivar a mano si el descarte fue un error, ver
    api_reactivar_email_sistema): borra las filas descartadas hace más de
    2 días. Solo toca filas ya descartadas Y no enviadas — nunca una fila
    real de correo enviado.
    """
    with app.app_context():
        try:
            cur = execute(
                "DELETE FROM emails_sistema_pendientes "
                "WHERE enviado = FALSE AND descartado_en IS NOT NULL "
                "  AND descartado_en < NOW() - INTERVAL '2 days'"
            )
            n = cur.rowcount
            get_db().commit()
            if n:
                log.info("[EMAILS-SISTEMA] Purgadas %d fila(s) descartadas hace más de 2 días.", n)
        except Exception as exc:
            log.error("[EMAILS-SISTEMA] Error purgando descartados: %s", exc)


def _job_avanzar_reinicio_emailjs():
    """
    (2026-09-01) Job diario: las 4 fechas `emailjs_reinicio_fecha_N` son
    puramente informativas — el admin las copiaba a mano desde el panel de
    cada cuenta en EmailJS.com para saber cuándo recupera su cupo mensual,
    y ningún otro código las lee (el cambio real de cuenta activa depende
    solo del contador de envíos, ver /api/emailjs/registrar-envio).

    Para no tener que entrar a mirar las 4 cuentas cada mes: en cuanto la
    fecha guardada de una cuenta ya ha pasado (hoy > fecha), se avanza ella
    sola +30 días (el ciclo gratuito de EmailJS es rolling de 30 días desde
    el último reinicio, no mes de calendario — por eso +30 días y no "+1
    mes", que además tropezaría con meses de distinta duración). Si el
    servidor ha estado parado más de un ciclo, avanza de 30 en 30 hasta que
    la fecha vuelva a caer en el futuro, para no arrastrar un desfase.

    Puramente informativo: si en algún momento el ciclo real de una cuenta
    en EmailJS.com se desvía (p.ej. tras un cambio manual de plan), basta
    con corregir la fecha a mano una vez en Admin y el job sigue avanzándola
    sola desde ahí.

    (2026-09-01, v12.30.93) Ampliado a la 4ª cuenta añadida en esa misma
    versión — sin más cambios que iterar también n=4, el resto de la
    lógica (parseo, +30 días, tope de reintentos) ya era genérica.
    """
    with app.app_context():
        try:
            db  = get_db()
            cur = db.cursor()
            c   = get_config()
            hoy = _hoy_canarias()
            for n in (1, 2, 3, 4):
                clave = f"emailjs_reinicio_fecha_{n}"
                valor = (c.get(clave) or "").strip()
                if not valor:
                    continue
                try:
                    fecha = datetime.strptime(valor, "%Y-%m-%d").date()
                except ValueError:
                    log.warning("[EMAILJS] %s con formato de fecha inválido: %r — se ignora", clave, valor)
                    continue
                original = fecha
                while fecha <= hoy:
                    fecha += timedelta(days=30)
                if fecha != original:
                    nueva = fecha.isoformat()
                    cur.execute("UPDATE config_alertas SET valor=%s WHERE clave=%s", (nueva, clave))
                    log.info("[EMAILJS] %s vencida (%s) — avanzada a %s (+30 días)", clave, original.isoformat(), nueva)
            db.commit()
        except Exception as exc:
            log.error("[EMAILJS] Error avanzando fechas de reinicio de cupo: %s", exc)


def _job_health_check(force: bool = False):
    """
    Job diario (07:05 hora Canarias): valida integridad operativa y envía
    Telegram al administrador si detecta problemas de configuración.
    Si force=True (llamada manual desde el botón admin), envía siempre
    aunque no haya problemas, para confirmar que el canal funciona.
    Nunca bloquea operaciones — solo alerta.
    """
    with app.app_context():
        _job_health_check_inner(force)
        _flush_egress_bytes()

def _job_health_check_inner(force: bool = False):
    log.info("▶ [HEALTH] Inicio job integridad operativa — %s", _date.today())
    resultado = _validar_integridad_operativa()

    # ── Destinatarios: configurados para el evento 'health_check' en
    #    Administrador → Configuración de Avisos ───────────────────────────
    admins_bd = _destinatarios_evento("health_check", "telegram")

    def _enviar_a_admins(texto_msg, titulo_bridge="🚨 [Integridad] Control Pedidos", nivel_bridge="urgente"):
        for adm in admins_bd:
            username = adm.get("username", "admin")
            chat_id  = adm.get("telegram_chat_id")
            if chat_id:
                res = _send_telegram(chat_id, texto_msg)
                log.info("[HEALTH] Telegram → %s (%s): %s",
                         username, chat_id,
                         "OK" if res.get("ok") else res.get("error"))
            else:
                log.warning("[HEALTH] %s configurado para Telegram pero sin telegram_chat_id", username)
            # ── Encolar en bridge agenda para este admin ─────────────────────
            # ANTI-REGRESION (bug 2026-07-14): este aviso es EXCLUSIVO de admin
            # (no tiene contrapartida de comprador) y hasta ahora solo se enviaba
            # por Telegram — nunca se encolaba para main_agenda, así que el
            # popup no aparecía nunca en el Organizador Princess del admin
            # aunque el mensaje sí llegara a su Telegram.
            _encolar_bridge_notificacion(
                usuario=username,
                tipo="integridad",
                titulo=titulo_bridge,
                mensaje=texto_msg.replace("*", ""),
                nivel=nivel_bridge,
                pedido_id=None,
            )

    if resultado.get("ok"):
        log.info("✅ [HEALTH] Integridad OK — sin problemas detectados")
        if force and admins_bd:
            _enviar_a_admins(
                "✅ *Control Pedidos — Integridad OK*\n\n"
                f"Sistema en buen estado — ningún problema detectado.\n"
                f"🏨 Hoteles activos: {resultado['resumen']['total_hoteles_activos']}\n"
                f"🛒 Compradores activos: {resultado['resumen']['total_compradores']}",
                titulo_bridge="✅ [Integridad] Control Pedidos — OK",
                nivel_bridge="aviso",
            )
        elif force:
            log.warning("[HEALTH] Sin admins con Telegram configurado — no se envió confirmación")
        return

    # Construir mensaje de alerta
    probs = resultado["problemas"]
    lineas = ["🚨 *ALERTA DE CONFIGURACIÓN — Control Pedidos*", ""]

    if probs["hoteles_sin_comprador"]:
        lineas.append(f"❌ *Hoteles sin comprador ({len(probs['hoteles_sin_comprador'])})* — CRÍTICO:")
        for h in probs["hoteles_sin_comprador"]:
            lineas.append(f"  · {h['hotel_codigo']} — {h['hotel_nombre']}")
        lineas.append("")

    if probs["compradores_sin_hoteles"]:
        lineas.append(f"⚠️ *Compradores sin hoteles ({len(probs['compradores_sin_hoteles'])})* :")
        for u in probs["compradores_sin_hoteles"]:
            lineas.append(f"  · {u['nombre']} ({u['username']})")
        lineas.append("")

    if probs["compradores_sin_telegram"]:
        lineas.append(f"⚠️ *Compradores sin Telegram ({len(probs['compradores_sin_telegram'])})* :")
        for u in probs["compradores_sin_telegram"]:
            lineas.append(f"  · {u['nombre']} ({u['username']})")
        lineas.append("")

    if probs["compradores_sin_email"]:
        lineas.append(f"⚠️ *Compradores sin email ({len(probs['compradores_sin_email'])})* :")
        for u in probs["compradores_sin_email"]:
            lineas.append(f"  · {u['nombre']} ({u['username']})")
        lineas.append("")

    if probs["admins_sin_email"]:
        lineas.append(f"⚠️ *Admins sin email ({len(probs['admins_sin_email'])})* :")
        for u in probs["admins_sin_email"]:
            lineas.append(f"  · {u['nombre']} ({u['username']})")
        lineas.append("")

    lineas.append(f"📋 Total problemas: *{resultado['resumen']['total_problemas']}*")
    lineas.append("— Accede al panel admin → Integridad para ver el detalle.")

    texto = "\n".join(lineas)

    # Enviar a todos los admins con Telegram configurado
    if admins_bd:
        _enviar_a_admins(texto)
    else:
        log.warning("[HEALTH] Sin admins con Telegram configurado — alerta solo en log")

    log.warning("[HEALTH] %d problema(s) de integridad detectados: %s",
                resultado["resumen"]["total_problemas"],
                {k: len(v) for k, v in probs.items() if v})


# ── Scheduler: alertas automáticas por Telegram ───────────────────────────────
# Corre dentro del mismo proceso gunicorn — sin Redis, sin Celery, sin workers.
# Cada 60 segundos, en horario 07:00-16:00 hora Canarias (todos los días),
# revisa todos los pedidos activos y envía Telegram si procede.
# La protección _ya_notificado_hoy() evita duplicados: aunque el job corra
# 540 veces al día, cada pedido solo recibe UNA alerta por día.

def _iniciar_scheduler():
    scheduler = BackgroundScheduler(timezone="Atlantic/Canary")
    # Intervalo: cada 60 segundos, solo entre las 07:00 y las 16:00 locales.
    # hour='7-15' → APScheduler ejecuta mientras hour esté en [7..15],
    # es decir desde las 07:00:00 hasta las 15:59:59 — el último ciclo
    # arranca a las 15:59 y el siguiente ya sería las 16:00, fuera de rango.
    scheduler.add_job(
        _job_alertas_diarias,
        trigger="cron",
        day_of_week="mon-fri",  # (2026-08-02) fin de semana: se retrasa a lunes,
                                 # ver guardián en _job_alertas_diarias_inner()
        hour="7-15",          # 07:00 → 15:59 (inclusive)
        minute="*",
        second="0",           # en punto de cada minuto
        id="alertas_cada_minuto",
        replace_existing=True,
        misfire_grace_time=60,
    )
    # Job de techo URGENTE a admins: cada 60 segundos, lun-vie, 07:00-16:59.
    # La lógica interna aplica deduplicación diaria y el ciclo de 2 días.
    scheduler.add_job(
        _job_techo_urgente_admins,
        trigger="cron",
        day_of_week="mon-fri",  # solo días laborables
        hour="7-16",            # 07:00 → 16:59 (función interna bloquea ≥ 17:00)
        minute="*",
        second="0",
        id="techo_urgente_admins",
        replace_existing=True,
        misfire_grace_time=60,
    )
    # Job de techo mensual: una vez al día a las 08:00 hora Canarias, lun-vie
    # (2026-08-02: alineado con techo urgente y familia repetida — no tiene
    # sentido notificar semáforo mensual en fin de semana, se retoma el lunes)
    scheduler.add_job(
        _job_alertas_techo_mensual,
        trigger="cron",
        day_of_week="mon-fri",
        hour="8",
        minute="0",
        second="0",
        id="alertas_techo_mensual",
        replace_existing=True,
        misfire_grace_time=3600,  # 1 hora — tolera reinicios de Render tras el cron de 08:00
    )
    # Job de familia/partida repetida: cada 60s, lun-vie 07:00-16:59.
    # Comprador: alerta diaria. Admins: alerta cada 2 días.
    scheduler.add_job(
        _job_familia_repetida,
        trigger="cron",
        day_of_week="mon-fri",
        hour="7-16",
        minute="*",
        second="0",
        id="familia_repetida",
        replace_existing=True,
        misfire_grace_time=60,
    )
    # Job de integridad: una vez al día a las 07:05 hora Canarias, lun-vie
    # (2026-08-17) fin de semana: envía Telegram + popup bridge a los
    # admins si detecta problemas — no tiene sentido avisar en fin de
    # semana de algo que puede esperar a que alguien esté trabajando; se
    # retoma el lunes con normalidad, mismo criterio que el resto de jobs
    # de alertas de esta lista.
    scheduler.add_job(
        _job_health_check,
        trigger="cron",
        day_of_week="mon-fri",
        hour="7",
        minute="5",
        second="0",
        id="health_check_diario",
        replace_existing=True,
        misfire_grace_time=300,
    )
    # Snapshot diario de tamaño de BD: a las 08:10 — se mantiene
    # independiente de la alerta (sirve también para el histórico de
    # tendencia en Admin → Integridad, no solo para avisar).
    scheduler.add_job(
        _job_db_size_tracking,
        trigger="cron",
        hour="8",
        minute="10",
        second="0",
        id="db_size_tracking_diario",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    # Migración de adjuntos de pedidos cerrados a Supabase Storage: a las
    # 03:00 (madrugada, sin prisa, sin interferir con nada). Por lotes de
    # 50 — si hay más pendientes, los coge al día siguiente.
    scheduler.add_job(
        _job_migrar_adjuntos_storage_diario,
        trigger="cron",
        hour="3",
        minute="0",
        second="0",
        id="migrar_adjuntos_storage_diario",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    # Job de alerta combinada de consumo Supabase (egress + tamaño BD): una
    # vez al día a las 08:30 hora Canarias, lun-vie, tras el snapshot de
    # tamaño de BD de las 08:10 — un único mensaje/popup si CUALQUIERA de
    # las dos cuotas se acerca o supera el límite del plan Free, en vez de
    # dos avisos separados sobre el mismo tema.
    #
    # Nota: egress usa el acumulado por día (tabla egress_tracking) y
    # refleja lo consumido HASTA AYER — si el umbral se cruza a media
    # tarde, el aviso no llega hasta la mañana siguiente. Tamaño de BD sí
    # se consulta en vivo en el momento del job.
    #
    # (2026-08-17) fin de semana: se retrasa a lunes, igual que el resto
    # de alertas por Telegram/popup de esta lista — el snapshot de tamaño
    # de BD (job aparte, arriba) sigue corriendo todos los días, así que
    # no se pierde histórico, solo se retrasa el AVISO si el umbral se
    # cruza durante el fin de semana.
    scheduler.add_job(
        _job_alerta_consumo,
        trigger="cron",
        day_of_week="mon-fri",
        hour="8",
        minute="30",
        second="0",
        id="alerta_consumo_diaria",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    # Recordatorio de emails de sistema en cola (incluye Fase 2 de
    # solicitudes de acceso): cada 10 min, 07:00-21:00, lun-vie — fuera de
    # ese rango (incluido el fin de semana, 2026-08-17) puede esperar a
    # que alguien vuelva a abrir la aplicación en horario laborable.
    scheduler.add_job(
        _job_recordar_emails_sistema_pendientes,
        trigger="cron",
        day_of_week="mon-fri",
        hour="7-21",
        minute="*/10",
        second="0",
        id="recordar_emails_sistema_pendientes",
        replace_existing=True,
        misfire_grace_time=120,
    )
    # Purga de correos de sistema descartados hace más de 2 días — diaria,
    # de madrugada (no es urgente, solo limpieza del panel de admin).
    scheduler.add_job(
        _job_purgar_emails_sistema_descartados,
        trigger="cron",
        hour="4",
        minute="0",
        second="0",
        id="purgar_emails_sistema_descartados",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    # Avance automático de las fechas de "reinicia cupo" de las 4 cuentas
    # EmailJS: diario a las 06:00, TODOS los días (el cupo de EmailJS
    # también se resetea en fin de semana, a diferencia del resto de jobs
    # de esta lista que sí respetan lun-vie).
    scheduler.add_job(
        _job_avanzar_reinicio_emailjs,
        trigger="cron",
        hour="6",
        minute="0",
        second="0",
        id="avanzar_reinicio_emailjs",
        replace_existing=True,
        misfire_grace_time=3600,
    )
    scheduler.start()
    log.info("✅ Scheduler iniciado — alertas cada 60s, lun-vie 07:00-16:00 (Atlantic/Canary)")
    log.info("✅ Scheduler — techo URGENTE admins cada 60s, lun-vie 07:00-16:59 (Atlantic/Canary)")
    log.info("✅ Scheduler — alertas techo mensual diarias, lun-vie 08:00 (Atlantic/Canary)")
    log.info("✅ Scheduler — familia/partida repetida cada 60s, lun-vie 07:00-16:59 (Atlantic/Canary)")
    log.info("✅ Scheduler — health check diario, lun-vie 07:05 (Atlantic/Canary)")
    log.info("✅ Scheduler — snapshot de tamaño de BD diario a las 08:10, todos los días (Atlantic/Canary)")
    log.info("✅ Scheduler — migración de adjuntos cerrados a Storage diaria a las 03:00, todos los días (Atlantic/Canary)")
    log.info("✅ Scheduler — alerta combinada de consumo (egress + tamaño BD) diaria, lun-vie 08:30 (Atlantic/Canary)")
    log.info("✅ Scheduler — recordatorio de emails de sistema pendientes cada 10 min, lun-vie 07:00-21:00 (Atlantic/Canary)")
    log.info("✅ Scheduler — purga de correos de sistema descartados hace >2 días, diaria a las 04:00, todos los días (Atlantic/Canary)")
    log.info("✅ Scheduler — avance automático de fechas 'reinicia cupo' EmailJS, diario a las 06:00, todos los días (Atlantic/Canary)")
    atexit.register(lambda: scheduler.shutdown(wait=False))

_iniciar_scheduler()
# ── Endpoint manual para forzar el job de alertas (admin only) ────────────────

@app.route("/api/admin/test-scheduler", methods=["POST"])
@admin_required
def test_scheduler():
    """
    Ejecuta el job de alertas inmediatamente.
    Útil para verificar que el scheduler funciona sin esperar a las 08:00/14:00.
    POST /api/admin/test-scheduler
    """
    import threading
    resultados = {"iniciado": True, "mensaje": "Job ejecutándose en segundo plano — revisa los móviles en unos segundos."}
    t = threading.Thread(target=_job_alertas_diarias, daemon=True)
    t.start()
    log.info("▶ [MANUAL] Job alertas lanzado manualmente por admin")
    return jsonify({"ok": True, **resultados})


@app.route("/api/admin/test-techo-mensual", methods=["POST"])
@admin_required
def test_techo_mensual():
    """
    Ejecuta el job de alertas de techo mensual inmediatamente.
    Útil para verificar que las notificaciones de techo funcionan sin esperar a las 08:00.
    POST /api/admin/test-techo-mensual
    """
    import threading
    t = threading.Thread(target=_job_alertas_techo_mensual, daemon=True)
    t.start()
    log.info("▶ [MANUAL] Job techo mensual lanzado manualmente por admin")
    return jsonify({
        "ok": True,
        "iniciado": True,
        "mensaje": "Job techo mensual ejecutándose en segundo plano — revisa los móviles en unos segundos."
    })


@app.route("/api/admin/test-techo-urgente", methods=["POST"])
@admin_required
def test_techo_urgente_admins():
    """
    Ejecuta el job de techo URGENTE a admins inmediatamente, ignorando
    la restricción de horario — útil para pruebas desde el panel de admin.
    POST /api/admin/test-techo-urgente
    """
    import threading
    t = threading.Thread(target=_job_techo_urgente_admins, daemon=True)
    t.start()
    log.info("\u25b6 [MANUAL] Job techo URGENTE admins lanzado manualmente por admin")
    return jsonify({
        "ok": True,
        "iniciado": True,
        "mensaje": "Job techo URGENTE admins ejecutándose en segundo plano."
    })


@app.route("/api/admin/test-familia-repetida", methods=["POST"])
@admin_required
def test_familia_repetida():
    """
    Lanza manualmente el job de alerta de familia/partida repetida.
    POST /api/admin/test-familia-repetida
    """
    import threading
    t = threading.Thread(target=_job_familia_repetida, daemon=True)
    t.start()
    log.info("\u25b6 [MANUAL] Job familia repetida lanzado manualmente por admin")
    return jsonify({
        "ok": True,
        "iniciado": True,
        "mensaje": "Job familia/partida repetida ejecutándose en segundo plano."
    })


# (2026-08-12) Nº de cuentas EmailJS en rotación — 1 (principal), 2
# (secundaria), 3 (terciaria), 4 (backup) desde v12.30.92 (antes solo
# hasta la 3). Único sitio a tocar si en el futuro se añade o se quita
# alguna: el resto del ciclo (registrar-envio, el aviso de Integridad,
# el job de avance de fechas y el panel de Admin) se calcula a partir de
# esta constante.
_EMAILJS_MAX_CUENTAS = 4

def _emailjs_cuenta_valida(valor) -> int:
    """Normaliza `emailjs_cuenta_activa` a un entero dentro de [1, _EMAILJS_MAX_CUENTAS];
    cualquier valor fuera de rango o no numérico cae a 1 (principal)."""
    try:
        n = int(valor)
    except (TypeError, ValueError):
        return 1
    return n if 1 <= n <= _EMAILJS_MAX_CUENTAS else 1

def _emailjs_siguiente_cuenta(activa: int) -> int:
    """Siguiente cuenta en el ciclo 1→2→3→4→1→... ."""
    return activa + 1 if activa < _EMAILJS_MAX_CUENTAS else 1


@app.route("/api/emailjs/config", methods=["GET"])
def api_emailjs_config():
    """
    v12.27.8 — Credenciales EmailJS activas (cuenta 1, 2, 3 o 4 según
    emailjs_cuenta_activa) + contador/umbral, para que el frontend inicialice
    emailjs.init() dinámicamente en vez de llevarlas hardcodeadas — así un
    cambio de cuenta (manual o automático al llegar al umbral) se aplica al
    momento, sin necesidad de desplegar nada.

    (2026-08-12) Generalizado de 2 a 3 cuentas (_EMAILJS_MAX_CUENTAS).
    (2026-09-01) Generalizado de 3 a 4 cuentas (_EMAILJS_MAX_CUENTAS) — sin
    más cambios en este endpoint, ya calculaba todo a partir de la constante.
    """
    c = get_config()
    activa = _emailjs_cuenta_valida(c.get("emailjs_cuenta_activa", 1))
    return jsonify({
        "ok": True,
        "cuenta_activa":  activa,
        "public_key":     c.get(f"emailjs_public_key_{activa}", "") or "",
        "service_id":     c.get(f"emailjs_service_id_{activa}", "") or "",
        "template_id":    c.get(f"emailjs_template_id_{activa}", "") or "",
        "contador":       int(c.get("emailjs_contador", 0) or 0),
        "umbral_cambio":  int(c.get("emailjs_umbral_cambio", 195) or 195),
    })


def _permite_registrar_envio_no_autenticado() -> bool:
    """
    (2026-08-11) Hay 3 flujos que envían un email REAL vía EmailJS desde un
    navegador SIN sesión iniciada todavía — recuperación de contraseña
    (solicitar_reset_password), código de verificación de login (login(),
    la sesión aún no existe en ese punto: _completar_login() no se ha
    llamado) y confirmación de Fase 2 de "solicitar acceso"
    (solicitar_usuario_fase2(), usuario nuevo sin cuenta). Antes de este
    fix, /api/emailjs/registrar-envio exigía sesión sin excepción — esos 3
    envíos SÍ consumían cuota real de EmailJS pero el contador nunca se
    enteraba, porque la llamada fallaba con 401 y el frontend se limita a
    loguearlo en consola sin más (a propósito, para no romper el envío ya
    hecho — ver enviarEmailJS() en templates/index.html).

    Arreglo: cada uno de esos 3 endpoints deja, justo antes de devolver los
    datos del email pendiente de enviar por el frontend, una marca de UN
    SOLO USO en la sesión (Flask permite `session` sin necesidad de
    "user_id" — no exige login por sí sola). Aquí se consume con `pop`
    (no `get`) para que no sirva más que para ese envío concreto — no es
    una puerta abierta a incrementar el contador a voluntad desde fuera;
    cada marca solo la puede haber puesto el propio backend, una vez, al
    preparar un envío real.
    """
    return bool(session.pop("pdte_registrar_envio_email", False))


@app.route("/api/emailjs/registrar-envio", methods=["POST"])
def api_emailjs_registrar_envio():
    """
    v12.27.8 — Llamado por el frontend justo después de cada emailjs.send()
    correcto (desde el helper central enviarEmailJS()). Incrementa el
    contador de forma atómica (UPDATE ... RETURNING, sin races entre
    usuarios concurrentes).

    (2026-08-11) Ya no lleva @login_required a secas: se acepta también
    sin sesión si el propio backend dejó la marca de un solo uso de
    _permite_registrar_envio_no_autenticado() (los 3 flujos de email sin
    login, ver esa función) — antes esos envíos reales no se contaban.

    v12.29.94 — Cambio de cuenta CÍCLICO entre 3 cuentas (antes solo
    bidireccional 1↔2): al llegar al umbral se avanza a la SIGUIENTE
    cuenta del ciclo (1→2→3→1→...) y el contador se reinicia a 0, para
    que cada cuenta cuente siempre de 1 a 195 en su propio ciclo,
    indefinidamente — pensado para que, cuando se vuelve a una cuenta ya
    usada, haya pasado tiempo de sobra para que se haya renovado del lado
    de EmailJS (su ciclo gratuito es mensual). Si la siguiente cuenta del
    ciclo no tiene las 3 credenciales completas, se prueba con la
    siguiente, y así hasta recorrer las 3; si ninguna otra cuenta tiene
    credenciales completas, NO cambia (evita dejar la app sin poder
    enviar) y se queda como aviso en Admin → Integridad para que un admin
    las rellene a tiempo.

    v12.30.92 — Ciclo ampliado a 4 cuentas (1→2→3→4→1→...), mismo
    mecanismo de siempre: solo cambió _EMAILJS_MAX_CUENTAS, esta función
    ya recorría "las _EMAILJS_MAX_CUENTAS-1 restantes" en un bucle
    genérico, sin ningún número de cuentas hardcodeado.
    """
    autenticado = "user_id" in session
    if autenticado and session.get("login_date") != _hoy_canarias().isoformat():
        session.clear()          # mismo criterio de caducidad diaria que login_required
        autenticado = False
    if not autenticado and not _permite_registrar_envio_no_autenticado():
        return jsonify({"error": "No autenticado"}), 401

    db  = get_db()
    cur = db.cursor()
    cur.execute("""
        UPDATE config_alertas SET valor = (COALESCE(valor,'0')::int + 1)::text
        WHERE clave='emailjs_contador' RETURNING valor
    """)
    row = cur.fetchone()
    contador = int((row[0] if isinstance(row, tuple) else row["valor"]) or 0)

    c = get_config()
    umbral = int(c.get("emailjs_umbral_cambio", 195) or 195)
    activa = _emailjs_cuenta_valida(c.get("emailjs_cuenta_activa", 1))
    cambiada = False
    destino = activa

    if contador >= umbral:
        candidato = activa
        encontrado = None
        for _ in range(_EMAILJS_MAX_CUENTAS - 1):
            candidato = _emailjs_siguiente_cuenta(candidato)
            pk_c  = (c.get(f"emailjs_public_key_{candidato}")  or "").strip()
            sid_c = (c.get(f"emailjs_service_id_{candidato}")  or "").strip()
            tid_c = (c.get(f"emailjs_template_id_{candidato}") or "").strip()
            if pk_c and sid_c and tid_c:
                encontrado = candidato
                break
        if encontrado is not None:
            destino = encontrado
            cur.execute("UPDATE config_alertas SET valor=%s WHERE clave='emailjs_cuenta_activa'", (str(destino),))
            cur.execute("UPDATE config_alertas SET valor='0' WHERE clave='emailjs_contador'")
            _ahora_es = datetime.now(pytz.timezone("Atlantic/Canary")).strftime("%d/%m/%Y %H:%M")
            cur.execute("UPDATE config_alertas SET valor=%s WHERE clave='emailjs_cambio_automatico_en'", (_ahora_es,))
            cambiada = True
            contador = 0
            log.warning("[EMAILJS] Contador alcanzó %s/%s — cambio automático de cuenta %s a cuenta %s (contador reiniciado)",
                        umbral, umbral, activa, destino)
        else:
            log.warning("[EMAILJS] Contador alcanzó %s/%s pero ninguna otra cuenta tiene credenciales completas — sin cambiar",
                        contador, umbral)
    db.commit()

    activa_final = destino if cambiada else activa
    c2 = get_config()  # releer tras el posible cambio, para devolver las credenciales correctas
    return jsonify({
        "ok": True,
        "contador":      contador,
        "umbral_cambio": umbral,
        "cuenta_activa": activa_final,
        "cambiada":      cambiada,
        "public_key":    c2.get(f"emailjs_public_key_{activa_final}", "") or "",
        "service_id":    c2.get(f"emailjs_service_id_{activa_final}", "") or "",
        "template_id":   c2.get(f"emailjs_template_id_{activa_final}", "") or "",
    })


@app.route("/api/admin/config-alertas", methods=["GET"])
@admin_required
def api_get_config_alertas():
    """Devuelve toda la configuración de alertas agrupada."""
    try:
        rows = rows_to_list(query(
            "SELECT clave, valor, tipo, label, grupo, orden FROM config_alertas ORDER BY grupo, orden"
        ))
        grupos = {}
        for r in rows:
            grupos.setdefault(r["grupo"], []).append(r)
        return jsonify({"ok": True, "config": rows, "grupos": grupos})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/config-alertas", methods=["PUT"])
@admin_required
def api_save_config_alertas():
    """Guarda uno o varios valores. Body: {clave: valor, ...}"""
    data = request.get_json() or {}
    if not data:
        return jsonify({"ok": False, "error": "Sin datos"}), 400
    try:
        db  = get_db()
        cur = db.cursor()
        for clave, valor in data.items():
            cur.execute("UPDATE config_alertas SET valor=%s WHERE clave=%s", (str(valor), clave))
        db.commit()
        log.info("[CONFIG] Configuración actualizada — claves: %s", list(data.keys()))
        return jsonify({"ok": True, "actualizadas": len(data)})
    except Exception as exc:
        log.error("[CONFIG] Error guardando config: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/config-avisos", methods=["GET"])
@admin_required
def api_get_config_avisos():
    """
    Devuelve el catálogo de eventos/causas, los hoteles (para el selector de
    eventos ligados a hotel), los usuarios disponibles (rol admin, compras
    o **hotel** — v12.17.0, antes solo admin/compras — activos, sin incluir
    ningún usuario técnico) y la configuración actual, para pintar la matriz
    ampliada en Administrador → Configuración de Avisos.

    GET /api/admin/config-avisos?hotel_id=<id>
    El parámetro hotel_id solo aplica a los eventos con requiere_hotel=TRUE;
    los eventos globales (requiere_hotel=FALSE) siempre muestran su config
    única (hotel_id NULL en BD), sea cual sea el hotel seleccionado.
    """
    try:
        eventos = rows_to_list(query(
            "SELECT codigo, nombre, descripcion, requiere_hotel FROM eventos_aviso ORDER BY orden, nombre"
        )) or []
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        )) or []
        usuarios = rows_to_list(query(
            "SELECT id, username, nombre, email, telegram_chat_id, rol FROM usuarios "
            "WHERE rol IN ('admin','compras','hotel') AND activo=1 "
            "ORDER BY CASE rol WHEN 'admin' THEN 0 WHEN 'compras' THEN 1 ELSE 2 END, nombre"
        )) or []

        hotel_id = None
        hotel_param = (request.args.get("hotel_id") or "").strip()
        if hotel_param and hotel_param.lower() != "global":
            try:
                hotel_id = int(hotel_param)
            except ValueError:
                hotel_id = None
        if hotel_id is None and hoteles:
            hotel_id = hoteles[0]["id"]  # selección por defecto: el primer hotel activo

        # Config de eventos globales (hotel_id NULL, siempre visible) +
        # config del hotel seleccionado para los eventos que lo requieren.
        filas = rows_to_list(query(
            """SELECT evento_codigo, hotel_id, usuario_id, telegram, email, popup
               FROM notificaciones_config
               WHERE hotel_id IS NULL OR hotel_id = %s""",
            (hotel_id,)
        )) or []
        config = {}
        for f in filas:
            config.setdefault(f["evento_codigo"], {})[str(f["usuario_id"])] = {
                "telegram": bool(f["telegram"]),
                "email": bool(f["email"]),
                "popup": bool(f["popup"]),
            }
        return jsonify({
            "ok": True, "eventos": eventos, "hoteles": hoteles, "usuarios": usuarios,
            "config": config, "hotel_id": hotel_id,
        })
    except Exception as exc:
        log.error("[CONFIG-AVISOS] Error GET: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/config-avisos", methods=["PUT"])
@admin_required
def api_save_config_avisos():
    """
    Guarda uno o varios cambios de la matriz de Configuración de Avisos.
    Body: {"cambios": [{"evento_codigo":"...", "hotel_id": 9|null, "usuario_id":1,
                        "telegram":true, "email":false, "popup":true}, ...]}
    hotel_id debe ser null para eventos globales (requiere_hotel=FALSE) y el
    id de hotel correspondiente para eventos ligados a hotel.
    PUT /api/admin/config-avisos
    """
    data = request.get_json() or {}
    cambios = data.get("cambios") or []
    if not cambios:
        return jsonify({"ok": False, "error": "Sin cambios"}), 400
    try:
        db  = get_db()
        cur = db.cursor()
        # Defensa en profundidad: el hotel_id correcto de cada evento se decide
        # aquí, en servidor, según requiere_hotel en eventos_aviso — nunca según
        # lo que mande el navegador. Así, aunque el frontend tuviera un fallo,
        # es físicamente imposible guardar un evento global con hotel_id real
        # (o uno por-hotel sin él).
        _req_hotel = {r["codigo"]: bool(r["requiere_hotel"]) for r in
                      (rows_to_list(query("SELECT codigo, requiere_hotel FROM eventos_aviso")) or [])}
        for c in cambios:
            evento_codigo = c.get("evento_codigo")
            usuario_id    = c.get("usuario_id")
            hotel_id      = c.get("hotel_id")  # None para eventos globales
            telegram      = bool(c.get("telegram", False))
            email         = bool(c.get("email", False))
            popup         = bool(c.get("popup", False))
            if not evento_codigo or not usuario_id:
                continue
            if evento_codigo not in _req_hotel:
                continue  # evento_codigo desconocido — se ignora en vez de fallar todo el lote
            if not _req_hotel[evento_codigo]:
                hotel_id = None  # evento global: nunca se guarda con hotel_id real, venga lo que venga
            elif hotel_id is None:
                continue  # evento por hotel sin hotel_id — nada que guardar, se ignora
            # DELETE + INSERT en vez de ON CONFLICT: en Postgres dos filas con
            # hotel_id NULL nunca "chocan" en un UNIQUE (cada NULL es distinto),
            # así que ON CONFLICT no sirve para deduplicar los eventos globales.
            # IS NOT DISTINCT FROM sí compara NULL=NULL como iguales.
            cur.execute(
                "DELETE FROM notificaciones_config "
                "WHERE evento_codigo=%s AND usuario_id=%s AND hotel_id IS NOT DISTINCT FROM %s",
                (evento_codigo, usuario_id, hotel_id)
            )
            if telegram or email or popup:
                cur.execute(
                    """INSERT INTO notificaciones_config
                       (evento_codigo, hotel_id, usuario_id, telegram, email, popup)
                       VALUES (%s,%s,%s,%s,%s,%s)""",
                    (evento_codigo, hotel_id, usuario_id, telegram, email, popup)
                )
        db.commit()
        log.info("[CONFIG-AVISOS] Actualizados %d destinatario(s) por admin", len(cambios))
        return jsonify({"ok": True, "actualizados": len(cambios)})
    except Exception as exc:
        log.error("[CONFIG-AVISOS] Error PUT: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/config-avisos/resolver", methods=["GET"])
@login_required
def api_resolver_config_avisos():
    """
    Consulta en tiempo real de destinatarios configurados para un evento, en
    el formato que necesita cualquier módulo que envíe avisos — incluido
    main_agenda (vía el bridge, con la misma sesión que /api/bridge/*).
    GET /api/config-avisos/resolver?evento=techo_urgente_admin&canal=telegram
    GET /api/config-avisos/resolver?evento=cambio_estado_pedido&canal=popup&hotel_id=9
    hotel_id es opcional — obligatorio en la práctica solo para eventos con
    requiere_hotel=TRUE (cambio_estado_pedido, alerta_pedido_hotel).
    Respuesta: {"ok": true, "destinatarios": [{"username":.., "telegram_chat_id":.., "email":..}, ...]}
    """
    evento_codigo = (request.args.get("evento") or "").strip()
    canal = (request.args.get("canal") or "telegram").strip().lower()
    hotel_param = (request.args.get("hotel_id") or "").strip()
    if not evento_codigo:
        return jsonify({"ok": False, "error": "Falta el parámetro 'evento'"}), 400
    if canal not in ("telegram", "email", "popup"):
        return jsonify({"ok": False, "error": "canal debe ser 'telegram', 'email' o 'popup'"}), 400
    hotel_id = None
    if hotel_param:
        try:
            hotel_id = int(hotel_param)
        except ValueError:
            return jsonify({"ok": False, "error": "hotel_id debe ser numérico"}), 400
    destinatarios = _resolver_notificacion(evento_codigo, canal, hotel_id=hotel_id)
    return jsonify({"ok": True, "evento": evento_codigo, "canal": canal, "hotel_id": hotel_id, "destinatarios": destinatarios})


# ── API Admin: correo de departamento por hotel (2026-08-28) ────────────────
# Ver PENDIENTES.md / _auto_migrate() (tabla departamento_hotel_email) /
# enviar_emails_estado() (uso real, correo interno de cambio de estado).
# `departamentos` es un catálogo único y global — esta pantalla es la
# primera administración que tiene ("Departamentos" no existía como vista
# hasta ahora), así que el GET también sirve para listar el catálogo en sí.

@app.route("/api/admin/departamentos-email", methods=["GET"])
@admin_required
def api_get_departamentos_email():
    """
    GET /api/admin/departamentos-email?hotel_id=<id>
    Devuelve el catálogo de departamentos y, para el hotel indicado (o el
    primer hotel activo si no se indica), el correo (o dos) ya registrado
    para cada uno — igual patrón que GET /api/admin/config-avisos.
    """
    try:
        departamentos = rows_to_list(query(
            "SELECT id, nombre FROM departamentos WHERE activo=1 ORDER BY nombre"
        )) or []
        hoteles = rows_to_list(query(
            "SELECT id, codigo, nombre FROM hoteles WHERE activo=1 ORDER BY codigo"
        )) or []

        hotel_id = None
        hotel_param = (request.args.get("hotel_id") or "").strip()
        if hotel_param:
            try:
                hotel_id = int(hotel_param)
            except ValueError:
                hotel_id = None
        if hotel_id is None and hoteles:
            hotel_id = hoteles[0]["id"]  # selección por defecto: el primer hotel activo

        config = {}
        if hotel_id is not None:
            filas = rows_to_list(query(
                "SELECT departamento_id, email, email2 FROM departamento_hotel_email WHERE hotel_id=%s",
                (hotel_id,)
            )) or []
            for f in filas:
                config[str(f["departamento_id"])] = {"email": f["email"] or "", "email2": f["email2"] or ""}

        return jsonify({
            "ok": True, "departamentos": departamentos, "hoteles": hoteles,
            "config": config, "hotel_id": hotel_id,
        })
    except Exception as exc:
        log.error("[DEPTO-EMAIL] Error GET: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/departamentos-email", methods=["PUT"])
@admin_required
def api_save_departamentos_email():
    """
    Guarda los correos de departamento de UN hotel de una vez (toda la
    tabla de esa pantalla, igual que el resto de pantallas de
    administración de esta app).
    Body: {"hotel_id": 9, "filas": [{"departamento_id": 3, "email": "...",
                                      "email2": "..."}, ...]}
    Una fila con email y email2 vacíos borra el registro (vuelve a "sin
    correo configurado" para ese departamento en ese hotel) en vez de
    dejar una fila vacía.
    """
    data = request.get_json() or {}
    hotel_id = data.get("hotel_id")
    filas    = data.get("filas") or []
    if not hotel_id:
        return jsonify({"ok": False, "error": "Falta hotel_id"}), 400
    try:
        hotel_id = int(hotel_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "hotel_id debe ser numérico"}), 400
    try:
        db  = get_db()
        cur = db.cursor()
        actualizados = 0
        for f in filas:
            depto_id = f.get("departamento_id")
            if not depto_id:
                continue
            email  = (f.get("email") or "").strip()
            email2 = (f.get("email2") or "").strip()
            if not email and not email2:
                cur.execute(
                    "DELETE FROM departamento_hotel_email WHERE hotel_id=%s AND departamento_id=%s",
                    (hotel_id, depto_id)
                )
            else:
                cur.execute(
                    """INSERT INTO departamento_hotel_email (hotel_id, departamento_id, email, email2)
                       VALUES (%s,%s,%s,%s)
                       ON CONFLICT (hotel_id, departamento_id)
                       DO UPDATE SET email=EXCLUDED.email, email2=EXCLUDED.email2""",
                    (hotel_id, depto_id, email or None, email2 or None)
                )
            actualizados += 1
        db.commit()
        log.info("[DEPTO-EMAIL] Hotel %s — %d departamento(s) actualizados por admin", hotel_id, actualizados)
        return jsonify({"ok": True, "actualizados": actualizados})
    except Exception as exc:
        log.error("[DEPTO-EMAIL] Error PUT: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ── API Admin: contactos adicionales de notificación (2026-08-28) ──────────
# A petición de Víctor: contactos sueltos (no son usuarios de la app, p. ej.
# "Chef Ejecutivo", "Director de Compras", "Administrativo A&B") que se
# ponen en copia en el correo interno de cambio de estado según el
# departamento del pedido y el estado nuevo — global para toda la cadena
# (decisión confirmada con Víctor, a diferencia del correo de departamento
# por hotel de arriba). Ver notificacion_contactos / notificacion_contacto_
# reglas (_auto_migrate/models.py) y el uso real en enviar_emails_estado().
# Los "estados" ofrecidos son ESTADOS_EMAIL_INTERNO — los únicos que de
# verdad disparan el correo interno de cambio de estado; no tendría efecto
# alguno configurar una regla para un estado fuera de esa lista.
_ESTADOS_EMAIL_INTERNO_ORDENADOS = [e for e in ESTADOS_VALIDOS if e in ESTADOS_EMAIL_INTERNO]

# (2026-08-28) Pseudo-estado EXCLUSIVO de "Notificaciones adicionales" — a
# petición de Víctor: además de las 5 combinaciones reales de
# ESTADOS_EMAIL_INTERNO, permite poner un contacto en copia
# ESPECÍFICAMENTE cuando el pedido que se acaba de enviar al proveedor
# había superado el techo de gastos del mes y tuvo que pasar por
# autorización de Dirección General (aprobar_expediente()) —
# independiente de si ese mismo contacto también está marcado para
# "ENVIADO AL PROVEEDOR" (el envío normal, sin exceso): las dos reglas
# son compatibles y no se excluyen entre sí. NUNCA es un estado real de
# un pedido — no está en ESTADOS_VALIDOS ni en ESTADOS_EMAIL_INTERNO, solo
# tiene sentido dentro de notificacion_contacto_reglas.estado. Se detecta
# en enviar_emails_estado() comprobando que el estado ANTERIOR del pedido
# era "PENDIENTE Vº Bº DIRECCIÓN GENERAL": ese estado solo lo pone
# create_pedido()/update_pedido() al superar el techo, y desde ahí el
# único destino posible es aprobar_expediente() (ver su propio
# docstring), así que esa combinación (estado_nuevo=ENVIADO AL PROVEEDOR,
# estado_antes=PENDIENTE Vº Bº DIRECCIÓN GENERAL) identifica sin
# ambigüedad un envío tras exceso de techo autorizado.
ESTADO_NOTIF_EXCESO_TECHO_DG = "EXCESO TECHO AUTORIZADO (DIRECCIÓN GENERAL)"
_ESTADOS_NOTIF_ADICIONAL_VALIDOS = set(ESTADOS_EMAIL_INTERNO) | {ESTADO_NOTIF_EXCESO_TECHO_DG}

@app.route("/api/admin/notificaciones-contactos", methods=["GET"])
@admin_required
def api_get_notificaciones_contactos():
    """
    Devuelve todos los contactos adicionales con sus reglas (departamento +
    estado) ya configuradas, el catálogo de departamentos y la lista de
    estados que tiene sentido ofrecer (ESTADOS_EMAIL_INTERNO) — todo lo que
    necesita el frontend para pintar, por cada contacto, una matriz de
    checkboxes Departamento × Estado. Se devuelve aparte, en
    "estado_exceso_techo", el pseudo-estado de "exceso de techo autorizado"
    (ver ESTADO_NOTIF_EXCESO_TECHO_DG más arriba) para que el frontend lo
    pinte como una columna extra, distinguida visualmente de las 5
    columnas de estado real.
    """
    try:
        contactos = rows_to_list(query(
            "SELECT id, nombre, email, email2, activo FROM notificacion_contactos ORDER BY nombre"
        )) or []
        reglas = rows_to_list(query(
            "SELECT contacto_id, departamento_id, estado FROM notificacion_contacto_reglas"
        )) or []
        reglas_por_contacto = {}
        for r in reglas:
            reglas_por_contacto.setdefault(r["contacto_id"], []).append(
                {"departamento_id": r["departamento_id"], "estado": r["estado"]}
            )
        for c in contactos:
            c["reglas"] = reglas_por_contacto.get(c["id"], [])

        departamentos = rows_to_list(query(
            "SELECT id, nombre FROM departamentos WHERE activo=1 ORDER BY nombre"
        )) or []

        return jsonify({
            "ok": True, "contactos": contactos, "departamentos": departamentos,
            "estados": _ESTADOS_EMAIL_INTERNO_ORDENADOS,
            "estado_exceso_techo": ESTADO_NOTIF_EXCESO_TECHO_DG,
        })
    except Exception as exc:
        log.error("[NOTIF-CONTACTOS] Error GET: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/notificaciones-contactos", methods=["POST"])
@admin_required
def api_crear_notificacion_contacto():
    """Crea un contacto nuevo, sin reglas todavía (se añaden con el PUT)."""
    data   = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"ok": False, "error": "El nombre es obligatorio"}), 400
    email  = (data.get("email") or "").strip() or None
    email2 = (data.get("email2") or "").strip() or None
    try:
        db  = get_db()
        cur = execute(
            "INSERT INTO notificacion_contactos (nombre, email, email2) VALUES (%s,%s,%s) RETURNING id",
            (nombre, email, email2)
        )
        cid = cur.fetchone()["id"]
        db.commit()
        log.info("[NOTIF-CONTACTOS] Contacto creado: %s (id=%s)", nombre, cid)
        return jsonify({"ok": True, "id": cid}), 201
    except Exception as exc:
        log.error("[NOTIF-CONTACTOS] Error POST: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/notificaciones-contactos/<int:cid>", methods=["PUT"])
@admin_required
def api_actualizar_notificacion_contacto(cid):
    """
    Actualiza nombre/email/email2/activo de un contacto Y reemplaza por
    completo su conjunto de reglas (departamento + estado) de una sola vez
    — mismo patrón "Guardar cambios" de una tabla entera que el resto de
    pantallas de administración de esta app. Body:
    {"nombre": "...", "email": "...", "email2": "...", "activo": true,
     "reglas": [{"departamento_id": 3, "estado": "ENVIADO AL PROVEEDOR"}, ...]}
    Una regla con un estado fuera de _ESTADOS_NOTIF_ADICIONAL_VALIDOS (los 5
    de ESTADOS_EMAIL_INTERNO + el pseudo-estado ESTADO_NOTIF_EXCESO_TECHO_DG)
    se descarta en silencio (no tendría ningún efecto — ver comentario de
    arriba).
    """
    data   = request.get_json(silent=True) or {}
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return jsonify({"ok": False, "error": "El nombre es obligatorio"}), 400
    email   = (data.get("email") or "").strip() or None
    email2  = (data.get("email2") or "").strip() or None
    activo  = 1 if data.get("activo", True) else 0
    reglas  = data.get("reglas") or []
    try:
        db  = get_db()
        cur = db.cursor()
        existe = query("SELECT id FROM notificacion_contactos WHERE id=%s", (cid,), one=True)
        if not existe:
            return jsonify({"ok": False, "error": "Contacto no encontrado"}), 404

        cur.execute(
            "UPDATE notificacion_contactos SET nombre=%s, email=%s, email2=%s, activo=%s WHERE id=%s",
            (nombre, email, email2, activo, cid)
        )
        cur.execute("DELETE FROM notificacion_contacto_reglas WHERE contacto_id=%s", (cid,))
        n_reglas = 0
        for r in reglas:
            depto_id = r.get("departamento_id")
            estado   = (r.get("estado") or "").strip()
            if not depto_id or estado not in _ESTADOS_NOTIF_ADICIONAL_VALIDOS:
                continue
            cur.execute(
                "INSERT INTO notificacion_contacto_reglas (contacto_id, departamento_id, estado) "
                "VALUES (%s,%s,%s) ON CONFLICT DO NOTHING",
                (cid, depto_id, estado)
            )
            n_reglas += 1
        db.commit()
        log.info("[NOTIF-CONTACTOS] Contacto %s actualizado — %d regla(s)", cid, n_reglas)
        return jsonify({"ok": True, "reglas_guardadas": n_reglas})
    except Exception as exc:
        log.error("[NOTIF-CONTACTOS] Error PUT %s: %s", cid, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/notificaciones-contactos/<int:cid>", methods=["DELETE"])
@admin_required
def api_eliminar_notificacion_contacto(cid):
    """Elimina un contacto y, en cascada (ON DELETE CASCADE), todas sus reglas."""
    try:
        db = get_db()
        execute("DELETE FROM notificacion_contactos WHERE id=%s", (cid,))
        db.commit()
        log.info("[NOTIF-CONTACTOS] Contacto %s eliminado", cid)
        return jsonify({"ok": True})
    except Exception as exc:
        log.error("[NOTIF-CONTACTOS] Error DELETE %s: %s", cid, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


# ─────────────────────────────────────────────────────────────────────────
# (2026-08-27) Puente de correos desde DALI (dali-sap-articulos-app) ───────
# Víctor: "podemos aprovechar la organizacion que tenemos actualmente en
# controlpendidos para el envio de correos y que los correos de
# dalisaparticulos utilice la misma infraestuctura? ... se podria generar
# dejar en cola y cuando alguien abra control de pedidos se lance". El
# catálogo DALI (backend en Node, sin SMTP propio tampoco) genera el HTML
# de sus correos de "documentación faltante" con el mismo diseño de
# cabecera que esta app (mismo logo, mismos colores — ver
# _email_header_html más abajo) y llama a este endpoint para dejarlo en la
# MISMA cola que ya usa esta app (emails_sistema_pendientes). El poller de
# siempre (_enviarEmailsSistemaPendientes, ver templates/index.html) lo
# recoge y lo envía por EmailJS la próxima vez que alguien con sesión
# admin/compras tenga esta app abierta (o dentro de los 5 minutos
# siguientes si ya la tiene abierta) — exactamente el mecanismo que pidió,
# sin tocar nada del propio poller/frontend. A petición de Víctor, estos
# correos comparten la cuenta EmailJS activa de esta app (sin cuenta
# dedicada aparte) — el cupo entra en la rotación normal entre las 3
# cuentas si hiciera falta.
#
# Se reutiliza el secreto YA compartido entre los dos servicios de Render
# (DALI_SSO_SECRET, hasta ahora solo usado para el SSO del menú lateral
# "Catálogo DALI") en vez de dar de alta uno nuevo — nada que configurar de
# más en Render, ya está desplegado igual en ambos servicios.
#
# El secreto en sí nunca viaja en la petición: DALI firma el cuerpo con
# HMAC-SHA256 (mismo esquema que _generar_token_sso_dali /
# _verificarTokenSso en authController.js, pero sin jti — como mucho, un
# replay duplica una fila en la cola, no compromete ninguna cuenta) y aquí
# se verifica en tiempo constante. Cabeceras esperadas:
#   X-Dali-Timestamp: <epoch segundos, UTC>
#   X-Dali-Signature: hmac_sha256_hex(DALI_SSO_SECRET, f"{timestamp}." + cuerpo_json_crudo)
# Se rechaza si la firma no coincide o si el timestamp se aleja de ahora
# más de _DALI_BRIDGE_MARGEN_SEGUNDOS (relojes desincronizados/latencia,
# pero también acota cuánto tiempo es reutilizable una firma capturada).
_DALI_BRIDGE_MARGEN_SEGUNDOS = 300  # 5 minutos

def _dali_bridge_firma_valida(cuerpo_crudo: bytes) -> bool:
    if not DALI_SSO_SECRET:
        return False
    timestamp = request.headers.get("X-Dali-Timestamp", "")
    firma_recibida = request.headers.get("X-Dali-Signature", "")
    if not timestamp or not firma_recibida:
        return False
    try:
        ts = int(timestamp)
    except ValueError:
        return False
    ahora = datetime.now(timezone.utc).timestamp()
    if abs(ahora - ts) > _DALI_BRIDGE_MARGEN_SEGUNDOS:
        return False
    mensaje = f"{timestamp}.".encode() + cuerpo_crudo
    firma_esperada = hmac.new(DALI_SSO_SECRET.encode(), mensaje, hashlib.sha256).hexdigest()
    return hmac.compare_digest(firma_recibida, firma_esperada)


@app.route("/api/externo/dali-sap/emails-pendientes", methods=["POST"])
def api_externo_dali_encolar_email():
    """
    POST /api/externo/dali-sap/emails-pendientes — encola un email en la
    MISMA cola que usa esta app (emails_sistema_pendientes) por cuenta del
    catálogo DALI (dali-sap-articulos-app). Ver comentario largo arriba.

    Body JSON: { destinatario, asunto, cuerpo_html, cuerpo_text?, cc_emails? }
    Requiere las cabeceras X-Dali-Timestamp / X-Dali-Signature (ver
    _dali_bridge_firma_valida). Sin sesión de usuario — es una llamada
    servidor a servidor, DALI no tiene (ni debe tener) cookie de esta app.
    """
    cuerpo_crudo = request.get_data()
    if not _dali_bridge_firma_valida(cuerpo_crudo):
        return jsonify({"ok": False, "error": "Firma inválida o caducada."}), 401

    datos = request.get_json(silent=True) or {}
    destinatario = (datos.get("destinatario") or "").strip()
    asunto = (datos.get("asunto") or "").strip()
    cuerpo_html = datos.get("cuerpo_html") or ""
    cuerpo_text = datos.get("cuerpo_text") or ""
    cc_emails = (datos.get("cc_emails") or "").strip() or None

    if not destinatario or not asunto or not cuerpo_html:
        return jsonify({"ok": False, "error": "Faltan destinatario, asunto o cuerpo_html."}), 400

    try:
        cur = execute(
            """INSERT INTO emails_sistema_pendientes
               (evento_codigo, destinatario, asunto, cuerpo_html, cuerpo_text, cc_emails, visible_en)
               VALUES ('dali_documentacion_faltante', %s, %s, %s, %s, %s, NOW())
               RETURNING id""",
            (destinatario, asunto, cuerpo_html, cuerpo_text, cc_emails)
        )
        nuevo_id = cur.fetchone()["id"]
        get_db().commit()
        log.info("[DALI-BRIDGE] Email de documentación faltante encolado (id=%s) para %s",
                  nuevo_id, destinatario)
        return jsonify({"ok": True, "id": nuevo_id})
    except Exception as exc:
        log.error("[DALI-BRIDGE] Error encolando email de DALI: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/externo/dali-sap/proveedores", methods=["GET"])
def api_externo_dali_proveedores():
    """
    GET /api/externo/dali-sap/proveedores — lista de proveedores activos
    de ESTA app con sus contactos (nombre/email/es_principal), para que
    DALI pueda usar directamente estos contactos como destinatario al
    encolar un email (ver api_externo_dali_encolar_email más arriba),
    en vez de mantener un único email de contacto por proveedor
    duplicado en su propia base de datos.

    (2026-08-27) Víctor: "como vamos a utilizar el sistema de envios de
    control_pedidos, podriamos utilizar tambien el apartado de
    proveedores con sus correos electronicos etc? de esta manera los
    tenemos unicamente en un unico punto y podemos incluir mas correos
    para el envio, ahora en articulos es solo uno". Decidió cruzar por
    NOMBRE exacto (va a mantener el nombre de cada proveedor aquí
    idéntico al de DALI a propósito — "trabajar sobre una única base")
    en vez de mantener un mapeo id-a-id aparte entre las dos apps; el
    cruce en sí lo hace DALI con la lista completa que devuelve este
    endpoint (comparando nombres normalizados), no hace falta que este
    lado sepa nada de los ids de DALI. `email_principal` es el mismo
    contacto "principal" (o el primero si no hay ninguno marcado) que ya
    usa el resto de esta app — mismo criterio que Documentación
    faltante de DALI usaba con su único email propio, así que quien no
    tenga contacto marcado como principal en Control de Pedidos no
    pierde exactamente nada respecto a antes.

    Sin sesión de usuario (llamada servidor a servidor) — misma firma
    HMAC que el endpoint de encolar, aquí con cuerpo vacío (GET).
    """
    if not _dali_bridge_firma_valida(request.get_data()):
        return jsonify({"ok": False, "error": "Firma inválida o caducada."}), 401
    try:
        rows = query("SELECT id,codigo,nombre,observaciones FROM proveedores WHERE activo=1 ORDER BY nombre")
        result = _prov_with_contactos(rows)
        salida = [
            {
                "nombre": p["nombre"],
                "email_principal": p["email"] or None,
                "contactos": [
                    {"nombre": c["nombre"], "email": c["email"], "es_principal": c["es_principal"]}
                    for c in p["contactos"]
                    if c.get("email")
                ],
            }
            for p in result
        ]
        return jsonify({"ok": True, "proveedores": salida})
    except Exception as exc:
        log.error("[DALI-BRIDGE] Error listando proveedores para DALI: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/externo/dali-sap/compradores", methods=["GET"])
def api_externo_dali_compradores():
    """
    GET /api/externo/dali-sap/compradores — nombre/email/móvil de los
    usuarios de ESTA app con rol 'compras' o 'admin', para que el
    catálogo DALI pueda firmar sus correos a proveedores (pantalla
    "Documentación faltante") con el mismo nombre/teléfono/email que ya
    usa esa persona en el resto de correos de Control de Pedidos, en vez
    de duplicar esos datos en la base de datos de DALI.

    (2026-08-29) Víctor, sobre la firma de esos correos: "¿puedes coger
    la info de la ficha usuarios control pedidos? los admin son los
    mismos y los compradores son admin en catalogo dali" — los
    administradores de Catálogo DALI son las mismas personas ya dadas de
    alta aquí como compradores (rol 'compras') o administradoras (rol
    'admin'); DALI cruza por EMAIL (el único identificador que comparte
    con la sesión de esa app) contra la lista que devuelve este
    endpoint, así que aquí basta con exponer ambos roles sin necesidad
    de saber nada de las cuentas de DALI. Se omiten los usuarios sin
    email (no habría con qué cruzarlos) y los inactivos.

    Sin sesión de usuario (llamada servidor a servidor) — misma firma
    HMAC que el resto del puente DALI (ver api_externo_dali_proveedores
    arriba), aquí también con cuerpo vacío (GET).

    (2026-08-31) Norma explícita a petición de Víctor, tras detectar una
    colisión real entre dos cuentas con el mismo email principal ("¿puedes
    coger la info de la ficha usuarios control pedidos?" derivó en que dos
    usuarios distintos tenían el mismo `email`, y DALI cogía el que no
    tenía móvil): esta consulta SOLO mira `email` (el principal, único por
    convención de uso aunque no forzado por una constraint UNIQUE), nunca
    `email2` — Víctor: "tener en cuenta que este mismo correo tambien es
    correo secundario en otro usuario, asi que podemos poner como norma
    que solo mire en el primer correo de cada usuario". `email2` puede
    repetirse a propósito entre varias cuentas (para que más de una
    persona reciba copia de avisos), así que cruzar también por ahí
    multiplicaría las colisiones en vez de evitarlas. Si en el futuro
    hiciera falta ampliar este cruce a `email2`, que sea una decisión
    deliberada y no un descuido de "más datos es mejor".
    """
    if not _dali_bridge_firma_valida(request.get_data()):
        return jsonify({"ok": False, "error": "Firma inválida o caducada."}), 401
    try:
        rows = query(
            """SELECT nombre, email, movil FROM usuarios
               WHERE activo = 1 AND rol IN ('compras', 'admin') AND email IS NOT NULL
               ORDER BY nombre"""
        )
        salida = [
            {"nombre": r["nombre"], "email": r["email"], "movil": r["movil"] or None}
            for r in rows
        ]
        return jsonify({"ok": True, "compradores": salida})
    except Exception as exc:
        log.error("[DALI-BRIDGE] Error listando compradores para DALI: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


# (2026-08-19) Nº máximo de intentos de envío de una fila de
# emails_sistema_pendientes antes de dejar de reintentarse sola — ver
# api_emails_sistema_pendientes() y api_emails_sistema_atascados().
# (2026-08-20) Bajado de 8 a 3: la migración que añade la columna
# `intentos` la rellena a 0 en las filas YA existentes (no puede saber
# cuántas veces se habían reintentado antes de esta versión) — así que
# las filas que ya llevaban atascadas desde antes del despliegue de este
# freno arrancan de nuevo con el cupo completo de intentos. Con 8 de
# margen, cada una de esas filas podía seguir descontando cupo de EmailJS
# unas cuantas veces más tras el propio despliegue del fix (reportado por
# Víctor: siguió bajando cupo justo después de desplegar el freno). Con
# 3 ese margen se acorta bastante sin cargarse los reintentos legítimos
# por fallos puntuales de red.
MAX_INTENTOS_EMAIL_SISTEMA = 3


@app.route("/api/emails-sistema-pendientes", methods=["GET"])
@login_required
def api_emails_sistema_pendientes():
    """
    Devuelve los emails de sistema pendientes de envío (encolados por jobs
    sin navegador abierto) para que el frontend los envíe vía EmailJS —
    y los RESERVA atómicamente en el mismo paso (v12.29.96).

    (2026-08-13) Antes esto era un SELECT simple: si dos pestañas/sesiones
    (o un poll normal solapado con una recarga de página) pedían la cola
    casi a la vez, ambas recibían la misma fila "pendiente" y ambas la
    mandaban de verdad por EmailJS antes de que ninguna llegara a marcarla
    como enviada — duplicados reales al destinatario (reportado por el
    usuario: pedido 39909, 2 correos idénticos el mismo minuto). La app
    corre con varios hilos (`--worker-class gthread --threads 4` en
    render.yaml), así que dos peticiones sí pueden ejecutarse en paralelo
    de verdad dentro del mismo proceso.

    Corrección: `UPDATE ... SELECT ... FOR UPDATE SKIP LOCKED ... RETURNING`
    en una sola sentencia atómica marca `en_proceso_desde = NOW()` en las
    filas que devuelve, así que una segunda petición concurrente ya no las
    ve como disponibles (el filtro excluye las reservadas hace menos de 2
    minutos) y con `SKIP LOCKED` ni siquiera se bloquea esperando a que la
    primera termine. Si la sesión que reservó una fila nunca confirma el
    envío (fallo de EmailJS, se cierra la pestaña a media faena...), la
    reserva caduca sola pasados 2 minutos y otra sesión puede reintentarla
    con normalidad — no se pierden envíos por este cambio.

    (2026-08-14) Filtro adicional `visible_en <= NOW()`: los correos de
    cambio de estado de pedido (evento_codigo 'cambio_estado_proveedor' /
    'cambio_estado_interno', encolados por _encolar_email_pedido_retrasado)
    se insertan con visible_en 5 minutos en el futuro y no deben devolverse
    hasta que se cumpla ese plazo — es la misma idea que el retraso ya
    aplicado al popup del bridge. El resto de eventos de esta cola
    (techo urgente, familias repetidas...) tienen visible_en = NOW() por
    defecto, así que siguen recogiéndose de inmediato, sin cambio.

    (2026-08-19) Freno de reintentos infinitos: cada vez que se reclama una
    fila se incrementa `intentos`; a partir de MAX_INTENTOS_EMAIL_SISTEMA
    ya no se vuelve a devolver — deja de reintentarse sola (antes, un
    correo que fallase SIEMPRE al enviarse, p. ej. por tamaño, reintentaba
    sin límite cada vez que caducaba la reserva de 2 minutos, descontando
    cupo de EmailJS en cada intento aunque nunca llegase a entregarse:
    reportado por Víctor, el contador de EmailJS subió de 54 a 71 sin que
    llegara ningún correo nuevo). También se excluyen las filas marcadas
    `descartado_en` (descarte manual, ver
    /api/admin/emails-sistema-atascados). Las filas que llegan al máximo
    de intentos sin descartar manualmente quedan visibles en ese mismo
    panel, no se pierden ni se borran solas.
    (v12.32.03) Si la conexión obtenida del pool resulta estar en modo
    solo-lectura (visto una vez en producción: `cannot execute UPDATE in
    a read-only transaction`, sin relación con nada que esta app
    configure explícitamente — probablemente una conexión reciclada del
    pool que quedó en ese estado tras un evento puntual de Supabase), se
    descarta esa conexión del pool y se reintenta UNA vez con una nueva,
    en vez de fallar directamente el listado completo de esa pasada.
    GET /api/emails-sistema-pendientes
    """
    def _listar_y_reservar():
        cur = execute(
            """
            UPDATE emails_sistema_pendientes
               SET en_proceso_desde = NOW(), intentos = intentos + 1
             WHERE id IN (
                 SELECT id FROM emails_sistema_pendientes
                  WHERE enviado = FALSE
                    AND visible_en <= NOW()
                    AND descartado_en IS NULL
                    AND intentos < %s
                    AND (en_proceso_desde IS NULL OR en_proceso_desde < NOW() - INTERVAL '2 minutes')
                  ORDER BY id
                  LIMIT 20
                  FOR UPDATE SKIP LOCKED
             )
            RETURNING id, evento_codigo, destinatario, asunto, cuerpo_html, cuerpo_text,
                      cc_emails, pedido_id
            """,
            (MAX_INTENTOS_EMAIL_SISTEMA,)
        )
        pendientes = rows_to_list(cur.fetchall()) or []
        pendientes.sort(key=lambda p: p["id"])
        get_db().commit()
        return pendientes

    try:
        pendientes = _listar_y_reservar()
        return jsonify({"ok": True, "pendientes": pendientes})
    except Exception as exc:
        if "read-only transaction" in str(exc).lower():
            log.warning("[EMAILS-SISTEMA] Conexión en modo solo-lectura, se descarta del pool y se reintenta: %s", exc)
            try:
                db = g.pop("db", None)
                if db is not None and _db_pool is not None:
                    try:
                        db.rollback()
                    except Exception:
                        pass
                    _db_pool.putconn(db, close=True)
                pendientes = _listar_y_reservar()
                return jsonify({"ok": True, "pendientes": pendientes})
            except Exception as exc2:
                log.error("[EMAILS-SISTEMA] Error listando/reservando pendientes tras reintentar: %s", exc2)
                return jsonify({"ok": False, "error": str(exc2)}), 500
        log.error("[EMAILS-SISTEMA] Error listando/reservando pendientes: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/emails-sistema-atascados", methods=["GET"])
@admin_required
def api_emails_sistema_atascados():
    """
    (2026-08-19) Lista los correos de la cola de sistema pendientes de
    enviar, para que un admin pueda ver qué se ha quedado atascado
    (típicamente por tamaño, EmailJS 413) y descartarlo si corresponde, en
    vez de que quede invisible drenando cupo de EmailJS en segundo plano.

    (2026-08-20) Ampliado para devolver TODA la cola pendiente (enviado =
    FALSE), no solo las filas que ya agotaron MAX_INTENTOS_EMAIL_SISTEMA.
    Motivo: al añadir la columna `intentos` (ADD COLUMN ... DEFAULT 0), las
    filas YA existentes en la cola (p. ej. correos de resumen de
    comparativas de antes de v12.30.20, con el HTML grande de antes del
    ajuste de recorte adaptativo) arrancan con `intentos = 0` — es decir,
    con cupo entero de reintentos por delante, y hasta que no lo agotan no
    aparecían en este panel. Resultado (reportado por Víctor): el contador
    de EmailJS seguía bajando tras desplegar el freno de reintentos, sin
    que el admin tuviera forma de ver ni descartar esas filas a mano hasta
    que fallaban varias veces más. Ahora se listan TODAS las pendientes,
    ordenadas por tamaño de HTML descendente (las más grandes — más
    sospechosas de ser las que fallan por 413 — arriba del todo) con un
    campo `atascado` (true si ya alcanzó MAX_INTENTOS_EMAIL_SISTEMA o fue
    descartada) para que el frontend distinga "aún reintentando" de
    "parada", pero el botón de descarte manual sigue disponible en ambos
    casos: no hace falta esperar a que se pare sola.
    GET /api/admin/emails-sistema-atascados
    """
    try:
        rows = rows_to_list(query(
            """SELECT id, evento_codigo, destinatario, asunto, LENGTH(cuerpo_html) as tam_html,
                      intentos, creado_en, descartado_en, enviado_no_confirmado,
                      (intentos >= %s OR descartado_en IS NOT NULL) AS atascado
                 FROM emails_sistema_pendientes
                WHERE enviado = FALSE
                ORDER BY tam_html DESC
                LIMIT 200""",
            (MAX_INTENTOS_EMAIL_SISTEMA,)
        )) or []
        return jsonify({"ok": True, "atascados": rows})
    except Exception as exc:
        log.error("[EMAILS-SISTEMA] Error listando atascados: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/emails-sistema-pendientes/<int:email_id>/descartar", methods=["POST"])
@admin_required
def api_descartar_email_sistema(email_id):
    """
    (2026-08-19) Descarta a mano una fila de la cola de emails de sistema
    (marca descartado_en=NOW()) — dejará de reintentarse. No se envía ni
    se borra el registro al momento; ver
    _job_purgar_emails_sistema_descartados (más abajo, corre a diario)
    para el borrado automático a los 2 días — antes se quedaba como
    constancia para siempre (2026-08-31, a petición de Víctor: "esto, una
    vez descartado no tiene sentido seguir llenado la pantalla").
    POST /api/admin/emails-sistema-pendientes/<id>/descartar
    """
    try:
        execute(
            "UPDATE emails_sistema_pendientes SET descartado_en = NOW() WHERE id=%s AND enviado = FALSE",
            (email_id,)
        )
        get_db().commit()
        return jsonify({"ok": True})
    except Exception as exc:
        log.error("[EMAILS-SISTEMA] Error descartando id=%s: %s", email_id, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/emails-sistema-pendientes/<int:email_id>/reactivar", methods=["POST"])
@admin_required
def api_reactivar_email_sistema(email_id):
    """
    (2026-08-31) Contrapartida del descarte manual — Víctor: "podemos
    poner otro botón para reactivar". Limpia descartado_en (para que
    api_emails_sistema_pendientes vuelva a considerarla candidata) y,
    si ya había agotado los reintentos (intentos >= MAX_INTENTOS_EMAIL_SISTEMA,
    caso típico: se descartó una fila que ya estaba "parada"), resetea
    intentos a 0 — si no se reseteara, "Reactivar" no reactivaría nada de
    verdad: seguiría excluida de la cola por haber agotado el cupo, solo
    que etiquetada "parado" en vez de "descartado" en el panel.
    POST /api/admin/emails-sistema-pendientes/<id>/reactivar
    """
    try:
        execute(
            "UPDATE emails_sistema_pendientes "
            "SET descartado_en = NULL, "
            "    intentos = CASE WHEN intentos >= %s THEN 0 ELSE intentos END "
            "WHERE id=%s AND enviado = FALSE",
            (MAX_INTENTOS_EMAIL_SISTEMA, email_id)
        )
        get_db().commit()
        return jsonify({"ok": True})
    except Exception as exc:
        log.error("[EMAILS-SISTEMA] Error reactivando id=%s: %s", email_id, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/emails-sistema-pendientes/<int:email_id>/marcar-enviado", methods=["POST"])
@login_required
def api_marcar_email_sistema_enviado(email_id):
    """POST /api/emails-sistema-pendientes/<id>/marcar-enviado — tras enviarlo vía EmailJS.

    (2026-08-31) A petición de Víctor: "cuando el correo interno de PEDIDO
    ENVIADO AL PROVEEDOR va con copia al departamento A&B se marque
    automáticamente la casilla y en todos los casos que se ponga en copia
    al responsable del departamento también se marque la correspondiente
    (...) solo con el envío del correo" — este es justo el único sitio
    donde se confirma que un correo se ha enviado DE VERDAD (lo llama el
    navegador tras el ACK de EmailJS), así que es el punto correcto para
    aplicar las marcas — nunca al encolar (ver enviar_emails_estado /
    _encolar_email_pedido_retrasado, que solo calculan y guardan la
    intención en marca_comunicado_ab/marca_comunicado_jefe_dep de esta
    misma fila). Se usa OR sobre el valor ya guardado en pedidos: una vez
    marcada, una casilla nunca se vuelve a desmarcar sola.
    """
    try:
        cur = execute(
            "UPDATE emails_sistema_pendientes SET enviado=TRUE, enviado_en=NOW() "
            "WHERE id=%s "
            "RETURNING pedido_id, marca_comunicado_ab, marca_comunicado_jefe_dep",
            (email_id,)
        )
        fila = cur.fetchone()
        if fila and fila.get("pedido_id") and (fila.get("marca_comunicado_ab") or fila.get("marca_comunicado_jefe_dep")):
            # (2026-09-01) FIX: pedidos.comunicado_ab / comunicado_jefe_dep son
            # columnas INTEGER (0/1) — igual que en el resto de la app (ver
            # api_guardar_pedido, api_actualizar_pedido: "1 if data.get(...)
            # else 0"). El `comunicado_ab OR %s` de antes aplicaba el
            # operador lógico OR de SQL directamente sobre un INTEGER, lo
            # cual PostgreSQL rechaza con un error de tipo — de forma
            # SIEMPRE determinista, nunca aleatoria, cada vez que esta rama
            # se ejecutaba (marca_comunicado_ab o marca_comunicado_jefe_dep
            # a True). Como esa rama SOLO se activa para el correo interno
            # de "ENVIADO AL PROVEEDOR" (el único evento que pone estas
            # marcas), esto explica por qué justo esos envíos — y solo esos
            # — devolvían 500 en marcar-enviado el 100% de las veces,
            # provocando el reenvío real duplicado que reportó Víctor.
            # GREATEST(entero, entero) da el mismo resultado que un OR
            # lógico para valores 0/1, sin el problema de tipos.
            execute(
                "UPDATE pedidos SET "
                "comunicado_ab = GREATEST(comunicado_ab, %s), "
                "comunicado_jefe_dep = GREATEST(comunicado_jefe_dep, %s) "
                "WHERE id=%s",
                (1 if fila.get("marca_comunicado_ab") else 0,
                 1 if fila.get("marca_comunicado_jefe_dep") else 0,
                 fila["pedido_id"])
            )
        get_db().commit()
        return jsonify({"ok": True})
    except Exception as exc:
        log.error("[EMAILS-SISTEMA] Error marcando enviado id=%s: %s", email_id, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/emails-sistema-pendientes/<int:email_id>/marcar-enviado-no-confirmado", methods=["POST"])
@login_required
def api_marcar_email_sistema_enviado_no_confirmado(email_id):
    """POST /api/emails-sistema-pendientes/<id>/marcar-enviado-no-confirmado

    (2026-09-01) El navegador llama a esto cuando emailjs.send() SÍ ha
    entregado el correo de verdad pero la confirmación normal
    (marcar-enviado, varios intentos) ha fallado igualmente. A propósito
    es una operación mínima y aislada — un único UPDATE sin tocar
    `pedidos` ni las columnas comunicado_ab/jefe_dep — para que tenga
    muchas más papeletas de funcionar incluso si el fallo de la
    confirmación normal viniera de ese bloque concreto.

    Sube `intentos` a MAX_INTENTOS_EMAIL_SISTEMA: dado que el correo YA
    se entregó, no debe volver a reclamarse ni reenviarse nunca — antes,
    al quedar enviado=FALSE, la reserva de 2 minutos caducaba y el
    siguiente sondeo (u otra pestaña) volvía a llamar a emailjs.send()
    de verdad, duplicando la entrega real cada vez que la confirmación
    fallaba. `enviado_no_confirmado=TRUE` dejar rastro para el panel de
    admin (ver api_emails_sistema_atascados): así se distingue de una
    fila que nunca llegó a enviarse (esas sí pueden "Reactivarse" con
    seguridad; estas no deben reenviarse jamás, solo cerrarse a mano con
    "Marcar como enviado").
    """
    try:
        execute(
            "UPDATE emails_sistema_pendientes "
            "SET intentos = %s, en_proceso_desde = NOW(), enviado_no_confirmado = TRUE "
            "WHERE id=%s AND enviado = FALSE",
            (MAX_INTENTOS_EMAIL_SISTEMA, email_id)
        )
        get_db().commit()
        return jsonify({"ok": True})
    except Exception as exc:
        log.error("[EMAILS-SISTEMA] Error marcando enviado-no-confirmado id=%s: %s", email_id, exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/techo-dedup-reset", methods=["POST"])
@admin_required
def techo_dedup_reset():
    """
    Borra los registros de deduplicación de techo del día actual en whatsapp_log,
    permitiendo que el job vuelva a enviar las alertas aunque ya lo haya hecho hoy.
    POST /api/admin/techo-dedup-reset
    """
    try:
        db  = get_db()
        cur = db.cursor()
        cur.execute(
            """DELETE FROM whatsapp_log
               WHERE tipo LIKE 'telegram_techo_mes_%'
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date"""
        )
        deleted = cur.rowcount
        db.commit()
        log.info("[TECHO-DEDUP-RESET] Eliminados %d registros de dedup del dia — forzado por admin", deleted)
        return jsonify({"ok": True, "eliminados": deleted,
                        "mensaje": f"{deleted} registros de deduplicacion eliminados. Ahora puedes lanzar test-techo-mensual."})
    except Exception as exc:
        log.error("[TECHO-DEDUP-RESET] Error: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/reset-alertas-hoy", methods=["POST"])
@admin_required
def reset_alertas_hoy():
    """
    Borra los registros de deduplicacion del dia actual para alertas de pedidos (telegram_auto),
    permitiendo que el job diario vuelva a enviar como si fuera la primera vez hoy.
    POST /api/admin/reset-alertas-hoy
    """
    try:
        db  = get_db()
        cur = db.cursor()
        cur.execute(
            """DELETE FROM whatsapp_log
               WHERE tipo = 'telegram_auto'
                 AND DATE(creado_en AT TIME ZONE 'Atlantic/Canary') =
                     (NOW() AT TIME ZONE 'Atlantic/Canary')::date"""
        )
        deleted = cur.rowcount
        db.commit()
        log.info("[ALERTAS-RESET] Eliminados %d registros telegram_auto del dia — forzado por admin", deleted)
        return jsonify({"ok": True, "eliminados": deleted,
                        "mensaje": f"{deleted} registros de dedup de pedidos eliminados. Ahora puedes lanzar test-scheduler."})
    except Exception as exc:
        log.error("[ALERTAS-RESET] Error: %s", exc)
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/admin/integridad", methods=["GET"])
@admin_required
def get_integridad():
    """
    Ejecuta _validar_integridad_operativa() en tiempo real y devuelve el resultado.
    Usado por el panel de admin para mostrar el badge de aviso y el detalle de problemas.
    GET /api/admin/integridad
    """
    resultado = _validar_integridad_operativa()
    return jsonify(resultado)


@app.route("/api/admin/db-size", methods=["GET"])
@admin_required
def get_db_size():
    """
    Historial de tamaño de base de datos (últimos 30 días registrados por
    _job_db_size_tracking) + el tamaño EN VIVO ahora mismo, calculado al
    vuelo — así la primera vez que se abre esta vista tras desplegar (antes
    de que corra el job de las 08:10) ya hay algo que mostrar.
    GET /api/admin/db-size
    """
    historial = query("""
        SELECT fecha, bytes_total, bytes_adjuntos
        FROM db_size_tracking
        ORDER BY fecha DESC
        LIMIT 30
    """)
    try:
        actual = query("""
            SELECT
                pg_database_size(current_database()) AS bytes_total,
                COALESCE(pg_total_relation_size('pedido_adjuntos'), 0) AS bytes_adjuntos
        """, one=True)
    except Exception:
        actual = None

    # Progreso de la migración a Storage (v12.8.0) — cuántos adjuntos de
    # pedidos cerrados ya están en Storage vs cuántos siguen en la BD.
    migracion = query("""
        SELECT
            COUNT(*) FILTER (WHERE pa.storage_path IS NOT NULL) AS migrados,
            COUNT(*) FILTER (WHERE pa.storage_path IS NULL AND p.estado = ANY(%s)) AS pendientes
        FROM pedido_adjuntos pa
        JOIN pedidos p ON p.id = pa.pedido_id
    """, (list(ESTADOS_CERRADOS),), one=True)

    ultimo_vacuum = query("""
        SELECT fecha, mb_antes, mb_despues, mb_liberados
        FROM db_vacuum_log
        ORDER BY fecha DESC
        LIMIT 1
    """, one=True)

    return jsonify({
        "ok": True,
        "actual": {
            "bytes_total": actual["bytes_total"] if actual else None,
            "bytes_adjuntos": actual["bytes_adjuntos"] if actual else None,
        },
        "storage": {
            "configurado": STORAGE_CONFIGURADO,
            "migrados": migracion["migrados"] if migracion else 0,
            "pendientes": migracion["pendientes"] if migracion else 0,
            "ultimo_vacuum": {
                "fecha": ultimo_vacuum["fecha"].isoformat(),
                "mb_antes": float(ultimo_vacuum["mb_antes"]),
                "mb_despues": float(ultimo_vacuum["mb_despues"]),
                "mb_liberados": float(ultimo_vacuum["mb_liberados"]),
            } if ultimo_vacuum else None,
        },
        "historial": [
            {
                "fecha": f["fecha"].isoformat(),
                "bytes_total": f["bytes_total"],
                "bytes_adjuntos": f["bytes_adjuntos"],
            }
            for f in historial
        ],
    })


@app.route("/api/admin/migrar-adjuntos-storage", methods=["POST"])
@admin_required
def migrar_adjuntos_storage_manual():
    """
    Lanza un lote de migración a Storage inmediatamente (mismo trabajo que
    el job de las 03:00), en primer plano — así el admin ve el resultado
    al momento en vez de esperar a la log. Útil para arrancar la migración
    nada más desplegar, sin esperar a la madrugada.

    A propósito NO compacta (VACUUM FULL) aquí, aunque el job nocturno sí
    lo haga tras migrar algo — VACUUM FULL bloquea la tabla por completo
    mientras dura, y este botón puede pulsarse en horario de oficina con
    gente usando la app. La compactación solo se dispara de madrugada.
    POST /api/admin/migrar-adjuntos-storage
    """
    if not STORAGE_CONFIGURADO:
        return jsonify({
            "ok": False,
            "error": "Storage no configurado — faltan SUPABASE_URL y/o SUPABASE_SERVICE_ROLE_KEY "
                     "en las variables de entorno de Render."
        }), 400
    resumen = _job_migrar_adjuntos_storage()
    log.info("▶ [MANUAL] Migración a Storage lanzada manualmente por admin: %s", resumen)
    return jsonify({"ok": True, **resumen})


@app.route("/api/admin/test-health", methods=["POST"])
@admin_required
def test_health_check():
    """
    Fuerza el job de health check inmediatamente (mismo que corre a las 07:05).
    Envía Telegram al admin si hay problemas.
    POST /api/admin/test-health
    """
    import threading
    t = threading.Thread(target=lambda: _job_health_check(force=True), daemon=True)
    t.start()
    log.info("▶ [MANUAL] Job health-check lanzado manualmente por admin")
    return jsonify({"ok": True, "mensaje": "Health check ejecutándose — revisa Telegram en unos segundos."})


@app.route("/api/admin/test-egress", methods=["POST"])
@admin_required
def test_egress_alerta():
    """
    Fuerza el job combinado de alerta de consumo (egress + tamaño BD)
    inmediatamente (mismo que corre a las 08:30), ignorando el umbral y la
    deduplicación diaria — sirve para confirmar que el canal de Telegram
    funciona. Ruta y nombre de función sin cambiar (por compatibilidad con
    el botón ya existente en el frontend), aunque desde Jul 2026 dispara
    la alerta combinada, no solo egress.
    POST /api/admin/test-egress
    """
    import threading
    t = threading.Thread(target=lambda: _job_alerta_consumo(force=True), daemon=True)
    t.start()
    log.info("▶ [MANUAL] Job alerta-consumo (egress + tamaño BD) lanzado manualmente por admin")
    return jsonify({"ok": True, "mensaje": "Alerta de consumo ejecutándose — revisa Telegram en unos segundos."})


# ── RESTAURACIÓN DE BACKUP ────────────────────────────────────────────────────

def _normalizar_ruta_backup(ruta):
    """Normaliza una ruta de carpeta para comparaciones fiables: sin espacios
    sobrantes, sin barra final y sin distinguir mayúsculas (Windows no
    distingue mayúsculas/minúsculas en rutas). Debe coincidir exactamente
    con la normalización que aplica restore_agent.py al escribir la caché."""
    return ruta.strip().rstrip("\\/").lower()


@app.route("/api/admin/backup/listar", methods=["POST"])
@admin_required
def backup_listar():
    """
    Devuelve la lista de backups disponibles para la ruta indicada.

    Fix v11.8.6: esta ruta dejó de intentar leer Path(ruta) directamente en
    el servidor — Render no tiene acceso a la red local de la oficina, el
    mismo motivo por el que /api/admin/backup/restaurar ya pasó a usar una
    cola (restore_queue) en vez de ejecutar acciones contra el filesystem.

    Ahora lee de `backups_cache`, una tabla que `restore_agent.py` mantiene
    sincronizada desde tu PC en cada ciclo (escanea BACKUP_DESTINO y sube el
    resultado a Supabase). Esta ruta nunca toca disco.
    """
    data = request.get_json(silent=True) or {}
    ruta = data.get("ruta", "").strip()

    if not ruta:
        return jsonify({"ok": False, "error": "Ruta no especificada"}), 400

    ruta_norm = _normalizar_ruta_backup(ruta)

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            SELECT nombre, fecha, mb, adjuntos, tiene_log, valido, tipo, actualizado_en
            FROM backups_cache
            WHERE ruta_normalizada = %s
            ORDER BY fecha_raw DESC NULLS LAST, nombre DESC
        """, (ruta_norm,))
        filas = cur.fetchall()

        # Fix "sin sincronizar" falso (15 jul 2026): `actualizado_en` de
        # backups_cache solo se toca cuando un backup cambia de verdad —
        # normal casi todo el día, ya que solo hay un backup diario a las
        # 17:00. Usarlo como "última vez que corrió el agente" generaba
        # avisos de agente caído durante horas aunque estuviera
        # sincronizando bien cada 5 minutos sin encontrar nada nuevo que
        # subir. `agente_heartbeat` (restore_agent.py ≥ 15 jul 2026) sí
        # registra cada ciclo, haya cambios o no. Si la tabla todavía no
        # existe (agente sin actualizar) o no tiene fila para esta ruta,
        # se usa el cálculo antiguo como red de seguridad.
        ultimo_heartbeat = None
        try:
            cur.execute(
                "SELECT visto_en FROM agente_heartbeat WHERE ruta_normalizada = %s",
                (ruta_norm,)
            )
            fila_hb = cur.fetchone()
            if fila_hb:
                ultimo_heartbeat = fila_hb["visto_en"]
        except Exception:
            db.rollback()

    if not filas:
        return jsonify({
            "ok": True,
            "backups": [],
            "aviso": (
                "El agente local (restore_agent.py) todavía no ha reportado ningún "
                "backup para esta ruta exacta. Comprueba que la tarea programada esté "
                "activa en el PC y que la ruta coincide con BACKUP_DESTINO en "
                "restore_agent.bat."
            ),
        })

    ultimo_escaneo = ultimo_heartbeat or max(f["actualizado_en"] for f in filas)
    minutos = int((datetime.now(timezone.utc) - ultimo_escaneo).total_seconds() // 60)

    resp = {
        "ok": True,
        "backups": [
            {
                "nombre":    f["nombre"],
                "fecha":     f["fecha"],
                "mb":        float(f["mb"]),
                "adjuntos":  f["adjuntos"],
                "tiene_log": f["tiene_log"],
                "valido":    f["valido"],
                "tipo":      f["tipo"],
            }
            for f in filas
        ],
        "ultimo_escaneo_minutos": minutos,
    }

    if minutos > 5:
        resp["aviso"] = (
            f"El agente local lleva {minutos} minutos sin sincronizar la caché. "
            "Si tu PC está apagado o la tarea programada está desactivada, esta "
            "lista puede no reflejar los backups más recientes."
        )

    return jsonify(resp)


@app.route("/api/admin/backup/log", methods=["POST"])
@admin_required
def backup_ver_log():
    """
    Devuelve el contenido del backup_log.txt de un backup concreto.
    POST /api/admin/backup/log
    Body JSON: { "ruta": "...", "nombre": "backup_20260616_1700" }

    Fix v11.8.6: igual que /api/admin/backup/listar, esta ruta dejó de leer
    el fichero directamente desde el filesystem de Render (sin acceso a la
    red local). El contenido se lee de `backups_cache`, donde lo deja
    restore_agent.py al sincronizar la lista de backups.
    """
    data   = request.get_json(silent=True) or {}
    ruta   = data.get("ruta",   "").strip()
    nombre = data.get("nombre", "").strip()

    if not ruta or not nombre:
        return jsonify({"ok": False, "error": "Faltan parámetros"}), 400

    ruta_norm = _normalizar_ruta_backup(ruta)

    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            SELECT tiene_log, log_contenido
            FROM backups_cache
            WHERE ruta_normalizada = %s AND nombre = %s
        """, (ruta_norm, nombre))
        fila = cur.fetchone()

    if not fila:
        return jsonify({
            "ok": False,
            "error": "Este backup no aparece en la caché del agente local. "
                     "Pulsa \"Actualizar lista\" para refrescarla."
        }), 404

    if not fila["tiene_log"] or not fila["log_contenido"]:
        return jsonify({"ok": False, "error": "El fichero backup_log.txt no existe en este backup"}), 404

    return jsonify({"ok": True, "log": fila["log_contenido"], "nombre": nombre})


@app.route("/api/admin/backup/restaurar", methods=["POST"])
@admin_required
def backup_restaurar():
    """
    OPCIÓN C — Cola de restauración desacoplada.

    Esta ruta NO restaura nada directamente (Render no tiene acceso a la
    carpeta de red local). En su lugar, registra una petición en la tabla
    `restore_queue`. Un agente local (restore_agent.py), ejecutado en el PC
    con acceso a \\shtabaiba\... y a Supabase, sondea esta tabla cada minuto,
    procesa la petición pendiente más antigua y marca el resultado.

    POST /api/admin/backup/restaurar
    Body JSON: { "ruta": "...", "nombre": "backup_20260616_1700", "modo": "pedidos" }
    """
    data   = request.get_json(silent=True) or {}
    ruta   = data.get("ruta",   "").strip()
    nombre = data.get("nombre", "").strip()
    modo   = data.get("modo",   "pedidos")

    if not ruta or not nombre:
        return jsonify({"ok": False, "error": "Faltan parámetros: ruta y nombre son obligatorios"}), 400
    if modo not in ("pedidos", "completo"):
        return jsonify({"ok": False, "error": "Modo no válido. Usa 'pedidos' o 'completo'"}), 400

    # Evitar encolar si ya hay una petición pendiente o en proceso
    db = get_db()
    with db.cursor() as cur:
        cur.execute(
            "SELECT id, backup_nombre, estado FROM restore_queue "
            "WHERE estado IN ('pendiente','en_proceso') "
            "ORDER BY solicitado_en DESC LIMIT 1"
        )
        existente = cur.fetchone()

        if existente:
            return jsonify({
                "ok": False,
                "error": f"Ya hay una restauración {existente['estado']} "
                         f"({existente['backup_nombre']}). Espera a que finalice."
            }), 409

        usuario_nombre = session.get("nombre") or session.get("username") or "admin"

        cur.execute("""
            INSERT INTO restore_queue (backup_nombre, backup_ruta, modo, estado, solicitado_por)
            VALUES (%s, %s, %s, 'pendiente', %s)
            RETURNING id
        """, (nombre, ruta, modo, usuario_nombre))
        nueva_id = cur.fetchone()["id"]
        db.commit()

    log.info("[RESTORE-QUEUE] Petición #%s encolada. backup=%s modo=%s por=%s",
              nueva_id, nombre, modo, usuario_nombre)

    return jsonify({
        "ok": True,
        "encolado": True,
        "queue_id": nueva_id,
        "mensaje": "Petición registrada. El agente local la procesará en menos de 1 minuto."
    })


@app.route("/api/admin/backup/estado", methods=["GET"])
@admin_required
def backup_estado_cola():
    """
    Devuelve el estado de la última petición de restauración encolada,
    para que el panel web haga polling y muestre el progreso en tiempo real.

    GET /api/admin/backup/estado
    """
    db = get_db()
    with db.cursor() as cur:
        cur.execute("""
            SELECT id, backup_nombre, modo, estado, solicitado_por,
                   solicitado_en, iniciado_en, completado_en, resumen, error_msg,
                   pre_restore_backup
            FROM restore_queue
            ORDER BY solicitado_en DESC
            LIMIT 1
        """)
        fila = cur.fetchone()

    if not fila:
        return jsonify({"ok": True, "hay_peticion": False})

    return jsonify({
        "ok": True,
        "hay_peticion": True,
        "id":                 fila["id"],
        "backup_nombre":      fila["backup_nombre"],
        "modo":               fila["modo"],
        "estado":             fila["estado"],
        "solicitado_por":     fila["solicitado_por"],
        "solicitado_en":      fila["solicitado_en"].isoformat() if fila["solicitado_en"] else None,
        "iniciado_en":        fila["iniciado_en"].isoformat()   if fila["iniciado_en"]   else None,
        "completado_en":      fila["completado_en"].isoformat() if fila["completado_en"] else None,
        "resumen":            fila["resumen"],
        "error_msg":          fila["error_msg"],
        "pre_restore_backup": fila["pre_restore_backup"],
    })


# ── Arranque ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    # v12.7.0: ya no hay Socket.IO en este proceso (el chat vive en su
    # propio servicio, ver control_pedidos_chat). app.run normal basta
    # para pruebas en local; en Render, gunicorn sirve app:app tal cual
    # — ver GUIA_DESPLIEGUE.md para el nuevo Start Command sin eventlet.
    app.run(host="0.0.0.0", port=port, debug=False)
